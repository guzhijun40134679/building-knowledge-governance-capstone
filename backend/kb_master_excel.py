from __future__ import annotations

import os
import tempfile
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from kb_db import BACKEND_DIR, DATA_DIR, connect_db, load_field_catalog, FIELD_GROUP_ORDER
from kb_unknowns import (
    normalize_booleanish,
    normalize_field_value,
    normalize_provider_text,
    normalize_requirement_choice,
    normalize_unknown_value,
)


# English is the canonical workbook format for this clean capstone edition.
MASTER_MAIN_SHEET = "Buildings"
MASTER_PLAN_SHEET = "Provider Plan Review"
MASTER_HELP_SHEET = "Field Guide"
STAGING_STATUS_HEADER = "Review Status"
LEGACY_SHEET_NAMES = {
    "楼宇主表": MASTER_MAIN_SHEET,
    "套餐人工确认表": MASTER_PLAN_SHEET,
    "字段说明": MASTER_HELP_SHEET,
}
# These persisted status codes are retained for compatibility with the original edition.
STAGING_STATUS_PENDING = "待补充"
STAGING_STATUS_ACTIVE = "临时"
STAGING_STATUS_MASTERED = "已入正式"
VALID_STAGING_STATUSES = {
    STAGING_STATUS_PENDING,
    STAGING_STATUS_ACTIVE,
    STAGING_STATUS_MASTERED,
}

MASTER_MAIN_HEADERS = [
    "Building Name",
    "Address",
    "Renters Insurance Required",
    "Insurance Coverage Amount",
    "COI Required",
    "COI Trigger",
    "Key Pickup Instructions",
    "Service Elevator Booking",
    "Electricity Account Required",
    "Electricity Provider",
    "Resident Internet Setup Required",
    "Verizon Supported",
    "Verizon Plan Tiers",
    "Verizon Notes / Contact",
    "Xfinity Supported",
    "Xfinity Plan Tiers",
    "Xfinity Notes / Contact",
    "Spectrum Supported",
    "Spectrum Plan Tiers",
    "Spectrum Notes / Contact",
    "Astound Supported",
    "Astound Plan Tiers",
    "Astound Notes / Contact",
    "Additional Internet Providers",
    "Internet Notes",
    "Move-In Notes",
]

STAGING_MAIN_HEADERS = [*MASTER_MAIN_HEADERS, STAGING_STATUS_HEADER]

PACKAGE_CONFIRM_HEADERS = [
    "Building Name",
    "Provider",
    "Supported",
    "Manually Confirmed Plans",
    "Confirmed By",
    "Confirmation Date",
    "Notes",
]

FIELD_HELP_ROWS = [
    ("Field", "Entry Rule"),
    ("Building Name", "Required. Use one row per building and keep the name stable."),
    ("Address", "Use the complete address when possible; it helps distinguish buildings with similar names."),
    ("Renters Insurance Required", "Enter Yes, No, or Optional. Leave blank when the source is unclear."),
    ("Insurance Coverage Amount", "Enter an amount only when the source states it explicitly."),
    ("COI Required", "Enter Yes, No, or Manual Review. Leave blank when the source does not say."),
    ("COI Trigger", "Examples: using movers, booking a service elevator, or receiving a large delivery."),
    ("Key Pickup Instructions", "Record only explicit key, fob, concierge, or leasing-office pickup instructions."),
    ("Service Elevator Booking", "Record only explicit service-elevator or move-in appointment instructions."),
    ("Electricity Account Required", "Enter Yes, No, or Optional. Leave blank when unclear."),
    ("Electricity Provider", "Enter only a utility explicitly named by the source."),
    ("Resident Internet Setup Required", "Enter Yes, No, or Optional. Leave blank when unclear."),
    ("Verizon Supported", "Enter Yes or No; leave blank when unclear."),
    ("Verizon Plan Tiers", "Enter manually verified Verizon plans, one per line when useful."),
    ("Verizon Notes / Contact", "Use only for Verizon contacts, URLs, setup instructions, or special notes."),
    ("Xfinity Supported", "Enter Yes or No; leave blank when unclear."),
    ("Xfinity Plan Tiers", "Enter manually verified Xfinity plans, one per line when useful."),
    ("Xfinity Notes / Contact", "Use only for Xfinity contacts, URLs, setup instructions, or special notes."),
    ("Spectrum Supported", "Enter Yes or No; leave blank when unclear."),
    ("Spectrum Plan Tiers", "Enter manually verified Spectrum plans, one per line when useful."),
    ("Spectrum Notes / Contact", "Use only for Spectrum contacts, URLs, setup instructions, or special notes."),
    ("Astound Supported", "Enter Yes or No; leave blank when unclear."),
    ("Astound Plan Tiers", "Enter manually verified Astound plans, one per line when useful."),
    ("Astound Notes / Contact", "Use only for Astound contacts, URLs, setup instructions, or special notes."),
    ("Additional Internet Providers", "List providers outside the four fixed providers, separated by commas."),
    ("Internet Notes", "Use for other providers or notes that do not fit the structured provider columns."),
    ("Move-In Notes", "Record move-in steps, insurance requirements, and other explicit reminders."),
]

HEADER_TO_FIELD_KEY = {
    "Building Name": "building_name",
    "Address": "address",
    "Renters Insurance Required": "insurance_required",
    "Insurance Coverage Amount": "insurance_coverage_amount",
    "COI Required": "insurance_coi_required",
    "COI Trigger": "insurance_coi_trigger",
    "Key Pickup Instructions": "key_pickup_notes",
    "Service Elevator Booking": "service_elevator_booking_notes",
    "Electricity Account Required": "electricity_required",
    "Electricity Provider": "electricity_provider",
    "Resident Internet Setup Required": "internet_self_setup_required",
    "Verizon Supported": "internet_verizon_supported",
    "Verizon Plan Tiers": "internet_verizon_plan_tiers",
    "Verizon Notes / Contact": "internet_verizon_notes",
    "Xfinity Supported": "internet_xfinity_supported",
    "Xfinity Plan Tiers": "internet_xfinity_plan_tiers",
    "Xfinity Notes / Contact": "internet_xfinity_notes",
    "Spectrum Supported": "internet_spectrum_supported",
    "Spectrum Plan Tiers": "internet_spectrum_plan_tiers",
    "Spectrum Notes / Contact": "internet_spectrum_notes",
    "Astound Supported": "internet_astound_supported",
    "Astound Plan Tiers": "internet_astound_plan_tiers",
    "Astound Notes / Contact": "internet_astound_notes",
    "Additional Internet Providers": "internet_provider",
    "Internet Notes": "internet_notes",
    "Move-In Notes": "move_in_notes",
    # Legacy Chinese headers remain accepted when importing existing workbooks.
    "大楼名称": "building_name",
    "地址": "address",
    "是否需要保险": "insurance_required",
    "保险保额": "insurance_coverage_amount",
    "是否需要COI": "insurance_coi_required",
    "COI触发条件": "insurance_coi_trigger",
    "钥匙领取说明": "key_pickup_notes",
    "货梯预约说明": "service_elevator_booking_notes",
    "是否需要开电": "electricity_required",
    "电力公司": "electricity_provider",
    "是否需要自己开网": "internet_self_setup_required",
    "Verizon是否支持": "internet_verizon_supported",
    "Verizon套餐档位": "internet_verizon_plan_tiers",
    "Verizon备注/联系人": "internet_verizon_notes",
    "Xfinity是否支持": "internet_xfinity_supported",
    "Xfinity套餐档位": "internet_xfinity_plan_tiers",
    "Xfinity备注/联系人": "internet_xfinity_notes",
    "Spectrum是否支持": "internet_spectrum_supported",
    "Spectrum套餐档位": "internet_spectrum_plan_tiers",
    "Spectrum备注/联系人": "internet_spectrum_notes",
    "Astound是否支持": "internet_astound_supported",
    "Astound套餐档位": "internet_astound_plan_tiers",
    "Astound备注/联系人": "internet_astound_notes",
    "额外网络运营商": "internet_provider",
    "网络备注": "internet_notes",
    "入住备注": "move_in_notes",
}

LEGACY_HEADER_TO_ENGLISH = {
    "大楼名称": "Building Name",
    "地址": "Address",
    "是否需要保险": "Renters Insurance Required",
    "保险保额": "Insurance Coverage Amount",
    "是否需要COI": "COI Required",
    "COI触发条件": "COI Trigger",
    "钥匙领取说明": "Key Pickup Instructions",
    "货梯预约说明": "Service Elevator Booking",
    "是否需要开电": "Electricity Account Required",
    "电力公司": "Electricity Provider",
    "是否需要自己开网": "Resident Internet Setup Required",
    "Verizon是否支持": "Verizon Supported",
    "Verizon套餐档位": "Verizon Plan Tiers",
    "Verizon备注/联系人": "Verizon Notes / Contact",
    "Xfinity是否支持": "Xfinity Supported",
    "Xfinity套餐档位": "Xfinity Plan Tiers",
    "Xfinity备注/联系人": "Xfinity Notes / Contact",
    "Spectrum是否支持": "Spectrum Supported",
    "Spectrum套餐档位": "Spectrum Plan Tiers",
    "Spectrum备注/联系人": "Spectrum Notes / Contact",
    "Astound是否支持": "Astound Supported",
    "Astound套餐档位": "Astound Plan Tiers",
    "Astound备注/联系人": "Astound Notes / Contact",
    "额外网络运营商": "Additional Internet Providers",
    "网络备注": "Internet Notes",
    "入住备注": "Move-In Notes",
    "审核状态": STAGING_STATUS_HEADER,
}

CORE_SNAPSHOT_FIELD_KEYS = {
    "building_name",
    "address",
    "insurance_required",
    "insurance_coverage_amount",
    "electricity_required",
    "electricity_provider",
    "internet_self_setup_required",
    "internet_provider",
    "internet_notes",
    "move_in_notes",
    "source_type",
    "source_file",
    "source_date",
    "info_cutoff_date",
}

INSURANCE_STATUS_FIELD_KEYS = {
    "insurance_coi_required",
    "insurance_renters_required",
    "insurance_personal_property_required",
    "insurance_personal_liability_required",
    "insurance_interested_party_required",
    "insurance_additional_insured_required",
    "insurance_certificate_holder_required",
}

GROUP_HEADER_ANCHORS = {
    "basic": "Address",
    "insurance": "Insurance Coverage Amount",
    "electricity": "Electricity Provider",
    "internet": "Internet Notes",
    "move_in": "Move-In Notes",
}

FIELD_HELP_RULE_OVERRIDES = dict(FIELD_HELP_ROWS[1:])

NETWORK_PROVIDER_FIELD_MAP = {
    "internet_verizon_supported": "Verizon",
    "internet_xfinity_supported": "Xfinity",
    "internet_spectrum_supported": "Spectrum",
    "internet_astound_supported": "Astound",
}
NETWORK_PLAN_FIELD_MAP = {
    "internet_verizon_plan_tiers": "Verizon",
    "internet_xfinity_plan_tiers": "Xfinity",
    "internet_spectrum_plan_tiers": "Spectrum",
    "internet_astound_plan_tiers": "Astound",
}
NETWORK_PROVIDER_NOTE_FIELD_MAP = {
    "internet_verizon_notes": "Verizon",
    "internet_xfinity_notes": "Xfinity",
    "internet_spectrum_notes": "Spectrum",
    "internet_astound_notes": "Astound",
}
FIXED_PROVIDER_LABELS = set(NETWORK_PROVIDER_FIELD_MAP.values())


def _catalog_entries_for_excel(*, include_staging_only: bool) -> List[dict]:
    try:
        conn = connect_db()
    except Exception:
        return []
    try:
        catalog = load_field_catalog(conn, statuses=("active",))
    finally:
        conn.close()
    entries: List[dict] = []
    for item in catalog:
        scope = (item.get("scope") or "master_and_staging").strip()
        if scope == "staging_only" and not include_staging_only:
            continue
        header = (normalize_unknown_value(item.get("excel_header_name")) or item.get("display_name") or item.get("field_key") or "").strip()
        if not header:
            continue
        entries.append(
            {
                **item,
                "group_key": item.get("group_key") or "custom",
                "display_order": int(item.get("display_order") or 900),
                "excel_header_name": header,
            }
        )
    entries.sort(
        key=lambda item: (
            FIELD_GROUP_ORDER.get(item["group_key"], 999),
            item["display_order"],
            item["display_name"],
        )
    )
    return entries


def current_master_headers() -> List[str]:
    headers = list(MASTER_MAIN_HEADERS)
    grouped: Dict[str, List[str]] = {}
    for item in _catalog_entries_for_excel(include_staging_only=False):
        header = item["excel_header_name"]
        if header in headers:
            continue
        grouped.setdefault(item["group_key"], []).append(header)
    for group_key in sorted(grouped, key=lambda key: FIELD_GROUP_ORDER.get(key, 999)):
        extra_headers = grouped[group_key]
        if group_key == "custom":
            headers.extend(extra_headers)
            continue
        anchor = GROUP_HEADER_ANCHORS.get(group_key)
        insert_at = headers.index(anchor) + 1 if anchor in headers else len(headers)
        for header in extra_headers:
            headers.insert(insert_at, header)
            insert_at += 1
    return headers


def current_staging_headers() -> List[str]:
    headers = list(current_master_headers())
    if STAGING_STATUS_HEADER not in headers:
        headers.append(STAGING_STATUS_HEADER)
    return headers


def current_header_to_field_key(*, include_staging_only: bool) -> Dict[str, str]:
    mapping = dict(HEADER_TO_FIELD_KEY)
    for item in _catalog_entries_for_excel(include_staging_only=include_staging_only):
        mapping[item["excel_header_name"]] = item["field_key"]
    return mapping


def current_field_help_rows(*, include_staging_only: bool) -> List[Tuple[str, str]]:
    rows: List[Tuple[str, str]] = [FIELD_HELP_ROWS[0]]
    entries_by_header = {
        item["excel_header_name"]: item for item in _catalog_entries_for_excel(include_staging_only=include_staging_only)
    }
    for header in current_master_headers():
        help_text = FIELD_HELP_RULE_OVERRIDES.get(header)
        if not help_text:
            entry = entries_by_header.get(header) or {}
            help_text = normalize_unknown_value(entry.get("description")) or "Leave blank when the source does not state a value explicitly."
        rows.append((header, help_text))
    if include_staging_only:
        rows.append((STAGING_STATUS_HEADER, "System-managed column. Display values are Needs Information, Staging, or Promoted to Master."))
    return rows


def extra_provider_text(value: object) -> str:
    provider_text = normalize_provider_text(value) or ""
    if not provider_text:
        return ""
    labels = [label for label in (part.strip() for part in provider_text.split(",")) if label]
    extra_only = [label for label in labels if label not in FIXED_PROVIDER_LABELS]
    return ", ".join(dict.fromkeys(extra_only))


def resolve_master_excel_path() -> Path:
    configured = os.getenv("WHITEPAPER_MASTER_XLSX_PATH", "").strip()
    if configured:
        path = Path(configured)
        if not path.is_absolute():
            path = (BACKEND_DIR / path).resolve()
        legacy_path = (DATA_DIR / "master_excel" / "楼宇知识库母表.xlsx").resolve()
        if (
            not path.exists()
            and path.name == "building_knowledge_master.xlsx"
            and legacy_path.exists()
        ):
            return legacy_path
        return path
    default_path = (DATA_DIR / "master_excel" / "building_knowledge_master.xlsx").resolve()
    legacy_path = (DATA_DIR / "master_excel" / "楼宇知识库母表.xlsx").resolve()
    return legacy_path if not default_path.exists() and legacy_path.exists() else default_path


def resolve_staging_excel_path() -> Path:
    configured = os.getenv("WHITEPAPER_STAGING_XLSX_PATH", "").strip()
    if configured:
        path = Path(configured)
        if not path.is_absolute():
            path = (BACKEND_DIR / path).resolve()
        legacy_path = (DATA_DIR / "master_excel" / "excel临时库.xlsx").resolve()
        if (
            not path.exists()
            and path.name == "building_knowledge_staging.xlsx"
            and legacy_path.exists()
        ):
            return legacy_path
        return path
    default_path = (DATA_DIR / "master_excel" / "building_knowledge_staging.xlsx").resolve()
    legacy_path = (DATA_DIR / "master_excel" / "excel临时库.xlsx").resolve()
    return legacy_path if not default_path.exists() and legacy_path.exists() else default_path


def bool_to_excel_text(value: Optional[bool]) -> str:
    if value is True:
        return "Yes"
    if value is False:
        return "No"
    return ""


def requirement_to_excel_text(value: object) -> str:
    choice = normalize_requirement_choice(value)
    if choice == "true":
        return "Yes"
    if choice == "false":
        return "No"
    if choice == "optional":
        return "Optional"
    return ""


def insurance_status_to_excel_text(value: object) -> str:
    text = (normalize_unknown_value(value) or "").strip().lower()
    if not text:
        return ""
    if text in {"yes", "true", "required", "是"}:
        return "Yes"
    if text in {"no", "false", "not_required", "否"}:
        return "No"
    if text in {"optional", "可选"}:
        return "Optional"
    if text in {"manual_review", "needs_manual_review", "需人工确认"}:
        return "Manual Review"
    return normalize_unknown_value(value) or ""


def insurance_status_from_excel_value(value: object) -> Optional[str]:
    text = (normalize_unknown_value(value) or "").strip().lower()
    if not text:
        return None
    if text in {"是", "yes", "true", "required"}:
        return "yes"
    if text in {"否", "no", "false", "not_required"}:
        return "no"
    if text in {"可选", "optional"}:
        return "optional"
    if text in {"需人工确认", "manual_review", "needs_manual_review"}:
        return "manual_review"
    return normalize_unknown_value(value)


def build_main_sheet_row(snapshot: dict) -> List[str]:
    row_by_header = build_main_sheet_row_values(snapshot)
    return [row_by_header.get(header, "") for header in current_master_headers()]


def normalize_staging_status(value: object) -> str:
    text = normalize_unknown_value(value) or ""
    english_to_code = {
        "needs information": STAGING_STATUS_PENDING,
        "staging": STAGING_STATUS_ACTIVE,
        "promoted to master": STAGING_STATUS_MASTERED,
    }
    if text.strip().lower() in english_to_code:
        return english_to_code[text.strip().lower()]
    if text in VALID_STAGING_STATUSES:
        return text
    return ""


def staging_status_to_excel_text(value: object) -> str:
    status = normalize_staging_status(value)
    return {
        STAGING_STATUS_PENDING: "Needs Information",
        STAGING_STATUS_ACTIVE: "Staging",
        STAGING_STATUS_MASTERED: "Promoted to Master",
    }.get(status, "")


def build_main_sheet_row_values(snapshot: dict) -> Dict[str, str]:
    extensions = snapshot.get("extensions", {}) or {}
    row_values = {
        "Building Name": normalize_unknown_value(snapshot.get("building_name")) or "",
        "Address": normalize_unknown_value(snapshot.get("address")) or "",
        "Renters Insurance Required": requirement_to_excel_text(snapshot.get("insurance_required")),
        "Insurance Coverage Amount": normalize_unknown_value(snapshot.get("insurance_coverage_amount")) or "",
        "COI Required": insurance_status_to_excel_text(
            snapshot.get("insurance_coi_required", extensions.get("insurance_coi_required"))
        ),
        "COI Trigger": normalize_unknown_value(
            snapshot.get("insurance_coi_trigger", extensions.get("insurance_coi_trigger"))
        )
        or "",
        "Key Pickup Instructions": normalize_unknown_value(
            snapshot.get("key_pickup_notes", extensions.get("key_pickup_notes"))
        )
        or "",
        "Service Elevator Booking": normalize_unknown_value(
            snapshot.get(
                "service_elevator_booking_notes",
                extensions.get("service_elevator_booking_notes"),
            )
        )
        or "",
        "Electricity Account Required": requirement_to_excel_text(snapshot.get("electricity_required")),
        "Electricity Provider": normalize_unknown_value(snapshot.get("electricity_provider")) or "",
        "Resident Internet Setup Required": requirement_to_excel_text(snapshot.get("internet_self_setup_required")),
        "Verizon Supported": bool_to_excel_text(normalize_booleanish(extensions.get("internet_verizon_supported"))),
        "Verizon Plan Tiers": normalize_unknown_value(extensions.get("internet_verizon_plan_tiers")) or "",
        "Verizon Notes / Contact": normalize_unknown_value(extensions.get("internet_verizon_notes")) or "",
        "Xfinity Supported": bool_to_excel_text(normalize_booleanish(extensions.get("internet_xfinity_supported"))),
        "Xfinity Plan Tiers": normalize_unknown_value(extensions.get("internet_xfinity_plan_tiers")) or "",
        "Xfinity Notes / Contact": normalize_unknown_value(extensions.get("internet_xfinity_notes")) or "",
        "Spectrum Supported": bool_to_excel_text(normalize_booleanish(extensions.get("internet_spectrum_supported"))),
        "Spectrum Plan Tiers": normalize_unknown_value(extensions.get("internet_spectrum_plan_tiers")) or "",
        "Spectrum Notes / Contact": normalize_unknown_value(extensions.get("internet_spectrum_notes")) or "",
        "Astound Supported": bool_to_excel_text(normalize_booleanish(extensions.get("internet_astound_supported"))),
        "Astound Plan Tiers": normalize_unknown_value(extensions.get("internet_astound_plan_tiers")) or "",
        "Astound Notes / Contact": normalize_unknown_value(extensions.get("internet_astound_notes")) or "",
        "Additional Internet Providers": extra_provider_text(snapshot.get("internet_provider")),
        "Internet Notes": normalize_unknown_value(snapshot.get("internet_notes")) or "",
        "Move-In Notes": normalize_unknown_value(snapshot.get("move_in_notes")) or "",
    }
    entries_by_field = {
        item["field_key"]: item for item in _catalog_entries_for_excel(include_staging_only=True)
    }
    for header, field_key in current_header_to_field_key(include_staging_only=True).items():
        if header in row_values or field_key == "library_status":
            continue
        entry = entries_by_field.get(field_key) or {}
        field_type = entry.get("field_type") or "text"
        value = snapshot.get(field_key)
        if value is None:
            value = extensions.get(field_key)
        if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
            row_values[header] = requirement_to_excel_text(value)
        elif field_key in INSURANCE_STATUS_FIELD_KEYS:
            row_values[header] = insurance_status_to_excel_text(value)
        elif field_key in NETWORK_PROVIDER_FIELD_MAP or field_type == "boolean":
            row_values[header] = bool_to_excel_text(normalize_booleanish(value))
        elif field_key == "internet_provider":
            row_values[header] = extra_provider_text(value)
        else:
            row_values[header] = normalize_unknown_value(value) or ""
    return row_values


def build_staging_sheet_row(snapshot: dict) -> List[str]:
    row_by_header = build_main_sheet_row_values(snapshot)
    row_by_header[STAGING_STATUS_HEADER] = staging_status_to_excel_text(snapshot.get("library_status"))
    return [row_by_header.get(header, "") for header in current_staging_headers()]


def create_standard_master_workbook(building_rows: Optional[Sequence[dict]] = None) -> Workbook:
    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = MASTER_MAIN_SHEET
    master_headers = current_master_headers()
    main_sheet.append(master_headers)
    main_sheet.freeze_panes = "A2"
    for cell in main_sheet[1]:
        cell.font = Font(bold=True)

    for snapshot in building_rows or []:
        main_sheet.append(build_main_sheet_row(snapshot))

    package_sheet = workbook.create_sheet(MASTER_PLAN_SHEET)
    package_sheet.append(PACKAGE_CONFIRM_HEADERS)
    for cell in package_sheet[1]:
        cell.font = Font(bold=True)

    help_sheet = workbook.create_sheet(MASTER_HELP_SHEET)
    for row in current_field_help_rows(include_staging_only=False):
        help_sheet.append(list(row))
    for cell in help_sheet[1]:
        cell.font = Font(bold=True)

    return workbook


def create_staging_workbook(building_rows: Optional[Sequence[dict]] = None) -> Workbook:
    workbook = create_standard_master_workbook()
    main_sheet = workbook[MASTER_MAIN_SHEET]
    if STAGING_STATUS_HEADER not in _extract_headers(main_sheet):
        column = main_sheet.max_column + 1
        main_sheet.cell(row=1, column=column, value=STAGING_STATUS_HEADER)
        main_sheet.cell(row=1, column=column).font = Font(bold=True)
    for snapshot in building_rows or []:
        main_sheet.append(build_staging_sheet_row(snapshot))
    return workbook


def _extract_headers(sheet) -> List[str]:
    return [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]


def _upgrade_legacy_workbook_labels(workbook: Workbook) -> bool:
    changed = False
    for legacy_name, english_name in LEGACY_SHEET_NAMES.items():
        if legacy_name in workbook.sheetnames and english_name not in workbook.sheetnames:
            workbook[legacy_name].title = english_name
            changed = True
    if MASTER_MAIN_SHEET in workbook.sheetnames:
        main_sheet = workbook[MASTER_MAIN_SHEET]
        for cell in main_sheet[1]:
            current = str(cell.value).strip() if cell.value is not None else ""
            replacement = LEGACY_HEADER_TO_ENGLISH.get(current)
            if replacement and replacement != current:
                cell.value = replacement
                changed = True
    return changed


def _normalized_identity(name: object, address: object) -> Tuple[str, str]:
    return (
        (normalize_unknown_value(name) or "").strip().lower(),
        (normalize_unknown_value(address) or "").strip().lower(),
    )


def validate_master_workbook(path: Path) -> dict:
    workbook = load_workbook(path)
    sheet_names = workbook.sheetnames
    missing_sheets = [
        sheet_name
        for sheet_name in (MASTER_MAIN_SHEET, MASTER_PLAN_SHEET, MASTER_HELP_SHEET)
        if sheet_name not in sheet_names
    ]

    main_headers: List[str] = []
    duplicate_headers: List[str] = []
    expected_headers = current_master_headers()
    missing_headers: List[str] = list(expected_headers)
    merged_ranges: List[str] = []
    row_count = 0

    if MASTER_MAIN_SHEET in sheet_names:
        main_sheet = workbook[MASTER_MAIN_SHEET]
        main_headers = _extract_headers(main_sheet)
        seen = set()
        for header in main_headers:
            if not header:
                continue
            if header in seen and header not in duplicate_headers:
                duplicate_headers.append(header)
            seen.add(header)
        missing_headers = [header for header in expected_headers if header not in main_headers]
        merged_ranges = [str(item) for item in main_sheet.merged_cells.ranges]
        row_count = max(main_sheet.max_row - 1, 0)

    workbook.close()
    ok = not missing_sheets and not duplicate_headers and not missing_headers and not merged_ranges
    return {
        "ok": ok,
        "path": str(path),
        "sheet_names": sheet_names,
        "missing_sheets": missing_sheets,
        "main_headers": main_headers,
        "missing_headers": missing_headers,
        "duplicate_headers": duplicate_headers,
        "merged_ranges": merged_ranges,
        "row_count": row_count,
    }


def validate_staging_workbook(path: Path) -> dict:
    workbook = load_workbook(path)
    sheet_names = workbook.sheetnames
    missing_sheets = [
        sheet_name
        for sheet_name in (MASTER_MAIN_SHEET, MASTER_PLAN_SHEET, MASTER_HELP_SHEET)
        if sheet_name not in sheet_names
    ]

    main_headers: List[str] = []
    duplicate_headers: List[str] = []
    expected_headers = current_staging_headers()
    missing_headers: List[str] = list(expected_headers)
    merged_ranges: List[str] = []
    row_count = 0

    if MASTER_MAIN_SHEET in sheet_names:
        main_sheet = workbook[MASTER_MAIN_SHEET]
        main_headers = _extract_headers(main_sheet)
        seen = set()
        for header in main_headers:
            if not header:
                continue
            if header in seen and header not in duplicate_headers:
                duplicate_headers.append(header)
            seen.add(header)
        missing_headers = [header for header in expected_headers if header not in main_headers]
        merged_ranges = [str(item) for item in main_sheet.merged_cells.ranges]
        row_count = max(main_sheet.max_row - 1, 0)

    workbook.close()
    ok = not missing_sheets and not duplicate_headers and not missing_headers and not merged_ranges
    return {
        "ok": ok,
        "path": str(path),
        "sheet_names": sheet_names,
        "missing_sheets": missing_sheets,
        "main_headers": main_headers,
        "missing_headers": missing_headers,
        "duplicate_headers": duplicate_headers,
        "merged_ranges": merged_ranges,
        "row_count": row_count,
    }


def write_workbook_atomically(workbook: Workbook, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    temp_file = tempfile.NamedTemporaryFile(
        prefix=f".{target_path.stem}_",
        suffix=target_path.suffix or ".xlsx",
        dir=target_path.parent,
        delete=False,
    )
    temp_path = Path(temp_file.name)
    temp_file.close()
    try:
        workbook.save(temp_path)
        os.replace(temp_path, target_path)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def upgrade_master_workbook(path: Path) -> bool:
    if not path.exists():
        return False
    workbook = load_workbook(path)
    changed = _upgrade_legacy_workbook_labels(workbook)

    if MASTER_MAIN_SHEET in workbook.sheetnames:
        main_sheet = workbook[MASTER_MAIN_SHEET]
        headers = _extract_headers(main_sheet)
        for header in current_master_headers():
            if header in headers:
                continue
            column = main_sheet.max_column + 1
            main_sheet.cell(row=1, column=column, value=header)
            main_sheet.cell(row=1, column=column).font = Font(bold=True)
            changed = True

    if MASTER_PLAN_SHEET not in workbook.sheetnames:
        package_sheet = workbook.create_sheet(MASTER_PLAN_SHEET)
        package_sheet.append(PACKAGE_CONFIRM_HEADERS)
        for cell in package_sheet[1]:
            cell.font = Font(bold=True)
        changed = True

    if MASTER_HELP_SHEET not in workbook.sheetnames:
        help_sheet = workbook.create_sheet(MASTER_HELP_SHEET)
        for row in current_field_help_rows(include_staging_only=False):
            help_sheet.append(list(row))
        for cell in help_sheet[1]:
            cell.font = Font(bold=True)
        changed = True
    else:
        help_sheet = workbook[MASTER_HELP_SHEET]
        existing_labels = {
            str(help_sheet.cell(row=row_index, column=1).value).strip()
            for row_index in range(1, help_sheet.max_row + 1)
            if help_sheet.cell(row=row_index, column=1).value
        }
        for row in current_field_help_rows(include_staging_only=False)[1:]:
            if row[0] in existing_labels:
                continue
            help_sheet.append(list(row))
            changed = True

    if changed:
        write_workbook_atomically(workbook, path)
    workbook.close()
    return changed


def ensure_master_workbook(building_rows: Optional[Sequence[dict]] = None) -> Path:
    workbook_path = resolve_master_excel_path()
    if workbook_path.exists():
        upgrade_master_workbook(workbook_path)
        return workbook_path
    workbook = create_standard_master_workbook(building_rows)
    write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path


def upgrade_staging_workbook(path: Path) -> bool:
    if not path.exists():
        return False
    workbook = load_workbook(path)
    changed = _upgrade_legacy_workbook_labels(workbook)

    if MASTER_MAIN_SHEET in workbook.sheetnames:
        main_sheet = workbook[MASTER_MAIN_SHEET]
        headers = _extract_headers(main_sheet)
        for header in current_staging_headers():
            if header in headers:
                continue
            column = main_sheet.max_column + 1
            main_sheet.cell(row=1, column=column, value=header)
            main_sheet.cell(row=1, column=column).font = Font(bold=True)
            changed = True

    if MASTER_PLAN_SHEET not in workbook.sheetnames:
        package_sheet = workbook.create_sheet(MASTER_PLAN_SHEET)
        package_sheet.append(PACKAGE_CONFIRM_HEADERS)
        for cell in package_sheet[1]:
            cell.font = Font(bold=True)
        changed = True

    if MASTER_HELP_SHEET not in workbook.sheetnames:
        help_sheet = workbook.create_sheet(MASTER_HELP_SHEET)
        for row in current_field_help_rows(include_staging_only=True):
            help_sheet.append(list(row))
        for cell in help_sheet[1]:
            cell.font = Font(bold=True)
        changed = True
    else:
        help_sheet = workbook[MASTER_HELP_SHEET]
        existing_labels = {
            str(help_sheet.cell(row=row_index, column=1).value).strip()
            for row_index in range(1, help_sheet.max_row + 1)
            if help_sheet.cell(row=row_index, column=1).value
        }
        for row in current_field_help_rows(include_staging_only=True)[1:]:
            if row[0] in existing_labels:
                continue
            help_sheet.append(list(row))
            changed = True

    if changed:
        write_workbook_atomically(workbook, path)
    workbook.close()
    return changed


def ensure_staging_workbook(building_rows: Optional[Sequence[dict]] = None) -> Path:
    workbook_path = resolve_staging_excel_path()
    if workbook_path.exists():
        upgrade_staging_workbook(workbook_path)
        return workbook_path
    workbook = create_staging_workbook(building_rows)
    write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path


def upsert_building_snapshot(
    snapshot: dict,
    *,
    lookup_keys: Optional[Iterable[Tuple[str, str]]] = None,
) -> Path:
    workbook_path = ensure_master_workbook()
    validation = validate_master_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError(
            "Master workbook validation failed: "
            f"missing sheets={validation['missing_sheets']}; "
            f"missing headers={validation['missing_headers']}; "
            f"duplicate headers={validation['duplicate_headers']}; "
            f"merged ranges={validation['merged_ranges']}"
        )

    workbook = load_workbook(workbook_path)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    column_by_header = {header: index + 1 for index, header in enumerate(headers) if header}

    target_identities = []
    current_identity = _normalized_identity(
        snapshot.get("building_name"),
        snapshot.get("address"),
    )
    if current_identity[0]:
        target_identities.append(current_identity)
    for item in lookup_keys or []:
        identity = _normalized_identity(item[0], item[1])
        if identity[0] and identity not in target_identities:
            target_identities.append(identity)

    target_row = None
    for row_index in range(2, main_sheet.max_row + 1):
        row_identity = _normalized_identity(
            main_sheet.cell(row=row_index, column=column_by_header["Building Name"]).value,
            main_sheet.cell(row=row_index, column=column_by_header["Address"]).value,
        )
        if row_identity in target_identities:
            target_row = row_index
            break

    if target_row is None:
        target_row = main_sheet.max_row + 1

    row_values = build_main_sheet_row(snapshot)
    for header, value in zip(current_master_headers(), row_values):
        main_sheet.cell(row=target_row, column=column_by_header[header], value=value)

    write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path


def upsert_staging_snapshot(
    snapshot: dict,
    *,
    lookup_keys: Optional[Iterable[Tuple[str, str]]] = None,
) -> Path:
    workbook_path = ensure_staging_workbook()
    validation = validate_staging_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError(
            "Staging workbook validation failed: "
            f"missing sheets={validation['missing_sheets']}; "
            f"missing headers={validation['missing_headers']}; "
            f"duplicate headers={validation['duplicate_headers']}; "
            f"merged ranges={validation['merged_ranges']}"
        )

    workbook = load_workbook(workbook_path)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    column_by_header = {header: index + 1 for index, header in enumerate(headers) if header}

    target_identities = []
    current_identity = _normalized_identity(
        snapshot.get("building_name"),
        snapshot.get("address"),
    )
    if current_identity[0]:
        target_identities.append(current_identity)
    for item in lookup_keys or []:
        identity = _normalized_identity(item[0], item[1])
        if identity[0] and identity not in target_identities:
            target_identities.append(identity)

    target_row = None
    for row_index in range(2, main_sheet.max_row + 1):
        row_identity = _normalized_identity(
            main_sheet.cell(row=row_index, column=column_by_header["Building Name"]).value,
            main_sheet.cell(row=row_index, column=column_by_header["Address"]).value,
        )
        if row_identity in target_identities:
            target_row = row_index
            break

    if target_row is None:
        target_row = main_sheet.max_row + 1

    row_values = build_staging_sheet_row(snapshot)
    for header, value in zip(current_staging_headers(), row_values):
        if header not in column_by_header:
            continue
        main_sheet.cell(row=target_row, column=column_by_header[header], value=value)

    write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path


def delete_building_snapshot(building_name: str, address: str = "") -> Path:
    workbook_path = ensure_master_workbook()
    validation = validate_master_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The Master workbook structure is invalid; the building row cannot be deleted.")

    workbook = load_workbook(workbook_path)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    column_by_header = {header: index + 1 for index, header in enumerate(headers) if header}
    target_identity = _normalized_identity(building_name, address)

    target_row = None
    for row_index in range(2, main_sheet.max_row + 1):
        row_identity = _normalized_identity(
            main_sheet.cell(row=row_index, column=column_by_header["Building Name"]).value,
            main_sheet.cell(row=row_index, column=column_by_header["Address"]).value,
        )
        if row_identity == target_identity:
            target_row = row_index
            break

    if target_row is not None:
        main_sheet.delete_rows(target_row, 1)

    write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path


def load_master_workbook_rows(path: Optional[Path] = None) -> List[dict]:
    workbook_path = path or ensure_master_workbook()
    upgrade_master_workbook(workbook_path)
    validation = validate_master_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The Master workbook structure is invalid and cannot be read.")

    workbook = load_workbook(workbook_path, data_only=True)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    header_to_field_key = current_header_to_field_key(include_staging_only=False)
    entries_by_field = {item["field_key"]: item for item in _catalog_entries_for_excel(include_staging_only=False)}
    rows: List[dict] = []

    for row_index in range(2, main_sheet.max_row + 1):
        raw_by_header = {
            header: main_sheet.cell(row=row_index, column=column_index + 1).value
            for column_index, header in enumerate(headers)
            if header
        }
        building_name = normalize_unknown_value(raw_by_header.get("Building Name"))
        if not building_name:
            continue

        snapshot = {"building_name": building_name, "extensions": {}, "_row_number": row_index}
        for header, raw_value in raw_by_header.items():
            field_key = header_to_field_key.get(header)
            if not field_key:
                continue
            entry = entries_by_field.get(field_key) or {}
            field_type = entry.get("field_type") or "text"
            if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
                normalized = normalize_field_value(field_key, "boolean", raw_value)
            elif field_key in INSURANCE_STATUS_FIELD_KEYS:
                normalized = insurance_status_from_excel_value(raw_value)
            elif field_key == "internet_provider":
                normalized = normalize_provider_text(raw_value)
            else:
                normalized = normalize_field_value(field_key, field_type, raw_value)
            if field_key in CORE_SNAPSHOT_FIELD_KEYS:
                snapshot[field_key] = normalized
            else:
                snapshot["extensions"][field_key] = normalized
        rows.append(snapshot)

    workbook.close()
    return rows


def load_staging_workbook_rows(path: Optional[Path] = None) -> List[dict]:
    workbook_path = path or ensure_staging_workbook()
    upgrade_staging_workbook(workbook_path)
    validation = validate_staging_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The Staging workbook structure is invalid and cannot be read.")

    workbook = load_workbook(workbook_path, data_only=True)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    header_to_field_key = current_header_to_field_key(include_staging_only=True)
    entries_by_field = {item["field_key"]: item for item in _catalog_entries_for_excel(include_staging_only=True)}
    rows: List[dict] = []

    for row_index in range(2, main_sheet.max_row + 1):
        raw_by_header = {
            header: main_sheet.cell(row=row_index, column=column_index + 1).value
            for column_index, header in enumerate(headers)
            if header
        }
        building_name = normalize_unknown_value(raw_by_header.get("Building Name"))
        if not building_name:
            continue

        snapshot = {
            "building_name": building_name,
            "extensions": {},
            "library_status": normalize_staging_status(raw_by_header.get(STAGING_STATUS_HEADER)),
            "_row_number": row_index,
        }
        for header, raw_value in raw_by_header.items():
            field_key = header_to_field_key.get(header)
            if not field_key:
                continue
            entry = entries_by_field.get(field_key) or {}
            field_type = entry.get("field_type") or "text"
            if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
                normalized = normalize_field_value(field_key, "boolean", raw_value)
            elif field_key in INSURANCE_STATUS_FIELD_KEYS:
                normalized = insurance_status_from_excel_value(raw_value)
            elif field_key == "internet_provider":
                normalized = normalize_provider_text(raw_value)
            else:
                normalized = normalize_field_value(field_key, field_type, raw_value)
            if field_key in CORE_SNAPSHOT_FIELD_KEYS:
                snapshot[field_key] = normalized
            else:
                snapshot["extensions"][field_key] = normalized
        rows.append(snapshot)

    workbook.close()
    return rows


def sync_staging_statuses(status_by_identity: Dict[Tuple[str, str], str]) -> Path:
    workbook_path = ensure_staging_workbook()
    validation = validate_staging_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The Staging workbook structure is invalid; review statuses cannot be written.")

    workbook = load_workbook(workbook_path)
    main_sheet = workbook[MASTER_MAIN_SHEET]
    headers = _extract_headers(main_sheet)
    column_by_header = {header: index + 1 for index, header in enumerate(headers) if header}
    status_column = column_by_header.get(STAGING_STATUS_HEADER)
    changed = False

    if status_column is not None:
        for row_index in range(2, main_sheet.max_row + 1):
            identity = _normalized_identity(
                main_sheet.cell(row=row_index, column=column_by_header["Building Name"]).value,
                main_sheet.cell(row=row_index, column=column_by_header["Address"]).value,
            )
            target_status = normalize_staging_status(status_by_identity.get(identity))
            current_status = normalize_staging_status(main_sheet.cell(row=row_index, column=status_column).value)
            if target_status != current_status:
                main_sheet.cell(row=row_index, column=status_column, value=staging_status_to_excel_text(target_status))
                changed = True

    if changed:
        write_workbook_atomically(workbook, workbook_path)
    workbook.close()
    return workbook_path
