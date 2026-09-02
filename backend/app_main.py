from __future__ import annotations

import asyncio
import base64
import hashlib
import io
import json
import os
import re
import secrets
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import time
import urllib.error
import urllib.request
import uuid
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Literal, Optional, Union
from urllib.parse import quote

from dotenv import load_dotenv
from fastapi import BackgroundTasks, Depends, FastAPI, File, Form, Header, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openai import AsyncOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PyPDF2 import PdfReader
from pydantic import BaseModel, Field

try:
    import fitz  # type: ignore
except ImportError:
    fitz = None

import legacy_demo as legacy
from kb_db import (
    DB_PATH,
    DATA_DIR,
    UPLOAD_ROOT,
    connect_db,
    db_connection,
    default_field_metadata,
    init_db,
    load_field_catalog,
    reset_standard_field_catalog,
    slugify_field_key,
)
from kb_master_excel import (
    MASTER_MAIN_SHEET,
    MASTER_PLAN_SHEET,
    MASTER_HELP_SHEET,
    NETWORK_PROVIDER_FIELD_MAP,
    NETWORK_PROVIDER_NOTE_FIELD_MAP,
    NETWORK_PLAN_FIELD_MAP,
    STAGING_STATUS_ACTIVE,
    STAGING_STATUS_HEADER,
    STAGING_STATUS_MASTERED,
    STAGING_STATUS_PENDING,
    delete_building_snapshot,
    ensure_master_workbook,
    ensure_staging_workbook,
    load_master_workbook_rows,
    load_staging_workbook_rows,
    resolve_master_excel_path,
    resolve_staging_excel_path,
    sync_staging_statuses,
    upsert_staging_snapshot,
    upsert_building_snapshot,
    upgrade_master_workbook,
    validate_master_workbook,
    validate_staging_workbook,
)
from kb_security import hash_password, hash_token, make_session_token, utc_now_iso, verify_password
from kb_unknowns import (
    display_value_or_unknown,
    extract_provider_names,
    normalize_booleanish,
    normalize_field_value,
    normalize_provider_name,
    normalize_provider_text,
    normalize_requirement_choice,
    normalize_unknown_value,
)
from ocr_services import (
    BaiduUnlimitedCloudOcrProvider,
    CallableVisionFieldExtractor,
    LocalOcrProvider,
    OcrResult,
    OcrRouter,
    UnlimitedOcrLocalHttpProvider,
)


BACKEND_DIR = Path(__file__).resolve().parent
DOTENV_PATH = BACKEND_DIR / ".env"
DOTENV_LOCAL_PATH = BACKEND_DIR / ".env.local"
load_dotenv(dotenv_path=DOTENV_PATH)
load_dotenv(dotenv_path=DOTENV_LOCAL_PATH, override=True)

FRONTEND_DIST_DIR = Path(os.getenv("FRONTEND_DIST_DIR", "../frontend/dist"))
if not FRONTEND_DIST_DIR.is_absolute():
    FRONTEND_DIST_DIR = (BACKEND_DIR / FRONTEND_DIST_DIR).resolve()
FRONTEND_INDEX_PATH = FRONTEND_DIST_DIR / "index.html"
FRONTEND_ASSETS_DIR = FRONTEND_DIST_DIR / "assets"
RUNTIME_MODE = (os.getenv("WHITEPAPER_RUNTIME_MODE", "dev") or "dev").strip().lower() or "dev"
RUNTIME_STATUS_PATH = Path(
    os.getenv("WHITEPAPER_RUNTIME_STATUS_PATH", str(DATA_DIR / "runtime_status.json"))
).resolve()
APP_DIR = BACKEND_DIR.parent
UPDATE_SCRIPT_PATH = Path(os.getenv("WHITEPAPER_UPDATE_SCRIPT", str(APP_DIR / "scripts" / "update_from_git.sh"))).resolve()
UPDATE_LOG_DIR = Path(os.getenv("WHITEPAPER_UPDATE_LOG_DIR", str(DATA_DIR / "update_logs"))).resolve()
UPDATE_RESTART_MARKER_PATH = Path(
    os.getenv("WHITEPAPER_UPDATE_RESTART_MARKER", str(DATA_DIR / "restart_requested"))
).resolve()
TUNNEL_HEALTH_URL = (os.getenv("WHITEPAPER_TUNNEL_HEALTH_URL", "") or "").strip()
TUNNEL_NAME = (os.getenv("WHITEPAPER_TUNNEL_NAME", "") or "").strip()
PUBLIC_BIND_HOST = (os.getenv("WHITEPAPER_BIND_HOST", "") or "").strip()
ALLOW_DEFAULT_PASSWORD_ON_LAN = (
    (os.getenv("WHITEPAPER_ALLOW_DEFAULT_PASSWORD_ON_LAN", "") or "").strip() == "1"
    or (os.getenv("WHITEPAPER_ALLOW_DEFAULT_PASSWORD_ON_LAN_RUNTIME", "") or "").strip() == "1"
)
MAX_UPLOAD_BYTES = int((os.getenv("WHITEPAPER_MAX_UPLOAD_BYTES", "52428800") or "52428800").strip())
MAX_IMAGE_UPLOAD_FILES = int((os.getenv("WHITEPAPER_MAX_IMAGE_UPLOAD_FILES", "12") or "12").strip())
MAX_PDF_UPLOAD_FILES = int((os.getenv("WHITEPAPER_MAX_PDF_UPLOAD_FILES", "6") or "6").strip())
PDF_RENDER_TIMEOUT_SECONDS = int((os.getenv("WHITEPAPER_PDF_RENDER_TIMEOUT_SECONDS", "60") or "60").strip())
OCR_TIMEOUT_SECONDS = int((os.getenv("WHITEPAPER_OCR_TIMEOUT_SECONDS", "45") or "45").strip())
VISION_TIMEOUT_SECONDS = int((os.getenv("WHITEPAPER_VISION_TIMEOUT_SECONDS", "45") or "45").strip())
APPLE_VISION_OCR_TIMEOUT_SECONDS = int(
    (os.getenv("WHITEPAPER_APPLE_VISION_OCR_TIMEOUT_SECONDS", "60") or "60").strip()
)
MAX_PDF_PARSE_PAGES = int((os.getenv("WHITEPAPER_MAX_PDF_PARSE_PAGES", "30") or "30").strip())
OCR_PROVIDER_ALIASES = {"baidu_unlimited": "baidu_unlimited_cloud"}
OCR_PROVIDER = OCR_PROVIDER_ALIASES.get(
    (os.getenv("OCR_PROVIDER", "local") or "local").strip().lower(),
    (os.getenv("OCR_PROVIDER", "local") or "local").strip().lower(),
)
OCR_FALLBACK_PROVIDER = OCR_PROVIDER_ALIASES.get(
    (os.getenv("OCR_FALLBACK_PROVIDER", "local") or "local").strip().lower(),
    (os.getenv("OCR_FALLBACK_PROVIDER", "local") or "local").strip().lower(),
)
BAIDU_OCR_API_KEY = (os.getenv("BAIDU_OCR_API_KEY", "") or "").strip()
BAIDU_OCR_SECRET_KEY = (os.getenv("BAIDU_OCR_SECRET_KEY", "") or "").strip()
BAIDU_OCR_BASE_URL = (os.getenv("BAIDU_OCR_BASE_URL", "https://aip.baidubce.com") or "https://aip.baidubce.com").strip()
BAIDU_OCR_POLL_INTERVAL_SECONDS = int((os.getenv("BAIDU_OCR_POLL_INTERVAL_SECONDS", "5") or "5").strip())
BAIDU_OCR_TIMEOUT_SECONDS = int((os.getenv("BAIDU_OCR_TIMEOUT_SECONDS", "300") or "300").strip())
UNLIMITED_OCR_LOCAL_BASE_URL = (
    os.getenv("UNLIMITED_OCR_LOCAL_BASE_URL", os.getenv("UNLIMITED_OCR_BASE_URL", "")) or ""
).strip()
UNLIMITED_OCR_LOCAL_MODEL = (
    os.getenv("UNLIMITED_OCR_LOCAL_MODEL", "baidu/Unlimited-OCR") or "baidu/Unlimited-OCR"
).strip()
UNLIMITED_OCR_LOCAL_API_KEY = (os.getenv("UNLIMITED_OCR_LOCAL_API_KEY", "") or "").strip()
UNLIMITED_OCR_LOCAL_TIMEOUT_SECONDS = int(
    (os.getenv("UNLIMITED_OCR_LOCAL_TIMEOUT_SECONDS", "1200") or "1200").strip()
)
UNLIMITED_OCR_LOCAL_MAX_TOKENS = int(
    (os.getenv("UNLIMITED_OCR_LOCAL_MAX_TOKENS", "8192") or "8192").strip()
)
AI_EXPLANATION_TIMEOUT_SECONDS = int((os.getenv("AI_EXPLANATION_TIMEOUT_SECONDS", "15") or "15").strip())
AI_EXPLANATION_MAX_TOKENS = min(
    600,
    max(100, int((os.getenv("AI_EXPLANATION_MAX_TOKENS", "300") or "300").strip())),
)

APPLE_VISION_OCR_SWIFT = r"""
import Foundation
import Vision
import ImageIO

if CommandLine.arguments.count < 2 {
    exit(2)
}

let url = URL(fileURLWithPath: CommandLine.arguments[1])
guard let source = CGImageSourceCreateWithURL(url as CFURL, nil),
      let image = CGImageSourceCreateImageAtIndex(source, 0, nil) else {
    exit(3)
}

let request = VNRecognizeTextRequest()
request.recognitionLevel = .accurate
request.usesLanguageCorrection = true
request.recognitionLanguages = ["en-US", "zh-Hans"]

let handler = VNImageRequestHandler(cgImage: image, options: [:])
do {
    try handler.perform([request])
    let lines = (request.results ?? []).compactMap { observation in
        observation.topCandidates(1).first?.string
    }
    print(lines.joined(separator: "\n"))
} catch {
    fputs(String(describing: error), stderr)
    exit(4)
}
"""
MAX_ANSWER_TOKENS_RAW = (os.getenv("MAX_ANSWER_TOKENS", "900") or "900").strip()
try:
    MAX_ANSWER_TOKENS = min(2000, max(300, int(MAX_ANSWER_TOKENS_RAW)))
except ValueError:
    MAX_ANSWER_TOKENS = 900
ALLOWED_EXCEL_UPLOAD_SUFFIXES = {".xlsx", ".xls"}
ALLOWED_PDF_UPLOAD_SUFFIXES = {".pdf"}
ALLOWED_IMAGE_UPLOAD_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
INTAKE_MODE_FULL_PACKAGE = "full_package"
INTAKE_MODE_SUPPLEMENT = "supplement"
INTAKE_MODES = {INTAKE_MODE_FULL_PACKAGE, INTAKE_MODE_SUPPLEMENT}
SUPPLEMENT_SCOPE_INSURANCE = "insurance"
SUPPLEMENT_SCOPE_ELECTRICITY = "electricity"
SUPPLEMENT_SCOPE_INTERNET = "internet"
SUPPLEMENT_SCOPE_MOVE_IN = "move_in"
SUPPLEMENT_SCOPE_ALL = "all"
SUPPLEMENT_SCOPES = {
    SUPPLEMENT_SCOPE_INSURANCE,
    SUPPLEMENT_SCOPE_ELECTRICITY,
    SUPPLEMENT_SCOPE_INTERNET,
    SUPPLEMENT_SCOPE_MOVE_IN,
    SUPPLEMENT_SCOPE_ALL,
}
BUILTIN_DEFAULT_USER_PASSWORDS = {
    "superadmin": (os.getenv("WHITEPAPER_SUPERADMIN_PASSWORD", "") or "").strip(),
    "admin": (os.getenv("WHITEPAPER_ADMIN_PASSWORD", "") or "").strip(),
    "employee": (os.getenv("WHITEPAPER_EMPLOYEE_PASSWORD", "") or "").strip(),
    "viewer": (os.getenv("WHITEPAPER_VIEWER_PASSWORD", "") or "").strip(),
}
LOGIN_FAILURE_WINDOW_SECONDS = 10 * 60
LOGIN_FAILURE_LIMIT = 8
LOGIN_FAILURES: Dict[str, List[float]] = {}

DEFAULT_CORS_ALLOW_ORIGINS = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]

CORE_MASTER_FIELD_KEYS = {
    "building_name",
    "address",
    "insurance_required",
    "insurance_coverage_amount",
    "electricity_required",
    "electricity_provider",
    "internet_self_setup_required",
    "internet_provider",
    "internet_notes",
    "move_in_notes",
    "source_type",
    "source_file",
    "source_date",
    "info_cutoff_date",
}

DETAILED_INSURANCE_FIELD_KEYS = {
    "document_type",
    "insurance_renters_required",
    "insurance_renters_minimum_coverage",
    "insurance_personal_property_required",
    "insurance_personal_property_minimum",
    "insurance_personal_liability_required",
    "insurance_personal_liability_per_occurrence",
    "insurance_personal_liability_aggregate",
    "insurance_coi_required",
    "insurance_coi_trigger",
    "insurance_interested_party_required",
    "insurance_additional_insured_required",
    "insurance_certificate_holder_required",
    "insurance_submission_method",
    "insurance_recipient",
    "insurance_alternative_program_or_penalty",
}

MOVE_IN_STRUCTURED_FIELD_KEYS = {
    "key_pickup_notes",
    "service_elevator_booking_notes",
}

BUILDING_CONTACT_FIELD_KEYS = {
    "building_management_contact",
    "building_front_desk_contact",
    "building_maintenance_contact",
}

WORKBOOK_EXTENSION_FIELD_KEYS = (
    *NETWORK_PROVIDER_FIELD_MAP.keys(),
    *NETWORK_PLAN_FIELD_MAP.keys(),
    *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
    *BUILDING_CONTACT_FIELD_KEYS,
    "insurance_coi_required",
    "insurance_coi_trigger",
    "key_pickup_notes",
    "service_elevator_booking_notes",
)

AI_WRITABLE_FIELD_KEYS = {
    "building_name",
    "address",
    "insurance_required",
    "insurance_coverage_amount",
    *DETAILED_INSURANCE_FIELD_KEYS,
    "electricity_required",
    "electricity_provider",
    "internet_self_setup_required",
    "internet_provider",
    "internet_notes",
    *NETWORK_PROVIDER_FIELD_MAP.keys(),
    *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
    *MOVE_IN_STRUCTURED_FIELD_KEYS,
    *BUILDING_CONTACT_FIELD_KEYS,
    "move_in_notes",
    "source_date",
    "info_cutoff_date",
    "source_type",
    "source_file",
}

SUPPLEMENT_IDENTITY_FIELD_KEYS = {"building_name", "address"}
SUPPLEMENT_SCOPE_LABELS = {
    SUPPLEMENT_SCOPE_INSURANCE: "insurance",
    SUPPLEMENT_SCOPE_ELECTRICITY: "electricity",
    SUPPLEMENT_SCOPE_INTERNET: "internet",
    SUPPLEMENT_SCOPE_MOVE_IN: "move-in",
    SUPPLEMENT_SCOPE_ALL: "full building review",
}
SUPPLEMENT_SCOPE_FIELD_KEYS = {
    SUPPLEMENT_SCOPE_INSURANCE: {
        "insurance_required",
        "insurance_coverage_amount",
        *DETAILED_INSURANCE_FIELD_KEYS,
    },
    SUPPLEMENT_SCOPE_ELECTRICITY: {
        "electricity_required",
        "electricity_provider",
    },
    SUPPLEMENT_SCOPE_INTERNET: {
        "internet_self_setup_required",
        "internet_provider",
        "internet_notes",
        *NETWORK_PROVIDER_FIELD_MAP.keys(),
        *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
        *NETWORK_PLAN_FIELD_MAP.keys(),
    },
    SUPPLEMENT_SCOPE_MOVE_IN: {
        *MOVE_IN_STRUCTURED_FIELD_KEYS,
        "move_in_notes",
    },
}
SUPPLEMENT_SCOPE_FIELD_KEYS[SUPPLEMENT_SCOPE_ALL] = AI_WRITABLE_FIELD_KEYS - SUPPLEMENT_IDENTITY_FIELD_KEYS

CHAT_BUILDING_FACT_FIELD_KEYS = {
    "insurance_required",
    "insurance_coverage_amount",
    *DETAILED_INSURANCE_FIELD_KEYS,
    "electricity_required",
    "electricity_provider",
    "internet_self_setup_required",
    "internet_provider",
    "internet_notes",
    *NETWORK_PROVIDER_FIELD_MAP.keys(),
    *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
    *MOVE_IN_STRUCTURED_FIELD_KEYS,
    *BUILDING_CONTACT_FIELD_KEYS,
    "move_in_notes",
}

FIELD_GROUP_KEYWORDS = {
    "insurance": {
        "保险",
        "insurance",
        "coverage",
        "保额",
        "renters",
        "liability",
        "coi",
        "certificate of insurance",
        "interested party",
        "additional insured",
        "additional interest",
    },
    "electric": {"开电", "electric", "electricity", "电力", "电表", "coned", "pseg"},
    "internet": {
        "网络",
        "wifi",
    "internet",
    "宽带",
    "开网",
    "套餐",
    "档位",
    "speed",
    "mbps",
    "gig",
    "plan",
    "运营商",
    "verizon",
    "xfinity",
        "astound",
        "spectrum",
        "honest",
    },
    "move_in": {"入住", "move in", "搬家", "搬入", "钥匙", "货梯", "elevator", "rello"},
    "contacts": {
        "联系方式",
        "联系人",
        "front desk",
        "concierge",
        "management",
        "property manager",
        "maintenance",
        "service request",
    },
    "summary": {"概况", "主要", "summary", "简介", "内容"},
}

FIELD_GROUP_OPTIONS = {"basic", "insurance", "electricity", "internet", "move_in", "contacts", "custom"}

INSURANCE_STATUS_LABELS = {
    "yes": "Yes",
    "no": "No",
    "optional": "Optional",
    "manual_review": "Manual Review",
}

STAGING_QUERY_STATUSES = (
    "pending",
    "ai_parsed",
    "employee_submitted",
    "needs_more_info",
    "conflict",
)

WELCOME_LETTER_FIELD_PROMPT = """You are a parser for an internal building-knowledge system. Extract only facts explicitly stated in the source. Never guess, complete missing facts from experience, or invent plans, prices, speeds, or recommended providers.

For insurance_required, electricity_required, and internet_self_setup_required, use true, false, or optional only when the source supports that choice. Keep Renters Insurance, Personal Property, Personal Liability, mover/delivery COI, Interested Party or Additional Interest, Additional Insured, and Certificate Holder separate. Omit fields that are not explicit. If the wording is ambiguous or combines incompatible concepts, use value=manual_review and preserve the supporting quote.

Return JSON only:
{
  "values": {
    "field_key": {"value": "concise English value", "confidence": 0.0, "evidence": "verbatim source quote"}
  }
}

Allowed fields:
building_name, address, insurance_required, insurance_coverage_amount, electricity_required,
electricity_provider, internet_self_setup_required, internet_provider, internet_notes,
key_pickup_notes, service_elevator_booking_notes, move_in_notes, source_date, info_cutoff_date, source_type, source_file,
building_management_contact, building_front_desk_contact, building_maintenance_contact,
document_type, insurance_renters_required, insurance_renters_minimum_coverage,
insurance_personal_property_required, insurance_personal_property_minimum,
insurance_personal_liability_required, insurance_personal_liability_per_occurrence,
insurance_personal_liability_aggregate, insurance_coi_required, insurance_coi_trigger,
insurance_interested_party_required, insurance_additional_insured_required,
insurance_certificate_holder_required, insurance_submission_method, insurance_recipient,
insurance_alternative_program_or_penalty
""".strip()

MOVE_IN_FIELD_PROMPT = """You classify move-in facts for an internal building-knowledge system. Use only the supplied source excerpts, which may be in English or Chinese. Do not guess.

Return only fields with direct evidence:
1. key_pickup_notes: where, from whom, when or under what conditions keys, mailbox keys, or fobs are collected; prerequisites and contact details. Mark any important missing slot as "Not stated in source."
2. service_elevator_booking_notes: how and where to book a service elevator or move-in appointment; recipient, email/phone/platform/form, contact, advance notice, time window, and COI requirement. Mark important missing slots as "Not stated in source."
3. move_in_notes: a complete operational checklist covering explicit prerequisites such as payment, forms, insurance, electricity, COI, elevators, and keys.

Do not treat lost/stolen-key rules, visitor registration, contact lists, or a bare mention of a concierge/front desk as key-pickup evidence. A source sentence may support both a focused field and move_in_notes, but each value must stay scoped to its field.

Return JSON only. Every value must include confidence and evidence_items with the original page and verbatim quote:
{"values":{"key_pickup_notes":{"value":"concise English value","confidence":0.0,"evidence_items":[{"page":1,"quote":"verbatim source"}],"manual_review_reason":""}}}
""".strip()

INSURANCE_FIELD_PROMPT = """You classify insurance facts for an internal building-knowledge system. Use only supplied source excerpts. Do not use general knowledge or guess.

Strictly distinguish renters insurance, personal property, personal liability, mover/delivery COI, Interested Party or Additional Interest, Additional Insured, and Certificate Holder. A mover COI page does not establish a tenant's renters-insurance limits. "Recommended but not required" means optional. The Guarantors, BuildingLink, Rello, email, and portals are submission or verification channels, not insurance products. Preserve the complete submission instruction when available.

Allowed fields:
insurance_renters_required, insurance_renters_minimum_coverage,
insurance_personal_property_required, insurance_personal_property_minimum,
insurance_personal_liability_required, insurance_personal_liability_per_occurrence,
insurance_personal_liability_aggregate, insurance_coi_required, insurance_coi_trigger,
insurance_interested_party_required, insurance_additional_insured_required,
insurance_certificate_holder_required, insurance_submission_method, insurance_recipient,
insurance_alternative_program_or_penalty

Omit unsupported fields. Use manual_review when the source combines concepts or does not support a reliable mapping. Return JSON only, with an English value, confidence, manual_review_reason, and evidence_items containing verbatim page quotes.
""".strip()

EVIDENCE_CLASSIFICATION_PROMPT = """Classify each supplied source excerpt by domain. Do not extract database fields and do not turn excerpts into final answers. Return every input id exactly once.

Allowed domains:
- renters_insurance: the tenant's own renters or tenant insurance.
- mover_coi: mover, delivery, move-in/move-out, or service-elevator Certificate of Insurance.
- electricity: electric utility, account setup, proof, provider, or included/sub-metered service.
- internet: Wi-Fi, internet, providers, service setup, SSID, or password.
- move_in: move-in process, appointments, forms, Rello, or checklist.
- key_pickup: explicit collection of keys, mailbox keys, or fobs.
- contact: stand-alone building, front desk, leasing, management, or maintenance contact.
- building_identity: building name, address, or explicit welcome-to identity.
- unknown: unrelated, table of contents, or insufficient context.

Keep mover COI requirements separate from renters insurance. Return JSON only:
{"items":[{"id":"p1_1","domain":"renters_insurance","confidence":0.0,"reason":"brief English reason"}]}
""".strip()

BUSINESS_SUMMARY_PROMPT = """Create a readable English Welcome Packet business summary using only the supplied excerpts. Do not create final database fields and do not use general knowledge.

Organize items into: building_identity, renters_insurance, moving_coi, internet, electricity, move_in_process, key_pickup, service_elevator, contacts, other_notes.

Every item must reference one evidence_id and return its original page and verbatim quote. The English summary value may paraphrase only facts in that quote. Use an empty array when a section has no evidence. Keep tenant renters insurance separate from mover/delivery COI. For internet and electricity, state only explicit included/setup/provider/SSID/password or account facts. Move-in is the broad checklist. Key pickup and service elevator are focused procedures. Contacts should separate management, front desk or concierge, and maintenance when the source permits. Optional details objects must remain fully grounded in the same quote.

Return JSON only:
{"business_summary":{"building_identity":[{"title":"Building Name","value":"English summary","evidence_id":"s1","page":1,"quote":"verbatim source","confidence":0.0}],"renters_insurance":[],"moving_coi":[],"internet":[],"electricity":[],"move_in_process":[],"key_pickup":[],"service_elevator":[],"contacts":[],"other_notes":[]}}
""".strip()

WELCOME_PACKET_ANALYSIS_PROMPT = """Analyze a Welcome Packet for an internal building-knowledge system. This pass creates an evidence-backed English business summary; it does not write to the database. Use only the supplied page text. Never guess or complete building rules from general knowledge.

1. document_classification must be one of full_welcome_letter, supplement, insurance_supplement, internet_supplement, electricity_supplement, coi_supplement, or unknown. For supplements, suggest a target building but do not claim it is already bound.
2. overall_summary is a short English human overview grounded in the source, not database evidence.
3. business_summary sections are building_identity, renters_insurance, moving_coi, internet, electricity, move_in_process, key_pickup, service_elevator, payments, contacts, other_notes.
4. workflow_hints may describe insurance, internet, electricity, and move_in facts, but must not change workflow state.

Every summary item requires id, English title, concise English value, page, verbatim quote, and confidence. Preserve the exact source language in quote. Keep tenant renters insurance separate from mover or delivery COI. Move-in is the broad checklist; key pickup and service elevator are focused procedures. Contacts should separate management, front desk or concierge, and maintenance where supported. Optional details must use only the same evidence. Use "Not stated in source" for a required procedural slot that the excerpt does not provide. Never turn a COI liability limit into renters-insurance coverage.

Return JSON only with document_classification, overall_summary, business_summary, workflow_hints, and validation_warnings.
""".strip()

WELCOME_PACKET_FIELD_MAPPING_PROMPT = """Map the supplied AI business summary and its source quotes to the allowed database fields. Do not re-summarize the whole document, use general knowledge, or output fields outside allowed_fields.

Each mapped field requires value, confidence, summary_item_id, evidence_items, and manual_review_reason. evidence_items.quote must be a verbatim source quote, never the summary value. overall_summary may provide context but is not evidence. Omit a field without direct evidence. Use true, false, or manual_review for booleans. Values must be concise operational English; preserve original text only in evidence quotes.

Keep renters-insurance fields separate from mover COI. COI trigger should contain only the triggering condition. Key pickup, service-elevator booking, and building contact fields may use only their corresponding grounded summary sections. move_in_notes may combine the grounded move-in, key, and elevator checklist. Remove OCR list-number artifacts from contact values. Use "Not stated in source" for important missing procedural slots rather than guessing.

Return JSON only:
{"values":{"field_key":{"value":"concise English value","confidence":0.0,"summary_item_id":"bs1","evidence_items":[{"page":1,"quote":"verbatim source"}],"manual_review_reason":""}},"validation_warnings":[]}
""".strip()

ELECTRICITY_FIELD_PROMPT = """Classify electricity-account facts using only the supplied source excerpts. Do not guess.

Return only:
1. electricity_required: true, false, or manual_review. Instructions to contact a utility, open an account, provide an account number, or provide proof support true. Explicitly included electricity supports false. General utility references do not establish a requirement.
2. electricity_provider: only a utility explicitly named in the source, such as Con Edison or PSE&G.

Every field requires confidence, manual_review_reason, and evidence_items with the original page and verbatim quote. Return JSON only.
""".strip()

ADDRESS_SUFFIX_PATTERN = (
    r"(?:Street|St|Avenue|Ave|Road|Rd|Drive|Dr|Lane|Ln|Boulevard|Blvd|Place|Pl|Court|Ct|Way|"
    r"Terrace|Ter|Parkway|Pkwy)"
)

FACT_EXPLANATION_PROMPT = """You explain structured records for an internal building-knowledge system in concise, natural English. Do not add any fact.

Use only the supplied structured fields, unknown markers, and pending-review notice. Never draw facts from the review queue, historical PDFs, network spreadsheets, general knowledge, or guesses. Do not invent internet plans, prices, speeds, or recommended providers.

If source_mode=master, call the information a "verified Master record." If source_mode=staging, call it a "current Staging record" and state that it is not yet a verified Master fact. If source_mode=crm, call it a "current CRM Case snapshot" and state that it supports staff execution rather than representing live building truth. Preserve every boundary such as Unknown, Staging view, pending review, or snapshot may need refresh.
""".strip()


def _parse_cors_allow_origins() -> List[str]:
    raw = (os.getenv("CORS_ALLOW_ORIGINS", "") or "").strip()
    if not raw:
        return DEFAULT_CORS_ALLOW_ORIGINS
    items = [item.strip() for item in raw.split(",") if item.strip()]
    return items or DEFAULT_CORS_ALLOW_ORIGINS


app = FastAPI(title="Whitepaper Internal Knowledge System")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_parse_cors_allow_origins(),
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=False,
)

if FRONTEND_ASSETS_DIR.is_dir():
    app.mount(
        "/assets",
        StaticFiles(directory=str(FRONTEND_ASSETS_DIR), check_dir=False),
        name="frontend-assets",
    )


class LoginRequest(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=1)


class PasswordChangeRequest(BaseModel):
    current_password: str = Field(min_length=1)
    new_password: str = Field(min_length=8)


class AdminUserCreateRequest(BaseModel):
    username: str = Field(min_length=2)
    display_name: str = Field(min_length=1)
    password: str = Field(min_length=8)
    role: Literal["super_admin", "admin", "employee", "viewer"] = "employee"
    is_active: bool = True


class AdminUserUpdateRequest(BaseModel):
    display_name: Optional[str] = None
    role: Optional[Literal["super_admin", "admin", "employee", "viewer"]] = None
    is_active: Optional[bool] = None


class AdminUserResetPasswordRequest(BaseModel):
    password: str = Field(min_length=8)


class SystemUpdateRequest(BaseModel):
    allow_dirty: bool = False
    restart_after_update: bool = True


class QueryAnswerRequest(BaseModel):
    building_id: Optional[str] = None
    staging_key: Optional[str] = None
    source_mode: Literal["master", "staging"] = "master"
    question: str = Field(min_length=1)
    include_ai: bool = True


class QueryExplanationRequest(BaseModel):
    building_id: Optional[str] = None
    staging_key: Optional[str] = None
    source_mode: Literal["master", "staging"] = "master"
    question: str = Field(min_length=1)


class ChatMessage(BaseModel):
    role: Literal["user", "assistant", "system"]
    content: str = Field(min_length=1)


class ChatRequest(BaseModel):
    building_id: Optional[str] = None
    staging_key: Optional[str] = None
    source_mode: Literal["master", "staging"] = "master"
    question: str = Field(min_length=1)
    history: List[ChatMessage] = Field(default_factory=list)


class HeaderMappingSelection(BaseModel):
    original_header: str
    mapped_field_key: Optional[str] = None
    action: Literal["map", "ignore", "create"] = "map"
    new_field_display_name: Optional[str] = None
    field_type: Optional[str] = "text"


class SheetImportConfirmation(BaseModel):
    sheet_name: str
    header_row_index: int = Field(ge=0)
    mappings: List[HeaderMappingSelection] = Field(default_factory=list)


class ImportConfirmRequest(BaseModel):
    batch_id: str = Field(min_length=1)
    sheets: List[SheetImportConfirmation] = Field(default_factory=list)


class ReviewRecordUpdate(BaseModel):
    record_id: str = Field(min_length=1)
    new_value: Optional[str] = None
    resolution: Optional[Literal["use_new", "use_old", "skip"]] = None


class ReviewDecisionRequest(BaseModel):
    action: Literal["approved", "rejected", "needs_more_info", "conflict", "mark_missing"]
    comment: str = ""
    updates: List[ReviewRecordUpdate] = Field(default_factory=list)


class ConfirmStagingBuildingRequest(BaseModel):
    staging_key: str = Field(min_length=1)


class ConfirmMasterBuildingRequest(BaseModel):
    building_id: str = Field(min_length=1)


class MasterUpdateRequest(BaseModel):
    updates: Dict[str, Optional[str]] = Field(default_factory=dict)
    note: str = ""


class StagingBuildingCreateRequest(BaseModel):
    building_name: str = Field(min_length=1)
    address: str = ""
    aliases: str = ""
    notes: str = ""
    insurance_required: Optional[Literal["", "true", "false", "optional"]] = ""
    electricity_required: Optional[Literal["", "true", "false", "optional"]] = ""
    internet_self_setup_required: Optional[Literal["", "true", "false", "optional"]] = ""


class RollbackRequest(BaseModel):
    audit_log_id: Optional[str] = None
    note: str = ""


class AliasCreateRequest(BaseModel):
    alias_name: str = Field(min_length=1)


class FieldDefinitionCreateRequest(BaseModel):
    field_key: str = Field(min_length=1)
    display_name: str = Field(min_length=1)
    field_type: str = Field(default="text")
    description: str = ""


class FieldRequestDraftFromTextRequest(BaseModel):
    display_name: str = Field(min_length=1)
    requirement_text: str = Field(min_length=1)


class FieldDraftPayload(BaseModel):
    field_key: str = Field(min_length=1)
    display_name: str = Field(min_length=1)
    field_type: str = Field(default="text")
    group_key: str = Field(default="custom")
    excel_header_name: str = Field(min_length=1)
    scope: Literal["master_and_staging", "staging_only"] = "master_and_staging"
    aliases: List[str] = Field(default_factory=list)
    query_keywords: List[str] = Field(default_factory=list)
    answer_template: str = ""
    visible_in_master_detail: bool = True
    visible_in_staging_detail: bool = True
    visible_in_query: bool = True
    description: str = ""


class FieldRequestCreateRequest(BaseModel):
    display_name: str = Field(min_length=1)
    requirement_text: str = ""
    draft: FieldDraftPayload
    apply_immediately: bool = False


class FieldRequestDecisionRequest(BaseModel):
    comment: str = ""
    draft: Optional[FieldDraftPayload] = None


class FieldDefinitionUpdateRequest(BaseModel):
    display_name: Optional[str] = None
    description: Optional[str] = None
    group_key: Optional[str] = None
    excel_header_name: Optional[str] = None
    visible_in_master_detail: Optional[bool] = None
    visible_in_staging_detail: Optional[bool] = None
    visible_in_query: Optional[bool] = None
    query_keywords: Optional[List[str]] = None
    answer_template: Optional[str] = None
    active: Optional[bool] = None


class CrmCaseGuestPayload(BaseModel):
    full_name: str = ""
    phone: str = ""
    email: str = ""
    wechat: str = ""
    notes: str = ""


class CrmCaseCreateRequest(BaseModel):
    group_name: str = Field(min_length=1)
    owner_user_id: Optional[str] = None
    unit: str = ""
    group_creator_name: str = ""
    group_creator_contact: str = ""
    agent_team_t: str = ""
    agent_team_m: str = ""
    lease_start_date: str = ""
    building_source: Literal["", "master", "staging"] = ""
    building_id: str = ""
    notes: str = ""
    network_earliest_start_note: str = ""
    guests: List[CrmCaseGuestPayload] = Field(default_factory=list)


class CrmCaseUpdateRequest(BaseModel):
    group_name: Optional[str] = None
    owner_user_id: Optional[str] = None
    unit: Optional[str] = None
    group_creator_name: Optional[str] = None
    group_creator_contact: Optional[str] = None
    agent_team_t: Optional[str] = None
    agent_team_m: Optional[str] = None
    lease_start_date: Optional[str] = None
    building_source: Optional[Literal["", "master", "staging"]] = None
    building_id: Optional[str] = None
    insurance_earliest_start_date: Optional[str] = None
    network_earliest_start_note: Optional[str] = None
    status: Optional[Literal["active", "paused", "completed", "cancelled"]] = None
    notes: Optional[str] = None


class CrmCaseDeleteRequest(BaseModel):
    reason: str = Field(min_length=1)


class CrmGuestCreateRequest(CrmCaseGuestPayload):
    pass


class CrmGuestUpdateRequest(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    wechat: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[Literal["active", "completed", "inactive"]] = None


class CrmServiceProgressUpdateRequest(BaseModel):
    status: Optional[Literal["pending", "open", "in_progress", "waiting_customer", "at_risk", "completed", "terminated", "not_needed"]] = None
    service_status: Optional[Literal["pending", "open", "in_progress", "waiting_customer", "at_risk", "completed", "terminated", "not_needed"]] = None
    active_flow_step_key: Optional[str] = None
    staff_flow_status: Optional[
        Literal[
            "not_introduced",
            "introduced",
            "following_up",
            "service_confirmed",
            "info_collected",
            "completed",
            "terminated",
        ]
    ] = None
    customer_flow_status: Optional[str] = None
    termination_reason: Optional[str] = None
    applicability: Optional[Literal["required", "optional", "not_needed", "unknown"]] = None
    service_scope: Optional[Literal["case_level", "customer_level"]] = None
    responsible_customer_id: Optional[str] = None
    covered_customer_ids: Optional[List[str]] = None
    responsibility_status: Optional[Literal["unassigned", "assigned", "confirmed", "declined", "changed"]] = None
    need_status: Optional[Literal["required", "optional", "not_needed", "unknown"]] = None
    submission_status: Optional[Literal["not_applicable", "not_submitted", "submitted", "rejected", "approved", "unknown"]] = None
    completion_status: Optional[Literal["not_applicable", "not_started", "in_progress", "completed", "waived", "failed", "unknown"]] = None
    intro_status: Optional[Literal["not_introduced", "introduced_to_group", "introduced_to_customer", "not_needed", "unknown"]] = None
    follow_up_status: Optional[Literal["not_required", "required", "scheduled", "overdue", "unknown"]] = None
    agent_completion_status: Optional[Literal["open", "pending_customer", "pending_external", "completed", "escalated", "unknown"]] = None
    blocked_reason: Optional[str] = None
    group_progress: Dict[str, Any] = Field(default_factory=dict)
    step_key: str = ""
    value: Dict[str, Any] = Field(default_factory=dict)
    note: str = ""


class CrmGuestProgressUpdateRequest(BaseModel):
    step_key: str = Field(min_length=1)
    value: Dict[str, Any] = Field(default_factory=dict)
    sensitive: Dict[str, Any] = Field(default_factory=dict)
    note: str = ""


class CrmTaskCreateRequest(BaseModel):
    title: str = Field(min_length=1)
    description: str = ""
    case_service_id: str = ""
    customer_id: str = ""
    target_customer_id: str = ""
    task_type: str = "manual"
    due_at: str = ""
    not_before_at: str = ""
    assigned_to: str = ""
    priority: Literal["low", "normal", "high", "urgent"] = "normal"
    status: Literal["open", "scheduled", "in_progress", "waiting_customer", "waiting_external", "completed", "overdue", "cancelled"] = "open"


class CrmTaskUpdateRequest(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    task_type: Optional[str] = None
    due_at: Optional[str] = None
    not_before_at: Optional[str] = None
    assigned_to: Optional[str] = None
    target_customer_id: Optional[str] = None
    priority: Optional[Literal["low", "normal", "high", "urgent"]] = None
    status: Optional[Literal["open", "scheduled", "in_progress", "waiting_customer", "waiting_external", "completed", "overdue", "cancelled"]] = None


class CrmCommunicationEventCreateRequest(BaseModel):
    case_service_id: str = ""
    customer_id: str = ""
    channel: Literal["wechat_group", "phone", "email", "internal_note", "ai_draft"] = "wechat_group"
    direction: Literal["inbound", "outbound", "internal"] = "internal"
    summary: str = Field(min_length=1)
    raw_ref: Dict[str, Any] = Field(default_factory=dict)


class CrmNotificationDraftRequest(BaseModel):
    case_id: str = Field(min_length=1)
    task_id: str = ""
    case_service_id: str = ""
    prompt: str = ""
    channel: Literal["wechat", "email", "sms", "internal"] = "wechat"
    recipient_type: Literal["group", "customer", "staff"] = "group"
    recipient_ref: str = ""


class CrmTemplateStepPayload(BaseModel):
    step_key: str = Field(min_length=1)
    title: str = Field(min_length=1)
    scope: Literal["group", "guest"] = "group"
    field_schema: List[dict] = Field(default_factory=list)
    display_order: int = 100
    active: bool = True


class CrmTemplateUpsertRequest(BaseModel):
    service_key: Optional[str] = None
    name: str = Field(min_length=1)
    description: str = ""
    category: str = "general"
    active: bool = True
    display_order: int = 100
    config: Dict[str, Any] = Field(default_factory=dict)
    steps: List[CrmTemplateStepPayload] = Field(default_factory=list)


@dataclass
class Actor:
    user_id: str
    username: str
    role: str


FIXED_NETWORK_PROVIDER_LABELS = list(NETWORK_PROVIDER_FIELD_MAP.values())
FIXED_NETWORK_PROVIDER_SET = set(FIXED_NETWORK_PROVIDER_LABELS)
KNOWN_EXTRA_NETWORK_PROVIDER_SET = {"Honest Networks"}
MASTER_EXCEL_SYNC_SOURCE_TYPE = "master_excel_sync"
STAGING_AB_SYNC_SOURCE_TYPE = "staging_excel_ab_sync"
STAGING_MANUAL_MASTER_SOURCE_TYPE = "staging_manual_master_submit"
STAGING_MANUAL_MASTER_PARSER_TYPE = "staging_manual_master_submit"
SYSTEM_SYNC_ACTOR = Actor(
    user_id="system_master_excel_sync",
    username="system",
    role="ai_system",
)

INTAKE_PARSE_STATUS_QUEUED = "queued"
INTAKE_PARSE_STATUS_RUNNING = "running"
INTAKE_PARSE_STATUS_COMPLETED = "completed"
INTAKE_PARSE_STATUS_FAILED = "failed"
INTAKE_PARSE_STATUSES = {
    INTAKE_PARSE_STATUS_QUEUED,
    INTAKE_PARSE_STATUS_RUNNING,
    INTAKE_PARSE_STATUS_COMPLETED,
    INTAKE_PARSE_STATUS_FAILED,
}
SUMMARY_CACHE_PROMPT_VERSION = "fact_summary_v1"

STAGING_LIBRARY_STATUSES = {
    STAGING_STATUS_PENDING,
    STAGING_STATUS_ACTIVE,
    STAGING_STATUS_MASTERED,
}

ACTIONABLE_REVIEW_STATUSES = {
    "pending",
    "ai_parsed",
    "employee_submitted",
    "needs_more_info",
    "conflict",
}

CRM_ALLOWED_ROLES = {"super_admin", "admin", "employee"}
CRM_SERVICE_STATUSES = {
    "pending",
    "open",
    "in_progress",
    "waiting_customer",
    "at_risk",
    "completed",
    "terminated",
    "not_needed",
}
CRM_STAFF_FLOW_STATUSES = {
    "not_introduced",
    "introduced",
    "following_up",
    "service_confirmed",
    "info_collected",
    "completed",
    "terminated",
}
CRM_CUSTOMER_FLOW_STATUSES = {
    "waiting_intro",
    "intent_unknown",
    "service_confirmed",
    "info_provided",
    "completed",
    "declined",
    "not_needed",
}
CRM_APPLICABILITY_VALUES = {"required", "optional", "not_needed", "unknown"}
CRM_SERVICE_SCOPES = {"case_level", "customer_level"}
CRM_RESPONSIBILITY_STATUSES = {
    "unassigned",
    "assigned",
    "confirmed",
    "declined",
    "changed",
}
CRM_NEED_STATUSES = {"required", "optional", "not_needed", "unknown"}
CRM_SUBMISSION_STATUSES = {
    "not_applicable",
    "not_submitted",
    "submitted",
    "rejected",
    "approved",
    "unknown",
}
CRM_COMPLETION_STATUSES = {
    "not_applicable",
    "not_started",
    "in_progress",
    "completed",
    "waived",
    "failed",
    "unknown",
}
CRM_INTRO_STATUSES = {
    "not_introduced",
    "introduced_to_group",
    "introduced_to_customer",
    "not_needed",
    "unknown",
}
CRM_FOLLOW_UP_STATUSES = {
    "not_required",
    "required",
    "scheduled",
    "overdue",
    "unknown",
}
CRM_AGENT_COMPLETION_STATUSES = {
    "open",
    "pending_customer",
    "pending_external",
    "completed",
    "escalated",
    "unknown",
}
CRM_TASK_STATUSES = {
    "open",
    "scheduled",
    "in_progress",
    "waiting_customer",
    "waiting_external",
    "completed",
    "overdue",
    "cancelled",
}
CRM_TASK_PRIORITIES = {"low", "normal", "high", "urgent"}
CRM_SERVICE_STATUS_LABELS = {
    "pending": "Not Started",
    "open": "Not Started",
    "in_progress": "In Progress",
    "waiting_customer": "Waiting for Customer",
    "at_risk": "At Risk / Incomplete",
    "completed": "Completed",
    "terminated": "Terminated",
    "not_needed": "Not Needed",
}
CRM_APPLICABILITY_LABELS = {
    "required": "Required",
    "optional": "Optional",
    "not_needed": "Not Needed",
    "unknown": "Pending Confirmation",
}
CRM_CASE_STATUS_LABELS = {
    "active": "Active",
    "paused": "Paused",
    "completed": "Completed",
    "cancelled": "Cancelled",
}
CRM_TASK_STATUS_LABELS = {
    "todo": "To Do",
    "open": "To Do",
    "scheduled": "Scheduled",
    "in_progress": "Active",
    "waiting_customer": "Waiting for Customer",
    "waiting_external": "Waiting for Third Party",
    "overdue": "Overdue",
    "completed": "Completed",
    "done": "Completed",
    "cancelled": "Cancelled",
}
CRM_TASK_PRIORITY_LABELS = {
    "low": "Low",
    "normal": "Normal",
    "high": "High",
    "urgent": "Urgent",
}
CRM_LOW_VALUE_SYSTEM_TASK_RULES = {
    "electricity_account:intro_now",
    "electricity_account:not_needed_notice",
    "internet_setup:available",
    "internet_setup:collect_21",
    "internet_setup:install_3",
    "internet_setup:included_notice",
    "phone_card:intro_now",
    "phone_card:follow_14",
    "phone_card:verify_3",
    "renters_insurance:available",
    "renters_insurance:follow_14",
    "renters_insurance:check_7",
    "renters_insurance:escalate_3",
}
CRM_SERVICE_SCOPE_LABELS = {
    "case_level": "Case-Level",
    "customer_level": "Per Customer",
}
CRM_SERVICE_DELIVERY_MODE_LABELS = {
    "sop_only": "SOP / Self-Service Guide Only",
    "assisted": "Assisted Service",
    "agency": "Team-Assisted Setup",
    "sales": "Team Sale / Manual Order Assistance",
    "not_needed": "Not Needed",
}

CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP = {
    "not_introduced": "waiting_intro",
    "introduced": "intent_unknown",
    "following_up": "intent_unknown",
    "service_confirmed": "service_confirmed",
    "info_collected": "info_provided",
    "completed": "completed",
    "terminated": "declined",
}

BASE_CRM_FLOW_PROFILE = {
    "staff_labels": {
        "not_introduced": "Not Introduced",
        "introduced": "Introduced",
        "following_up": "Following Up",
        "service_confirmed": "Customer Confirmed Service",
        "info_collected": "Information Collected",
        "completed": "Support Work Completed",
        "terminated": "Terminated",
    },
    "customer_labels": {
        "waiting_intro": "Awaiting Introduction",
        "intent_unknown": "Customer Intent Unknown",
        "service_confirmed": "Service Confirmed",
        "info_provided": "Information Provided",
        "completed": "Completed",
        "declined": "Declined",
        "not_needed": "Not Needed",
    },
    "skip_stages": [],
    "staff_to_customer_map": CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
    "required_fields_by_stage": {},
    "flow_steps": [],
    "terminal_rules": [
        {
            "customer_flow_status": "declined",
            "service_status": "terminated",
            "cancel_open_tasks": True,
            "require_reason": True,
        }
    ],
}

DEFAULT_CRM_FLOW_PROFILES = {
    "renters_insurance": {
        **BASE_CRM_FLOW_PROFILE,
        "staff_labels": {
            **BASE_CRM_FLOW_PROFILE["staff_labels"],
            "not_introduced": "Send Insurance SOP",
            "introduced": "Insurance SOP Sent",
            "following_up": "Pre-Deadline Check",
            "service_confirmed": "Customer Will Self-Serve",
            "completed": "Insurance Completion Confirmed",
            "terminated": "Customer Declined / Manual Handling Required",
        },
        "customer_labels": {
            **BASE_CRM_FLOW_PROFILE["customer_labels"],
            "intent_unknown": "Confirm Whether SOP Was Followed",
            "service_confirmed": "Customer Will Self-Serve",
            "info_provided": "Customer Provided Insurance Submission Details",
            "completed": "Insurance Completed / Submitted",
            "declined": "Customer Declined",
        },
        "skip_stages": ["staff:service_confirmed", "staff:info_collected"],
        "required_fields_by_stage": {},
        "flow_steps": [
            {
                "step_key": "pending_sop",
                "enabled": True,
                "staff_flow_status": "not_introduced",
                "staff_label": "Send Insurance SOP",
                "customer_flow_status": "waiting_intro",
                "customer_label": "Awaiting Introduction",
                "service_status": "pending",
                "required_fields": [],
                "description": "The insurance SOP has not yet been sent.",
            },
            {
                "step_key": "sop_sent",
                "enabled": True,
                "staff_flow_status": "introduced",
                "staff_label": "Insurance SOP Sent",
                "customer_flow_status": "intent_unknown",
                "customer_label": "Confirm Whether SOP Was Followed",
                "service_status": "in_progress",
                "required_fields": [],
                "description": "The customer purchases and submits insurance using the SOP.",
            },
            {
                "step_key": "insurance_completed",
                "enabled": True,
                "staff_flow_status": "completed",
                "staff_label": "Insurance Completion Confirmed",
                "customer_flow_status": "completed",
                "customer_label": "Insurance Completed / Submitted",
                "service_status": "completed",
                "required_fields": [],
                "is_completion": True,
                "description": "Support staff only confirm that the customer completed the purchase or submission.",
            },
            {
                "step_key": "insurance_at_risk",
                "enabled": True,
                "staff_flow_status": "following_up",
                "staff_label": "At Risk / Incomplete",
                "customer_flow_status": "intent_unknown",
                "customer_label": "Incomplete; Risk Reminder Required",
                "service_status": "at_risk",
                "required_fields": [],
                "is_risk": True,
                "description": "The insurance deadline has arrived or passed, but completion is still unconfirmed.",
            },
        ],
    },
    "electricity_account": {
        **BASE_CRM_FLOW_PROFILE,
        "staff_labels": {
            **BASE_CRM_FLOW_PROFILE["staff_labels"],
            "introduced": "Electricity Guide Sent",
            "following_up": "Follow Up on Electricity Setup",
            "service_confirmed": "Customer Confirmed Setup Method",
            "info_collected": "Assistance Information Collected",
            "completed": "Electricity Setup Confirmed",
        },
        "customer_labels": {
            **BASE_CRM_FLOW_PROFILE["customer_labels"],
            "intent_unknown": "Confirm Whether Electricity Setup Is Needed",
            "service_confirmed": "Customer Confirmed Electricity Setup",
            "info_provided": "Customer Provided Electricity Information",
            "completed": "Electricity Account Completed",
        },
        "required_fields_by_stage": {
            "info_collected": ["account_holder", "phone", "email"]
        },
    },
    "internet_setup": {
        **BASE_CRM_FLOW_PROFILE,
        "staff_labels": {
            **BASE_CRM_FLOW_PROFILE["staff_labels"],
            "introduced": "Internet Options Introduced",
            "following_up": "Follow Up on Internet Decision",
            "service_confirmed": "Customer Confirmed Our Internet Assistance",
            "info_collected": "Verification-Code Window Scheduled",
            "completed": "Order / Installation Confirmed",
        },
        "customer_labels": {
            **BASE_CRM_FLOW_PROFILE["customer_labels"],
            "intent_unknown": "Confirm Whether Customer Wants Internet",
            "service_confirmed": "Customer Confirmed Internet Setup",
            "info_provided": "Customer Provided Account Information",
            "completed": "Internet Activated",
        },
        "required_fields_by_stage": {
            "info_collected": ["verification_window", "account_holder", "phone"]
        },
        "flow_steps": [
            {
                "step_key": "pending_intro",
                "enabled": True,
                "staff_flow_status": "not_introduced",
                "staff_label": "Not Introduced",
                "customer_flow_status": "waiting_intro",
                "customer_label": "Awaiting Introduction",
                "service_status": "pending",
                "required_fields": [],
                "description": "Internet options have not yet been introduced to the customer.",
                "display_order": 10,
            },
            {
                "step_key": "package_introduced",
                "enabled": True,
                "staff_flow_status": "introduced",
                "staff_label": "Internet Options Introduced",
                "customer_flow_status": "intent_unknown",
                "customer_label": "Confirm Whether Customer Wants Internet",
                "service_status": "in_progress",
                "required_fields": [],
                "description": "Available internet options have been introduced in the group or directly to the customer.",
                "display_order": 20,
            },
            {
                "step_key": "confirm_internet_order",
                "enabled": True,
                "staff_flow_status": "service_confirmed",
                "staff_label": "Customer Confirmed Our Internet Assistance",
                "customer_flow_status": "service_confirmed",
                "customer_label": "Customer Confirmed Internet Setup",
                "service_status": "in_progress",
                "required_fields": [],
                "description": "The customer confirmed they want our help arranging internet service; collect account-opening information next.",
                "display_order": 40,
            },
            {
                "step_key": "network_info_collected",
                "enabled": True,
                "staff_flow_status": "info_collected",
                "staff_label": "Account Information Collected",
                "customer_flow_status": "info_provided",
                "customer_label": "Customer Provided Account Information",
                "service_status": "in_progress",
                "required_fields": ["account_holder", "phone", "notes"],
                "description": "The account holder, phone number, and customer notes have been recorded.",
                "display_order": 45,
            },
            {
                "step_key": "verification_window_scheduled",
                "enabled": True,
                "staff_flow_status": "info_collected",
                "staff_label": "Verification-Code Window Scheduled",
                "customer_flow_status": "info_provided",
                "customer_label": "Customer Provided Account Information",
                "service_status": "in_progress",
                "required_fields": ["verification_window"],
                "description": "A start time has been arranged for the customer to receive a verification code.",
                "display_order": 50,
            },
            {
                "step_key": "network_completed",
                "enabled": True,
                "staff_flow_status": "completed",
                "staff_label": "Order / Installation Confirmed",
                "customer_flow_status": "completed",
                "customer_label": "Internet Activated",
                "service_status": "completed",
                "required_fields": [],
                "is_completion": True,
                "description": "The internet order, installation, or activation result has been confirmed.",
                "display_order": 60,
            },
            {
                "step_key": "internet_terminated",
                "enabled": True,
                "staff_flow_status": "terminated",
                "staff_label": "Terminated",
                "customer_flow_status": "declined",
                "customer_label": "Declined",
                "service_status": "terminated",
                "required_fields": [],
                "is_terminal": True,
                "description": "The customer declined the service or staff ended it manually.",
                "display_order": 70,
            },
        ],
    },
    "phone_card": {
        **BASE_CRM_FLOW_PROFILE,
        "staff_labels": {
            **BASE_CRM_FLOW_PROFILE["staff_labels"],
            "introduced": "Phone Options Introduced",
            "following_up": "Follow Up on Customer Intent",
            "service_confirmed": "Customer Confirmed Phone Service",
            "info_collected": "Account / Number-Transfer Information Collected",
            "completed": "Order / Activation Confirmed",
        },
        "customer_labels": {
            **BASE_CRM_FLOW_PROFILE["customer_labels"],
            "intent_unknown": "Confirm Customer Intent",
            "service_confirmed": "Customer Confirmed Service",
            "info_provided": "Customer Provided Information",
            "completed": "Phone Service Completed",
            "declined": "Customer Declined Service",
        },
        "required_fields_by_stage": {
            "info_collected": ["first_name", "last_name", "phone", "number_type"]
        },
    },
}

DEFAULT_CRM_SERVICE_TEMPLATES = [
    {
        "service_key": "renters_insurance",
        "name": "Renters Insurance",
        "description": "Guide insurance introduction, purchase, and submission using verified building requirements.",
        "category": "building",
        "display_order": 10,
        "config": {
            "building_driven": True,
            "auto_source": "insurance",
            "service_scope": "case_level",
            "service_delivery_mode": "sop_only",
            "flow_profile": DEFAULT_CRM_FLOW_PROFILES["renters_insurance"],
        },
        "steps": [
            {
                "step_key": "intro",
                "title": "Introduce Insurance Requirements",
                "scope": "group",
                "display_order": 10,
                "field_schema": [
                    {"key": "introduced", "label": "Insurance Requirements Introduced to the Group", "type": "checkbox"},
                    {"key": "intro_note", "label": "Introduction Notes", "type": "textarea"},
                ],
            },
            {
                "step_key": "guest_status",
                "title": "Customer Insurance Progress",
                "scope": "guest",
                "display_order": 20,
                "field_schema": [
                    {
                        "key": "status",
                        "label": "Service Status",
                        "type": "select",
                        "options": ["Not Started", "Introduced", "Customer Is Purchasing", "Purchased", "Submitted to Building", "Not Needed"],
                    },
                    {"key": "policy_info", "label": "Policy / Submission Information", "type": "textarea"},
                ],
            },
        ],
    },
    {
        "service_key": "electricity_account",
        "name": "Electricity Account",
        "description": "Guide electricity-account setup using verified building rules and utility information.",
        "category": "building",
        "display_order": 20,
        "config": {
            "building_driven": True,
            "auto_source": "electricity",
            "service_scope": "case_level",
            "service_delivery_mode": "sop_only",
            "flow_profile": DEFAULT_CRM_FLOW_PROFILES["electricity_account"],
        },
        "steps": [
            {
                "step_key": "intro",
                "title": "Introduce Electricity Requirements",
                "scope": "group",
                "display_order": 10,
                "field_schema": [
                    {"key": "introduced", "label": "Electricity Setup Requirements Introduced", "type": "checkbox"},
                    {"key": "provider_note", "label": "Utility / Setup Notes", "type": "textarea"},
                ],
            },
            {
                "step_key": "guest_status",
                "title": "Customer Electricity Progress",
                "scope": "guest",
                "display_order": 20,
                "field_schema": [
                    {
                        "key": "status",
                        "label": "Service Status",
                        "type": "select",
                        "options": ["Not Started", "Guide Sent", "Customer Self-Service", "Team-Assisted Setup", "Activated", "Not Needed"],
                    },
                    {"key": "account_note", "label": "Account / Confirmation Notes", "type": "textarea"},
                ],
            },
        ],
    },
    {
        "service_key": "internet_setup",
        "name": "Internet Setup",
        "description": "Guide internet setup using verified building providers, plans, and notes.",
        "category": "building",
        "display_order": 30,
        "config": {
            "building_driven": True,
            "auto_source": "internet",
            "service_scope": "case_level",
            "service_delivery_mode": "sales",
            "flow_profile": DEFAULT_CRM_FLOW_PROFILES["internet_setup"],
        },
        "steps": [
            {
                "step_key": "intro",
                "title": "Introduce Internet Options",
                "scope": "group",
                "display_order": 10,
                "field_schema": [
                    {"key": "introduced", "label": "Available Providers / Rules Introduced", "type": "checkbox"},
                    {"key": "network_note", "label": "Internet Notes", "type": "textarea"},
                ],
            },
            {
                "step_key": "guest_order",
                "title": "Customer Order / Installation Progress",
                "scope": "guest",
                "display_order": 20,
                "field_schema": [
                    {
                        "key": "status",
                        "label": "Service Status",
                        "type": "select",
                        "options": ["Not Started", "Interested", "Provider Selected", "Ordered", "Installation Scheduled", "Installed", "Not Needed"],
                    },
                    {"key": "provider", "label": "Selected Provider / Plan", "type": "text"},
                    {"key": "install_time", "label": "Installation Time", "type": "text"},
                ],
            },
        ],
    },
    {
        "service_key": "phone_card",
        "name": "Phone Service",
        "description": "First determine whether the customer already bought through another channel, then collect information only when appropriate.",
        "category": "general",
        "display_order": 40,
        "config": {
            "building_driven": False,
            "service_scope": "customer_level",
            "service_delivery_mode": "sales",
            "flow_profile": DEFAULT_CRM_FLOW_PROFILES["phone_card"],
        },
        "steps": [
            {
                "step_key": "channel_check",
                "title": "Channel Check",
                "scope": "group",
                "display_order": 10,
                "field_schema": [
                    {
                        "key": "other_channel_purchased",
                        "label": "Already Purchased Through Another Channel",
                        "type": "select",
                        "options": ["Unknown", "No", "Yes"],
                    },
                    {"key": "channel_note", "label": "Channel Notes", "type": "textarea"},
                ],
            },
            {
                "step_key": "attitude",
                "title": "Customer Intent",
                "scope": "group",
                "display_order": 20,
                "field_schema": [
                    {
                        "key": "customer_attitude",
                        "label": "Customer Intent",
                        "type": "select",
                        "options": ["Unknown", "Customer Wants to Order", "Agent Already Handled It", "No Response After Introduction", "Chose Another Provider"],
                    },
                    {"key": "attitude_note", "label": "Follow-Up Notes", "type": "textarea"},
                ],
            },
            {
                "step_key": "collect_info",
                "title": "Order Information",
                "scope": "guest",
                "display_order": 30,
                "field_schema": [
                    {"key": "first_name", "label": "First name", "type": "text"},
                    {"key": "last_name", "label": "Last name", "type": "text"},
                    {"key": "ssn", "label": "SSN", "type": "sensitive"},
                    {"key": "dob", "label": "DOB", "type": "date"},
                    {"key": "number_type", "label": "Number Type", "type": "select", "options": ["New Number", "Port Existing Number"]},
                    {"key": "sim_type", "label": "SIM Type", "type": "select", "options": ["Physical SIM", "eSIM"]},
                    {"key": "imei", "label": "IMEI", "type": "text"},
                    {"key": "number", "label": "Number", "type": "text"},
                    {"key": "carrier", "label": "Carrier", "type": "text"},
                    {"key": "acc", "label": "ACC", "type": "text"},
                    {"key": "transfer_pin", "label": "Transfer PIN", "type": "sensitive"},
                    {"key": "billing_zip", "label": "Billing Zip", "type": "text"},
                    {"key": "address", "label": "Address", "type": "textarea"},
                    {"key": "activation_date", "label": "Activation Date", "type": "date"},
                ],
            },
        ],
    },
]

PROCESSED_REVIEW_STATUSES = {
    "migrated_to_staging",
    "migrated_to_master",
    "rejected",
}

APPROVAL_STAGE_TO_STAGING = "to_staging"
APPROVAL_STAGE_TO_MASTER = "to_master"
APPROVAL_STAGES = {APPROVAL_STAGE_TO_STAGING, APPROVAL_STAGE_TO_MASTER}


def normalize_review_approval_stage(value: object) -> str:
    stage = str(value or "").strip().lower()
    return stage if stage in APPROVAL_STAGES else APPROVAL_STAGE_TO_STAGING


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def json_loads_safe(value: object, default: Any) -> Any:
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    try:
        return json.loads(text)
    except Exception:
        return default


def canonical_json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def request_payload_hash(value: Any) -> str:
    return hashlib.sha256(canonical_json_dumps(value).encode("utf-8")).hexdigest()


def ensure_runtime_status_dir() -> None:
    RUNTIME_STATUS_PATH.parent.mkdir(parents=True, exist_ok=True)


def read_runtime_status() -> dict:
    ensure_runtime_status_dir()
    if not RUNTIME_STATUS_PATH.is_file():
        return {}
    try:
        return json.loads(RUNTIME_STATUS_PATH.read_text("utf-8"))
    except Exception:
        return {}


def write_runtime_status(update: dict) -> None:
    ensure_runtime_status_dir()
    current = read_runtime_status()
    current.update({key: value for key, value in update.items() if value is not None})
    current.setdefault("runtime_mode", RUNTIME_MODE)
    try:
        RUNTIME_STATUS_PATH.write_text(json.dumps(current, ensure_ascii=False, indent=2), "utf-8")
    except Exception:
        return


def ensure_excel_mirror_ready_for_write() -> None:
    runtime_status = read_runtime_status()
    if runtime_status.get("excel_mirror_healthy") is False:
        raise HTTPException(
            status_code=503,
            detail="The Excel mirror is unhealthy, so critical writes are paused. Ask an administrator to refresh or repair the mirror first.",
        )


def probe_tunnel_health() -> Optional[dict]:
    if not TUNNEL_HEALTH_URL:
        return None
    result = {
        "configured": True,
        "name": TUNNEL_NAME or None,
        "url": TUNNEL_HEALTH_URL,
        "healthy": False,
        "status_code": None,
        "error": None,
    }
    try:
        with urllib.request.urlopen(TUNNEL_HEALTH_URL, timeout=2.5) as response:
            result["status_code"] = getattr(response, "status", None)
            result["healthy"] = 200 <= int(result["status_code"] or 0) < 400
    except urllib.error.HTTPError as exc:
        result["status_code"] = exc.code
        result["error"] = str(exc)
    except Exception as exc:
        result["error"] = str(exc)
    return result


def run_system_command(
    args: List[str],
    *,
    cwd: Optional[Path] = None,
    timeout: int = 30,
) -> dict:
    try:
        result = subprocess.run(
            args,
            cwd=str(cwd or APP_DIR),
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=timeout,
            check=False,
        )
        return {
            "ok": result.returncode == 0,
            "returncode": result.returncode,
            "stdout": (result.stdout or "").strip(),
            "stderr": (result.stderr or "").strip(),
        }
    except subprocess.TimeoutExpired as exc:
        return {
            "ok": False,
            "returncode": -1,
            "stdout": (exc.stdout or "").strip() if isinstance(exc.stdout, str) else "",
            "stderr": f"Command timed out after {timeout}s.",
        }
    except Exception as exc:
        return {"ok": False, "returncode": -1, "stdout": "", "stderr": str(exc)}


def git_root_path() -> Optional[Path]:
    result = run_system_command(["git", "-C", str(APP_DIR), "rev-parse", "--show-toplevel"], timeout=10)
    if not result["ok"] or not result["stdout"]:
        return None
    return Path(result["stdout"].splitlines()[-1]).resolve()


def git_text(root: Path, args: List[str], *, timeout: int = 20) -> str:
    result = run_system_command(["git", "-C", str(root), *args], timeout=timeout)
    return result["stdout"].strip() if result["ok"] else ""


def build_update_status(*, check_remote: bool = False) -> dict:
    root = git_root_path()
    if not root:
        return {
            "enabled": False,
            "reason": "The current directory is not a Git repository, so remote updates are unavailable.",
            "update_available": False,
        }
    fetch_result = None
    if check_remote:
        fetch_result = run_system_command(["git", "-C", str(root), "fetch", "--prune"], timeout=45)
    branch = git_text(root, ["rev-parse", "--abbrev-ref", "HEAD"]) or ""
    current_commit = git_text(root, ["rev-parse", "HEAD"]) or ""
    current_short = git_text(root, ["rev-parse", "--short", "HEAD"]) or ""
    upstream = git_text(root, ["rev-parse", "--abbrev-ref", "--symbolic-full-name", "@{u}"]) or ""
    remote_commit = git_text(root, ["rev-parse", upstream or f"origin/{branch}"]) if branch else ""
    remote_short = git_text(root, ["rev-parse", "--short", upstream or f"origin/{branch}"]) if branch else ""
    dirty_output = git_text(root, ["status", "--porcelain"], timeout=20)
    ahead = behind = 0
    if upstream:
        count_text = git_text(root, ["rev-list", "--left-right", "--count", f"HEAD...{upstream}"])
        parts = count_text.split()
        if len(parts) >= 2:
            ahead = int(parts[0] or 0)
            behind = int(parts[1] or 0)
    return {
        "enabled": True,
        "git_root": str(root),
        "branch": branch,
        "upstream": upstream,
        "current_commit": current_commit,
        "current_short": current_short,
        "remote_commit": remote_commit,
        "remote_short": remote_short,
        "ahead": ahead,
        "behind": behind,
        "update_available": bool(remote_commit and current_commit and remote_commit != current_commit and behind > 0),
        "dirty": bool(dirty_output.strip()),
        "dirty_preview": dirty_output.splitlines()[:12],
        "fetch_ok": None if fetch_result is None else fetch_result["ok"],
        "fetch_error": None if fetch_result is None or fetch_result["ok"] else fetch_result["stderr"],
        "update_script": str(UPDATE_SCRIPT_PATH),
        "restart_marker": str(UPDATE_RESTART_MARKER_PATH),
        "last_update": read_runtime_status().get("last_update"),
        "restart_required": bool(read_runtime_status().get("restart_required")),
    }


def delayed_process_exit(delay_seconds: float = 1.0) -> None:
    time.sleep(delay_seconds)
    os._exit(75)


def is_public_bind_host() -> bool:
    host = (PUBLIC_BIND_HOST or "").strip().lower()
    return host in {"0.0.0.0", "::", "[::]"} or (host and host not in {"127.0.0.1", "localhost", "::1"})


def default_password_accounts(conn: sqlite3.Connection) -> List[str]:
    accounts = []
    for username, default_password in BUILTIN_DEFAULT_USER_PASSWORDS.items():
        if not default_password:
            continue
        row = conn.execute("SELECT password_hash FROM users WHERE username = ?", (username,)).fetchone()
        if row and verify_password(default_password, row["password_hash"]):
            accounts.append(username)
    return accounts


def assert_default_passwords_safe(conn: sqlite3.Connection) -> None:
    accounts = default_password_accounts(conn)
    exposed = is_public_bind_host()
    if not accounts:
        write_runtime_status(
            {
                "default_password_accounts": [],
                "default_password_risk": False,
            }
        )
        return
    write_runtime_status(
        {
            "default_password_accounts": accounts,
            "default_password_risk": exposed,
        }
    )
    if exposed and not ALLOW_DEFAULT_PASSWORD_ON_LAN:
        raise RuntimeError(
            "Default account passwords are still active while the service is listening on a LAN or public interface. "
            "Change the passwords first, or temporarily set WHITEPAPER_ALLOW_DEFAULT_PASSWORD_ON_LAN=1 to acknowledge the risk."
        )


def login_rate_key(request: Request, username: str) -> str:
    client_host = request.client.host if request.client else ""
    return f"{client_host}:{username.strip().lower()}"


def prune_login_failures(now: float) -> None:
    cutoff = now - LOGIN_FAILURE_WINDOW_SECONDS
    for key in list(LOGIN_FAILURES):
        LOGIN_FAILURES[key] = [item for item in LOGIN_FAILURES[key] if item >= cutoff]
        if not LOGIN_FAILURES[key]:
            LOGIN_FAILURES.pop(key, None)


def check_login_rate_limit(request: Request, username: str) -> None:
    now = time.time()
    prune_login_failures(now)
    failures = LOGIN_FAILURES.get(login_rate_key(request, username), [])
    if len(failures) >= LOGIN_FAILURE_LIMIT:
        raise HTTPException(status_code=429, detail="Too many failed sign-in attempts. Please try again later.")


def record_login_failure(request: Request, username: str) -> None:
    now = time.time()
    prune_login_failures(now)
    key = login_rate_key(request, username)
    LOGIN_FAILURES.setdefault(key, []).append(now)


def clear_login_failures(request: Request, username: str) -> None:
    LOGIN_FAILURES.pop(login_rate_key(request, username), None)


def begin_idempotent_request(
    conn: sqlite3.Connection,
    *,
    actor: "Actor",
    scope: str,
    idempotency_key: Optional[str],
    payload: Any,
) -> Optional[dict]:
    key = (idempotency_key or "").strip()
    if not key:
        return None
    request_hash = request_payload_hash(payload)
    existing = conn.execute(
        """
        SELECT id, request_hash, status, response_json, status_code
        FROM idempotency_keys
        WHERE user_id = ? AND scope = ? AND idempotency_key = ?
        """,
        (actor.user_id, scope, key),
    ).fetchone()
    if existing:
        if existing["request_hash"] != request_hash:
            raise HTTPException(status_code=409, detail="The same idempotency key was used for a different request. Refresh and try again.")
        if existing["status"] == "completed" and existing["response_json"]:
            return {
                "replay": True,
                "record_id": existing["id"],
                "status_code": existing["status_code"] or 200,
                "response": json_loads_safe(existing["response_json"], {}),
            }
        raise HTTPException(status_code=409, detail="The same operation is already in progress. Refresh shortly to view the result.")

    record_id = f"idem_{uuid.uuid4().hex}"
    timestamp = now_iso()
    conn.execute(
        """
        INSERT INTO idempotency_keys(
          id, user_id, scope, idempotency_key, request_hash, status, response_json, status_code, created_at, updated_at
        ) VALUES(?, ?, ?, ?, ?, 'in_progress', '', 200, ?, ?)
        """,
        (record_id, actor.user_id, scope, key, request_hash, timestamp, timestamp),
    )
    return {"replay": False, "record_id": record_id, "key": key, "scope": scope}


def complete_idempotent_request(
    conn: sqlite3.Connection,
    *,
    actor: "Actor",
    scope: str,
    idempotency_key: Optional[str],
    response_payload: Any,
    status_code: int = 200,
) -> None:
    key = (idempotency_key or "").strip()
    if not key:
        return
    conn.execute(
        """
        UPDATE idempotency_keys
        SET status = 'completed', response_json = ?, status_code = ?, updated_at = ?
        WHERE user_id = ? AND scope = ? AND idempotency_key = ?
        """,
        (
            json.dumps(response_payload, ensure_ascii=False),
            status_code,
            now_iso(),
            actor.user_id,
            scope,
            key,
        ),
    )


def get_db() -> sqlite3.Connection:
    return connect_db()


def get_actor(row: dict) -> Actor:
    return Actor(user_id=row["id"], username=row["username"], role=row["role"])


def system_actor() -> Actor:
    return SYSTEM_SYNC_ACTOR


def iso_now_local_date() -> str:
    return datetime.now().date().isoformat()


def fixed_provider_field_status(value: object) -> Optional[bool]:
    return normalize_booleanish(value)


def fixed_provider_labels_from_values(values: Dict[str, Optional[str]]) -> List[str]:
    labels: List[str] = []
    for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items():
        if fixed_provider_field_status(values.get(field_key)) is True:
            labels.append(label)
    return labels


def extra_provider_labels_from_text(value: object) -> List[str]:
    return [
        label
        for label in extract_provider_names(value)
        if label not in FIXED_NETWORK_PROVIDER_SET
    ]


def known_extra_provider_labels_from_text(value: object) -> List[str]:
    return [
        label
        for label in extract_provider_names(value)
        if label in KNOWN_EXTRA_NETWORK_PROVIDER_SET
    ]


def merge_extra_provider_text(*values: object) -> Optional[str]:
    labels: List[str] = []
    for value in values:
        for label in extra_provider_labels_from_text(value):
            if label not in labels:
                labels.append(label)
    return ", ".join(labels) if labels else None


def ensure_known_extra_providers_from_notes(provider_value: object, notes_value: object) -> Optional[str]:
    labels = extra_provider_labels_from_text(provider_value)
    for label in known_extra_provider_labels_from_text(notes_value):
        if label not in labels:
            labels.append(label)
    return ", ".join(labels) if labels else None


def build_network_provider_text_from_values(values: Dict[str, Optional[str]]) -> Optional[str]:
    combined = fixed_provider_labels_from_values(values)
    provider_value = ensure_known_extra_providers_from_notes(
        values.get("internet_provider"),
        values.get("internet_notes"),
    )
    for label in extra_provider_labels_from_text(provider_value):
        if label not in combined:
            combined.append(label)
    return ", ".join(combined) if combined else None


def parse_bearer_token(authorization: Optional[str]) -> Optional[str]:
    if not authorization:
        return None
    parts = authorization.split(" ", 1)
    if len(parts) != 2 or parts[0].lower() != "bearer":
        return None
    return parts[1].strip() or None


async def get_current_user(
    authorization: Optional[str] = Header(default=None),
) -> dict:
    token = parse_bearer_token(authorization)
    if not token:
        raise HTTPException(status_code=401, detail="You are not signed in, or the session is no longer valid.")

    token_hash_value = hash_token(token)
    last_locked_error = ""
    for attempt in range(3):
        try:
            with db_connection() as conn:
                session = conn.execute(
                    """
                    SELECT sessions.id, sessions.user_id, sessions.expires_at, users.username, users.display_name, users.role, users.is_active
                    FROM sessions
                    JOIN users ON users.id = sessions.user_id
                    WHERE sessions.token_hash = ?
                    """,
                    (token_hash_value,),
                ).fetchone()
                if not session:
                    raise HTTPException(status_code=401, detail="The session is invalid. Please sign in again.")
                if not session["is_active"]:
                    raise HTTPException(status_code=403, detail="This account has been disabled.")
                if session["expires_at"] <= now_iso():
                    try:
                        conn.execute("DELETE FROM sessions WHERE id = ?", (session["id"],))
                    except sqlite3.OperationalError as exc:
                        if "locked" not in str(exc).lower():
                            raise
                    raise HTTPException(status_code=401, detail="The session has expired. Please sign in again.")
                try:
                    conn.execute(
                        "UPDATE sessions SET last_used_at = ? WHERE id = ?",
                        (now_iso(), session["id"]),
                    )
                except sqlite3.OperationalError as exc:
                    if "locked" not in str(exc).lower():
                        raise
                return {
                    "id": session["user_id"],
                    "username": session["username"],
                    "display_name": session["display_name"],
                    "role": session["role"],
                }
        except sqlite3.OperationalError as exc:
            if "locked" not in str(exc).lower():
                raise
            last_locked_error = str(exc)
            if attempt < 2:
                await asyncio.sleep(0.25 * (attempt + 1))
                continue
    raise HTTPException(
        status_code=503,
        detail=f"The database is busy with background parsing. Please try again shortly. {last_locked_error}",
    )


def require_roles(*roles: str) -> Callable[[dict], dict]:
    async def _dependency(user: dict = Depends(get_current_user)) -> dict:
        if user["role"] not in roles:
            raise HTTPException(status_code=403, detail="Your account does not have permission to perform this action.")
        return user

    return _dependency


def request_ip(request: Request) -> str:
    if not request.client:
        return ""
    return request.client.host or ""


def write_audit_log(
    conn: sqlite3.Connection,
    request: Optional[Request],
    actor: Actor,
    *,
    action_type: str,
    target_table: str,
    target_record_id: Optional[str] = None,
    building_name: Optional[str] = None,
    field_name: Optional[str] = None,
    old_value: Optional[str] = None,
    new_value: Optional[str] = None,
    source: Optional[str] = None,
    note: str = "",
) -> None:
    created_at = now_iso()
    conn.execute(
        """
        INSERT INTO audit_logs(
          id, user_id, user_role, action_type, target_table, target_record_id, building_name,
          field_name, old_value, new_value, source, ip_address, user_agent, note, created_at
        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            f"audit_{uuid.uuid4().hex}",
            actor.user_id,
            actor.role,
            action_type,
            target_table,
            target_record_id,
            building_name or "",
            field_name or "",
            old_value or "",
            new_value or "",
            source or "",
            request_ip(request) if request else "",
            request.headers.get("user-agent", "") if request else "",
            note,
            created_at,
        ),
    )


def field_catalog(conn: sqlite3.Connection) -> List[dict]:
    return load_field_catalog(conn, statuses=("active",))


def find_field_definition(conn: sqlite3.Connection, field_key: str) -> Optional[dict]:
    return conn.execute(
        """
        SELECT
          id,
          field_key,
          display_name,
          field_type,
          required,
          description,
          is_core,
          active,
          scope,
          group_key,
          display_order,
          excel_header_name,
          visible_in_master_detail,
          visible_in_staging_detail,
          visible_in_query,
          query_keywords_json,
          answer_template,
          status
        FROM field_definitions
        WHERE field_key = ?
        """,
        (field_key,),
    ).fetchone()


def canonical_header(value: str) -> str:
    return re.sub(r"[\s\-_./\\()（）:：;；&|]+", "", (value or "").strip().lower())


def quick_mapping_candidates(conn: sqlite3.Connection, header: str) -> List[dict]:
    normalized_header = canonical_header(header)
    definitions = field_catalog(conn)
    exact: List[dict] = []
    fuzzy: List[tuple[float, dict]] = []

    for item in definitions:
        candidates = [item["display_name"], item["field_key"], *item.get("aliases", [])]
        for candidate in candidates:
            normalized_candidate = canonical_header(candidate)
            if normalized_candidate == normalized_header:
                exact.append(
                    {
                        "field_key": item["field_key"],
                        "display_name": item["display_name"],
                        "match_method": "exact_match"
                        if candidate in {item["display_name"], item["field_key"]}
                        else "alias_match",
                        "confidence": 1.0,
                    }
                )
                break
            ratio = SequenceMatcher(None, normalized_header, normalized_candidate).ratio()
            if ratio >= 0.72:
                fuzzy.append(
                    (
                        ratio,
                        {
                            "field_key": item["field_key"],
                            "display_name": item["display_name"],
                            "match_method": "fuzzy_match",
                            "confidence": round(float(ratio), 3),
                        },
                    )
                )
    if exact:
        exact.sort(key=lambda item: (item["match_method"] != "exact_match", item["display_name"]))
        return exact[:3]
    fuzzy.sort(key=lambda pair: (-pair[0], pair[1]["display_name"]))
    return [item for _, item in fuzzy[:3]]


def optional_ai_client() -> Optional[AsyncOpenAI]:
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip().strip('"').strip("'")
    if not api_key or api_key.lower() in {"your_key_here", "your_real_key", "change_me", "xxx"}:
        return None
    return AsyncOpenAI(api_key=api_key, base_url=deepseek_base_url())


def deepseek_model() -> str:
    return (os.getenv("DEEPSEEK_MODEL", "") or "deepseek-chat").strip().strip('"').strip("'") or "deepseek-chat"


def deepseek_base_url() -> str:
    return (
        (os.getenv("DEEPSEEK_BASE_URL", "") or os.getenv("DEEPSEEK_API_BASE", "") or "https://api.deepseek.com")
        .strip()
        .strip('"')
        .strip("'")
        or "https://api.deepseek.com"
    )


def env_bool(name: str, default: bool = False) -> bool:
    raw = (os.getenv(name, "") or "").strip().strip('"').strip("'").lower()
    if not raw:
        return default
    if raw in {"1", "true", "yes", "y", "on", "enabled"}:
        return True
    if raw in {"0", "false", "no", "n", "off", "disabled"}:
        return False
    return default


def deepseek_thinking_enabled() -> bool:
    return env_bool("DEEPSEEK_THINKING_ENABLED", True)


def deepseek_thinking_reasoning_effort() -> str:
    value = (os.getenv("DEEPSEEK_THINKING_REASONING_EFFORT", "") or "high").strip().strip('"').strip("'").lower()
    return value if value in {"low", "medium", "high"} else "high"


def write_llm_call_log(
    conn: sqlite3.Connection,
    *,
    source_document_id: Optional[str],
    stage: str,
    model: str,
    system_prompt: str,
    user_payload: object,
    raw_response: str = "",
    parsed_response: object = None,
    error: str = "",
) -> None:
    try:
        conn.execute(
            """
            INSERT INTO llm_call_logs(
              id, source_document_id, stage, model, system_prompt, user_payload_json,
              raw_response, parsed_response_json, error, created_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                f"llm_log_{uuid.uuid4().hex}",
                source_document_id,
                stage,
                model,
                system_prompt,
                json_dumps(user_payload),
                raw_response or "",
                json_dumps(parsed_response if parsed_response is not None else {}),
                error or "",
                now_iso(),
            ),
        )
        conn.commit()
    except sqlite3.Error:
        # LLM logging must never block parsing or review intake.
        return


def ai_explanation_enabled() -> bool:
    return optional_ai_client() is not None


async def generate_fact_explanation(
    *,
    question: str,
    snapshot: dict,
    field_keys: List[str],
    fact_answer: str,
    source_mode: Literal["master", "staging", "crm"] = "master",
) -> Optional[str]:
    client = optional_ai_client()
    if not client or not fact_answer.strip():
        return None
    source_labels = {
        "master": "Master",
        "staging": "Current Staging record",
        "crm": "Current CRM Case snapshot",
    }

    fact_payload = {
        "source_mode": source_mode,
        "source_label": source_labels.get(source_mode, "Master"),
        "building_name": snapshot.get("building_name"),
        "address": snapshot.get("address"),
        "field_keys": field_keys,
        "fact_answer": fact_answer,
        "pending_count": snapshot.get("pending_count", 0),
        "facts": {
            "insurance_required": snapshot.get("insurance_required"),
            "insurance_coverage_amount": snapshot.get("insurance_coverage_amount"),
            "electricity_required": snapshot.get("electricity_required"),
            "electricity_provider": snapshot.get("electricity_provider"),
            "internet_self_setup_required": snapshot.get("internet_self_setup_required"),
            "internet_provider": snapshot.get("internet_provider"),
            "internet_notes": snapshot.get("internet_notes"),
            "move_in_notes": snapshot.get("move_in_notes"),
            **{
                field_key: insurance_field_value(snapshot, field_key)
                for field_key in DETAILED_INSURANCE_FIELD_KEYS
            },
        },
    }
    try:
        response = await asyncio.wait_for(
            client.chat.completions.create(
                model=deepseek_model(),
                temperature=0.2,
                messages=[
                    {"role": "system", "content": FACT_EXPLANATION_PROMPT},
                    {
                        "role": "user",
                        "content": json_dumps(
                            {
                                "question": question,
                                "structured_record_payload": fact_payload,
                            }
                        ),
                    },
                ],
                max_tokens=AI_EXPLANATION_MAX_TOKENS,
            ),
            timeout=AI_EXPLANATION_TIMEOUT_SECONDS,
        )
        content = normalize_unknown_value(response.choices[0].message.content or "")
        return content
    except Exception:
        return None


async def ai_suggest_field_mapping(header: str, catalog: List[dict]) -> Optional[dict]:
    client = optional_ai_client()
    if not client:
        return None

    prompt = {
        "header": header,
        "field_options": [
            {
                "field_key": item["field_key"],
                "display_name": item["display_name"],
                "aliases": item.get("aliases", [])[:8],
            }
            for item in catalog[:80]
        ],
    }
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {
                    "role": "system",
                    "content": "You map spreadsheet headers to schema fields. Return a candidate only when confidence is high. Output JSON: {\"field_key\": string|null, \"confidence\": 0-1}.",
                },
                {"role": "user", "content": json_dumps(prompt)},
            ],
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        data = json.loads(content)
    except Exception:
        return None

    field_key = (data.get("field_key") or "").strip()
    if not field_key:
        return None
    match = next((item for item in catalog if item["field_key"] == field_key), None)
    if not match:
        return None
    confidence = float(data.get("confidence") or 0)
    if confidence <= 0:
        confidence = 0.55
    return {
        "field_key": field_key,
        "display_name": match["display_name"],
        "match_method": "ai_suggested",
        "confidence": round(confidence, 3),
    }


def infer_field_group(display_name: str, requirement_text: str) -> str:
    haystack = " ".join([display_name or "", requirement_text or ""]).lower()
    if any(token in haystack for token in FIELD_GROUP_KEYWORDS["insurance"]):
        return "insurance"
    if any(token in haystack for token in FIELD_GROUP_KEYWORDS["electric"]):
        return "electricity"
    if any(token in haystack for token in FIELD_GROUP_KEYWORDS["internet"]):
        return "internet"
    if any(token in haystack for token in FIELD_GROUP_KEYWORDS["move_in"]):
        return "move_in"
    if any(token in haystack for token in FIELD_GROUP_KEYWORDS["contacts"]):
        return "contacts"
    if any(token in haystack for token in ("地址", "楼名", "building", "address", "location")):
        return "basic"
    return "custom"


def infer_field_type(display_name: str, requirement_text: str) -> str:
    haystack = " ".join([display_name or "", requirement_text or ""]).lower()
    boolean_tokens = (
        "是否",
        "需不需要",
        "required",
        "support",
        "supports",
        "可办理",
        "available",
        "need",
        "must",
        "can",
        "able",
    )
    return "boolean" if any(token in haystack for token in boolean_tokens) else "text"


def infer_scope(requirement_text: str) -> str:
    text = (requirement_text or "").lower()
    return "staging_only" if "仅临时" in requirement_text or "staging only" in text else "master_and_staging"


def normalize_query_keywords(display_name: str, query_keywords: List[str]) -> List[str]:
    seen: set[str] = set()
    values: List[str] = []
    for raw in [display_name, *query_keywords]:
        text = normalize_unknown_value(raw)
        if not text:
            continue
        text = text.strip()
        lowered = text.lower()
        if lowered in seen:
            continue
        seen.add(lowered)
        values.append(text)
    return values[:12]


def next_display_order_for_group(conn: sqlite3.Connection, group_key: str) -> int:
    row = conn.execute(
        """
        SELECT MAX(display_order) AS max_order
        FROM field_definitions
        WHERE group_key = ?
        """,
        (group_key,),
    ).fetchone()
    max_order = int((row or {}).get("max_order") or 0)
    if max_order:
        return max_order + 5
    metadata = default_field_metadata(f"draft_{group_key}", group_key, "text", 0)
    return int(metadata.get("display_order") or 900)


def heuristic_field_draft(display_name: str, requirement_text: str) -> dict:
    group_key = infer_field_group(display_name, requirement_text)
    field_type = infer_field_type(display_name, requirement_text)
    metadata = default_field_metadata(
        slugify_field_key(display_name or requirement_text or "custom_field"),
        display_name,
        field_type,
        0,
    )
    query_keywords = normalize_query_keywords(
        display_name,
        re.split(r"[，,；;、/\n]+", requirement_text or ""),
    )
    return {
        "field_key": slugify_field_key(display_name or requirement_text or "custom_field"),
        "display_name": display_name,
        "field_type": field_type,
        "group_key": group_key,
        "excel_header_name": display_name,
        "scope": infer_scope(requirement_text),
        "aliases": [display_name],
        "query_keywords": query_keywords,
        "answer_template": metadata.get("answer_template") or "{display_name}：{value}",
        "visible_in_master_detail": True,
        "visible_in_staging_detail": True,
        "visible_in_query": True,
        "description": requirement_text.strip() or f"Business field for {display_name}",
    }


def normalize_field_draft_payload(payload: Union[dict, FieldDraftPayload]) -> dict:
    raw = payload.model_dump() if isinstance(payload, FieldDraftPayload) else dict(payload or {})
    display_name = (normalize_unknown_value(raw.get("display_name")) or "").strip()
    if not display_name:
        raise HTTPException(status_code=400, detail="The field display name cannot be empty.")
    requirement_text = normalize_unknown_value(raw.get("description")) or ""
    field_type = (normalize_unknown_value(raw.get("field_type")) or infer_field_type(display_name, requirement_text)).strip().lower()
    if field_type not in {"text", "boolean"}:
        field_type = "text"
    group_key = (normalize_unknown_value(raw.get("group_key")) or infer_field_group(display_name, requirement_text)).strip()
    if group_key not in FIELD_GROUP_OPTIONS:
        group_key = "custom"
    field_key = slugify_field_key(
        normalize_unknown_value(raw.get("field_key")) or display_name
    )
    metadata = default_field_metadata(field_key, display_name, field_type, 0)
    scope = (normalize_unknown_value(raw.get("scope")) or infer_scope(requirement_text)).strip()
    if scope not in {"master_and_staging", "staging_only"}:
        scope = "master_and_staging"
    excel_header_name = (normalize_unknown_value(raw.get("excel_header_name")) or display_name).strip()
    aliases = normalize_query_keywords(
        display_name,
        [*(raw.get("aliases") or []), display_name],
    )
    query_keywords = normalize_query_keywords(
        display_name,
        raw.get("query_keywords") or [],
    )
    answer_template = (normalize_unknown_value(raw.get("answer_template")) or metadata.get("answer_template") or "{display_name}：{value}").strip()
    description = (normalize_unknown_value(raw.get("description")) or "").strip()
    return {
        "field_key": field_key,
        "display_name": display_name,
        "field_type": field_type,
        "group_key": group_key,
        "excel_header_name": excel_header_name,
        "scope": scope,
        "aliases": aliases,
        "query_keywords": query_keywords,
        "answer_template": answer_template,
        "visible_in_master_detail": bool(raw.get("visible_in_master_detail", metadata.get("visible_in_master_detail", 1))),
        "visible_in_staging_detail": bool(raw.get("visible_in_staging_detail", metadata.get("visible_in_staging_detail", 1))),
        "visible_in_query": bool(raw.get("visible_in_query", metadata.get("visible_in_query", 1))),
        "description": description or f"Business field for {display_name}",
    }


async def ai_generate_field_draft(display_name: str, requirement_text: str) -> tuple[dict, bool]:
    heuristic = heuristic_field_draft(display_name, requirement_text)
    client = optional_ai_client()
    if not client:
        return normalize_field_draft_payload(heuristic), False
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0.2,
            response_format={"type": "json_object"},
            messages=[
                {
                    "role": "system",
                    "content": (
                        "You are an internal knowledge-base field-design assistant. Use the field name and requirement description to draft a field as JSON. "
                        "Return JSON only, without explanation. group_key must be basic/insurance/electricity/internet/move_in/custom. "
                        "field_type must be text or boolean. scope must be master_and_staging or staging_only."
                    ),
                },
                {
                    "role": "user",
                    "content": json_dumps(
                        {
                            "display_name": display_name,
                            "requirement_text": requirement_text,
                            "expected_shape": {
                                "field_key": "snake_case",
                                "display_name": "Display name",
                                "field_type": "text|boolean",
                                "group_key": "basic|insurance|electricity|internet|move_in|custom",
                                "excel_header_name": "Excel column name",
                                "scope": "master_and_staging|staging_only",
                                "aliases": ["alias 1", "alias 2"],
                                "query_keywords": ["query keyword 1", "query keyword 2"],
                                "answer_template": "{display_name}：{value}",
                                "visible_in_master_detail": True,
                                "visible_in_staging_detail": True,
                                "visible_in_query": True,
                                "description": "Field description",
                            },
                        }
                    ),
                },
            ],
        )
        content = response.choices[0].message.content or "{}"
        draft = json.loads(content)
        merged = {**heuristic, **(draft or {})}
        return normalize_field_draft_payload(merged), True
    except Exception:
        return normalize_field_draft_payload(heuristic), False


def load_field_change_requests(
    conn: sqlite3.Connection,
    *,
    status: Optional[str] = None,
    requested_by: Optional[str] = None,
) -> List[dict]:
    where_parts: List[str] = []
    params: List[object] = []
    if status:
        where_parts.append("status = ?")
        params.append(status)
    if requested_by:
        where_parts.append("requested_by = ?")
        params.append(requested_by)
    sql = """
        SELECT *
        FROM field_change_requests
    """
    if where_parts:
        sql += " WHERE " + " AND ".join(where_parts)
    sql += " ORDER BY created_at DESC"
    rows = conn.execute(sql, tuple(params)).fetchall()
    items: List[dict] = []
    for row in rows:
        item = dict(row)
        item["draft"] = json_loads_safe(item.get("draft_payload_json"), {})
        items.append(item)
    return items


def rename_excel_header_everywhere(old_header: str, new_header: str) -> None:
    if not old_header or old_header == new_header:
        return
    for workbook_path in [resolve_master_excel_path(), resolve_staging_excel_path()]:
        if not workbook_path.exists():
            continue
        workbook = load_workbook(workbook_path)
        changed = False
        if MASTER_MAIN_SHEET in workbook.sheetnames:
            main_sheet = workbook[MASTER_MAIN_SHEET]
            for column in range(1, main_sheet.max_column + 1):
                cell = main_sheet.cell(row=1, column=column)
                if str(cell.value or "").strip() == old_header:
                    cell.value = new_header
                    changed = True
                    break
        if MASTER_HELP_SHEET in workbook.sheetnames:
            help_sheet = workbook[MASTER_HELP_SHEET]
            for row_index in range(2, help_sheet.max_row + 1):
                cell = help_sheet.cell(row=row_index, column=1)
                if str(cell.value or "").strip() == old_header:
                    cell.value = new_header
                    changed = True
                    break
        if changed:
            tmp_dir = tempfile.mkdtemp(prefix="whitepaper_field_header_")
            tmp_path = Path(tmp_dir) / workbook_path.name
            workbook.save(tmp_path)
            tmp_path.replace(workbook_path)
        workbook.close()


def apply_field_draft(
    conn: sqlite3.Connection,
    *,
    draft_payload: Union[dict, FieldDraftPayload],
    actor: Actor,
    request: Optional[Request] = None,
    action_note: str = "",
) -> dict:
    draft = normalize_field_draft_payload(draft_payload)
    existing = find_field_definition(conn, draft["field_key"])
    now = now_iso()
    display_order = next_display_order_for_group(conn, draft["group_key"])
    if existing:
        display_order = int(existing.get("display_order") or display_order)
        old_header = normalize_unknown_value(existing.get("excel_header_name")) or existing.get("display_name") or ""
        conn.execute(
            """
            UPDATE field_definitions
            SET display_name = ?, field_type = ?, description = ?, active = 1, scope = ?, group_key = ?,
                display_order = ?, excel_header_name = ?, visible_in_master_detail = ?, visible_in_staging_detail = ?,
                visible_in_query = ?, query_keywords_json = ?, answer_template = ?, status = 'active', updated_at = ?
            WHERE field_key = ?
            """,
            (
                draft["display_name"],
                draft["field_type"],
                draft["description"],
                draft["scope"],
                draft["group_key"],
                display_order,
                draft["excel_header_name"],
                1 if draft["visible_in_master_detail"] else 0,
                1 if draft["visible_in_staging_detail"] else 0,
                1 if draft["visible_in_query"] else 0,
                json_dumps(draft["query_keywords"]),
                draft["answer_template"],
                now,
                draft["field_key"],
            ),
        )
        rename_excel_header_everywhere(old_header, draft["excel_header_name"])
        action_type = "field_definition_updated"
    else:
        conn.execute(
            """
            INSERT INTO field_definitions(
              id, field_key, display_name, field_type, required, description, is_core, active,
              scope, group_key, display_order, excel_header_name, visible_in_master_detail,
              visible_in_staging_detail, visible_in_query, query_keywords_json, answer_template,
              status, created_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, 0, ?, 0, 1, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?, ?)
            """,
            (
                f"field_{draft['field_key']}",
                draft["field_key"],
                draft["display_name"],
                draft["field_type"],
                draft["description"],
                draft["scope"],
                draft["group_key"],
                display_order,
                draft["excel_header_name"],
                1 if draft["visible_in_master_detail"] else 0,
                1 if draft["visible_in_staging_detail"] else 0,
                1 if draft["visible_in_query"] else 0,
                json_dumps(draft["query_keywords"]),
                draft["answer_template"],
                actor.user_id,
                now,
                now,
            ),
        )
        action_type = "field_definition_created"

    alias_rows = conn.execute(
        "SELECT alias_name FROM field_aliases WHERE field_key = ?",
        (draft["field_key"],),
    ).fetchall()
    existing_aliases = {str(row["alias_name"]).strip().lower() for row in alias_rows}
    for alias in draft["aliases"]:
        alias_text = (normalize_unknown_value(alias) or "").strip()
        if not alias_text or alias_text.lower() in existing_aliases:
            continue
        existing_aliases.add(alias_text.lower())
        conn.execute(
            """
            INSERT INTO field_aliases(id, field_key, alias_name, language, confidence, created_by, created_at)
            VALUES(?, ?, ?, 'mixed', 1.0, ?, ?)
            """,
            (f"alias_{draft['field_key']}_{uuid.uuid4().hex[:8]}", draft["field_key"], alias_text, actor.user_id, now),
        )

    ensure_master_workbook_from_db(conn)
    ensure_staging_workbook_from_sources(conn)
    sync_excel_mirrors(conn, actor=actor, request=request)
    write_audit_log(
        conn,
        request,
        actor,
        action_type=action_type,
        target_table="field_definitions",
        target_record_id=draft["field_key"],
        field_name=draft["field_key"],
        new_value=json_dumps(draft),
        note=action_note or "Applied field draft.",
    )
    return find_field_definition(conn, draft["field_key"]) or draft


def detect_header_row(rows: List[List[object]], conn: sqlite3.Connection) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:10]):
        headers = [legacy._format_spreadsheet_value(value) for value in row]
        score = 0
        populated = 0
        for header in headers:
            if not header:
                continue
            populated += 1
            score += len(quick_mapping_candidates(conn, header))
        if populated == 0:
            continue
        total_score = score * 10 + populated
        if total_score > best_score:
            best_score = total_score
            best_index = index
    return best_index


async def preview_excel_file(file_path: Path, conn: sqlite3.Connection) -> dict:
    sheets = []
    catalog = field_catalog(conn)
    all_headers: List[str] = []
    for sheet_name, rows in legacy._load_tabular_sheet_rows(file_path):
        header_row_index = detect_header_row(rows, conn)
        header_row = rows[header_row_index] if header_row_index < len(rows) else []
        headers = [legacy._format_spreadsheet_value(value) for value in header_row]
        preview_headers = []
        for header in headers:
            if not header:
                continue
            all_headers.append(header)
            candidates = quick_mapping_candidates(conn, header)
            if not candidates:
                ai_candidate = await ai_suggest_field_mapping(header, catalog)
                if ai_candidate:
                    candidates = [ai_candidate]
            preview_headers.append(
                {
                    "original_header": header,
                    "suggested": candidates[0] if candidates else None,
                    "candidates": candidates,
                }
            )

        sample_rows = []
        for row in rows[header_row_index + 1 : header_row_index + 4]:
            sample_rows.append([legacy._format_spreadsheet_value(value) for value in row[: len(headers)]])

        sheets.append(
            {
                "sheet_name": sheet_name,
                "header_row_index": header_row_index,
                "headers": preview_headers,
                "sample_rows": sample_rows,
            }
        )
    return {"sheets": sheets, "available_fields": field_catalog(conn)}


def validate_standard_master_upload(file_path: Path) -> dict:
    validation = validate_master_workbook(file_path)
    if not validation["ok"]:
        detail_parts = []
        if validation["missing_sheets"]:
            detail_parts.append(f"Missing sheet(s): {', '.join(validation['missing_sheets'])}")
        if validation["missing_headers"]:
            detail_parts.append(f"Missing Master header(s): {', '.join(validation['missing_headers'])}")
        if validation["duplicate_headers"]:
            detail_parts.append(f"Duplicate header(s): {', '.join(validation['duplicate_headers'])}")
        if validation["merged_ranges"]:
            detail_parts.append(f"The Master sheet contains merged cell range(s): {', '.join(validation['merged_ranges'])}")
        detail = "; ".join(detail_parts) or "Master workbook validation failed."
        raise HTTPException(status_code=400, detail=detail)
    return validation


def validate_upload_file(upload_file: UploadFile, allowed_suffixes: set[str], label: str) -> str:
    file_name = upload_file.filename or "upload.bin"
    suffix = Path(file_name).suffix.lower()
    if suffix not in allowed_suffixes:
        allowed = ", ".join(sorted(allowed_suffixes))
        raise HTTPException(status_code=400, detail=f"Unsupported {label.lower()} file type. Upload one of: {allowed}.")
    content_type = (upload_file.content_type or "").lower().strip()
    if content_type and content_type != "application/octet-stream":
        if allowed_suffixes == ALLOWED_PDF_UPLOAD_SUFFIXES and "pdf" not in content_type:
            raise HTTPException(status_code=400, detail="The PDF MIME type does not match the file.")
        if allowed_suffixes == ALLOWED_IMAGE_UPLOAD_SUFFIXES and not content_type.startswith("image/"):
            raise HTTPException(status_code=400, detail="The image MIME type does not match the file.")
        if allowed_suffixes == ALLOWED_EXCEL_UPLOAD_SUFFIXES and not any(
            token in content_type for token in ("spreadsheet", "excel", "sheet", "octet-stream")
        ):
            raise HTTPException(status_code=400, detail="The Excel MIME type does not match the file.")
    return file_name


def validate_stored_upload_magic(path: Path, allowed_suffixes: Optional[set[str]], label: str) -> None:
    if not allowed_suffixes:
        return
    suffix = path.suffix.lower()
    try:
        with path.open("rb") as handle:
            header = handle.read(16)
    except Exception:
        raise HTTPException(status_code=400, detail=f"The {label.lower()} file could not be read.")
    if allowed_suffixes == ALLOWED_PDF_UPLOAD_SUFFIXES and not header.startswith(b"%PDF"):
        raise HTTPException(status_code=400, detail="The PDF content does not match its extension.")
    if allowed_suffixes == ALLOWED_IMAGE_UPLOAD_SUFFIXES:
        signatures = {
            ".png": lambda data: data.startswith(b"\x89PNG\r\n\x1a\n"),
            ".jpg": lambda data: data.startswith(b"\xff\xd8\xff"),
            ".jpeg": lambda data: data.startswith(b"\xff\xd8\xff"),
            ".webp": lambda data: data.startswith(b"RIFF") and data[8:12] == b"WEBP",
            ".bmp": lambda data: data.startswith(b"BM"),
            ".tif": lambda data: data.startswith(b"II*\x00") or data.startswith(b"MM\x00*"),
            ".tiff": lambda data: data.startswith(b"II*\x00") or data.startswith(b"MM\x00*"),
        }
        if suffix in signatures and not signatures[suffix](header):
            raise HTTPException(status_code=400, detail="The image content does not match its extension.")


def save_upload_file(
    upload_dir: Path,
    upload_file: UploadFile,
    *,
    allowed_suffixes: Optional[set[str]] = None,
    label: str = "Upload",
    name_prefix: str = "",
) -> tuple[Path, str]:
    upload_dir.mkdir(parents=True, exist_ok=True)
    file_name = (
        validate_upload_file(upload_file, allowed_suffixes, label)
        if allowed_suffixes
        else upload_file.filename or "upload.bin"
    )
    safe_name = re.sub(r"[^\w.\-\u4e00-\u9fff]+", "_", file_name).strip("_") or "upload.bin"
    safe_prefix = re.sub(r"[^\w.\-]+", "_", name_prefix).strip("_")
    stored_name = f"{uuid.uuid4().hex}_{safe_name}"
    if safe_prefix:
        stored_name = f"{safe_prefix}_{stored_name}"
    stored_path = upload_dir / stored_name
    hasher = hashlib.sha256()
    total_bytes = 0
    with stored_path.open("wb") as handle:
        while True:
            chunk = upload_file.file.read(1024 * 1024)
            if not chunk:
                break
            total_bytes += len(chunk)
            if total_bytes > MAX_UPLOAD_BYTES:
                handle.close()
                stored_path.unlink(missing_ok=True)
                raise HTTPException(
                    status_code=413,
                    detail=f"The {label.lower()} file is too large. Each file must be no larger than {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
                )
            hasher.update(chunk)
            handle.write(chunk)
    try:
        validate_stored_upload_magic(stored_path, allowed_suffixes, label)
    except HTTPException:
        stored_path.unlink(missing_ok=True)
        raise
    upload_file.file.seek(0)
    return stored_path, hasher.hexdigest()


IMAGE_SOURCE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
PDF_SOURCE_SUFFIXES = {".pdf"}
SOURCE_ARCHIVE_BUILDING_DIR = "by_building"
SOURCE_ARCHIVE_NEEDS_REVIEW_DIR = "_needs_review"


def list_source_image_files(stored_path: Optional[Union[str, Path]]) -> List[Path]:
    if not stored_path:
        return []
    try:
        resolved = Path(stored_path).resolve()
    except Exception:
        return []
    if resolved.is_file():
        return [resolved]
    if resolved.is_dir():
        return sorted(
            [
                item
                for item in resolved.iterdir()
                if item.is_file() and item.suffix.lower() in IMAGE_SOURCE_SUFFIXES
            ],
            key=lambda item: item.name.lower(),
        )
    return []


def list_source_pdf_files(stored_path: Optional[Union[str, Path]]) -> List[Path]:
    if not stored_path:
        return []
    try:
        resolved = Path(stored_path).resolve()
    except Exception:
        return []
    if resolved.is_file() and resolved.suffix.lower() in PDF_SOURCE_SUFFIXES:
        return [resolved]
    if resolved.is_dir():
        return sorted(
            [
                item
                for item in resolved.iterdir()
                if item.is_file() and item.suffix.lower() in PDF_SOURCE_SUFFIXES
            ],
            key=lambda item: item.name.lower(),
        )
    return []


def list_source_files(stored_path: Optional[Union[str, Path]]) -> List[Path]:
    return list_source_image_files(stored_path) or list_source_pdf_files(stored_path)


def path_is_within(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def source_archive_record_dir_name(source_document_id: str) -> str:
    return source_document_id if source_document_id.startswith("source_") else f"source_{source_document_id}"


def source_archive_slug(*parts: object) -> str:
    raw = " ".join(normalize_unknown_value(part) for part in parts if normalize_unknown_value(part))
    cleaned = re.sub(r"[^\w\u4e00-\u9fff]+", "_", raw.lower()).strip("_")
    return (cleaned[:120].strip("_") or "unknown")


def unique_archive_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for index in range(2, 1000):
        candidate = path.with_name(f"{stem}_{index}{suffix}")
        if not candidate.exists():
            return candidate
    return path.with_name(f"{stem}_{uuid.uuid4().hex[:8]}{suffix}")


def archive_target_for_source(
    conn: sqlite3.Connection,
    parse_result: Optional[dict],
    *,
    source_document_id: str,
    force_needs_review: bool = False,
) -> tuple[Path, dict]:
    sources_root = UPLOAD_ROOT / "sources"
    record_dir_name = source_archive_record_dir_name(source_document_id)
    result = parse_result or {}
    target_staging_key = normalize_unknown_value(result.get("target_staging_key")) or ""
    staging_key = normalize_unknown_value(result.get("staging_key")) or ""
    building_id = normalize_unknown_value(result.get("building_id")) or ""
    building_name = normalize_unknown_value(result.get("building_name")) or ""
    address = ""

    supplement_without_target = (
        result.get("intake_mode") == INTAKE_MODE_SUPPLEMENT
        and not target_staging_key
        and not staging_key
    )
    if force_needs_review or supplement_without_target:
        return (
            sources_root / SOURCE_ARCHIVE_NEEDS_REVIEW_DIR / record_dir_name,
            {
                "archive_status": "needs_review",
                "archive_label": "Awaiting archive",
                "building_name": building_name,
                "building_key": "",
            },
        )

    resolved_key = target_staging_key or staging_key
    if resolved_key:
        snapshot = load_staging_building_snapshot(conn, resolved_key)
        if snapshot:
            building_name = normalize_unknown_value(snapshot.get("building_name")) or building_name
            address = normalize_unknown_value(snapshot.get("address")) or ""
    elif building_id:
        snapshot = load_master_building_snapshot(conn, building_id)
        if snapshot:
            building_name = normalize_unknown_value(snapshot.get("building_name")) or building_name
            address = normalize_unknown_value(snapshot.get("address")) or ""

    if not building_name:
        return (
            sources_root / SOURCE_ARCHIVE_NEEDS_REVIEW_DIR / record_dir_name,
            {
                "archive_status": "needs_review",
                "archive_label": "Awaiting archive",
                "building_name": "",
                "building_key": "",
            },
        )

    slug = source_archive_slug(building_name, address, resolved_key or building_id)
    return (
        sources_root / SOURCE_ARCHIVE_BUILDING_DIR / slug / record_dir_name,
        {
            "archive_status": "archived",
            "archive_label": f"Archived by building: {building_name}",
            "building_name": building_name,
            "building_address": address,
            "building_key": resolved_key or building_id,
            "building_slug": slug,
        },
    )


def update_source_archive_artifact(
    conn: sqlite3.Connection,
    source_document_id: str,
    archive_info: dict,
) -> None:
    row = conn.execute(
        "SELECT parse_artifacts_json FROM source_documents WHERE id = ?",
        (source_document_id,),
    ).fetchone()
    artifacts = json_loads_safe(row.get("parse_artifacts_json") if row else "{}", {})
    if not isinstance(artifacts, dict):
        artifacts = {}
    artifacts["file_archive"] = archive_info
    conn.execute(
        "UPDATE source_documents SET parse_artifacts_json = ?, updated_at = ? WHERE id = ?",
        (json_dumps(artifacts), now_iso(), source_document_id),
    )


def archive_source_document_files(
    conn: sqlite3.Connection,
    source_document_id: str,
    parse_result: Optional[dict],
    *,
    actor: Actor,
    force_needs_review: bool = False,
) -> Optional[str]:
    row = source_document_row_for_job(conn, source_document_id)
    if not row:
        return None
    stored_path_raw = normalize_unknown_value(row.get("stored_path"))
    if not stored_path_raw:
        return None
    source_path = Path(stored_path_raw)
    if not source_path.exists():
        return f"The source file does not exist and could not be archived: {stored_path_raw}"

    sources_root = (UPLOAD_ROOT / "sources").resolve()
    building_archive_root = (UPLOAD_ROOT / "sources" / SOURCE_ARCHIVE_BUILDING_DIR).resolve()
    needs_review_root = (UPLOAD_ROOT / "sources" / SOURCE_ARCHIVE_NEEDS_REVIEW_DIR).resolve()
    if path_is_within(source_path, building_archive_root) or path_is_within(source_path, needs_review_root):
        return None

    try:
        target_dir, archive_info = archive_target_for_source(
            conn,
            parse_result,
            source_document_id=source_document_id,
            force_needs_review=force_needs_review,
        )
        target_dir.mkdir(parents=True, exist_ok=True)
        original_parent = source_path.parent if source_path.is_file() else source_path
        moved_files: List[str] = []

        if source_path.is_file():
            target_file = unique_archive_path(target_dir / source_path.name)
            shutil.move(str(source_path), str(target_file))
            moved_files.append(str(target_file))
            new_stored_path = target_file
        else:
            for child in sorted(source_path.iterdir(), key=lambda item: item.name.lower()):
                if not child.is_file() or child.name == ".DS_Store":
                    continue
                target_file = unique_archive_path(target_dir / child.name)
                shutil.move(str(child), str(target_file))
                moved_files.append(str(target_file))
            new_stored_path = target_dir

        if original_parent.exists() and path_is_within(original_parent, sources_root):
            try:
                original_parent.rmdir()
            except OSError:
                pass

        if not moved_files:
            return f"The source directory is empty; no files were moved: {stored_path_raw}"

        archive_info.update(
            {
                "archive_path": str(new_stored_path),
                "file_count": len(moved_files),
                "archived_at": now_iso(),
            }
        )
        conn.execute(
            "UPDATE source_documents SET stored_path = ?, updated_at = ? WHERE id = ?",
            (str(new_stored_path), now_iso(), source_document_id),
        )
        update_source_archive_artifact(conn, source_document_id, archive_info)
        write_audit_log(
            conn,
            None,
            actor,
            action_type="source_document_archived",
            target_table="source_documents",
            target_record_id=source_document_id,
            building_name=archive_info.get("building_name") or "",
            source=row.get("source_file") or "",
            note=archive_info.get("archive_label") or "",
        )
        return None
    except Exception as exc:
        warning = f"Source-file archiving failed: {exc}"
        update_source_archive_artifact(
            conn,
            source_document_id,
            {
                "archive_status": "failed",
                "archive_label": "Archive failed",
                "archive_error": warning,
                "archive_path": stored_path_raw,
                "archived_at": now_iso(),
            },
        )
        write_audit_log(
            conn,
            None,
            actor,
            action_type="source_document_archive_failed",
            target_table="source_documents",
            target_record_id=source_document_id,
            source=row.get("source_file") or "",
            note=warning,
        )
        return warning


def build_source_file_url(stored_path: Optional[Union[str, Path]]) -> str:
    if not stored_path:
        return ""
    try:
        resolved = Path(stored_path).resolve()
        if resolved.is_dir():
            source_files = list_source_files(resolved)
            if not source_files:
                return ""
            resolved = source_files[0]
        relative = resolved.relative_to(UPLOAD_ROOT.resolve())
    except Exception:
        return ""
    return "/source-files/" + "/".join(relative.parts)


def build_source_file_urls(stored_path: Optional[Union[str, Path]]) -> List[str]:
    if not stored_path:
        return []
    try:
        resolved = Path(stored_path).resolve()
        if resolved.is_dir():
            return [build_source_file_url(item) for item in list_source_files(resolved)]
    except Exception:
        return []
    url = build_source_file_url(stored_path)
    return [url] if url else []


def resolve_protected_file(root: Path, relative_path: str) -> Path:
    try:
        root_resolved = root.resolve()
        target = (root_resolved / relative_path).resolve()
        target.relative_to(root_resolved)
    except Exception:
        raise HTTPException(status_code=404, detail="The file does not exist.")
    if not target.is_file():
        raise HTTPException(status_code=404, detail="The file does not exist.")
    return target


def path_contains_path(container: Path, target: Path) -> bool:
    try:
        container_resolved = container.resolve()
        target_resolved = target.resolve()
        if container_resolved == target_resolved:
            return True
        target_resolved.relative_to(container_resolved)
        return True
    except Exception:
        return False


def normalize_intake_mode(value: object) -> str:
    normalized = (normalize_unknown_value(value) or INTAKE_MODE_FULL_PACKAGE).strip().lower()
    return normalized if normalized in INTAKE_MODES else INTAKE_MODE_FULL_PACKAGE


def normalize_supplement_scope(value: object) -> str:
    normalized = (normalize_unknown_value(value) or SUPPLEMENT_SCOPE_ALL).strip().lower()
    return normalized if normalized in SUPPLEMENT_SCOPES else SUPPLEMENT_SCOPE_ALL


def build_intake_source_metadata(
    *,
    intake_mode: str,
    supplement_scope: str,
    target_staging_key: str,
    source_kind: str = "",
    case_id: str = "",
    communication_event_id: str = "",
    captured_at: str = "",
) -> Dict[str, str]:
    mode = normalize_intake_mode(intake_mode)
    if mode != INTAKE_MODE_SUPPLEMENT:
        return {}
    metadata = {
        "intake_mode": INTAKE_MODE_SUPPLEMENT,
        "supplement_scope": normalize_supplement_scope(supplement_scope),
        "target_staging_key": normalize_unknown_value(target_staging_key) or "",
    }
    for key, value in {
        "source_kind": source_kind,
        "case_id": case_id,
        "communication_event_id": communication_event_id,
        "captured_at": captured_at,
    }.items():
        normalized = normalize_unknown_value(value)
        if normalized:
            metadata[key] = normalized
    return metadata


def parse_intake_source_metadata(value: object) -> Dict[str, str]:
    data = json_loads_safe(value, {})
    if not isinstance(data, dict):
        return {}
    mode = normalize_intake_mode(data.get("intake_mode"))
    if mode != INTAKE_MODE_SUPPLEMENT:
        return {}
    return build_intake_source_metadata(
        intake_mode=mode,
        supplement_scope=data.get("supplement_scope"),
        target_staging_key=data.get("target_staging_key"),
        source_kind=data.get("source_kind"),
        case_id=data.get("case_id"),
        communication_event_id=data.get("communication_event_id"),
        captured_at=data.get("captured_at"),
    )


def read_intake_source_options(value: object) -> Dict[str, str]:
    data = json_loads_safe(value, {})
    if not isinstance(data, dict):
        return build_intake_source_metadata(
            intake_mode=INTAKE_MODE_FULL_PACKAGE,
            supplement_scope=SUPPLEMENT_SCOPE_ALL,
            target_staging_key="",
        )
    return build_intake_source_metadata(
        intake_mode=data.get("intake_mode"),
        supplement_scope=data.get("supplement_scope"),
        target_staging_key=data.get("target_staging_key"),
        source_kind=data.get("source_kind"),
        case_id=data.get("case_id"),
        communication_event_id=data.get("communication_event_id"),
        captured_at=data.get("captured_at"),
    )


def review_group_intake_metadata(rows: List[dict]) -> Dict[str, str]:
    for row in rows:
        metadata = parse_intake_source_metadata(row.get("source_content"))
        if metadata:
            if not metadata.get("target_staging_key"):
                metadata["target_staging_key"] = normalize_unknown_value(row.get("target_staging_key")) or ""
            return metadata
    return {}


def filter_parsed_payloads_for_supplement_scope(
    parsed: Dict[str, dict],
    supplement_scope: str,
) -> Dict[str, dict]:
    scope = normalize_supplement_scope(supplement_scope)
    allowed_fields = SUPPLEMENT_SCOPE_FIELD_KEYS.get(scope, SUPPLEMENT_SCOPE_FIELD_KEYS[SUPPLEMENT_SCOPE_ALL])
    return {
        field_key: payload
        for field_key, payload in (parsed or {}).items()
        if field_key in allowed_fields and field_key not in SUPPLEMENT_IDENTITY_FIELD_KEYS
    }


def filter_chat_building_payloads(parsed: Dict[str, dict]) -> Dict[str, dict]:
    building_contact_markers = (
        "物业",
        "管理处",
        "前台",
        "礼宾",
        "维修",
        "building management",
        "property management",
        "front desk",
        "concierge",
        "maintenance",
        "leasing office",
    )
    public_service_markers = (
        *building_contact_markers,
        "pseg",
        "pse&g",
        "con edison",
        "verizon",
        "xfinity",
        "spectrum",
        "astound",
        "honest networks",
        "official",
        "官网",
        "客服",
    )
    private_markers = (
        "验证码",
        "短信码",
        "身份证",
        "护照",
        "社会安全号",
        "verification code",
        "one-time password",
        " otp ",
        "passport",
        "social security",
        "driver license",
    )
    filtered: Dict[str, dict] = {}
    for field_key, payload in (parsed or {}).items():
        if field_key not in CHAT_BUILDING_FACT_FIELD_KEYS or not isinstance(payload, dict):
            continue
        evidence_text = " ".join(
            [
                normalize_unknown_value(payload.get("value")) or "",
                normalize_unknown_value(payload.get("evidence")) or "",
                *[
                    normalize_unknown_value(item.get("quote")) or ""
                    for item in (payload.get("evidence_items") or [])
                    if isinstance(item, dict)
                ],
            ]
        ).lower()
        if any(marker in evidence_text for marker in private_markers):
            continue
        contains_contact_value = bool(
            re.search(r"\b[\w.+-]+@[\w.-]+\.[a-z]{2,}\b", evidence_text, flags=re.IGNORECASE)
            or re.search(r"(?<!\d)(?:\+?1[-.\s]?)?\(?\d{3}\)?[-.\s]\d{3}[-.\s]\d{4}(?!\d)", evidence_text)
        )
        if contains_contact_value and not any(marker in evidence_text for marker in public_service_markers):
            continue
        if field_key in BUILDING_CONTACT_FIELD_KEYS and not any(
            marker in evidence_text for marker in building_contact_markers
        ):
            continue
        cloned = dict(payload)
        flags = list(cloned.get("review_flags") or [])
        if "chat_source" not in flags:
            flags.append("chat_source")
        cloned["review_flags"] = flags
        filtered[field_key] = cloned
    return filtered


def actor_can_access_source_path(conn: sqlite3.Connection, actor: Actor, target_path: Path) -> bool:
    if actor.role in {"super_admin", "admin"}:
        return True
    if actor.role != "employee":
        return False
    rows = conn.execute(
        """
        SELECT stored_path
        FROM source_documents
        WHERE created_by = ? AND stored_path != ''
        """,
        (actor.user_id,),
    ).fetchall()
    return any(path_contains_path(Path(row["stored_path"]), target_path) for row in rows)


def require_source_path_access(conn: sqlite3.Connection, actor: Actor, target_path: Path) -> None:
    if not actor_can_access_source_path(conn, actor, target_path):
        raise HTTPException(status_code=403, detail="You do not have permission to access this source file.")


def serialize_source_document(row: Optional[dict]) -> Optional[dict]:
    if not row:
        return None
    item = dict(row)
    source_metadata = parse_intake_source_metadata(item.get("source_content"))
    if source_metadata:
        item["source_metadata"] = source_metadata
        item["source_content"] = ""
    item["source_url"] = build_source_file_url(item.get("stored_path"))
    item["source_urls"] = build_source_file_urls(item.get("stored_path"))
    item["pages"] = json_loads_safe(item.get("extracted_pages_json"), [])
    item["parse_artifacts"] = json_loads_safe(item.get("parse_artifacts_json"), {})
    item["file_archive"] = item["parse_artifacts"].get("file_archive") if isinstance(item["parse_artifacts"], dict) else {}
    return item


def serialize_intake_job(row: Optional[dict]) -> Optional[dict]:
    if not row:
        return None
    item = dict(row)
    source_metadata = read_intake_source_options(item.get("source_content"))
    return {
        "source_document_id": item.get("id"),
        "source_file": item.get("source_file") or "",
        "raw_input_type": item.get("raw_input_type") or "",
        "parser_type": item.get("parser_type") or "",
        "source_type": item.get("source_type") or "",
        "parse_status": item.get("parse_status") or INTAKE_PARSE_STATUS_COMPLETED,
        "parse_started_at": item.get("parse_started_at") or "",
        "parse_completed_at": item.get("parse_completed_at") or "",
        "parse_error": item.get("parse_error") or "",
        "submission_group_id": item.get("submission_group_id") or "",
        "created_by": item.get("created_by") or "",
        "created_at": item.get("created_at") or "",
        "updated_at": item.get("updated_at") or "",
        "source_metadata": source_metadata,
    }


def serialize_llm_call_log(row: dict) -> dict:
    item = dict(row)
    item["user_payload"] = json_loads_safe(item.get("user_payload_json"), {})
    item["parsed_response"] = json_loads_safe(item.get("parsed_response_json"), {})
    return item


def lookup_master_building(conn: sqlite3.Connection, building_name: str, address: str = "") -> Optional[dict]:
    rows = conn.execute(
        "SELECT * FROM master_building_info ORDER BY building_name ASC"
    ).fetchall()
    name_norm = legacy._normalize_text(building_name)
    address_norm = legacy._normalize_text(address)
    if not name_norm:
        return None
    for row in rows:
        row_name_norm = legacy._normalize_text(row["building_name"])
        row_address_norm = legacy._normalize_text(row["address"] or "")
        if name_norm != row_name_norm:
            continue
        if address_norm:
            if address_norm == row_address_norm:
                return row
            continue
        if not row_address_norm:
            return row
    return None


def get_master_field_value(conn: sqlite3.Connection, building_id: str, field_key: str) -> Optional[str]:
    if field_key in CORE_MASTER_FIELD_KEYS:
        row = conn.execute(
            f"SELECT {field_key} AS value FROM master_building_info WHERE id = ?",
            (building_id,),
        ).fetchone()
        if not row:
            return None
        value = row["value"]
        if isinstance(value, int) and field_key in {
            "insurance_required",
            "electricity_required",
            "internet_self_setup_required",
        }:
            if value == 2:
                return "optional"
            return "true" if value else "false"
        return normalize_unknown_value(value)

    row = conn.execute(
        """
        SELECT value_text
        FROM master_building_field_values
        WHERE building_id = ? AND field_key = ?
        """,
        (building_id, field_key),
    ).fetchone()
    if not row:
        return None
    return normalize_unknown_value(row["value_text"])


def get_staging_snapshot_field_value(snapshot: Optional[dict], field_key: str) -> Optional[str]:
    if not snapshot:
        return None
    if field_key in {
        "insurance_required",
        "electricity_required",
        "internet_self_setup_required",
    }:
        return normalize_requirement_choice(snapshot.get(field_key))
    if field_key in {
        "building_name",
        "address",
        "insurance_coverage_amount",
        "electricity_provider",
        "internet_provider",
        "internet_notes",
        "move_in_notes",
        "source_type",
        "source_file",
        "source_date",
        "info_cutoff_date",
    }:
        return normalize_unknown_value(snapshot.get(field_key))
    extensions = snapshot.get("extensions") or {}
    if field_key in NETWORK_PROVIDER_FIELD_MAP:
        return normalize_field_value(field_key, "boolean", extensions.get(field_key))
    return normalize_unknown_value(extensions.get(field_key))


def missing_required_detail_for_group(values_by_field: Dict[str, Optional[str]]) -> bool:
    insurance_required = normalize_requirement_choice(values_by_field.get("insurance_required"))
    electricity_required = normalize_requirement_choice(values_by_field.get("electricity_required"))
    if insurance_required == "true" and not normalize_unknown_value(values_by_field.get("insurance_coverage_amount")):
        return True
    if electricity_required == "true" and not normalize_unknown_value(values_by_field.get("electricity_provider")):
        return True
    return False


def missing_required_detail_for_field(field_key: str, values_by_field: Dict[str, Optional[str]]) -> bool:
    insurance_required = normalize_requirement_choice(values_by_field.get("insurance_required"))
    electricity_required = normalize_requirement_choice(values_by_field.get("electricity_required"))
    renters_required = normalize_insurance_status(values_by_field.get("insurance_renters_required"))
    liability_required = normalize_insurance_status(values_by_field.get("insurance_personal_liability_required"))
    if field_key in {"insurance_coverage_amount", "insurance_renters_minimum_coverage"}:
        if insurance_required == "true" and not normalize_unknown_value(values_by_field.get("insurance_coverage_amount")):
            return True
        if (
            renters_required == "yes"
            and not normalize_unknown_value(values_by_field.get("insurance_renters_minimum_coverage"))
            and not normalize_unknown_value(values_by_field.get("insurance_personal_liability_per_occurrence"))
        ):
            return True
    if field_key == "electricity_provider":
        return electricity_required == "true" and not normalize_unknown_value(values_by_field.get("electricity_provider"))
    if field_key in {"insurance_personal_liability_per_occurrence", "insurance_personal_liability_aggregate"}:
        return liability_required == "yes" and not (
            normalize_unknown_value(values_by_field.get("insurance_personal_liability_per_occurrence"))
            or normalize_unknown_value(values_by_field.get("insurance_personal_liability_aggregate"))
        )
    return False


def manual_review_reasons_for_group(values_by_field: Dict[str, Optional[str]]) -> List[str]:
    reasons: List[str] = []
    insurance_required = normalize_requirement_choice(values_by_field.get("insurance_required"))
    electricity_required = normalize_requirement_choice(values_by_field.get("electricity_required"))
    renters_required = normalize_insurance_status(values_by_field.get("insurance_renters_required"))
    renters_coverage = normalize_unknown_value(values_by_field.get("insurance_renters_minimum_coverage"))
    liability_required = normalize_insurance_status(values_by_field.get("insurance_personal_liability_required"))
    liability_per_occurrence = normalize_unknown_value(
        values_by_field.get("insurance_personal_liability_per_occurrence")
    )
    liability_aggregate = normalize_unknown_value(values_by_field.get("insurance_personal_liability_aggregate"))
    if insurance_required == "true" and not normalize_unknown_value(values_by_field.get("insurance_coverage_amount")):
        reasons.append("Insurance is required, but the general minimum renters coverage is not stated clearly.")
    if electricity_required == "true" and not normalize_unknown_value(values_by_field.get("electricity_provider")):
        reasons.append("Electricity setup is required, but the utility provider is not stated clearly.")
    if renters_required == "yes" and not renters_coverage and not liability_per_occurrence:
        reasons.append("Renters Insurance is required, but no minimum coverage amount could be mapped directly.")
    if liability_required == "yes" and not liability_per_occurrence and not liability_aggregate:
        reasons.append("Personal Liability is required, but the per-occurrence and aggregate limits remain unclear.")
    if liability_aggregate and not liability_per_occurrence:
        reasons.append("An aggregate liability limit was found, but the per-occurrence limit is unclear.")
    return reasons


def create_staging_request(
    conn: sqlite3.Connection,
    *,
    submission_group_id: str,
    building_name: str,
    building_id: Optional[str],
    field_name: str,
    old_value: Optional[str],
    new_value: Optional[str],
    source_type: str,
    source_content: str,
    source_file: str,
    submitted_by: str,
    ai_confidence: Optional[float],
    review_status: str,
    import_batch_id: Optional[str],
    parser_type: str,
    raw_input_type: str,
    source_document_id: Optional[str],
    conflict_with_long_term: bool,
    low_confidence: bool,
    missing_required_detail: bool,
    approval_stage: str = "to_staging",
    target_staging_key: str = "",
    evidence_json: Optional[List[dict]] = None,
    manual_review_reason: str = "",
    review_flags_json: Optional[List[str]] = None,
    priority: str = "normal",
) -> None:
    record_id = f"stage_{uuid.uuid4().hex}"
    created_at = now_iso()
    conn.execute(
        """
        INSERT INTO staging_update_requests(
          id, record_id, submission_group_id, building_name, building_id, field_name, old_value, new_value,
          normalized_new_value, source_type, source_content, source_file, source_document_id, approval_stage, target_staging_key,
          submitted_by, submitted_at, ai_confidence, review_status, reviewer, reviewed_at,
          review_comment, conflict_with_long_term, priority, import_batch_id, parser_type, raw_input_type,
          low_confidence, missing_required_detail, evidence_json, manual_review_reason, review_flags_json, created_at, updated_at
        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '', NULL, '', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            record_id,
            record_id,
            submission_group_id,
            building_name,
            building_id,
            field_name,
            old_value or "",
            new_value or "",
            new_value or "",
            source_type,
            source_content,
            source_file,
            source_document_id,
            normalize_review_approval_stage(approval_stage),
            target_staging_key or "",
            submitted_by,
            created_at,
            ai_confidence,
            review_status,
            1 if conflict_with_long_term else 0,
            priority,
            import_batch_id,
            parser_type,
            raw_input_type,
            1 if low_confidence else 0,
            1 if missing_required_detail else 0,
            json_dumps(evidence_json or []),
            manual_review_reason.strip(),
            json_dumps(review_flags_json or []),
            created_at,
            created_at,
        ),
    )


def persist_import_to_staging(
    conn: sqlite3.Connection,
    *,
    batch_id: str,
    file_name: str,
    sheets: List[SheetImportConfirmation],
    file_path: Path,
    actor: Actor,
    request: Request,
    source_type: str = "excel_import",
    parser_type: str = "excel_header_mapping",
    raw_input_type: str = "excel",
) -> dict:
    workbook = {sheet_name: rows for sheet_name, rows in legacy._load_tabular_sheet_rows(file_path)}
    imported_rows = 0
    imported_fields = 0

    for sheet in sheets:
        rows = workbook.get(sheet.sheet_name, [])
        if not rows:
            continue
        header_row = rows[sheet.header_row_index] if sheet.header_row_index < len(rows) else []
        header_values = [legacy._format_spreadsheet_value(value) for value in header_row]
        mapping_by_header = {}

        for mapping in sheet.mappings:
            if mapping.action == "ignore":
                mapped_field_key = None
                match_method = "manual_selected"
            elif mapping.action == "create":
                if actor.role != "super_admin":
                    raise HTTPException(status_code=403, detail="Only a Super Admin can create a new field during import.")
                display_name = mapping.new_field_display_name or mapping.original_header
                field_key = re.sub(r"[^a-z0-9_]+", "_", mapping.mapped_field_key or mapping.original_header.lower()).strip("_")
                field_key = field_key or f"field_{uuid.uuid4().hex[:8]}"
                conn.execute(
                    """
                    INSERT OR IGNORE INTO field_definitions(
                      id, field_key, display_name, field_type, required, description, is_core, active, created_by, created_at, updated_at
                    ) VALUES(?, ?, ?, ?, 0, 'Created from import mapping', 0, 1, ?, ?, ?)
                    """,
                    (f"field_{field_key}", field_key, display_name, mapping.field_type or "text", actor.user_id, now_iso(), now_iso()),
                )
                conn.execute(
                    """
                    INSERT OR IGNORE INTO field_aliases(id, field_key, alias_name, language, confidence, created_by, created_at)
                    VALUES(?, ?, ?, 'mixed', 1.0, ?, ?)
                    """,
                    (f"alias_{field_key}_{uuid.uuid4().hex[:8]}", field_key, mapping.original_header, actor.user_id, now_iso()),
                )
                mapped_field_key = field_key
                match_method = "manual_selected"
            else:
                mapped_field_key = mapping.mapped_field_key
                match_method = "manual_selected"

            mapping_by_header[mapping.original_header] = mapped_field_key
            conn.execute(
                """
                INSERT INTO import_header_mappings(
                  id, import_batch_id, sheet_name, original_header, mapped_field_key, match_method, confidence, confirmed_by_admin, ignored, created_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"mapping_{uuid.uuid4().hex}",
                    batch_id,
                    sheet.sheet_name,
                    mapping.original_header,
                    mapped_field_key,
                    match_method,
                    1.0 if mapping.action != "ignore" else 0.0,
                    actor.user_id,
                    1 if mapping.action == "ignore" else 0,
                    now_iso(),
                ),
            )

        for row_number, raw_row in enumerate(rows[sheet.header_row_index + 1 :], start=sheet.header_row_index + 2):
            formatted_row = [legacy._format_spreadsheet_value(value) for value in raw_row]
            if not any(formatted_row):
                continue

            row_values: Dict[str, Optional[str]] = {}
            building_name = ""
            address = ""
            for index, original_header in enumerate(header_values):
                if not original_header:
                    continue
                field_key = mapping_by_header.get(original_header)
                if not field_key:
                    continue
                field_definition = find_field_definition(conn, field_key)
                if not field_definition:
                    continue
                raw_value = formatted_row[index] if index < len(formatted_row) else ""
                normalized_value = normalize_field_value(field_key, field_definition["field_type"], raw_value)
                row_values[field_key] = normalized_value
                if field_key == "building_name" and normalized_value:
                    building_name = normalized_value
                if field_key == "address" and normalized_value:
                    address = normalized_value

            if not building_name:
                continue

            imported_rows += 1
            submission_group_id = f"import:{batch_id}:{sheet.sheet_name}:{row_number}"
            master_match = lookup_master_building(conn, building_name, address)
            building_id = master_match["id"] if master_match else None
            row_source = json_dumps(
                {
                    "sheet_name": sheet.sheet_name,
                    "row_number": row_number,
                    "headers": header_values,
                    "values": formatted_row,
                }
            )
            for field_key, normalized_value in row_values.items():
                if normalized_value is None:
                    continue
                old_value = get_master_field_value(conn, building_id, field_key) if building_id else None
                create_staging_request(
                    conn,
                    submission_group_id=submission_group_id,
                    building_name=building_name,
                    building_id=building_id,
                    field_name=field_key,
                    old_value=old_value,
                    new_value=normalized_value,
                    source_type=source_type,
                    source_content=row_source,
                    source_file=file_name,
                    submitted_by=actor.user_id,
                    ai_confidence=None,
                    review_status="pending",
                    import_batch_id=batch_id,
                    parser_type=parser_type,
                    raw_input_type=raw_input_type,
                    source_document_id=None,
                    conflict_with_long_term=bool(old_value and normalized_value != old_value),
                    low_confidence=False,
                    missing_required_detail=missing_required_detail_for_field(field_key, row_values),
                )
                imported_fields += 1

    conn.execute(
        "UPDATE import_batches SET status = 'imported', confirmed_by = ?, updated_at = ? WHERE id = ?",
        (actor.user_id, now_iso(), batch_id),
    )
    write_audit_log(
        conn,
        request,
        actor,
        action_type="excel_import_confirmed",
        target_table="import_batches",
        target_record_id=batch_id,
        source=file_name,
        note=f"Imported {imported_rows} rows / {imported_fields} fields into staging.",
    )
    return {"rows": imported_rows, "fields": imported_fields}


def supported_bool_to_int(value: Optional[str]) -> Optional[int]:
    bool_value = normalize_booleanish(value)
    if bool_value is None:
        return None
    return 1 if bool_value else 0


def requirement_choice_to_int(value: Optional[str]) -> Optional[int]:
    choice = normalize_requirement_choice(value)
    if choice == "true":
        return 1
    if choice == "false":
        return 0
    if choice == "optional":
        return 2
    return None


def requirement_state(value: object) -> str:
    choice = normalize_requirement_choice(value)
    if choice == "true":
        return "required"
    if choice == "false":
        return "not_required"
    if choice == "optional":
        return "optional"
    if value == 1:
        return "required"
    if value == 0:
        return "not_required"
    if value == 2:
        return "optional"
    return "unknown"


def calculate_master_completeness(snapshot: dict) -> dict:
    extensions = snapshot.get("extensions") or {}
    score = 0
    building_name = normalize_unknown_value(snapshot.get("building_name"))
    address = normalize_unknown_value(snapshot.get("address"))
    score += 10 if building_name else 0
    score += 10 if address else 0

    insurance_state = requirement_state(snapshot.get("insurance_required"))
    insurance_detail = any(
        normalize_unknown_value(value)
        for value in (
            snapshot.get("insurance_coverage_amount"),
            extensions.get("insurance_renters_minimum_coverage"),
            extensions.get("insurance_personal_liability_per_occurrence"),
            extensions.get("insurance_submission_method"),
        )
    )
    insurance_complete = insurance_state != "unknown" and (
        insurance_state != "required" or insurance_detail
    )
    if insurance_state != "unknown":
        score += 15
        score += 10 if insurance_complete else 0

    electricity_state = requirement_state(snapshot.get("electricity_required"))
    electricity_provider = normalize_unknown_value(snapshot.get("electricity_provider"))
    electricity_complete = electricity_state != "unknown" and (
        electricity_state != "required" or bool(electricity_provider)
    )
    if electricity_state != "unknown":
        score += 15
        score += 10 if electricity_complete else 0

    internet_state = requirement_state(snapshot.get("internet_self_setup_required"))
    internet_detail = any(
        normalize_unknown_value(value)
        for value in (
            snapshot.get("internet_provider"),
            snapshot.get("internet_notes"),
            *[extensions.get(field_key) for field_key in NETWORK_PROVIDER_FIELD_MAP],
        )
    )
    internet_complete = internet_state != "unknown" and (
        internet_state != "required" or internet_detail
    )
    if internet_state != "unknown":
        score += 10
        score += 10 if internet_complete else 0

    move_in_complete = any(
        normalize_unknown_value(value)
        for value in (
            snapshot.get("move_in_notes"),
            extensions.get("key_pickup_notes"),
            extensions.get("service_elevator_booking_notes"),
        )
    )
    score += 10 if move_in_complete else 0

    complete = bool(
        building_name
        and address
        and insurance_complete
        and electricity_complete
        and internet_complete
        and score >= 80
    )
    return {
        "completeness_status": "verified_complete" if complete else "verified_partial",
        "completeness_score": min(100, score),
        "completeness_dimensions": {
            "identity": bool(building_name and address),
            "insurance": insurance_complete,
            "electricity": electricity_complete,
            "internet": internet_complete,
            "move_in": move_in_complete,
        },
    }


def refresh_master_completeness(
    conn: sqlite3.Connection,
    building_id: str,
    *,
    verification_note: Optional[str] = None,
) -> dict:
    snapshot = load_master_building_snapshot(conn, building_id)
    if not snapshot:
        raise HTTPException(status_code=404, detail="The Master building record does not exist.")
    completeness = calculate_master_completeness(snapshot)
    assignments = ["completeness_status = ?", "completeness_score = ?"]
    params: List[Any] = [
        completeness["completeness_status"],
        completeness["completeness_score"],
    ]
    if verification_note is not None:
        assignments.append("verification_note = ?")
        params.append(verification_note.strip())
    assignments.append("updated_at = ?")
    params.extend([now_iso(), building_id])
    conn.execute(
        f"UPDATE master_building_info SET {', '.join(assignments)} WHERE id = ?",
        params,
    )
    return completeness


def ensure_master_building(
    conn: sqlite3.Connection,
    *,
    building_name: str,
    address: str,
    actor: Actor,
) -> dict:
    existing = lookup_master_building(conn, building_name, address)
    if existing:
        return existing
    created_at = now_iso()
    building_id = f"master_{uuid.uuid4().hex}"
    conn.execute(
        """
        INSERT INTO master_building_info(
          id, building_name, address, updated_by, version, status, created_at, updated_at
        ) VALUES(?, ?, ?, ?, 1, 'active', ?, ?)
        """,
        (building_id, building_name, address or "", actor.user_id, created_at, created_at),
    )
    return conn.execute("SELECT * FROM master_building_info WHERE id = ?", (building_id,)).fetchone()


def upsert_master_field(
    conn: sqlite3.Connection,
    *,
    building_id: str,
    field_key: str,
    new_value: Optional[str],
    actor: Actor,
    source_type: str,
    source_file: str,
    source_date: Optional[str],
    info_cutoff_date: Optional[str],
) -> tuple[Optional[str], Optional[str]]:
    old_value = get_master_field_value(conn, building_id, field_key)
    timestamp = now_iso()
    effective_source_date = source_date or iso_now_local_date()

    if field_key in CORE_MASTER_FIELD_KEYS:
        if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
            stored_value: Any = requirement_choice_to_int(new_value)
        elif field_key == "internet_provider":
            stored_value = normalize_provider_text(new_value)
        else:
            stored_value = normalize_unknown_value(new_value)
        assignments = [f"{field_key} = ?", "updated_by = ?", "updated_at = ?", "last_verified_at = ?"]
        params: List[Any] = [stored_value, actor.user_id, timestamp, timestamp]
        if field_key != "source_type":
            assignments.append("source_type = ?")
            params.append(source_type or None)
        if field_key != "source_file":
            assignments.append("source_file = ?")
            params.append(source_file or None)
        if field_key != "source_date":
            assignments.append("source_date = ?")
            params.append(effective_source_date)
        if field_key != "info_cutoff_date":
            assignments.append("info_cutoff_date = ?")
            params.append(info_cutoff_date or None)
        params.append(building_id)
        conn.execute(
            f"UPDATE master_building_info SET {', '.join(assignments)} WHERE id = ?",
            params,
        )
        return old_value, normalize_unknown_value(new_value)

    existing = conn.execute(
        """
        SELECT id, value_text, version
        FROM master_building_field_values
        WHERE building_id = ? AND field_key = ?
        """,
        (building_id, field_key),
    ).fetchone()
    normalized_value = normalize_unknown_value(new_value)
    if existing:
        conn.execute(
            """
            UPDATE master_building_field_values
            SET value_text = ?, source_type = ?, source_file = ?, source_date = ?, info_cutoff_date = ?,
                last_verified_at = ?, updated_by = ?, version = version + 1, updated_at = ?
            WHERE id = ?
            """,
            (
                normalized_value,
                source_type,
                source_file,
                effective_source_date,
                info_cutoff_date,
                timestamp,
                actor.user_id,
                timestamp,
                existing["id"],
            ),
        )
    else:
        conn.execute(
            """
            INSERT INTO master_building_field_values(
              id, building_id, field_key, value_text, value_json, source_type, source_file,
              source_date, info_cutoff_date, last_verified_at, updated_by, version, status, created_at, updated_at
            ) VALUES(?, ?, ?, ?, '', ?, ?, ?, ?, ?, ?, 1, 'active', ?, ?)
            """,
            (
                f"field_value_{uuid.uuid4().hex}",
                building_id,
                field_key,
                normalized_value,
                source_type,
                source_file,
                effective_source_date,
                info_cutoff_date,
                timestamp,
                actor.user_id,
                timestamp,
                timestamp,
            ),
        )
    return old_value, normalized_value


def bump_master_version(conn: sqlite3.Connection, building_id: str, actor: Actor) -> None:
    conn.execute(
        """
        UPDATE master_building_info
        SET version = version + 1, updated_by = ?, updated_at = ?, last_verified_at = ?
        WHERE id = ?
        """,
        (actor.user_id, now_iso(), now_iso(), building_id),
    )


def fetch_group_records(conn: sqlite3.Connection, submission_group_id: str) -> List[dict]:
    return conn.execute(
        """
        SELECT *
        FROM staging_update_requests
        WHERE submission_group_id = ?
        ORDER BY created_at ASC, field_name ASC
        """,
        (submission_group_id,),
    ).fetchall()


def review_group_manual_master_building_id(rows: List[dict]) -> str:
    for row in rows:
        source_content = json_loads_safe(row.get("source_content"), {})
        if isinstance(source_content, dict):
            building_id = normalize_unknown_value(source_content.get("manual_master_building_id"))
            if building_id:
                return building_id
    return ""


INSURANCE_DATABASE_COLUMNS = [
    "Building Name",
    "Address",
    "Renters Insurance Required",
    "Minimum Renters Insurance Coverage",
    "Personal Liability Required",
    "Personal Liability per occurrence",
    "Personal Liability aggregate",
    "COI Required",
    "COI Trigger",
    "Interested Party Required",
    "Submission Method",
    "Alternative Program or Penalty",
    "Parse Status",
]


def serialize_review_record(row: dict) -> dict:
    item = dict(row)
    item["evidence_items"] = json_loads_safe(item.get("evidence_json"), [])
    item["review_flags"] = json_loads_safe(item.get("review_flags_json"), [])
    item["manual_review_reason"] = (item.get("manual_review_reason") or "").strip()
    item["approval_stage"] = review_group_approval_stage([item])
    return item


def review_group_approval_stage(rows: List[dict]) -> str:
    if not rows:
        return APPROVAL_STAGE_TO_STAGING
    source_types = {str(row.get("source_type") or "") for row in rows}
    if STAGING_AB_SYNC_SOURCE_TYPE in source_types or MASTER_EXCEL_SYNC_SOURCE_TYPE in source_types:
        return APPROVAL_STAGE_TO_MASTER
    for row in rows:
        stage = normalize_review_approval_stage(row.get("approval_stage"))
        if stage in APPROVAL_STAGES:
            return stage
    return APPROVAL_STAGE_TO_STAGING


def can_actor_decide_review_stage(actor: Actor, stage: str) -> bool:
    stage = normalize_review_approval_stage(stage)
    if stage == APPROVAL_STAGE_TO_MASTER:
        return actor.role == "super_admin"
    if stage == APPROVAL_STAGE_TO_STAGING:
        return actor.role in {"super_admin", "admin"}
    return False


def review_group_target_staging_key(rows: List[dict]) -> str:
    for row in rows:
        value = normalize_unknown_value(row.get("target_staging_key"))
        if value:
            return value
    return ""


def review_group_value_map(rows: List[dict], building: Optional[dict] = None) -> Dict[str, Optional[str]]:
    values: Dict[str, Optional[str]] = {}
    for row in rows:
        field_name = row.get("field_name") or ""
        if not field_name or field_name.startswith("__"):
            continue
        normalized = normalize_unknown_value(row.get("new_value"))
        if normalized is not None:
            values[field_name] = normalized
    if building:
        for field_key in CORE_MASTER_FIELD_KEYS:
            values.setdefault(field_key, normalize_unknown_value(building.get(field_key)))
        for field_key, value in (building.get("extensions") or {}).items():
            values.setdefault(field_key, normalize_unknown_value(value))
    return values


def insurance_parse_status_from_records(rows: List[dict]) -> str:
    reasons = [row.get("manual_review_reason") for row in rows if row.get("manual_review_reason")]
    if reasons:
        return "Human review required"
    has_insurance_values = any(
        normalize_unknown_value(row.get("new_value"))
        for row in rows
        if (row.get("field_name") or "").startswith("insurance_")
    )
    return "Partially resolved" if has_insurance_values else "Not mentioned"


def insurance_manual_reasons_from_records(rows: List[dict]) -> List[str]:
    reasons: List[str] = []
    seen: set[str] = set()
    for row in rows:
        reason = (row.get("manual_review_reason") or "").strip().strip("；")
        if not reason or reason in seen:
            continue
        reasons.append(reason)
        seen.add(reason)
    return reasons


def insurance_original_quotes_from_records(rows: List[dict]) -> List[str]:
    quotes: List[str] = []
    seen: set[str] = set()
    for row in rows:
        for evidence in row.get("evidence_items") or []:
            quote = normalize_unknown_value((evidence or {}).get("quote"))
            if not quote or quote in seen:
                continue
            quotes.append(quote)
            seen.add(quote)
    return quotes[:6]


def build_insurance_database_view(values: Dict[str, Optional[str]], rows: List[dict]) -> dict:
    row = {
        "Building Name": values.get("building_name") or "—",
        "Address": values.get("address") or "—",
        "Renters Insurance Required": insurance_status_label(values.get("insurance_renters_required")),
        "Minimum Renters Insurance Coverage": values.get("insurance_renters_minimum_coverage") or "—",
        "Personal Liability Required": insurance_status_label(values.get("insurance_personal_liability_required")),
        "Personal Liability per occurrence": values.get("insurance_personal_liability_per_occurrence") or "—",
        "Personal Liability aggregate": values.get("insurance_personal_liability_aggregate") or "—",
        "COI Required": insurance_status_label(values.get("insurance_coi_required")),
        "COI Trigger": values.get("insurance_coi_trigger") or "—",
        "Interested Party Required": insurance_status_label(values.get("insurance_interested_party_required")),
        "Submission Method": values.get("insurance_submission_method") or "—",
        "Alternative Program or Penalty": values.get("insurance_alternative_program_or_penalty") or "—",
        "Parse Status": insurance_parse_status_from_records(rows),
    }
    return {
        "columns": INSURANCE_DATABASE_COLUMNS,
        "rows": [row],
        "manual_review_reasons": list(
            dict.fromkeys(
                [
                    *insurance_manual_reasons_from_records(rows),
                    *manual_review_reasons_for_group(values),
                ]
            )
        ),
        "original_quotes": insurance_original_quotes_from_records(rows),
    }


def build_insurance_mapping_view(values: Dict[str, Optional[str]], rows: List[dict]) -> dict:
    mapping_rows: List[dict] = []

    def push(original: str, category: str, page_option: str, suggestion: str, manual: bool) -> None:
        mapping_rows.append(
            {
                "Original Building Requirement": original,
                "System Category": category,
                "Matching Insurance Page Option": page_option,
                "Recommended Action": suggestion,
                "Human Review Required": "Yes" if manual else "No",
            }
        )

    quotes = insurance_original_quotes_from_records(rows)
    quote_lookup = "\n".join(quotes)
    renters_status = normalize_insurance_status(values.get("insurance_renters_required"))
    liability_status = normalize_insurance_status(values.get("insurance_personal_liability_required"))
    coi_status = normalize_insurance_status(values.get("insurance_coi_required"))
    party_status = normalize_insurance_status(values.get("insurance_interested_party_required"))
    if renters_status:
        suggestion = "Confirm whether coverage is mandatory before the customer buys it. If it is optional, do not present it as required."
        if values.get("insurance_renters_minimum_coverage"):
            suggestion = f"Confirm that renters coverage is at least {values['insurance_renters_minimum_coverage']}."
        push(
            quote_lookup or "Renters Insurance requirement",
            "Renters Insurance",
            "Renters Insurance",
            suggestion,
            renters_status == "manual_review",
        )
    if values.get("insurance_personal_property_minimum") or normalize_insurance_status(values.get("insurance_personal_property_required")):
        push(
            quote_lookup or "Personal Property / belongings coverage",
            "Personal Property",
            "Personal Property",
            "Map this only when the source explicitly requires personal-property or belongings coverage; otherwise leave it unset.",
            normalize_insurance_status(values.get("insurance_personal_property_required")) == "manual_review",
        )
    if liability_status or values.get("insurance_personal_liability_per_occurrence") or values.get("insurance_personal_liability_aggregate"):
        manual = bool(values.get("insurance_personal_liability_aggregate")) or liability_status == "manual_review"
        suggestion = "Map the liability requirement to Personal Liability."
        if values.get("insurance_personal_liability_per_occurrence"):
            suggestion = f"Confirm Personal Liability of at least {values['insurance_personal_liability_per_occurrence']} per occurrence."
        if values.get("insurance_personal_liability_aggregate"):
            suggestion += f" The source also specifies an aggregate of {values['insurance_personal_liability_aggregate']}; request human review if the page has no aggregate field."
        push(
            quote_lookup or "Personal Liability / liability coverage",
            "Personal Liability",
            "Personal Liability",
            suggestion,
            manual,
        )
    if coi_status:
        push(
            quote_lookup or "COI",
            "COI",
            "Not a Lemonade renters-insurance option",
            "A moving or delivery company normally supplies the COI. Do not tell the customer to select it on a renters-insurance page.",
            coi_status == "manual_review",
        )
    if party_status or normalize_insurance_status(values.get("insurance_additional_insured_required")) or normalize_insurance_status(values.get("insurance_certificate_holder_required")):
        manual = party_status == "manual_review"
        suggestion = "Use Interested Party or Additional Interest only when the source names that type. If it merely says to add the property, request human review."
        page_option = "Interested Party / Additional Interest"
        if normalize_insurance_status(values.get("insurance_additional_insured_required")) == "yes":
            page_option = "Additional Insured"
        elif normalize_insurance_status(values.get("insurance_certificate_holder_required")) == "yes":
            page_option = "Certificate Holder"
        push(
            quote_lookup or "Interested Party / Additional Insured / Certificate Holder",
            "Policy-related party",
            page_option,
            suggestion,
            manual,
        )
    if values.get("insurance_submission_method"):
        push(
            quote_lookup or "Proof / verification / upload",
            "Submission process",
            "Download proof of insurance or the policy after purchase",
            f"Follow the source submission method: {values['insurance_submission_method']}. This is a verification channel, not an insurance purchase option.",
            False,
        )
    return {
        "columns": ["Original Building Requirement", "System Category", "Matching Insurance Page Option", "Recommended Action", "Human Review Required"],
        "rows": mapping_rows,
    }


def build_insurance_staff_explanation(values: Dict[str, Optional[str]], rows: List[dict]) -> str:
    lines: List[str] = []
    renters_status = normalize_insurance_status(values.get("insurance_renters_required"))
    liability_status = normalize_insurance_status(values.get("insurance_personal_liability_required"))
    coi_status = normalize_insurance_status(values.get("insurance_coi_required"))
    reasons = insurance_manual_reasons_from_records(rows)

    if renters_status == "yes":
        coverage = values.get("insurance_renters_minimum_coverage")
        if coverage:
            lines.append(f"This building explicitly requires renters insurance. The minimum stated renters coverage is {coverage}.")
        else:
            lines.append("This building explicitly requires renters insurance, but the minimum renters coverage is not stated clearly.")
    elif renters_status == "optional":
        lines.append("Renters insurance is optional for this building. Do not tell the customer that it is mandatory.")
    elif renters_status == "no":
        lines.append("The source does not require the customer to purchase renters insurance.")
    elif renters_status == "manual_review":
        lines.append("The source mentions insurance, but a person must confirm whether it means renters insurance.")
    else:
        lines.append("The source does not clearly state whether renters insurance is required.")

    if liability_status == "yes":
        details = []
        if values.get("insurance_personal_liability_per_occurrence"):
            details.append(f"per occurrence {values['insurance_personal_liability_per_occurrence']}")
        if values.get("insurance_personal_liability_aggregate"):
            details.append(f"aggregate {values['insurance_personal_liability_aggregate']}")
        if details:
            lines.append(f"The customer should also verify Personal Liability. The current requirement is {', '.join(details)}.")
        else:
            lines.append("The source mentions Personal Liability, but the limits still need confirmation.")

    if coi_status == "yes":
        trigger = values.get("insurance_coi_trigger") or "The source does not fully explain the trigger."
        lines.append(f"The source also mentions a COI. Trigger: {trigger} This is not something the customer buys directly on a renters-insurance page.")

    if values.get("insurance_submission_method"):
        lines.append(f"The stated insurance or verification submission method is: {values['insurance_submission_method']}.")
    if values.get("insurance_alternative_program_or_penalty"):
        lines.append(f"If the customer does not submit on time, the source mentions this alternative program or penalty: {values['insurance_alternative_program_or_penalty']}.")
    if reasons:
        lines.append("Items requiring human review: " + "; ".join(reasons))
    else:
        lines.append("The information that can directly guide the customer is reasonably clear in this source.")
    return "\n".join(lines)


def insurance_values_from_snapshot(snapshot: dict) -> Dict[str, Optional[str]]:
    extensions = snapshot.get("extensions") or {}
    renters_required = normalize_unknown_value(extensions.get("insurance_renters_required"))
    if not renters_required:
        coarse_state = requirement_state(snapshot.get("insurance_required"))
        renters_required = {
            "required": "yes",
            "not_required": "no",
            "optional": "optional",
        }.get(coarse_state)
    renters_minimum = normalize_unknown_value(extensions.get("insurance_renters_minimum_coverage")) or normalize_unknown_value(snapshot.get("insurance_coverage_amount"))
    return {
        "insurance_renters_required": renters_required,
        "insurance_renters_minimum_coverage": renters_minimum,
        "insurance_personal_property_required": normalize_unknown_value(extensions.get("insurance_personal_property_required")),
        "insurance_personal_property_minimum": normalize_unknown_value(extensions.get("insurance_personal_property_minimum")),
        "insurance_personal_liability_required": normalize_unknown_value(extensions.get("insurance_personal_liability_required")),
        "insurance_personal_liability_per_occurrence": normalize_unknown_value(extensions.get("insurance_personal_liability_per_occurrence")),
        "insurance_personal_liability_aggregate": normalize_unknown_value(extensions.get("insurance_personal_liability_aggregate")),
        "insurance_coi_required": normalize_unknown_value(extensions.get("insurance_coi_required")),
        "insurance_coi_trigger": normalize_unknown_value(extensions.get("insurance_coi_trigger")),
        "insurance_interested_party_required": normalize_unknown_value(extensions.get("insurance_interested_party_required")),
        "insurance_additional_insured_required": normalize_unknown_value(extensions.get("insurance_additional_insured_required")),
        "insurance_certificate_holder_required": normalize_unknown_value(extensions.get("insurance_certificate_holder_required")),
        "insurance_submission_method": normalize_unknown_value(extensions.get("insurance_submission_method")),
        "insurance_recipient": normalize_unknown_value(extensions.get("insurance_recipient")),
        "insurance_alternative_program_or_penalty": normalize_unknown_value(extensions.get("insurance_alternative_program_or_penalty")),
    }


def move_in_values_from_snapshot(snapshot: dict) -> Dict[str, Optional[str]]:
    extensions = snapshot.get("extensions") or {}
    return {
        "key_pickup_notes": normalize_unknown_value(
            snapshot.get("key_pickup_notes") or extensions.get("key_pickup_notes")
        ),
        "service_elevator_booking_notes": normalize_unknown_value(
            snapshot.get("service_elevator_booking_notes")
            or extensions.get("service_elevator_booking_notes")
        ),
        "move_in_notes": normalize_unknown_value(snapshot.get("move_in_notes")),
    }


def network_plan_values_from_snapshot(snapshot: dict) -> Dict[str, Optional[str]]:
    extensions = snapshot.get("extensions") or {}
    return {
        field_key: normalize_unknown_value(extensions.get(field_key))
        for field_key in NETWORK_PLAN_FIELD_MAP
    }


def network_provider_note_values_from_snapshot(snapshot: dict) -> Dict[str, Optional[str]]:
    extensions = snapshot.get("extensions") or {}
    return {
        field_key: normalize_unknown_value(extensions.get(field_key))
        for field_key in NETWORK_PROVIDER_NOTE_FIELD_MAP
    }


def format_network_plan_text(value: Optional[str]) -> Optional[str]:
    text = normalize_unknown_value(value)
    if not text:
        return None
    parts = [part.strip() for part in re.split(r"[\n,]+", text) if part.strip()]
    return "；".join(dict.fromkeys(parts)) if parts else None


def payload_field_names(payload: BaseModel) -> set[str]:
    return set(getattr(payload, "model_fields_set", getattr(payload, "__fields_set__", set())))


def parse_iso_date(value: object) -> Optional[datetime.date]:
    text = (normalize_unknown_value(value) or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def days_in_month(year: int, month: int) -> int:
    if month == 2:
        return 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28
    if month in {4, 6, 9, 11}:
        return 30
    return 31


def subtract_months_from_date(value: str, months: int) -> str:
    parsed = parse_iso_date(value)
    if not parsed:
        return ""
    month_index = parsed.year * 12 + parsed.month - 1 - months
    year = month_index // 12
    month = month_index % 12 + 1
    day = min(parsed.day, days_in_month(year, month))
    return datetime(year, month, day).date().isoformat()


def lease_days_from_today(value: str) -> Optional[int]:
    parsed = parse_iso_date(value)
    if not parsed:
        return None
    return (parsed - datetime.now().date()).days


def ensure_crm_actor(user: dict) -> Actor:
    actor = get_actor(user)
    if actor.role not in CRM_ALLOWED_ROLES:
        raise HTTPException(status_code=403, detail="Your account does not have CRM permission.")
    return actor


def ensure_crm_default_templates(conn: sqlite3.Connection) -> None:
    timestamp = now_iso()
    for template in DEFAULT_CRM_SERVICE_TEMPLATES:
        existing = conn.execute(
            "SELECT * FROM crm_service_templates WHERE service_key = ?",
            (template["service_key"],),
        ).fetchone()
        if existing:
            template_id = existing["id"]
            current_config = json_loads_safe(existing.get("config_json"), {})
            default_config = template.get("config") or {}
            merged_config = {**default_config, **current_config}
            if not isinstance(current_config.get("flow_profile"), dict):
                merged_config["flow_profile"] = default_config.get("flow_profile") or DEFAULT_CRM_FLOW_PROFILES.get(
                    template["service_key"],
                    BASE_CRM_FLOW_PROFILE,
                )
            else:
                current_profile = current_config.get("flow_profile") or {}
                default_profile = default_config.get("flow_profile") or DEFAULT_CRM_FLOW_PROFILES.get(
                    template["service_key"],
                    BASE_CRM_FLOW_PROFILE,
                )
                merged_profile = {
                    **default_profile,
                    **current_profile,
                    "staff_labels": {
                        **(default_profile.get("staff_labels") or {}),
                        **(current_profile.get("staff_labels") if isinstance(current_profile.get("staff_labels"), dict) else {}),
                    },
                    "customer_labels": {
                        **(default_profile.get("customer_labels") or {}),
                        **(current_profile.get("customer_labels") if isinstance(current_profile.get("customer_labels"), dict) else {}),
                    },
                    "staff_to_customer_map": {
                        **CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
                        **(
                            current_profile.get("staff_to_customer_map")
                            if isinstance(current_profile.get("staff_to_customer_map"), dict)
                            else {}
                        ),
                    },
                }
                if template["service_key"] == "renters_insurance" and set(current_profile.get("skip_stages") or []) == {
                    "info_collected"
                }:
                    merged_profile["skip_stages"] = ["staff:service_confirmed", "staff:info_collected"]
                merged_config["flow_profile"] = merged_profile
            if "service_scope" not in merged_config and default_config.get("service_scope"):
                merged_config["service_scope"] = default_config.get("service_scope")
            if merged_config != current_config:
                conn.execute(
                    """
                    UPDATE crm_service_templates
                    SET config_json = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (json_dumps(merged_config), timestamp, template_id),
                )
        else:
            template_id = f"crm_tpl_{uuid.uuid4().hex}"
            conn.execute(
                """
                INSERT INTO crm_service_templates(
                  id, service_key, name, description, category, active, display_order,
                  config_json, created_by, created_at, updated_at
                ) VALUES(?, ?, ?, ?, ?, 1, ?, ?, 'system', ?, ?)
                """,
                (
                    template_id,
                    template["service_key"],
                    template["name"],
                    template.get("description", ""),
                    template.get("category", "general"),
                    int(template.get("display_order") or 100),
                    json_dumps(template.get("config") or {}),
                    timestamp,
                    timestamp,
                ),
            )
        for step in template.get("steps") or []:
            conn.execute(
                """
                INSERT OR IGNORE INTO crm_service_steps(
                  id, template_id, step_key, title, scope, field_schema_json,
                  display_order, active, created_at, updated_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                """,
                (
                    f"crm_step_{uuid.uuid4().hex}",
                    template_id,
                    step["step_key"],
                    step["title"],
                    step.get("scope") or "group",
                    json_dumps(step.get("field_schema") or []),
                    int(step.get("display_order") or 100),
                    timestamp,
                    timestamp,
                ),
            )


def list_crm_template_rows(conn: sqlite3.Connection, *, include_inactive: bool = False) -> List[dict]:
    ensure_crm_default_templates(conn)
    where_clause = "" if include_inactive else "WHERE active = 1"
    return conn.execute(
        f"""
        SELECT *
        FROM crm_service_templates
        {where_clause}
        ORDER BY display_order ASC, name ASC
        """
    ).fetchall()


def active_crm_templates(conn: sqlite3.Connection) -> List[dict]:
    return list_crm_template_rows(conn, include_inactive=False)


def crm_template_steps(conn: sqlite3.Connection, template_id: str) -> List[dict]:
    rows = conn.execute(
        """
        SELECT *
        FROM crm_service_steps
        WHERE template_id = ? AND active = 1
        ORDER BY display_order ASC, title ASC
        """,
        (template_id,),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "step_key": row["step_key"],
            "title": row["title"],
            "scope": row["scope"],
            "field_schema": json_loads_safe(row.get("field_schema_json"), []),
            "display_order": row["display_order"],
            "active": bool(row["active"]),
        }
        for row in rows
    ]


def serialize_crm_template(conn: sqlite3.Connection, row: dict) -> dict:
    config = json_loads_safe(row.get("config_json"), {})
    return {
        "id": row["id"],
        "service_key": row["service_key"],
        "name": row["name"],
        "description": row.get("description") or "",
        "category": row.get("category") or "general",
        "active": bool(row.get("active")),
        "display_order": row.get("display_order") or 100,
        "service_scope": crm_template_service_scope(row),
        "service_delivery_mode": config.get("service_delivery_mode") or "",
        "flow_profile": crm_flow_snapshot_for_template(row),
        "task_rules": (
            config.get("task_rules")
            if isinstance(config.get("task_rules"), list) and config.get("task_rules_configured")
            else crm_service_task_rules(row["service_key"], "required", config)
        ),
        "config": config,
        "steps": crm_template_steps(conn, row["id"]),
    }


def crm_template_service_scope(template_row: dict) -> str:
    config = json_loads_safe(template_row.get("config_json"), {})
    configured = config.get("service_scope")
    if configured in CRM_SERVICE_SCOPES:
        return configured
    return crm_service_scope(template_row["service_key"])


def crm_normalize_responsibility_status(value: str, fallback: str = "unassigned") -> str:
    normalized = value or fallback
    return normalized if normalized in CRM_RESPONSIBILITY_STATUSES else fallback


def crm_active_customer_rows(conn: sqlite3.Connection, case_id: str) -> List[dict]:
    return conn.execute(
        "SELECT * FROM crm_case_guests WHERE case_id = ? AND status != 'inactive' ORDER BY created_at ASC",
        (case_id,),
    ).fetchall()


def crm_service_requires_responsible(service_scope: str, applicability: str, completion_status: str = "") -> bool:
    if service_scope != "case_level":
        return False
    if applicability == "not_needed":
        return False
    if crm_normalize_completion_status(completion_status or "") in {"waived", "not_applicable"}:
        return False
    return True


def crm_case_service_responsibility_state(
    customer_rows: List[dict],
    *,
    service_scope: str,
    applicability: str,
    completion_status: str = "",
    current_responsible_customer_id: str = "",
    current_responsibility_status: str = "",
) -> dict:
    covered_customer_ids = [row["id"] for row in customer_rows] if service_scope == "case_level" else []
    if not crm_service_requires_responsible(service_scope, applicability, completion_status):
        return {
            "responsible_customer_id": "",
            "covered_customer_ids": covered_customer_ids,
            "responsibility_status": "unassigned",
        }
    valid_customer_ids = set(covered_customer_ids)
    responsible_customer_id = current_responsible_customer_id if current_responsible_customer_id in valid_customer_ids else ""
    if not responsible_customer_id and len(customer_rows) == 1:
        responsible_customer_id = customer_rows[0]["id"]
    responsibility_status = crm_normalize_responsibility_status(current_responsibility_status)
    if responsible_customer_id:
        if responsibility_status in {"unassigned", ""}:
            responsibility_status = "assigned"
    else:
        responsibility_status = "unassigned"
    return {
        "responsible_customer_id": responsible_customer_id,
        "covered_customer_ids": covered_customer_ids,
        "responsibility_status": responsibility_status,
    }


def crm_task_title_for_responsible_customer(title: str, customer_name: str, customer_rows: List[dict]) -> str:
    base_title = (title or "").strip()
    for row in customer_rows:
        full_name = (row.get("full_name") or "").strip()
        if full_name and base_title.endswith(f" - {full_name}"):
            base_title = base_title[: -len(f" - {full_name}")].strip()
            break
    if customer_name and not base_title.endswith(f" - {customer_name}"):
        return f"{base_title} - {customer_name}".strip()
    return base_title


def crm_reassign_service_tasks_to_responsible_customer(
    conn: sqlite3.Connection,
    *,
    case_service_id: str,
    responsible_customer_id: str,
    customer_rows: List[dict],
    timestamp: str,
) -> None:
    customer = next((row for row in customer_rows if row["id"] == responsible_customer_id), None)
    customer_name = (customer.get("full_name") or "").strip() if customer else ""
    conn.execute(
        """
        UPDATE crm_tasks
        SET status = 'completed', completed_at = COALESCE(completed_at, ?), updated_at = ?
        WHERE case_service_id = ?
          AND task_type = 'assign_responsible_customer'
          AND status NOT IN ('completed', 'cancelled', 'done')
        """,
        (timestamp, timestamp, case_service_id),
    )
    task_rows = conn.execute(
        """
        SELECT *
        FROM crm_tasks
        WHERE case_service_id = ?
          AND task_type != 'assign_responsible_customer'
          AND status NOT IN ('completed', 'cancelled', 'done')
        """,
        (case_service_id,),
    ).fetchall()
    for task in task_rows:
        conn.execute(
            """
            UPDATE crm_tasks
            SET target_customer_id = ?, customer_id = ?,
                title = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                responsible_customer_id,
                responsible_customer_id,
                crm_task_title_for_responsible_customer(
                    task.get("title") or "",
                    customer_name,
                    customer_rows,
                ),
                timestamp,
                task["id"],
            ),
        )


def crm_requirement_applicability(value: object) -> str:
    state = requirement_state(value)
    if state == "required":
        return "required"
    if state == "not_required":
        return "not_needed"
    if state == "optional":
        return "optional"
    return "unknown"


def crm_insurance_applicability(snapshot: Optional[dict]) -> str:
    if not snapshot:
        return "unknown"
    crm_insurance = snapshot.get("insurance")
    if isinstance(crm_insurance, dict):
        value = crm_insurance.get("required")
        if value in CRM_APPLICABILITY_VALUES:
            return value
    insurance_values = insurance_values_from_snapshot(snapshot)
    status = normalize_insurance_status(insurance_values.get("insurance_renters_required"))
    if status == "yes":
        return "required"
    if status == "no":
        return "not_needed"
    if status == "optional":
        return "optional"
    return crm_requirement_applicability(snapshot.get("insurance_required"))


def crm_service_applicability(service_key: str, snapshot: Optional[dict]) -> str:
    snapshot = snapshot or {}
    if service_key == "phone_card":
        return "optional"
    if service_key == "renters_insurance":
        return crm_insurance_applicability(snapshot)
    if service_key == "electricity_account":
        crm_electricity = snapshot.get("electricity")
        if isinstance(crm_electricity, dict) and crm_electricity.get("required") in CRM_APPLICABILITY_VALUES:
            return crm_electricity["required"]
        return crm_requirement_applicability(snapshot.get("electricity_required"))
    if service_key == "internet_setup":
        crm_internet = snapshot.get("internet")
        if isinstance(crm_internet, dict) and crm_internet.get("required") in CRM_APPLICABILITY_VALUES:
            return crm_internet["required"]
        return crm_requirement_applicability(snapshot.get("internet_self_setup_required"))
    return "optional"


def crm_default_service_status(applicability: str) -> str:
    return "not_needed" if applicability == "not_needed" else "pending"


def crm_normalize_need_status(value: str, fallback: str = "unknown") -> str:
    legacy = {
        "needed": "required",
        "waived": "not_needed",
    }
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_NEED_STATUSES else fallback


def crm_normalize_submission_status(value: str, fallback: str = "unknown") -> str:
    legacy = {
        "not_started": "not_submitted",
        "info_requested": "not_submitted",
        "info_collected": "not_submitted",
        "submitted_by_customer": "submitted",
        "submitted_by_staff": "submitted",
    }
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_SUBMISSION_STATUSES else fallback


def crm_normalize_completion_status(value: str, fallback: str = "unknown") -> str:
    legacy = {
        "blocked": "failed",
        "cancelled": "failed",
    }
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_COMPLETION_STATUSES else fallback


def crm_normalize_intro_status(value: str, fallback: str = "unknown") -> str:
    legacy = {"not_required": "not_needed"}
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_INTRO_STATUSES else fallback


def crm_normalize_follow_up_status(value: str, fallback: str = "unknown") -> str:
    legacy = {
        "not_started": "required",
        "waiting_customer": "required",
        "waiting_agent": "required",
        "waiting_provider": "required",
        "followed_up": "scheduled",
        "no_response": "overdue",
    }
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_FOLLOW_UP_STATUSES else fallback


def crm_normalize_agent_completion_status(value: str, fallback: str = "unknown") -> str:
    legacy = {
        "not_started": "open",
        "action_required": "open",
        "in_progress": "open",
        "waiting_customer": "pending_customer",
        "waiting_provider": "pending_external",
        "verified": "completed",
        "closed": "completed",
        "blocked": "escalated",
        "not_required": "completed",
    }
    normalized = legacy.get(value or "", value or fallback)
    return normalized if normalized in CRM_AGENT_COMPLETION_STATUSES else fallback


def crm_normalize_staff_flow_status(value: str, fallback: str = "not_introduced") -> str:
    normalized = value or fallback
    return normalized if normalized in CRM_STAFF_FLOW_STATUSES else fallback


def crm_normalize_customer_flow_status(value: str, fallback: str = "waiting_intro") -> str:
    normalized = value or fallback
    return normalized if normalized in CRM_CUSTOMER_FLOW_STATUSES else fallback


def crm_customer_flow_base_status(value: str, flow_profile: Optional[dict] = None, fallback: str = "intent_unknown") -> str:
    return crm_normalize_customer_flow_status(value, fallback)


def crm_normalize_customer_flow_status_for_profile(value: str, fallback: str = "waiting_intro", flow_profile: Optional[dict] = None) -> str:
    return crm_normalize_customer_flow_status(value, fallback)


def crm_normalize_service_status(value: str, fallback: str = "pending") -> str:
    normalized = value or fallback
    return normalized if normalized in CRM_SERVICE_STATUSES else fallback


def crm_flow_stage_is_enabled(skip_stages: object, group_key: str, status_key: str) -> bool:
    stage_set = set(skip_stages if isinstance(skip_stages, list) else [])
    return status_key not in stage_set and f"{group_key}:{status_key}" not in stage_set


def crm_normalize_flow_step(raw_step: dict, index: int, profile: dict) -> dict:
    staff_status = crm_normalize_staff_flow_status(
        str(raw_step.get("staff_flow_status") or raw_step.get("status_key") or raw_step.get("step_key") or ""),
        "not_introduced",
    )
    customer_status = crm_normalize_customer_flow_status_for_profile(
        str(raw_step.get("customer_flow_status") or ""),
        crm_customer_flow_from_staff_flow(staff_status, "waiting_intro", profile),
        profile,
    )
    step_key = str(raw_step.get("step_key") or staff_status or f"step_{index + 1}").strip() or f"step_{index + 1}"
    staff_labels = profile.get("staff_labels") if isinstance(profile.get("staff_labels"), dict) else {}
    customer_labels = profile.get("customer_labels") if isinstance(profile.get("customer_labels"), dict) else {}
    service_status = str(raw_step.get("service_status") or "").strip()
    if service_status and service_status not in CRM_SERVICE_STATUSES:
        service_status = ""
    required_fields = raw_step.get("required_fields")
    if not isinstance(required_fields, list):
        by_stage = profile.get("required_fields_by_stage") if isinstance(profile.get("required_fields_by_stage"), dict) else {}
        required_fields = by_stage.get(staff_status) if isinstance(by_stage.get(staff_status), list) else []
    enabled = raw_step.get("enabled")
    if enabled is None:
        enabled = raw_step.get("active")
    if enabled is None:
        enabled = crm_flow_stage_is_enabled(profile.get("skip_stages") or [], "staff", staff_status)
    return {
        "step_key": step_key,
        "enabled": bool(enabled),
        "staff_flow_status": staff_status,
        "staff_label": str(raw_step.get("staff_label") or raw_step.get("label") or staff_labels.get(staff_status) or staff_status).strip(),
        "customer_flow_status": customer_status,
        "customer_label": str(
            raw_step.get("customer_label") or customer_labels.get(customer_status) or customer_status
        ).strip(),
        "service_status": service_status,
        "required_fields": [str(item).strip() for item in required_fields if str(item).strip()],
        "description": str(raw_step.get("description") or "").strip(),
        "is_completion": bool(raw_step.get("is_completion")),
        "is_risk": bool(raw_step.get("is_risk")),
        "is_terminal": bool(raw_step.get("is_terminal")),
        "task_rules": raw_step.get("task_rules") if isinstance(raw_step.get("task_rules"), list) else [],
        "display_order": int(raw_step.get("display_order") or (index + 1) * 10),
    }


def crm_flow_steps_for_profile(profile: Optional[dict]) -> List[dict]:
    profile = profile if isinstance(profile, dict) else BASE_CRM_FLOW_PROFILE
    raw_steps = profile.get("flow_steps")
    if isinstance(raw_steps, list) and raw_steps:
        return [
            crm_normalize_flow_step(step, index, profile)
            for index, step in enumerate(raw_steps)
            if isinstance(step, dict)
        ]
    steps = []
    staff_labels = profile.get("staff_labels") if isinstance(profile.get("staff_labels"), dict) else {}
    customer_labels = profile.get("customer_labels") if isinstance(profile.get("customer_labels"), dict) else {}
    for index, status_key in enumerate(["not_introduced", "introduced", "following_up", "service_confirmed", "info_collected", "completed", "terminated"]):
        customer_status = crm_customer_flow_from_staff_flow(status_key, "waiting_intro", profile)
        steps.append(
            crm_normalize_flow_step(
                {
                    "step_key": status_key,
                    "enabled": crm_flow_stage_is_enabled(profile.get("skip_stages") or [], "staff", status_key),
                    "staff_flow_status": status_key,
                    "staff_label": staff_labels.get(status_key) or status_key,
                    "customer_flow_status": customer_status,
                    "customer_label": customer_labels.get(customer_status) or customer_status,
                    "required_fields": (profile.get("required_fields_by_stage") or {}).get(status_key, []),
                    "display_order": (index + 1) * 10,
                },
                index,
                profile,
            )
        )
    return steps


def crm_enabled_flow_steps(profile: Optional[dict]) -> List[dict]:
    return [step for step in crm_flow_steps_for_profile(profile) if step.get("enabled")]


def crm_flow_step_by_key(profile: Optional[dict], step_key: str) -> Optional[dict]:
    for step in crm_flow_steps_for_profile(profile):
        if step.get("step_key") == step_key:
            return step
    return None


def crm_default_flow_step_key(profile: Optional[dict], staff_flow_status: str = "not_introduced") -> str:
    enabled_steps = crm_enabled_flow_steps(profile)
    risk_match = ""
    for step in enabled_steps:
        if step.get("staff_flow_status") == staff_flow_status:
            if staff_flow_status == "following_up" and step.get("is_risk"):
                risk_match = step.get("step_key") or ""
                continue
            return step.get("step_key") or ""
    if staff_flow_status in {"service_confirmed", "info_collected", "following_up"}:
        for step in enabled_steps:
            if step.get("staff_flow_status") == "introduced" and not step.get("is_risk"):
                return step.get("step_key") or ""
    if staff_flow_status == "completed":
        for step in enabled_steps:
            if step.get("is_completion") or step.get("staff_flow_status") == "completed":
                return step.get("step_key") or ""
    if staff_flow_status == "terminated":
        for step in enabled_steps:
            if step.get("is_terminal") or step.get("staff_flow_status") == "terminated":
                return step.get("step_key") or ""
    if risk_match:
        return risk_match
    return enabled_steps[0].get("step_key") if enabled_steps else ""


def crm_flow_profile_for_template(template_row: Optional[dict]) -> dict:
    if not template_row:
        return BASE_CRM_FLOW_PROFILE
    config = json_loads_safe(template_row.get("config_json"), {})
    profile = config.get("flow_profile")
    if isinstance(profile, dict):
        merged = {
            **BASE_CRM_FLOW_PROFILE,
            **profile,
            "staff_labels": {
                **BASE_CRM_FLOW_PROFILE["staff_labels"],
                **(profile.get("staff_labels") if isinstance(profile.get("staff_labels"), dict) else {}),
            },
            "customer_labels": {
                **BASE_CRM_FLOW_PROFILE["customer_labels"],
                **(profile.get("customer_labels") if isinstance(profile.get("customer_labels"), dict) else {}),
            },
            "staff_to_customer_map": {
                **CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
                **(profile.get("staff_to_customer_map") if isinstance(profile.get("staff_to_customer_map"), dict) else {}),
            },
            "flow_steps": profile.get("flow_steps") if isinstance(profile.get("flow_steps"), list) else [],
        }
        return merged
    return DEFAULT_CRM_FLOW_PROFILES.get(template_row.get("service_key") or "", BASE_CRM_FLOW_PROFILE)


def crm_flow_snapshot_for_template(template_row: Optional[dict]) -> dict:
    profile = crm_flow_profile_for_template(template_row)
    flow_steps = crm_flow_steps_for_profile(profile)
    return {
        "staff_labels": profile.get("staff_labels") or {},
        "customer_labels": profile.get("customer_labels") or {},
        "skip_stages": profile.get("skip_stages") or [],
        "staff_to_customer_map": {
            **CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
            **(profile.get("staff_to_customer_map") if isinstance(profile.get("staff_to_customer_map"), dict) else {}),
        },
        "required_fields_by_stage": profile.get("required_fields_by_stage") or {},
        "flow_steps": flow_steps,
        "terminal_rules": profile.get("terminal_rules") or BASE_CRM_FLOW_PROFILE["terminal_rules"],
    }


def crm_effective_service_flow_snapshot(stored_snapshot: object, template_row: Optional[dict]) -> dict:
    stored = stored_snapshot if isinstance(stored_snapshot, dict) else json_loads_safe(stored_snapshot, {})
    template_snapshot = crm_flow_snapshot_for_template(template_row)
    if not isinstance(stored, dict) or not stored:
        return template_snapshot
    merged = {
        **template_snapshot,
        **stored,
        "staff_labels": {
            **(template_snapshot.get("staff_labels") or {}),
            **(stored.get("staff_labels") if isinstance(stored.get("staff_labels"), dict) else {}),
        },
        "customer_labels": {
            **(template_snapshot.get("customer_labels") or {}),
            **(stored.get("customer_labels") if isinstance(stored.get("customer_labels"), dict) else {}),
        },
        "staff_to_customer_map": {
            **CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
            **(template_snapshot.get("staff_to_customer_map") if isinstance(template_snapshot.get("staff_to_customer_map"), dict) else {}),
            **(stored.get("staff_to_customer_map") if isinstance(stored.get("staff_to_customer_map"), dict) else {}),
        },
    }
    if not isinstance(stored.get("flow_steps"), list) or not stored.get("flow_steps"):
        merged["flow_steps"] = template_snapshot.get("flow_steps") or []
    return merged


def crm_default_flow_fields(service_key: str, applicability: str) -> dict:
    if applicability == "not_needed":
        return {
            "staff_flow_status": "completed",
            "customer_flow_status": "not_needed",
            "service_status": "not_needed",
        }
    if service_key == "phone_card":
        return {
            "staff_flow_status": "not_introduced",
            "customer_flow_status": "waiting_intro",
            "service_status": "pending",
        }
    return {
        "staff_flow_status": "not_introduced",
        "customer_flow_status": "waiting_intro",
        "service_status": crm_default_service_status(applicability),
    }


def crm_customer_flow_from_staff_flow(
    staff_flow_status: str,
    fallback: str = "waiting_intro",
    flow_profile: Optional[dict] = None,
) -> str:
    profile_mapping = {}
    if isinstance(flow_profile, dict) and isinstance(flow_profile.get("staff_to_customer_map"), dict):
        profile_mapping = flow_profile.get("staff_to_customer_map") or {}
    mapping = {
        **CRM_DEFAULT_STAFF_TO_CUSTOMER_MAP,
        **profile_mapping,
    }
    return crm_normalize_customer_flow_status_for_profile(mapping.get(staff_flow_status) or "", fallback, flow_profile)


def crm_service_status_from_flows(
    staff_flow_status: str,
    customer_flow_status: str,
    applicability: str,
    flow_profile: Optional[dict] = None,
) -> str:
    customer_base_status = crm_customer_flow_base_status(customer_flow_status, flow_profile)
    if applicability == "not_needed" or customer_base_status == "not_needed":
        return "not_needed"
    if staff_flow_status == "terminated" or customer_base_status == "declined":
        return "terminated"
    if staff_flow_status == "completed" or customer_base_status == "completed":
        return "completed"
    if staff_flow_status in {"service_confirmed", "info_collected"} or customer_base_status in {
        "service_confirmed",
        "info_provided",
    }:
        return "in_progress"
    if staff_flow_status == "following_up" or customer_base_status == "intent_unknown":
        return "waiting_customer"
    if staff_flow_status == "introduced":
        return "in_progress"
    return crm_default_service_status(applicability)


def crm_flow_fields_from_step_key(
    flow_profile: Optional[dict],
    active_flow_step_key: str,
    *,
    applicability: str,
    fallback_staff_flow_status: str = "not_introduced",
    fallback_customer_flow_status: str = "waiting_intro",
) -> dict:
    step = crm_flow_step_by_key(flow_profile, active_flow_step_key)
    if not step or not step.get("enabled"):
        active_flow_step_key = crm_default_flow_step_key(flow_profile, fallback_staff_flow_status)
        step = crm_flow_step_by_key(flow_profile, active_flow_step_key)
    if not step:
        staff_flow_status = crm_normalize_staff_flow_status(fallback_staff_flow_status)
        customer_flow_status = crm_normalize_customer_flow_status_for_profile(
            fallback_customer_flow_status,
            "waiting_intro",
            flow_profile,
        )
        service_status = crm_service_status_from_flows(staff_flow_status, customer_flow_status, applicability, flow_profile)
        return {
            "active_flow_step_key": "",
            "staff_flow_status": staff_flow_status,
            "customer_flow_status": customer_flow_status,
            "service_status": service_status,
        }
    staff_flow_status = crm_normalize_staff_flow_status(step.get("staff_flow_status") or "", fallback_staff_flow_status)
    customer_flow_status = crm_normalize_customer_flow_status_for_profile(
        step.get("customer_flow_status") or "",
        fallback_customer_flow_status,
        flow_profile,
    )
    step_service_status = ""
    if step.get("is_completion"):
        step_service_status = "completed"
    elif step.get("is_risk"):
        step_service_status = "at_risk"
    elif step.get("is_terminal"):
        step_service_status = "terminated"
    else:
        step_service_status = step.get("service_status") or ""
    service_status = crm_normalize_service_status(
        step_service_status,
        crm_service_status_from_flows(staff_flow_status, customer_flow_status, applicability, flow_profile),
    )
    return {
        "active_flow_step_key": step.get("step_key") or "",
        "staff_flow_status": staff_flow_status,
        "customer_flow_status": customer_flow_status,
        "service_status": service_status,
    }


def crm_flows_from_legacy(service_row: dict, applicability: str) -> dict:
    if applicability == "not_needed" or service_row.get("status") == "not_needed":
        return crm_default_flow_fields(service_row.get("service_key") or "", "not_needed")
    completion_status = crm_normalize_completion_status(service_row.get("completion_status") or "")
    submission_status = crm_normalize_submission_status(service_row.get("submission_status") or "")
    intro_status = crm_normalize_intro_status(service_row.get("intro_status") or "")
    follow_up_status = crm_normalize_follow_up_status(service_row.get("follow_up_status") or "")
    agent_completion_status = crm_normalize_agent_completion_status(service_row.get("agent_completion_status") or "")
    status = crm_normalize_service_status(service_row.get("status") or "")

    if status == "terminated" or completion_status == "failed":
        return {
            "staff_flow_status": "terminated",
            "customer_flow_status": "declined",
            "service_status": "terminated",
        }
    if completion_status in {"waived", "not_applicable"}:
        return crm_default_flow_fields(service_row.get("service_key") or "", "not_needed")
    if status == "completed" or completion_status == "completed" or agent_completion_status == "completed":
        return {
            "staff_flow_status": "completed",
            "customer_flow_status": "completed",
            "service_status": "completed",
        }
    if submission_status in {"submitted", "approved"}:
        return {
            "staff_flow_status": "info_collected",
            "customer_flow_status": "info_provided",
            "service_status": "in_progress",
        }
    if agent_completion_status == "pending_external":
        return {
            "staff_flow_status": "info_collected",
            "customer_flow_status": "info_provided",
            "service_status": "in_progress",
        }
    if agent_completion_status == "pending_customer":
        return {
            "staff_flow_status": "service_confirmed",
            "customer_flow_status": "service_confirmed",
            "service_status": "in_progress",
        }
    if follow_up_status in {"required", "overdue", "scheduled"}:
        if intro_status in {"introduced_to_group", "introduced_to_customer"}:
            return {
                "staff_flow_status": "following_up",
                "customer_flow_status": "intent_unknown",
                "service_status": "waiting_customer",
            }
    if intro_status in {"introduced_to_group", "introduced_to_customer"}:
        return {
            "staff_flow_status": "introduced",
            "customer_flow_status": "intent_unknown",
            "service_status": "in_progress",
        }
    return crm_default_flow_fields(service_row.get("service_key") or "", applicability)


def crm_legacy_fields_from_flows(
    *,
    service_key: str,
    applicability: str,
    staff_flow_status: str,
    customer_flow_status: str,
    current: Optional[dict] = None,
    flow_profile: Optional[dict] = None,
) -> dict:
    current = current or {}
    customer_base_status = crm_customer_flow_base_status(customer_flow_status, flow_profile)
    default_fields = crm_initial_status_fields(service_key, applicability)
    result = dict(default_fields)
    if customer_base_status == "not_needed":
        result.update(
            {
                "need_status": "not_needed",
                "submission_status": "not_applicable",
                "completion_status": "waived",
                "intro_status": "not_needed",
                "follow_up_status": "not_required",
                "agent_completion_status": "completed",
            }
        )
        return result
    if customer_base_status == "declined" or staff_flow_status == "terminated":
        result.update(
            {
                "need_status": current.get("need_status") if current.get("need_status") in {"required", "optional"} else default_fields["need_status"],
                "submission_status": "not_submitted",
                "completion_status": "failed",
                "intro_status": "introduced_to_group"
                if staff_flow_status != "not_introduced"
                else "not_introduced",
                "follow_up_status": "not_required",
                "agent_completion_status": "completed",
            }
        )
        return result
    if staff_flow_status in {"introduced", "following_up", "service_confirmed", "info_collected", "completed"}:
        result["intro_status"] = "introduced_to_group"
    if staff_flow_status in {"following_up", "service_confirmed"}:
        result["follow_up_status"] = "required"
        result["agent_completion_status"] = "pending_customer" if staff_flow_status == "service_confirmed" else "open"
    if staff_flow_status == "info_collected":
        result["follow_up_status"] = "scheduled"
        result["agent_completion_status"] = "pending_external"
    if staff_flow_status == "completed":
        result["follow_up_status"] = "not_required"
        result["agent_completion_status"] = "completed"

    if customer_base_status in {"service_confirmed", "info_provided"}:
        result["completion_status"] = "in_progress"
    if customer_base_status == "info_provided":
        result["submission_status"] = "submitted"
    if customer_base_status == "completed":
        result["submission_status"] = "approved" if result["submission_status"] != "not_applicable" else "not_applicable"
        result["completion_status"] = "completed"
        result["follow_up_status"] = "not_required"
        result["agent_completion_status"] = "completed"
    return result


def crm_service_scope(service_key: str) -> str:
    return "customer_level" if service_key == "phone_card" else "case_level"


def crm_initial_status_fields(service_key: str, applicability: str) -> dict:
    if service_key == "phone_card":
        return {
            "need_status": "optional",
            "submission_status": "not_submitted",
            "completion_status": "not_started",
            "intro_status": "not_introduced",
            "follow_up_status": "required",
            "agent_completion_status": "open",
        }
    if applicability == "not_needed":
        return {
            "need_status": "not_needed",
            "submission_status": "not_applicable",
            "completion_status": "waived",
            "intro_status": "not_needed",
            "follow_up_status": "not_required",
            "agent_completion_status": "completed",
        }
    if applicability == "optional":
        need_status = "optional"
    elif applicability == "required":
        need_status = "required"
    else:
        need_status = "unknown"
    return {
        "need_status": need_status,
        "submission_status": "not_submitted",
        "completion_status": "not_started",
        "intro_status": "not_introduced",
        "follow_up_status": "required" if applicability == "required" else "unknown",
        "agent_completion_status": "open",
    }


def crm_completion_from_service_status(status: str, fallback: str) -> str:
    if status == "completed":
        return "completed"
    if status == "not_needed":
        return "waived"
    if status == "terminated":
        return "failed"
    if status in {"in_progress", "waiting_customer", "at_risk"}:
        return "in_progress"
    return fallback or "not_started"


def crm_task_due_at(base_date: Optional[datetime.date], *, days_before: int = 0, due_hour: int = 9, due_minute: int = 0) -> str:
    due_date = (base_date or datetime.now().date()) - timedelta(days=days_before)
    try:
        due_hour = max(0, min(23, int(due_hour)))
    except (TypeError, ValueError):
        due_hour = 9
    try:
        due_minute = max(0, min(59, int(due_minute)))
    except (TypeError, ValueError):
        due_minute = 0
    return datetime.combine(due_date, datetime.min.time()).replace(hour=due_hour, minute=due_minute).isoformat()


def crm_task_due_at_for_rule(base_date: Optional[datetime.date], rule: dict) -> str:
    if rule.get("timing") == "immediate":
        return datetime.combine(datetime.now().date(), datetime.min.time()).replace(
            hour=int(rule.get("due_hour", 9) or 9),
            minute=int(rule.get("due_minute", 0) or 0),
        ).isoformat()
    return crm_task_due_at(
        base_date,
        days_before=int(rule.get("days") or 0),
        due_hour=int(rule.get("due_hour", 9) if rule.get("due_hour") is not None else 9),
        due_minute=int(rule.get("due_minute", 0) if rule.get("due_minute") is not None else 0),
    )


def crm_task_initial_status(due_at: str, status: str = "open") -> str:
    if status in CRM_TASK_STATUSES and status != "open":
        return status
    due_date = parse_iso_date((due_at or "")[:10])
    if due_date and due_date > datetime.now().date():
        return "scheduled"
    return "open"


def crm_normalize_task_status(status: str) -> str:
    legacy = {
        "todo": "open",
        "done": "completed",
    }
    normalized = legacy.get(status or "", status or "open")
    return normalized if normalized in CRM_TASK_STATUSES else "open"


def crm_task_is_overdue(row: dict) -> bool:
    status = crm_normalize_task_status(row.get("status") or "")
    if status in {"completed", "cancelled"}:
        return False
    due_date = parse_iso_date((row.get("due_at") or "")[:10])
    return bool(due_date and due_date < datetime.now().date())


def normalize_crm_task_rows(conn: sqlite3.Connection) -> None:
    timestamp = now_iso()
    conn.execute(
        """
        UPDATE crm_tasks
        SET status = 'open', updated_at = ?
        WHERE status = 'todo'
        """,
        (timestamp,),
    )
    conn.execute(
        """
        UPDATE crm_tasks
        SET status = 'completed', updated_at = ?, completed_at = COALESCE(NULLIF(completed_at, ''), ?)
        WHERE status = 'done'
        """,
        (timestamp, timestamp),
    )
    conn.execute(
        """
        UPDATE crm_tasks
        SET assigned_to = assigned_user_id, updated_at = ?
        WHERE COALESCE(assigned_to, '') = '' AND COALESCE(assigned_user_id, '') != ''
        """,
        (timestamp,),
    )
    conn.execute(
        """
        UPDATE crm_tasks
        SET target_customer_id = customer_id, updated_at = ?
        WHERE target_customer_id IS NULL AND customer_id IS NOT NULL
        """,
        (timestamp,),
    )
    conn.execute(
        """
        UPDATE crm_tasks
        SET not_before_at = due_at, updated_at = ?
        WHERE COALESCE(not_before_at, '') = '' AND COALESCE(due_at, '') != ''
        """,
        (timestamp,),
    )
    conn.execute(
        f"""
        UPDATE crm_tasks
        SET status = 'cancelled', updated_at = ?
        WHERE status NOT IN ('completed', 'cancelled', 'done')
          AND created_from_rule IN ({",".join("?" for _ in CRM_LOW_VALUE_SYSTEM_TASK_RULES)})
        """,
        (timestamp, *sorted(CRM_LOW_VALUE_SYSTEM_TASK_RULES)),
    )


def cancel_low_value_system_tasks(
    conn: sqlite3.Connection,
    *,
    case_id: str = "",
    service_id: str = "",
    timestamp: Optional[str] = None,
) -> None:
    where_parts = [
        "status NOT IN ('completed', 'cancelled', 'done')",
        f"created_from_rule IN ({','.join('?' for _ in CRM_LOW_VALUE_SYSTEM_TASK_RULES)})",
    ]
    params: List[Any] = [*(sorted(CRM_LOW_VALUE_SYSTEM_TASK_RULES))]
    if case_id:
        where_parts.append("case_id = ?")
        params.append(case_id)
    if service_id:
        where_parts.append("case_service_id = ?")
        params.append(service_id)
    conn.execute(
        f"""
        UPDATE crm_tasks
        SET status = 'cancelled', updated_at = ?
        WHERE {' AND '.join(where_parts)}
        """,
        (timestamp or now_iso(), *params),
    )


def crm_insert_task_if_absent(
    conn: sqlite3.Connection,
    *,
    case_id: str,
    case_service_id: str = "",
    customer_id: str = "",
    title: str,
    description: str = "",
    task_type: str = "follow_up",
    due_at: str,
    not_before_at: str = "",
    priority: str = "normal",
    assigned_user_id: str = "",
    source: str = "system",
    created_from_rule: str = "",
    dedupe_key: str,
) -> None:
    timestamp = now_iso()
    task_status = crm_task_initial_status(due_at)
    conn.execute(
        """
        INSERT OR IGNORE INTO crm_tasks(
          id, case_id, case_service_id, customer_id, title, description, task_type,
          due_at, not_before_at, priority, status, assigned_user_id, assigned_to,
          target_customer_id, source, created_from_rule, dedupe_key, created_at, updated_at
        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            f"crm_task_{uuid.uuid4().hex}",
            case_id,
            case_service_id or None,
            customer_id or None,
            title,
            description,
            task_type,
            due_at,
            not_before_at or due_at,
            priority,
            task_status,
            assigned_user_id,
            assigned_user_id,
            customer_id or None,
            source,
            created_from_rule,
            dedupe_key,
            timestamp,
            timestamp,
        ),
    )
    conn.execute(
        """
        UPDATE crm_tasks
        SET title = ?, description = ?, task_type = ?, due_at = ?, not_before_at = ?,
            priority = ?, assigned_user_id = COALESCE(NULLIF(assigned_user_id, ''), ?),
            assigned_to = COALESCE(NULLIF(assigned_to, ''), ?),
            target_customer_id = ?, customer_id = ?, created_from_rule = ?, updated_at = ?
        WHERE dedupe_key = ?
          AND status NOT IN ('completed', 'cancelled', 'done')
        """,
        (
            title,
            description,
            task_type,
            due_at,
            not_before_at or due_at,
            priority,
            assigned_user_id,
            assigned_user_id,
            customer_id or None,
            customer_id or None,
            created_from_rule,
            timestamp,
            dedupe_key,
        ),
    )


def crm_sim_card_sent_record_due_at(sent_date: str) -> str:
    parsed = parse_iso_date((sent_date or "")[:10])
    if not parsed:
        return ""
    return datetime.combine(parsed, datetime.min.time()).replace(hour=9).isoformat()


def sync_phone_card_sent_record_task(
    conn: sqlite3.Connection,
    *,
    case_row: dict,
    service_row: dict,
    guest_row: dict,
    progress_value: dict,
    actor: Actor,
    timestamp: Optional[str] = None,
) -> None:
    if service_row.get("service_key") != "phone_card":
        return
    sent_date = str(progress_value.get("sent_date") or progress_value.get("sim_sent_date") or "").strip()
    dedupe_key = f"{case_row['id']}:{service_row['id']}:{guest_row['id']}:sim_card_sent_record"
    timestamp = timestamp or now_iso()
    if not sent_date:
        conn.execute(
            """
            UPDATE crm_tasks
            SET status = 'cancelled', updated_at = ?
            WHERE dedupe_key = ?
              AND task_type = 'sim_card_sent_record'
              AND status != 'cancelled'
            """,
            (timestamp, dedupe_key),
        )
        return
    due_at = crm_sim_card_sent_record_due_at(sent_date)
    if not due_at:
        return
    customer_name = (guest_row.get("full_name") or guest_row.get("wechat") or "Customer").strip()
    sim_type = str(progress_value.get("sim_type") or "unknown").strip()
    delivery_method = str(progress_value.get("delivery_method") or "").strip()
    pickup_location = str(progress_value.get("pickup_location") or "").strip()
    tracking_number = str(progress_value.get("tracking_number") or "").strip()
    mailing_address = str(progress_value.get("mailing_address") or "").strip()
    notes = [
        "SIM-card shipment record for task-center history; it does not count as an open task.",
        f"Card type: {sim_type or 'To be confirmed'}",
        f"Delivery method: {delivery_method or 'To be confirmed'}",
        f"Pickup location: {pickup_location}" if pickup_location else "",
        f"Mailing address: {mailing_address}" if mailing_address else "",
        f"Tracking：{tracking_number}" if tracking_number else "",
    ]
    title = f"SIM card sent — {customer_name}"
    existing = conn.execute("SELECT id FROM crm_tasks WHERE dedupe_key = ?", (dedupe_key,)).fetchone()
    if existing:
        conn.execute(
            """
            UPDATE crm_tasks
            SET title = ?, description = ?, due_at = ?, not_before_at = ?, priority = 'normal',
                status = 'completed', completed_at = ?, assigned_user_id = ?, assigned_to = ?,
                customer_id = ?, target_customer_id = ?, source = 'system',
                created_from_rule = 'phone_card:sim_card_sent_record', updated_at = ?
            WHERE dedupe_key = ?
            """,
            (
                title,
                "\n".join(item for item in notes if item),
                due_at,
                due_at,
                due_at,
                case_row.get("owner_user_id") or actor.user_id,
                case_row.get("owner_user_id") or actor.user_id,
                guest_row["id"],
                guest_row["id"],
                timestamp,
                dedupe_key,
            ),
        )
    else:
        conn.execute(
            """
            INSERT INTO crm_tasks(
              id, case_id, case_service_id, customer_id, title, description, task_type,
              due_at, not_before_at, priority, status, assigned_user_id, assigned_to,
              target_customer_id, source, created_from_rule, dedupe_key, created_at, updated_at, completed_at
            ) VALUES(?, ?, ?, ?, ?, ?, 'sim_card_sent_record', ?, ?, 'normal', 'completed', ?, ?, ?, 'system',
                     'phone_card:sim_card_sent_record', ?, ?, ?, ?)
            """,
            (
                f"crm_task_{uuid.uuid4().hex}",
                case_row["id"],
                service_row["id"],
                guest_row["id"],
                title,
                "\n".join(item for item in notes if item),
                due_at,
                due_at,
                case_row.get("owner_user_id") or actor.user_id,
                case_row.get("owner_user_id") or actor.user_id,
                guest_row["id"],
                dedupe_key,
                timestamp,
                timestamp,
                due_at,
            ),
        )


def crm_normalize_template_task_rules(raw_rules: object) -> List[dict]:
    if not isinstance(raw_rules, list):
        return []
    normalized_rules = []
    for index, raw_rule in enumerate(raw_rules):
        if not isinstance(raw_rule, dict):
            continue
        title = str(raw_rule.get("title") or "").strip()
        if not title:
            continue
        task_type = str(raw_rule.get("task_type") or "follow_up").strip() or "follow_up"
        priority = str(raw_rule.get("priority") or "normal").strip() or "normal"
        if priority not in CRM_TASK_PRIORITIES:
            priority = "normal"
        timing = str(raw_rule.get("timing") or "").strip()
        days_value = raw_rule.get("days", raw_rule.get("days_before", 0))
        try:
            days = int(days_value or 0)
        except (TypeError, ValueError):
            days = 0
        try:
            due_hour = int(raw_rule.get("due_hour", 9) if raw_rule.get("due_hour") is not None else 9)
        except (TypeError, ValueError):
            due_hour = 9
        try:
            due_minute = int(raw_rule.get("due_minute", 0) if raw_rule.get("due_minute") is not None else 0)
        except (TypeError, ValueError):
            due_minute = 0
        normalized_rule = {
            "key": str(raw_rule.get("key") or f"rule_{index + 1}").strip() or f"rule_{index + 1}",
            "title": title,
            "task_type": task_type,
            "priority": priority,
            "description": str(raw_rule.get("description") or "").strip(),
            "flow_step_key": str(raw_rule.get("flow_step_key") or "").strip(),
            "due_hour": max(0, min(23, due_hour)),
            "due_minute": max(0, min(59, due_minute)),
        }
        if timing == "immediate":
            normalized_rule["timing"] = "immediate"
        else:
            normalized_rule["days"] = days
        normalized_rules.append(normalized_rule)
    return normalized_rules


def crm_service_task_rules(service_key: str, applicability: str, template_config: Optional[dict] = None) -> List[dict]:
    template_config = template_config or {}
    configured_rules = crm_normalize_template_task_rules(template_config.get("task_rules"))
    if template_config.get("task_rules_configured") or configured_rules:
        return configured_rules
    flow_profile = template_config.get("flow_profile") if isinstance(template_config.get("flow_profile"), dict) else {}
    step_rules: List[dict] = []
    for step in crm_enabled_flow_steps(flow_profile):
        for rule in crm_normalize_template_task_rules(step.get("task_rules")):
            rule["flow_step_key"] = rule.get("flow_step_key") or step.get("step_key") or ""
            step_rules.append(rule)
    if step_rules:
        return step_rules
    if service_key == "renters_insurance":
        return [
            {
                "key": "insurance_ddl_reminder_3",
                "title": "Insurance deadline reminder — 3 days before",
                "days": 3,
                "due_hour": 9,
                "task_type": "deadline_check",
                "priority": "high",
                "description": "The customer purchases or submits insurance under the SOP; staff only confirms completion.",
            },
            {
                "key": "insurance_ddl_due",
                "title": "Renters-insurance deadline",
                "days": 0,
                "due_hour": 0,
                "task_type": "deadline_check",
                "priority": "urgent",
                "description": "The insurance deadline defaults to the lease start date. The customer purchases or submits under the SOP; staff only confirms completion.",
            },
        ]
    if service_key == "electricity_account":
        if applicability == "not_needed":
            return []
        return [
            {"key": "check_7", "title": "Check electricity account before move-in", "days": 7, "task_type": "deadline_check", "priority": "high"},
            {"key": "escalate_3", "title": "Escalate incomplete electricity account", "days": 3, "task_type": "escalation", "priority": "urgent"},
        ]
    if service_key == "internet_setup":
        if applicability == "not_needed":
            return []
        return []
    if service_key == "phone_card":
        return []
    return []


def cancel_legacy_insurance_tasks(conn: sqlite3.Connection, case_id: str, service_id: str) -> None:
    timestamp = now_iso()
    legacy_rules = {
        "renters_insurance:available",
        "renters_insurance:follow_14",
        "renters_insurance:check_7",
        "renters_insurance:escalate_3",
    }
    conn.execute(
        f"""
        UPDATE crm_tasks
        SET status = 'cancelled', updated_at = ?
        WHERE case_id = ?
          AND case_service_id = ?
          AND status NOT IN ('completed', 'cancelled', 'done')
          AND created_from_rule IN ({",".join("?" for _ in legacy_rules)})
        """,
        (timestamp, case_id, service_id, *sorted(legacy_rules)),
    )


def cancel_open_crm_service_tasks(conn: sqlite3.Connection, service_id: str, timestamp: Optional[str] = None) -> None:
    conn.execute(
        """
        UPDATE crm_tasks
        SET status = 'cancelled', updated_at = ?
        WHERE case_service_id = ?
          AND status NOT IN ('completed', 'cancelled', 'done')
        """,
        (timestamp or now_iso(), service_id),
    )


def apply_insurance_overdue_risk_if_needed(
    conn: sqlite3.Connection,
    *,
    case_id: str,
    service: dict,
    customer_rows: List[dict],
    assigned_user_id: str,
    lease_date: Optional[datetime.date],
) -> None:
    if service.get("service_key") != "renters_insurance":
        return
    if service.get("applicability") != "required" or not lease_date:
        return
    due_at = crm_task_due_at(lease_date, days_before=0, due_hour=0)
    due_dt = datetime.fromisoformat(due_at)
    if due_dt > datetime.now():
        return
    service_status = crm_normalize_service_status(service.get("service_status") or service.get("status") or "")
    completion_status = crm_normalize_completion_status(service.get("completion_status") or "")
    if service_status in {"completed", "terminated", "not_needed"} or completion_status in {"completed", "waived", "not_applicable"}:
        return
    timestamp = now_iso()
    flow_snapshot = json_loads_safe(service.get("flow_snapshot_json"), {})
    risk_step = next((step for step in crm_enabled_flow_steps(flow_snapshot) if step.get("is_risk")), None)
    if not risk_step:
        risk_step = crm_flow_step_by_key(flow_snapshot, "insurance_at_risk")
    if risk_step:
        active_flow_step_key = risk_step.get("step_key") or ""
        staff_flow_status = crm_normalize_staff_flow_status(risk_step.get("staff_flow_status") or "", "following_up")
        customer_flow_status = crm_normalize_customer_flow_status_for_profile(
            risk_step.get("customer_flow_status") or "",
            "intent_unknown",
            flow_snapshot,
        )
    else:
        active_flow_step_key = ""
        staff_flow_status = "following_up"
        customer_flow_status = "intent_unknown"
    legacy_fields = crm_legacy_fields_from_flows(
        service_key="renters_insurance",
        applicability=service.get("applicability") or "required",
        staff_flow_status=staff_flow_status,
        customer_flow_status=customer_flow_status,
        flow_profile=flow_snapshot,
        current=dict(service),
    )
    conn.execute(
        """
        UPDATE crm_case_services
        SET status = 'at_risk', service_status = 'at_risk', active_flow_step_key = ?,
            staff_flow_status = ?, customer_flow_status = ?,
            need_status = ?, submission_status = ?, completion_status = ?,
            intro_status = ?, follow_up_status = 'overdue', agent_completion_status = ?,
            blocked_reason = COALESCE(NULLIF(blocked_reason, ''), 'Insurance deadline reached without confirmed completion'),
            updated_at = ?
        WHERE id = ?
          AND COALESCE(service_status, status) NOT IN ('completed', 'terminated', 'not_needed')
        """,
        (
            active_flow_step_key,
            staff_flow_status,
            customer_flow_status,
            legacy_fields["need_status"],
            legacy_fields["submission_status"],
            legacy_fields["completion_status"],
            legacy_fields["intro_status"],
            legacy_fields["agent_completion_status"],
            timestamp,
            service["id"],
        ),
    )
    responsible_customer_id = service.get("responsible_customer_id") or ""
    responsible = next((row for row in customer_rows if row["id"] == responsible_customer_id), None)
    customer_name = (responsible.get("full_name") or "").strip() if responsible else ""
    task_due_at = datetime.combine(datetime.now().date(), datetime.min.time()).replace(hour=9).isoformat()
    crm_insert_task_if_absent(
        conn,
        case_id=case_id,
        case_service_id=service["id"],
        customer_id=responsible_customer_id,
        title=crm_task_title_for_responsible_customer(
            "Renters-insurance deadline passed; confirm risk or incompletion",
            customer_name,
            customer_rows,
        ),
        description="The insurance deadline has arrived or passed, but the service is not confirmed complete. Staff must record the risk and continue follow-up.",
        task_type="escalation",
        due_at=task_due_at,
        not_before_at=task_due_at,
        priority="urgent",
        assigned_user_id=assigned_user_id,
        created_from_rule="renters_insurance:insurance_ddl_risk",
        dedupe_key=f"{case_id}:{service['id']}:insurance_ddl_risk",
    )


def ensure_crm_tasks_for_case(conn: sqlite3.Connection, case_id: str) -> None:
    case_row = crm_case_row(conn, case_id)
    if not case_row:
        return
    lease_date = parse_iso_date(case_row.get("lease_start_date"))
    assigned_user_id = case_row.get("owner_user_id") or ""
    service_rows = conn.execute(
        """
        SELECT crm_case_services.*, crm_service_templates.config_json AS template_config_json
        FROM crm_case_services
        LEFT JOIN crm_service_templates ON crm_service_templates.id = crm_case_services.template_id
        WHERE crm_case_services.case_id = ?
        """,
        (case_id,),
    ).fetchall()
    customer_rows = conn.execute(
        "SELECT * FROM crm_case_guests WHERE case_id = ? AND status != 'inactive' ORDER BY created_at ASC",
        (case_id,),
    ).fetchall()
    customers_by_id = {row["id"]: row for row in customer_rows}
    for service in service_rows:
        if service["service_key"] == "renters_insurance":
            cancel_legacy_insurance_tasks(conn, case_id, service["id"])
        cancel_low_value_system_tasks(conn, case_id=case_id, service_id=service["id"])
        template_config = json_loads_safe(service.get("template_config_json"), {})
        rules = crm_service_task_rules(service["service_key"], service["applicability"], template_config)
        service_scope = service.get("service_scope") or crm_service_scope(service["service_key"])
        completion_status = crm_normalize_completion_status(service.get("completion_status") or "")
        if completion_status in {"completed", "failed"}:
            continue
        if completion_status == "waived" and service["service_key"] not in {"internet_setup", "electricity_account"}:
            continue
        if crm_service_requires_responsible(service_scope, service["applicability"], completion_status):
            responsible_customer_id = service.get("responsible_customer_id") or ""
            responsible_customer = customers_by_id.get(responsible_customer_id)
            if not responsible_customer:
                due_at = datetime.combine(datetime.now().date(), datetime.min.time()).replace(hour=9).isoformat()
                crm_insert_task_if_absent(
                    conn,
                    case_id=case_id,
                    case_service_id=service["id"],
                    title=f"Assign an owner for [{service['service_name']}]",
                    description="A group-level service needs one specific customer assigned to submit, open the account, or coordinate.",
                    task_type="assign_responsible_customer",
                    due_at=due_at,
                    not_before_at=due_at,
                    priority="high",
                    assigned_user_id=assigned_user_id,
                    created_from_rule=f"{service['service_key']}:assign_responsible_customer",
                    dedupe_key=f"{case_id}:{service['id']}:assign_responsible_customer",
                )
                targets = [None] if service["service_key"] == "renters_insurance" else []
                if not targets:
                    continue
            else:
                targets = [responsible_customer]
        else:
            targets = customer_rows if service_scope == "customer_level" else [None]
        for target in targets:
            customer_id = target["id"] if target else ""
            customer_suffix = f" - {target['full_name']}" if target else ""
            for rule in rules:
                due_at = crm_task_due_at_for_rule(lease_date, rule)
                due_day = due_at[:10]
                crm_insert_task_if_absent(
                    conn,
                    case_id=case_id,
                    case_service_id=service["id"],
                    customer_id=customer_id,
                    title=f"{rule['title']}{customer_suffix}",
                    description=rule.get("description") or "",
                    task_type=rule.get("task_type") or "follow_up",
                    due_at=due_at,
                    not_before_at=due_at,
                    priority=rule.get("priority") or "normal",
                    assigned_user_id=assigned_user_id,
                    created_from_rule=f"{service['service_key']}:{rule['key']}",
                    dedupe_key=f"{case_id}:{service['id']}:{customer_id}:{rule['key']}:{due_day}",
                )
        cancel_low_value_system_tasks(conn, case_id=case_id, service_id=service["id"])
        apply_insurance_overdue_risk_if_needed(
            conn,
            case_id=case_id,
            service=service,
            customer_rows=customer_rows,
            assigned_user_id=assigned_user_id,
            lease_date=lease_date,
        )

    for rule in [
        {"key": "case_check_7", "title": "Review all incomplete items 7 days before move-in", "days": 7, "priority": "high"},
        {"key": "case_escalate_3", "title": "Escalate all incomplete items 3 days before move-in", "days": 3, "priority": "urgent"},
    ]:
        due_at = crm_task_due_at(lease_date, days_before=rule["days"])
        crm_insert_task_if_absent(
            conn,
            case_id=case_id,
            title=rule["title"],
            description="Review insurance, electricity, internet, and SIM-card items that remain incomplete or blocked.",
            task_type="deadline_check" if rule["days"] == 7 else "escalation",
            due_at=due_at,
            priority=rule["priority"],
            assigned_user_id=assigned_user_id,
            created_from_rule=f"case:{rule['key']}",
            dedupe_key=f"{case_id}:{rule['key']}:{due_at[:10]}",
        )


def build_crm_building_snapshot(
    conn: sqlite3.Connection,
    building_source: str,
    building_id: str,
) -> dict:
    source = (building_source or "").strip()
    if source == "master":
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="The Master building record does not exist.")
    elif source == "staging":
        snapshot = load_staging_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="The Staging building record does not exist.")
    else:
        return {}

    insurance = insurance_values_from_snapshot(snapshot)
    move_in = move_in_values_from_snapshot(snapshot)
    network_notes = network_provider_note_values_from_snapshot(snapshot)
    network_plans = network_plan_values_from_snapshot(snapshot)
    return {
        "source": source,
        "source_label": "Master" if source == "master" else "Staging (not yet in Master)",
        "building_id": snapshot.get("id") or snapshot.get("staging_key") or building_id,
        "building_name": snapshot.get("building_name") or "",
        "address": snapshot.get("address") or "",
        "is_staging": source == "staging",
        "insurance": {
            "required": crm_insurance_applicability(snapshot),
            "renters_required": insurance.get("insurance_renters_required") or "",
            "renters_minimum_coverage": insurance.get("insurance_renters_minimum_coverage") or "",
            "coi_required": insurance.get("insurance_coi_required") or "",
            "coi_trigger": insurance.get("insurance_coi_trigger") or "",
            "submission_method": insurance.get("insurance_submission_method") or "",
            "recipient": insurance.get("insurance_recipient") or "",
        },
        "electricity": {
            "required": crm_requirement_applicability(snapshot.get("electricity_required")),
            "provider": normalize_unknown_value(snapshot.get("electricity_provider")) or "",
        },
        "internet": {
            "required": crm_requirement_applicability(snapshot.get("internet_self_setup_required")),
            "providers": normalize_unknown_value(snapshot.get("internet_provider")) or "",
            "notes": normalize_unknown_value(snapshot.get("internet_notes")) or "",
            "provider_notes": {
                key: value for key, value in network_notes.items() if value
            },
            "plan_tiers": {
                key: format_network_plan_text(value) for key, value in network_plans.items() if value
            },
        },
        "move_in": {
            "key_pickup_notes": move_in.get("key_pickup_notes") or "",
            "service_elevator_booking_notes": move_in.get("service_elevator_booking_notes") or "",
            "move_in_notes": move_in.get("move_in_notes") or "",
        },
        "source_date": snapshot.get("source_date") or "",
        "updated_at": snapshot.get("updated_at") or "",
        "library_status": snapshot.get("library_status") or ("Master" if source == "master" else "Needs enrichment"),
    }


def ensure_crm_case_services(
    conn: sqlite3.Connection,
    *,
    case_id: str,
    building_snapshot: Optional[dict],
) -> None:
    timestamp = now_iso()
    customer_rows = crm_active_customer_rows(conn, case_id)
    existing = {
        row["service_key"]: row
        for row in conn.execute(
            "SELECT * FROM crm_case_services WHERE case_id = ?",
            (case_id,),
        ).fetchall()
    }
    for template in active_crm_templates(conn):
        applicability = crm_service_applicability(template["service_key"], building_snapshot)
        service_scope = crm_template_service_scope(template)
        default_fields = crm_initial_status_fields(template["service_key"], applicability)
        default_flow_fields = crm_default_flow_fields(template["service_key"], applicability)
        flow_snapshot = crm_flow_snapshot_for_template(template)
        current = existing.get(template["service_key"])
        if current:
            current_flow_snapshot = crm_effective_service_flow_snapshot(current.get("flow_snapshot_json"), template)
            legacy_flow_fields = crm_flows_from_legacy(current, applicability)
            next_staff_flow = crm_normalize_staff_flow_status(
                current.get("staff_flow_status") or "",
                legacy_flow_fields["staff_flow_status"],
            )
            next_customer_flow = crm_normalize_customer_flow_status_for_profile(
                current.get("customer_flow_status") or "",
                legacy_flow_fields["customer_flow_status"],
                current_flow_snapshot,
            )
            next_service_status = crm_normalize_service_status(
                current.get("service_status") or current.get("status") or "",
                legacy_flow_fields["service_status"],
            )
            if applicability == "not_needed":
                next_staff_flow = default_flow_fields["staff_flow_status"]
                next_customer_flow = default_flow_fields["customer_flow_status"]
                next_service_status = default_flow_fields["service_status"]
                next_active_flow_step_key = ""
            else:
                next_active_flow_step_key = current.get("active_flow_step_key") or crm_default_flow_step_key(
                    current_flow_snapshot,
                    next_staff_flow,
                )
                if current.get("active_flow_step_key"):
                    step_fields = crm_flow_fields_from_step_key(
                        current_flow_snapshot,
                        next_active_flow_step_key,
                        applicability=applicability,
                        fallback_staff_flow_status=next_staff_flow,
                        fallback_customer_flow_status=next_customer_flow,
                    )
                    next_active_flow_step_key = step_fields["active_flow_step_key"]
                    next_staff_flow = step_fields["staff_flow_status"]
                    next_customer_flow = step_fields["customer_flow_status"]
                    next_service_status = step_fields["service_status"]
            next_status = next_service_status or current["status"]
            if applicability == "not_needed" and current["status"] in {"pending", "not_needed", "terminated"}:
                next_status = "not_needed"
            elif current["status"] in {"pending", "not_needed"}:
                next_status = crm_service_status_from_flows(next_staff_flow, next_customer_flow, applicability, current_flow_snapshot)
            current_need = current.get("need_status") or ""
            current_submission = current.get("submission_status") or ""
            current_completion = current.get("completion_status") or ""
            current_intro = current.get("intro_status") or ""
            current_follow_up = current.get("follow_up_status") or ""
            current_agent_completion = current.get("agent_completion_status") or ""

            next_need = crm_normalize_need_status(current_need, default_fields["need_status"])
            if current_need in {"", "unknown", "needed", "waived"} or applicability == "not_needed":
                next_need = default_fields["need_status"]

            next_submission = crm_normalize_submission_status(
                current_submission,
                default_fields["submission_status"],
            )
            if current_submission in {"", "not_started"} and default_fields["submission_status"] != "not_submitted":
                next_submission = default_fields["submission_status"]

            mapped_completion = crm_completion_from_service_status(
                next_status,
                default_fields["completion_status"],
            )
            next_completion = crm_normalize_completion_status(current_completion, mapped_completion)
            if current_completion in {"", "not_started"} and mapped_completion != "not_started":
                next_completion = mapped_completion

            next_intro = crm_normalize_intro_status(current_intro, default_fields["intro_status"])
            if current_intro in {"", "not_introduced"} and default_fields["intro_status"] != "not_introduced":
                next_intro = default_fields["intro_status"]

            next_follow_up = crm_normalize_follow_up_status(
                current_follow_up,
                default_fields["follow_up_status"],
            )
            if current_follow_up in {"", "not_started"} and default_fields["follow_up_status"] != "required":
                next_follow_up = default_fields["follow_up_status"]

            next_agent_completion = crm_normalize_agent_completion_status(
                current_agent_completion,
                default_fields["agent_completion_status"],
            )
            if current_agent_completion in {"", "not_started", "action_required"} and default_fields["agent_completion_status"] != "open":
                next_agent_completion = default_fields["agent_completion_status"]
            responsibility = crm_case_service_responsibility_state(
                customer_rows,
                service_scope=service_scope,
                applicability=applicability,
                completion_status=next_completion,
                current_responsible_customer_id=current.get("responsible_customer_id") or "",
                current_responsibility_status=current.get("responsibility_status") or "",
            )
            conn.execute(
                """
                UPDATE crm_case_services
                SET template_id = ?, service_name = ?, service_scope = ?, applicability = ?,
                    responsible_customer_id = ?, covered_customer_ids = ?, responsibility_status = ?,
                    status = ?, active_flow_step_key = ?, staff_flow_status = ?, customer_flow_status = ?, service_status = ?,
                    flow_snapshot_json = ?,
                    need_status = ?, submission_status = ?, completion_status = ?,
                    intro_status = ?, follow_up_status = ?, agent_completion_status = ?,
                    updated_at = ?
                WHERE id = ?
                """,
                (
                    template["id"],
                    template["name"],
                    service_scope,
                    applicability,
                    responsibility["responsible_customer_id"],
                    json_dumps(responsibility["covered_customer_ids"]),
                    responsibility["responsibility_status"],
                    next_status,
                    next_active_flow_step_key,
                    next_staff_flow,
                    next_customer_flow,
                    next_status,
                    json_dumps(current_flow_snapshot),
                    next_need,
                    next_submission,
                    next_completion,
                    next_intro,
                    next_follow_up,
                    next_agent_completion,
                    timestamp,
                    current["id"],
                ),
            )
            continue
        responsibility = crm_case_service_responsibility_state(
            customer_rows,
            service_scope=service_scope,
            applicability=applicability,
            completion_status=default_fields["completion_status"],
        )
        service_status = default_flow_fields["service_status"]
        active_flow_step_key = "" if applicability == "not_needed" else crm_default_flow_step_key(
            flow_snapshot,
            default_flow_fields["staff_flow_status"],
        )
        conn.execute(
            """
            INSERT INTO crm_case_services(
              id, case_id, template_id, service_key, service_name, service_scope,
              responsible_customer_id, covered_customer_ids, responsibility_status,
              active_flow_step_key, staff_flow_status, customer_flow_status, service_status, flow_snapshot_json,
              applicability, status, need_status, submission_status, completion_status,
              intro_status, follow_up_status, agent_completion_status,
              group_progress_json, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, '{}', ?, ?)
            """,
            (
                f"crm_service_{uuid.uuid4().hex}",
                case_id,
                template["id"],
                template["service_key"],
                template["name"],
                service_scope,
                responsibility["responsible_customer_id"],
                json_dumps(responsibility["covered_customer_ids"]),
                responsibility["responsibility_status"],
                active_flow_step_key,
                default_flow_fields["staff_flow_status"],
                default_flow_fields["customer_flow_status"],
                service_status,
                json_dumps(flow_snapshot),
                applicability,
                service_status,
                default_fields["need_status"],
                default_fields["submission_status"],
                default_fields["completion_status"],
                default_fields["intro_status"],
                default_fields["follow_up_status"],
                default_fields["agent_completion_status"],
                timestamp,
                timestamp,
            ),
        )
    ensure_crm_tasks_for_case(conn, case_id)


def crm_case_row(conn: sqlite3.Connection, case_id: str) -> Optional[dict]:
    return conn.execute(
        """
        SELECT
          crm_cases.*,
          owner.display_name AS owner_display_name,
          owner.username AS owner_username,
          creator.display_name AS creator_display_name,
          creator.username AS creator_username
        FROM crm_cases
        LEFT JOIN users AS owner ON owner.id = crm_cases.owner_user_id
        LEFT JOIN users AS creator ON creator.id = crm_cases.created_by
        WHERE crm_cases.id = ?
        """,
        (case_id,),
    ).fetchone()


def crm_case_is_deleted(case_row: dict) -> bool:
    return bool(normalize_unknown_value(case_row.get("deleted_at")))


def crm_actor_can_access_case(actor: Actor, case_row: dict) -> bool:
    if actor.role == "super_admin":
        return True
    return case_row.get("owner_user_id") == actor.user_id


def require_crm_case_access(conn: sqlite3.Connection, case_id: str, actor: Actor) -> dict:
    case_row = crm_case_row(conn, case_id)
    if not case_row:
        raise HTTPException(status_code=404, detail="The CRM Case does not exist.")
    if not crm_actor_can_access_case(actor, case_row):
        raise HTTPException(status_code=403, detail="You can view and manage only CRM Cases assigned to you.")
    return case_row


def serialize_crm_case(row: dict) -> dict:
    snapshot = json_loads_safe(row.get("building_snapshot_json"), {})
    return {
        "id": row["id"],
        "group_name": row.get("group_name") or "",
        "owner_user_id": row.get("owner_user_id") or "",
        "owner_display_name": row.get("owner_display_name") or row.get("owner_username") or "",
        "unit": row.get("unit") or "",
        "group_creator_name": row.get("group_creator_name") or "",
        "group_creator_contact": row.get("group_creator_contact") or "",
        "agent_team_t": row.get("agent_team_t") or "",
        "agent_team_m": row.get("agent_team_m") or "",
        "lease_start_date": row.get("lease_start_date") or "",
        "lease_days_from_today": lease_days_from_today(row.get("lease_start_date") or ""),
        "building_source": row.get("building_source") or "",
        "building_id": row.get("building_id") or "",
        "building_name": row.get("building_name") or "",
        "building_address": row.get("building_address") or "",
        "building_snapshot": snapshot,
        "insurance_earliest_start_date": row.get("insurance_earliest_start_date") or "",
        "network_earliest_start_note": row.get("network_earliest_start_note") or "",
        "status": row.get("status") or "active",
        "notes": row.get("notes") or "",
        "created_by": row.get("created_by") or "",
        "creator_display_name": row.get("creator_display_name") or row.get("creator_username") or "",
        "open_task_count": int(row.get("open_task_count") or 0),
        "deleted_at": row.get("deleted_at") or "",
        "deleted_by": row.get("deleted_by") or "",
        "delete_reason": row.get("delete_reason") or "",
        "is_deleted": bool(row.get("deleted_at")),
        "created_at": row.get("created_at") or "",
        "updated_at": row.get("updated_at") or "",
    }


def serialize_crm_service(conn: sqlite3.Connection, service_row: dict) -> dict:
    template = conn.execute(
        "SELECT * FROM crm_service_templates WHERE id = ?",
        (service_row["template_id"],),
    ).fetchone()
    group_progress_rows = conn.execute(
        """
        SELECT *
        FROM crm_case_service_progress
        WHERE case_service_id = ?
        ORDER BY updated_at DESC
        """,
        (service_row["id"],),
    ).fetchall()
    flow_snapshot = crm_effective_service_flow_snapshot(service_row.get("flow_snapshot_json"), template)
    return {
        "id": service_row["id"],
        "case_id": service_row["case_id"],
        "template_id": service_row["template_id"],
        "service_key": service_row["service_key"],
        "service_name": service_row["service_name"],
        "service_scope": service_row.get("service_scope") or crm_service_scope(service_row["service_key"]),
        "responsible_customer_id": service_row.get("responsible_customer_id") or "",
        "covered_customer_ids": json_loads_safe(service_row.get("covered_customer_ids"), []),
        "responsibility_status": crm_normalize_responsibility_status(service_row.get("responsibility_status") or ""),
        "active_flow_step_key": service_row.get("active_flow_step_key") or "",
        "applicability": service_row["applicability"],
        "staff_flow_status": crm_normalize_staff_flow_status(
            service_row.get("staff_flow_status") or "",
            crm_flows_from_legacy(service_row, service_row.get("applicability") or "unknown")["staff_flow_status"],
        ),
        "customer_flow_status": crm_normalize_customer_flow_status_for_profile(
            service_row.get("customer_flow_status") or "",
            crm_flows_from_legacy(service_row, service_row.get("applicability") or "unknown")["customer_flow_status"],
            flow_snapshot,
        ),
        "service_status": crm_normalize_service_status(
            service_row.get("service_status") or service_row.get("status") or "",
            crm_flows_from_legacy(service_row, service_row.get("applicability") or "unknown")["service_status"],
        ),
        "status": service_row.get("service_status") or service_row["status"],
        "termination_reason": service_row.get("termination_reason") or "",
        "flow_snapshot": flow_snapshot,
        "need_status": crm_normalize_need_status(service_row.get("need_status") or ""),
        "submission_status": crm_normalize_submission_status(service_row.get("submission_status") or ""),
        "completion_status": crm_normalize_completion_status(service_row.get("completion_status") or ""),
        "intro_status": crm_normalize_intro_status(service_row.get("intro_status") or ""),
        "follow_up_status": crm_normalize_follow_up_status(service_row.get("follow_up_status") or ""),
        "agent_completion_status": crm_normalize_agent_completion_status(service_row.get("agent_completion_status") or ""),
        "blocked_reason": service_row.get("blocked_reason") or "",
        "group_progress": json_loads_safe(service_row.get("group_progress_json"), {}),
        "step_progress": [
            {
                "id": row["id"],
                "step_key": row["step_key"],
                "scope": row["scope"],
                "value": json_loads_safe(row.get("value_json"), {}),
                "note": row.get("note") or "",
                "updated_by": row.get("updated_by") or "",
                "updated_at": row.get("updated_at") or "",
            }
            for row in group_progress_rows
        ],
        "template": serialize_crm_template(conn, template) if template else None,
        "created_at": service_row.get("created_at") or "",
        "updated_at": service_row.get("updated_at") or "",
    }


def serialize_crm_guest(row: dict, progress_rows: List[dict], *, reveal_sensitive: bool) -> dict:
    return {
        "id": row["id"],
        "case_id": row["case_id"],
        "full_name": row.get("full_name") or "",
        "phone": row.get("phone") or "",
        "email": row.get("email") or "",
        "wechat": row.get("wechat") or "",
        "notes": row.get("notes") or "",
        "status": row.get("status") or "active",
        "service_progress": [
            {
                "id": progress["id"],
                "service_id": progress["case_service_id"],
                "service_key": progress.get("service_key") or "",
                "step_key": progress["step_key"],
                "value": json_loads_safe(progress.get("value_json"), {}),
                "sensitive": json_loads_safe(progress.get("sensitive_json"), {}) if reveal_sensitive else {},
                "sensitive_masked": bool(json_loads_safe(progress.get("sensitive_json"), {})) and not reveal_sensitive,
                "note": progress.get("note") or "",
                "updated_by": progress.get("updated_by") or "",
                "updated_at": progress.get("updated_at") or "",
            }
            for progress in progress_rows
        ],
        "created_at": row.get("created_at") or "",
        "updated_at": row.get("updated_at") or "",
    }


def serialize_crm_task(row: dict) -> dict:
    normalized_status = crm_normalize_task_status(row.get("status") or "")
    is_overdue = crm_task_is_overdue(row)
    assigned_to = row.get("assigned_to") or row.get("assigned_user_id") or ""
    customer_id = row.get("target_customer_id") or row.get("customer_id") or ""
    return {
        "id": row["id"],
        "task_id": row["id"],
        "case_id": row["case_id"],
        "case_service_id": row.get("case_service_id") or "",
        "customer_id": customer_id,
        "target_customer_id": customer_id,
        "title": row.get("title") or "",
        "task_title": row.get("title") or "",
        "description": row.get("description") or "",
        "task_type": row.get("task_type") or "manual",
        "due_at": row.get("due_at") or "",
        "not_before_at": row.get("not_before_at") or "",
        "priority": row.get("priority") or "normal",
        "status": normalized_status,
        "task_status": normalized_status,
        "raw_status": row.get("status") or "",
        "is_overdue": is_overdue,
        "assigned_user_id": assigned_to,
        "assigned_to": assigned_to,
        "assigned_to_name": row.get("assigned_to_name") or row.get("assigned_username") or "",
        "source": row.get("source") or "system",
        "created_from_rule": row.get("created_from_rule") or "",
        "service_key": row.get("service_key") or "",
        "service_type": row.get("service_key") or "",
        "service_name": row.get("service_name") or "",
        "customer_name": row.get("customer_name") or "",
        "target_customer_name": row.get("customer_name") or "",
        "group_name": row.get("group_name") or "",
        "case_group_name": row.get("group_name") or "",
        "case_status": row.get("case_status") or "",
        "case_building_name": row.get("case_building_name") or "",
        "case_unit": row.get("case_unit") or "",
        "service_need_status": crm_normalize_need_status(row.get("service_need_status") or ""),
        "service_submission_status": crm_normalize_submission_status(row.get("service_submission_status") or ""),
        "service_completion_status": crm_normalize_completion_status(row.get("service_completion_status") or ""),
        "service_intro_status": crm_normalize_intro_status(row.get("service_intro_status") or ""),
        "service_follow_up_status": crm_normalize_follow_up_status(row.get("service_follow_up_status") or ""),
        "service_agent_completion_status": crm_normalize_agent_completion_status(row.get("service_agent_completion_status") or ""),
        "service_staff_flow_status": crm_normalize_staff_flow_status(row.get("service_staff_flow_status") or ""),
        "service_customer_flow_status": row.get("service_customer_flow_status") or "",
        "service_status": crm_normalize_service_status(row.get("service_status") or row.get("service_legacy_status") or ""),
        "created_at": row.get("created_at") or "",
        "updated_at": row.get("updated_at") or "",
        "completed_at": row.get("completed_at") or "",
    }


def serialize_crm_communication_event(row: dict) -> dict:
    return {
        "id": row["id"],
        "case_id": row["case_id"],
        "case_service_id": row.get("case_service_id") or "",
        "customer_id": row.get("customer_id") or "",
        "channel": row.get("channel") or "wechat_group",
        "direction": row.get("direction") or "internal",
        "summary": row.get("summary") or "",
        "raw_ref": json_loads_safe(row.get("raw_ref_json"), {}),
        "created_by": row.get("created_by") or "",
        "service_name": row.get("service_name") or "",
        "customer_name": row.get("customer_name") or "",
        "created_at": row.get("created_at") or "",
    }


def serialize_crm_notification(row: dict) -> dict:
    return {
        "id": row["id"],
        "case_id": row["case_id"],
        "task_id": row.get("task_id") or "",
        "case_service_id": row.get("case_service_id") or "",
        "channel": row.get("channel") or "wechat",
        "recipient_type": row.get("recipient_type") or "group",
        "recipient_ref": row.get("recipient_ref") or "",
        "content": row.get("content") or "",
        "status": row.get("status") or "draft",
        "generated_by": row.get("generated_by") or "system",
        "service_name": row.get("service_name") or "",
        "task_title": row.get("task_title") or "",
        "sent_at": row.get("sent_at") or "",
        "created_at": row.get("created_at") or "",
        "updated_at": row.get("updated_at") or "",
    }


CRM_TIMELINE_ACTION_LABELS = {
    "crm_case_created": "Created Case",
    "crm_case_updated": "Updated Case",
    "crm_case_building_bound": "Linked building",
    "crm_case_building_snapshot_refreshed": "Refreshed building snapshot",
    "crm_case_services_generated": "Generated services and tasks",
    "crm_service_progress_updated": "Updated service status",
    "crm_guest_service_progress_updated": "Updated customer service progress",
    "crm_task_created": "Created task",
    "crm_task_updated": "Updated task",
    "crm_notification_approved": "Approved notification",
    "crm_notification_sent": "Marked notification as sent",
}


def crm_timeline_actor(row: dict) -> tuple[str, str]:
    actor_id = row.get("user_id") or row.get("created_by") or row.get("generated_by") or ""
    actor_name = row.get("actor_display_name") or row.get("actor_username") or row.get("created_by_name") or actor_id
    return actor_id, actor_name or "System"


def crm_timeline_audit_title(row: dict) -> str:
    action = row.get("action_type") or ""
    if action == "crm_task_updated" and row.get("new_value") == "completed":
        return "Completed task"
    if action == "crm_service_progress_updated":
        service_name = row.get("service_name") or ""
        return f"Updated service status: {service_name}" if service_name else "Updated service status"
    return CRM_TIMELINE_ACTION_LABELS.get(action, action or "System record")


def crm_timeline_audit_summary(row: dict) -> str:
    action = row.get("action_type") or ""
    if action == "crm_task_created":
        return row.get("task_title") or row.get("new_value") or "Created a task."
    if action == "crm_task_updated":
        status = row.get("new_value") or ""
        if status == "completed":
            return f"Task completed: {row.get('task_title') or ''}".strip()
        return f"Task status updated: {status}" if status else row.get("task_title") or ""
    if action == "crm_service_progress_updated":
        old_value = normalize_unknown_value(row.get("old_value")) or ""
        new_value = normalize_unknown_value(row.get("new_value")) or ""
        if new_value:
            return f"Service status updated: {new_value[:220]}"
        if old_value:
            return f"Service status changed: {old_value[:220]}"
    return normalize_unknown_value(row.get("note")) or normalize_unknown_value(row.get("new_value")) or ""


def load_crm_case_timeline(conn: sqlite3.Connection, case_id: str) -> List[dict]:
    audit_rows = conn.execute(
        """
        SELECT audit_logs.*, users.display_name AS actor_display_name, users.username AS actor_username,
               crm_case_services.service_name,
               crm_tasks.title AS task_title,
               crm_case_guests.full_name AS customer_name
        FROM audit_logs
        LEFT JOIN users ON users.id = audit_logs.user_id
        LEFT JOIN crm_case_services
          ON audit_logs.target_table = 'crm_case_services'
         AND crm_case_services.id = audit_logs.target_record_id
        LEFT JOIN crm_tasks
          ON audit_logs.target_table = 'crm_tasks'
         AND crm_tasks.id = audit_logs.target_record_id
        LEFT JOIN crm_case_guests
          ON crm_case_guests.id = COALESCE(crm_tasks.target_customer_id, crm_tasks.customer_id)
        WHERE (
          audit_logs.target_table = 'crm_cases' AND audit_logs.target_record_id = ?
        ) OR (
          audit_logs.target_table = 'crm_case_services'
          AND audit_logs.target_record_id IN (SELECT id FROM crm_case_services WHERE case_id = ?)
        ) OR (
          audit_logs.target_table = 'crm_tasks'
          AND audit_logs.target_record_id IN (SELECT id FROM crm_tasks WHERE case_id = ?)
        ) OR (
          audit_logs.target_table = 'crm_notifications'
          AND audit_logs.action_type != 'crm_notification_draft_created'
          AND audit_logs.target_record_id IN (SELECT id FROM crm_notifications WHERE case_id = ?)
        )
        ORDER BY datetime(audit_logs.created_at) DESC
        LIMIT 100
        """,
        (case_id, case_id, case_id, case_id),
    ).fetchall()
    timeline: List[dict] = []
    for row in audit_rows:
        actor_id, actor_name = crm_timeline_actor(row)
        timeline.append(
            {
                "id": row["id"],
                "event_type": row.get("action_type") or "audit",
                "title": crm_timeline_audit_title(row),
                "actor_id": actor_id,
                "actor_name": actor_name,
                "occurred_at": row.get("created_at") or "",
                "service_name": row.get("service_name") or "",
                "task_title": row.get("task_title") or "",
                "customer_name": row.get("customer_name") or "",
                "summary": crm_timeline_audit_summary(row),
            }
        )
    communication_rows = conn.execute(
        """
        SELECT crm_communication_events.*, users.display_name AS created_by_name,
               crm_case_services.service_name, crm_case_guests.full_name AS customer_name
        FROM crm_communication_events
        LEFT JOIN users ON users.id = crm_communication_events.created_by
        LEFT JOIN crm_case_services ON crm_case_services.id = crm_communication_events.case_service_id
        LEFT JOIN crm_case_guests ON crm_case_guests.id = crm_communication_events.customer_id
        WHERE crm_communication_events.case_id = ?
          AND crm_communication_events.channel != 'ai_draft'
        ORDER BY datetime(crm_communication_events.created_at) DESC
        LIMIT 60
        """,
        (case_id,),
    ).fetchall()
    for row in communication_rows:
        actor_id, actor_name = crm_timeline_actor(row)
        timeline.append(
            {
                "id": row["id"],
                "event_type": "communication",
                "title": "Recorded communication",
                "actor_id": actor_id,
                "actor_name": actor_name,
                "occurred_at": row.get("created_at") or "",
                "service_name": row.get("service_name") or "",
                "task_title": "",
                "customer_name": row.get("customer_name") or "",
                "summary": row.get("summary") or "",
            }
        )
    notification_rows = conn.execute(
        """
        SELECT crm_notifications.*, users.display_name AS created_by_name,
               crm_case_services.service_name, crm_tasks.title AS task_title
        FROM crm_notifications
        LEFT JOIN users ON users.id = crm_notifications.generated_by
        LEFT JOIN crm_case_services ON crm_case_services.id = crm_notifications.case_service_id
        LEFT JOIN crm_tasks ON crm_tasks.id = crm_notifications.task_id
        WHERE crm_notifications.case_id = ?
        ORDER BY datetime(crm_notifications.created_at) DESC
        LIMIT 40
        """,
        (case_id,),
    ).fetchall()
    for row in notification_rows:
        timeline.append(
            {
                "id": row["id"],
                "event_type": "notification",
                "title": "Generated notification draft" if row.get("status") == "draft" else "Notification record",
                "actor_id": row.get("generated_by") or "",
                "actor_name": row.get("generated_by") or "System",
                "occurred_at": row.get("created_at") or "",
                "service_name": row.get("service_name") or "",
                "task_title": row.get("task_title") or "",
                "customer_name": "",
                "summary": (row.get("content") or "")[:240],
            }
        )
    timeline.sort(key=lambda item: item.get("occurred_at") or "", reverse=True)
    return timeline[:120]


def load_crm_case_detail(conn: sqlite3.Connection, case_id: str, actor: Actor) -> dict:
    normalize_crm_task_rows(conn)
    case_row = require_crm_case_access(conn, case_id, actor)
    if not crm_case_is_deleted(case_row):
        ensure_crm_case_services(
            conn,
            case_id=case_id,
            building_snapshot=json_loads_safe(case_row.get("building_snapshot_json"), {}),
        )
        case_row = crm_case_row(conn, case_id) or case_row
    service_rows = conn.execute(
        """
        SELECT crm_case_services.*
        FROM crm_case_services
        LEFT JOIN crm_service_templates ON crm_service_templates.id = crm_case_services.template_id
        WHERE crm_case_services.case_id = ?
        ORDER BY crm_service_templates.display_order ASC, crm_case_services.service_name ASC
        """,
        (case_id,),
    ).fetchall()
    guest_rows = conn.execute(
        """
        SELECT *
        FROM crm_case_guests
        WHERE case_id = ?
        ORDER BY created_at ASC
        """,
        (case_id,),
    ).fetchall()
    progress_rows = conn.execute(
        """
        SELECT crm_guest_service_progress.*, crm_case_services.service_key
        FROM crm_guest_service_progress
        JOIN crm_case_services ON crm_case_services.id = crm_guest_service_progress.case_service_id
        JOIN crm_case_guests ON crm_case_guests.id = crm_guest_service_progress.guest_id
        WHERE crm_case_guests.case_id = ?
        ORDER BY crm_guest_service_progress.updated_at DESC
        """,
        (case_id,),
    ).fetchall()
    progress_by_guest: Dict[str, List[dict]] = {}
    for progress in progress_rows:
        progress_by_guest.setdefault(progress["guest_id"], []).append(progress)
    task_rows = conn.execute(
        """
        SELECT crm_tasks.*, crm_case_services.service_key, crm_case_services.service_name,
               crm_case_services.need_status AS service_need_status,
               crm_case_services.submission_status AS service_submission_status,
               crm_case_services.completion_status AS service_completion_status,
               crm_case_services.intro_status AS service_intro_status,
               crm_case_services.follow_up_status AS service_follow_up_status,
               crm_case_services.agent_completion_status AS service_agent_completion_status,
               crm_case_services.staff_flow_status AS service_staff_flow_status,
               crm_case_services.customer_flow_status AS service_customer_flow_status,
               crm_case_services.service_status,
               crm_case_services.status AS service_legacy_status,
               crm_case_guests.full_name AS customer_name,
               crm_cases.group_name, crm_cases.status AS case_status,
               crm_cases.building_name AS case_building_name, crm_cases.unit AS case_unit,
               assignee.display_name AS assigned_to_name, assignee.username AS assigned_username
        FROM crm_tasks
        JOIN crm_cases ON crm_cases.id = crm_tasks.case_id
        LEFT JOIN crm_case_services ON crm_case_services.id = crm_tasks.case_service_id
        LEFT JOIN crm_case_guests ON crm_case_guests.id = COALESCE(crm_tasks.target_customer_id, crm_tasks.customer_id)
        LEFT JOIN users AS assignee ON assignee.id = COALESCE(NULLIF(crm_tasks.assigned_to, ''), crm_tasks.assigned_user_id)
        WHERE crm_tasks.case_id = ?
        ORDER BY datetime(crm_tasks.due_at) ASC, crm_tasks.priority DESC
        """,
        (case_id,),
    ).fetchall()
    communication_rows = conn.execute(
        """
        SELECT crm_communication_events.*, crm_case_services.service_name,
               crm_case_guests.full_name AS customer_name
        FROM crm_communication_events
        LEFT JOIN crm_case_services ON crm_case_services.id = crm_communication_events.case_service_id
        LEFT JOIN crm_case_guests ON crm_case_guests.id = crm_communication_events.customer_id
        WHERE crm_communication_events.case_id = ?
        ORDER BY datetime(crm_communication_events.created_at) DESC
        """,
        (case_id,),
    ).fetchall()
    notification_rows = conn.execute(
        """
        SELECT crm_notifications.*, crm_case_services.service_name, crm_tasks.title AS task_title
        FROM crm_notifications
        LEFT JOIN crm_case_services ON crm_case_services.id = crm_notifications.case_service_id
        LEFT JOIN crm_tasks ON crm_tasks.id = crm_notifications.task_id
        WHERE crm_notifications.case_id = ?
        ORDER BY datetime(crm_notifications.created_at) DESC
        """,
        (case_id,),
    ).fetchall()
    reveal_sensitive = actor.role == "super_admin" or case_row.get("owner_user_id") == actor.user_id
    return {
        "case": serialize_crm_case(case_row),
        "services": [serialize_crm_service(conn, row) for row in service_rows],
        "guests": [
            serialize_crm_guest(row, progress_by_guest.get(row["id"], []), reveal_sensitive=reveal_sensitive)
            for row in guest_rows
        ],
        "customers": [
            serialize_crm_guest(row, progress_by_guest.get(row["id"], []), reveal_sensitive=reveal_sensitive)
            for row in guest_rows
        ],
        "tasks": [serialize_crm_task(row) for row in task_rows],
        "communication_events": [serialize_crm_communication_event(row) for row in communication_rows],
        "notifications": [serialize_crm_notification(row) for row in notification_rows],
        "timeline": load_crm_case_timeline(conn, case_id),
    }


def crm_summary_value(value: object, default: str = "Unknown") -> str:
    normalized = normalize_unknown_value(value)
    return normalized or default


def crm_summary_excerpt(value: object, limit: int = 220) -> str:
    text = crm_summary_value(value, "")
    if not text:
        return ""
    collapsed = re.sub(r"\s+", " ", text).strip()
    if len(collapsed) <= limit:
        return collapsed
    return f"{collapsed[:limit].rstrip()}..."


def crm_format_lease_days(days: object) -> str:
    if days is None:
        return "Unknown"
    try:
        day_count = int(days)
    except (TypeError, ValueError):
        return "Unknown"
    if day_count > 0:
        return f"Starts in {day_count} days"
    if day_count == 0:
        return "Starts today"
    return f"Started {abs(day_count)} days ago"


def crm_service_summary_label(service: Optional[dict]) -> str:
    if not service:
        return "Not enabled"
    status_label = CRM_SERVICE_STATUS_LABELS.get(service.get("status"), service.get("status") or "Unknown")
    applicability_label = CRM_APPLICABILITY_LABELS.get(
        service.get("applicability"),
        service.get("applicability") or "To be confirmed",
    )
    return f"{status_label} (building assessment: {applicability_label})"


def build_crm_case_fact_summary(detail: dict) -> str:
    case_item = detail.get("case") or {}
    snapshot = case_item.get("building_snapshot") or {}
    guests = detail.get("guests") or []
    services = detail.get("services") or []
    service_by_key = {service.get("service_key"): service for service in services}

    guest_names = [
        crm_summary_value(guest.get("full_name"), "")
        for guest in guests
        if crm_summary_value(guest.get("full_name"), "")
    ]
    guest_label = f"{len(guest_names)} people"
    if guest_names:
        guest_label += f" ({', '.join(guest_names[:6])}{', and others' if len(guest_names) > 6 else ''})"

    lines = [
        f"CRM Case: {crm_summary_value(case_item.get('group_name'), 'Unnamed Case')} (status: {CRM_CASE_STATUS_LABELS.get(case_item.get('status'), case_item.get('status') or 'Unknown')})",
        f"Assigned support owner: {crm_summary_value(case_item.get('owner_display_name'))}",
        f"Group guests: {guest_label}",
        f"Group creator: {crm_summary_value(case_item.get('group_creator_name'))}; Team T: {crm_summary_value(case_item.get('agent_team_t'))}; Team M: {crm_summary_value(case_item.get('agent_team_m'))}",
        f"Lease start date: {crm_summary_value(case_item.get('lease_start_date'))} ({crm_format_lease_days(case_item.get('lease_days_from_today'))})",
    ]

    if snapshot.get("building_name"):
        source_label = snapshot.get("source_label") or ("Master" if snapshot.get("source") == "master" else "Staging")
        lines.append(
            f"Linked building: {source_label} — {crm_summary_value(snapshot.get('building_name'))}; address: {crm_summary_value(snapshot.get('address'))}"
        )
        if snapshot.get("is_staging"):
            lines.append("The building comes from Staging and is not yet in Master; use this Case summary only as a working service reference.")
    else:
        lines.append("No building is linked, so building-specific insurance, electricity, and internet rules cannot be identified automatically.")

    insurance = snapshot.get("insurance") if isinstance(snapshot.get("insurance"), dict) else {}
    electricity = snapshot.get("electricity") if isinstance(snapshot.get("electricity"), dict) else {}
    internet = snapshot.get("internet") if isinstance(snapshot.get("internet"), dict) else {}
    move_in = snapshot.get("move_in") if isinstance(snapshot.get("move_in"), dict) else {}

    lines.append(
        "Renters insurance: "
        f"{crm_service_summary_label(service_by_key.get('renters_insurance'))}；"
        f"earliest start: {crm_summary_value(case_item.get('insurance_earliest_start_date'))}; "
        f"minimum coverage: {crm_summary_value(insurance.get('renters_minimum_coverage'))}; "
        f"COI：{crm_summary_value(insurance.get('coi_required'))}"
    )
    if insurance.get("coi_trigger"):
        lines.append(f"COI trigger: {crm_summary_excerpt(insurance.get('coi_trigger'))}")
    if insurance.get("submission_method") or insurance.get("recipient"):
        lines.append(
            f"Insurance submission: {crm_summary_value(insurance.get('submission_method'))}; recipient: {crm_summary_value(insurance.get('recipient'))}"
        )

    lines.append(
        "Electricity account: "
        f"{crm_service_summary_label(service_by_key.get('electricity_account'))}；"
        f"provider: {crm_summary_value(electricity.get('provider'))}"
    )
    lines.append(
        "Internet setup: "
        f"{crm_service_summary_label(service_by_key.get('internet_setup'))}；"
        f"available providers: {crm_summary_value(internet.get('providers'))}; "
        f"earliest start: {crm_summary_value(case_item.get('network_earliest_start_note'))}"
    )
    if internet.get("notes"):
        lines.append(f"Internet notes: {crm_summary_excerpt(internet.get('notes'))}")

    if move_in.get("key_pickup_notes"):
        lines.append(f"Key pickup: {crm_summary_excerpt(move_in.get('key_pickup_notes'))}")
    if move_in.get("service_elevator_booking_notes"):
        lines.append(f"Service-elevator booking: {crm_summary_excerpt(move_in.get('service_elevator_booking_notes'))}")
    if move_in.get("move_in_notes"):
        lines.append(f"Move-in notes: {crm_summary_excerpt(move_in.get('move_in_notes'))}")

    service_statuses = [
        f"{service.get('service_name') or service.get('service_key')}={CRM_SERVICE_STATUS_LABELS.get(service.get('status'), service.get('status') or 'Unknown')}"
        for service in services
    ]
    if service_statuses:
        lines.append(f"Service-line status: {'; '.join(service_statuses)}")
    lines.append("Building information is stored as a fixed CRM snapshot. If the source may have changed, refresh it manually from the building library.")
    return "\n".join(lines)


def build_crm_case_ai_snapshot(detail: dict) -> dict:
    case_item = detail.get("case") or {}
    snapshot = case_item.get("building_snapshot") or {}
    insurance = snapshot.get("insurance") if isinstance(snapshot.get("insurance"), dict) else {}
    electricity = snapshot.get("electricity") if isinstance(snapshot.get("electricity"), dict) else {}
    internet = snapshot.get("internet") if isinstance(snapshot.get("internet"), dict) else {}
    move_in = snapshot.get("move_in") if isinstance(snapshot.get("move_in"), dict) else {}
    guests = detail.get("guests") or []
    services = detail.get("services") or []
    return {
        "building_name": case_item.get("building_name") or snapshot.get("building_name") or case_item.get("group_name"),
        "address": case_item.get("building_address") or snapshot.get("address") or "",
        "pending_count": 0,
        "insurance_required": insurance.get("required"),
        "insurance_coverage_amount": insurance.get("renters_minimum_coverage"),
        "electricity_required": electricity.get("required"),
        "electricity_provider": electricity.get("provider"),
        "internet_self_setup_required": internet.get("required"),
        "internet_provider": internet.get("providers"),
        "internet_notes": internet.get("notes"),
        "move_in_notes": move_in.get("move_in_notes"),
        "crm_group_name": case_item.get("group_name") or "",
        "crm_owner": case_item.get("owner_display_name") or "",
        "crm_lease_start_date": case_item.get("lease_start_date") or "",
        "crm_guests_count": len(guests),
        "crm_service_statuses": {
            service.get("service_key"): service.get("status")
            for service in services
            if service.get("service_key")
        },
    }


def crm_service_row_for_case(
    conn: sqlite3.Connection,
    case_id: str,
    service_id: str,
) -> Optional[dict]:
    return conn.execute(
        """
        SELECT *
        FROM crm_case_services
        WHERE case_id = ? AND (id = ? OR service_key = ?)
        """,
        (case_id, service_id, service_id),
    ).fetchone()


def valid_crm_owner(conn: sqlite3.Connection, owner_user_id: str) -> Optional[dict]:
    return conn.execute(
        """
        SELECT id, username, display_name, role
        FROM users
        WHERE id = ? AND is_active = 1 AND role IN ('super_admin', 'admin', 'employee')
        """,
        (owner_user_id,),
    ).fetchone()


def normalize_crm_owner(conn: sqlite3.Connection, actor: Actor, requested_owner_id: Optional[str]) -> str:
    if actor.role == "super_admin" and requested_owner_id:
        if not valid_crm_owner(conn, requested_owner_id):
            raise HTTPException(status_code=400, detail="The selected support owner does not exist or lacks CRM permission.")
        return requested_owner_id
    return actor.user_id


def normalize_crm_guest_payloads(guests: List[CrmCaseGuestPayload]) -> List[dict]:
    normalized: List[dict] = []
    for guest in guests:
        item = {
            "full_name": guest.full_name.strip(),
            "phone": guest.phone.strip(),
            "email": guest.email.strip(),
            "wechat": guest.wechat.strip(),
            "notes": guest.notes.strip(),
        }
        if not any(item.values()):
            continue
        if not item["full_name"]:
            raise HTTPException(status_code=400, detail="Each guest needs at least a name when a Case is created.")
        normalized.append(item)
    if not normalized:
        raise HTTPException(status_code=400, detail="Register at least one group guest when creating a Case.")
    return normalized


REVIEW_RESOLUTION_USE_NEW = "use_new"
REVIEW_RESOLUTION_USE_OLD = "use_old"
REVIEW_RESOLUTION_SKIP = "skip"
REVIEW_RESOLUTIONS = {
    REVIEW_RESOLUTION_USE_NEW,
    REVIEW_RESOLUTION_USE_OLD,
    REVIEW_RESOLUTION_SKIP,
}


def review_record_is_writable_field(row: dict) -> bool:
    field_name = str(row.get("field_name") or "")
    return bool(field_name) and not field_name.startswith("__") and field_name != "source_date"


def review_value_is_effective_for_field(
    conn: sqlite3.Connection,
    field_key: str,
    value: object,
) -> bool:
    definition = find_field_definition(conn, field_key)
    normalized = normalize_field_value(
        field_key,
        definition["field_type"] if definition else "text",
        value,
    )
    if normalized is None:
        return False
    if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
        return requirement_choice_to_int(normalized) is not None
    if field_key in NETWORK_PROVIDER_FIELD_MAP:
        return normalize_booleanish(normalized) is not None
    return normalize_unknown_value(normalized) is not None


def collect_effective_review_updates(
    conn: sqlite3.Connection,
    rows: List[dict],
    updates: Dict[str, Optional[str]],
    resolutions: Dict[str, Optional[str]],
) -> tuple[Dict[str, Optional[str]], List[str], Dict[str, dict], dict]:
    field_updates: Dict[str, Optional[str]] = {}
    row_by_field: Dict[str, dict] = {}
    ordered_fields: List[str] = []
    unresolved_conflicts: List[str] = []
    skipped_records = 0
    kept_old_records = 0
    empty_skipped_records = 0

    for row in rows:
        if not review_record_is_writable_field(row):
            continue
        record_id = row["record_id"]
        field_name = row["field_name"]
        resolution = resolutions.get(record_id)
        if resolution not in REVIEW_RESOLUTIONS:
            if row.get("conflict_with_long_term"):
                unresolved_conflicts.append(field_name)
                continue
            resolution = REVIEW_RESOLUTION_USE_NEW

        row_by_field.setdefault(field_name, row)
        if resolution == REVIEW_RESOLUTION_SKIP:
            skipped_records += 1
            continue
        if resolution == REVIEW_RESOLUTION_USE_OLD:
            kept_old_records += 1
            continue

        new_value = updates.get(record_id, row["new_value"])
        if not review_value_is_effective_for_field(conn, field_name, new_value):
            empty_skipped_records += 1
            continue
        field_updates[field_name] = new_value
        if field_name not in ordered_fields:
            ordered_fields.append(field_name)

    stats = {
        "unresolved_conflicts": sorted(set(unresolved_conflicts)),
        "skipped_records": skipped_records,
        "kept_old_records": kept_old_records,
        "empty_skipped_records": empty_skipped_records,
        "write_fields": len(ordered_fields),
    }
    return field_updates, ordered_fields, row_by_field, stats


def close_approved_review_without_effective_changes(
    conn: sqlite3.Connection,
    *,
    group_id: str,
    rows: List[dict],
    approval_stage: str,
    comment: str,
    actor: Actor,
    request: Request,
) -> dict:
    timestamp = now_iso()
    final_status = (
        "migrated_to_staging"
        if approval_stage == APPROVAL_STAGE_TO_STAGING
        else "migrated_to_master"
    )
    target_building_id = next((row["building_id"] for row in rows if row.get("building_id")), None)
    conn.execute(
        """
        UPDATE staging_update_requests
        SET building_id = COALESCE(?, building_id),
            review_status = ?, reviewer = ?, reviewed_at = ?, review_comment = ?, updated_at = ?
        WHERE submission_group_id = ?
        """,
        (
            target_building_id,
            final_status,
            actor.user_id,
            timestamp,
            comment or "No valid fields were written during approval; existing values were retained and blank values were skipped.",
            timestamp,
            group_id,
        ),
    )
    write_audit_log(
        conn,
        request,
        actor,
        action_type="review_group_approved_no_effective_changes",
        target_table="staging_update_requests",
        target_record_id=group_id,
        building_name=rows[0].get("building_name") or "",
        source=rows[0].get("source_file") or "",
        note=comment or "Approved with no effective field changes.",
    )
    return {
        "status": final_status,
        "building_id": target_building_id,
        "no_effective_changes": True,
    }


def apply_review_decision(
    conn: sqlite3.Connection,
    *,
    group_id: str,
    action: str,
    comment: str,
    updates: Dict[str, Optional[str]],
    resolutions: Dict[str, Optional[str]],
    actor: Actor,
    request: Request,
) -> dict:
    rows = fetch_group_records(conn, group_id)
    if not rows:
        raise HTTPException(status_code=404, detail="The review group was not found.")

    timestamp = now_iso()
    approval_stage = review_group_approval_stage(rows)
    if not can_actor_decide_review_stage(actor, approval_stage):
        if approval_stage == APPROVAL_STAGE_TO_MASTER:
            raise HTTPException(status_code=403, detail="Only a Super Admin can approve promotion into Master.")
        raise HTTPException(status_code=403, detail="Only an Admin or Super Admin can approve entry into Staging.")
    building_name = rows[0]["building_name"]
    building_address = next(
        (row["new_value"] for row in rows if row["field_name"] == "address" and row["new_value"]),
        "",
    )

    if action == "approved":
        if approval_stage == APPROVAL_STAGE_TO_STAGING:
            target_staging_key = review_group_target_staging_key(rows)
            target_staging_snapshot = (
                load_staging_building_snapshot(conn, target_staging_key) if target_staging_key else None
            )
            allow_existing_match = source_group_allows_existing_match(rows, updates)
            field_updates, _ordered_fields, _row_by_field, update_stats = collect_effective_review_updates(
                conn,
                rows,
                updates,
                resolutions,
            )
            if update_stats["unresolved_conflicts"]:
                raise HTTPException(
                    status_code=400,
                    detail="For every conflicting field, choose either Use New Value or Keep Existing Value: "
                    + "、".join(update_stats["unresolved_conflicts"]),
                )

            if not field_updates:
                return close_approved_review_without_effective_changes(
                    conn,
                    group_id=group_id,
                    rows=rows,
                    approval_stage=approval_stage,
                    comment=comment,
                    actor=actor,
                    request=request,
                )

            if target_staging_snapshot:
                field_updates["building_name"] = target_staging_snapshot.get("building_name")
                field_updates["address"] = target_staging_snapshot.get("address")
                allow_existing_match = True

            staging_snapshot = sync_source_values_into_staging_library(
                conn,
                building_name=normalize_unknown_value(field_updates.get("building_name")) or building_name,
                address=normalize_unknown_value(field_updates.get("address")) or building_address,
                values_by_field=field_updates,
                actor=actor,
                request=request,
                source_file=rows[0]["source_file"],
                audit_action="review_group_approved_to_staging",
                audit_note=comment or "The reviewed information was written to Staging.",
                allow_existing_match=allow_existing_match,
            )
            conn.execute(
                """
                UPDATE staging_update_requests
                SET review_status = 'migrated_to_staging', reviewer = ?, reviewed_at = ?, review_comment = ?, updated_at = ?
                WHERE submission_group_id = ?
                """,
                (actor.user_id, timestamp, comment, timestamp, group_id),
            )
            write_audit_log(
                conn,
                request,
                actor,
                action_type="review_group_approved_to_staging",
                target_table="staging_update_requests",
                target_record_id=group_id,
                building_name=building_name,
                source=rows[0]["source_file"],
                note=comment or "Approved and written into staging library.",
            )
            return {
                "staging_key": staging_snapshot["staging_key"] if staging_snapshot else None,
                "status": "migrated_to_staging",
            }

        has_direct_evidence = any(
            bool(row.get("source_document_id"))
            or bool(json_loads_safe(row.get("evidence_json"), []))
            for row in rows
        )
        if not has_direct_evidence and not comment.strip():
            raise HTTPException(
                status_code=400,
                detail="This promotion has no direct source evidence. Add an administrator verification note before approval.",
            )

        field_updates, ordered_fields, row_by_field, update_stats = collect_effective_review_updates(
            conn,
            rows,
            updates,
            resolutions,
        )
        if update_stats["unresolved_conflicts"]:
            raise HTTPException(
                status_code=400,
                detail="For every conflicting field, choose either Use New Value or Keep Existing Value: "
                + "、".join(update_stats["unresolved_conflicts"]),
            )
        if not field_updates:
            return close_approved_review_without_effective_changes(
                conn,
                group_id=group_id,
                rows=rows,
                approval_stage=approval_stage,
                comment=comment,
                actor=actor,
                request=request,
            )

        approved_building_name = normalize_unknown_value(field_updates.get("building_name")) or building_name
        approved_building_address = (
            normalize_unknown_value(field_updates.get("address"))
            if "address" in field_updates
            else normalize_unknown_value(building_address)
        ) or ""
        manual_master_building_id = review_group_manual_master_building_id(rows)
        if manual_master_building_id:
            master = load_master_building_snapshot(conn, manual_master_building_id)
            if not master:
                raise HTTPException(status_code=404, detail="The manually selected Master building does not exist.")
            approved_building_name = normalize_unknown_value(master.get("building_name")) or ""
            approved_building_address = normalize_unknown_value(master.get("address")) or ""
            field_updates["building_name"] = approved_building_name
            field_updates["address"] = approved_building_address
            for identity_field in ["building_name", "address"]:
                if identity_field not in ordered_fields:
                    ordered_fields.insert(0 if identity_field == "building_name" else min(1, len(ordered_fields)), identity_field)
        else:
            master = ensure_master_building(
                conn,
                building_name=approved_building_name,
                address=approved_building_address,
                actor=actor,
            )
        building_id = master["id"]
        previous_identity = (
            normalize_unknown_value(master.get("building_name")) or "",
            normalize_unknown_value(master.get("address")) or "",
        )

        if any(field_key in field_updates for field_key in DETAILED_INSURANCE_FIELD_KEYS):
            detail_payloads = {}
            current_snapshot = load_master_building_snapshot(conn, building_id) or {"extensions": {}}
            current_extensions = current_snapshot.get("extensions") or {}
            for field_key in DETAILED_INSURANCE_FIELD_KEYS:
                if field_key in field_updates:
                    detail_payloads[field_key] = {"value": field_updates.get(field_key), "confidence": 1.0}
                else:
                    current_value = current_extensions.get(field_key)
                    if current_value is not None:
                        detail_payloads[field_key] = {"value": current_value, "confidence": 1.0}
            derived = derive_legacy_insurance_fields_from_detailed(detail_payloads)
            for field_key, payload in derived.items():
                field_updates[field_key] = payload.get("value")
                if field_key not in ordered_fields:
                    ordered_fields.append(field_key)
            if "insurance_renters_required" in field_updates and not normalize_unknown_value(field_updates.get("insurance_renters_required")):
                field_updates["insurance_required"] = ""
                if "insurance_required" not in ordered_fields:
                    ordered_fields.append("insurance_required")
            if "insurance_renters_minimum_coverage" in field_updates and not normalize_unknown_value(field_updates.get("insurance_renters_minimum_coverage")):
                field_updates["insurance_coverage_amount"] = ""
                if "insurance_coverage_amount" not in ordered_fields:
                    ordered_fields.append("insurance_coverage_amount")

        touched_network_fields = False
        for field_key in ordered_fields:
            row = row_by_field.get(field_key, rows[0])
            new_value = field_updates.get(field_key)
            if field_key in NETWORK_PROVIDER_FIELD_MAP or field_key in {"internet_provider", "internet_notes"}:
                touched_network_fields = True
            old_value, stored_value = upsert_master_field(
                conn,
                building_id=building_id,
                field_key=field_key,
                new_value=new_value,
                actor=actor,
                source_type=row["source_type"],
                source_file=row["source_file"],
                source_date=None,
                info_cutoff_date=None,
            )
            write_audit_log(
                conn,
                request,
                actor,
                action_type="staging_approved_to_master",
                target_table="master_building_info"
                if field_key in CORE_MASTER_FIELD_KEYS
                else "master_building_field_values",
                target_record_id=building_id,
                building_name=building_name,
                field_name=field_key,
                old_value=old_value or "",
                new_value=stored_value or "",
                source=row["source_file"],
                note=comment,
            )
        if touched_network_fields:
            refresh_master_network_provider_text(
                conn,
                building_id=building_id,
                actor=actor,
                source_type=rows[0]["source_type"],
                source_file=rows[0]["source_file"],
            )
        refresh_master_completeness(
            conn,
            building_id,
            verification_note=comment or "Approved by a Super Admin.",
        )
        bump_master_version(conn, building_id, actor)
        sync_master_workbook_for_building(conn, building_id, previous_identity=previous_identity)
        refresh_after_master_excel_write(conn, actor=actor, request=request)
        target_staging_key = review_group_target_staging_key(rows)
        if target_staging_key:
            mark_staging_snapshot_mastered(conn, staging_key=target_staging_key, actor=actor, request=request)
        refreshed_master = (
            load_master_building_snapshot(conn, building_id)
            if manual_master_building_id
            else lookup_master_building(conn, approved_building_name, approved_building_address)
        )
        refreshed_building_id = refreshed_master["id"] if refreshed_master else building_id
        conn.execute(
            """
            UPDATE staging_update_requests
            SET building_id = ?, review_status = 'migrated_to_master', reviewer = ?, reviewed_at = ?, review_comment = ?, updated_at = ?
            WHERE submission_group_id = ?
            """,
            (refreshed_building_id, actor.user_id, timestamp, comment, timestamp, group_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="review_group_approved",
            target_table="staging_update_requests",
            target_record_id=group_id,
            building_name=building_name,
            note=comment or "Approved and migrated to master.",
        )
        return {"building_id": refreshed_building_id, "status": "migrated_to_master"}

    target_status = {
        "rejected": "rejected",
        "needs_more_info": "needs_more_info",
        "conflict": "conflict",
        "mark_missing": "needs_more_info",
    }[action]
    conn.execute(
        """
        UPDATE staging_update_requests
        SET review_status = ?, reviewer = ?, reviewed_at = ?, review_comment = ?, updated_at = ?
        WHERE submission_group_id = ?
        """,
        (target_status, actor.user_id, timestamp, comment, timestamp, group_id),
    )
    write_audit_log(
        conn,
        request,
        actor,
        action_type=f"review_group_{action}",
        target_table="staging_update_requests",
        target_record_id=group_id,
        building_name=building_name,
        note=comment,
    )
    return {"status": target_status}


MAILING_ADDRESS_CONTEXT_TOKENS = (
    "p.o. box",
    "po box",
    "p o box",
    "interested party",
    "interested party information",
    "additional interest",
    "certificate holder",
    "certificate holder information",
    "additional insured",
)


def text_has_mailing_address_context(text: str) -> bool:
    lowered = (text or "").lower()
    if any(token in lowered for token in MAILING_ADDRESS_CONTEXT_TOKENS):
        return True
    return bool(re.search(r"\bp\.?\s*o\.?\s*box\b|\bpo\s*box\b", lowered))


def is_weak_mailing_address_candidate(text: str, value: Optional[str] = None) -> bool:
    combined = " ".join(part for part in (text or "", value or "") if part).lower()
    if not combined:
        return False
    if text_has_mailing_address_context(combined):
        return True
    if "property management" in combined and re.search(r"\bbox\s+\d+\b", combined):
        return True
    return False


def is_suspicious_ocr_address_value(value: str) -> bool:
    cleaned = re.sub(r"\s+", " ", value or "").strip()
    # OCR can read a logo like "100 Example Avenue" as "1 1100 Example Avenue".
    if re.match(r"^\d{1,2}\s+\d{5,6}\s+", cleaned):
        return True
    street_number = re.match(r"^(\d{6,})\s+", cleaned)
    if street_number:
        return True
    return False


def is_non_street_address_value(text: str, value: str) -> bool:
    combined = " ".join(part for part in (text or "", value or "") if part).lower()
    value_lowered = (value or "").lower()
    if "floor terrace" in value_lowered or "rooftop terrace" in value_lowered:
        return True
    if "terrace" in value_lowered and any(token in combined for token in ("amenities include", "amenities package", "rooftop")):
        return True
    return False


def parse_address_from_text(text: str) -> Optional[str]:
    cleaned = text or ""
    patterns = [
        rf"\b(\d{{1,6}}\s+[A-Za-z0-9.'#/&\- ]{{2,80}}?\b{ADDRESS_SUFFIX_PATTERN}\b(?:\s+(?:Apt|Apartment|Unit|#)\s*[A-Za-z0-9\-]+)?)",
        rf"\b(\d{{1,6}}\s+[A-Za-z0-9.'#/&\- ]{{2,80}}?\b{ADDRESS_SUFFIX_PATTERN}\b)",
    ]
    for pattern in patterns:
        matches = list(re.finditer(pattern, cleaned, flags=re.IGNORECASE))
        for match in matches:
            value = match.group(1).strip().strip(" ,.;:!?-—")
            window = cleaned[max(0, match.start() - 120) : min(len(cleaned), match.end() + 120)]
            if is_suspicious_ocr_address_value(value):
                continue
            if is_non_street_address_value(window, value):
                continue
            if is_weak_mailing_address_candidate(window, value):
                continue
            normalized = normalize_unknown_value(value)
            if normalized:
                return normalized
    return None


def sanitize_address_payloads(parsed: Dict[str, dict]) -> Dict[str, dict]:
    cleaned = dict(parsed or {})
    payload = cleaned.get("address")
    if not isinstance(payload, dict):
        return cleaned
    value = normalize_unknown_value(payload.get("value"))
    evidence_text = normalize_unknown_value(payload.get("evidence")) or ""
    if not evidence_text:
        evidence_text = " ".join(
            normalize_unknown_value(item.get("quote")) or ""
            for item in (payload.get("evidence_items") or [])
            if isinstance(item, dict)
        ).strip()
    if value and is_weak_mailing_address_candidate(evidence_text, value):
        cleaned.pop("address", None)
        return cleaned
    if value:
        normalized = normalize_unknown_value(value)
        if normalized and parse_address_from_text(normalized):
            payload["value"] = normalized
            cleaned["address"] = payload
            return cleaned
    cleaned.pop("address", None)
    return cleaned


def parse_building_name_from_text(text: str) -> Optional[str]:
    cleaned = (text or "").replace("\r\n", "\n")
    patterns = [
        r"welcome to\s+([^\n—\-:,]{3,120})",
        r"\bnew home at\s+([^\n—\-:,!?.]{3,120})",
        r"\bhome at\s+([^\n—\-:,!?.]{3,120})",
        r"subject:\s.*?\bat\s+([^\n—\-:,]{3,120})",
        r"(?:^|[\n.;:]\s*|account\s+number\s+)([A-Za-z0-9][A-Za-z0-9.'& ]{2,80}?)\s+will\s+activate\s+water\b",
        r"interested\s+party\s+information\s*:?\s*([A-Za-z0-9][A-Za-z0-9.'& ]{2,80}?)(?:\s*[-–]\s*[A-Za-z]|\s+P\.?\s*O\.?\s*BOX\b)",
        r"\blist\s+(?:the\s+)?([A-Za-z0-9][A-Za-z0-9.'& ]{2,80}?)\s+as\s+(?:an?\s+)?interested\s+party\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, cleaned, flags=re.IGNORECASE)
        if not match:
            continue
        candidate = normalize_unknown_value(match.group(1))
        if candidate:
            candidate = re.sub(r"^.*\baccount\s+number\s+", "", candidate, flags=re.IGNORECASE)
            return candidate.strip(" .,:;!?")
    return parse_address_from_text(cleaned)


def normalize_money_text(value: object) -> Optional[str]:
    text = normalize_unknown_value(value)
    if not text:
        return None
    match = re.search(r"\$\s?\d[\d,]*(?:\.\d{2})?", text)
    if match:
        return match.group(0).replace(" ", "")
    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"\$?\d[\d,]*(?:\.\d{2})?", compact):
        return compact if compact.startswith("$") else f"${compact}"
    return None


def normalize_insurance_status(value: object) -> Optional[str]:
    text = (normalize_unknown_value(value) or "").strip().lower()
    if not text:
        return None
    if text in {"yes", "true", "required", "是"}:
        return "yes"
    if text in {"no", "false", "not_required", "否"}:
        return "no"
    if text in {"optional", "可选"}:
        return "optional"
    if text in {"manual_review", "needs_manual_review", "需人工确认"}:
        return "manual_review"
    return None


def insurance_status_label(value: object) -> str:
    status = normalize_insurance_status(value)
    if status:
        return INSURANCE_STATUS_LABELS[status]
    return "Not mentioned"


def insurance_field_value(snapshot: dict, field_key: str) -> Optional[str]:
    if field_key in CORE_MASTER_FIELD_KEYS:
        return normalize_unknown_value(snapshot.get(field_key))
    return normalize_unknown_value((snapshot.get("extensions") or {}).get(field_key))


def summary_field_keys() -> List[str]:
    return [
        "insurance_required",
        "insurance_coverage_amount",
        *sorted(DETAILED_INSURANCE_FIELD_KEYS),
        "electricity_required",
        "electricity_provider",
        "internet_self_setup_required",
        "internet_provider",
        "internet_notes",
        *NETWORK_PROVIDER_FIELD_MAP.keys(),
        *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
        "key_pickup_notes",
        "service_elevator_booking_notes",
        "move_in_notes",
    ]


def summary_snapshot_payload(snapshot: dict, field_keys: List[str]) -> dict:
    extensions = snapshot.get("extensions") or {}
    facts: Dict[str, Any] = {}
    for field_key in field_keys:
        if field_key in snapshot:
            facts[field_key] = snapshot.get(field_key)
        else:
            facts[field_key] = extensions.get(field_key)
    return {
        "id": snapshot.get("id"),
        "building_name": snapshot.get("building_name"),
        "address": snapshot.get("address"),
        "source_type": snapshot.get("source_type"),
        "source_file": snapshot.get("source_file"),
        "info_cutoff_date": snapshot.get("info_cutoff_date"),
        "updated_at": snapshot.get("updated_at"),
        "pending_count": snapshot.get("pending_count"),
        "facts": facts,
    }


def summary_snapshot_hash(snapshot: dict, field_keys: List[str]) -> str:
    payload = summary_snapshot_payload(snapshot, field_keys)
    encoded = json_dumps(payload).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


async def generate_cached_fact_explanation(
    *,
    source_mode: Literal["master", "staging"],
    record_id: str,
    question: str,
    snapshot: dict,
    fact_answer: str,
) -> tuple[Optional[str], str, str]:
    field_keys = summary_field_keys()
    if not ai_explanation_enabled() or not fact_answer.strip():
        return None, "disabled", ""
    snapshot_hash = summary_snapshot_hash(snapshot, field_keys)
    model_name = deepseek_model()
    normalized_question = legacy._normalize_text(question)
    question_hash = hashlib.sha256(normalized_question.encode("utf-8")).hexdigest()[:20]
    cache_key = (
        f"{source_mode}:{record_id}:{snapshot_hash}:{question_hash}:"
        f"{SUMMARY_CACHE_PROMPT_VERSION}:{model_name}"
    )
    with db_connection() as conn:
        cached = conn.execute(
            """
            SELECT ai_summary
            FROM ai_summary_cache
            WHERE cache_key = ? AND snapshot_hash = ? AND prompt_version = ?
            """,
            (cache_key, snapshot_hash, SUMMARY_CACHE_PROMPT_VERSION),
        ).fetchone()
        if cached and normalize_unknown_value(cached.get("ai_summary")):
            return cached["ai_summary"], "hit", cache_key

    ai_summary = await generate_fact_explanation(
        question=question,
        snapshot=snapshot,
        field_keys=field_keys,
        fact_answer=fact_answer,
        source_mode=source_mode,
    )
    if not ai_summary:
        return None, "miss", cache_key

    timestamp = now_iso()
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO ai_summary_cache(
              id, cache_key, source_mode, record_id, snapshot_hash, prompt_version,
              fact_summary, ai_summary, model, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(cache_key) DO UPDATE SET
              fact_summary = excluded.fact_summary,
              ai_summary = excluded.ai_summary,
              model = excluded.model,
              updated_at = excluded.updated_at
            """,
            (
                f"summary_cache_{uuid.uuid4().hex}",
                cache_key,
                source_mode,
                record_id,
                snapshot_hash,
                SUMMARY_CACHE_PROMPT_VERSION,
                fact_answer,
                ai_summary,
                model_name,
                timestamp,
                timestamp,
            ),
        )
    return ai_summary, "miss", cache_key


def pages_from_text(text: str) -> List[dict]:
    cleaned = normalize_unknown_value(text)
    if not cleaned:
        return []
    return [{"page": 1, "text": cleaned}]


def extract_pdf_pages(path: Path, max_pages: int = MAX_PDF_PARSE_PAGES) -> List[dict]:
    if not path.is_file():
        return []
    try:
        pdf = PdfReader(str(path))
    except Exception:
        return []
    pages: List[dict] = []
    page_count = min(max_pages, len(pdf.pages))
    for page_index in range(page_count):
        page = pdf.pages[page_index]
        page_text = legacy._normalize_extracted_pdf_text(page.extract_text() or "")
        if page_text:
            pages.append({"page": page_index + 1, "text": page_text})
    return pages


def render_pdf_pages_to_images(
    path: Path,
    *,
    max_pages: int = 8,
    output_dir: Optional[Path] = None,
) -> List[tuple[int, Path]]:
    if not path.is_file():
        return []

    if shutil.which("pdftoppm") is not None and output_dir is not None:
        prefix = output_dir / "page"
        try:
            subprocess.run(
                ["pdftoppm", "-png", "-f", "1", "-l", str(max_pages), str(path), str(prefix)],
                check=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=PDF_RENDER_TIMEOUT_SECONDS,
            )
        except Exception:
            pass
        else:
            rendered: List[tuple[int, Path]] = []
            for image_path in sorted(output_dir.glob("page-*.png"))[:max_pages]:
                match = re.search(r"page-(\d+)\.png$", image_path.name)
                page_number = int(match.group(1)) if match else len(rendered) + 1
                rendered.append((page_number, image_path))
            if rendered:
                return rendered

    if fitz is None or output_dir is None:
        return []

    rendered: List[tuple[int, Path]] = []
    try:
        document = fitz.open(str(path))
    except Exception:
        return []
    try:
        for page_index in range(min(max_pages, len(document))):
            page = document.load_page(page_index)
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            image_path = output_dir / f"page-{page_index + 1}.png"
            pix.save(str(image_path))
            rendered.append((page_index + 1, image_path))
    finally:
        document.close()
    return rendered


def extract_text_from_pdf_via_ocr_pages(path: Path) -> List[dict]:
    if shutil.which("tesseract") is None or not path.is_file():
        return []
    with tempfile.TemporaryDirectory(prefix="whitepaper_pdf_ocr_") as temp_dir:
        rendered_pages = render_pdf_pages_to_images(
            path,
            max_pages=MAX_PDF_PARSE_PAGES,
            output_dir=Path(temp_dir),
        )
        if not rendered_pages:
            return []
        pages: List[dict] = []
        for page_number, image_path in rendered_pages:
            text = extract_text_from_image_locally(image_path)
            if not text:
                continue
            pages.append({"page": page_number, "text": text})
        return pages


def join_page_texts(pages: List[dict]) -> str:
    return "\n\n".join(
        normalize_unknown_value(item.get("text")) or ""
        for item in pages
        if normalize_unknown_value(item.get("text"))
    ).strip()


def pdf_text_needs_visual_fallback(text: str, pages: Optional[List[dict]] = None) -> bool:
    cleaned = clean_welcome_letter_text(text)
    if len(cleaned.strip()) < 40:
        return True
    if not pages:
        return False
    normalized_page_texts = []
    page_lengths = []
    for item in pages:
        page_text = clean_welcome_letter_text(item.get("text") or "")
        if not page_text:
            continue
        normalized_page_texts.append(legacy._normalize_text(page_text))
        page_lengths.append(len(page_text))
    if not normalized_page_texts:
        return True
    unique_page_texts = {item for item in normalized_page_texts if item}
    max_page_length = max(page_lengths or [0])
    substantive_candidates = [
        candidate
        for candidate in build_sentence_candidates(pages)
        if len(candidate["quote"].split()) >= 6
    ]
    if len(unique_page_texts) <= max(1, len(normalized_page_texts) // 2) and max_page_length < 160:
        return True
    if len(substantive_candidates) < 3 and max_page_length < 220:
        return True
    if len(re.findall(r"[A-Za-z]{3,}", cleaned)) < 24:
        return True
    return False


def clean_welcome_letter_text(text: str) -> str:
    cleaned = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


TOC_SEGMENT_PATTERN = re.compile(
    r"^(contents?|welcome|contacts?|move-?in(?: & deliveries)?|rent payment|building services|amenities|general building information|insurancepets\d*)$",
    re.IGNORECASE,
)


def is_toc_like_segment(segment: str) -> bool:
    cleaned = (segment or "").strip(" .,:;!-").strip()
    if not cleaned:
        return True
    if TOC_SEGMENT_PATTERN.fullmatch(cleaned):
        return True
    if len(cleaned.split()) <= 4 and cleaned.lower() in {
        "contents",
        "welcome",
        "contacts",
        "rent payment",
        "general building information",
    }:
        return True
    return False


def build_sentence_candidates(pages: List[dict]) -> List[dict]:
    candidates: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for item in pages:
        page_number = int(item.get("page") or 1)
        text = clean_welcome_letter_text(item.get("text") or "")
        if not text:
            continue
        raw_lines = [line.strip() for line in re.split(r"\n{1,2}", text) if line.strip()]
        merged_segments: List[str] = []
        buffer = ""
        for raw_line in raw_lines:
            line = re.sub(r"\s+", " ", raw_line).strip()
            if not line:
                continue
            if not buffer:
                buffer = line
            else:
                buffer = f"{buffer} {line}".strip()
            if re.search(r"[.!?;:]$", line):
                merged_segments.append(buffer)
                buffer = ""
        if buffer:
            merged_segments.append(buffer)
        raw_segments: List[str] = []
        for segment in merged_segments:
            raw_segments.extend(re.split(r"(?<=[.!?;])\s+", segment))
        for segment in raw_segments:
            quote = re.sub(r"\s+", " ", segment).strip(" \t\n-•")
            if len(quote) < 4 or is_toc_like_segment(quote):
                continue
            normalized = legacy._normalize_text(quote)
            if not normalized:
                continue
            marker = (page_number, normalized)
            if marker in seen:
                continue
            seen.add(marker)
            candidates.append(
                {
                    "page": page_number,
                    "quote": quote,
                    "normalized": normalized,
                    "lowered": quote.lower(),
                }
            )
    return candidates


def locate_evidence_items(candidates: List[dict], evidence: object) -> List[dict]:
    text = normalize_unknown_value(evidence)
    if not text:
        return []
    normalized = legacy._normalize_text(text)
    if not normalized:
        return []
    matches: List[dict] = []
    for candidate in candidates:
        candidate_normalized = candidate["normalized"]
        if normalized in candidate_normalized or candidate_normalized in normalized:
            matches.append(candidate_evidence_item(candidate))
    if matches:
        return matches[:3]
    best_ratio = 0.0
    best_candidate = None
    for candidate in candidates:
        ratio = SequenceMatcher(None, normalized, candidate["normalized"]).ratio()
        if ratio > best_ratio:
            best_ratio = ratio
            best_candidate = candidate
    if best_candidate and best_ratio >= 0.62:
        return [candidate_evidence_item(best_candidate)]
    return []


def first_candidate(candidates: List[dict], predicate: Callable[[dict], bool]) -> Optional[dict]:
    for candidate in candidates:
        if predicate(candidate):
            return candidate
    return None


def all_candidates(candidates: List[dict], predicate: Callable[[dict], bool]) -> List[dict]:
    return [candidate for candidate in candidates if predicate(candidate)]


NON_IDENTITY_BUILDING_CONTEXT_TOKENS = (
    "interested party",
    "additional interest",
    "additional insured",
    "certificate holder",
    "utility information",
    "utilities/electric",
    "water, sewer",
    "activate water",
    "policy must include",
    "payable to",
    "cashier's checks",
    "bank checks",
    "security deposit",
    "p.o. box",
    "po box",
    "p o box",
)


def is_non_identity_building_context(text: str) -> bool:
    lowered = (text or "").lower()
    return any(token in lowered for token in NON_IDENTITY_BUILDING_CONTEXT_TOKENS)


def is_building_identity_from_service_or_party_context(text: str, building_name: str) -> bool:
    lowered = (text or "").lower()
    normalized_name = (building_name or "").lower().strip()
    if not lowered or not normalized_name or normalized_name not in lowered:
        return False
    return "will activate water" in lowered or "interested party information" in lowered or "as an interested party" in lowered


def address_payload_allows_existing_match(payload: Optional[dict]) -> bool:
    if not payload:
        return False
    address_value = normalize_unknown_value(payload.get("value")) or ""
    if not address_value:
        return False

    evidence_text = normalize_unknown_value(payload.get("evidence")) or ""
    if not evidence_text:
        evidence_text = " ".join(
            normalize_unknown_value(item.get("quote")) or ""
            for item in (payload.get("evidence_items") or [])
        ).strip()
    building_name = normalize_unknown_value(payload.get("value")) or ""
    if is_building_identity_from_service_or_party_context(evidence_text, building_name):
        return True
    if is_non_identity_building_context(evidence_text):
        return False

    lowered_value = address_value.lower()
    if is_weak_mailing_address_candidate(evidence_text, address_value) or "p.o. box" in lowered_value or "po box" in lowered_value or "p o box" in lowered_value:
        return False

    return bool(parse_address_from_text(address_value))


def building_payload_allows_existing_match(
    payload: Optional[dict],
    address_payload: Optional[dict] = None,
) -> bool:
    if address_payload_allows_existing_match(address_payload):
        return True
    if not payload:
        return False

    review_flags = set(payload.get("review_flags") or [])
    if "fallback_from_filename" in review_flags:
        return False

    evidence_text = normalize_unknown_value(payload.get("evidence")) or ""
    if not evidence_text:
        evidence_text = " ".join(
            normalize_unknown_value(item.get("quote")) or ""
            for item in (payload.get("evidence_items") or [])
        ).strip()
    if is_non_identity_building_context(evidence_text):
        return False

    confidence = float(payload.get("confidence") or 0)
    return confidence >= 0.9


def review_row_building_payload(row: Optional[dict]) -> Optional[dict]:
    if not row:
        return None
    evidence_items = json_loads_safe(row.get("evidence_json"), [])
    primary_quote = ""
    if evidence_items:
        primary_quote = normalize_unknown_value(evidence_items[0].get("quote")) or ""
    return {
        "value": normalize_unknown_value(row.get("new_value")),
        "confidence": float(row.get("ai_confidence") or 0),
        "evidence": primary_quote,
        "evidence_items": evidence_items,
        "review_flags": json_loads_safe(row.get("review_flags_json"), []),
    }


def review_row_field_payload(
    row: Optional[dict],
    override_value: Optional[str] = None,
) -> Optional[dict]:
    if not row:
        return None
    evidence_items = json_loads_safe(row.get("evidence_json"), [])
    primary_quote = ""
    if evidence_items:
        primary_quote = normalize_unknown_value(evidence_items[0].get("quote")) or ""
    return {
        "value": normalize_unknown_value(override_value) if override_value is not None else normalize_unknown_value(row.get("new_value")),
        "confidence": float(row.get("ai_confidence") or 0),
        "evidence": primary_quote,
        "evidence_items": evidence_items,
        "review_flags": json_loads_safe(row.get("review_flags_json"), []),
    }


def source_group_allows_existing_match(rows: List[dict], updates: Optional[Dict[str, Optional[str]]] = None) -> bool:
    building_row = next((row for row in rows if row.get("field_name") == "building_name"), None)
    address_row = next((row for row in rows if row.get("field_name") == "address"), None)
    building_payload = review_row_building_payload(building_row)
    address_override = None
    if updates and address_row and address_row.get("record_id") in updates:
        address_override = updates.get(address_row["record_id"])
    address_payload = review_row_field_payload(address_row, override_value=address_override)
    return building_payload_allows_existing_match(building_payload, address_payload)


def is_key_pickup_candidate(item: dict) -> bool:
    lowered = item.get("lowered") or ""
    quote = item.get("quote") or ""
    if not lowered:
        return False
    if "contents welcome contacts" in lowered:
        return False
    if any(
        token in lowered
        for token in (
            "lost or stolen",
            "visitor",
            "visitors",
            "visitors must be announced",
            "visitor must be announced",
            "contact list",
            "maintenance assistance",
        )
    ):
        return False

    key_tokens = ("key", "keys", "fob", "fobs", "mailbox")
    if not any(token in lowered for token in key_tokens):
        return False

    pickup_tokens = (
        "key pickup",
        "pick up",
        "pickup",
        "collect",
        "collected",
        "obtain",
        "receive",
        "receiving",
        "before receiving",
        "prior to receiving",
        "management office",
        "leasing office",
        "front desk",
        "concierge",
    )
    if not any(token in lowered for token in pickup_tokens):
        return False

    looks_like_contact_row = bool(re.search(r"[\w.+-]+@[\w.-]+\.\w+|\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b", quote))
    has_pickup_action = any(
        token in lowered
        for token in (
            "key pickup",
            "pick up",
            "pickup",
            "collect",
            "collected",
            "receive",
            "receiving",
            "before receiving",
            "prior to receiving",
        )
    )
    if looks_like_contact_row and not has_pickup_action:
        return False
    return True


MOVE_IN_LLM_FIELD_KEYS = {
    "key_pickup_notes",
    "service_elevator_booking_notes",
    "move_in_notes",
}

MOVE_IN_CANDIDATE_TOKENS = (
    "move-in",
    "move in",
    "move-ins",
    "deliveries",
    "delivery",
    "key",
    "keys",
    "fob",
    "fobs",
    "mailbox",
    "service elevator",
    "elevator",
    "rello",
    "time slot",
    "appointment",
    "pre-move checklist",
    "walkthrough",
    "lease start date",
    "waiver",
    "concierge",
    "front desk",
    "management office",
    "leasing office",
    "property administrator",
    "delay",
    "delays",
)


def is_service_elevator_candidate(item: dict) -> bool:
    lowered = item.get("lowered") or ""
    if not lowered or "contents welcome contacts" in lowered:
        return False
    return any(
        token in lowered
        for token in (
            "service elevator",
            "move-in appointment",
            "move in appointment",
            "reserve elevator",
            "reserving elevator",
            "schedule elevator",
            "booking window",
            "time slot",
            "rello",
            "pre-move checklist",
            "schedule",
        )
    )


def is_move_in_candidate(item: dict) -> bool:
    lowered = item.get("lowered") or ""
    if not lowered or "contents welcome contacts" in lowered:
        return False
    if any(token in lowered for token in ("maintenance assistance", "contact list")):
        return False
    return any(token in lowered for token in MOVE_IN_CANDIDATE_TOKENS)


def move_in_candidate_segments(candidates: List[dict], *, max_items: int = 36) -> List[dict]:
    selected: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for candidate in candidates:
        if not is_move_in_candidate(candidate):
            continue
        marker = (candidate["page"], candidate["normalized"])
        if marker in seen:
            continue
        seen.add(marker)
        selected.append(
            {
                "id": f"p{candidate['page']}_{len(selected) + 1}",
                "page": candidate["page"],
                "quote": candidate["quote"],
                "domain": candidate.get("domain") or "unknown",
            }
        )
        if len(selected) >= max_items:
            break
    return selected


def is_valid_key_pickup_quote(quote: str) -> bool:
    return is_key_pickup_candidate(
        {
            "quote": quote,
            "lowered": (quote or "").lower(),
            "normalized": legacy._normalize_text(quote or ""),
            "page": 1,
        }
    )


def is_valid_service_elevator_quote(quote: str) -> bool:
    return is_service_elevator_candidate(
        {
            "quote": quote,
            "lowered": (quote or "").lower(),
            "normalized": legacy._normalize_text(quote or ""),
            "page": 1,
        }
    )


def is_elevator_only_move_in_quote(quote: str) -> bool:
    lowered = (quote or "").lower()
    if not is_valid_service_elevator_quote(quote):
        return False
    elevator_markers = (
        "elevator",
        "service elevator",
        "reserve elevator",
        "reserving elevator",
        "time slot",
        "move-in appointment",
        "move in appointment",
        "rello",
    )
    if not any(token in lowered for token in elevator_markers):
        return False
    broader_move_in_markers = (
        "key",
        "keys",
        "fob",
        "lease start",
        "walkthrough",
        "self move waiver",
        "certificate of occupancy",
        "move-in checklist",
        "pre-move checklist",
        "before receiving your keys",
        "prior to receiving your keys",
    )
    return not any(token in lowered for token in broader_move_in_markers)


def service_elevator_submission_target_is_missing(text: str) -> bool:
    lowered = (text or "").lower()
    if not any(token in lowered for token in ("form", "send back", "submit", "reserve", "reserving", "schedule", "book")):
        return False
    target_markers = (
        "@",
        "http",
        "www.",
        "front desk",
        "concierge",
        "leasing office",
        "management office",
        "property management",
        "property manager",
        "portal",
        "rello",
        "buildinglink",
        "link:",
        "contact",
        "call",
        "phone",
    )
    return not any(marker in lowered for marker in target_markers) and not re.search(r"\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b", text or "")


def sanitize_move_in_payloads(
    payloads: Dict[str, dict],
    candidates: List[dict],
) -> Dict[str, dict]:
    sanitized: Dict[str, dict] = {}
    for field_key, payload in (payloads or {}).items():
        if field_key not in MOVE_IN_LLM_FIELD_KEYS or not isinstance(payload, dict):
            continue
        value = normalize_unknown_value(payload.get("value"))
        if value is None:
            continue
        value = clean_ai_field_value_text(str(value))
        evidence_items = payload_evidence_items(payload, candidates)
        evidence_items = filter_evidence_items_for_field(field_key, evidence_items)

        manual_reasons: List[str] = []
        confidence_cap: Optional[float] = None
        if field_key == "key_pickup_notes":
            valid_items = [item for item in evidence_items if is_valid_key_pickup_quote(item["quote"])]
            if len(valid_items) != len(evidence_items):
                manual_reasons.append("Content unrelated to key pickup, such as guest rules, lost-key rules, or contact lists, was filtered out.")
            evidence_items = valid_items
        elif field_key == "service_elevator_booking_notes":
            valid_items = [item for item in evidence_items if is_valid_service_elevator_quote(item["quote"])]
            if len(valid_items) != len(evidence_items):
                manual_reasons.append("Content unrelated to service-elevator or move-in booking was filtered out.")
            evidence_items = valid_items
            service_elevator_context = " ".join([value, *(item["quote"] for item in evidence_items)])
            if service_elevator_submission_target_is_missing(service_elevator_context):
                if not any(marker in value for marker in ("Submission recipient is not stated in the source", "提交对象原文未明确", "提交对象：原文未说明")):
                    value = f"{value} (Submission recipient is not stated in the source.)"
                manual_reasons.append("The source requires booking material to be submitted or returned but does not identify the recipient or channel.")
                confidence_cap = 0.7
        elif field_key == "move_in_notes":
            # move_in_notes is intentionally broad: it is the full move-in
            # checklist/remarks field, so key pickup and elevator evidence may
            # also appear here as long as original quotes are preserved.
            pass

        review_flags = [*(payload.get("review_flags") or []), *evidence_domain_review_flags(evidence_items)]
        confidence = float(payload.get("confidence") or 0)
        if not evidence_items:
            review_flags.append("no_direct_evidence")
            manual_reasons.append("No directly supporting source sentence was found.")
            confidence = min(confidence or 0.55, 0.55)
        else:
            confidence = min(max(confidence or 0.82, 0.72), 0.96)
        if confidence_cap is not None:
            confidence = min(confidence, confidence_cap)
        existing_reason = normalize_unknown_value(payload.get("manual_review_reason"))
        if existing_reason:
            manual_reasons.append(existing_reason)
        sanitized[field_key] = {
            "value": value,
            "confidence": confidence,
            "evidence": evidence_items[0]["quote"] if evidence_items else (payload.get("evidence") or ""),
            "evidence_items": evidence_items,
            "manual_review_reason": "; ".join(dict.fromkeys(reason for reason in manual_reasons if reason)),
            "review_flags": list(dict.fromkeys(review_flags)),
        }
    return sanitized


KEY_PICKUP_DETAIL_ORDER = (
    ("location", "Pickup Location"),
    ("contact_person_or_team", "Pickup Contact or Team"),
    ("channel", "Pickup or Booking Channel"),
    ("contact_info", "Contact Information"),
    ("prerequisites", "Prerequisites"),
    ("timing", "Pickup Timing or Conditions"),
    ("missing_info", "Missing Information"),
)

SERVICE_ELEVATOR_DETAIL_ORDER = (
    ("booking_method", "Booking Method"),
    ("submit_to", "Submission Recipient"),
    ("channel", "Booking Channel"),
    ("contact_info", "Contact Information"),
    ("advance_notice", "Advance Notice"),
    ("available_windows", "Available Booking Windows"),
    ("coi_required", "COI Required"),
    ("missing_info", "Missing Information"),
)

MOVE_IN_PROCESS_DETAIL_ORDER = (
    ("checklist", "Pre-Move-In Checklist"),
    ("deadlines", "Deadlines"),
    ("required_before_keys", "Requirements Before Key Pickup"),
    ("coordination_notes", "Move-In Coordination Notes"),
)

CONTACT_DETAIL_ORDER = (
    ("contact_type", "Contact Type"),
    ("person_or_team", "Person or Team"),
    ("phone", "Phone"),
    ("email", "Email"),
    ("channel", "Channel"),
    ("purpose", "Purpose"),
    ("hours", "Hours or Availability"),
    ("missing_info", "Missing Information"),
)

CONTACT_EMAIL_RE = re.compile(r"[\w.+-]+@[\w.-]+\.\w+")
CONTACT_PHONE_RE = re.compile(
    r"(?:\+?1[\s.\-]?)?(?:\(?\d{3}\)?[\s.\-]?\d{3}[\s.\-]?\d{4})"
)

CONTACT_TYPE_LABELS = {
    "leasing": "Leasing Office",
    "management": "Property Management",
    "front_desk": "Front Desk or Concierge",
    "maintenance": "Maintenance or Service",
    "general": "General Building Contact",
    "unknown": "Building Contact",
}

CONTACT_TYPE_FIELDS = {
    "management": "building_management_contact",
    "front_desk": "building_front_desk_contact",
    "maintenance": "building_maintenance_contact",
}


def contact_bits_from_text(text: str) -> tuple[List[str], List[str]]:
    emails = list(dict.fromkeys(CONTACT_EMAIL_RE.findall(text or "")))
    phones: List[str] = []
    for match in CONTACT_PHONE_RE.findall(text or ""):
        phone = re.sub(r"\s+", " ", match).strip(" .;,")
        if phone and phone not in phones:
            phones.append(phone)
    return emails, phones


def contact_details_from_quote(text: str) -> dict:
    emails, phones = contact_bits_from_text(text)
    contact_type = normalized_contact_type_from_context(text, {})
    return {
        "contact_type": contact_type,
        "person_or_team": CONTACT_TYPE_LABELS.get(contact_type, CONTACT_TYPE_LABELS["unknown"]),
        "email": " / ".join(emails) if emails else "Not stated in source",
        "phone": " / ".join(phones) if phones else "Not stated in source",
        "channel": "email" if emails else ("phone" if phones else "Not stated in source"),
        "purpose": "Purpose is not stated in the source" if contact_type in {"general", "unknown"} else CONTACT_TYPE_LABELS.get(contact_type, ""),
        "hours": "Not stated in source",
        "missing_info": "Purpose is not stated in the source" if contact_type in {"general", "unknown"} else "Not stated in source",
    }


def normalized_contact_type_from_context(context: str, details: Optional[dict] = None) -> str:
    detail_type = normalize_unknown_value((details or {}).get("contact_type"))
    if detail_type:
        lowered_detail = detail_type.lower()
        if any(token in lowered_detail for token in ("leasing", "lease", "rental")):
            return "leasing"
        if any(token in lowered_detail for token in ("management", "manager", "property")):
            return "management"
        if any(token in lowered_detail for token in ("front", "concierge", "doorman", "desk", "lobby")):
            return "front_desk"
        if any(token in lowered_detail for token in ("maintenance", "service request", "repair")):
            return "maintenance"
        if any(token in lowered_detail for token in ("general", "office")):
            return "general"
    lowered = (context or "").lower()
    if any(token in lowered for token in ("maintenance", "service request", "repairs", "repair request", "work order")):
        return "maintenance"
    if any(token in lowered for token in ("front desk", "concierge", "doorman", "lobby desk")):
        return "front_desk"
    if any(token in lowered for token in ("leasing office", "leasing team", "leasing agent", "rental office")):
        return "leasing"
    if any(token in lowered for token in ("management office", "property management", "property manager", "resident manager")):
        return "management"
    if any(token in lowered for token in ("office", "contact", "phone", "email")):
        return "general"
    return "unknown"


def contact_line_from_summary_item(item: dict) -> tuple[str, str, List[str], List[str]]:
    details = item.get("details") if isinstance(item.get("details"), dict) else {}
    context = " ".join(
        part
        for part in [
            normalize_unknown_value(item.get("title")) or "",
            normalize_unknown_value(item.get("value")) or "",
            normalize_unknown_value(item.get("quote")) or "",
            normalize_summary_detail_value(details),
        ]
        if part
    )
    contact_type = normalized_contact_type_from_context(context, details)
    emails, phones = contact_bits_from_text(context)
    parts: List[str] = []
    for detail_key, label in CONTACT_DETAIL_ORDER:
        if detail_key == "contact_type":
            continue
        text = normalize_unknown_value(details.get(detail_key))
        if text:
            parts.append(f"{label}: {text}")
    if not parts:
        value = normalize_unknown_value(item.get("value")) or normalize_unknown_value(item.get("quote")) or ""
        if value:
            parts.append(value)
    label = CONTACT_TYPE_LABELS.get(contact_type, CONTACT_TYPE_LABELS["unknown"])
    line = f"{label}: {'; '.join(dict.fromkeys(parts))}" if parts else label
    if contact_type in {"general", "unknown"} and (emails or phones):
        line = f"{line}; purpose is not stated in the source"
    return contact_type, line, emails, phones


def contact_items_from_business_summary_or_candidates(
    business_summary: Dict[str, List[dict]],
    candidates: List[dict],
) -> List[dict]:
    items = list((business_summary or {}).get("contacts") or [])
    if items:
        return items
    fallback_items: List[dict] = []
    for candidate in candidates or []:
        if normalize_unknown_value(candidate.get("domain")) != "contact":
            continue
        quote = normalize_unknown_value(candidate.get("quote")) or ""
        if not quote:
            continue
        fallback_items.append(
            {
                "id": f"contact_rule_{len(fallback_items) + 1}",
                "title": "Building Contact Information",
                "value": quote,
                "page": int(candidate.get("page") or 1),
                "quote": quote,
                "domain": "contact",
                "confidence": float(candidate.get("domain_confidence") or 0.72),
                "details": contact_details_from_quote(quote),
            }
        )
        if len(fallback_items) >= 8:
            break
    return fallback_items


def stable_contact_payloads_from_business_summary(
    business_summary: Dict[str, List[dict]],
    candidates: List[dict],
) -> Dict[str, dict]:
    contact_items = contact_items_from_business_summary_or_candidates(business_summary, candidates)
    if not contact_items:
        return {}
    field_lines: Dict[str, List[str]] = {field_key: [] for field_key in BUILDING_CONTACT_FIELD_KEYS}
    field_items: Dict[str, List[dict]] = {field_key: [] for field_key in BUILDING_CONTACT_FIELD_KEYS}
    for item in contact_items[:10]:
        contact_type, line, emails, phones = contact_line_from_summary_item(item)
        target_field = CONTACT_TYPE_FIELDS.get(contact_type)
        if target_field and line:
            field_lines[target_field].append(line)
            field_items[target_field].append(item)
    payloads: Dict[str, dict] = {}
    for field_key, lines in field_lines.items():
        unique_lines = [line for line in dict.fromkeys(normalize_unknown_value(line) or "" for line in lines) if line]
        if not unique_lines:
            continue
        evidence_items = business_summary_evidence_items(field_items.get(field_key) or contact_items, candidates)
        payloads[field_key] = {
            "value": "\n".join(unique_lines),
            "confidence": 0.92,
            "evidence": evidence_items[0]["quote"] if evidence_items else "",
            "evidence_items": evidence_items,
            "manual_review_reason": "",
            "review_flags": ["summary_stable_contact_mapping"],
        }
    return sanitize_contact_payloads(payloads, candidates)


def sanitize_contact_payloads(
    payloads: Dict[str, dict],
    candidates: List[dict],
) -> Dict[str, dict]:
    sanitized: Dict[str, dict] = {}
    for field_key, payload in (payloads or {}).items():
        if field_key not in BUILDING_CONTACT_FIELD_KEYS or not isinstance(payload, dict):
            continue
        value = normalize_unknown_value(payload.get("value"))
        if value is None:
            continue
        value = clean_ai_field_value_text(str(value), contact_context=True)
        evidence_items = payload_evidence_items(payload, candidates)
        evidence_items = filter_evidence_items_for_field(field_key, evidence_items)
        review_flags = [*(payload.get("review_flags") or []), *evidence_domain_review_flags(evidence_items)]
        manual_reasons: List[str] = []
        if not evidence_items:
            manual_reasons.append("No directly supporting building-contact quotation was found in the source.")
            review_flags.append("no_direct_evidence")
        existing_reason = normalize_unknown_value(payload.get("manual_review_reason"))
        if existing_reason:
            manual_reasons.append(existing_reason)
        try:
            confidence = float(payload.get("confidence") or 0.82)
        except (TypeError, ValueError):
            confidence = 0.82
        confidence = min(max(confidence, 0.72), 0.96) if evidence_items else min(confidence, 0.55)
        sanitized[field_key] = {
            "value": value,
            "confidence": confidence,
            "evidence": evidence_items[0]["quote"] if evidence_items else (payload.get("evidence") or ""),
            "evidence_items": evidence_items,
            "manual_review_reason": "; ".join(dict.fromkeys(reason for reason in manual_reasons if reason)),
            "review_flags": list(dict.fromkeys(review_flags)),
        }
    return sanitized


def summary_items_with_fallback(
    business_summary: Dict[str, List[dict]],
    section: str,
    *,
    legacy_predicate: Optional[Callable[[str], bool]] = None,
) -> List[dict]:
    items = list((business_summary or {}).get(section) or [])
    if items or not legacy_predicate:
        return items
    legacy_items: List[dict] = []
    for item in (business_summary or {}).get("move_in") or []:
        quote = normalize_unknown_value(item.get("quote")) or normalize_unknown_value(item.get("value")) or ""
        if quote and legacy_predicate(quote):
            legacy_items.append(item)
    return legacy_items


def summary_item_details_line(item: dict, detail_order: tuple[tuple[str, str], ...]) -> str:
    details = item.get("details") if isinstance(item.get("details"), dict) else {}
    parts: List[str] = []
    for detail_key, label in detail_order:
        text = normalize_unknown_value(details.get(detail_key))
        if text:
            parts.append(f"{label}: {text}")
    if parts:
        return "; ".join(dict.fromkeys(parts))
    value = normalize_unknown_value(item.get("value")) or normalize_unknown_value(item.get("quote")) or ""
    title = normalize_unknown_value(item.get("title")) or ""
    if title and value and title not in value:
        return f"{title}: {value}"
    return value


def text_has_contact_or_channel(text: str) -> bool:
    lowered = (text or "").lower()
    if re.search(r"\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b", text or "") or re.search(r"[\w.+-]+@[\w.-]+\.\w+", text or ""):
        return True
    return any(
        marker in lowered
        for marker in (
            "front desk",
            "concierge",
            "leasing office",
            "management office",
            "property management",
            "property manager",
            "portal",
            "rello",
            "buildinglink",
            "link:",
            "contact",
            "call",
            "phone",
            "email",
            "submit",
            "send",
        )
    )


def append_missing_move_in_details(field_key: str, line: str, context: str) -> str:
    additions: List[str] = []
    lowered_line = line.lower()
    if field_key == "key_pickup_notes":
        if not any(marker in lowered_line for marker in ("pickup location", "领取地点")) and not any(token in context.lower() for token in ("front desk", "concierge", "leasing office", "management office", "office")):
            additions.append("Pickup Location: Not stated in source")
        if not any(marker in lowered_line for marker in ("pickup contact", "领取对象")) and not any(token in context.lower() for token in ("front desk", "concierge", "leasing office", "property manager", "property management")):
            additions.append("Pickup Contact or Team: Not stated in source")
        if not any(marker in lowered_line for marker in ("contact information", "联系方式")) and not text_has_contact_or_channel(context):
            additions.append("Contact Information: Not stated in source")
    elif field_key == "service_elevator_booking_notes":
        if not any(marker in lowered_line for marker in ("booking method", "预约方式")) and not any(token in context.lower() for token in ("form", "portal", "rello", "buildinglink", "email", "call", "phone", "submit", "schedule", "reserve", "book")):
            additions.append("Booking Method: Not stated in source")
        if not any(marker in lowered_line for marker in ("submission recipient", "提交对象")) and service_elevator_submission_target_is_missing(context):
            additions.append("Submission Recipient: Not stated in source")
        if not any(marker in lowered_line for marker in ("contact information", "联系方式")) and not text_has_contact_or_channel(context):
            additions.append("Contact Information: Not stated in source")
    if additions:
        line = "; ".join([line, *additions] if line else additions)
    return line.strip("; ")


def business_summary_evidence_items(items: List[dict], candidates: List[dict]) -> List[dict]:
    evidence_items: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for item in items or []:
        quote = normalize_unknown_value(item.get("quote")) or ""
        if not quote:
            continue
        located = locate_evidence_items(candidates, quote)
        if not located:
            try:
                page = int(item.get("page") or 1)
            except (TypeError, ValueError):
                page = 1
            located = [{"page": page, "quote": quote, "domain": normalize_unknown_value(item.get("domain")) or ""}]
        for evidence in located:
            marker = (int(evidence.get("page") or 1), normalize_unknown_value(evidence.get("quote")) or "")
            if marker[1] and marker not in seen:
                seen.add(marker)
                evidence_items.append(evidence)
    return evidence_items


def stable_move_in_payload_from_summary(
    field_key: str,
    items: List[dict],
    candidates: List[dict],
    detail_order: tuple[tuple[str, str], ...],
    *,
    prefix: str = "",
) -> Optional[dict]:
    if not items:
        return None
    lines: List[str] = []
    for item in items[:6]:
        line = summary_item_details_line(item, detail_order)
        context = " ".join(
            part
            for part in [
                normalize_unknown_value(item.get("value")) or "",
                normalize_unknown_value(item.get("quote")) or "",
                normalize_summary_detail_value(item.get("details")),
            ]
            if part
        )
        line = append_missing_move_in_details(field_key, line, context)
        if line:
            lines.append(line)
    value = "\n".join(dict.fromkeys(lines))
    if prefix and value and not value.startswith(prefix):
        value = f"{prefix}{value}"
    if not value:
        return None
    evidence_items = business_summary_evidence_items(items, candidates)
    return {
        "value": value,
        "confidence": 0.94,
        "evidence": evidence_items[0]["quote"] if evidence_items else "",
        "evidence_items": evidence_items,
        "manual_review_reason": "",
        "review_flags": ["summary_stable_mapping"],
    }


def stable_move_in_payloads_from_business_summary(
    business_summary: Dict[str, List[dict]],
    candidates: List[dict],
) -> Dict[str, dict]:
    key_items = summary_items_with_fallback(
        business_summary,
        "key_pickup",
        legacy_predicate=is_valid_key_pickup_quote,
    )
    elevator_items = summary_items_with_fallback(
        business_summary,
        "service_elevator",
        legacy_predicate=is_valid_service_elevator_quote,
    )
    process_items = list((business_summary or {}).get("move_in_process") or [])
    payloads: Dict[str, dict] = {}
    key_payload = stable_move_in_payload_from_summary(
        "key_pickup_notes",
        key_items,
        candidates,
        KEY_PICKUP_DETAIL_ORDER,
    )
    if key_payload:
        payloads["key_pickup_notes"] = key_payload
    elevator_payload = stable_move_in_payload_from_summary(
        "service_elevator_booking_notes",
        elevator_items,
        candidates,
        SERVICE_ELEVATOR_DETAIL_ORDER,
    )
    if elevator_payload:
        payloads["service_elevator_booking_notes"] = elevator_payload
    move_in_items = [*process_items, *key_items, *elevator_items]
    move_in_payload = stable_move_in_payload_from_summary(
        "move_in_notes",
        move_in_items,
        candidates,
        MOVE_IN_PROCESS_DETAIL_ORDER,
        prefix="Complete move-in notes: ",
    )
    if move_in_payload:
        payloads["move_in_notes"] = move_in_payload
    return sanitize_move_in_payloads(payloads, candidates)


INSURANCE_LLM_FIELD_KEYS = {
    "insurance_required",
    "insurance_coverage_amount",
    "insurance_renters_required",
    "insurance_renters_minimum_coverage",
    "insurance_personal_property_required",
    "insurance_personal_property_minimum",
    "insurance_personal_liability_required",
    "insurance_personal_liability_per_occurrence",
    "insurance_personal_liability_aggregate",
    "insurance_coi_required",
    "insurance_coi_trigger",
    "insurance_interested_party_required",
    "insurance_additional_insured_required",
    "insurance_certificate_holder_required",
    "insurance_submission_method",
    "insurance_recipient",
    "insurance_alternative_program_or_penalty",
}

INSURANCE_REQUIREMENT_FIELD_KEYS = {
    "insurance_renters_required",
    "insurance_personal_property_required",
    "insurance_personal_liability_required",
    "insurance_coi_required",
    "insurance_interested_party_required",
    "insurance_additional_insured_required",
    "insurance_certificate_holder_required",
}

ELECTRICITY_LLM_FIELD_KEYS = {"electricity_required", "electricity_provider"}

EVIDENCE_DOMAINS = {
    "renters_insurance",
    "mover_coi",
    "electricity",
    "internet",
    "move_in",
    "key_pickup",
    "contact",
    "building_identity",
    "unknown",
}

BUSINESS_SUMMARY_KEYS = (
    "building_identity",
    "renters_insurance",
    "moving_coi",
    "insurance",
    "internet",
    "electricity",
    "move_in_process",
    "key_pickup",
    "service_elevator",
    "move_in",
    "payments",
    "contacts",
    "other_notes",
)

BUSINESS_SUMMARY_PRIMARY_KEYS = (
    "building_identity",
    "renters_insurance",
    "moving_coi",
    "internet",
    "electricity",
    "move_in_process",
    "key_pickup",
    "service_elevator",
    "payments",
    "contacts",
    "other_notes",
)

BUSINESS_SUMMARY_LEGACY_KEYS = {"insurance", "move_in"}

FIELD_ALLOWED_EVIDENCE_DOMAINS: Dict[str, set[str]] = {
    "insurance_required": {"renters_insurance"},
    "insurance_coverage_amount": {"renters_insurance"},
    "insurance_renters_required": {"renters_insurance"},
    "insurance_renters_minimum_coverage": {"renters_insurance"},
    "insurance_personal_property_required": {"renters_insurance"},
    "insurance_personal_property_minimum": {"renters_insurance"},
    "insurance_personal_liability_required": {"renters_insurance"},
    "insurance_personal_liability_per_occurrence": {"renters_insurance"},
    "insurance_personal_liability_aggregate": {"renters_insurance"},
    "insurance_interested_party_required": {"renters_insurance"},
    "insurance_additional_insured_required": {"renters_insurance"},
    "insurance_certificate_holder_required": {"renters_insurance"},
    "insurance_submission_method": {"renters_insurance"},
    "insurance_recipient": {"renters_insurance"},
    "insurance_alternative_program_or_penalty": {"renters_insurance"},
    "insurance_coi_required": {"mover_coi", "renters_insurance", "move_in"},
    "insurance_coi_trigger": {"mover_coi", "renters_insurance", "move_in"},
    "electricity_required": {"electricity"},
    "electricity_provider": {"electricity"},
    "key_pickup_notes": {"key_pickup"},
    "service_elevator_booking_notes": {"move_in", "mover_coi", "contact"},
    "move_in_notes": {"move_in", "mover_coi", "key_pickup", "contact"},
}
FIELD_ALLOWED_EVIDENCE_DOMAINS.update(
    {
        field_key: {"contact", "move_in", "key_pickup", "mover_coi", "renters_insurance"}
        for field_key in BUILDING_CONTACT_FIELD_KEYS
    }
)

FIELD_ALLOWED_SUMMARY_SECTIONS: Dict[str, set[str]] = {
    "insurance_required": {"renters_insurance", "insurance"},
    "insurance_coverage_amount": {"renters_insurance", "insurance"},
    "insurance_renters_required": {"renters_insurance", "insurance"},
    "insurance_renters_minimum_coverage": {"renters_insurance", "insurance"},
    "insurance_personal_property_required": {"renters_insurance", "insurance"},
    "insurance_personal_property_minimum": {"renters_insurance", "insurance"},
    "insurance_personal_liability_required": {"renters_insurance", "insurance"},
    "insurance_personal_liability_per_occurrence": {"renters_insurance", "insurance"},
    "insurance_personal_liability_aggregate": {"renters_insurance", "insurance"},
    "insurance_interested_party_required": {"renters_insurance", "insurance"},
    "insurance_additional_insured_required": {"renters_insurance", "insurance"},
    "insurance_certificate_holder_required": {"renters_insurance", "insurance"},
    "insurance_submission_method": {"renters_insurance", "insurance"},
    "insurance_recipient": {"renters_insurance", "insurance"},
    "insurance_alternative_program_or_penalty": {"renters_insurance", "insurance"},
    "insurance_coi_required": {"moving_coi", "insurance"},
    "insurance_coi_trigger": {"moving_coi", "insurance"},
    "key_pickup_notes": {"key_pickup", "move_in"},
    "service_elevator_booking_notes": {"service_elevator", "move_in"},
    "move_in_notes": {"move_in_process", "key_pickup", "service_elevator", "move_in"},
}
FIELD_ALLOWED_SUMMARY_SECTIONS.update(
    {field_key: {"contacts"} for field_key in BUILDING_CONTACT_FIELD_KEYS}
)

INSURANCE_CANDIDATE_TOKENS = (
    "insurance",
    "renter",
    "renters",
    "liability",
    "coverage",
    "coi",
    "certificate of insurance",
    "interested party",
    "additional interest",
    "additional insured",
    "certificate holder",
    "declaration page",
    "the guarantors",
    "guarantors",
    "proof of insurance",
    "policy",
    "protection program",
)

ELECTRICITY_CANDIDATE_TOKENS = (
    "electric",
    "electricity",
    "utility",
    "utilities/electric",
    "con edison",
    "coned",
    "con-ed",
    "pseg",
    "pse&g",
    "account number",
    "proof of electric",
    "proof of your account",
)


def domain_candidate_segments(
    candidates: List[dict],
    predicate: Callable[[dict], bool],
    *,
    max_items: int = 42,
) -> List[dict]:
    selected: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for candidate in candidates:
        if not predicate(candidate):
            continue
        marker = (candidate["page"], candidate["normalized"])
        if marker in seen:
            continue
        seen.add(marker)
        selected.append(
            {
                "id": f"p{candidate['page']}_{len(selected) + 1}",
                "page": candidate["page"],
                "quote": candidate["quote"],
                "domain": candidate.get("domain") or "unknown",
            }
        )
        if len(selected) >= max_items:
            break
    return selected


def evidence_classification_segments(candidates: List[dict], *, max_items: int = 90) -> List[dict]:
    segments: List[dict] = []
    for index, candidate in enumerate(candidates[:max_items], start=1):
        segments.append(
            {
                "id": f"c{index}",
                "page": candidate.get("page") or 1,
                "quote": candidate.get("quote") or "",
                "rule_domain": candidate.get("domain") or "unknown",
                "rule_reason": candidate.get("domain_reason") or "",
            }
        )
    return segments


def business_summary_segments(candidates: List[dict], *, max_items: int = 120) -> List[dict]:
    segments: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for candidate in candidates:
        domain = normalize_unknown_value(candidate.get("domain")) or "unknown"
        if domain == "unknown" and not any(
            token in (candidate.get("lowered") or "")
            for token in ("insurance", "electric", "internet", "wifi", "wi-fi", "address", "contact", "move")
        ):
            continue
        marker = (int(candidate.get("page") or 1), candidate.get("normalized") or "")
        if marker in seen:
            continue
        seen.add(marker)
        segments.append(
            {
                "id": f"s{len(segments) + 1}",
                "page": int(candidate.get("page") or 1),
                "quote": candidate.get("quote") or "",
                "domain": domain if domain in EVIDENCE_DOMAINS else "unknown",
            }
        )
        if len(segments) >= max_items:
            break
    return segments


def business_summary_key_for_domain(domain: str, quote: str = "") -> str:
    if domain == "building_identity":
        return "building_identity"
    if domain == "renters_insurance":
        return "renters_insurance"
    if domain == "mover_coi":
        return "moving_coi"
    if domain == "internet":
        return "internet"
    if domain == "electricity":
        return "electricity"
    if domain == "key_pickup":
        return "key_pickup"
    if domain == "move_in":
        if is_valid_service_elevator_quote(quote):
            return "service_elevator"
        return "move_in_process"
    if domain == "contact":
        return "contacts"
    return "other_notes"


def business_summary_title_for_domain(domain: str, quote: str) -> str:
    lowered = (quote or "").lower()
    if domain == "renters_insurance":
        return "Renters Insurance"
    if domain == "mover_coi":
        return "Moving COI"
    if domain == "internet":
        if "ssid" in lowered:
            return "Wi-Fi Name"
        if "password" in lowered:
            return "Wi-Fi Password"
        if any(token in lowered for token in ("included", "honest", "verizon", "xfinity", "spectrum", "astound")):
            return "Internet Information"
        return "Internet"
    if domain == "electricity":
        if any(token in lowered for token in ("included", "sub-meter", "submeter", "monthly billing")):
            return "Electricity Billing"
        return "Electricity"
    if domain == "building_identity":
        return "Building Information"
    if domain == "key_pickup":
        return "Key Pickup"
    if domain == "move_in":
        if is_valid_service_elevator_quote(quote):
            return "Service-Elevator Booking"
        return "Move-In Process"
    if domain == "contact":
        return "Contacts"
    return "Other Notes"


def sanitize_business_summary(summary: object, segment_by_id: Dict[str, dict]) -> Dict[str, List[dict]]:
    result: Dict[str, List[dict]] = {key: [] for key in BUSINESS_SUMMARY_KEYS}
    raw_summary = summary.get("business_summary") if isinstance(summary, dict) else summary
    if not isinstance(raw_summary, dict):
        return result
    seen: set[tuple[str, int, str]] = set()
    seen_ids: set[str] = set()
    for key in BUSINESS_SUMMARY_KEYS:
        items = raw_summary.get(key) or []
        if not isinstance(items, list):
            continue
        for item in items[:8]:
            if not isinstance(item, dict):
                continue
            evidence_id = normalize_unknown_value(item.get("evidence_id")) or ""
            source = segment_by_id.get(evidence_id)
            quote = normalize_unknown_value(item.get("quote")) or ""
            page = item.get("page")
            if source:
                quote = source.get("quote") or quote
                page = source.get("page") or page
            if not quote:
                continue
            try:
                page_number = int(page or 1)
            except (TypeError, ValueError):
                page_number = 1
            marker = (key, page_number, quote)
            if marker in seen:
                continue
            seen.add(marker)
            try:
                confidence = float(item.get("confidence") or 0.72)
            except (TypeError, ValueError):
                confidence = 0.72
            summary_id = normalize_unknown_value(item.get("id")) or evidence_id or f"{key}_{len(result[key]) + 1}"
            if summary_id in seen_ids:
                summary_id = f"{key}_{len(seen_ids) + 1}"
            seen_ids.add(summary_id)
            summary_payload = {
                "id": summary_id,
                "title": normalize_unknown_value(item.get("title")) or business_summary_title_for_domain(
                    source.get("domain") if source else "", quote
                ),
                "value": normalize_unknown_value(item.get("value")) or quote,
                "evidence_id": evidence_id,
                "page": page_number,
                "quote": quote,
                "domain": (source.get("domain") if source else normalize_unknown_value(item.get("domain"))) or "",
                "confidence": max(0.0, min(confidence, 1.0)),
            }
            details = normalize_summary_details(item.get("details"))
            if details:
                summary_payload["details"] = details
            result[key].append(summary_payload)
    return result


def build_rule_business_summary(candidates: List[dict]) -> Dict[str, List[dict]]:
    result: Dict[str, List[dict]] = {key: [] for key in BUSINESS_SUMMARY_KEYS}
    seen: set[tuple[str, int, str]] = set()
    for candidate in candidates:
        domain = normalize_unknown_value(candidate.get("domain")) or "unknown"
        quote = normalize_unknown_value(candidate.get("quote")) or ""
        key = business_summary_key_for_domain(domain, quote)
        if key == "other_notes" and domain == "unknown":
            continue
        if len(result[key]) >= 6:
            continue
        if not quote:
            continue
        page = int(candidate.get("page") or 1)
        marker = (key, page, quote)
        if marker in seen:
            continue
        seen.add(marker)
        item = {
            "title": business_summary_title_for_domain(domain, quote),
            "value": quote,
            "evidence_id": "",
            "page": page,
            "quote": quote,
            "domain": domain,
            "confidence": float(candidate.get("domain_confidence") or 0.72),
        }
        if key == "contacts":
            item["details"] = contact_details_from_quote(quote)
        result[key].append(item)
    return result


LLM_SUPPLEMENT_DOCUMENT_TYPES = {
    "supplement",
    "insurance_supplement",
    "internet_supplement",
    "electricity_supplement",
    "coi_supplement",
}


def compact_source_pages_for_llm(
    pages: List[dict],
    *,
    max_pages: int = 30,
    max_chars_per_page: int = 2500,
    max_total_chars: int = 42000,
) -> List[dict]:
    compacted: List[dict] = []
    total = 0
    for item in pages[:max_pages]:
        text = clean_welcome_letter_text(item.get("text") or "")
        if not text:
            continue
        remaining = max_total_chars - total
        if remaining <= 0:
            break
        clipped = text[: min(max_chars_per_page, remaining)]
        compacted.append({"page": int(item.get("page") or len(compacted) + 1), "text": clipped})
        total += len(clipped)
    return compacted


def normalize_document_classification(raw: object) -> dict:
    data = raw if isinstance(raw, dict) else {}
    document_type = normalize_unknown_value(data.get("document_type")) or "unknown"
    if document_type not in {
        "full_welcome_letter",
        "supplement",
        "insurance_supplement",
        "internet_supplement",
        "electricity_supplement",
        "coi_supplement",
        "unknown",
    }:
        document_type = "unknown"
    try:
        confidence = float(data.get("confidence") or 0)
    except (TypeError, ValueError):
        confidence = 0.0
    target = data.get("suggested_target_building") if isinstance(data.get("suggested_target_building"), dict) else {}
    affected = [
        str(item).strip()
        for item in (data.get("affected_domains") or [])
        if str(item).strip() in {"insurance", "internet", "electricity", "move_in", "payments", "contacts", "building_identity"}
    ]
    return {
        "document_type": document_type,
        "confidence": max(0.0, min(confidence, 1.0)),
        "suggested_target_building": {
            "building_name": normalize_unknown_value(target.get("building_name")) or "",
            "address": normalize_unknown_value(target.get("address")) or "",
            "reason": normalize_unknown_value(target.get("reason")) or "",
        },
        "affected_domains": affected,
        "reason": normalize_unknown_value(data.get("reason")) or "",
    }


def normalize_overall_summary(raw: object) -> str:
    if isinstance(raw, str):
        return normalize_unknown_value(raw) or ""
    if isinstance(raw, list):
        parts: List[str] = []
        for item in raw[:12]:
            if isinstance(item, dict):
                title = normalize_unknown_value(item.get("title")) or ""
                value = normalize_unknown_value(item.get("value")) or normalize_unknown_value(item.get("summary")) or ""
                text = f"{title}: {value}" if title and value else (value or title)
            else:
                text = normalize_unknown_value(item) or ""
            if text:
                parts.append(text)
        return "\n".join(dict.fromkeys(parts))
    if isinstance(raw, dict):
        parts = []
        for key, value in raw.items():
            text = normalize_unknown_value(value)
            if text:
                parts.append(f"{key}: {text}")
        return "\n".join(parts)
    return ""


def normalize_summary_detail_value(raw: object) -> str:
    if isinstance(raw, list):
        parts = [normalize_unknown_value(item) or "" for item in raw]
        return "; ".join(dict.fromkeys(part for part in parts if part))
    if isinstance(raw, dict):
        parts = []
        for key, value in raw.items():
            text = normalize_summary_detail_value(value)
            if text:
                parts.append(f"{key}: {text}")
        return "; ".join(parts)
    return normalize_unknown_value(raw) or ""


def normalize_summary_details(raw: object) -> dict:
    if not isinstance(raw, dict):
        return {}
    details: dict = {}
    for key, value in raw.items():
        normalized_key = normalize_unknown_value(key)
        normalized_value = normalize_summary_detail_value(value)
        if normalized_key and normalized_value:
            details[normalized_key] = normalized_value
    return details


def build_rule_overall_summary(summary: Dict[str, List[dict]]) -> str:
    labels = {
        "building_identity": "Building Information",
        "renters_insurance": "Renters Insurance",
        "moving_coi": "Moving or Delivery COI",
        "internet": "Internet",
        "electricity": "Electricity",
        "move_in_process": "Move-In Process",
        "key_pickup": "Key Pickup",
        "service_elevator": "Service-Elevator Booking",
        "payments": "Payments",
        "contacts": "Contacts",
        "other_notes": "Other Notes",
    }
    parts: List[str] = []
    for key in BUSINESS_SUMMARY_PRIMARY_KEYS:
        items = summary.get(key) or []
        if not items:
            continue
        first = items[0]
        value = normalize_unknown_value(first.get("value")) or normalize_unknown_value(first.get("quote")) or ""
        if value:
            parts.append(f"{labels.get(key, key)}：{crm_summary_excerpt(value, 120)}")
    return "\n".join(parts)


def sanitize_llm_business_summary(raw_summary: object, candidates: List[dict]) -> tuple[Dict[str, List[dict]], List[str]]:
    result: Dict[str, List[dict]] = {key: [] for key in BUSINESS_SUMMARY_KEYS}
    warnings: List[str] = []
    raw = raw_summary if isinstance(raw_summary, dict) else {}
    seen: set[tuple[str, int, str]] = set()
    seen_ids: set[str] = set()
    running_index = 1
    for key in BUSINESS_SUMMARY_KEYS:
        items = raw.get(key) or []
        if not isinstance(items, list):
            continue
        for item in items[:10]:
            if not isinstance(item, dict):
                continue
            quote = normalize_unknown_value(item.get("quote"))
            if not quote:
                warnings.append(f"{key}: summary item skipped because quote is missing")
                continue
            evidence_items = locate_evidence_items(candidates, quote)
            if evidence_items:
                page = int(evidence_items[0].get("page") or item.get("page") or 1)
                quote = evidence_items[0].get("quote") or quote
                domain = normalize_unknown_value(evidence_items[0].get("domain")) or normalize_unknown_value(item.get("domain")) or ""
            else:
                try:
                    page = int(item.get("page") or 1)
                except (TypeError, ValueError):
                    page = 1
                domain = normalize_unknown_value(item.get("domain")) or ""
                warnings.append(f"{key}: quote was not matched exactly to extracted candidates")
            marker = (key, page, quote)
            if marker in seen:
                continue
            seen.add(marker)
            try:
                confidence = float(item.get("confidence") or 0.72)
            except (TypeError, ValueError):
                confidence = 0.72
            summary_id = normalize_unknown_value(item.get("id")) or f"bs{running_index}"
            running_index += 1
            if summary_id in seen_ids:
                summary_id = f"{key}_{running_index}"
                running_index += 1
            seen_ids.add(summary_id)
            summary_payload = {
                "id": summary_id,
                "title": normalize_unknown_value(item.get("title")) or business_summary_title_for_domain(domain, quote),
                "value": normalize_unknown_value(item.get("value")) or quote,
                "evidence_id": normalize_unknown_value(item.get("evidence_id")) or summary_id,
                "page": page,
                "quote": quote,
                "domain": domain,
                "confidence": max(0.0, min(confidence, 1.0)),
            }
            details = normalize_summary_details(item.get("details"))
            if details:
                summary_payload["details"] = details
            result[key].append(summary_payload)
    return result, warnings


def flatten_business_summary_items(summary: Dict[str, List[dict]]) -> List[dict]:
    flattened: List[dict] = []
    for section, items in (summary or {}).items():
        for index, item in enumerate(items or [], start=1):
            summary_id = normalize_unknown_value(item.get("id")) or normalize_unknown_value(item.get("evidence_id")) or f"{section}_{index}"
            flattened.append(
                {
                    "id": summary_id,
                    "section": section,
                    "title": item.get("title") or "",
                    "value": item.get("value") or "",
                    "page": item.get("page") or 1,
                    "quote": item.get("quote") or "",
                    "domain": item.get("domain") or "",
                    "confidence": item.get("confidence") or 0,
                    "details": item.get("details") or {},
                }
            )
    return flattened


def normalize_workflow_hints(raw: object) -> dict:
    data = raw if isinstance(raw, dict) else {}
    result = {}
    for key in ("insurance", "internet", "electricity", "move_in"):
        value = data.get(key)
        result[key] = value if isinstance(value, dict) else {}
    return result


def llm_field_options_for_mapping(conn: sqlite3.Connection) -> List[dict]:
    definitions = {item["field_key"]: item for item in field_catalog(conn)}
    options: List[dict] = []
    for field_key in sorted(AI_WRITABLE_FIELD_KEYS):
        definition = definitions.get(field_key) or {}
        options.append(
            {
                "field_key": field_key,
                "display_name": definition.get("display_name") or field_key,
                "field_type": definition.get("field_type") or "text",
                "description": definition.get("description") or "",
                "group_key": definition.get("group_key") or "",
            }
        )
    return options


def sanitize_llm_field_mapping_values(
    raw_values: object,
    candidates: List[dict],
    summary_section_by_id: Optional[Dict[str, str]] = None,
) -> tuple[Dict[str, dict], List[str]]:
    values = raw_values if isinstance(raw_values, dict) else {}
    sanitized: Dict[str, dict] = {}
    warnings: List[str] = []
    summary_section_by_id = summary_section_by_id or {}
    for field_key, payload in values.items():
        if field_key not in AI_WRITABLE_FIELD_KEYS or not isinstance(payload, dict):
            continue
        value = normalize_unknown_value(payload.get("value"))
        if value is None:
            continue
        evidence_items = payload_evidence_items(payload, candidates)
        if not evidence_items and field_key not in {"source_type", "source_file"}:
            warnings.append(f"{field_key}: skipped because no original quote evidence was provided")
            continue
        evidence_items = filter_evidence_items_for_field(field_key, evidence_items)
        if not evidence_items and field_key not in {"source_type", "source_file"}:
            warnings.append(f"{field_key}: skipped because evidence domain is not allowed for this field")
            continue
        try:
            confidence = float(payload.get("confidence") or 0.78)
        except (TypeError, ValueError):
            confidence = 0.78
        summary_item_id = normalize_unknown_value(payload.get("summary_item_id")) or ""
        summary_section = summary_section_by_id.get(summary_item_id, "")
        allowed_sections = FIELD_ALLOWED_SUMMARY_SECTIONS.get(field_key)
        if summary_section and allowed_sections and summary_section not in allowed_sections:
            warnings.append(
                f"{field_key}: skipped because summary section {summary_section} is not allowed for this field"
            )
            continue
        if field_key in BUILDING_CONTACT_FIELD_KEYS:
            value = clean_ai_field_value_text(str(value), contact_context=True)
        review_flags = ["llm_primary_mapping", *evidence_domain_review_flags(evidence_items)]
        if summary_item_id:
            review_flags.append(f"summary_item:{summary_item_id}")
        if summary_section:
            review_flags.append(f"summary_section:{summary_section}")
        sanitized[field_key] = {
            "value": value,
            "confidence": max(0.0, min(confidence, 0.97)),
            "evidence": evidence_items[0]["quote"] if evidence_items else "",
            "evidence_items": evidence_items,
            "manual_review_reason": normalize_unknown_value(payload.get("manual_review_reason")) or "",
            "review_flags": list(dict.fromkeys(review_flags)),
        }
    return sanitized, warnings


def merge_with_primary_preferred(primary: Dict[str, dict], *fallbacks: Dict[str, dict]) -> Dict[str, dict]:
    merged: Dict[str, dict] = {}
    for payload_set in fallbacks:
        for field_key, payload in (payload_set or {}).items():
            if field_key not in merged:
                merged[field_key] = payload
    for field_key, payload in (primary or {}).items():
        merged[field_key] = payload
    return merged


def is_insurance_candidate(item: dict) -> bool:
    lowered = item.get("lowered") or ""
    if not lowered or "contents welcome contacts" in lowered:
        return False
    return any(token in lowered for token in INSURANCE_CANDIDATE_TOKENS)


def is_electricity_candidate(item: dict) -> bool:
    lowered = item.get("lowered") or ""
    if not lowered or "contents welcome contacts" in lowered:
        return False
    return any(token in lowered for token in ELECTRICITY_CANDIDATE_TOKENS)


def is_contact_candidate(item: dict) -> bool:
    quote = item.get("quote") or ""
    lowered = item.get("lowered") or quote.lower()
    if not lowered or "contents welcome contacts" in lowered:
        return False
    has_contact = bool(CONTACT_EMAIL_RE.search(quote) or CONTACT_PHONE_RE.search(quote))
    return has_contact or any(
        token in lowered
        for token in (
            "front desk",
            "concierge",
            "doorman",
            "leasing office",
            "management office",
            "property management",
            "property manager",
            "maintenance",
            "service request",
            "resident services",
            "contact",
        )
    )


def is_building_identity_candidate(item: dict) -> bool:
    quote = item.get("quote") or ""
    lowered = item.get("lowered") or quote.lower()
    if not lowered:
        return False
    if is_non_identity_building_context(quote):
        return False
    return bool(parse_address_from_text(quote)) or any(
        token in lowered for token in ("welcome to", "new home at", "home at", "subject:")
    )


def classify_candidate_domain_by_rules(candidate: dict, mover_coi_pages: Optional[set[int]] = None) -> tuple[str, float, str]:
    quote = candidate.get("quote") or ""
    lowered = candidate.get("lowered") or quote.lower()
    if not lowered:
        return "unknown", 0.5, "empty candidate"
    if candidate_has_mover_coi_context(candidate, mover_coi_pages):
        return "mover_coi", 0.98, "moving COI or page-level moving certificate context"
    if quote_has_renters_insurance_context(quote):
        return "renters_insurance", 0.98, "explicit renters / tenant insurance context"
    if "interested party information" in lowered or re.search(
        r"\blist\s+.+?\s+as\s+an?\s+interested\s+party\b",
        lowered,
    ):
        return "renters_insurance", 0.9, "renters insurance interested party block"
    if "proof of insurance" in lowered and not any(token in lowered for token in ("mover", "moving company", "move in", "move-in", "move out", "move-out")):
        return "renters_insurance", 0.86, "proof of insurance without moving COI context"
    if "insurance" in lowered and "leasing" in lowered and re.search(r"[\w.+-]+@[\w.-]+\.\w+", quote):
        return "renters_insurance", 0.82, "insurance submission to leasing contact"
    if "personal liability" in lowered and any(token in lowered for token in ("renter", "tenant", "policy", "lease")):
        return "renters_insurance", 0.86, "personal liability tied to renter / tenant policy context"
    if is_electricity_candidate(candidate):
        return "electricity", 0.9, "electricity / utility account signal"
    if network_provider_parts_from_candidate(candidate) or any(
        token in lowered for token in ("internet", "wifi", "wi-fi", "verizon", "xfinity", "astound", "spectrum", "honest")
    ):
        return "internet", 0.88, "internet / provider signal"
    if is_key_pickup_candidate(candidate):
        return "key_pickup", 0.9, "key / fob pickup signal"
    if is_move_in_candidate(candidate) or is_service_elevator_candidate(candidate):
        return "move_in", 0.86, "move-in / service elevator signal"
    if is_building_identity_candidate(candidate):
        return "building_identity", 0.84, "building identity / address signal"
    if is_contact_candidate(candidate):
        return "contact", 0.78, "contact-only signal"
    if is_insurance_candidate(candidate):
        return "unknown", 0.56, "insurance signal without clear renters vs mover COI boundary"
    return "unknown", 0.5, "no domain signal"


def candidate_evidence_item(candidate: dict) -> dict:
    item = {"page": candidate.get("page") or 1, "quote": candidate.get("quote") or ""}
    domain = normalize_unknown_value(candidate.get("domain"))
    if domain:
        item["domain"] = domain
    domain_confidence = candidate.get("domain_confidence")
    if domain_confidence is not None:
        item["domain_confidence"] = domain_confidence
    domain_reason = normalize_unknown_value(candidate.get("domain_reason"))
    if domain_reason:
        item["domain_reason"] = domain_reason
    return item


def classify_evidence_candidates_by_rules(candidates: List[dict]) -> List[dict]:
    mover_coi_pages = mover_coi_pages_from_candidates(candidates)
    classified: List[dict] = []
    for candidate in candidates or []:
        domain, confidence, reason = classify_candidate_domain_by_rules(candidate, mover_coi_pages)
        next_candidate = dict(candidate)
        next_candidate["domain"] = domain if domain in EVIDENCE_DOMAINS else "unknown"
        next_candidate["domain_confidence"] = confidence
        next_candidate["domain_reason"] = reason
        classified.append(next_candidate)
    return classified


def evidence_item_domains(evidence_items: List[dict]) -> set[str]:
    return {
        domain
        for item in evidence_items or []
        for domain in [normalize_unknown_value(item.get("domain"))]
        if domain in EVIDENCE_DOMAINS and domain != "unknown"
    }


def evidence_domain_review_flags(evidence_items: List[dict]) -> List[str]:
    return [f"evidence_domain:{domain}" for domain in sorted(evidence_item_domains(evidence_items))]


def filter_evidence_items_for_field(field_key: str, evidence_items: List[dict]) -> List[dict]:
    allowed_domains = FIELD_ALLOWED_EVIDENCE_DOMAINS.get(field_key)
    if not allowed_domains:
        return evidence_items
    filtered: List[dict] = []
    for item in evidence_items or []:
        domain = normalize_unknown_value(item.get("domain"))
        if not domain or domain in allowed_domains:
            filtered.append(item)
    return filtered


def candidate_has_allowed_domain(item: dict, allowed_domains: set[str]) -> bool:
    domain = normalize_unknown_value(item.get("domain"))
    return not domain or domain in allowed_domains


def payload_evidence_items(payload: dict, candidates: List[dict]) -> List[dict]:
    def safe_page(value: object) -> int:
        try:
            return int(value or 1)
        except (TypeError, ValueError):
            return 1

    evidence_items: List[dict] = []
    for item in payload.get("evidence_items") or []:
        if not isinstance(item, dict):
            continue
        quote = normalize_unknown_value(item.get("quote"))
        if not quote:
            continue
        located_items = locate_evidence_items(candidates, quote)
        if located_items:
            evidence_items.extend(located_items)
        else:
            # Vision/OCR can produce field evidence even when PDF text extraction only
            # found a table of contents. Keep the model-provided quote for reviewers.
            fallback_item = {"page": safe_page(item.get("page")), "quote": quote}
            domain = normalize_unknown_value(item.get("domain"))
            if domain in EVIDENCE_DOMAINS:
                fallback_item["domain"] = domain
            evidence_items.append(fallback_item)
    if not evidence_items:
        evidence_items = locate_evidence_items(candidates, payload.get("evidence") or payload.get("value"))
    if not evidence_items:
        quote = normalize_unknown_value(payload.get("evidence"))
        if quote:
            fallback_item = {"page": safe_page(payload.get("page")), "quote": quote}
            domain = normalize_unknown_value(payload.get("domain"))
            if domain in EVIDENCE_DOMAINS:
                fallback_item["domain"] = domain
            evidence_items.append(fallback_item)

    deduped: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for item in evidence_items:
        quote = normalize_unknown_value(item.get("quote"))
        if not quote:
            continue
        marker = (safe_page(item.get("page")), quote)
        if marker in seen:
            continue
        seen.add(marker)
        deduped_item = {"page": marker[0], "quote": marker[1]}
        domain = normalize_unknown_value(item.get("domain"))
        if domain in EVIDENCE_DOMAINS:
            deduped_item["domain"] = domain
        domain_confidence = item.get("domain_confidence")
        if domain_confidence is not None:
            deduped_item["domain_confidence"] = domain_confidence
        domain_reason = normalize_unknown_value(item.get("domain_reason"))
        if domain_reason:
            deduped_item["domain_reason"] = domain_reason
        deduped.append(deduped_item)
    return deduped[:4]


def evidence_text_from_items(items: List[dict]) -> str:
    return "\n".join(normalize_unknown_value(item.get("quote")) or "" for item in items).strip()


def quote_has_personal_liability_amount_context(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        token in lowered
        for token in (
            "personal liability",
            "liability coverage",
            "bodily injury",
            "property damage",
            "per occurrence",
            "aggregate",
        )
    )


def quote_has_property_coverage_context(text: str) -> bool:
    lowered = (text or "").lower()
    return any(token in lowered for token in ("personal property", "property coverage", "personal belongings", "belongings coverage"))


def quote_has_renters_insurance_context(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        token in lowered
        for token in (
            "renters insurance",
            "renter's insurance",
            "renter’s insurance",
            "renter insurance",
            "tenant insurance",
            "tenant's insurance",
            "tenant’s insurance",
            "proof of renters insurance",
        )
    )


def clean_insurance_recipient_text(text: str) -> str:
    cleaned = re.sub(r"\s+", " ", normalize_unknown_value(text) or "").strip(" .;:")
    if not cleaned:
        return ""
    lowered = cleaned.lower()
    boundary_tokens = (
        "if your insurance agent",
        "pet requirements",
        "utility information",
        "must have renters",
        "your policy must include",
        "all leaseholders listed",
        "obtain a minimum",
        "before receiving",
        "schedule a time slot",
    )
    for token in boundary_tokens:
        position = lowered.find(token)
        if position > 0:
            cleaned = cleaned[:position].strip(" .;:")
            lowered = cleaned.lower()
    cleaned = re.sub(r"^(information|info)\s*:\s*", "", cleaned, flags=re.IGNORECASE).strip(" .;:")
    return cleaned


def extract_interested_party_recipient_value(text: str) -> str:
    cleaned = normalize_unknown_value(text) or ""
    if not cleaned:
        return ""
    block_match = re.search(
        r"interested\s+party\s+information\s*:?\s*(.+)",
        cleaned,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if block_match:
        value = clean_insurance_recipient_text(block_match.group(1))
        if value:
            return value
    list_match = re.search(
        r"\blist\s+(.{2,220}?)\s+as\s+(?:an?\s+)?interested\s+party\b",
        cleaned,
        flags=re.IGNORECASE | re.DOTALL,
    )
    if list_match:
        return clean_insurance_recipient_text(list_match.group(1))
    return ""


def expanded_interested_party_quote(candidates: List[dict], candidate: dict) -> str:
    quote = normalize_unknown_value(candidate.get("quote")) or ""
    if not quote:
        return ""
    lowered = quote.lower()
    if "interested party information" not in lowered:
        return quote
    page = int(candidate.get("page") or 1)
    try:
        start_index = candidates.index(candidate)
    except ValueError:
        return quote
    parts = [quote]
    for neighbor in candidates[start_index + 1 : start_index + 5]:
        if int(neighbor.get("page") or 1) != page:
            break
        neighbor_quote = normalize_unknown_value(neighbor.get("quote")) or ""
        neighbor_lowered = neighbor_quote.lower()
        if not neighbor_quote:
            continue
        if any(
            token in neighbor_lowered
            for token in (
                "pet requirements",
                "utility information",
                "service elevator",
                "move in",
                "move-in",
                "parking",
            )
        ):
            break
        parts.append(neighbor_quote)
        if re.search(r"\b\d{5}(?:-\d{4})?\b", neighbor_quote):
            break
    return " ".join(parts)


def quote_has_mover_coi_context(text: str) -> bool:
    lowered = (text or "").lower()
    has_certificate = any(
        token in lowered
        for token in (
            "coi",
            "certificate of insurance",
            "certificate of liability insurance",
            "certificate of \ninsurance",
        )
    )
    has_move_context = any(
        token in lowered
        for token in (
            "mover",
            "movers",
            "moving company",
            "move-in",
            "move in",
            "move-out",
            "move out",
            "schedule a move",
            "schedule move",
            "reserving elevator",
            "reserve elevator",
            "date of move",
            "certificate of insurance for moving",
        )
    )
    return has_certificate and has_move_context and not quote_has_renters_insurance_context(text)


def mover_coi_pages_from_candidates(candidates: List[dict]) -> set[int]:
    pages: set[int] = set()
    for item in candidates or []:
        quote = item.get("quote") or ""
        lowered = item.get("lowered") or quote.lower()
        if "certificate of insurance for moving" in lowered:
            try:
                pages.add(int(item.get("page") or 1))
            except (TypeError, ValueError):
                pages.add(1)
    return pages


def candidate_has_mover_coi_context(item: dict, mover_coi_pages: Optional[set[int]] = None) -> bool:
    quote = item.get("quote") or ""
    if quote_has_mover_coi_context(quote):
        return True
    try:
        page = int(item.get("page") or 1)
    except (TypeError, ValueError):
        page = 1
    return bool(mover_coi_pages and page in mover_coi_pages and not quote_has_renters_insurance_context(quote))


def coi_trigger_value_from_quote(quote: str) -> str:
    text = re.sub(r"\s+", " ", normalize_unknown_value(quote) or "").strip()
    lowered = text.lower()
    if "moving company" in lowered and "one (1) week" in lowered:
        return (
            "Moving company must provide a completed certificate of liability insurance "
            "at least one (1) week before scheduling a move in or out."
        )
    if ("movers must" in lowered or "mover must" in lowered) and "reserving elevator" in lowered:
        return "Movers must provide COI before reserving the elevator / scheduling move-in."
    if "certificate of insurance for moving" in lowered:
        return "Certificate of Insurance is required for moving."
    return text


def normalize_insurance_requirement_for_field(field_key: str, value: object) -> Optional[str]:
    status = normalize_insurance_status(value)
    if status is None:
        return None
    if status == "optional" and field_key != "insurance_renters_required":
        return "manual_review"
    return status


def insurance_requirement_status_from_evidence(
    field_key: str,
    evidence_text: str,
    proposed_status: str,
) -> str:
    lowered = (evidence_text or "").lower()
    if field_key == "insurance_renters_required":
        if any(
            marker in lowered
            for marker in (
                "recommended but not required",
                "strongly recommended but not required",
                "not required but recommended",
                "optional",
                "may simply ignore",
                "may ignore",
            )
        ):
            return "optional"
        if any(marker in lowered for marker in ("no renters insurance required", "no insurance required")):
            return "no"
    if field_key in INSURANCE_REQUIREMENT_FIELD_KEYS and evidence_text:
        if sentence_has_required_language(evidence_text):
            return "yes"
    return proposed_status


INSURANCE_SUBMISSION_TOKENS = (
    "submit",
    "submitted",
    "upload",
    "provide",
    "verification",
    "proof",
    "declaration page",
    "policy",
    "portal",
    "guarantors",
    "leasing office",
)


INSURANCE_SUBMISSION_CUTOFF_PATTERNS = (
    r"\brenter[’']?s insurance personal liability insurance\b",
    r"\brenters insurance personal liability insurance\b",
    r"\bpersonal liability insurance\b",
    r"\brenters insurance requirements\b",
    r"\byour policy must include\b",
    r"\bcoverage against all claims\b",
    r"\bpersonal property coverage\b",
)


def clean_insurance_submission_quote(quote: str) -> Optional[str]:
    text = normalize_unknown_value(quote)
    if not text:
        return None
    text = re.sub(r"\s+", " ", text).strip()
    lowered = text.lower()
    for pattern in INSURANCE_SUBMISSION_CUTOFF_PATTERNS:
        match = re.search(pattern, lowered, flags=re.IGNORECASE)
        if match and match.start() > 0:
            text = text[: match.start()].strip(" \t\n.;")
            lowered = text.lower()
            break
    if not any(token in lowered for token in INSURANCE_SUBMISSION_TOKENS):
        return None
    sentence_parts = [part.strip(" \t\n") for part in re.split(r"(?<=[.!?])\s+", text) if part.strip()]
    submission_sentences = [
        part
        for part in sentence_parts
        if any(token in part.lower() for token in INSURANCE_SUBMISSION_TOKENS)
    ]
    if submission_sentences:
        return " ".join(dict.fromkeys(submission_sentences[:2])).strip()
    return text


def clean_ai_field_value_text(value: str, *, contact_context: bool = False) -> str:
    text = normalize_unknown_value(value) or ""
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    # Some models keep a trailing OCR/list number after a contact row.
    if contact_context or text_has_network_contact_signal(text):
        text = re.sub(r"\s+(?:\d{1,2}|[A-Z])\.$", "", text).strip()
    return text.strip(" \t\n,;")


def clean_insurance_coi_trigger_value(value: str, evidence_text: str) -> str:
    source = " ".join(
        part for part in [normalize_unknown_value(value) or "", normalize_unknown_value(evidence_text) or ""] if part
    )
    lowered = source.lower()
    pieces: List[str] = []
    if "moving company" in lowered or "mover" in lowered:
        pieces.append("moving company")
    if "move in or out" in lowered or "move-in or move-out" in lowered:
        pieces.append("move-in / move-out")
    elif "move in" in lowered or "move-in" in lowered:
        pieces.append("move-in")
    elif "move out" in lowered or "move-out" in lowered:
        pieces.append("move-out")
    if "delivery" in lowered:
        pieces.append("large-item delivery")
    if "service elevator" in lowered or "elevator" in lowered:
        pieces.append("service-elevator use")
    if "one (1) week" in lowered or "1 week" in lowered or "one week" in lowered:
        deadline = "at least one week"
    elif "72 hours" in lowered or "three business days" in lowered or "3 business days" in lowered:
        deadline = "at least 3 business days"
    elif "48 hours" in lowered or "two business days" in lowered or "2 business days" in lowered:
        deadline = "at least 2 business days"
    else:
        deadline = ""
    if pieces:
        unique_pieces = list(dict.fromkeys(pieces))
        if "moving company" in unique_pieces:
            actions = [piece for piece in unique_pieces if piece != "moving company"]
            trigger = f"moving-company booking for {' / '.join(actions)}" if actions else "moving company"
        else:
            trigger = " / ".join(unique_pieces)
        if not deadline:
            return f"COI required for {trigger}"
        if trigger == "moving company":
            return f"Submit the moving-company COI {deadline} in advance"
        return f"Submit the COI {deadline} before {trigger}"
    return clean_ai_field_value_text(value)


def sanitize_insurance_payloads(
    payloads: Dict[str, dict],
    candidates: List[dict],
) -> Dict[str, dict]:
    sanitized: Dict[str, dict] = {}
    mover_coi_pages = mover_coi_pages_from_candidates(candidates)
    for field_key, payload in (payloads or {}).items():
        if field_key not in INSURANCE_LLM_FIELD_KEYS or not isinstance(payload, dict):
            continue
        value = normalize_unknown_value(payload.get("value"))
        if value is None:
            continue

        evidence_items = payload_evidence_items(payload, candidates)
        evidence_items = filter_evidence_items_for_field(field_key, evidence_items)
        evidence_text = evidence_text_from_items(evidence_items)
        lowered_evidence = evidence_text.lower()
        review_flags = [*(payload.get("review_flags") or []), *evidence_domain_review_flags(evidence_items)]
        manual_reasons: List[str] = []
        confidence = float(payload.get("confidence") or 0.75)
        renters_context = quote_has_renters_insurance_context(evidence_text)
        evidence_pages: set[int] = set()
        for item in evidence_items:
            try:
                evidence_pages.add(int(item.get("page") or 1))
            except (TypeError, ValueError):
                evidence_pages.add(1)
        mover_coi_context = quote_has_mover_coi_context(evidence_text) or (
            bool(evidence_pages & mover_coi_pages) and not renters_context
        )

        if mover_coi_context and field_key in {
            "insurance_required",
            "insurance_coverage_amount",
            "insurance_renters_required",
            "insurance_renters_minimum_coverage",
            "insurance_personal_property_required",
            "insurance_personal_property_minimum",
            "insurance_personal_liability_required",
            "insurance_personal_liability_per_occurrence",
            "insurance_personal_liability_aggregate",
            "insurance_interested_party_required",
            "insurance_additional_insured_required",
            "insurance_certificate_holder_required",
            "insurance_submission_method",
            "insurance_recipient",
            "insurance_alternative_program_or_penalty",
        }:
            continue

        if field_key == "insurance_required":
            requirement_value = normalize_requirement_choice(value)
            if requirement_value is None:
                continue
            if evidence_items:
                if any(
                    marker in lowered_evidence
                    for marker in (
                        "recommended but not required",
                        "strongly recommended but not required",
                        "not required but recommended",
                        "optional",
                        "may simply ignore",
                        "may ignore",
                    )
                ):
                    requirement_value = "optional"
                elif any(marker in lowered_evidence for marker in ("no renters insurance required", "no insurance required")):
                    requirement_value = "false"
            value = requirement_value

        if field_key in INSURANCE_REQUIREMENT_FIELD_KEYS:
            normalized_status = normalize_insurance_requirement_for_field(field_key, value)
            if normalized_status is None:
                continue
            value = insurance_requirement_status_from_evidence(field_key, evidence_text, normalized_status)

        if field_key in {
            "insurance_renters_minimum_coverage",
            "insurance_coverage_amount",
            "insurance_personal_property_minimum",
            "insurance_personal_liability_per_occurrence",
            "insurance_personal_liability_aggregate",
        }:
            money_value = normalize_money_text(value)
            if not money_value:
                continue
            value = money_value

        if field_key in {"insurance_required", "insurance_coverage_amount", "insurance_renters_required", "insurance_renters_minimum_coverage"} and evidence_items:
            if not renters_context and (
                "coi" in lowered_evidence or "certificate of insurance" in lowered_evidence
            ):
                continue

        if field_key in {"insurance_renters_minimum_coverage", "insurance_coverage_amount"} and quote_has_personal_liability_amount_context(evidence_text):
            continue
        if field_key == "insurance_personal_property_minimum" and quote_has_personal_liability_amount_context(evidence_text):
            continue
        if field_key in {
            "insurance_personal_liability_required",
            "insurance_personal_liability_per_occurrence",
            "insurance_personal_liability_aggregate",
        } and evidence_items:
            if not renters_context and "personal liability" not in lowered_evidence:
                continue
        if field_key in {"insurance_personal_liability_per_occurrence", "insurance_personal_liability_aggregate"}:
            if evidence_items and not quote_has_personal_liability_amount_context(evidence_text):
                review_flags.append("ambiguous_liability_amount")
                manual_reasons.append("The amount refers to insurance coverage but does not explicitly identify Personal Liability.")

        if field_key == "insurance_coi_required" and evidence_items:
            if (
                "coi" not in lowered_evidence
                and "certificate of insurance" not in lowered_evidence
                and "certificate of liability insurance" not in lowered_evidence
            ):
                continue
            if value == "yes" and not sentence_has_required_language(evidence_text):
                value = "manual_review"
                review_flags.append("ambiguous_coi_requirement")
                manual_reasons.append("The source mentions a COI, but a person must confirm whether it is mandatory.")

        if field_key == "insurance_coi_trigger" and evidence_items:
            if (
                "coi" not in lowered_evidence
                and "certificate of insurance" not in lowered_evidence
                and "certificate of liability insurance" not in lowered_evidence
            ):
                continue
            if not any(
                token in lowered_evidence
                for token in (
                    "if",
                    "when",
                    "using",
                    "moving company",
                    "delivery",
                    "service elevator",
                    "required",
                    "must",
                    "provide",
                    "complete",
                )
            ):
                continue
            value = clean_insurance_coi_trigger_value(value, evidence_text)

        if field_key == "insurance_interested_party_required" and evidence_items:
            if "interested party" not in lowered_evidence and "additional interest" not in lowered_evidence:
                if value == "yes":
                    value = "manual_review"
                    review_flags.append("ambiguous_party_type")
                    manual_reasons.append("The source does not explicitly name an Interested Party or Additional Interest.")

        if field_key == "insurance_additional_insured_required" and evidence_items and "additional insured" not in lowered_evidence:
            continue
        if field_key == "insurance_certificate_holder_required" and evidence_items and "certificate holder" not in lowered_evidence:
            continue
        if field_key in {
            "insurance_submission_method",
            "insurance_recipient",
            "insurance_alternative_program_or_penalty",
        } and evidence_items:
            if not any(
                token in lowered_evidence
                for token in (
                    "insurance",
                    "renters",
                    "renter's",
                    "renter’s",
                    "guarantors",
                    "policy",
                    "declaration page",
                    "protection program",
                    "interested party",
                    "additional interest",
                )
            ):
                continue

        if field_key == "insurance_submission_method" and evidence_items:
            submission_quotes = [
                cleaned_quote
                for item in evidence_items
                for cleaned_quote in [clean_insurance_submission_quote(item["quote"])]
                if cleaned_quote
            ]
            if submission_quotes:
                value = "\n".join(dict.fromkeys(submission_quotes[:2]))

        if not evidence_items:
            review_flags.append("no_direct_evidence")
            manual_reasons.append("No directly supporting source sentence was found.")
            confidence = min(confidence, 0.55)
        else:
            confidence = min(max(confidence, 0.72), 0.96)

        existing_reason = normalize_unknown_value(payload.get("manual_review_reason"))
        if existing_reason:
            manual_reasons.append(existing_reason)

        sanitized[field_key] = {
            "value": value,
            "confidence": confidence,
            "evidence": evidence_items[0]["quote"] if evidence_items else (payload.get("evidence") or ""),
            "evidence_items": evidence_items,
            "manual_review_reason": "; ".join(dict.fromkeys(reason for reason in manual_reasons if reason)),
            "review_flags": list(dict.fromkeys(review_flags)),
        }
    return sanitized


def extract_electric_provider(text: str) -> Optional[str]:
    lowered = (text or "").lower()
    providers: List[str] = []
    if any(token in lowered for token in ("con edison", "coned", "con-ed")):
        providers.append("Con Edison")
    if any(token in lowered for token in ("pse&g", "pseg", "pse g")):
        providers.append("PSE&G")
    if "national grid" in lowered:
        providers.append("National Grid")
    return ", ".join(dict.fromkeys(providers)) or None


def evidence_says_electric_included(text: str) -> bool:
    lowered = (text or "").lower()
    return (
        ("electric" in lowered or "electricity" in lowered)
        and any(token in lowered for token in ("included", "provided", "building covers", "landlord pays"))
    )


def evidence_says_electric_required(text: str) -> bool:
    lowered = (text or "").lower()
    if not ("electric" in lowered or "electricity" in lowered or "con edison" in lowered or "pseg" in lowered or "pse&g" in lowered):
        return False
    return any(
        token in lowered
        for token in (
            "set up",
            "open",
            "contact",
            "provide verification",
            "proof of electric",
            "proof of your account",
            "account number",
            "prior to move-in",
            "before your move-in",
            "electric service",
            "utility information",
            "utilities/electric",
            "electric provider",
            "utility",
            "utilities",
            "account",
            "proof",
            "upload",
            "rello",
            "pre move",
            "pre-move",
        )
    )


def find_electricity_requirement_candidate(
    candidates: List[dict],
    provider: Optional[str] = None,
) -> Optional[dict]:
    provider_tokens: List[str] = []
    if provider == "Con Edison":
        provider_tokens = ["con edison", "coned", "con-ed"]
    elif provider == "PSE&G":
        provider_tokens = ["pse&g", "pseg", "pse g"]
    elif provider == "National Grid":
        provider_tokens = ["national grid"]

    for candidate in candidates or []:
        if not candidate_has_allowed_domain(candidate, {"electricity"}):
            continue
        lowered = candidate.get("lowered") or (candidate.get("quote") or "").lower()
        has_provider = any(token in lowered for token in provider_tokens) if provider_tokens else False
        has_electric = "electric" in lowered or "electricity" in lowered
        if not has_provider and not has_electric:
            continue
        if evidence_says_electric_included(lowered):
            continue
        if evidence_says_electric_required(lowered) or (
            has_provider
            and any(
                token in lowered
                for token in (
                    "utility",
                    "utilities",
                    "move-in",
                    "move in",
                    "setup",
                    "set up",
                    "account",
                    "proof",
                    "upload",
                    "rello",
                    "pre move",
                    "pre-move",
                    "service",
                    "contact",
                )
            )
        ):
            return candidate
    return None


def sanitize_electricity_payloads(
    payloads: Dict[str, dict],
    candidates: List[dict],
) -> Dict[str, dict]:
    sanitized: Dict[str, dict] = {}
    domain_evidence_items: List[dict] = []
    for source_field_key, source_payload in (payloads or {}).items():
        if source_field_key not in ELECTRICITY_LLM_FIELD_KEYS or not isinstance(source_payload, dict):
            continue
        domain_evidence_items.extend(
            filter_evidence_items_for_field(source_field_key, payload_evidence_items(source_payload, candidates))
        )
    domain_evidence_items = domain_evidence_items[:4]
    domain_evidence_text = evidence_text_from_items(domain_evidence_items)
    domain_provider_hint = extract_electric_provider(domain_evidence_text)
    domain_provider_candidate = find_electricity_requirement_candidate(candidates, domain_provider_hint)

    for field_key, payload in (payloads or {}).items():
        if field_key not in ELECTRICITY_LLM_FIELD_KEYS or not isinstance(payload, dict):
            continue
        value = normalize_unknown_value(payload.get("value"))
        if value is None:
            continue
        evidence_items = payload_evidence_items(payload, candidates)
        evidence_items = filter_evidence_items_for_field(field_key, evidence_items)
        evidence_text = evidence_text_from_items(evidence_items)
        provider_hint = (
            extract_electric_provider(evidence_text)
            or extract_electric_provider(value)
            or extract_electric_provider(str(payload.get("evidence") or ""))
            or domain_provider_hint
        )
        provider_candidate = find_electricity_requirement_candidate(candidates, provider_hint) or domain_provider_candidate
        if provider_candidate and not evidence_items:
            evidence_items = [{"page": provider_candidate.get("page") or 1, "quote": provider_candidate.get("quote") or ""}]
            evidence_text = evidence_text_from_items(evidence_items)
        combined_electric_text = "\n".join(
            part
            for part in (
                evidence_text,
                provider_candidate.get("quote") if provider_candidate else "",
                domain_evidence_text,
            )
            if part
        )
        review_flags = [*(payload.get("review_flags") or []), *evidence_domain_review_flags(evidence_items)]
        manual_reasons: List[str] = []
        confidence = float(payload.get("confidence") or 0.75)

        if field_key == "electricity_required":
            if evidence_says_electric_included(combined_electric_text):
                value = "false"
            elif evidence_says_electric_required(combined_electric_text) or provider_candidate:
                value = "true"
                if not evidence_items and domain_evidence_items:
                    evidence_items = domain_evidence_items[:4]
                    evidence_text = evidence_text_from_items(evidence_items)
            else:
                value = "manual_review"
                review_flags.append("ambiguous_electricity_requirement")
                manual_reasons.append("The source does not directly state whether the tenant must open an electricity account.")

        if field_key == "electricity_provider":
            provider = extract_electric_provider(evidence_text) or extract_electric_provider(value)
            if not provider:
                continue
            value = provider
            if not evidence_items and domain_evidence_items:
                evidence_items = domain_evidence_items[:4]

        if not evidence_items:
            review_flags.append("no_direct_evidence")
            manual_reasons.append("No directly supporting source sentence was found.")
            confidence = min(confidence, 0.55)
        else:
            confidence = min(max(confidence, 0.72), 0.96)

        existing_reason = normalize_unknown_value(payload.get("manual_review_reason"))
        if existing_reason:
            manual_reasons.append(existing_reason)

        sanitized[field_key] = {
            "value": value,
            "confidence": confidence,
            "evidence": evidence_items[0]["quote"] if evidence_items else (payload.get("evidence") or ""),
            "evidence_items": evidence_items,
            "manual_review_reason": "; ".join(dict.fromkeys(reason for reason in manual_reasons if reason)),
            "review_flags": list(dict.fromkeys(review_flags)),
        }
    return sanitized


def amounts_in_text(text: str) -> List[str]:
    seen: List[str] = []
    for match in re.finditer(r"\$\s?\d[\d,]*(?:\.\d{2})?", text or ""):
        amount = match.group(0).replace(" ", "")
        if amount not in seen:
            seen.append(amount)
    return seen


def sentence_has_required_language(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        token in lowered
        for token in (
            "required",
            "require",
            "requires",
            "must",
            "need to",
            "needs to",
            "shall",
            "proof of",
            "maintain",
            "obtain",
            "provide",
            "submit",
            "submitted",
            "complete",
            "completion",
            "will be required",
        )
    )


def sentence_has_optional_language(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        token in lowered
        for token in (
            "optional",
            "recommended but not required",
            "strongly recommended but not required",
            "not required but recommended",
            "may ignore",
            "not required",
        )
    )


def sentence_has_negative_language(text: str) -> bool:
    lowered = (text or "").lower()
    return any(token in lowered for token in ("not required", "no insurance required", "may ignore"))


def payload_with_evidence(
    *,
    value: Optional[str],
    quote: Optional[str],
    page: Optional[int],
    confidence: float = 1.0,
    manual_review_reason: str = "",
    review_flags: Optional[List[str]] = None,
) -> Optional[dict]:
    normalized_value = normalize_unknown_value(value)
    if normalized_value is None:
        normalized_value = value if isinstance(value, str) and value.strip() else None
    if normalized_value is None:
        return None
    evidence_items = []
    if normalize_unknown_value(quote):
        evidence_items.append({"page": page or 1, "quote": normalize_unknown_value(quote)})
    return {
        "value": normalized_value,
        "confidence": confidence,
        "evidence": normalize_unknown_value(quote) or "",
        "evidence_items": evidence_items,
        "manual_review_reason": manual_review_reason.strip(),
        "review_flags": review_flags or [],
    }


def merge_field_payload(result: Dict[str, dict], field_key: str, payload: Optional[dict]) -> None:
    if not payload:
        return
    current = result.get(field_key)
    if not current or float(payload.get("confidence") or 0) >= float(current.get("confidence") or 0):
        result[field_key] = payload


def classify_document_type(raw_input_type: str, source_file: str, full_text: str) -> str:
    lowered_text = (full_text or "").lower()
    lowered_name = (source_file or "").lower()
    if "welcome packet" in lowered_text or "welcome packet" in lowered_name:
        return "welcome packet"
    if "welcome letter" in lowered_text or "welcome letter" in lowered_name:
        return "welcome letter"
    if raw_input_type == "text" and any(token in lowered_text for token in ("from:", "sent:", "subject:")):
        return "email"
    if raw_input_type == "image":
        return "OCR text"
    if raw_input_type in {"pdf", "pdf_package"}:
        return "PDF"
    if raw_input_type == "text":
        return "email"
    return "unknown"


def provider_from_sentence(candidate: dict) -> Optional[str]:
    if not candidate:
        return None
    quote = candidate["quote"]
    lowered = candidate["lowered"]
    if not any(token in lowered for token in ("internet", "wifi", "provider", "verizon", "xfinity", "astound", "spectrum", "honest")):
        return None
    if any(token in lowered for token in ("terms and conditions", "license", "internet rider", "amenities", "subject to honest")):
        return None
    providers = extract_provider_names(quote)
    providers = [
        label
        for label in providers
        if label in FIXED_NETWORK_PROVIDER_SET or label == "Honest Networks"
    ]
    if not providers:
        return None
    return ", ".join(providers)


def provider_note_from_quote(quote: str, provider_label: str) -> str:
    text = normalize_unknown_value(quote) or ""
    if not text:
        return ""
    escaped = re.escape(provider_label)
    match = re.search(
        rf"{escaped}(?:\s+Fios)?(?:\s*\([^)]*\))?\s*[:：\-]\s*(.+)$",
        text,
        flags=re.IGNORECASE,
    )
    if match and normalize_unknown_value(match.group(1)):
        return normalize_unknown_value(match.group(1)) or ""
    return text


def provider_segment_from_quote(quote: str, provider_label: str) -> str:
    text = normalize_unknown_value(quote) or ""
    if not text:
        return ""
    lowered = text.lower()
    raw_tokens_by_label = {
        "Verizon": ("verizon",),
        "Xfinity": ("xfinity", "xifinity"),
        "Spectrum": ("spectrum",),
        "Astound": ("astound",),
        "Honest Networks": ("honest networks", "honest"),
    }
    start_positions = [
        lowered.find(token)
        for token in raw_tokens_by_label.get(provider_label, (provider_label.lower(),))
        if lowered.find(token) >= 0
    ]
    if not start_positions:
        return ""
    start = min(start_positions)
    boundary_positions: List[int] = []
    for label, tokens in raw_tokens_by_label.items():
        if label == provider_label:
            continue
        for token in tokens:
            pos = lowered.find(token, start + 1)
            if pos > start:
                boundary_positions.append(pos)
    for token in (
        "renters insurance",
        "renter's insurance",
        "renter’s insurance",
        "insurance requirements",
        "pet requirements",
        "interested party",
        "utility information",
        "water, sewer",
        "must have renters",
    ):
        pos = lowered.find(token, start + 1)
        if pos > start:
            boundary_positions.append(pos)
    end = min(boundary_positions) if boundary_positions else len(text)
    return text[start:end].strip(" \n\t,;")


def leading_network_contact_segment(quote: str, current_label: str) -> str:
    text = normalize_unknown_value(quote) or ""
    if not text:
        return ""
    lowered = text.lower()
    boundary_positions: List[int] = []
    raw_tokens_by_label = {
        "Verizon": ("verizon",),
        "Xfinity": ("xfinity", "xifinity"),
        "Spectrum": ("spectrum",),
        "Astound": ("astound",),
        "Honest Networks": ("honest networks", "honest"),
    }
    for label, tokens in raw_tokens_by_label.items():
        if label == current_label:
            continue
        for token in tokens:
            pos = lowered.find(token)
            if pos > 0:
                boundary_positions.append(pos)
    for token in (
        "renters insurance",
        "renter's insurance",
        "renter’s insurance",
        "insurance requirements",
        "pet requirements",
        "interested party",
        "utility information",
        "water, sewer",
        "must have renters",
        "signed lease before receiving keys",
    ):
        pos = lowered.find(token)
        if pos > 0:
            boundary_positions.append(pos)
    end = min(boundary_positions) if boundary_positions else len(text)
    segment = text[:end].strip(" \n\t,;")
    return segment if text_has_network_contact_signal(segment) else ""


def extend_network_provider_part_with_neighbors(part: dict, candidates: List[dict], index: int) -> dict:
    page = int(part.get("page") or 1)
    segments = dict(part.get("segments") or {})
    for label in part.get("fixed") or []:
        segment = normalize_unknown_value(segments.get(label)) or ""
        if text_has_network_contact_signal(segment):
            continue
        contact_pieces: List[str] = []
        for neighbor in candidates[index + 1 : index + 3]:
            if int(neighbor.get("page") or 1) != page:
                break
            contact_piece = leading_network_contact_segment(neighbor.get("quote") or "", label)
            if contact_piece:
                if label != "Honest Networks" and "http" in contact_piece.lower() and not re.search(r"[\w.+-]+@[\w.-]+\.\w+|\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b", contact_piece):
                    break
                contact_pieces.append(contact_piece)
                continue
            break
        if contact_pieces:
            segments[label] = " ".join([segment, *contact_pieces]).strip()
    part["segments"] = segments
    return part


def network_extra_note_items_from_candidates(candidates: List[dict]) -> List[dict]:
    items: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for index, candidate in enumerate(candidates):
        lowered = candidate.get("lowered") or ""
        if "honest" not in lowered:
            continue
        page = int(candidate.get("page") or 1)
        initial_piece = provider_segment_from_quote(candidate.get("quote") or "", "Honest Networks")
        pieces = [initial_piece or candidate.get("quote") or ""]
        for neighbor in candidates[index + 1 : index + 3]:
            if int(neighbor.get("page") or 1) != page:
                break
            neighbor_quote = neighbor.get("quote") or ""
            neighbor_lowered = neighbor.get("lowered") or neighbor_quote.lower()
            if any(label.lower() in neighbor_lowered for label in FIXED_NETWORK_PROVIDER_SET):
                break
            if any(
                token in neighbor_lowered
                for token in (
                    "http",
                    "www.",
                    "@",
                    "wifi",
                    "wi-fi",
                    "internet",
                    "contact",
                    "phone",
                    "agent",
                )
            ) or text_has_network_contact_signal(neighbor_quote):
                pieces.append(neighbor_quote)
                continue
            break
        quote = re.sub(r"\s+", " ", " ".join(piece for piece in pieces if piece)).strip()
        if not quote:
            continue
        marker = (page, quote)
        if marker in seen:
            continue
        seen.add(marker)
        items.append({"page": page, "quote": quote})
    return items


def network_provider_parts_from_candidate(candidate: dict) -> Optional[dict]:
    if not candidate:
        return None
    quote = normalize_unknown_value(candidate.get("quote")) or ""
    lowered = candidate.get("lowered") or quote.lower()
    if not quote:
        return None
    if not any(token in lowered for token in ("internet", "wifi", "wi-fi", "provider", "verizon", "xfinity", "astound", "spectrum", "honest")):
        return None
    if any(token in lowered for token in ("terms and conditions", "license", "internet rider", "amenities", "subject to honest")):
        return None
    providers = extract_provider_names(quote)
    providers = [
        label
        for label in providers
        if label in FIXED_NETWORK_PROVIDER_SET or label == "Honest Networks"
    ]
    if not providers:
        return None
    fixed = [label for label in providers if label in FIXED_NETWORK_PROVIDER_SET]
    extras = [label for label in providers if label not in FIXED_NETWORK_PROVIDER_SET]
    segments = {label: provider_segment_from_quote(quote, label) for label in providers}
    return {
        "fixed": fixed,
        "extras": extras,
        "segments": segments,
        "quote": quote,
        "page": candidate.get("page") or 1,
    }


def network_payload_evidence_items(payload: Optional[dict]) -> List[dict]:
    if not payload:
        return []
    items = []
    for item in payload.get("evidence_items") or []:
        quote = normalize_unknown_value(item.get("quote"))
        if quote:
            items.append({"page": item.get("page") or 1, "quote": quote})
    if not items and normalize_unknown_value(payload.get("evidence")):
        items.append({"page": 1, "quote": normalize_unknown_value(payload.get("evidence"))})
    if not items and normalize_unknown_value(payload.get("value")):
        items.append({"page": 1, "quote": normalize_unknown_value(payload.get("value"))})
    return items


def ensure_known_extra_provider_payload_from_notes(cleaned: Dict[str, dict], *payloads: Optional[dict]) -> Dict[str, dict]:
    source_parts: List[str] = []
    evidence_items: List[dict] = []
    for payload in payloads:
        if not isinstance(payload, dict):
            continue
        source_parts.append(normalize_unknown_value(payload.get("value")) or "")
        source_parts.append(normalize_unknown_value(payload.get("evidence")) or "")
        for item in payload.get("evidence_items") or []:
            if not isinstance(item, dict):
                continue
            quote = normalize_unknown_value(item.get("quote")) or ""
            source_parts.append(quote)
            if quote:
                evidence_items.append({"page": item.get("page") or 1, "quote": quote})

    notes_payload = cleaned.get("internet_notes")
    if isinstance(notes_payload, dict):
        source_parts.append(normalize_unknown_value(notes_payload.get("value")) or "")
        source_parts.append(normalize_unknown_value(notes_payload.get("evidence")) or "")
        for item in notes_payload.get("evidence_items") or []:
            if not isinstance(item, dict):
                continue
            quote = normalize_unknown_value(item.get("quote")) or ""
            source_parts.append(quote)
            if quote:
                evidence_items.append({"page": item.get("page") or 1, "quote": quote})

    provider_payload = cleaned.get("internet_provider")
    existing_labels = (
        extra_provider_labels_from_text(provider_payload.get("value"))
        if isinstance(provider_payload, dict)
        else []
    )
    for label in known_extra_provider_labels_from_text("\n".join(source_parts)):
        if label not in existing_labels:
            existing_labels.append(label)
    if not existing_labels:
        return cleaned

    merge_field_payload(
        cleaned,
        "internet_provider",
        {
            "value": ", ".join(existing_labels),
            "confidence": max(float((provider_payload or {}).get("confidence") or 0), 0.9)
            if isinstance(provider_payload, dict)
            else 0.9,
            "evidence": evidence_items[0]["quote"] if evidence_items else "\n".join(part for part in source_parts if part),
            "evidence_items": evidence_items[:4],
            "manual_review_reason": "",
            "review_flags": [],
        },
    )
    return cleaned


def text_has_network_included_meaning(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        marker in lowered
        for marker in (
            "internet included",
            "included internet",
            "wifi included",
            "wi-fi included",
            "wifi is included",
            "wi-fi is included",
            "wifi is being provided",
            "wi-fi is being provided",
            "internet is being provided",
            "being provided by",
            "provided by the building",
            "楼内已包含网络",
            "大楼自带网络",
            "网络已包含",
            "包含网络",
        )
    )


def text_has_network_contact_signal(text: str) -> bool:
    lowered = (text or "").lower()
    return bool(
        re.search(r"[\w.+-]+@[\w.-]+\.[a-z]{2,}", text or "", flags=re.IGNORECASE)
        or re.search(r"https?://|www\.", lowered)
        or re.search(r"\b\d{3}[-.\s]\d{3}[-.\s]\d{4}\b", text or "")
        or any(token in lowered for token in ("contact", "call", "phone", "agent", "representative"))
    )


def provider_note_has_actionable_detail(note: str, provider_label: str) -> bool:
    text = normalize_unknown_value(note) or ""
    if not text:
        return False
    lowered = text.lower()
    provider_terms = [
        provider_label.lower(),
        provider_label.lower().replace(" ", ""),
        "fios",
        "cable",
        "internet",
    ]
    remainder = lowered
    for term in provider_terms:
        remainder = remainder.replace(term, " ")
    remainder = re.sub(r"[\s,.;:()&/-]+", " ", remainder).strip()
    if not remainder:
        return False
    if text_has_network_contact_signal(text):
        return True
    return any(
        token in lowered
        for token in (
            "http",
            "www.",
            "mbps",
            "gbps",
            "gig",
            "$",
            "month",
            "free",
            "install",
            "setup",
            "set up",
            "router",
            "wifi",
            "wi-fi",
            "mobile",
            "instant-on",
            "instant on",
        )
    )


def actionable_provider_note_from_segment(segment: str, provider_label: str) -> str:
    note = provider_note_from_quote(segment, provider_label)
    note = clean_ai_field_value_text(note, contact_context=True)
    if note and note != provider_label and provider_note_has_actionable_detail(note, provider_label):
        return note
    return ""


def provider_segments_from_evidence(evidence_items: List[dict], provider_label: str) -> List[dict]:
    segments: List[dict] = []
    seen: set[tuple[int, str]] = set()
    for item in evidence_items:
        if not isinstance(item, dict):
            continue
        quote = normalize_unknown_value(item.get("quote"))
        if not quote:
            continue
        segment = provider_segment_from_quote(quote, provider_label)
        if not segment:
            continue
        marker = (int(item.get("page") or 1), segment)
        if marker in seen:
            continue
        seen.add(marker)
        segments.append({"page": marker[0], "quote": segment})
    return segments


def ensure_network_self_setup_from_contact_payloads(parsed: Dict[str, dict]) -> Dict[str, dict]:
    cleaned = dict(parsed or {})
    current = cleaned.get("internet_self_setup_required") or {}
    current_state = normalize_requirement_choice(current.get("value"))
    if current_state in {"true", "false", "optional"}:
        return cleaned

    provider_note_fields = {field_key: label for field_key, label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items()}
    provider_support_fields = {field_key: label for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items()}
    evidence_items: List[dict] = []

    for field_key, label in provider_support_fields.items():
        payload = cleaned.get(field_key) or {}
        support_state = normalize_requirement_choice(payload.get("value"))
        if support_state != "true" and normalize_booleanish(payload.get("value")) is not True:
            continue
        evidence_items.extend(network_payload_evidence_items(payload))
        if not network_payload_evidence_items(payload):
            evidence_items.append({"page": 1, "quote": label})

    for field_key, label in provider_note_fields.items():
        payload = cleaned.get(field_key) or {}
        note_value = normalize_unknown_value(payload.get("value"))
        if not note_value:
            continue
        note_evidence = network_payload_evidence_items(payload)
        combined = "\n".join([note_value, *(item["quote"] for item in note_evidence)])
        if label.lower() in combined.lower() or text_has_network_contact_signal(combined):
            evidence_items.extend(note_evidence or [{"page": 1, "quote": note_value}])

    notes_payload = cleaned.get("internet_notes") or {}
    notes_value = normalize_unknown_value(notes_payload.get("value"))
    if notes_value:
        provider_in_notes = any(
            label.lower() in notes_value.lower()
            for label in (*FIXED_NETWORK_PROVIDER_SET, "Honest Networks", "Honest")
        )
        if provider_in_notes and text_has_network_contact_signal(notes_value):
            evidence_items.extend(network_payload_evidence_items(notes_payload) or [{"page": 1, "quote": notes_value}])

    if not evidence_items:
        return cleaned
    deduped_evidence: List[dict] = []
    seen_evidence: set[tuple[int, str]] = set()
    for item in evidence_items:
        quote = normalize_unknown_value(item.get("quote"))
        if not quote:
            continue
        key = (int(item.get("page") or 1), quote)
        if key in seen_evidence:
            continue
        seen_evidence.add(key)
        deduped_evidence.append({"page": key[0], "quote": quote})
    evidence_items = deduped_evidence
    if not evidence_items:
        return cleaned
    combined_evidence = "\n".join(item["quote"] for item in evidence_items)
    if text_has_network_included_meaning(combined_evidence):
        return cleaned

    first = evidence_items[0]
    cleaned["internet_self_setup_required"] = {
        "value": "true",
        "confidence": max(float(current.get("confidence") or 0), 0.96),
        "evidence": first["quote"],
        "evidence_items": evidence_items[:4],
        "manual_review_reason": "",
        "review_flags": ["derived_from_supported_network_provider_contact"],
    }
    return cleaned


def sanitize_network_payloads(parsed: Dict[str, dict]) -> Dict[str, dict]:
    cleaned = dict(parsed or {})
    for note_field in NETWORK_PROVIDER_NOTE_FIELD_MAP.keys():
        payload = cleaned.get(note_field)
        if isinstance(payload, dict) and normalize_unknown_value(payload.get("value")):
            payload = dict(payload)
            payload["value"] = clean_ai_field_value_text(str(payload.get("value")), contact_context=True)
            cleaned[note_field] = payload
    provider_payload = cleaned.pop("internet_provider", None)
    if not provider_payload:
        return ensure_network_self_setup_from_contact_payloads(
            ensure_known_extra_provider_payload_from_notes(cleaned)
        )
    provider_value = normalize_unknown_value(provider_payload.get("value"))
    raw_evidence_items = provider_payload.get("evidence_items") or []
    provider_source_text = "\n".join(
        [
            provider_value or "",
            *[
                normalize_unknown_value(item.get("quote")) or ""
                for item in raw_evidence_items
                if isinstance(item, dict)
            ],
            normalize_unknown_value(provider_payload.get("evidence")) or "",
        ]
    )
    providers = extract_provider_names(provider_source_text)
    providers = [
        label
        for label in providers
        if label in FIXED_NETWORK_PROVIDER_SET or label == "Honest Networks"
    ]
    if not providers:
        return ensure_network_self_setup_from_contact_payloads(
            ensure_known_extra_provider_payload_from_notes(cleaned, provider_payload)
        )

    evidence_items = [item for item in raw_evidence_items if isinstance(item, dict)]
    if not evidence_items and normalize_unknown_value(provider_payload.get("evidence")):
        evidence_items = [{"page": 1, "quote": normalize_unknown_value(provider_payload.get("evidence"))}]
    if not evidence_items:
        evidence_items = [{"page": 1, "quote": provider_value}]
    primary_quote = normalize_unknown_value(evidence_items[0].get("quote")) or provider_value or ""
    fixed_label_to_field = {label: field_key for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items()}
    fixed_label_to_note_field = {label: field_key for field_key, label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items()}
    extra_notes: List[dict] = []
    extra_provider_labels: List[str] = []
    for label in providers:
        provider_segments = provider_segments_from_evidence(evidence_items, label)
        if not provider_segments:
            provider_segment = provider_segment_from_quote(primary_quote, label)
            provider_segments = [{"page": evidence_items[0].get("page") or 1, "quote": provider_segment}]
        provider_segment = provider_segments[0]["quote"]
        provider_evidence = provider_segments[:4]
        if label in FIXED_NETWORK_PROVIDER_SET:
            support_field = fixed_label_to_field.get(label)
            if support_field:
                merge_field_payload(
                    cleaned,
                    support_field,
                    {
                        "value": "true",
                        "confidence": float(provider_payload.get("confidence") or 0.85),
                        "evidence": provider_segment,
                        "evidence_items": provider_evidence,
                        "manual_review_reason": "",
                        "review_flags": [],
                    },
                )
            note_field = fixed_label_to_note_field.get(label)
            note_segment = next(
                (
                    item["quote"]
                    for item in provider_segments
                    if actionable_provider_note_from_segment(item["quote"], label)
                ),
                "",
            )
            note_text = actionable_provider_note_from_segment(note_segment, label)
            if note_field and note_text:
                merge_field_payload(
                    cleaned,
                    note_field,
                    {
                        "value": note_text,
                        "confidence": float(provider_payload.get("confidence") or 0.8),
                        "evidence": note_segment,
                        "evidence_items": [
                            item for item in provider_segments[:4] if actionable_provider_note_from_segment(item["quote"], label)
                        ]
                        or provider_evidence,
                        "manual_review_reason": "",
                        "review_flags": [],
                    },
                )
        else:
            if label not in extra_provider_labels:
                extra_provider_labels.append(label)
            extra_note_lines = [provider_segment] if provider_segment else []
            if label == "Honest Networks":
                for item in evidence_items:
                    quote = normalize_unknown_value(item.get("quote"))
                    if not quote:
                        continue
                    lowered_quote = quote.lower()
                    if "honest" in lowered_quote:
                        extra_note_lines.append(provider_segment_from_quote(quote, label))
                    elif "honest.net" in lowered_quote:
                        extra_note_lines.append(quote)
            extra_note_value = " ".join(dict.fromkeys(line for line in extra_note_lines if line)) or label
            extra_notes.append({"value": extra_note_value, "evidence_items": provider_evidence})
    if extra_provider_labels:
        merge_field_payload(
            cleaned,
            "internet_provider",
            {
                "value": ", ".join(extra_provider_labels),
                "confidence": float(provider_payload.get("confidence") or 0.9),
                "evidence": primary_quote,
                "evidence_items": evidence_items[:4],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )
    if extra_notes:
        existing_notes = normalize_unknown_value((cleaned.get("internet_notes") or {}).get("value"))
        note_lines = [*(existing_notes.split("\n") if existing_notes else []), *(item["value"] for item in extra_notes)]
        extra_evidence_items: List[dict] = []
        for item in extra_notes:
            extra_evidence_items.extend(item.get("evidence_items") or [])
        merge_field_payload(
            cleaned,
            "internet_notes",
            {
                "value": "\n".join(dict.fromkeys(line for line in note_lines if line)),
                "confidence": float(provider_payload.get("confidence") or 0.8),
                "evidence": extra_evidence_items[0]["quote"] if extra_evidence_items else primary_quote,
                "evidence_items": extra_evidence_items[:4] or evidence_items[:4],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )
    return ensure_network_self_setup_from_contact_payloads(
        ensure_known_extra_provider_payload_from_notes(cleaned, provider_payload)
    )


def parse_welcome_letter_fields_from_pages(
    *,
    raw_input_type: str,
    source_file: str,
    pages: List[dict],
    classified_candidates: Optional[List[dict]] = None,
) -> Dict[str, dict]:
    full_text = join_page_texts(pages)
    candidates = classified_candidates or classify_evidence_candidates_by_rules(build_sentence_candidates(pages))
    mover_coi_pages = mover_coi_pages_from_candidates(candidates)
    result: Dict[str, dict] = {}

    merge_field_payload(
        result,
        "document_type",
        payload_with_evidence(
            value=classify_document_type(raw_input_type, source_file, full_text),
            quote=candidates[0]["quote"] if candidates else source_file,
            page=candidates[0]["page"] if candidates else 1,
            confidence=1.0,
        ),
    )

    building_candidate = first_candidate(
        candidates,
        lambda item: (
            (
                any(
                    token in item["lowered"]
                    for token in ("welcome to", "new home at", "home at", "subject:")
                )
                and candidate_has_allowed_domain(item, {"building_identity"})
            )
            or any(
                token in item["lowered"]
                for token in ("will activate water", "interested party information", "as an interested party")
            )
        )
        and (
            not is_non_identity_building_context(item["quote"])
            or any(
                token in item["lowered"]
                for token in ("will activate water", "interested party information", "as an interested party")
            )
        ),
    )
    building_name = parse_building_name_from_text(building_candidate["quote"] if building_candidate else full_text)
    merge_field_payload(
        result,
        "building_name",
        payload_with_evidence(
            value=building_name or normalize_unknown_value(source_file.rsplit(".", 1)[0]),
            quote=building_candidate["quote"] if building_candidate else source_file,
            page=building_candidate["page"] if building_candidate else 1,
            confidence=0.95 if building_candidate else 0.55,
            manual_review_reason="" if building_candidate else "No explicit welcome phrase was found; the filename was used as a fallback.",
            review_flags=[] if building_candidate else ["fallback_from_filename"],
        ),
    )

    address_candidate = first_candidate(
        candidates,
        lambda item: bool(parse_address_from_text(item["quote"]))
        and candidate_has_allowed_domain(item, {"building_identity"}),
    )
    if not address_candidate:
        address_candidate = first_candidate(
            candidates,
            lambda item: bool(parse_address_from_text(item["quote"]))
            and not candidate_has_mover_coi_context(item, mover_coi_pages),
        )
    merge_field_payload(
        result,
        "address",
        payload_with_evidence(
            value=parse_address_from_text(address_candidate["quote"] if address_candidate else full_text),
            quote=address_candidate["quote"] if address_candidate else "",
            page=address_candidate["page"] if address_candidate else 1,
            confidence=0.92 if address_candidate else 0.65,
        ),
    )

    electricity_candidate = first_candidate(
        candidates,
        lambda item: any(token in item["lowered"] for token in ("electric account", "con edison", "pseg", "pse&g"))
        and not any(token in item["lowered"] for token in ("included", "billing statement")),
    )
    electricity_included_candidate = first_candidate(
        candidates,
        lambda item: "electric" in item["lowered"] and "included" in item["lowered"],
    )
    if electricity_candidate:
        merge_field_payload(
            result,
            "electricity_required",
            payload_with_evidence(value="true", quote=electricity_candidate["quote"], page=electricity_candidate["page"]),
        )
        provider_value = extract_electric_provider(electricity_candidate["quote"])
        merge_field_payload(
            result,
            "electricity_provider",
            payload_with_evidence(
                value=provider_value,
                quote=electricity_candidate["quote"],
                page=electricity_candidate["page"],
            ),
        )
    elif electricity_included_candidate:
        merge_field_payload(
            result,
            "electricity_required",
            payload_with_evidence(value="false", quote=electricity_included_candidate["quote"], page=electricity_included_candidate["page"]),
        )

    internet_provider_parts: List[dict] = []
    for candidate_index, candidate in enumerate(candidates):
        if not any(
            token in candidate["lowered"]
            for token in ("internet", "wifi", "provider", "verizon", "xfinity", "astound", "spectrum", "honest")
        ):
            continue
        part = network_provider_parts_from_candidate(candidate)
        if not part:
            continue
        if not candidate_has_allowed_domain(candidate, {"internet"}) and not (part.get("fixed") or part.get("extras")):
            continue
        internet_provider_parts.append(extend_network_provider_part_with_neighbors(part, candidates, candidate_index))
    if internet_provider_parts:
        fixed_evidence: Dict[str, List[dict]] = {}
        fixed_notes: Dict[str, List[str]] = {}
        extra_notes: List[dict] = []
        for item in internet_provider_parts:
            for label in item["fixed"]:
                segment = item.get("segments", {}).get(label) or item["quote"]
                fixed_evidence.setdefault(label, []).append({"page": item["page"], "quote": segment})
                note = actionable_provider_note_from_segment(segment, label)
                if note:
                    fixed_notes.setdefault(label, []).append(note)
            for label in item["extras"]:
                segment = item.get("segments", {}).get(label) or item["quote"]
                extra_notes.append({"page": item["page"], "quote": segment})

        fixed_label_to_field = {label: field_key for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items()}
        fixed_label_to_note_field = {label: field_key for field_key, label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items()}
        for label, evidence_items in fixed_evidence.items():
            field_key = fixed_label_to_field.get(label)
            if not field_key:
                continue
            merge_field_payload(
                result,
                field_key,
                {
                    "value": "true",
                    "confidence": 0.94,
                    "evidence": evidence_items[0]["quote"],
                    "evidence_items": evidence_items[:4],
                    "manual_review_reason": "",
                    "review_flags": [],
                },
            )
            note_field_key = fixed_label_to_note_field.get(label)
            note_values = [note for note in dict.fromkeys(fixed_notes.get(label, [])) if note and note != label]
            if note_field_key and note_values:
                merge_field_payload(
                    result,
                    note_field_key,
                    {
                        "value": "\n".join(note_values[:4]),
                        "confidence": 0.92,
                        "evidence": evidence_items[0]["quote"],
                        "evidence_items": evidence_items[:4],
                        "manual_review_reason": "",
                        "review_flags": [],
                    },
                )
        if extra_notes:
            merge_field_payload(
                result,
                "internet_notes",
                {
                    "value": "\n".join(dict.fromkeys(item["quote"] for item in extra_notes[:4])),
                    "confidence": 0.88,
                    "evidence": extra_notes[0]["quote"],
                    "evidence_items": extra_notes[:4],
                    "manual_review_reason": "",
                    "review_flags": [],
                },
            )
        self_setup_evidence = None
        for evidence_items in fixed_evidence.values():
            if evidence_items:
                self_setup_evidence = evidence_items[0]
                break
        if self_setup_evidence is None and extra_notes:
            self_setup_evidence = extra_notes[0]
        if self_setup_evidence is None:
            self_setup_evidence = {
                "page": internet_provider_parts[0]["page"],
                "quote": internet_provider_parts[0]["quote"],
            }
        merge_field_payload(
            result,
            "internet_self_setup_required",
            payload_with_evidence(
                value="true",
                quote=self_setup_evidence["quote"],
                page=self_setup_evidence["page"],
            ),
        )

    extra_network_note_items = network_extra_note_items_from_candidates(candidates)
    if extra_network_note_items:
        existing_notes = normalize_unknown_value((result.get("internet_notes") or {}).get("value"))
        note_lines = [*(existing_notes.split("\n") if existing_notes else [])]
        note_lines.extend(item["quote"] for item in extra_network_note_items)
        merge_field_payload(
            result,
            "internet_notes",
            {
                "value": "\n".join(dict.fromkeys(line for line in note_lines if line)),
                "confidence": 0.9,
                "evidence": extra_network_note_items[0]["quote"],
                "evidence_items": extra_network_note_items[:4],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )
        if "internet_self_setup_required" not in result:
            merge_field_payload(
                result,
                "internet_self_setup_required",
                payload_with_evidence(
                    value="true",
                    quote=extra_network_note_items[0]["quote"],
                    page=extra_network_note_items[0]["page"],
                    confidence=0.9,
                ),
            )

    internet_included_candidate = first_candidate(
        candidates,
        lambda item: (
            any(token in item["lowered"] for token in ("internet included", "wifi included", "wi-fi is being provided", "wifi is being provided", "wi-fiis being provided"))
            or ("being provided" in item["lowered"] and ("wifi" in item["lowered"] or "wi-fi" in item["lowered"]))
        )
        and candidate_has_allowed_domain(item, {"internet"}),
    )
    if internet_included_candidate:
        if "included" in internet_included_candidate["lowered"] or "provided" in internet_included_candidate["lowered"]:
            merge_field_payload(
                result,
                "internet_self_setup_required",
                payload_with_evidence(
                    value="false",
                    quote=internet_included_candidate["quote"],
                    page=internet_included_candidate["page"],
                ),
            )

    key_pickup_candidates = all_candidates(
        candidates,
        lambda item: is_key_pickup_candidate(item) and candidate_has_allowed_domain(item, {"key_pickup"}),
    )
    if key_pickup_candidates:
        merge_field_payload(
            result,
            "key_pickup_notes",
            {
                "value": "\n".join(item["quote"] for item in key_pickup_candidates[:4]),
                "confidence": 0.9,
                "evidence": key_pickup_candidates[0]["quote"],
                "evidence_items": [
                    {"page": item["page"], "quote": item["quote"]}
                    for item in key_pickup_candidates[:4]
                ],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )

    service_elevator_candidates = all_candidates(
        candidates,
        lambda item: is_service_elevator_candidate(item)
        and candidate_has_allowed_domain(item, {"move_in", "mover_coi", "contact"}),
    )
    if service_elevator_candidates:
        merge_field_payload(
            result,
            "service_elevator_booking_notes",
            {
                "value": "\n".join(item["quote"] for item in service_elevator_candidates[:3]),
                "confidence": 0.86,
                "evidence": service_elevator_candidates[0]["quote"],
                "evidence_items": [
                    {"page": item["page"], "quote": item["quote"]}
                    for item in service_elevator_candidates[:3]
                ],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )

    move_in_candidates = all_candidates(
        candidates,
        lambda item: any(
            token in item["lowered"]
            for token in ("move-in", "move in", "move-ins", "delays", "lease start date", "apartment walkthrough")
        )
        and "contents welcome contacts" not in item["lowered"]
        and candidate_has_allowed_domain(item, {"move_in", "mover_coi", "key_pickup"}),
    )
    if move_in_candidates:
        merge_field_payload(
            result,
            "move_in_notes",
            {
                "value": "\n".join(item["quote"] for item in move_in_candidates[:3]),
                "confidence": 0.82,
                "evidence": move_in_candidates[0]["quote"],
                "evidence_items": [
                    {"page": item["page"], "quote": item["quote"]}
                    for item in move_in_candidates[:3]
                ],
                "manual_review_reason": "",
                "review_flags": [],
            },
        )

    renters_sentence = first_candidate(
        candidates,
        lambda item: any(
            token in item["lowered"]
            for token in ("renters insurance", "renter's insurance", "renter’s insurance", "renter insurance")
        )
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    renters_amount_sentence = first_candidate(
        candidates,
        lambda item: any(
            token in item["lowered"]
            for token in ("renters insurance", "renter's insurance", "renter’s insurance", "renter insurance")
        )
        and bool(amounts_in_text(item["quote"]))
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    generic_insurance_sentence = first_candidate(
        candidates,
        lambda item: (
            "insurance" in item["lowered"]
            and "coi" not in item["lowered"]
            and "certificate of insurance" not in item["lowered"]
            and "certificate of liability insurance" not in item["lowered"]
            and not candidate_has_mover_coi_context(item, mover_coi_pages)
            and candidate_has_allowed_domain(item, {"renters_insurance"})
        ),
    )
    if renters_sentence:
        if sentence_has_optional_language(renters_sentence["quote"]) and not sentence_has_required_language(renters_sentence["quote"]):
            renters_status = "optional"
        elif sentence_has_negative_language(renters_sentence["quote"]) and not sentence_has_required_language(renters_sentence["quote"]):
            renters_status = "no"
        elif sentence_has_required_language(renters_sentence["quote"]):
            renters_status = "yes"
        else:
            renters_status = "manual_review"
        merge_field_payload(
            result,
            "insurance_renters_required",
            payload_with_evidence(
                value=renters_status,
                quote=renters_sentence["quote"],
                page=renters_sentence["page"],
                manual_review_reason="The source mentions renters insurance, but the requirement is not stated strongly enough." if renters_status == "manual_review" else "",
                review_flags=["ambiguous_requirement"] if renters_status == "manual_review" else [],
            ),
        )
        renters_amount_source = renters_amount_sentence or renters_sentence
        renters_amounts = amounts_in_text(renters_amount_source["quote"])
        if renters_amounts:
            merge_field_payload(
                result,
                "insurance_renters_minimum_coverage",
                payload_with_evidence(
                    value=renters_amounts[0],
                    quote=renters_amount_source["quote"],
                    page=renters_amount_source["page"],
                    manual_review_reason=(
                        "Only general coverage was identified; Property and Liability could not be separated."
                        if "property and liability" in renters_amount_source["lowered"]
                        else ""
                    ),
                    review_flags=(["ambiguous_property_liability_split"] if "property and liability" in renters_amount_source["lowered"] else []),
                ),
            )
    elif generic_insurance_sentence and sentence_has_optional_language(generic_insurance_sentence["quote"]):
        merge_field_payload(
            result,
            "insurance_renters_required",
            payload_with_evidence(
                value="optional",
                quote=generic_insurance_sentence["quote"],
                page=generic_insurance_sentence["page"],
            ),
        )

    property_sentence = first_candidate(
        candidates,
        lambda item: any(token in item["lowered"] for token in ("personal property", "property coverage", "personal belongings"))
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if property_sentence:
        merge_field_payload(
            result,
            "insurance_personal_property_required",
            payload_with_evidence(value="yes", quote=property_sentence["quote"], page=property_sentence["page"]),
        )
        property_amounts = amounts_in_text(property_sentence["quote"])
        if property_amounts:
            merge_field_payload(
                result,
                "insurance_personal_property_minimum",
                payload_with_evidence(value=property_amounts[0], quote=property_sentence["quote"], page=property_sentence["page"]),
            )

    liability_candidates = all_candidates(
        candidates,
        lambda item: (
            "personal liability" in item["lowered"]
            or (
                quote_has_renters_insurance_context(item["quote"])
                and any(token in item["lowered"] for token in ("liability coverage", "bodily injury", "property damage", "aggregate liability", "per occurrence"))
            )
        )
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if liability_candidates:
        merge_field_payload(
            result,
            "insurance_personal_liability_required",
            payload_with_evidence(value="yes", quote=liability_candidates[0]["quote"], page=liability_candidates[0]["page"]),
        )
        for item in liability_candidates:
            amounts = amounts_in_text(item["quote"])
            if not amounts:
                continue
            if "per occurrence" in item["lowered"] or "occurrence" in item["lowered"]:
                merge_field_payload(
                    result,
                    "insurance_personal_liability_per_occurrence",
                    payload_with_evidence(value=amounts[0], quote=item["quote"], page=item["page"]),
                )
            if "aggregate" in item["lowered"]:
                merge_field_payload(
                    result,
                    "insurance_personal_liability_aggregate",
                    payload_with_evidence(value=amounts[-1], quote=item["quote"], page=item["page"]),
                )
            elif "liability" in item["lowered"]:
                merge_field_payload(
                    result,
                    "insurance_personal_liability_per_occurrence",
                    payload_with_evidence(value=amounts[0], quote=item["quote"], page=item["page"]),
                )

    coi_candidates = all_candidates(
        candidates,
        lambda item: (
            "coi" in item["lowered"]
            or "certificate of insurance" in item["lowered"]
            or "certificate of liability insurance" in item["lowered"]
        )
        and candidate_has_allowed_domain(item, {"mover_coi", "renters_insurance", "move_in"}),
    )
    coi_candidate = next((item for item in coi_candidates if candidate_has_mover_coi_context(item, mover_coi_pages)), None)
    if coi_candidate and "one (1) week" not in coi_candidate["lowered"]:
        detailed_mover_coi = next(
            (
                item
                for item in coi_candidates
                if candidate_has_mover_coi_context(item, mover_coi_pages)
                and ("one (1) week" in item["lowered"] or "moving company" in item["lowered"])
            ),
            None,
        )
        if detailed_mover_coi:
            coi_candidate = detailed_mover_coi
    if not coi_candidate and coi_candidates:
        coi_candidate = coi_candidates[0]
    if coi_candidate:
        merge_field_payload(
            result,
            "insurance_coi_required",
            payload_with_evidence(
                value="yes" if sentence_has_required_language(coi_candidate["quote"]) else "manual_review",
                quote=coi_candidate["quote"],
                page=coi_candidate["page"],
                manual_review_reason="" if sentence_has_required_language(coi_candidate["quote"]) else "The source mentions a COI, but a person must confirm whether it is mandatory.",
                review_flags=[] if sentence_has_required_language(coi_candidate["quote"]) else ["ambiguous_coi_requirement"],
            ),
        )
        trigger_match = re.search(r"(?:if|when|for)\s+(.+)", coi_candidate["quote"], flags=re.IGNORECASE)
        merge_field_payload(
            result,
            "insurance_coi_trigger",
            payload_with_evidence(
                value=trigger_match.group(1).strip(" .") if trigger_match else coi_trigger_value_from_quote(coi_candidate["quote"]),
                quote=coi_candidate["quote"],
                page=coi_candidate["page"],
            ),
        )

    interested_party_candidate = first_candidate(
        candidates,
        lambda item: ("interested party" in item["lowered"] or "additional interest" in item["lowered"])
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    interested_party_info_candidate = first_candidate(
        candidates,
        lambda item: "interested party information" in item["lowered"]
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if interested_party_candidate:
        merge_field_payload(
            result,
            "insurance_interested_party_required",
            payload_with_evidence(value="yes", quote=interested_party_candidate["quote"], page=interested_party_candidate["page"]),
        )
        recipient_source_candidate = interested_party_info_candidate or interested_party_candidate
        interested_party_quote = expanded_interested_party_quote(candidates, recipient_source_candidate)
        interested_party_recipient = extract_interested_party_recipient_value(interested_party_quote)
        if interested_party_recipient:
            merge_field_payload(
                result,
                "insurance_recipient",
                payload_with_evidence(
                    value=interested_party_recipient,
                    quote=interested_party_quote,
                    page=recipient_source_candidate["page"],
                    confidence=0.86,
                ),
            )
    ambiguous_added_party_candidate = first_candidate(
        candidates,
        lambda item: any(token in item["lowered"] for token in ("landlord", "property manager", "management"))
        and "add" in item["lowered"]
        and "insured" not in item["lowered"]
        and "interest" not in item["lowered"]
        and "holder" not in item["lowered"]
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if ambiguous_added_party_candidate and not interested_party_candidate:
        merge_field_payload(
            result,
            "insurance_interested_party_required",
            payload_with_evidence(
                value="manual_review",
                quote=ambiguous_added_party_candidate["quote"],
                page=ambiguous_added_party_candidate["page"],
                manual_review_reason="The source requires the property or landlord to be added to the policy but does not specify Interested Party, Additional Interest, or another type.",
                review_flags=["ambiguous_party_type"],
            ),
        )

    additional_insured_candidate = first_candidate(
        candidates,
        lambda item: "additional insured" in item["lowered"]
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if additional_insured_candidate:
        merge_field_payload(
            result,
            "insurance_additional_insured_required",
            payload_with_evidence(value="yes", quote=additional_insured_candidate["quote"], page=additional_insured_candidate["page"]),
        )

    certificate_holder_candidate = first_candidate(
        candidates,
        lambda item: "certificate holder" in item["lowered"]
        and not candidate_has_mover_coi_context(item, mover_coi_pages)
        and candidate_has_allowed_domain(item, {"renters_insurance"}),
    )
    if certificate_holder_candidate:
        merge_field_payload(
            result,
            "insurance_certificate_holder_required",
            payload_with_evidence(value="yes", quote=certificate_holder_candidate["quote"], page=certificate_holder_candidate["page"]),
        )

    submission_candidates = all_candidates(
        candidates,
        lambda item: (
            any(token in item["lowered"] for token in ("upload", "submit", "portal", "buildinglink", "rello", "the guarantors", "proof of insurance", "proof of renters insurance"))
            or ("insurance policy" in item["lowered"] and ("submit" in item["lowered"] or "submitted" in item["lowered"]))
        )
        and any(token in item["lowered"] for token in ("insurance", "guarantors", "proof of insurance", "proof of renters insurance", "insurance policy")),
    )
    submission_candidates = [
        item
        for item in submission_candidates
        if (
            quote_has_renters_insurance_context(item["quote"])
            or "proof of insurance" in item["lowered"]
            or not candidate_has_mover_coi_context(item, mover_coi_pages)
        )
        and candidate_has_allowed_domain(item, {"renters_insurance"})
    ]
    if submission_candidates:
        methods: List[str] = []
        recipients: List[str] = []
        penalty_quotes: List[str] = []
        recipient_sources: List[dict] = []
        for item in submission_candidates:
            lowered = item["lowered"]
            if "buildinglink" in lowered:
                methods.append("BuildingLink")
            if "rello" in lowered:
                methods.append("Rello")
            if "the guarantors" in lowered:
                methods.append("The Guarantors")
            if "portal" in lowered:
                methods.append("portal")
            email_matches = re.findall(r"[\w.\-+]+@[\w.\-]+\.\w+", item["quote"])
            recipients.extend(email_matches)
            if email_matches:
                recipient_sources.append(item)
            if any(token in lowered for token in ("protection program", "automatically be enrolled", "fee of $", "monthly fee")):
                penalty_quotes.append(item["quote"])
        if methods:
            merge_field_payload(
                result,
                "insurance_submission_method",
                {
                    "value": ", ".join(dict.fromkeys(methods)),
                    "confidence": 0.9,
                    "evidence": submission_candidates[0]["quote"],
                    "evidence_items": [{"page": item["page"], "quote": item["quote"]} for item in submission_candidates[:4]],
                    "manual_review_reason": "",
                    "review_flags": [],
                },
            )
        if recipients:
            recipient_source = recipient_sources[0] if recipient_sources else submission_candidates[0]
            merge_field_payload(
                result,
                "insurance_recipient",
                {
                    "value": ", ".join(dict.fromkeys(recipients)),
                    "confidence": 0.85,
                    "evidence": recipient_source["quote"],
                    "evidence_items": [{"page": item["page"], "quote": item["quote"]} for item in recipient_sources[:4]]
                    or [{"page": recipient_source["page"], "quote": recipient_source["quote"]}],
                    "manual_review_reason": "",
                    "review_flags": [],
                },
            )
        if penalty_quotes:
            merge_field_payload(
                result,
                "insurance_alternative_program_or_penalty",
                {
                    "value": "\n".join(dict.fromkeys(penalty_quotes)),
                    "confidence": 0.92,
                    "evidence": penalty_quotes[0],
                    "evidence_items": [{"page": item["page"], "quote": item["quote"]} for item in submission_candidates if item["quote"] in penalty_quotes][:3],
                    "manual_review_reason": "",
                    "review_flags": [],
                },
            )

    if (
        result.get("insurance_renters_minimum_coverage")
        and result.get("insurance_personal_liability_per_occurrence")
        and normalize_money_text(result["insurance_renters_minimum_coverage"]["value"])
        == normalize_money_text(result["insurance_personal_liability_per_occurrence"]["value"])
    ):
        result["insurance_personal_liability_per_occurrence"]["manual_review_reason"] = (
            result["insurance_personal_liability_per_occurrence"].get("manual_review_reason") or ""
        )

    return result


def parse_insurance_coverage_from_text(text: str) -> Optional[str]:
    cleaned = text or ""
    patterns = [
        r"(?:coverage|coverage amount|liability|insured for|insurance amount|保额)[^$\d]{0,20}(\$?\s?\d[\d,]*(?:\.\d{2})?)",
        r"(\$?\s?\d[\d,]*(?:\.\d{2})?)\s*(?:coverage|liability|insured)",
    ]
    for pattern in patterns:
        match = re.search(pattern, cleaned, flags=re.IGNORECASE)
        if not match:
            continue
        value = normalize_unknown_value(match.group(1).replace(" ", ""))
        if value:
            return value
    return None


def infer_internet_self_setup_required(text: str) -> Optional[str]:
    lowered = (text or "").lower()
    false_markers = (
        "internet included",
        "included internet",
        "wifi included",
        "网络已包含",
        "楼内已包含网络",
        "大楼自带网络",
    )
    true_markers = (
        "arrange internet",
        "arrange internet / cable service",
        "internet / cable service",
        "available providers include",
        "contact verizon",
        "contact spectrum",
        "contact astound",
        "contact honest",
        "please arrange internet",
        "set up internet",
        "open your internet account",
        "honest networks",
        "verizon",
        "spectrum",
        "astound",
    )
    if any(marker in lowered for marker in false_markers):
        return "false"
    if any(marker in lowered for marker in true_markers):
        return "true"
    return None


def sanitize_welcome_letter_fields(parsed: Dict[str, dict], text: str, source_file: str) -> Dict[str, dict]:
    cleaned = dict(parsed)
    provider_names = extract_provider_names(text)
    provider_text = ", ".join(provider_names) if provider_names else None
    parsed_building_name = parse_building_name_from_text(text)
    parsed_address = parse_address_from_text(text)
    bad_address_tokens = ("welcome", "excited", "attached", "community", "please find")

    if parsed_building_name and not normalize_unknown_value((cleaned.get("building_name") or {}).get("value")):
        cleaned["building_name"] = {
            "value": parsed_building_name,
            "confidence": 0.9,
            "evidence": parsed_building_name,
        }

    current_address = normalize_unknown_value((cleaned.get("address") or {}).get("value"))
    if current_address:
        lowered = current_address.lower()
        if any(token in lowered for token in bad_address_tokens):
            current_address = None
    if parsed_address and current_address != parsed_address:
        cleaned["address"] = {
            "value": parsed_address,
            "confidence": max(float((cleaned.get("address") or {}).get("confidence") or 0), 0.9),
            "evidence": parsed_address,
        }
    elif current_address is None:
        cleaned.pop("address", None)

    insurance_required = normalize_requirement_choice((cleaned.get("insurance_required") or {}).get("value"))
    coverage_value = parse_insurance_coverage_from_text(text)
    if insurance_required != "true":
        cleaned.pop("insurance_coverage_amount", None)
    elif coverage_value:
        cleaned["insurance_coverage_amount"] = {
            "value": coverage_value,
            "confidence": max(float((cleaned.get("insurance_coverage_amount") or {}).get("confidence") or 0), 0.8),
            "evidence": coverage_value,
        }
    else:
        cleaned.pop("insurance_coverage_amount", None)

    if provider_text:
        cleaned["internet_provider"] = {
            "value": provider_text,
            "confidence": max(float((cleaned.get("internet_provider") or {}).get("confidence") or 0), 0.9),
            "evidence": provider_text,
        }

    inferred_self_setup = infer_internet_self_setup_required(text)
    if inferred_self_setup:
        cleaned["internet_self_setup_required"] = {
            "value": inferred_self_setup,
            "confidence": max(
                float((cleaned.get("internet_self_setup_required") or {}).get("confidence") or 0),
                0.85,
            ),
            "evidence": "internet instructions",
        }

    building_name = normalize_unknown_value((cleaned.get("building_name") or {}).get("value"))
    if not building_name:
        fallback_name = normalize_unknown_value(source_file.rsplit(".", 1)[0]) or parsed_building_name or parsed_address
        if fallback_name:
            cleaned["building_name"] = {
                "value": fallback_name,
                "confidence": 0.5,
                "evidence": source_file,
            }

    return sanitize_address_payloads(cleaned)


def heuristic_welcome_letter_parse(text: str) -> Dict[str, dict]:
    parsed: Dict[str, dict] = {}
    cleaned = text or ""
    lowered = cleaned.lower()
    building_name = parse_building_name_from_text(cleaned)
    if building_name:
        parsed["building_name"] = {
            "value": building_name,
            "confidence": 0.9,
            "evidence": building_name,
        }

    if "renters insurance" in lowered or "insurance" in lowered:
        if any(
            token in lowered
            for token in ("optional", "recommended but not required", "strongly recommended but not required", "not required but recommended")
        ):
            parsed["insurance_required"] = {
                "value": "optional",
                "confidence": 0.98,
                "evidence": "optional / recommended but not required",
            }
        elif "not required" in lowered:
            parsed["insurance_required"] = {
                "value": "false",
                "confidence": 0.85,
                "evidence": "not required",
            }
        elif any(token in lowered for token in ("insurance required", "must provide insurance", "proof of insurance")):
            parsed["insurance_required"] = {
                "value": "true",
                "confidence": 0.85,
                "evidence": "insurance required",
            }
        coverage_value = parse_insurance_coverage_from_text(cleaned)
        if coverage_value:
            parsed["insurance_coverage_amount"] = {
                "value": coverage_value,
                "confidence": 0.8,
                "evidence": coverage_value,
            }
    if any(token in lowered for token in ("con edison", "coned", "pseg", "pse&g", "electricity")):
        parsed["electricity_required"] = {
            "value": "true",
            "confidence": 0.9,
            "evidence": "electricity",
        }
        if "con edison" in lowered or "coned" in lowered:
            parsed["electricity_provider"] = {
                "value": "Con Edison",
                "confidence": 0.95,
                "evidence": "Con Edison",
            }
        if "pseg" in lowered or "pse&g" in lowered:
            parsed["electricity_provider"] = {
                "value": "PSE&G",
                "confidence": 0.8,
                "evidence": "PSE&G",
            }
    if any(token in lowered for token in ("internet", "wifi", "verizon", "xfinity", "astound", "spectrum", "honest")):
        self_setup = infer_internet_self_setup_required(cleaned)
        if self_setup:
            parsed["internet_self_setup_required"] = {
                "value": self_setup,
                "confidence": 0.85,
                "evidence": "internet instructions",
            }
        provider_names = extract_provider_names(cleaned)
        if provider_names:
            provider_text = ", ".join(provider_names)
            parsed["internet_provider"] = {
                "value": provider_text,
                "confidence": 0.9,
                "evidence": provider_text,
            }
    address = parse_address_from_text(cleaned)
    if address:
        parsed["address"] = {"value": address, "confidence": 0.6, "evidence": address}
    return parsed


async def ai_extract_fields_from_text(text: str, source_file: str) -> Dict[str, dict]:
    client = optional_ai_client()
    if not client:
        return {}
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {"role": "system", "content": WELCOME_LETTER_FIELD_PROMPT},
                {
                    "role": "user",
                    "content": json_dumps({"source_file": source_file, "text": text[:12000]}),
                },
            ],
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        data = json.loads(content)
        values = data.get("values", {})
        if not isinstance(values, dict):
            return {}
        result = {}
        for field_key, payload in values.items():
            if field_key not in AI_WRITABLE_FIELD_KEYS or not isinstance(payload, dict):
                continue
            value = normalize_unknown_value(payload.get("value"))
            if value is None:
                continue
            result[field_key] = {
                "value": value,
                "confidence": float(payload.get("confidence") or 0),
                "evidence": payload.get("evidence") or "",
            }
        return result
    except Exception:
        return {}


async def ai_classify_evidence_domains(candidates: List[dict], source_file: str) -> List[dict]:
    classified = classify_evidence_candidates_by_rules(candidates)
    client = optional_ai_client()
    if not client or not classified:
        return classified

    ambiguous_candidates = [
        candidate
        for candidate in classified
        if candidate.get("domain") == "unknown"
        or (
            is_insurance_candidate(candidate)
            and candidate.get("domain") not in {"renters_insurance", "mover_coi"}
        )
    ]
    if not ambiguous_candidates:
        return classified

    candidate_segments = evidence_classification_segments(ambiguous_candidates, max_items=90)
    segment_by_id = {
        segment["id"]: ambiguous_candidates[index]
        for index, segment in enumerate(candidate_segments)
        if index < len(ambiguous_candidates)
    }
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {"role": "system", "content": EVIDENCE_CLASSIFICATION_PROMPT},
                {
                    "role": "user",
                    "content": json_dumps(
                        {
                            "source_file": source_file,
                            "candidate_segments": candidate_segments,
                        }
                    ),
                },
            ],
            response_format={"type": "json_object"},
        )
        data = json.loads(response.choices[0].message.content or "{}")
        items = data.get("items", [])
        if not isinstance(items, list):
            return classified
        marker_updates: Dict[tuple[int, str], dict] = {}
        for item in items:
            if not isinstance(item, dict):
                continue
            candidate = segment_by_id.get(str(item.get("id") or ""))
            if not candidate:
                continue
            domain = normalize_unknown_value(item.get("domain"))
            if domain not in EVIDENCE_DOMAINS:
                continue
            try:
                confidence = float(item.get("confidence") or candidate.get("domain_confidence") or 0.6)
            except (TypeError, ValueError):
                confidence = float(candidate.get("domain_confidence") or 0.6)
            reason = normalize_unknown_value(item.get("reason")) or "AI evidence classification"
            marker_updates[(int(candidate.get("page") or 1), candidate.get("normalized") or "")] = {
                "domain": domain,
                "domain_confidence": max(0.0, min(confidence, 1.0)),
                "domain_reason": reason,
            }
        if not marker_updates:
            return classified
        merged: List[dict] = []
        for candidate in classified:
            update = marker_updates.get((int(candidate.get("page") or 1), candidate.get("normalized") or ""))
            if update:
                next_candidate = dict(candidate)
                next_candidate.update(update)
                merged.append(next_candidate)
            else:
                merged.append(candidate)
        return merged
    except Exception:
        return classified


async def ai_build_business_summary(candidates: List[dict], source_file: str) -> Dict[str, List[dict]]:
    rule_summary = build_rule_business_summary(candidates)
    client = optional_ai_client()
    segments = business_summary_segments(candidates, max_items=120)
    if not client or not segments:
        return rule_summary
    segment_by_id = {segment["id"]: segment for segment in segments}
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {"role": "system", "content": BUSINESS_SUMMARY_PROMPT},
                {
                    "role": "user",
                    "content": json_dumps(
                        {
                            "source_file": source_file,
                            "candidate_segments": segments,
                        }
                    ),
                },
            ],
            response_format={"type": "json_object"},
        )
        data = json.loads(response.choices[0].message.content or "{}")
        ai_summary = sanitize_business_summary(data, segment_by_id)
        if any(ai_summary.get(key) for key in BUSINESS_SUMMARY_KEYS):
            return ai_summary
    except Exception:
        pass
    return rule_summary


async def ai_classify_document_and_summarize(
    conn: sqlite3.Connection,
    *,
    source_document_id: str,
    source_file: str,
    raw_input_type: str,
    parser_type: str,
    pages: List[dict],
    candidates: List[dict],
    intake_metadata: dict,
) -> dict:
    rule_summary = build_rule_business_summary(candidates)
    rule_overall_summary = build_rule_overall_summary(rule_summary)
    client = optional_ai_client()
    model = deepseek_model()
    user_payload = {
        "source_file": source_file,
        "raw_input_type": raw_input_type,
        "parser_type": parser_type,
        "intake_metadata": intake_metadata or {},
        "pages": compact_source_pages_for_llm(pages),
    }
    if not client or not user_payload["pages"]:
        return {
            "document_classification": normalize_document_classification(
                {
                    "document_type": "full_welcome_letter",
                    "confidence": 0.0,
                    "reason": "LLM unavailable; used rule fallback summary.",
                }
            ),
            "overall_summary": rule_overall_summary,
            "business_summary": rule_summary,
            "workflow_hints": normalize_workflow_hints({}),
            "validation_warnings": ["llm_analysis_unavailable"],
        }
    thinking_enabled = deepseek_thinking_enabled()
    thinking_effort = deepseek_thinking_reasoning_effort()
    raw_response = ""
    data: dict = {}
    used_thinking = False
    fallback_warnings: List[str] = []
    try:
        messages = [
            {"role": "system", "content": WELCOME_PACKET_ANALYSIS_PROMPT},
            {"role": "user", "content": json_dumps(user_payload)},
        ]
        if thinking_enabled:
            try:
                response = await client.chat.completions.create(
                    model=model,
                    messages=messages,
                    response_format={"type": "json_object"},
                    reasoning_effort=thinking_effort,
                    extra_body={"thinking": {"type": "enabled"}},
                )
                raw_response = normalize_unknown_value(response.choices[0].message.content) or ""
                if not raw_response:
                    raise ValueError("thinking mode returned empty content")
                used_thinking = True
            except Exception as thinking_exc:
                fallback_warnings.append(f"thinking_mode_fallback:{str(thinking_exc)[:160]}")
                raw_response = ""
        if not raw_response:
            response = await client.chat.completions.create(
                model=model,
                temperature=0,
                messages=messages,
                response_format={"type": "json_object"},
            )
            raw_response = normalize_unknown_value(response.choices[0].message.content) or "{}"
        data = json.loads(raw_response)
        overall_summary = normalize_overall_summary(data.get("overall_summary")) or rule_overall_summary
        business_summary, warnings = sanitize_llm_business_summary(data.get("business_summary"), candidates)
        if not any(business_summary.get(key) for key in BUSINESS_SUMMARY_KEYS):
            business_summary = rule_summary
            warnings.append("llm_business_summary_empty_rule_fallback_used")
        artifacts = {
            "document_classification": normalize_document_classification(data.get("document_classification")),
            "overall_summary": overall_summary,
            "business_summary": business_summary,
            "workflow_hints": normalize_workflow_hints(data.get("workflow_hints")),
            "validation_warnings": [
                *(data.get("validation_warnings") if isinstance(data.get("validation_warnings"), list) else []),
                *fallback_warnings,
                *warnings,
            ],
            "llm_options": {
                "thinking_requested": thinking_enabled,
                "thinking_enabled": used_thinking,
                "reasoning_effort": thinking_effort if used_thinking else "",
            },
        }
        write_llm_call_log(
            conn,
            source_document_id=source_document_id,
            stage="document_classification_and_business_summary",
            model=model,
            system_prompt=WELCOME_PACKET_ANALYSIS_PROMPT,
            user_payload={
                **user_payload,
                "_llm_options": {
                    "thinking_requested": thinking_enabled,
                    "thinking_enabled": used_thinking,
                    "reasoning_effort": thinking_effort if used_thinking else "",
                },
            },
            raw_response=raw_response,
            parsed_response=artifacts,
        )
        return artifacts
    except Exception as exc:
        write_llm_call_log(
            conn,
            source_document_id=source_document_id,
            stage="document_classification_and_business_summary",
            model=model,
            system_prompt=WELCOME_PACKET_ANALYSIS_PROMPT,
            user_payload=user_payload,
            error=str(exc),
        )
        return {
            "document_classification": normalize_document_classification(
                {
                    "document_type": "unknown",
                    "confidence": 0.0,
                    "reason": "LLM analysis failed; used rule fallback summary.",
                }
            ),
            "overall_summary": rule_overall_summary,
            "business_summary": rule_summary,
            "workflow_hints": normalize_workflow_hints({}),
            "validation_warnings": [f"llm_analysis_error:{str(exc)[:160]}"],
        }


async def ai_map_summary_to_fields(
    conn: sqlite3.Connection,
    *,
    source_document_id: str,
    source_file: str,
    overall_summary: str,
    business_summary: Dict[str, List[dict]],
    document_classification: dict,
    workflow_hints: dict,
    candidates: List[dict],
) -> tuple[Dict[str, dict], List[str], dict]:
    client = optional_ai_client()
    model = deepseek_model()
    summary_items = flatten_business_summary_items(business_summary)
    user_payload = {
        "source_file": source_file,
        "document_classification": document_classification or {},
        "overall_summary": overall_summary or "",
        "business_summary_items": summary_items,
        "workflow_hints": workflow_hints or {},
        "allowed_fields": llm_field_options_for_mapping(conn),
    }
    if not client or not summary_items:
        return {}, ["llm_field_mapping_unavailable"], {}
    try:
        response = await client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[
                {"role": "system", "content": WELCOME_PACKET_FIELD_MAPPING_PROMPT},
                {"role": "user", "content": json_dumps(user_payload)},
            ],
            response_format={"type": "json_object"},
        )
        raw_response = response.choices[0].message.content or "{}"
        data = json.loads(raw_response)
        summary_section_by_id = {
            normalize_unknown_value(item.get("id")) or "": normalize_unknown_value(item.get("section")) or ""
            for item in summary_items
            if normalize_unknown_value(item.get("id"))
        }
        values, warnings = sanitize_llm_field_mapping_values(data.get("values"), candidates, summary_section_by_id)
        if isinstance(data.get("validation_warnings"), list):
            warnings = [*data.get("validation_warnings"), *warnings]
        parsed = {"values": values, "validation_warnings": warnings}
        write_llm_call_log(
            conn,
            source_document_id=source_document_id,
            stage="business_summary_to_field_mapping",
            model=model,
            system_prompt=WELCOME_PACKET_FIELD_MAPPING_PROMPT,
            user_payload=user_payload,
            raw_response=raw_response,
            parsed_response=parsed,
        )
        return values, warnings, parsed
    except Exception as exc:
        write_llm_call_log(
            conn,
            source_document_id=source_document_id,
            stage="business_summary_to_field_mapping",
            model=model,
            system_prompt=WELCOME_PACKET_FIELD_MAPPING_PROMPT,
            user_payload=user_payload,
            error=str(exc),
        )
        return {}, [f"llm_field_mapping_error:{str(exc)[:160]}"], {}


async def ai_extract_move_in_fields_from_candidates(
    candidates: List[dict],
    source_file: str,
    business_summary: Optional[Dict[str, List[dict]]] = None,
) -> Dict[str, dict]:
    client = optional_ai_client()
    if not client:
        return {}
    candidate_segments = move_in_candidate_segments(candidates)
    if not candidate_segments:
        return {}
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {"role": "system", "content": MOVE_IN_FIELD_PROMPT},
                {
                    "role": "user",
                    "content": json_dumps(
                        {
                            "source_file": source_file,
                            "business_summary": business_summary or {},
                            "candidate_segments": candidate_segments,
                        }
                    ),
                },
            ],
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        data = json.loads(content)
        values = data.get("values", {})
        if not isinstance(values, dict):
            return {}
        result: Dict[str, dict] = {}
        for field_key, payload in values.items():
            if field_key not in MOVE_IN_LLM_FIELD_KEYS or not isinstance(payload, dict):
                continue
            value = normalize_unknown_value(payload.get("value"))
            if value is None:
                continue
            evidence_items = []
            for item in payload.get("evidence_items") or []:
                if not isinstance(item, dict):
                    continue
                quote = normalize_unknown_value(item.get("quote"))
                if not quote:
                    continue
                evidence_items.append({"page": int(item.get("page") or 1), "quote": quote})
            result[field_key] = {
                "value": value,
                "confidence": float(payload.get("confidence") or 0.82),
                "evidence": evidence_items[0]["quote"] if evidence_items else payload.get("evidence") or "",
                "evidence_items": evidence_items,
                "manual_review_reason": payload.get("manual_review_reason") or "",
                "review_flags": [],
            }
        return sanitize_move_in_payloads(result, candidates)
    except Exception:
        return {}


async def ai_extract_domain_fields_from_candidates(
    *,
    candidates: List[dict],
    source_file: str,
    prompt: str,
    field_keys: set[str],
    candidate_segments: List[dict],
    business_summary: Optional[Dict[str, List[dict]]] = None,
) -> Dict[str, dict]:
    client = optional_ai_client()
    if not client or not candidate_segments:
        return {}
    try:
        response = await client.chat.completions.create(
            model=deepseek_model(),
            temperature=0,
            messages=[
                {"role": "system", "content": prompt},
                {
                    "role": "user",
                    "content": json_dumps(
                        {
                            "source_file": source_file,
                            "business_summary": business_summary or {},
                            "candidate_segments": candidate_segments,
                        }
                    ),
                },
            ],
            response_format={"type": "json_object"},
        )
        content = response.choices[0].message.content or "{}"
        data = json.loads(content)
        values = data.get("values", {})
        if not isinstance(values, dict):
            return {}
        result: Dict[str, dict] = {}
        for field_key, payload in values.items():
            if field_key not in field_keys or not isinstance(payload, dict):
                continue
            value = normalize_unknown_value(payload.get("value"))
            if value is None:
                continue
            evidence_items = []
            for item in payload.get("evidence_items") or []:
                if not isinstance(item, dict):
                    continue
                quote = normalize_unknown_value(item.get("quote"))
                if not quote:
                    continue
                evidence_items.append({"page": int(item.get("page") or 1), "quote": quote})
            result[field_key] = {
                "value": value,
                "confidence": float(payload.get("confidence") or 0.8),
                "evidence": evidence_items[0]["quote"] if evidence_items else payload.get("evidence") or "",
                "evidence_items": evidence_items,
                "manual_review_reason": payload.get("manual_review_reason") or "",
                "review_flags": [],
            }
        return result
    except Exception:
        return {}


async def ai_extract_insurance_fields_from_candidates(
    candidates: List[dict],
    source_file: str,
    business_summary: Optional[Dict[str, List[dict]]] = None,
) -> Dict[str, dict]:
    insurance_domains = {"renters_insurance", "mover_coi", "move_in"}
    candidate_segments = domain_candidate_segments(
        candidates,
        lambda item: is_insurance_candidate(item) and candidate_has_allowed_domain(item, insurance_domains),
        max_items=48,
    )
    raw_values = await ai_extract_domain_fields_from_candidates(
        candidates=candidates,
        source_file=source_file,
        prompt=INSURANCE_FIELD_PROMPT,
        field_keys=INSURANCE_LLM_FIELD_KEYS,
        candidate_segments=candidate_segments,
        business_summary=business_summary,
    )
    return sanitize_insurance_payloads(raw_values, candidates)


async def ai_extract_electricity_fields_from_candidates(
    candidates: List[dict],
    source_file: str,
    business_summary: Optional[Dict[str, List[dict]]] = None,
) -> Dict[str, dict]:
    candidate_segments = domain_candidate_segments(
        candidates,
        lambda item: is_electricity_candidate(item) and candidate_has_allowed_domain(item, {"electricity"}),
        max_items=30,
    )
    raw_values = await ai_extract_domain_fields_from_candidates(
        candidates=candidates,
        source_file=source_file,
        prompt=ELECTRICITY_FIELD_PROMPT,
        field_keys=ELECTRICITY_LLM_FIELD_KEYS,
        candidate_segments=candidate_segments,
        business_summary=business_summary,
    )
    return sanitize_electricity_payloads(raw_values, candidates)


VISION_LAST_ERROR: Optional[str] = None


def resolved_vision_model_name() -> str:
    model = os.getenv("VISION_MODEL", "").strip()
    if not model:
        return ""
    base_url = (os.getenv("VISION_BASE_URL", "") or "").strip().lower()
    if "xiaomimimo.com" in base_url:
        return model.lower()
    return model


def vision_client() -> Optional[AsyncOpenAI]:
    global VISION_LAST_ERROR
    api_key = os.getenv("VISION_API_KEY", "").strip()
    model = resolved_vision_model_name()
    if not api_key or not model:
        return None
    base_url = os.getenv("VISION_BASE_URL", "").strip() or None
    VISION_LAST_ERROR = None
    return AsyncOpenAI(api_key=api_key, base_url=base_url, timeout=VISION_TIMEOUT_SECONDS)


async def ai_extract_fields_from_image(
    path: Path,
    source_file: str,
    *,
    page_number: Optional[int] = None,
) -> Dict[str, dict]:
    global VISION_LAST_ERROR
    client = vision_client()
    model = resolved_vision_model_name()
    if not client or not model or not path.is_file():
        return {}
    mime = "image/png"
    suffix = path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        mime = "image/jpeg"
    elif suffix == ".webp":
        mime = "image/webp"
    if path.stat().st_size > MAX_UPLOAD_BYTES:
        return {}
    image_b64 = base64.b64encode(path.read_bytes()).decode("utf-8")
    try:
        response = await client.chat.completions.create(
            model=model,
            temperature=0,
            messages=[
                {"role": "system", "content": WELCOME_LETTER_FIELD_PROMPT},
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": json_dumps({"source_file": source_file})},
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime};base64,{image_b64}"},
                        },
                    ],
                },
            ],
            response_format={"type": "json_object"},
        )
        data = json.loads(response.choices[0].message.content or "{}")
        values = data.get("values", {})
        result = {}
        for field_key, payload in values.items():
            if field_key not in AI_WRITABLE_FIELD_KEYS or not isinstance(payload, dict):
                continue
            value = normalize_unknown_value(payload.get("value"))
            if value is None:
                continue
            evidence = normalize_unknown_value(payload.get("evidence")) or ""
            evidence_items = []
            if evidence and page_number:
                evidence_items.append({"page": page_number, "quote": evidence})
            result[field_key] = {
                "value": value,
                "confidence": float(payload.get("confidence") or 0.65),
                "evidence": evidence,
                "evidence_items": evidence_items,
            }
        VISION_LAST_ERROR = None
        return result
    except Exception as exc:
        VISION_LAST_ERROR = str(exc)
        return {}


async def ai_extract_fields_from_pdf_images(
    path: Path,
    source_file: str,
    *,
    max_pages: int = 6,
) -> Dict[str, dict]:
    if not vision_client() or not path.is_file():
        return {}
    with tempfile.TemporaryDirectory(prefix="whitepaper_pdf_vision_") as temp_dir:
        rendered_pages = render_pdf_pages_to_images(path, max_pages=max_pages, output_dir=Path(temp_dir))
        if not rendered_pages:
            return {}
        merged: Dict[str, dict] = {}
        extractor = configured_vision_field_extractor()
        for page_number, image_path in rendered_pages:
            payloads = await extractor.extract_fields(
                image_path,
                f"{source_file}#page-{page_number or 1}",
                page_number=page_number,
            )
            merged = merge_parsed_payloads(merged, payloads)
        return merged


async def prepare_pdf_source_payload(
    path: Path,
    source_file: str,
    *,
    source_document_id: str = "",
    input_index: int = 1,
) -> tuple[List[dict], str, str, Dict[str, dict]]:
    extracted_pages = extract_pdf_pages(path)
    extracted_text = join_page_texts(extracted_pages)
    parser_parts = ["welcome_letter_pdf"]

    if pdf_text_needs_visual_fallback(extracted_text, extracted_pages):
        ocr_result = await extract_with_ocr_router(
            path,
            source_file=source_file,
            source_document_id=source_document_id,
            input_index=input_index,
        )
        ocr_pages = [
            {"page": page.page_number, "text": page.text, "blocks": page.blocks}
            for page in ocr_result.pages
            if normalize_unknown_value(page.text)
        ]
        ocr_text = join_page_texts(ocr_pages)
        if ocr_text:
            extracted_pages = ocr_pages
            extracted_text = ocr_text
            parser_parts.append(f"ocr_{ocr_result.provider}")

    vision_values: Dict[str, dict] = {}
    if pdf_text_needs_visual_fallback(extracted_text, extracted_pages):
        vision_values = await ai_extract_fields_from_pdf_images(path, source_file)
        if vision_values:
            parser_parts.append("vision")

    return extracted_pages, extracted_text, "_".join(parser_parts), vision_values


def offset_payload_evidence_pages(payload: dict, page_offset: int) -> dict:
    if not page_offset or not isinstance(payload, dict):
        return payload
    cloned = dict(payload)
    evidence_items = []
    for item in cloned.get("evidence_items") or []:
        if not isinstance(item, dict):
            continue
        next_item = dict(item)
        try:
            next_item["page"] = int(next_item.get("page") or 1) + page_offset
        except (TypeError, ValueError):
            next_item["page"] = page_offset + 1
        evidence_items.append(next_item)
    if evidence_items:
        cloned["evidence_items"] = evidence_items
    if cloned.get("page") is not None:
        try:
            cloned["page"] = int(cloned.get("page") or 1) + page_offset
        except (TypeError, ValueError):
            cloned["page"] = page_offset + 1
    return cloned


def offset_payload_set_evidence_pages(payloads: Dict[str, dict], page_offset: int) -> Dict[str, dict]:
    return {
        field_key: offset_payload_evidence_pages(payload, page_offset)
        for field_key, payload in (payloads or {}).items()
    }


def summarize_pdf_source_files(files: List[UploadFile], stored_paths: List[Path]) -> str:
    names: List[str] = []
    for index, upload in enumerate(files):
        name = normalize_unknown_value(upload.filename or "")
        if not name and index < len(stored_paths):
            name = stored_paths[index].name
        if name:
            names.append(name)
    if not names:
        return f"PDF document package ({len(stored_paths)} files)"
    if len(names) == 1:
        return names[0]
    if len(names) <= 4:
        return " / ".join(names)
    return f"{names[0]} and {len(names) - 1} more PDF files"


async def prepare_pdf_package_source_payload(
    pdf_paths: List[Path],
    source_file: str,
    *,
    source_document_id: str = "",
) -> tuple[List[dict], str, str, Dict[str, dict]]:
    extracted_pages: List[dict] = []
    vision_values: Dict[str, dict] = {}
    parser_parts = ["welcome_letter_pdf"]
    if len(pdf_paths) > 1:
        parser_parts.append("multi")

    for index, pdf_path in enumerate(pdf_paths, start=1):
        page_offset = len(extracted_pages)
        pages, _text, parser_type, pdf_vision_values = await prepare_pdf_source_payload(
            pdf_path,
            f"{source_file}#pdf-{index}",
            source_document_id=source_document_id,
            input_index=index,
        )
        if "ocr" in parser_type and "ocr" not in parser_parts:
            parser_parts.append("ocr")
        if "vision" in parser_type and "vision" not in parser_parts:
            parser_parts.append("vision")
        for page in pages:
            page_text = normalize_unknown_value(page.get("text"))
            if not page_text:
                continue
            extracted_pages.append(
                {
                    **page,
                    "page": len(extracted_pages) + 1,
                    "source_file": pdf_path.name,
                    "source_page": page.get("page") or len(extracted_pages) + 1,
                    "text": page_text,
                }
            )
        vision_values = merge_parsed_payloads(
            vision_values,
            offset_payload_set_evidence_pages(pdf_vision_values, page_offset),
        )

    return extracted_pages, join_page_texts(extracted_pages), "_".join(parser_parts), vision_values


def summarize_image_source_files(files: List[UploadFile], stored_paths: List[Path]) -> str:
    names: List[str] = []
    for index, upload in enumerate(files):
        name = normalize_unknown_value(upload.filename or "")
        if not name and index < len(stored_paths):
            name = stored_paths[index].name
        if name:
            names.append(name)
    if not names:
        return f"Image group ({len(stored_paths)} images)"
    if len(names) == 1:
        return names[0]
    return f"{names[0]} and {len(names) - 1} more images"


async def prepare_image_source_payload(
    image_paths: List[Path],
    source_file: str,
    *,
    source_document_id: str = "",
) -> tuple[List[dict], str, str, Dict[str, dict]]:
    extracted_pages: List[dict] = []
    vision_values: Dict[str, dict] = {}
    ocr_providers: List[str] = []
    vision_extractor = configured_vision_field_extractor()
    for index, image_path in enumerate(image_paths, start=1):
        ocr_result = await extract_with_ocr_router(
            image_path,
            source_file=f"{source_file}#image-{index}",
            source_document_id=source_document_id,
            input_index=index,
            page_number=index,
        )
        ocr_text = ocr_result.document_text
        ocr_providers.append(ocr_result.provider)
        normalized_ocr = normalize_unknown_value(ocr_text)
        if normalized_ocr:
            extracted_pages.append(
                {
                    "page": index,
                    "text": normalized_ocr,
                    "blocks": ocr_result.pages[0].blocks if ocr_result.pages else [],
                }
            )
        image_vision_values = await vision_extractor.extract_fields(
            image_path,
            f"{source_file}#image-{index}",
            page_number=index,
        )
        vision_values = merge_parsed_payloads(vision_values, image_vision_values)
    parser_parts = ["welcome_letter_image"]
    if len(image_paths) > 1:
        parser_parts.append("multi")
    if ocr_providers:
        parser_parts.append("ocr_" + "_".join(dict.fromkeys(ocr_providers)))
    if vision_values:
        parser_parts.append("vision")
    return extracted_pages, join_page_texts(extracted_pages), "_".join(parser_parts), vision_values


def enrich_ai_payload_with_evidence(
    ai_values: Dict[str, dict],
    candidates: List[dict],
    *,
    allowed_fields: Optional[set[str]] = None,
) -> Dict[str, dict]:
    enriched: Dict[str, dict] = {}
    for field_key, payload in (ai_values or {}).items():
        if allowed_fields and field_key not in allowed_fields:
            continue
        evidence_items = payload_evidence_items(payload, candidates)
        manual_review_reason = ""
        review_flags: List[str] = []
        if not evidence_items:
            manual_review_reason = "No directly supporting source sentence was found."
            review_flags.append("no_direct_evidence")
        review_flags.extend(evidence_domain_review_flags(evidence_items))
        enriched[field_key] = {
            "value": payload.get("value"),
            "confidence": float(payload.get("confidence") or 0),
            "evidence": evidence_items[0]["quote"] if evidence_items else (payload.get("evidence") or ""),
            "evidence_items": evidence_items,
            "manual_review_reason": manual_review_reason,
            "review_flags": review_flags,
        }
    return enriched


def derive_legacy_insurance_fields_from_detailed(parsed: Dict[str, dict]) -> Dict[str, dict]:
    derived: Dict[str, dict] = {}
    renters_status = normalize_insurance_status((parsed.get("insurance_renters_required") or {}).get("value"))
    if renters_status in {"yes", "no", "optional"}:
        mapped = {"yes": "true", "no": "false", "optional": "optional"}[renters_status]
        payload = parsed.get("insurance_renters_required") or {}
        derived["insurance_required"] = {
            "value": mapped,
            "confidence": float(payload.get("confidence") or 0.9),
            "evidence": payload.get("evidence") or "",
            "evidence_items": payload.get("evidence_items") or [],
            "manual_review_reason": payload.get("manual_review_reason") or "",
            "review_flags": payload.get("review_flags") or [],
        }

    renters_coverage_payload = parsed.get("insurance_renters_minimum_coverage") or {}
    renters_coverage = normalize_money_text(renters_coverage_payload.get("value"))
    if renters_coverage:
        derived["insurance_coverage_amount"] = {
            "value": renters_coverage,
            "confidence": float(renters_coverage_payload.get("confidence") or 0.9),
            "evidence": renters_coverage_payload.get("evidence") or "",
            "evidence_items": renters_coverage_payload.get("evidence_items") or [],
            "manual_review_reason": renters_coverage_payload.get("manual_review_reason") or "",
            "review_flags": renters_coverage_payload.get("review_flags") or [],
        }
    return derived


def merge_parsed_payloads(*payload_sets: Dict[str, dict]) -> Dict[str, dict]:
    merged: Dict[str, dict] = {}
    for payload_set in payload_sets:
        for field_key, payload in (payload_set or {}).items():
            merge_field_payload(merged, field_key, payload)
    return merged


def extract_text_from_image_with_tesseract(path: Path) -> str:
    if shutil.which("tesseract") is None or not path.is_file():
        return ""
    for language in ("eng+chi_sim", "eng"):
        try:
            result = subprocess.run(
                ["tesseract", str(path), "stdout", "-l", language],
                check=False,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=OCR_TIMEOUT_SECONDS,
            )
        except Exception:
            continue
        if result.returncode == 0 and result.stdout.strip():
            return result.stdout.strip()
    return ""


def extract_text_from_image_with_apple_vision(path: Path) -> str:
    if sys.platform != "darwin" or shutil.which("swift") is None or not path.is_file():
        return ""
    try:
        result = subprocess.run(
            ["swift", "-", str(path)],
            input=APPLE_VISION_OCR_SWIFT,
            check=False,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=APPLE_VISION_OCR_TIMEOUT_SECONDS,
        )
    except Exception:
        return ""
    if result.returncode != 0:
        return ""
    return result.stdout.strip()


def extract_text_from_image_locally(path: Path) -> str:
    for extractor in (extract_text_from_image_with_tesseract, extract_text_from_image_with_apple_vision):
        text = normalize_unknown_value(extractor(path))
        if text:
            return text
    return ""


def extract_text_from_pdf_via_ocr(path: Path) -> str:
    return join_page_texts(extract_text_from_pdf_via_ocr_pages(path))


OCR_ROUTER_INSTANCE: Optional[OcrRouter] = None
VISION_FIELD_EXTRACTOR_INSTANCE: Optional[CallableVisionFieldExtractor] = None


def configured_ocr_router() -> OcrRouter:
    global OCR_ROUTER_INSTANCE
    if OCR_ROUTER_INSTANCE is None:
        local_provider = LocalOcrProvider(
            image_extractor=extract_text_from_image_locally,
            pdf_extractor=extract_text_from_pdf_via_ocr_pages,
        )
        baidu_provider = BaiduUnlimitedCloudOcrProvider(
            api_key=BAIDU_OCR_API_KEY,
            secret_key=BAIDU_OCR_SECRET_KEY,
            base_url=BAIDU_OCR_BASE_URL,
            poll_interval_seconds=BAIDU_OCR_POLL_INTERVAL_SECONDS,
            timeout_seconds=BAIDU_OCR_TIMEOUT_SECONDS,
        )
        unlimited_ocr_local_provider = UnlimitedOcrLocalHttpProvider(
            base_url=UNLIMITED_OCR_LOCAL_BASE_URL,
            model=UNLIMITED_OCR_LOCAL_MODEL,
            api_key=UNLIMITED_OCR_LOCAL_API_KEY,
            timeout_seconds=UNLIMITED_OCR_LOCAL_TIMEOUT_SECONDS,
            max_tokens=UNLIMITED_OCR_LOCAL_MAX_TOKENS,
            max_pdf_pages=MAX_PDF_PARSE_PAGES,
            pdf_renderer=render_pdf_pages_to_images,
        )
        OCR_ROUTER_INSTANCE = OcrRouter(
            providers={
                "local": local_provider,
                "baidu_unlimited_cloud": baidu_provider,
                "unlimited_ocr_local_http": unlimited_ocr_local_provider,
            },
            primary=OCR_PROVIDER,
            fallback=OCR_FALLBACK_PROVIDER,
        )
    return OCR_ROUTER_INSTANCE


def configured_vision_field_extractor() -> CallableVisionFieldExtractor:
    global VISION_FIELD_EXTRACTOR_INSTANCE
    if VISION_FIELD_EXTRACTOR_INSTANCE is None:
        VISION_FIELD_EXTRACTOR_INSTANCE = CallableVisionFieldExtractor(
            "xiaomi_mimo",
            ai_extract_fields_from_image,
        )
    return VISION_FIELD_EXTRACTOR_INSTANCE


def ocr_artifact_dir(source_document_id: str) -> Optional[Path]:
    if not source_document_id:
        return None
    return UPLOAD_ROOT / "ocr_artifacts" / source_document_id


def resumable_ocr_task_id(source_document_id: str, input_index: int, provider: str) -> str:
    if not source_document_id or provider != "baidu_unlimited_cloud":
        return ""
    with db_connection() as conn:
        row = conn.execute(
            """
            SELECT external_task_id
            FROM ocr_jobs
            WHERE source_document_id = ? AND input_index = ?
              AND provider IN ('baidu_unlimited_cloud', 'baidu_unlimited')
              AND status IN ('pending', 'processing', 'queued')
            ORDER BY CASE provider WHEN 'baidu_unlimited_cloud' THEN 0 ELSE 1 END
            LIMIT 1
            """,
            (source_document_id, input_index),
        ).fetchone()
    return normalize_unknown_value(row.get("external_task_id") if row else "") or ""


def update_ocr_job(
    source_document_id: str,
    input_index: int,
    provider: str,
    payload: dict,
) -> None:
    if not source_document_id:
        return
    timestamp = now_iso()
    status = normalize_unknown_value(payload.get("status")) or "processing"
    external_task_id = normalize_unknown_value(payload.get("external_task_id")) or ""
    duration_ms = int(payload.get("duration_ms") or 0)
    result_artifact_path = normalize_unknown_value(payload.get("result_artifact_path")) or ""
    error_code = normalize_unknown_value(payload.get("error_code")) or ""
    error_message = re.sub(
        r"((?:access_token|client_id|client_secret|secret_key|api_key)=)[^&\s]+",
        r"\1=[redacted]",
        normalize_unknown_value(payload.get("error_message")) or "",
        flags=re.IGNORECASE,
    )[:1000]
    attempt_increment = max(0, int(payload.get("attempt_increment") or 0))
    completed_at = timestamp if status in {"completed", "success", "failed", "empty"} else None
    next_poll_at = (
        (datetime.now(timezone.utc) + timedelta(seconds=BAIDU_OCR_POLL_INTERVAL_SECONDS)).isoformat()
        if status in {"pending", "processing", "queued"}
        else None
    )
    metadata = {
        key: payload.get(key)
        for key in ("warnings", "fallback_used", "raw_artifact_sha256")
        if payload.get(key) not in (None, "", [])
    }
    with db_connection() as conn:
        existing = conn.execute(
            """
            SELECT id, attempt_count FROM ocr_jobs
            WHERE source_document_id = ? AND input_index = ? AND provider = ?
            """,
            (source_document_id, input_index, provider),
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE ocr_jobs
                SET external_task_id = COALESCE(NULLIF(?, ''), external_task_id), status = ?,
                    attempt_count = attempt_count + ?, next_poll_at = ?, completed_at = ?,
                    duration_ms = ?, result_artifact_path = ?, error_code = ?, error_message = ?,
                    metadata_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    external_task_id,
                    status,
                    attempt_increment,
                    next_poll_at,
                    completed_at,
                    duration_ms,
                    result_artifact_path,
                    error_code,
                    error_message,
                    json_dumps(metadata),
                    timestamp,
                    existing["id"],
                ),
            )
        else:
            conn.execute(
                """
                INSERT INTO ocr_jobs(
                  id, source_document_id, input_index, provider, external_task_id, status,
                  attempt_count, submitted_at, next_poll_at, completed_at, duration_ms,
                  result_artifact_path, error_code, error_message, metadata_json, created_at, updated_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"ocr_{uuid.uuid4().hex}",
                    source_document_id,
                    input_index,
                    provider,
                    external_task_id,
                    status,
                    attempt_increment,
                    timestamp,
                    next_poll_at,
                    completed_at,
                    duration_ms,
                    result_artifact_path,
                    error_code,
                    error_message,
                    json_dumps(metadata),
                    timestamp,
                    timestamp,
                ),
            )


def record_ocr_result(source_document_id: str, input_index: int, result: OcrResult) -> None:
    artifact_hash = ""
    artifact_path = Path(result.raw_artifact_path) if result.raw_artifact_path else None
    if artifact_path and artifact_path.is_file():
        artifact_hash = hashlib.sha256(artifact_path.read_bytes()).hexdigest()
    update_ocr_job(
        source_document_id,
        input_index,
        result.provider,
        {
            "status": "completed" if result.ok else result.status,
            "external_task_id": result.external_task_ids[0] if result.external_task_ids else "",
            "duration_ms": result.duration_ms,
            "result_artifact_path": result.raw_artifact_path,
            "error_code": result.error_code,
            "error_message": result.error_message,
            "warnings": result.warnings,
            "fallback_used": result.fallback_used,
            "raw_artifact_sha256": artifact_hash,
        },
    )
    if not source_document_id:
        return
    with db_connection() as conn:
        row = conn.execute(
            "SELECT parse_artifacts_json FROM source_documents WHERE id = ?",
            (source_document_id,),
        ).fetchone()
        artifacts = json_loads_safe(row.get("parse_artifacts_json") if row else "{}", {})
        if not isinstance(artifacts, dict):
            artifacts = {}
        ocr_items = artifacts.get("ocr") if isinstance(artifacts.get("ocr"), list) else []
        public_item = {"input_index": input_index, **result.public_metadata()}
        ocr_items = [item for item in ocr_items if int(item.get("input_index") or 0) != input_index]
        ocr_items.append(public_item)
        artifacts["ocr"] = sorted(ocr_items, key=lambda item: int(item.get("input_index") or 0))
        conn.execute(
            "UPDATE source_documents SET parse_artifacts_json = ?, updated_at = ? WHERE id = ?",
            (json_dumps(artifacts), now_iso(), source_document_id),
        )


def ocr_status_hook(source_document_id: str, input_index: int, provider: str) -> Callable[[dict], None]:
    def hook(payload: dict) -> None:
        update_ocr_job(source_document_id, input_index, provider, payload)

    return hook


async def extract_with_ocr_router(
    path: Path,
    *,
    source_file: str,
    source_document_id: str,
    input_index: int,
    page_number: int = 1,
) -> OcrResult:
    router = configured_ocr_router()
    if (
        router.primary == "baidu_unlimited_cloud"
        and path.suffix.lower() != ".pdf"
        and path.is_file()
        and path.stat().st_size > 10 * 1024 * 1024
    ):
        local_provider = router.providers["local"]
        result = await local_provider.extract(
            path,
            source_file=source_file,
            page_number=page_number,
            artifact_dir=ocr_artifact_dir(source_document_id),
        )
        result.fallback_used = True
        result.warnings.append("baidu_skipped_image_size_limit")
        record_ocr_result(source_document_id, input_index, result)
        return result
    if router.primary == "baidu_unlimited_cloud" and path.suffix.lower() == ".pdf":
        try:
            page_count = len(PdfReader(str(path)).pages)
        except Exception:
            page_count = 0
        if page_count > MAX_PDF_PARSE_PAGES:
            local_provider = router.providers["local"]
            result = await local_provider.extract(
                path,
                source_file=source_file,
                page_number=page_number,
                artifact_dir=ocr_artifact_dir(source_document_id),
            )
            result.fallback_used = True
            result.warnings.append(
                f"baidu_skipped_page_limit:{page_count}>{MAX_PDF_PARSE_PAGES}"
            )
            record_ocr_result(source_document_id, input_index, result)
            return result
    resume_task_id = resumable_ocr_task_id(source_document_id, input_index, router.primary)
    result = await router.extract(
        path,
        source_file=source_file,
        page_number=page_number,
        artifact_dir=ocr_artifact_dir(source_document_id),
        status_hook=ocr_status_hook(source_document_id, input_index, router.primary),
        resume_task_id=resume_task_id,
    )
    record_ocr_result(source_document_id, input_index, result)
    return result


async def parse_source_to_staging(
    conn: sqlite3.Connection,
    *,
    source_document_id: str,
    source_file: str,
    raw_input_type: str,
    parser_type: str,
    source_type: str,
    source_content: str,
    extracted_text: str,
    extracted_pages: Optional[List[dict]],
    actor: Actor,
    request: Request,
    vision_values: Optional[Dict[str, dict]] = None,
    intake_mode: str = INTAKE_MODE_FULL_PACKAGE,
    supplement_scope: str = SUPPLEMENT_SCOPE_ALL,
    target_staging_key: str = "",
) -> dict:
    normalized_intake_mode = normalize_intake_mode(intake_mode)
    normalized_supplement_scope = normalize_supplement_scope(supplement_scope)
    normalized_target_staging_key = normalize_unknown_value(target_staging_key) or ""
    stored_source_metadata = parse_intake_source_metadata(source_content)
    source_metadata = build_intake_source_metadata(
        intake_mode=normalized_intake_mode,
        supplement_scope=normalized_supplement_scope,
        target_staging_key=normalized_target_staging_key,
        source_kind=stored_source_metadata.get("source_kind", ""),
        case_id=stored_source_metadata.get("case_id", ""),
        communication_event_id=stored_source_metadata.get("communication_event_id", ""),
        captured_at=stored_source_metadata.get("captured_at", ""),
    )
    source_artifact_row = conn.execute(
        "SELECT parse_artifacts_json FROM source_documents WHERE id = ?",
        (source_document_id,),
    ).fetchone()
    existing_source_artifacts = json_loads_safe(
        source_artifact_row.get("parse_artifacts_json") if source_artifact_row else "{}",
        {},
    )
    existing_ocr_artifacts = (
        existing_source_artifacts.get("ocr", [])
        if isinstance(existing_source_artifacts, dict)
        else []
    )
    row_source_content = json_dumps(source_metadata) if source_metadata else (source_content or extracted_text)

    source_pages = [dict(item) for item in (extracted_pages or []) if normalize_unknown_value(item.get("text"))]
    if not source_pages:
        source_pages = pages_from_text(extracted_text or source_content)
    flattened_text = join_page_texts(source_pages) or extracted_text or source_content
    candidates = await ai_classify_evidence_domains(build_sentence_candidates(source_pages), source_file)
    analysis_artifacts = await ai_classify_document_and_summarize(
        conn,
        source_document_id=source_document_id,
        source_file=source_file,
        raw_input_type=raw_input_type,
        parser_type=parser_type,
        pages=source_pages,
        candidates=candidates,
        intake_metadata=source_metadata,
    )
    document_classification = analysis_artifacts.get("document_classification") or {}
    overall_summary = normalize_overall_summary(analysis_artifacts.get("overall_summary"))
    business_summary = analysis_artifacts.get("business_summary") or build_rule_business_summary(candidates)
    workflow_hints = analysis_artifacts.get("workflow_hints") or {}
    validation_warnings = list(analysis_artifacts.get("validation_warnings") or [])
    domain_counts: Dict[str, int] = {}
    for candidate in candidates:
        domain = normalize_unknown_value(candidate.get("domain")) or "unknown"
        domain_counts[domain] = domain_counts.get(domain, 0) + 1
    ai_detected_supplement = (
        document_classification.get("document_type") in LLM_SUPPLEMENT_DOCUMENT_TYPES
        and normalized_intake_mode != INTAKE_MODE_SUPPLEMENT
    )
    if ai_detected_supplement and not normalized_target_staging_key:
        suggested_target = document_classification.get("suggested_target_building") or {}
        notice_building_name = (
            normalize_unknown_value(suggested_target.get("building_name"))
            or source_file.rsplit(".", 1)[0]
        )
        notice_metadata = build_intake_source_metadata(
            intake_mode=INTAKE_MODE_SUPPLEMENT,
            supplement_scope=SUPPLEMENT_SCOPE_ALL,
            target_staging_key="",
        )
        notice_value = "AI classified this as supplemental material. Link it to a target Staging building and parse it again."
        artifacts = {
            "ocr": existing_ocr_artifacts,
            "document_classification": document_classification,
            "overall_summary": overall_summary,
            "business_summary": business_summary,
            "field_mapping": {},
            "workflow_hints": workflow_hints,
            "validation_warnings": [
                *validation_warnings,
                "ai_detected_supplement_requires_target_building",
            ],
            "evidence_domain_counts": domain_counts,
        }
        conn.execute(
            "UPDATE source_documents SET parse_artifacts_json = ?, updated_at = ? WHERE id = ?",
            (json_dumps(artifacts), now_iso(), source_document_id),
        )
        create_staging_request(
            conn,
            submission_group_id=f"source:{source_document_id}",
            building_name=notice_building_name,
            building_id=None,
            field_name="__parser_notice__",
            old_value=None,
            new_value=notice_value,
            source_type=source_type,
            source_content=json_dumps(notice_metadata),
            source_file=source_file,
            submitted_by=actor.user_id,
            ai_confidence=float(document_classification.get("confidence") or 0),
            review_status="needs_more_info",
            import_batch_id=None,
            parser_type=parser_type,
            raw_input_type=raw_input_type,
            source_document_id=source_document_id,
            approval_stage=APPROVAL_STAGE_TO_STAGING,
            target_staging_key="",
            conflict_with_long_term=False,
            low_confidence=True,
            missing_required_detail=False,
            evidence_json=[],
            manual_review_reason=document_classification.get("reason") or notice_value,
            review_flags_json=[
                "ai_detected_supplement",
                f"document_type:{document_classification.get('document_type') or 'unknown'}",
                *[
                    f"affected_domain:{domain}"
                    for domain in (document_classification.get("affected_domains") or [])
                ],
            ],
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="source_parsed_to_review_queue",
            target_table="staging_update_requests",
            target_record_id=f"source:{source_document_id}",
            building_name=notice_building_name,
            source=source_file,
            note=f"raw_input_type={raw_input_type}, parser_type={parser_type}, ai_detected_supplement=1",
        )
        return {
            "submission_group_id": f"source:{source_document_id}",
            "building_name": notice_building_name,
            "building_id": None,
            "staging_key": None,
            "intake_mode": INTAKE_MODE_SUPPLEMENT,
            "supplement_scope": SUPPLEMENT_SCOPE_ALL,
            "target_staging_key": "",
            "staging_synced": False,
            "message": notice_value,
        }

    effective_intake_mode = normalized_intake_mode
    effective_supplement_scope = normalized_supplement_scope
    effective_target_staging_key = normalized_target_staging_key
    if ai_detected_supplement and normalized_target_staging_key:
        effective_intake_mode = INTAKE_MODE_SUPPLEMENT
        effective_supplement_scope = normalized_supplement_scope or SUPPLEMENT_SCOPE_ALL
    effective_row_source_content = row_source_content
    if effective_intake_mode == INTAKE_MODE_SUPPLEMENT and not source_metadata:
        effective_row_source_content = json_dumps(
            build_intake_source_metadata(
                intake_mode=effective_intake_mode,
                supplement_scope=effective_supplement_scope,
                target_staging_key=effective_target_staging_key,
            )
        )
    conn.execute(
        "UPDATE source_documents SET parse_artifacts_json = ?, updated_at = ? WHERE id = ?",
        (
            json_dumps(
                {
                    "ocr": existing_ocr_artifacts,
                    "document_classification": document_classification,
                    "overall_summary": overall_summary,
                    "business_summary": business_summary,
                    "field_mapping": {},
                    "workflow_hints": workflow_hints,
                    "validation_warnings": validation_warnings,
                    "evidence_domain_counts": domain_counts,
                }
            ),
            now_iso(),
            source_document_id,
        ),
    )
    conn.commit()

    heuristic_raw_values = parse_welcome_letter_fields_from_pages(
        raw_input_type=raw_input_type,
        source_file=source_file,
        pages=source_pages,
        classified_candidates=candidates,
    )
    heuristic_values = merge_parsed_payloads(
        {
            field_key: payload
            for field_key, payload in heuristic_raw_values.items()
            if field_key not in INSURANCE_LLM_FIELD_KEYS and field_key not in ELECTRICITY_LLM_FIELD_KEYS
            and field_key not in MOVE_IN_LLM_FIELD_KEYS
        },
        sanitize_insurance_payloads(heuristic_raw_values, candidates),
        sanitize_electricity_payloads(heuristic_raw_values, candidates),
        sanitize_move_in_payloads(heuristic_raw_values, candidates),
    )
    llm_mapping_values, mapping_warnings, field_mapping_artifact = await ai_map_summary_to_fields(
        conn,
        source_document_id=source_document_id,
        source_file=source_file,
        overall_summary=overall_summary,
        business_summary=business_summary,
        document_classification=document_classification,
        workflow_hints=workflow_hints,
        candidates=candidates,
    )
    validation_warnings.extend(mapping_warnings)
    if llm_mapping_values:
        llm_mapping_values = merge_parsed_payloads(
            {
                field_key: payload
                for field_key, payload in llm_mapping_values.items()
                if field_key not in INSURANCE_LLM_FIELD_KEYS
                and field_key not in ELECTRICITY_LLM_FIELD_KEYS
                and field_key not in MOVE_IN_LLM_FIELD_KEYS
            },
            sanitize_insurance_payloads(llm_mapping_values, candidates),
            sanitize_electricity_payloads(llm_mapping_values, candidates),
            sanitize_move_in_payloads(llm_mapping_values, candidates),
    )
    stable_move_in_values = stable_move_in_payloads_from_business_summary(business_summary, candidates)
    stable_contact_values = stable_contact_payloads_from_business_summary(business_summary, candidates)

    insurance_llm_values: Dict[str, dict] = {}
    electricity_llm_values: Dict[str, dict] = {}
    move_in_llm_values: Dict[str, dict] = {}
    ai_values: Dict[str, dict] = {}
    if not llm_mapping_values:
        insurance_llm_values = await ai_extract_insurance_fields_from_candidates(candidates, source_file, business_summary)
        electricity_llm_values = await ai_extract_electricity_fields_from_candidates(candidates, source_file, business_summary)
        move_in_llm_values = await ai_extract_move_in_fields_from_candidates(candidates, source_file, business_summary)

        ai_raw_values = await ai_extract_fields_from_text(flattened_text, source_file)
        ai_values = enrich_ai_payload_with_evidence(
            ai_raw_values,
            candidates,
            allowed_fields={"building_name", "address", "info_cutoff_date", "source_date", "document_type"},
        )
    vision_enriched = enrich_ai_payload_with_evidence(
        vision_values or {},
        candidates,
        allowed_fields=set(AI_WRITABLE_FIELD_KEYS),
    )
    vision_enriched = merge_parsed_payloads(
        {
            field_key: payload
            for field_key, payload in vision_enriched.items()
            if field_key not in MOVE_IN_LLM_FIELD_KEYS
            and field_key not in INSURANCE_LLM_FIELD_KEYS
            and field_key not in ELECTRICITY_LLM_FIELD_KEYS
        },
        sanitize_insurance_payloads(vision_enriched, candidates),
        sanitize_electricity_payloads(vision_enriched, candidates),
        sanitize_move_in_payloads(vision_enriched, candidates),
    )

    if llm_mapping_values:
        merged = merge_with_primary_preferred(
            llm_mapping_values,
            heuristic_values,
            vision_enriched,
        )
    else:
        merged = merge_parsed_payloads(
            heuristic_values,
            insurance_llm_values,
            electricity_llm_values,
            move_in_llm_values,
            ai_values,
            vision_enriched,
        )
    if stable_move_in_values:
        merged = merge_with_primary_preferred(stable_move_in_values, merged)
    if stable_contact_values:
        merged = merge_with_primary_preferred(stable_contact_values, merged)
    merged = sanitize_address_payloads(merged)
    merged = sanitize_network_payloads(merged)
    merged = merge_parsed_payloads(merged, derive_legacy_insurance_fields_from_detailed(merged))
    if source_type.startswith("chat_"):
        merged = filter_chat_building_payloads(merged)
    artifacts = {
        "ocr": existing_ocr_artifacts,
        "document_classification": document_classification,
        "overall_summary": overall_summary,
        "business_summary": business_summary,
        "field_mapping": field_mapping_artifact or {"values": llm_mapping_values},
        "stable_move_in_mapping": stable_move_in_values,
        "stable_contact_mapping": stable_contact_values,
        "workflow_hints": workflow_hints,
        "validation_warnings": list(dict.fromkeys(str(item) for item in validation_warnings if item)),
        "evidence_domain_counts": domain_counts,
    }
    conn.execute(
        "UPDATE source_documents SET parse_artifacts_json = ?, updated_at = ? WHERE id = ?",
        (json_dumps(artifacts), now_iso(), source_document_id),
    )

    target_staging_snapshot = None
    if effective_intake_mode == INTAKE_MODE_SUPPLEMENT:
        if not effective_target_staging_key:
            raise HTTPException(status_code=400, detail="Supplemental material must be linked to a Staging building first.")
        target_staging_snapshot = load_staging_building_snapshot(conn, effective_target_staging_key)
        if not target_staging_snapshot:
            raise HTTPException(status_code=404, detail="The target Staging building for this supplemental material does not exist.")
        merged = filter_parsed_payloads_for_supplement_scope(merged, effective_supplement_scope)

    building_payload = merged.get("building_name") or {}
    address_payload = merged.get("address") or {}
    if target_staging_snapshot:
        building_name = normalize_unknown_value(target_staging_snapshot.get("building_name")) or source_file.rsplit(".", 1)[0]
        address = normalize_unknown_value(target_staging_snapshot.get("address")) or ""
        allow_existing_match = True
        staging_match = target_staging_snapshot
        target_staging_key = target_staging_snapshot["staging_key"]
    else:
        building_name = normalize_unknown_value(building_payload.get("value")) or source_file.rsplit(".", 1)[0]
        address = normalize_unknown_value((merged.get("address") or {}).get("value")) or ""
        allow_existing_match = building_payload_allows_existing_match(building_payload, address_payload)
        staging_match = (
            resolve_staging_snapshot_for_source_sync(conn, building_name, address)
            if allow_existing_match
            else None
        )
        target_staging_key = staging_match["staging_key"] if staging_match else ""
    building_id = None
    submission_group_id = f"source:{source_document_id}"

    values_by_field = {
        field_key: normalize_field_value(
            field_key,
            (find_field_definition(conn, field_key) or {}).get("field_type", "text"),
            payload.get("value"),
        )
        for field_key, payload in merged.items()
        if field_key in AI_WRITABLE_FIELD_KEYS
    }
    if not values_by_field:
        scope_label = SUPPLEMENT_SCOPE_LABELS.get(effective_supplement_scope, "selected")
        notice_value = (
            f"No importable fields were identified within the {scope_label} scope"
            if effective_intake_mode == INTAKE_MODE_SUPPLEMENT
            else "More information or human input is required"
        )
        create_staging_request(
            conn,
            submission_group_id=submission_group_id,
            building_name=building_name,
            building_id=building_id,
            field_name="__parser_notice__",
            old_value=None,
            new_value=notice_value,
            source_type=source_type,
            source_content=effective_row_source_content,
            source_file=source_file,
            submitted_by=actor.user_id,
            ai_confidence=0.0,
            review_status="needs_more_info",
            import_batch_id=None,
            parser_type=parser_type,
            raw_input_type=raw_input_type,
            source_document_id=source_document_id,
            approval_stage=APPROVAL_STAGE_TO_STAGING,
            target_staging_key=target_staging_key,
            conflict_with_long_term=False,
            low_confidence=True,
            missing_required_detail=False,
            evidence_json=[],
            manual_review_reason=f"{notice_value}; no writable fields were generated.",
            review_flags_json=[
                "no_parsed_fields",
                *(
                    [f"supplement_scope:{effective_supplement_scope}"]
                    if effective_intake_mode == INTAKE_MODE_SUPPLEMENT
                    else []
                ),
            ],
        )
    else:
        for field_key, payload in merged.items():
            if field_key not in AI_WRITABLE_FIELD_KEYS:
                continue
            definition = find_field_definition(conn, field_key)
            normalized_value = normalize_field_value(
                field_key,
                definition["field_type"] if definition else "text",
                payload.get("value"),
            )
            if normalized_value is None:
                continue
            old_value = get_staging_snapshot_field_value(staging_match, field_key)
            confidence = float(payload.get("confidence") or 0)
            evidence_items = payload.get("evidence_items") or []
            manual_reason = (payload.get("manual_review_reason") or "").strip()
            review_flags = list(payload.get("review_flags") or [])
            if effective_intake_mode == INTAKE_MODE_SUPPLEMENT:
                scope_flag = f"supplement_scope:{effective_supplement_scope}"
                if scope_flag not in review_flags:
                    review_flags.append(scope_flag)
            field_missing_detail = missing_required_detail_for_field(field_key, values_by_field)
            if not evidence_items and field_key not in {"source_type", "source_file"}:
                if "no_direct_evidence" not in review_flags:
                    review_flags.append("no_direct_evidence")
                if not manual_reason:
                    manual_reason = "No directly supporting source sentence was found."
            create_staging_request(
                conn,
                submission_group_id=submission_group_id,
                building_name=building_name,
                building_id=building_id,
                field_name=field_key,
                old_value=old_value,
                new_value=normalized_value,
                source_type=source_type,
                source_content=effective_row_source_content,
                source_file=source_file,
                submitted_by=actor.user_id,
                ai_confidence=confidence,
                review_status="ai_parsed" if confidence > 0 else "needs_more_info",
                import_batch_id=None,
                parser_type=parser_type,
                raw_input_type=raw_input_type,
                source_document_id=source_document_id,
                approval_stage=APPROVAL_STAGE_TO_STAGING,
                target_staging_key=target_staging_key,
                conflict_with_long_term=bool(old_value and old_value != normalized_value),
                low_confidence=confidence < 0.65 or bool(manual_reason),
                missing_required_detail=field_missing_detail,
                evidence_json=evidence_items,
                manual_review_reason=manual_reason,
                review_flags_json=review_flags,
            )

    write_audit_log(
        conn,
        request,
        actor,
        action_type="source_parsed_to_review_queue",
        target_table="staging_update_requests",
        target_record_id=submission_group_id,
        building_name=building_name,
        source=source_file,
        note=f"raw_input_type={raw_input_type}, parser_type={parser_type}",
    )
    return {
        "submission_group_id": submission_group_id,
        "building_name": building_name,
        "building_id": building_id,
        "staging_key": staging_match["staging_key"] if staging_match else None,
        "intake_mode": effective_intake_mode,
        "supplement_scope": effective_supplement_scope if effective_intake_mode == INTAKE_MODE_SUPPLEMENT else "",
        "target_staging_key": target_staging_key,
        "staging_synced": False,
        "message": "The parsed result entered the review queue and may be written to Staging after approval.",
    }


def resolve_master_building(conn: sqlite3.Connection, query: str) -> tuple[Optional[dict], List[dict]]:
    ranked = rank_master_buildings(conn, query)
    if not ranked:
        return None, []
    return ranked[0][1], [row for _, row in ranked[:5]]


def load_master_building_snapshot(conn: sqlite3.Connection, building_id: str) -> Optional[dict]:
    row = conn.execute("SELECT * FROM master_building_info WHERE id = ?", (building_id,)).fetchone()
    if not row:
        return None
    ext_rows = conn.execute(
        """
        SELECT field_key, value_text
        FROM master_building_field_values
        WHERE building_id = ? AND status = 'active'
        ORDER BY field_key ASC
        """,
        (building_id,),
    ).fetchall()
    pending_count = conn.execute(
        """
        SELECT COUNT(*) AS total
        FROM staging_update_requests
        WHERE building_id = ? AND review_status IN ('pending', 'ai_parsed', 'employee_submitted', 'needs_more_info', 'conflict')
        """,
        (building_id,),
    ).fetchone()["total"]
    data = dict(row)
    extensions = {item["field_key"]: item["value_text"] for item in ext_rows}
    data["extensions"] = extensions
    data["internet_provider"] = build_network_provider_text_from_values(
        {
            **{field_key: extensions.get(field_key) for field_key in NETWORK_PROVIDER_FIELD_MAP},
            "internet_provider": data.get("internet_provider"),
            "internet_notes": data.get("internet_notes"),
        }
    )
    data["pending_count"] = pending_count
    return data


def build_staging_snapshot_key(building_name: str, address: str = "") -> str:
    identity = f"{legacy._normalize_text(building_name)}|{legacy._normalize_text(address)}"
    return f"staging_{hashlib.sha1(identity.encode('utf-8')).hexdigest()[:16]}"


def resolve_staging_snapshot_for_source_sync(
    conn: sqlite3.Connection,
    building_name: str,
    address: str = "",
) -> Optional[dict]:
    normalized_name = legacy._normalize_text(building_name)
    normalized_address = legacy._normalize_text(address or "")
    if not normalized_name:
        return None
    snapshots = load_staging_building_snapshots(conn)
    exact_match = next(
        (
            snapshot
            for snapshot in snapshots
            if legacy._normalize_text(snapshot.get("building_name")) == normalized_name
            and legacy._normalize_text(snapshot.get("address") or "") == normalized_address
        ),
        None,
    )
    if exact_match:
        return exact_match
    same_name = [
        snapshot
        for snapshot in snapshots
        if legacy._normalize_text(snapshot.get("building_name")) == normalized_name
    ]
    if len(same_name) == 1:
        return same_name[0]

    ranked = rank_staging_buildings(conn, " ".join(part for part in [building_name, address] if part).strip())
    if not ranked:
        return None
    top_score, top_snapshot = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0
    if top_score >= 180 and (not second_score or top_score >= second_score + 25):
        return top_snapshot
    if normalized_address and top_score >= 150 and legacy._normalize_text(top_snapshot.get("building_name")) == normalized_name:
        return top_snapshot
    return None


def master_building_doc(row: dict) -> legacy.BuildingDoc:
    return legacy.BuildingDoc(
        building_id=row["id"],
        building_name=row["building_name"],
        address=row["address"] or "",
        category="building",
        file_name="master.db",
        file_path=DB_PATH,
        relative_path="master",
        source_type="master",
        content="",
    )


def staging_building_doc(snapshot: dict) -> legacy.BuildingDoc:
    return legacy.BuildingDoc(
        building_id=snapshot["id"],
        building_name=snapshot["building_name"],
        address=snapshot.get("address") or "",
        category="staging",
        file_name=snapshot.get("source_file") or "staging",
        file_path=Path("."),
        relative_path="staging",
        source_type="staging",
        content="",
    )


def rank_master_buildings(conn: sqlite3.Connection, query: str) -> List[tuple[int, dict]]:
    rows = conn.execute("SELECT * FROM master_building_info ORDER BY building_name ASC").fetchall()
    if not query.strip():
        return []
    ranked: List[tuple[int, dict]] = []
    for row in rows:
        score = legacy._score_building_match(query, master_building_doc(row))
        if score > 0:
            ranked.append((score, row))
    ranked.sort(key=lambda item: (-item[0], item[1]["building_name"]))
    return ranked


def rank_staging_buildings(conn: sqlite3.Connection, query: str) -> List[tuple[int, dict]]:
    snapshots = load_staging_building_snapshots(conn)
    if not query.strip():
        return []
    ranked: List[tuple[int, dict]] = []
    for snapshot in snapshots:
        score = legacy._score_building_match(query, staging_building_doc(snapshot))
        if score > 0:
            ranked.append((score, snapshot))
    ranked.sort(key=lambda item: (-item[0], item[1]["building_name"]))
    return ranked


def serialize_query_building(record: Optional[dict], source_mode: Literal["master", "staging"]) -> Optional[dict]:
    if not record:
        return None
    return {
        "id": record.get("id"),
        "staging_key": record.get("staging_key") or record.get("id"),
        "building_name": record.get("building_name") or "",
        "address": record.get("address") or "",
        "source_mode": source_mode,
        "completeness_status": record.get("completeness_status"),
        "completeness_score": record.get("completeness_score"),
    }


def detect_query_selection_conflict(
    *,
    query: str,
    source_mode: Literal["master", "staging"],
    current_record: Optional[dict],
    ranked_records: List[tuple[int, dict]],
) -> tuple[Optional[dict], str]:
    if not current_record or not ranked_records:
        return None, ""

    top_score, top_record = ranked_records[0]
    current_key = current_record.get("staging_key") if source_mode == "staging" else current_record.get("id")
    top_key = top_record.get("staging_key") if source_mode == "staging" else top_record.get("id")
    if current_key == top_key:
        return None, ""

    current_score = legacy._score_building_match(
        query,
        staging_building_doc(current_record) if source_mode == "staging" else master_building_doc(current_record),
    )
    second_score = ranked_records[1][0] if len(ranked_records) > 1 else 0
    if top_score < 160:
        return None, ""
    if current_score >= top_score - 15:
        return None, ""
    if second_score and top_score < second_score + 35 and top_score < 220:
        return None, ""

    candidate = serialize_query_building(top_record, source_mode)
    message = (
        f"The selected building is {current_record.get('building_name') or 'the current building'}, "
        f"but the question appears to refer to {top_record.get('building_name') or 'another building'}. "
        "Would you like to switch buildings and run the query again?"
    )
    return candidate, message


def staging_snapshot_record_total(snapshot: dict) -> int:
    total = 0
    if normalize_unknown_value(snapshot.get("building_name")):
        total += 1
    if normalize_unknown_value(snapshot.get("address")):
        total += 1
    for field_key in (
        "insurance_required",
        "insurance_coverage_amount",
        "electricity_required",
        "electricity_provider",
        "internet_self_setup_required",
        "internet_provider",
        "internet_notes",
        "key_pickup_notes",
        "service_elevator_booking_notes",
        "move_in_notes",
    ):
        if normalize_unknown_value(snapshot.get(field_key)) is not None:
            total += 1
    for field_key, value in (snapshot.get("extensions") or {}).items():
        if workbook_extension_value({"extensions": {field_key: value}}, field_key) is not None:
            total += 1
    return total


def load_staging_building_snapshot(conn: sqlite3.Connection, staging_key: str) -> Optional[dict]:
    row = conn.execute(
        "SELECT * FROM staging_building_info WHERE id = ?",
        (staging_key,),
    ).fetchone()
    if not row:
        return None
    ext_rows = conn.execute(
        """
        SELECT field_key, value_text
        FROM staging_building_field_values
        WHERE building_id = ?
        ORDER BY field_key ASC
        """,
        (staging_key,),
    ).fetchall()
    data = dict(row)
    data["id"] = row["id"]
    data["staging_key"] = row["id"]
    data["source_type"] = row.get("source_type") or "staging_excel"
    data["status"] = row.get("status") or "staging"
    data["extensions"] = {item["field_key"]: item["value_text"] for item in ext_rows}
    data["internet_provider"] = build_network_provider_text_from_values(
        {
            **{field_key: data["extensions"].get(field_key) for field_key in NETWORK_PROVIDER_FIELD_MAP},
            "internet_provider": data.get("internet_provider"),
            "internet_notes": data.get("internet_notes"),
        }
    )
    data["pending_count"] = staging_snapshot_record_total(data)
    return data


def load_staging_building_snapshots(conn: sqlite3.Connection) -> List[dict]:
    rows = conn.execute(
        """
        SELECT id
        FROM staging_building_info
        ORDER BY building_name ASC, address ASC, created_at ASC
        """
    ).fetchall()
    snapshots = []
    for row in rows:
        snapshot = load_staging_building_snapshot(conn, row["id"])
        if snapshot:
            snapshots.append(snapshot)
    return snapshots


def resolve_staging_building(conn: sqlite3.Connection, query: str) -> tuple[Optional[dict], List[dict]]:
    ranked = rank_staging_buildings(conn, query)
    if not ranked:
        snapshots = load_staging_building_snapshots(conn)
        return (None, snapshots[:5]) if not query.strip() else (None, [])
    return ranked[0][1], [row for _, row in ranked[:5]]


def load_all_master_building_snapshots(conn: sqlite3.Connection) -> List[dict]:
    rows = conn.execute(
        """
        SELECT id
        FROM master_building_info
        ORDER BY building_name ASC, address ASC, created_at ASC
        """
    ).fetchall()
    snapshots = []
    for row in rows:
        snapshot = load_master_building_snapshot(conn, row["id"])
        if snapshot:
            snapshots.append(snapshot)
    return snapshots


def ensure_master_workbook_from_db(conn: sqlite3.Connection) -> Path:
    return ensure_master_workbook(load_all_master_building_snapshots(conn))


def ensure_staging_workbook_from_sources(conn: sqlite3.Connection) -> Path:
    workbook_path = resolve_staging_excel_path()
    if workbook_path.exists():
        ensure_staging_workbook()
        return workbook_path
    # SQLite may already contain the only surviving staging mirror after a workbook
    # was moved or lost. Recreate the workbook from those rows before considering
    # master data, otherwise startup would silently erase the staging library.
    seed_rows = load_staging_building_snapshots(conn)
    master_workbook_path = resolve_master_excel_path()
    if not seed_rows and master_workbook_path.exists():
        try:
            seed_rows = load_master_workbook_rows(master_workbook_path)
        except Exception:
            seed_rows = []
    if not seed_rows:
        seed_rows = load_all_master_building_snapshots(conn)
    return ensure_staging_workbook(seed_rows)


def sync_master_workbook_for_building(
    conn: sqlite3.Connection,
    building_id: str,
    *,
    previous_identity: Optional[tuple[str, str]] = None,
) -> Path:
    workbook_path = resolve_master_excel_path()
    if not workbook_path.exists():
        ensure_master_workbook_from_db(conn)
    snapshot = load_master_building_snapshot(conn, building_id)
    if not snapshot:
        raise HTTPException(status_code=404, detail="The Master record does not exist, so the master workbook cannot be synchronized.")
    lookup_keys: List[tuple[str, str]] = []
    if previous_identity and (previous_identity[0] or previous_identity[1]):
        lookup_keys.append(previous_identity)
    return upsert_building_snapshot(snapshot, lookup_keys=lookup_keys)


def sync_staging_workbook_for_building(
    snapshot: dict,
    *,
    previous_identity: Optional[tuple[str, str]] = None,
) -> Path:
    lookup_keys: List[tuple[str, str]] = []
    if previous_identity and (previous_identity[0] or previous_identity[1]):
        lookup_keys.append(previous_identity)
    return upsert_staging_snapshot(snapshot, lookup_keys=lookup_keys)


def create_manual_staging_building_snapshot(
    conn: sqlite3.Connection,
    *,
    building_name: str,
    address: str,
    aliases: str,
    notes: str,
    insurance_required: Optional[str],
    electricity_required: Optional[str],
    internet_self_setup_required: Optional[str],
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    normalized_name = normalize_unknown_value(building_name) or ""
    normalized_address = normalize_unknown_value(address) or ""
    if not normalized_name:
        raise HTTPException(status_code=400, detail="A building name is required to add a Staging record.")
    existing = resolve_staging_snapshot_for_source_sync(conn, normalized_name, normalized_address)
    if existing:
        raise HTTPException(status_code=409, detail="A similar building already exists in Staging. Search and confirm before adding another one.")

    extension_field_keys = set(workbook_catalog_extension_field_keys(conn, include_staging_only=True))
    extensions = {field_key: None for field_key in extension_field_keys}
    extra_notes = []
    if normalize_unknown_value(aliases):
        extra_notes.append(f"Aliases/search keywords: {normalize_unknown_value(aliases)}")
    if normalize_unknown_value(notes):
        extra_notes.append(f"Notes: {normalize_unknown_value(notes)}")
    if "manual_notes" in extension_field_keys and extra_notes:
        extensions["manual_notes"] = "\n".join(extra_notes)

    workbook_snapshot = {
        "building_name": normalized_name,
        "address": normalized_address,
        "insurance_required": requirement_choice_to_int(insurance_required),
        "insurance_coverage_amount": None,
        "electricity_required": requirement_choice_to_int(electricity_required),
        "electricity_provider": None,
        "internet_self_setup_required": requirement_choice_to_int(internet_self_setup_required),
        "internet_provider": None,
        "internet_notes": None,
        "move_in_notes": "\n".join(extra_notes) if extra_notes else None,
        "extensions": extensions,
        "source_type": "manual_staging_create",
        "source_file": "manual_staging_create",
        "library_status": STAGING_STATUS_PENDING,
    }
    master_identities = {snapshot_identity(item) for item in load_all_master_building_snapshots(conn)}
    workbook_snapshot["library_status"] = derive_staging_library_status(workbook_snapshot, master_identities)
    sync_staging_workbook_for_building(workbook_snapshot)
    refresh_after_staging_excel_write(conn, actor=actor, request=request)
    staging_key = build_staging_snapshot_key(normalized_name, normalized_address)
    snapshot = load_staging_building_snapshot(conn, staging_key)
    if not snapshot:
        raise HTTPException(status_code=409, detail="The Staging workbook was updated, but the building was not found after refreshing the mirror.")
    write_audit_log(
        conn,
        request,
        actor,
        action_type="staging_manual_building_created",
        target_table="staging_building_info",
        target_record_id=staging_key,
        building_name=snapshot.get("building_name") or normalized_name,
        source="manual_staging_create",
        note=notes or "Building added to Staging manually.",
    )
    return snapshot


def extra_provider_text(value: object) -> Optional[str]:
    labels = extra_provider_labels_from_text(value)
    return ", ".join(labels) if labels else None


WORKBOOK_SYNC_FIELD_ORDER = [
    "building_name",
    "address",
    "insurance_required",
    "insurance_coverage_amount",
    "insurance_coi_required",
    "insurance_coi_trigger",
    "key_pickup_notes",
    "service_elevator_booking_notes",
    "electricity_required",
    "electricity_provider",
    "internet_self_setup_required",
    "internet_provider",
    "internet_notes",
    "move_in_notes",
    *NETWORK_PROVIDER_FIELD_MAP.keys(),
    *NETWORK_PLAN_FIELD_MAP.keys(),
    *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
]


def workbook_catalog_field_keys(
    conn: sqlite3.Connection,
    *,
    include_staging_only: bool,
) -> List[str]:
    keys: List[str] = list(WORKBOOK_SYNC_FIELD_ORDER)
    for item in field_catalog(conn):
        if item["field_key"] in keys:
            continue
        if not include_staging_only and item.get("scope") == "staging_only":
            continue
        keys.append(item["field_key"])
    return keys


def workbook_catalog_extension_field_keys(
    conn: sqlite3.Connection,
    *,
    include_staging_only: bool,
) -> List[str]:
    return [
        field_key
        for field_key in workbook_catalog_field_keys(conn, include_staging_only=include_staging_only)
        if field_key not in CORE_MASTER_FIELD_KEYS
    ]


def workbook_extension_value(snapshot: Optional[dict], field_key: str) -> Optional[str]:
    snapshot = snapshot or {}
    extensions = snapshot.get("extensions", {}) or {}
    if field_key in NETWORK_PROVIDER_FIELD_MAP:
        return normalize_field_value(field_key, "boolean", extensions.get(field_key))
    if field_key == "insurance_coi_required":
        return normalize_insurance_status(extensions.get(field_key))
    return normalize_unknown_value(extensions.get(field_key))


def snapshot_identity(snapshot: dict) -> tuple[str, str]:
    return (
        legacy._normalize_text(snapshot.get("building_name")),
        legacy._normalize_text(snapshot.get("address") or ""),
    )


def snapshot_has_substantive_business_data(snapshot: dict) -> bool:
    if normalize_requirement_choice(snapshot.get("insurance_required")):
        return True
    if normalize_requirement_choice(snapshot.get("electricity_required")):
        return True
    if normalize_requirement_choice(snapshot.get("internet_self_setup_required")):
        return True
    for field_key in (
        "insurance_coverage_amount",
        "electricity_provider",
        "internet_provider",
        "internet_notes",
        "key_pickup_notes",
        "service_elevator_booking_notes",
        "move_in_notes",
    ):
        if normalize_unknown_value(snapshot.get(field_key)) is not None:
            return True
    extensions = snapshot.get("extensions") or {}
    for field_key, value in extensions.items():
        if workbook_extension_value({"extensions": {field_key: value}}, field_key) is not None:
            return True
    return False


def normalize_staging_library_status(value: object) -> str:
    text = normalize_unknown_value(value) or ""
    if text in STAGING_LIBRARY_STATUSES:
        return text
    return STAGING_STATUS_PENDING


def derive_staging_library_status(snapshot: dict, master_identities: set[tuple[str, str]]) -> str:
    if snapshot_identity(snapshot) in master_identities:
        return STAGING_STATUS_MASTERED
    if snapshot_has_substantive_business_data(snapshot):
        return STAGING_STATUS_ACTIVE
    return STAGING_STATUS_PENDING


def has_explicit_snapshot_value(field_key: str, value: object) -> bool:
    if field_key in {"insurance_required", "electricity_required", "internet_self_setup_required"}:
        return normalize_requirement_choice(value) is not None
    if field_key in NETWORK_PROVIDER_FIELD_MAP:
        return normalize_booleanish(value) is not None
    return normalize_unknown_value(value) is not None


def merge_staging_excel_rows(rows: List[dict]) -> tuple[List[dict], int]:
    merged_by_identity: Dict[tuple[str, str], dict] = {}
    duplicate_count = 0
    workbook_fields = [
        "building_name",
        "address",
        "insurance_required",
        "insurance_coverage_amount",
        "electricity_required",
        "electricity_provider",
        "internet_self_setup_required",
        "internet_provider",
        "internet_notes",
        "move_in_notes",
    ]
    for row in rows:
        identity = snapshot_identity(row)
        existing = merged_by_identity.get(identity)
        if not existing:
            merged_by_identity[identity] = {
                **row,
                "extensions": dict(row.get("extensions") or {}),
            }
            continue
        duplicate_count += 1
        merged = {
            **existing,
            "extensions": dict(existing.get("extensions") or {}),
        }
        for field_key in workbook_fields:
            if has_explicit_snapshot_value(field_key, merged.get(field_key)):
                continue
            candidate = row.get(field_key)
            if has_explicit_snapshot_value(field_key, candidate):
                merged[field_key] = candidate
        candidate_extensions = row.get("extensions") or {}
        extension_keys = set(merged["extensions"]) | set(candidate_extensions) | set(WORKBOOK_EXTENSION_FIELD_KEYS)
        for field_key in extension_keys:
            if has_explicit_snapshot_value(field_key, merged["extensions"].get(field_key)):
                continue
            candidate = candidate_extensions.get(field_key)
            if has_explicit_snapshot_value(field_key, candidate):
                merged["extensions"][field_key] = candidate
        existing_row_number = existing.get("_row_number") or 10**9
        candidate_row_number = row.get("_row_number") or 10**9
        if candidate_row_number < existing_row_number:
            merged["_row_number"] = row.get("_row_number")
        merged_by_identity[identity] = merged
    return list(merged_by_identity.values()), duplicate_count


def comparable_snapshot_for_excel_diff(snapshot: Optional[dict], extension_field_keys: Optional[Iterable[str]] = None) -> dict:
    snapshot = snapshot or {}
    extension_keys = set(extension_field_keys or []) | set((snapshot.get("extensions") or {}).keys()) | set(WORKBOOK_EXTENSION_FIELD_KEYS)
    return {
        "building_name": normalize_unknown_value(snapshot.get("building_name")),
        "address": normalize_unknown_value(snapshot.get("address")),
        "insurance_required": normalize_field_value("insurance_required", "boolean", snapshot.get("insurance_required")),
        "insurance_coverage_amount": normalize_unknown_value(snapshot.get("insurance_coverage_amount")),
        "electricity_required": normalize_field_value("electricity_required", "boolean", snapshot.get("electricity_required")),
        "electricity_provider": normalize_unknown_value(snapshot.get("electricity_provider")),
        "internet_self_setup_required": normalize_field_value(
            "internet_self_setup_required", "boolean", snapshot.get("internet_self_setup_required")
        ),
        "internet_provider": extra_provider_text(snapshot.get("internet_provider")),
        "internet_notes": normalize_unknown_value(snapshot.get("internet_notes")),
        "move_in_notes": normalize_unknown_value(snapshot.get("move_in_notes")),
        "extensions": {
            field_key: workbook_extension_value(snapshot, field_key)
            for field_key in sorted(extension_keys)
        },
    }


def workbook_compare_signature(diff_fields: List[tuple[str, Optional[str]]]) -> str:
    return hashlib.sha1(json_dumps(diff_fields).encode("utf-8")).hexdigest()[:12]


def refresh_master_mirror_from_excel(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    workbook_path = resolve_master_excel_path()
    if not workbook_path.exists():
        workbook_path = ensure_master_workbook_from_db(conn)
    upgrade_master_workbook(workbook_path)
    validation = validate_master_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The master workbook structure is invalid, so the Master mirror cannot be synchronized.")

    workbook_rows = [row for row in load_master_workbook_rows(workbook_path) if snapshot_has_substantive_business_data(row)]
    desired_identities = {snapshot_identity(row) for row in workbook_rows}
    existing_rows = load_all_master_building_snapshots(conn)
    existing_by_identity = {snapshot_identity(item): item for item in existing_rows}
    sync_field_keys = workbook_catalog_field_keys(conn, include_staging_only=False)
    extension_field_keys = workbook_catalog_extension_field_keys(conn, include_staging_only=False)

    created = 0
    updated = 0
    deleted = 0

    for existing in existing_rows:
        identity = snapshot_identity(existing)
        if identity in desired_identities:
            continue
        conn.execute(
            "UPDATE staging_update_requests SET building_id = NULL, updated_at = ? WHERE building_id = ?",
            (now_iso(), existing["id"]),
        )
        conn.execute("DELETE FROM master_building_field_values WHERE building_id = ?", (existing["id"],))
        conn.execute("DELETE FROM master_building_info WHERE id = ?", (existing["id"],))
        deleted += 1

    for workbook_row in workbook_rows:
        identity = snapshot_identity(workbook_row)
        existing = existing_by_identity.get(identity) or lookup_master_building(
            conn,
            workbook_row["building_name"],
            workbook_row.get("address") or "",
        )
        if existing:
            building_id = existing["id"]
            previous_identity = (
                normalize_unknown_value(existing.get("building_name")) or "",
                normalize_unknown_value(existing.get("address")) or "",
            )
        else:
            master = ensure_master_building(
                conn,
                building_name=workbook_row["building_name"],
                address=workbook_row.get("address") or "",
                actor=actor,
            )
            building_id = master["id"]
            previous_identity = (
                normalize_unknown_value(master.get("building_name")) or "",
                normalize_unknown_value(master.get("address")) or "",
            )
            created += 1

        current_snapshot = load_master_building_snapshot(conn, building_id) or {"extensions": {}}
        current_values = comparable_snapshot_for_excel_diff(current_snapshot, extension_field_keys)
        target_values = comparable_snapshot_for_excel_diff(workbook_row, extension_field_keys)
        changed = current_values != target_values or previous_identity != (
            workbook_row["building_name"],
            workbook_row.get("address") or "",
        )
        if not changed:
            refresh_master_completeness(conn, building_id)
            continue

        for field_key in sync_field_keys:
            if field_key in extension_field_keys:
                new_value = target_values["extensions"].get(field_key)
            else:
                new_value = target_values.get(field_key)
            upsert_master_field(
                conn,
                building_id=building_id,
                field_key=field_key,
                new_value=new_value,
                actor=actor,
                source_type=MASTER_EXCEL_SYNC_SOURCE_TYPE,
                source_file=workbook_path.name,
                source_date=None,
                info_cutoff_date=None,
            )
        refresh_master_completeness(conn, building_id)
        bump_master_version(conn, building_id, actor)
        updated += 1

    conn.execute(
        "INSERT OR REPLACE INTO app_meta(key, value) VALUES('master_excel_reconciled_at', ?)",
        (now_iso(),),
    )
    if created or updated or deleted:
        write_audit_log(
            conn,
            request,
            actor,
            action_type="master_excel_mirror_refreshed",
            target_table="master_building_info",
            source=workbook_path.name,
            note=f"created={created}, updated={updated}, deleted={deleted}",
        )
    return {
        "created": created,
        "updated": updated,
        "deleted": deleted,
        "workbook_path": str(workbook_path),
    }


def refresh_staging_mirror_from_excel(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    workbook_path = ensure_staging_workbook_from_sources(conn)
    validation = validate_staging_workbook(workbook_path)
    if not validation["ok"]:
        raise ValueError("The Staging workbook structure is invalid, so the Staging mirror cannot be synchronized.")
    extension_field_keys = workbook_catalog_extension_field_keys(conn, include_staging_only=True)

    workbook_rows, duplicate_rows = merge_staging_excel_rows(load_staging_workbook_rows(workbook_path))
    master_identities = {snapshot_identity(item) for item in load_all_master_building_snapshots(conn)}
    status_by_identity: Dict[tuple[str, str], str] = {}
    normalized_rows: List[dict] = []
    for row in workbook_rows:
        row["library_status"] = derive_staging_library_status(row, master_identities)
        status_by_identity[snapshot_identity(row)] = row["library_status"]
        normalized_rows.append(row)
    sync_staging_statuses(status_by_identity)

    conn.execute("DELETE FROM staging_building_field_values")
    conn.execute("DELETE FROM staging_building_info")

    timestamp = now_iso()
    effective_source_date = iso_now_local_date()
    inserted = 0
    for row in normalized_rows:
        building_id = build_staging_snapshot_key(row["building_name"], row.get("address") or "")
        conn.execute(
            """
            INSERT INTO staging_building_info(
              id, building_name, address, insurance_required, insurance_coverage_amount,
              electricity_required, electricity_provider, internet_self_setup_required,
              internet_provider, internet_notes, move_in_notes, source_type, source_file,
              source_date, info_cutoff_date, updated_by, library_status, version, status, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 'active', ?, ?)
            """,
            (
                building_id,
                normalize_unknown_value(row.get("building_name")) or "",
                normalize_unknown_value(row.get("address")) or "",
                requirement_choice_to_int(row.get("insurance_required")),
                normalize_unknown_value(row.get("insurance_coverage_amount")),
                requirement_choice_to_int(row.get("electricity_required")),
                normalize_unknown_value(row.get("electricity_provider")),
                requirement_choice_to_int(row.get("internet_self_setup_required")),
                normalize_provider_text(row.get("internet_provider")),
                normalize_unknown_value(row.get("internet_notes")),
                normalize_unknown_value(row.get("move_in_notes")),
                "staging_excel",
                workbook_path.name,
                effective_source_date,
                None,
                actor.user_id,
                row.get("library_status") or STAGING_STATUS_PENDING,
                timestamp,
                timestamp,
            ),
        )
        for field_key in extension_field_keys:
            definition = find_field_definition(conn, field_key)
            field_type = definition["field_type"] if definition else ("boolean" if field_key in NETWORK_PROVIDER_FIELD_MAP else "text")
            value = normalize_field_value(field_key, field_type, (row.get("extensions") or {}).get(field_key))
            if value is None:
                continue
            conn.execute(
                """
                INSERT INTO staging_building_field_values(
                  id, building_id, field_key, value_text, value_json, source_type, source_file,
                  source_date, info_cutoff_date, updated_by, version, status, created_at, updated_at
                ) VALUES(?, ?, ?, ?, '', ?, ?, ?, ?, ?, 1, 'active', ?, ?)
                """,
                (
                    f"staging_field_value_{uuid.uuid4().hex}",
                    building_id,
                    field_key,
                    value,
                    "staging_excel",
                    workbook_path.name,
                    effective_source_date,
                    None,
                    actor.user_id,
                    timestamp,
                    timestamp,
                ),
            )
        inserted += 1

    if inserted:
        write_audit_log(
            conn,
            request,
            actor,
            action_type="staging_excel_mirror_refreshed",
            target_table="staging_building_info",
            source=workbook_path.name,
            note=f"inserted={inserted}, merged_duplicates={duplicate_rows}",
        )
    return {
        "inserted": inserted,
        "merged_duplicates": duplicate_rows,
        "workbook_path": str(workbook_path),
    }


def delete_generated_ab_review_queue(conn: sqlite3.Connection) -> int:
    return conn.execute(
        """
        DELETE FROM staging_update_requests
        WHERE source_type IN (?, ?)
        """,
        (MASTER_EXCEL_SYNC_SOURCE_TYPE, STAGING_AB_SYNC_SOURCE_TYPE),
    ).rowcount


def ensure_manual_master_review_group_for_staging_snapshot(
    conn: sqlite3.Connection,
    *,
    staging_snapshot: dict,
    actor: Actor,
) -> dict:
    if not snapshot_has_substantive_business_data(staging_snapshot):
        raise HTTPException(status_code=400, detail="This record contains only basic directory information and cannot be promoted to Master.")

    staging_key = staging_snapshot.get("staging_key") or staging_snapshot.get("id") or ""
    unfinished_statuses = {"pending", "ai_parsed", "employee_submitted", "needs_more_info", "conflict"}
    sync_field_keys = workbook_catalog_field_keys(conn, include_staging_only=True)
    extension_field_keys = set(workbook_catalog_extension_field_keys(conn, include_staging_only=True))
    values_by_field: Dict[str, Optional[str]] = {}
    ordered_fields: List[str] = []
    for field_key in sync_field_keys:
        if field_key in {"source_type", "source_file", "source_date", "info_cutoff_date", "library_status"}:
            continue
        if field_key in extension_field_keys:
            candidate_value = (staging_snapshot.get("extensions") or {}).get(field_key)
        else:
            candidate_value = staging_snapshot.get(field_key)
        if not has_explicit_snapshot_value(field_key, candidate_value):
            continue
        values_by_field[field_key] = normalize_unknown_value(candidate_value) or ""
        ordered_fields.append(field_key)

    building_name = normalize_unknown_value(staging_snapshot.get("building_name")) or ""
    address = normalize_unknown_value(staging_snapshot.get("address")) or ""
    if not building_name:
        raise HTTPException(status_code=400, detail="The Staging record has no building name and cannot be submitted for Master review.")
    values_by_field.setdefault("building_name", building_name)
    if address:
        values_by_field.setdefault("address", address)
    ordered_fields = [
        field_key
        for field_key in ["building_name", "address", *ordered_fields]
        if field_key in values_by_field
    ]
    if not any(field_key not in {"building_name", "address"} for field_key in ordered_fields):
        raise HTTPException(status_code=400, detail="This Staging record has no business fields that can be submitted for Master review.")

    master_snapshot = lookup_master_building(conn, building_name, address)
    group_signature = workbook_compare_signature(
        [(field_key, values_by_field.get(field_key)) for field_key in ordered_fields]
    )
    group_id = f"manual_master_{hashlib.sha1(f'{staging_key}:{group_signature}'.encode('utf-8')).hexdigest()[:20]}"
    existing_rows = fetch_group_records(conn, group_id)
    if existing_rows and any(row["review_status"] in unfinished_statuses for row in existing_rows):
        status = next((row["review_status"] for row in existing_rows if row["review_status"] in unfinished_statuses), "pending")
        return {"group_id": group_id, "created": False, "reused": True, "review_status": status}

    conn.execute(
        f"""
        UPDATE staging_update_requests
        SET review_status = 'rejected',
            reviewer = ?,
            reviewed_at = ?,
            review_comment = ?,
            updated_at = ?
        WHERE source_type = ?
          AND target_staging_key = ?
          AND submission_group_id != ?
          AND review_status IN ({",".join("?" for _ in unfinished_statuses)})
        """,
        (
            actor.user_id,
            now_iso(),
            "The Staging record was resubmitted for Master review; the previous pending group is now obsolete.",
            now_iso(),
            STAGING_MANUAL_MASTER_SOURCE_TYPE,
            staging_key,
            group_id,
            *unfinished_statuses,
        ),
    )

    if existing_rows:
        conn.execute("DELETE FROM staging_update_requests WHERE submission_group_id = ?", (group_id,))

    row_source = json_dumps(
        {
            "staging_key": staging_key,
            "building_name": building_name,
            "address": address,
            "values": values_by_field,
        }
    )
    building_id = master_snapshot["id"] if master_snapshot else None
    source_file = staging_snapshot.get("source_file") or Path(resolve_staging_excel_path()).name
    for field_key in ordered_fields:
        new_value = values_by_field.get(field_key)
        old_value = get_master_field_value(conn, building_id, field_key) if building_id else None
        create_staging_request(
            conn,
            submission_group_id=group_id,
            building_name=building_name,
            building_id=building_id,
            field_name=field_key,
            old_value=old_value,
            new_value=new_value,
            source_type=STAGING_MANUAL_MASTER_SOURCE_TYPE,
            source_content=row_source,
            source_file=source_file,
            submitted_by=actor.user_id,
            ai_confidence=None,
            review_status="pending",
            import_batch_id=None,
            parser_type=STAGING_MANUAL_MASTER_PARSER_TYPE,
            raw_input_type="staging",
            source_document_id=None,
            approval_stage=APPROVAL_STAGE_TO_MASTER,
            target_staging_key=staging_key,
            conflict_with_long_term=bool(old_value is not None and old_value != (new_value or "")),
            low_confidence=False,
            missing_required_detail=missing_required_detail_for_field(field_key, values_by_field),
            priority="high" if not master_snapshot else "normal",
        )
    return {"group_id": group_id, "created": True, "reused": False, "review_status": "pending"}


def sync_excel_mirrors(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    try:
        master_result = refresh_master_mirror_from_excel(conn, actor=actor, request=request)
        staging_result = refresh_staging_mirror_from_excel(conn, actor=actor, request=request)
        cleared = delete_generated_ab_review_queue(conn)
        conn.execute(
            "INSERT OR REPLACE INTO app_meta(key, value) VALUES('excel_mirrors_refreshed_at', ?)",
            (now_iso(),),
        )
        write_runtime_status({"excel_mirror_healthy": True, "excel_mirror_last_error": ""})
        return {
            "master": master_result,
            "staging": staging_result,
            "cleared_auto_review_records": cleared,
        }
    except Exception as exc:
        write_runtime_status({"excel_mirror_healthy": False, "excel_mirror_last_error": str(exc)})
        raise


def refresh_after_master_excel_write(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    try:
        master_result = refresh_master_mirror_from_excel(conn, actor=actor, request=request)
        staging_result = refresh_staging_mirror_from_excel(conn, actor=actor, request=request)
        cleared = delete_generated_ab_review_queue(conn)
        conn.execute(
            "INSERT OR REPLACE INTO app_meta(key, value) VALUES('excel_mirrors_refreshed_at', ?)",
            (now_iso(),),
        )
        write_runtime_status({"excel_mirror_healthy": True, "excel_mirror_last_error": ""})
        return {
            "master": master_result,
            "staging": staging_result,
            "cleared_auto_review_records": cleared,
        }
    except Exception as exc:
        write_runtime_status({"excel_mirror_healthy": False, "excel_mirror_last_error": str(exc)})
        raise


def refresh_after_staging_excel_write(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request] = None,
) -> dict:
    try:
        staging_result = refresh_staging_mirror_from_excel(conn, actor=actor, request=request)
        cleared = delete_generated_ab_review_queue(conn)
        conn.execute(
            "INSERT OR REPLACE INTO app_meta(key, value) VALUES('excel_mirrors_refreshed_at', ?)",
            (now_iso(),),
        )
        write_runtime_status({"excel_mirror_healthy": True, "excel_mirror_last_error": ""})
        return {
            "staging": staging_result,
            "cleared_auto_review_records": cleared,
        }
    except Exception as exc:
        write_runtime_status({"excel_mirror_healthy": False, "excel_mirror_last_error": str(exc)})
        raise


def mark_staging_snapshot_mastered(
    conn: sqlite3.Connection,
    *,
    staging_key: str,
    actor: Actor,
    request: Optional[Request] = None,
) -> Optional[dict]:
    snapshot = load_staging_building_snapshot(conn, staging_key)
    if not snapshot:
        return None
    previous_identity = (
        normalize_unknown_value(snapshot.get("building_name")) or "",
        normalize_unknown_value(snapshot.get("address")) or "",
    )
    workbook_snapshot = {
        "building_name": previous_identity[0],
        "address": previous_identity[1],
        "insurance_required": snapshot.get("insurance_required"),
        "insurance_coverage_amount": snapshot.get("insurance_coverage_amount"),
        "electricity_required": snapshot.get("electricity_required"),
        "electricity_provider": snapshot.get("electricity_provider"),
        "internet_self_setup_required": snapshot.get("internet_self_setup_required"),
        "internet_provider": snapshot.get("internet_provider"),
        "internet_notes": snapshot.get("internet_notes"),
        "move_in_notes": snapshot.get("move_in_notes"),
        "extensions": dict(snapshot.get("extensions") or {}),
        "library_status": STAGING_STATUS_MASTERED,
    }
    sync_staging_workbook_for_building(workbook_snapshot, previous_identity=previous_identity)
    refresh_after_staging_excel_write(conn, actor=actor, request=request)
    return load_staging_building_snapshot(conn, staging_key)


def sync_source_values_into_staging_library(
    conn: sqlite3.Connection,
    *,
    building_name: str,
    address: str,
    values_by_field: Dict[str, Optional[str]],
    actor: Actor,
    request: Optional[Request] = None,
    source_file: str = "",
    audit_action: str = "source_synced_to_staging_library",
    audit_note: str = "",
    allow_existing_match: bool = True,
) -> Optional[dict]:
    normalized_name = normalize_unknown_value(building_name) or ""
    normalized_address = normalize_unknown_value(address) or ""
    if not normalized_name:
        return None

    sync_field_keys = set(workbook_catalog_field_keys(conn, include_staging_only=True))
    extension_field_keys = set(workbook_catalog_extension_field_keys(conn, include_staging_only=True))

    existing_snapshot = (
        resolve_staging_snapshot_for_source_sync(conn, normalized_name, normalized_address)
        if allow_existing_match
        else None
    )
    previous_identity = None
    if existing_snapshot:
        previous_identity = (
            normalize_unknown_value(existing_snapshot.get("building_name")) or "",
            normalize_unknown_value(existing_snapshot.get("address")) or "",
        )

    workbook_snapshot = {
        "building_name": normalized_name,
        "address": normalized_address,
        "insurance_required": None,
        "insurance_coverage_amount": None,
        "electricity_required": None,
        "electricity_provider": None,
        "internet_self_setup_required": None,
        "internet_provider": None,
        "internet_notes": None,
        "move_in_notes": None,
        "extensions": {field_key: None for field_key in extension_field_keys},
        "library_status": STAGING_STATUS_PENDING,
    }

    if existing_snapshot:
        workbook_snapshot.update(
            {
                "building_name": previous_identity[0] or normalized_name,
                "address": previous_identity[1] or normalized_address,
                "insurance_required": existing_snapshot.get("insurance_required"),
                "insurance_coverage_amount": existing_snapshot.get("insurance_coverage_amount"),
                "electricity_required": existing_snapshot.get("electricity_required"),
                "electricity_provider": existing_snapshot.get("electricity_provider"),
                "internet_self_setup_required": existing_snapshot.get("internet_self_setup_required"),
                "internet_provider": existing_snapshot.get("internet_provider"),
                "internet_notes": existing_snapshot.get("internet_notes"),
                "move_in_notes": existing_snapshot.get("move_in_notes"),
                "extensions": dict(existing_snapshot.get("extensions") or {}),
                "library_status": existing_snapshot.get("library_status") or STAGING_STATUS_PENDING,
            }
        )

    touched_network_fields = False
    for field_key, raw_value in values_by_field.items():
        if field_key not in sync_field_keys:
            continue
        field_definition = find_field_definition(conn, field_key)
        field_type = (
            field_definition["field_type"]
            if field_definition and field_definition.get("field_type")
            else "boolean"
            if field_key in NETWORK_PROVIDER_FIELD_MAP
            else "text"
        )
        normalized_value = normalize_field_value(field_key, field_type, raw_value)
        if not has_explicit_snapshot_value(field_key, normalized_value):
            continue
        if field_key in {"building_name", "address"} and existing_snapshot:
            continue
        if field_key in NETWORK_PROVIDER_FIELD_MAP:
            workbook_snapshot["extensions"][field_key] = normalized_value
            touched_network_fields = True
            continue
        if field_key in extension_field_keys:
            workbook_snapshot["extensions"][field_key] = normalized_value
            continue
        workbook_snapshot[field_key] = normalized_value
        if field_key in {"internet_provider", "internet_notes"}:
            touched_network_fields = True

    parsed_provider_names = set(extract_provider_names(workbook_snapshot.get("internet_provider")))
    for field_key, provider_label in NETWORK_PROVIDER_FIELD_MAP.items():
        if provider_label in parsed_provider_names and not has_explicit_snapshot_value(
            field_key,
            workbook_snapshot["extensions"].get(field_key),
        ):
            workbook_snapshot["extensions"][field_key] = "true"
            touched_network_fields = True

    if touched_network_fields:
        combined_network_values: Dict[str, Optional[str]] = {
            field_key: workbook_snapshot["extensions"].get(field_key)
            for field_key in NETWORK_PROVIDER_FIELD_MAP
        }
        combined_network_values["internet_provider"] = workbook_snapshot.get("internet_provider")
        combined_network_values["internet_notes"] = workbook_snapshot.get("internet_notes")
        workbook_snapshot["internet_provider"] = build_network_provider_text_from_values(combined_network_values)

    master_identities = {snapshot_identity(item) for item in load_all_master_building_snapshots(conn)}
    workbook_snapshot["library_status"] = derive_staging_library_status(workbook_snapshot, master_identities)

    sync_staging_workbook_for_building(workbook_snapshot, previous_identity=previous_identity)
    refresh_after_staging_excel_write(conn, actor=actor, request=request)

    refreshed_key = build_staging_snapshot_key(
        normalize_unknown_value(workbook_snapshot.get("building_name")) or "",
        normalize_unknown_value(workbook_snapshot.get("address")) or "",
    )
    refreshed_snapshot = load_staging_building_snapshot(conn, refreshed_key)
    if refreshed_snapshot:
        write_audit_log(
            conn,
            request,
            actor,
            action_type=audit_action,
            target_table="staging_building_info",
            target_record_id=refreshed_snapshot["staging_key"],
            building_name=refreshed_snapshot["building_name"],
            source=source_file or Path(resolve_staging_excel_path()).name,
            note=audit_note or "Parsed source results were synchronized to Staging.",
        )
    return refreshed_snapshot


def extension_boolean_value(snapshot: dict, field_key: str) -> Optional[bool]:
    extensions = snapshot.get("extensions", {}) or {}
    return normalize_booleanish(extensions.get(field_key))


def supported_network_providers(snapshot: dict) -> List[str]:
    supported: List[str] = []
    for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items():
        if extension_boolean_value(snapshot, field_key) is True:
            supported.append(label)
    return supported


def provider_status_for_snapshot(snapshot: dict, provider_label: str) -> Optional[bool]:
    normalized = normalize_provider_name(provider_label)
    for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items():
        if label == normalized:
            extension_status = extension_boolean_value(snapshot, field_key)
            if extension_status is not None:
                return extension_status
            explicit_providers = extract_provider_names(snapshot.get("internet_provider"))
            if label in explicit_providers:
                return True
            return None
    return None


def effective_network_provider_text(snapshot: dict) -> Optional[str]:
    values: Dict[str, Optional[str]] = {
        field_key: snapshot.get("extensions", {}).get(field_key)
        for field_key in NETWORK_PROVIDER_FIELD_MAP
    }
    values["internet_provider"] = snapshot.get("internet_provider")
    return build_network_provider_text_from_values(values)


def refresh_master_network_provider_text(
    conn: sqlite3.Connection,
    *,
    building_id: str,
    actor: Actor,
    source_type: str,
    source_file: str,
) -> None:
    snapshot = load_master_building_snapshot(conn, building_id)
    if not snapshot:
        return
    provider_text = effective_network_provider_text(snapshot)
    upsert_master_field(
        conn,
        building_id=building_id,
        field_key="internet_provider",
        new_value=provider_text,
        actor=actor,
        source_type=source_type,
        source_file=source_file,
        source_date=None,
        info_cutoff_date=None,
    )


def master_network_wifi_mode(snapshot: dict) -> str:
    self_setup = requirement_state(snapshot.get("internet_self_setup_required"))
    if self_setup == "required":
        return "Resident must arrange internet service"
    if self_setup == "not_required":
        return "Internet is included or no resident setup is required"
    if self_setup == "optional":
        return "Resident setup is optional"
    return "Not specified"


def build_network_payload_from_master(snapshot: dict) -> Optional[dict]:
    provider_rows = []
    has_any_structured_value = False
    explicit_provider_labels = extract_provider_names(snapshot.get("internet_provider"))
    explicit_providers = set(explicit_provider_labels)
    extra_providers = [
        label for label in explicit_provider_labels if label not in FIXED_NETWORK_PROVIDER_SET
    ]
    plan_values = network_plan_values_from_snapshot(snapshot)
    note_values = network_provider_note_values_from_snapshot(snapshot)
    for field_key, label in NETWORK_PROVIDER_FIELD_MAP.items():
        status_value = extension_boolean_value(snapshot, field_key)
        if status_value is True:
            status = "supported"
            has_any_structured_value = True
        elif status_value is False:
            status = "unsupported"
            has_any_structured_value = True
        elif label in explicit_providers:
            status = "supported"
            has_any_structured_value = True
        else:
            status = "pending"
        provider_rows.append(
            {
                "key": label.lower(),
                "label": label,
                "status": status,
                "plans": plan_values.get(
                    next(
                        (
                            plan_field_key
                            for plan_field_key, plan_label in NETWORK_PLAN_FIELD_MAP.items()
                            if plan_label == label
                        ),
                        "",
                    )
                )
                or "",
                "note": note_values.get(
                    next(
                        (
                            note_field_key
                            for note_field_key, note_label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items()
                            if note_label == label
                        ),
                        "",
                    )
                )
                or "",
            }
        )

    explicit_provider = normalize_provider_text(snapshot.get("internet_provider"))
    notes = normalize_unknown_value(snapshot.get("internet_notes"))
    if any(plan_values.values()) or any(note_values.values()):
        has_any_structured_value = True
    if explicit_provider or notes or requirement_state(snapshot.get("internet_self_setup_required")) != "unknown":
        has_any_structured_value = True
    if not has_any_structured_value:
        return None

    return {
        "building_name": normalize_unknown_value(snapshot.get("building_name")) or "",
        "address": normalize_unknown_value(snapshot.get("address")) or "",
        "website": "",
        "contact": "",
        "wifi_mode": master_network_wifi_mode(snapshot),
        "mode_detail": effective_network_provider_text(snapshot) or "",
        "notes": notes or "",
        "extra_providers": extra_providers,
        "providers": provider_rows,
        "source_file": Path(resolve_master_excel_path()).name,
        "source_sheets": [MASTER_MAIN_SHEET],
        "has_conflict": False,
        "conflict_fields": [],
        "reference_notice": "The internet table below is generated from structured Master data and the master workbook.",
        "inferred_self_setup_required": normalize_booleanish(snapshot.get("internet_self_setup_required")),
    }


def matched_field_keys_for_question(conn: sqlite3.Connection, question: str) -> List[str]:
    lowered = question.lower()
    matched: List[str] = []
    if any(token in lowered for token in FIELD_GROUP_KEYWORDS["insurance"]):
        matched.extend(
            [
                "insurance_required",
                "insurance_coverage_amount",
                "insurance_renters_required",
                "insurance_renters_minimum_coverage",
                "insurance_personal_property_required",
                "insurance_personal_property_minimum",
                "insurance_personal_liability_required",
                "insurance_personal_liability_per_occurrence",
                "insurance_personal_liability_aggregate",
                "insurance_coi_required",
                "insurance_coi_trigger",
                "insurance_interested_party_required",
                "insurance_additional_insured_required",
                "insurance_certificate_holder_required",
                "insurance_submission_method",
                "insurance_recipient",
                "insurance_alternative_program_or_penalty",
            ]
        )
    if any(token in lowered for token in FIELD_GROUP_KEYWORDS["electric"]):
        matched.extend(["electricity_required", "electricity_provider"])
    if any(token in lowered for token in FIELD_GROUP_KEYWORDS["internet"]):
        matched.extend(
            [
                "internet_self_setup_required",
                "internet_provider",
                "internet_notes",
                *NETWORK_PROVIDER_FIELD_MAP.keys(),
                *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
            ]
        )
    if any(token in lowered for token in FIELD_GROUP_KEYWORDS["move_in"]):
        matched.extend(["key_pickup_notes", "service_elevator_booking_notes", "move_in_notes"])
    if matched:
        return list(dict.fromkeys(matched))

    catalog = field_catalog(conn)
    scored: List[tuple[float, str]] = []
    normalized_question = legacy._normalize_text(question)
    for item in catalog:
        if not item.get("visible_in_query"):
            continue
        candidates = [
            item["display_name"],
            item["field_key"],
            *(item.get("query_keywords") or []),
            *item.get("aliases", []),
        ]
        best = 0.0
        for candidate in candidates:
            normalized_candidate = legacy._normalize_text(candidate)
            if not normalized_candidate:
                continue
            if normalized_candidate and normalized_candidate in normalized_question:
                best = max(best, 1.0)
            else:
                best = max(best, SequenceMatcher(None, normalized_question, normalized_candidate).ratio())
        if best >= 0.65:
            scored.append((best, item["field_key"]))
    scored.sort(key=lambda pair: -pair[0])
    return [field_key for _, field_key in scored[:4]]


def build_summary_answer(snapshot: dict, source_mode: Literal["master", "staging"] = "master") -> str:
    provider_text = effective_network_provider_text(snapshot)
    insurance_values = insurance_values_from_snapshot(snapshot)
    move_in_values = move_in_values_from_snapshot(snapshot)
    plan_values = network_plan_values_from_snapshot(snapshot)
    note_values = network_provider_note_values_from_snapshot(snapshot)
    renters_status = normalize_insurance_status(insurance_values.get("insurance_renters_required"))
    electricity_state = requirement_state(snapshot.get("electricity_required"))
    internet_state = requirement_state(snapshot.get("internet_self_setup_required"))
    lines = [
        f"Building: {snapshot['building_name']}",
        f"Address: {display_value_or_unknown(snapshot.get('address'))}",
        f"Renters Insurance: {insurance_status_label(renters_status)}",
        f"Minimum Renters Insurance Coverage: {display_value_or_unknown(insurance_values.get('insurance_renters_minimum_coverage'))}",
        f"Electricity Setup: {'Required' if electricity_state == 'required' else 'Not required' if electricity_state == 'not_required' else 'Optional' if electricity_state == 'optional' else 'Unknown'}",
        f"Electric Utility: {display_value_or_unknown(snapshot.get('electricity_provider'))}",
        f"Internet Setup: {'Resident must arrange service' if internet_state == 'required' else 'Included or no resident setup required' if internet_state == 'not_required' else 'Optional' if internet_state == 'optional' else 'Unknown'}",
        f"Internet Providers: {provider_text or 'Unknown'}",
    ]
    if insurance_values.get("insurance_personal_liability_per_occurrence") or insurance_values.get("insurance_personal_liability_aggregate"):
        liability_bits = []
        if insurance_values.get("insurance_personal_liability_per_occurrence"):
            liability_bits.append(f"per occurrence {insurance_values['insurance_personal_liability_per_occurrence']}")
        if insurance_values.get("insurance_personal_liability_aggregate"):
            liability_bits.append(f"aggregate {insurance_values['insurance_personal_liability_aggregate']}")
        lines.append(f"Personal Liability: {'; '.join(liability_bits)}")
    if insurance_values.get("insurance_coi_required"):
        lines.append(
            f"COI: {insurance_status_label(insurance_values.get('insurance_coi_required'))}"
            + (f" ({insurance_values['insurance_coi_trigger']})" if insurance_values.get("insurance_coi_trigger") else "")
        )
    if move_in_values.get("key_pickup_notes"):
        lines.append(f"Key Pickup: {move_in_values['key_pickup_notes']}")
    elif move_in_values.get("move_in_notes") and any(
        token in move_in_values["move_in_notes"].lower() for token in ("key", "钥匙")
    ):
        lines.append(f"Additional Key-Pickup Note: {move_in_values['move_in_notes']}")
    if move_in_values.get("service_elevator_booking_notes"):
        lines.append(f"Service-Elevator Reservation: {move_in_values['service_elevator_booking_notes']}")
    elif move_in_values.get("move_in_notes") and any(
        token in move_in_values["move_in_notes"].lower() for token in ("elevator", "货梯", "rello", "time slot")
    ):
        lines.append(f"Additional Service-Elevator Note: {move_in_values['move_in_notes']}")
    if snapshot.get("internet_notes"):
        lines.append(f"Internet Notes: {snapshot['internet_notes']}")
    for field_key, provider_label in NETWORK_PLAN_FIELD_MAP.items():
        plan_text = format_network_plan_text(plan_values.get(field_key))
        if plan_text:
            lines.append(f"{provider_label} Plans: {plan_text}")
    for field_key, provider_label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items():
        note_text = normalize_unknown_value(note_values.get(field_key))
        if note_text:
            lines.append(f"{provider_label} Notes/Contact: {note_text}")
    if move_in_values.get("move_in_notes"):
        lines.append(f"Move-In Notes: {move_in_values['move_in_notes']}")
    if source_mode == "master" and snapshot.get("pending_count"):
        lines.append("Additional information is awaiting review; this answer uses approved Master data only.")
    if source_mode == "master" and snapshot.get("completeness_status") != "verified_complete":
        lines.append("This building record is verified but incomplete; blank fields are treated as unknown.")
    if source_mode == "staging":
        lines.append("This is the Staging view and shows the currently submitted information directly.")
    return "\n".join(lines)


def format_bool_for_ui(value: object) -> str:
    state = requirement_state(value)
    if state == "required":
        return "Yes"
    if state == "not_required":
        return "No"
    if state == "optional":
        return "Optional"
    normalized = normalize_booleanish(value)
    if normalized is True:
        return "Yes"
    if normalized is False:
        return "No"
    return normalize_unknown_value(value) or "Unknown"


def render_dynamic_field_answer(field_definition: Optional[dict], value: Optional[str], pending_note: str) -> str:
    if not field_definition:
        return f"Unknown field: {value if value else 'Unknown'}{pending_note if not value else ''}"
    display_name = field_definition.get("display_name") or field_definition["field_key"]
    template = (field_definition.get("answer_template") or "{display_name}: {value}").strip()
    if field_definition.get("field_type") == "boolean":
        rendered_value = format_bool_for_ui(value)
    else:
        rendered_value = value or "Unknown"
    rendered = template.replace("{display_name}", display_name).replace("{value}", rendered_value)
    if not value:
        rendered += pending_note
    return rendered


def build_structured_answer(
    conn: sqlite3.Connection,
    snapshot: dict,
    question: str,
    field_keys: List[str],
    source_mode: Literal["master", "staging"] = "master",
) -> str:
    if not field_keys or any(token in question.lower() for token in FIELD_GROUP_KEYWORDS["summary"]):
        return build_summary_answer(snapshot, source_mode)

    extension_values = snapshot.get("extensions", {})
    insurance_values = insurance_values_from_snapshot(snapshot)
    move_in_values = move_in_values_from_snapshot(snapshot)
    plan_values = network_plan_values_from_snapshot(snapshot)
    responses: List[str] = []
    pending_note = (
        " (pending information exists; the following still uses approved Master data)"
        if source_mode == "master" and snapshot.get("pending_count")
        else ""
    )
    record_label = "current Staging record" if source_mode == "staging" else "Master record"

    if set(field_keys) & (
        {"insurance_required", "insurance_coverage_amount"} | DETAILED_INSURANCE_FIELD_KEYS
    ):
        renters_status = normalize_insurance_status(insurance_values.get("insurance_renters_required"))
        renters_coverage = insurance_values.get("insurance_renters_minimum_coverage")
        liability_status = normalize_insurance_status(insurance_values.get("insurance_personal_liability_required"))
        liability_per_occurrence = insurance_values.get("insurance_personal_liability_per_occurrence")
        liability_aggregate = insurance_values.get("insurance_personal_liability_aggregate")
        coi_status = normalize_insurance_status(insurance_values.get("insurance_coi_required"))
        coi_trigger = insurance_values.get("insurance_coi_trigger")
        submission_method = insurance_values.get("insurance_submission_method")
        recipient = insurance_values.get("insurance_recipient")
        penalty = insurance_values.get("insurance_alternative_program_or_penalty")

        if renters_status == "yes":
            if renters_coverage:
                responses.append(f"This building requires renters insurance. The recorded minimum coverage is {renters_coverage}{pending_note}.")
            else:
                responses.append(f"This building requires renters insurance, but the minimum coverage amount is not clearly stated{pending_note}.")
        elif renters_status == "optional":
            responses.append(f"According to the {record_label}, renters insurance is optional for this building{pending_note}.")
        elif renters_status == "no":
            responses.append(f"According to the {record_label}, renters insurance is not required for this building{pending_note}.")
        elif renters_status == "manual_review":
            responses.append(f"The source mentions insurance, but a person must confirm whether it specifically means renters insurance{pending_note}.")
        else:
            responses.append(f"The record does not state whether renters insurance is required{pending_note}.")

        if liability_status == "yes":
            liability_parts = []
            if liability_per_occurrence:
                liability_parts.append(f"per occurrence {liability_per_occurrence}")
            if liability_aggregate:
                liability_parts.append(f"aggregate {liability_aggregate}")
            if liability_parts:
                responses.append(f"The source also mentions personal liability, recorded as {'; '.join(liability_parts)}{pending_note}.")
            else:
                responses.append(f"The source mentions personal liability, but the limits still require human confirmation{pending_note}.")
        elif liability_status == "manual_review":
            responses.append(f"The source may mention personal liability, but it still requires human confirmation{pending_note}.")

        if coi_status == "yes":
            if coi_trigger:
                responses.append(f"The source also requires a COI when: {coi_trigger}{pending_note}. A COI is not the same as the customer's own renters insurance.")
            else:
                responses.append(f"The source mentions a COI, but the triggering conditions are not fully specified{pending_note}.")
        elif coi_status == "manual_review":
            responses.append(f"The source mentions a COI, but whether it is mandatory and when it applies still require human confirmation{pending_note}.")

        if normalize_insurance_status(insurance_values.get("insurance_interested_party_required")) == "yes":
            responses.append(f"The current record also requires an Interested Party / Additional Interest{pending_note}.")
        elif normalize_insurance_status(insurance_values.get("insurance_interested_party_required")) == "manual_review":
            responses.append(f"The source requires adding the property manager or landlord to the policy, but the exact designation still requires human confirmation{pending_note}.")

        if normalize_insurance_status(insurance_values.get("insurance_additional_insured_required")) == "yes":
            responses.append(f"The current record explicitly requires an Additional Insured{pending_note}.")
        if normalize_insurance_status(insurance_values.get("insurance_certificate_holder_required")) == "yes":
            responses.append(f"The current record explicitly requires a Certificate Holder{pending_note}.")
        if submission_method:
            responses.append(f"Insurance submission/verification method: {submission_method}{pending_note}.")
        if recipient:
            responses.append(f"Insurance recipient: {recipient}{pending_note}.")
        if penalty:
            responses.append(f"Alternative program or penalty when insurance is not submitted: {penalty}{pending_note}.")

    if set(field_keys) & {"electricity_required", "electricity_provider"}:
        required = requirement_state(snapshot.get("electricity_required"))
        provider = normalize_unknown_value(snapshot.get("electricity_provider"))
        if required == "required":
            if provider:
                responses.append(f"Residents must arrange electricity service. The recorded utility is {provider}{pending_note}.")
            else:
                responses.append(f"Residents must arrange electricity service, but the utility is currently unknown{pending_note}.")
        elif required == "not_required":
            responses.append(f"According to the {record_label}, residents do not need to arrange electricity service{pending_note}.")
        elif required == "optional":
            if provider:
                responses.append(f"According to the {record_label}, electricity setup is optional. If needed, contact {provider}{pending_note}.")
            else:
                responses.append(f"According to the {record_label}, electricity setup is optional{pending_note}.")
        else:
            responses.append(f"Whether the resident must arrange electricity service is currently unknown{pending_note}.")

    network_detail_field_keys = {
        "internet_self_setup_required",
        "internet_provider",
        "internet_notes",
        *NETWORK_PROVIDER_FIELD_MAP.keys(),
        *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
    }
    if set(field_keys) & network_detail_field_keys:
        required = requirement_state(snapshot.get("internet_self_setup_required"))
        provider = effective_network_provider_text(snapshot)
        notes = normalize_unknown_value(snapshot.get("internet_notes"))
        lowered_question = (question or "").lower()
        mentioned_providers = [
            label
            for label in NETWORK_PROVIDER_FIELD_MAP.values()
            if label.lower() in lowered_question
        ]
        for label in mentioned_providers:
            status = provider_status_for_snapshot(snapshot, label)
            if status is True:
                responses.append(f"According to the {record_label}, {label} is available{pending_note}.")
            elif status is False:
                responses.append(f"According to the {record_label}, {label} is not currently marked as available{pending_note}.")
            else:
                responses.append(f"Whether {label} is supported is currently unknown{pending_note}.")
            note_field_key = next(
                (
                    field_key
                    for field_key, note_label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items()
                    if note_label == label
                ),
                "",
            )
            note_text = normalize_unknown_value(note_values.get(note_field_key))
            if note_text:
                responses.append(f"{label} notes/contact: {note_text}.")
        if required == "required":
            if provider:
                responses.append(f"Residents must arrange internet service. Providers available in the {record_label} include {provider}{pending_note}.")
            else:
                responses.append(f"Residents must arrange internet service, but the specific providers are currently unknown{pending_note}.")
        elif required == "not_required":
            responses.append(f"According to the {record_label}, residents do not need to arrange internet service, or service is already included{pending_note}.")
            if provider:
                responses.append(f"Additional provider/support information currently recorded: {provider}.")
        elif required == "optional":
            if provider:
                responses.append(f"According to the {record_label}, arranging internet service is optional. Reference providers include {provider}{pending_note}.")
            else:
                responses.append(f"According to the {record_label}, arranging internet service is optional{pending_note}.")
        else:
            if provider:
                responses.append(f"Whether the resident must arrange internet service is unknown, but providers available in the {record_label} include {provider}{pending_note}.")
            else:
                responses.append(f"Whether the resident must arrange internet service is currently unknown{pending_note}.")
        if notes:
            responses.append(f"Internet notes: {notes}")
        plan_lines = []
        for plan_field_key, provider_label in NETWORK_PLAN_FIELD_MAP.items():
            plan_text = format_network_plan_text(plan_values.get(plan_field_key))
            if not plan_text:
                continue
            if mentioned_providers and provider_label not in mentioned_providers:
                continue
            plan_lines.append(f"{provider_label} plans: {plan_text}")
        responses.extend(plan_lines)
        for note_field_key, provider_label in NETWORK_PROVIDER_NOTE_FIELD_MAP.items():
            if mentioned_providers and provider_label not in mentioned_providers:
                continue
            note_text = normalize_unknown_value(note_values.get(note_field_key))
            if note_text and not any(f"{provider_label} notes/contact" in item for item in responses):
                responses.append(f"{provider_label} notes/contact: {note_text}")

    if set(field_keys) & {"key_pickup_notes", "service_elevator_booking_notes", "move_in_notes"}:
        key_pickup = move_in_values.get("key_pickup_notes")
        service_elevator = move_in_values.get("service_elevator_booking_notes")
        move_in_note = move_in_values.get("move_in_notes")
        if "key_pickup_notes" in field_keys:
            key_pickup_answer = key_pickup or move_in_note or "Not currently recorded"
            responses.append(f"Key pickup: {key_pickup_answer}{pending_note if key_pickup_answer == 'Not currently recorded' else ''}")
        if "service_elevator_booking_notes" in field_keys:
            service_elevator_answer = service_elevator or move_in_note or "Not currently recorded"
            responses.append(
                f"Service-elevator reservation: {service_elevator_answer}{pending_note if service_elevator_answer == 'Not currently recorded' else ''}"
            )
        if "move_in_notes" in field_keys:
            fallback = move_in_note
            if not fallback and (key_pickup or service_elevator):
                fallback = "Key-pickup and service-elevator instructions are recorded in their dedicated fields."
            responses.append(f"Move-in notes: {fallback if fallback else 'Unknown'}{pending_note if not fallback else ''}")

    catalog_map = {item["field_key"]: item for item in field_catalog(conn)}
    for field_key in field_keys:
        if field_key in {
            "insurance_required",
            "insurance_coverage_amount",
            "electricity_required",
            "electricity_provider",
            "internet_self_setup_required",
            "internet_provider",
            "internet_notes",
            *NETWORK_PROVIDER_FIELD_MAP.keys(),
            *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
            "key_pickup_notes",
            "service_elevator_booking_notes",
            "move_in_notes",
        } | DETAILED_INSURANCE_FIELD_KEYS:
            continue
        value = normalize_unknown_value(snapshot.get(field_key))
        if value is None:
            value = normalize_unknown_value(extension_values.get(field_key))
        responses.append(render_dynamic_field_answer(catalog_map.get(field_key), value, pending_note))

    if source_mode == "staging" and responses:
        responses.append("This is the Staging view and shows the currently submitted information directly.")
    if source_mode == "master" and snapshot.get("completeness_status") != "verified_complete":
        responses.append("This building record is verified but incomplete; blank fields are treated as unknown.")

    return "\n".join(dict.fromkeys(responses)) if responses else build_summary_answer(snapshot, source_mode)


def is_network_question(question: str, field_keys: List[str]) -> bool:
    lowered = (question or "").lower()
    if any(token in lowered for token in FIELD_GROUP_KEYWORDS["internet"]):
        return True
    return bool(
        set(field_keys)
        & {
            "internet_self_setup_required",
            "internet_provider",
            "internet_notes",
            *NETWORK_PROVIDER_FIELD_MAP.keys(),
            *NETWORK_PROVIDER_NOTE_FIELD_MAP.keys(),
        }
    )


def infer_network_self_setup_from_record(record: legacy.InternetRecord) -> Optional[bool]:
    lowered = legacy._normalize_text(" ".join(filter(None, [record.wifi_mode, record.mode_detail, record.notes])))
    if any(token in lowered for token in ("自行", "self", "choose", "自己开网", "self setup")):
        return True
    if any(token in lowered for token in ("自带", "包含", "included", "楼内已含", "free", "mandatory")):
        return False
    return None


def resolve_network_for_snapshot(snapshot: dict) -> Optional[dict]:
    master_payload = build_network_payload_from_master(snapshot)
    if master_payload:
        return master_payload

    queries = []
    building_name = normalize_unknown_value(snapshot.get("building_name")) or ""
    address = normalize_unknown_value(snapshot.get("address")) or ""
    if building_name and address:
        queries.append(f"{building_name} {address}")
    if building_name:
        queries.append(building_name)
    if address:
        queries.append(address)

    matched_record: Optional[legacy.InternetRecord] = None
    for query in queries:
        matched, _ = legacy._resolve_internet(query)
        if matched:
            matched_record = matched
            break
    if not matched_record:
        return None

    payload = legacy._internet_result_payload(matched_record)
    provider_labels = [
        normalize_provider_name(item.get("label")) or item.get("label") or ""
        for item in payload["providers"]
        if item.get("status") in {"supported", "pending"}
    ]
    conflict_fields: List[str] = []
    master_providers = [
        label for label in extract_provider_names(snapshot.get("internet_provider"))
        if label in NETWORK_PROVIDER_FIELD_MAP.values()
    ]
    if master_providers and provider_labels and not set(master_providers) & set(provider_labels):
        conflict_fields.append("internet_provider")

    inferred_self_setup = infer_network_self_setup_from_record(matched_record)
    master_self_setup = snapshot.get("internet_self_setup_required")
    if master_self_setup in {0, 1} and inferred_self_setup is not None and master_self_setup != int(inferred_self_setup):
        conflict_fields.append("internet_self_setup_required")

    payload["has_conflict"] = bool(conflict_fields)
    payload["conflict_fields"] = conflict_fields
    payload.setdefault("extra_providers", [])
    payload["reference_notice"] = (
        "The internet table is a historical structured reference; the authoritative answer still comes from Master fields."
        if conflict_fields
        else "The internet table below is supplemental reference material and does not override the Master answer."
    )
    payload["inferred_self_setup_required"] = inferred_self_setup
    return payload


def stream_text_response(content: str):
    async def event_stream():
        yield f"data: {json_dumps({'type': 'token', 'content': content})}\n\n"
        yield f"data: {json_dumps({'type': 'done'})}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


def frontend_candidate(path: str) -> Optional[Path]:
    if not path:
        return FRONTEND_INDEX_PATH if FRONTEND_INDEX_PATH.is_file() else None
    candidate = (FRONTEND_DIST_DIR / path.lstrip("/")).resolve()
    try:
        candidate.relative_to(FRONTEND_DIST_DIR)
    except ValueError:
        return None
    if candidate.is_file():
        return candidate
    return None


def serve_frontend(path: str = "") -> FileResponse:
    if not FRONTEND_INDEX_PATH.is_file():
        raise HTTPException(status_code=404, detail="Frontend static assets have not been built.")
    candidate = frontend_candidate(path)
    if candidate is not None:
        return FileResponse(candidate)
    return FileResponse(FRONTEND_INDEX_PATH)


def bootstrap_legacy_headers() -> None:
    return None


def mark_interrupted_intake_jobs(conn: sqlite3.Connection) -> None:
    timestamp = now_iso()
    baidu_resume_enabled = int(
        OCR_PROVIDER == "baidu_unlimited_cloud" and bool(BAIDU_OCR_API_KEY and BAIDU_OCR_SECRET_KEY)
    )
    conn.execute(
        """
        UPDATE source_documents
        SET parse_status = ?, parse_completed_at = ?, parse_error = ?, updated_at = ?
        WHERE parse_status IN (?, ?)
          AND NOT EXISTS (
            SELECT 1 FROM ocr_jobs
            WHERE ocr_jobs.source_document_id = source_documents.id
              AND ? = 1
              AND ocr_jobs.provider IN ('baidu_unlimited_cloud', 'baidu_unlimited')
              AND ocr_jobs.external_task_id != ''
              AND ocr_jobs.status IN ('queued', 'pending', 'processing')
          )
        """,
        (
            INTAKE_PARSE_STATUS_FAILED,
            timestamp,
            "The background parsing task was interrupted by a service restart. Please submit the file again.",
            timestamp,
            INTAKE_PARSE_STATUS_QUEUED,
            INTAKE_PARSE_STATUS_RUNNING,
            baidu_resume_enabled,
        ),
    )


def resumable_intake_jobs(conn: sqlite3.Connection) -> List[dict]:
    return conn.execute(
        """
        SELECT DISTINCT sd.id AS source_document_id,
               COALESCE(u.id, sd.created_by, 'system_intake_parser') AS user_id,
               COALESCE(u.username, 'system') AS username,
               COALESCE(u.role, 'ai_system') AS role
        FROM source_documents sd
        JOIN ocr_jobs oj ON oj.source_document_id = sd.id
        LEFT JOIN users u ON u.id = sd.created_by
        WHERE sd.parse_status IN (?, ?)
          AND oj.provider IN ('baidu_unlimited_cloud', 'baidu_unlimited')
          AND oj.external_task_id != ''
          AND oj.status IN ('queued', 'pending', 'processing')
        ORDER BY sd.created_at ASC
        """,
        (INTAKE_PARSE_STATUS_QUEUED, INTAKE_PARSE_STATUS_RUNNING),
    ).fetchall()


@app.on_event("startup")
async def startup_event():
    init_db()
    write_runtime_status(
        {
            "runtime_mode": RUNTIME_MODE,
            "backend_healthy": True,
            "frontend_healthy": FRONTEND_INDEX_PATH.is_file() if RUNTIME_MODE == "daemon" else None,
            "last_boot_at": now_iso(),
        }
    )
    jobs_to_resume: List[dict] = []
    with db_connection() as conn:
        mark_interrupted_intake_jobs(conn)
        if OCR_PROVIDER == "baidu_unlimited_cloud" and BAIDU_OCR_API_KEY and BAIDU_OCR_SECRET_KEY:
            jobs_to_resume = [dict(item) for item in resumable_intake_jobs(conn)]
        assert_default_passwords_safe(conn)
        ensure_crm_default_templates(conn)
        normalize_crm_task_rows(conn)
        try:
            ensure_master_workbook_from_db(conn)
            ensure_staging_workbook_from_sources(conn)
            sync_excel_mirrors(
                conn,
                actor=system_actor(),
                request=None,
            )
            write_runtime_status({"excel_mirror_healthy": True, "excel_mirror_last_error": ""})
        except Exception as exc:
            write_runtime_status({"excel_mirror_healthy": False, "excel_mirror_last_error": str(exc)})
            print(f"[whitepaper] excel mirror init warning: {exc}")
    legacy.reload_building_store()
    legacy.reload_internet_store()
    legacy.reload_script_store()
    for item in jobs_to_resume:
        asyncio.create_task(
            run_intake_parse_job(
                item["source_document_id"],
                {
                    "user_id": item["user_id"],
                    "username": item["username"],
                    "role": item["role"],
                },
            )
        )


@app.post("/auth/login")
async def login(payload: LoginRequest, request: Request):
    username = payload.username.strip()
    check_login_rate_limit(request, username)
    with db_connection() as conn:
        user = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (username,),
        ).fetchone()
        if not user or not user["is_active"] or not verify_password(payload.password, user["password_hash"]):
            record_login_failure(request, username)
            raise HTTPException(status_code=401, detail="Incorrect username or password.")
        clear_login_failures(request, username)
        token, token_hash_value, expires_at = make_session_token()
        session_id = f"session_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO sessions(id, user_id, token_hash, expires_at, created_at, last_used_at)
            VALUES(?, ?, ?, ?, ?, ?)
            """,
            (session_id, user["id"], token_hash_value, expires_at, now_iso(), now_iso()),
        )
        actor = get_actor(user)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="login",
            target_table="users",
            target_record_id=user["id"],
        )
        return {
            "token": token,
            "user": {
                "id": user["id"],
                "username": user["username"],
                "display_name": user["display_name"],
                "role": user["role"],
            },
            "expires_at": expires_at,
        }


@app.get("/auth/me")
async def auth_me(user: dict = Depends(get_current_user)):
    return {"user": user}


@app.post("/auth/logout")
async def logout(
    request: Request,
    authorization: Optional[str] = Header(default=None),
    user: dict = Depends(get_current_user),
):
    token = parse_bearer_token(authorization)
    with db_connection() as conn:
        conn.execute("DELETE FROM sessions WHERE token_hash = ?", (hash_token(token or ""),))
        write_audit_log(
            conn,
            request,
            get_actor(user),
            action_type="logout",
            target_table="users",
            target_record_id=user["id"],
        )
    return {"ok": True}


def serialize_admin_user(row: dict) -> dict:
    return {
        "id": row["id"],
        "username": row.get("username") or "",
        "display_name": row.get("display_name") or "",
        "role": row.get("role") or "employee",
        "is_active": bool(row.get("is_active")),
        "created_at": row.get("created_at") or "",
        "updated_at": row.get("updated_at") or "",
    }


def ensure_not_last_active_super_admin(
    conn: sqlite3.Connection,
    *,
    user_id: str,
    next_role: Optional[str] = None,
    next_is_active: Optional[bool] = None,
) -> None:
    current = conn.execute("SELECT role, is_active FROM users WHERE id = ?", (user_id,)).fetchone()
    if not current:
        return
    current_is_super = current["role"] == "super_admin" and bool(current["is_active"])
    next_is_super = (next_role or current["role"]) == "super_admin" and (
        bool(current["is_active"]) if next_is_active is None else bool(next_is_active)
    )
    if current_is_super and not next_is_super:
        active_super_count = conn.execute(
            "SELECT COUNT(*) AS total FROM users WHERE role = 'super_admin' AND is_active = 1"
        ).fetchone()["total"]
        if active_super_count <= 1:
            raise HTTPException(status_code=400, detail="At least one active Super Admin must remain.")


@app.post("/auth/change-password")
async def change_password(
    payload: PasswordChangeRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    timestamp = now_iso()
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user["id"],)).fetchone()
        if not row or not verify_password(payload.current_password, row["password_hash"]):
            raise HTTPException(status_code=400, detail="The current password is incorrect.")
        conn.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?",
            (hash_password(payload.new_password), timestamp, user["id"]),
        )
        write_audit_log(
            conn,
            request,
            get_actor(row),
            action_type="password_changed",
            target_table="users",
            target_record_id=user["id"],
        )
    return {"ok": True}


@app.get("/admin/users")
async def list_admin_users(user: dict = Depends(require_roles("super_admin"))):
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, username, display_name, role, is_active, created_at, updated_at
            FROM users
            ORDER BY is_active DESC, role ASC, display_name ASC, username ASC
            """
        ).fetchall()
    return {"users": [serialize_admin_user(row) for row in rows]}


@app.post("/admin/users")
async def create_admin_user(
    payload: AdminUserCreateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    username = payload.username.strip().lower()
    if not re.fullmatch(r"[a-z0-9_.-]{2,64}", username):
        raise HTTPException(status_code=400, detail="Usernames may contain only lowercase letters, numbers, periods, underscores, and hyphens.")
    timestamp = now_iso()
    actor = get_actor(user)
    with db_connection() as conn:
        if conn.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone():
            raise HTTPException(status_code=409, detail="That username already exists.")
        user_id = f"user_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO users(id, username, display_name, password_hash, role, is_active, created_at, updated_at)
            VALUES(?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                username,
                payload.display_name.strip(),
                hash_password(payload.password),
                payload.role,
                1 if payload.is_active else 0,
                timestamp,
                timestamp,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="admin_user_created",
            target_table="users",
            target_record_id=user_id,
            new_value=username,
        )
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return {"user": serialize_admin_user(row)}


@app.patch("/admin/users/{user_id}")
async def update_admin_user(
    user_id: str,
    payload: AdminUserUpdateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    fields = payload_field_names(payload)
    timestamp = now_iso()
    actor = get_actor(user)
    with db_connection() as conn:
        current = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Account not found.")
        ensure_not_last_active_super_admin(
            conn,
            user_id=user_id,
            next_role=payload.role if "role" in fields else None,
            next_is_active=payload.is_active if "is_active" in fields else None,
        )
        updates: List[str] = []
        params: List[Any] = []
        if "display_name" in fields and payload.display_name is not None:
            updates.append("display_name = ?")
            params.append(payload.display_name.strip())
        if "role" in fields and payload.role is not None:
            updates.append("role = ?")
            params.append(payload.role)
        if "is_active" in fields and payload.is_active is not None:
            updates.append("is_active = ?")
            params.append(1 if payload.is_active else 0)
        if updates:
            updates.append("updated_at = ?")
            params.extend([timestamp, user_id])
            conn.execute(f"UPDATE users SET {', '.join(updates)} WHERE id = ?", params)
            if "is_active" in fields and payload.is_active is False:
                conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
            write_audit_log(
                conn,
                request,
                actor,
                action_type="admin_user_updated",
                target_table="users",
                target_record_id=user_id,
                old_value=current.get("display_name") or current.get("username") or "",
                new_value=json_dumps(payload.dict(exclude_unset=True)),
            )
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return {"user": serialize_admin_user(row)}


@app.post("/admin/users/{user_id}/reset-password")
async def reset_admin_user_password(
    user_id: str,
    payload: AdminUserResetPasswordRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    timestamp = now_iso()
    actor = get_actor(user)
    with db_connection() as conn:
        current = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Account not found.")
        conn.execute(
            "UPDATE users SET password_hash = ?, updated_at = ? WHERE id = ?",
            (hash_password(payload.password), timestamp, user_id),
        )
        conn.execute("DELETE FROM sessions WHERE user_id = ?", (user_id,))
        write_audit_log(
            conn,
            request,
            actor,
            action_type="admin_user_password_reset",
            target_table="users",
            target_record_id=user_id,
            new_value=current.get("username") or "",
        )
        row = conn.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    return {"user": serialize_admin_user(row)}


@app.get("/dashboard/overview")
async def dashboard_overview(user: dict = Depends(get_current_user)):
    actor = get_actor(user)
    workbook_path = resolve_master_excel_path()
    workbook_exists = workbook_path.exists()
    staging_workbook_path = resolve_staging_excel_path()
    staging_workbook_exists = staging_workbook_path.exists()
    with db_connection() as conn:
        overview = {
            "master_buildings": conn.execute(
                "SELECT COUNT(*) AS total FROM master_building_info"
            ).fetchone()["total"],
            "staging_buildings": conn.execute(
                "SELECT COUNT(*) AS total FROM staging_building_info"
            ).fetchone()["total"],
            "staging_pending": None
            if actor.role == "viewer"
            else conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM staging_update_requests
                WHERE review_status IN ('pending', 'ai_parsed', 'employee_submitted', 'needs_more_info', 'conflict')
                """
            ).fetchone()["total"],
            "import_batches": conn.execute(
                "SELECT COUNT(*) AS total FROM import_batches"
            ).fetchone()["total"],
            "source_documents": conn.execute(
                "SELECT COUNT(*) AS total FROM source_documents"
            ).fetchone()["total"],
            "audit_logs": conn.execute(
                "SELECT COUNT(*) AS total FROM audit_logs"
            ).fetchone()["total"],
            "master_excel_exists": workbook_exists,
            "master_excel_path": str(workbook_path),
            "staging_excel_exists": staging_workbook_exists,
            "staging_excel_path": str(staging_workbook_path),
        }
    return overview


@app.get("/crm/service-templates")
async def list_crm_service_templates(
    include_inactive: bool = Query(default=False),
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    include_inactive = include_inactive and actor.role in {"super_admin", "admin"}
    with db_connection() as conn:
        templates = [
            serialize_crm_template(conn, row)
            for row in list_crm_template_rows(conn, include_inactive=include_inactive)
        ]
    return {"templates": templates}


@app.post("/crm/service-templates")
async def create_crm_service_template(
    payload: CrmTemplateUpsertRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    timestamp = now_iso()
    service_key = (payload.service_key or slugify_field_key(payload.name)).strip()
    if not service_key:
        raise HTTPException(status_code=400, detail="The service-line key cannot be empty.")
    with db_connection() as conn:
        existing = conn.execute(
            "SELECT id FROM crm_service_templates WHERE service_key = ?",
            (service_key,),
        ).fetchone()
        if existing:
            raise HTTPException(status_code=409, detail="That service-line key already exists.")
        template_id = f"crm_tpl_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO crm_service_templates(
              id, service_key, name, description, category, active, display_order,
              config_json, created_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                template_id,
                service_key,
                payload.name.strip(),
                payload.description.strip(),
                payload.category.strip() or "general",
                1 if payload.active else 0,
                payload.display_order,
                json_dumps(payload.config),
                actor.user_id,
                timestamp,
                timestamp,
            ),
        )
        for step in payload.steps:
            conn.execute(
                """
                INSERT INTO crm_service_steps(
                  id, template_id, step_key, title, scope, field_schema_json,
                  display_order, active, created_at, updated_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    f"crm_step_{uuid.uuid4().hex}",
                    template_id,
                    step.step_key.strip(),
                    step.title.strip(),
                    step.scope,
                    json_dumps(step.field_schema),
                    step.display_order,
                    1 if step.active else 0,
                    timestamp,
                    timestamp,
                ),
            )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_template_created",
            target_table="crm_service_templates",
            target_record_id=template_id,
            field_name=service_key,
            new_value=payload.name,
        )
        template = conn.execute("SELECT * FROM crm_service_templates WHERE id = ?", (template_id,)).fetchone()
        return {"template": serialize_crm_template(conn, template)}


@app.patch("/crm/service-templates/{template_id}")
async def update_crm_service_template(
    template_id: str,
    payload: CrmTemplateUpsertRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        current = conn.execute(
            "SELECT * FROM crm_service_templates WHERE id = ?",
            (template_id,),
        ).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="CRM service template not found.")
        conn.execute(
            """
            UPDATE crm_service_templates
            SET name = ?, description = ?, category = ?, active = ?, display_order = ?,
                config_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                payload.name.strip(),
                payload.description.strip(),
                payload.category.strip() or "general",
                1 if payload.active else 0,
                payload.display_order,
                json_dumps(payload.config),
                timestamp,
                template_id,
            ),
        )
        if payload.steps:
            conn.execute("DELETE FROM crm_service_steps WHERE template_id = ?", (template_id,))
            for step in payload.steps:
                conn.execute(
                    """
                    INSERT INTO crm_service_steps(
                      id, template_id, step_key, title, scope, field_schema_json,
                      display_order, active, created_at, updated_at
                    ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        f"crm_step_{uuid.uuid4().hex}",
                        template_id,
                        step.step_key.strip(),
                        step.title.strip(),
                        step.scope,
                        json_dumps(step.field_schema),
                        step.display_order,
                        1 if step.active else 0,
                        timestamp,
                        timestamp,
                    ),
                )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_template_updated",
            target_table="crm_service_templates",
            target_record_id=template_id,
            field_name=current["service_key"],
            old_value=current["name"],
            new_value=payload.name,
        )
        template = conn.execute("SELECT * FROM crm_service_templates WHERE id = ?", (template_id,)).fetchone()
        return {"template": serialize_crm_template(conn, template)}


@app.get("/crm/cases")
async def list_crm_cases(
    scope: str = Query(default="my"),
    q: str = Query(default=""),
    status: str = Query(default=""),
    owner_user_id: str = Query(default=""),
    include_deleted: bool = Query(default=False),
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        ensure_crm_default_templates(conn)
        sql = """
            SELECT
              crm_cases.*,
              owner.display_name AS owner_display_name,
              owner.username AS owner_username,
              creator.display_name AS creator_display_name,
              creator.username AS creator_username,
              (
                SELECT COUNT(*)
                FROM crm_tasks
                WHERE crm_tasks.case_id = crm_cases.id
                  AND crm_tasks.status NOT IN ('completed', 'cancelled', 'done')
              ) AS open_task_count
            FROM crm_cases
            LEFT JOIN users AS owner ON owner.id = crm_cases.owner_user_id
            LEFT JOIN users AS creator ON creator.id = crm_cases.created_by
            WHERE 1 = 1
        """
        params: List[Any] = []
        normalized_status = status.strip()
        if normalized_status == "deleted":
            if actor.role not in {"super_admin", "admin"}:
                raise HTTPException(status_code=403, detail="Only administrators can view deleted Cases.")
            sql += " AND COALESCE(crm_cases.deleted_at, '') != ''"
        elif not include_deleted:
            sql += " AND COALESCE(crm_cases.deleted_at, '') = ''"
        if actor.role != "super_admin" or scope != "all":
            sql += " AND crm_cases.owner_user_id = ?"
            params.append(actor.user_id)
        elif owner_user_id.strip():
            sql += " AND crm_cases.owner_user_id = ?"
            params.append(owner_user_id.strip())
        if normalized_status and normalized_status != "deleted":
            sql += " AND crm_cases.status = ?"
            params.append(normalized_status)
        if q.strip():
            query = f"%{q.strip().lower()}%"
            sql += """
                AND (
                  lower(crm_cases.group_name) LIKE ?
                  OR lower(crm_cases.building_name) LIKE ?
                  OR lower(crm_cases.building_address) LIKE ?
                  OR lower(crm_cases.group_creator_name) LIKE ?
                )
            """
            params.extend([query, query, query, query])
        sql += " ORDER BY crm_cases.created_at DESC LIMIT 300"
        rows = conn.execute(sql, params).fetchall()
        if actor.role == "super_admin":
            owners = conn.execute(
                """
                SELECT id, display_name, username, role
                FROM users
                WHERE is_active = 1 AND role IN ('super_admin', 'admin', 'employee')
                ORDER BY display_name ASC, username ASC
                """
            ).fetchall()
        else:
            owners = conn.execute(
                """
                SELECT id, display_name, username, role
                FROM users
                WHERE id = ?
                """,
                (actor.user_id,),
            ).fetchall()
    return {"cases": [serialize_crm_case(row) for row in rows], "owners": owners}


@app.post("/crm/cases")
async def create_crm_case(
    payload: CrmCaseCreateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        normalized_guests = normalize_crm_guest_payloads(payload.guests)
        owner_user_id = normalize_crm_owner(conn, actor, payload.owner_user_id)
        building_snapshot: dict = {}
        building_name = ""
        building_address = ""
        if not payload.building_source or not payload.building_id.strip():
            raise HTTPException(status_code=400, detail="A new CRM Case must be linked to a Master or Staging building.")
        building_snapshot = build_crm_building_snapshot(
            conn,
            payload.building_source,
            payload.building_id.strip(),
        )
        building_name = building_snapshot.get("building_name") or ""
        building_address = building_snapshot.get("address") or ""
        case_id = f"crm_case_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO crm_cases(
              id, group_name, owner_user_id, unit, group_creator_name, group_creator_contact, agent_team_t, agent_team_m,
              lease_start_date, building_source, building_id, building_name, building_address,
              building_snapshot_json, insurance_earliest_start_date, network_earliest_start_note,
              status, notes, created_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'active', ?, ?, ?, ?)
            """,
            (
                case_id,
                payload.group_name.strip(),
                owner_user_id,
                payload.unit.strip(),
                payload.group_creator_name.strip(),
                payload.group_creator_contact.strip(),
                payload.agent_team_t.strip(),
                payload.agent_team_m.strip(),
                payload.lease_start_date.strip(),
                payload.building_source if building_snapshot else "",
                payload.building_id.strip() if building_snapshot else "",
                building_name,
                building_address,
                json_dumps(building_snapshot),
                subtract_months_from_date(payload.lease_start_date, 3),
                payload.network_earliest_start_note.strip(),
                payload.notes.strip(),
                actor.user_id,
                timestamp,
                timestamp,
            ),
        )
        for guest in normalized_guests:
            conn.execute(
                """
                INSERT INTO crm_case_guests(
                  id, case_id, full_name, phone, email, wechat, notes, status, created_at, updated_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
                """,
                (
                    f"crm_guest_{uuid.uuid4().hex}",
                    case_id,
                    guest["full_name"],
                    guest["phone"],
                    guest["email"],
                    guest["wechat"],
                    guest["notes"],
                    timestamp,
                    timestamp,
                ),
            )
        ensure_crm_case_services(conn, case_id=case_id, building_snapshot=building_snapshot)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_case_created",
            target_table="crm_cases",
            target_record_id=case_id,
            building_name=building_name or payload.group_name.strip(),
            new_value=payload.group_name.strip(),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_case_guests_created",
            target_table="crm_case_guests",
            target_record_id=case_id,
            building_name=building_name or payload.group_name.strip(),
            new_value=", ".join(guest["full_name"] for guest in normalized_guests),
            note=f"Created {len(normalized_guests)} guests with case.",
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.get("/crm/cases/{case_id}")
async def get_crm_case(case_id: str, user: dict = Depends(get_current_user)):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        return load_crm_case_detail(conn, case_id, actor)


@app.get("/crm/cases/{case_id}/summary")
async def get_crm_case_summary(case_id: str, user: dict = Depends(get_current_user)):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        detail = load_crm_case_detail(conn, case_id, actor)
    fact_summary = build_crm_case_fact_summary(detail)
    ai_summary = await generate_fact_explanation(
        question="Create a concise service summary for this CRM Case. Highlight what must be completed, what is not required, and what still needs confirmation.",
        snapshot=build_crm_case_ai_snapshot(detail),
        field_keys=[
            "crm_group_name",
            "crm_owner",
            "crm_lease_start_date",
            "crm_guests_count",
            "building_name",
            "address",
            "insurance_required",
            "insurance_coverage_amount",
            "electricity_required",
            "electricity_provider",
            "internet_self_setup_required",
            "internet_provider",
            "internet_notes",
            "move_in_notes",
            "crm_service_statuses",
        ],
        fact_answer=fact_summary,
        source_mode="crm",
    )
    return {
        "case_id": case_id,
        "fact_summary": fact_summary,
        "ai_summary": ai_summary,
        "ai_enabled": ai_explanation_enabled(),
        "message": "CRM Case service summary generated.",
    }


@app.patch("/crm/cases/{case_id}")
async def update_crm_case(
    case_id: str,
    payload: CrmCaseUpdateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    fields = payload_field_names(payload)
    timestamp = now_iso()
    with db_connection() as conn:
        current = require_crm_case_access(conn, case_id, actor)
        updates: List[str] = []
        params: List[Any] = []

        def set_column(column: str, value: Any) -> None:
            updates.append(f"{column} = ?")
            params.append(value)

        if "group_name" in fields and payload.group_name is not None:
            next_name = payload.group_name.strip()
            if not next_name:
                raise HTTPException(status_code=400, detail="The group name cannot be empty.")
            set_column("group_name", next_name)
        if "owner_user_id" in fields and payload.owner_user_id is not None:
            if actor.role != "super_admin":
                raise HTTPException(status_code=403, detail="Only a Super Admin can change the owner.")
            set_column("owner_user_id", normalize_crm_owner(conn, actor, payload.owner_user_id))
        for column, value in [
            ("unit", payload.unit),
            ("group_creator_name", payload.group_creator_name),
            ("group_creator_contact", payload.group_creator_contact),
            ("agent_team_t", payload.agent_team_t),
            ("agent_team_m", payload.agent_team_m),
            ("lease_start_date", payload.lease_start_date),
            ("insurance_earliest_start_date", payload.insurance_earliest_start_date),
            ("network_earliest_start_note", payload.network_earliest_start_note),
            ("status", payload.status),
            ("notes", payload.notes),
        ]:
            if column in fields and value is not None:
                set_column(column, value.strip() if isinstance(value, str) else value)

        building_snapshot = json_loads_safe(current.get("building_snapshot_json"), {})
        building_source = payload.building_source if "building_source" in fields else current.get("building_source")
        building_id = payload.building_id if "building_id" in fields else current.get("building_id")
        if "building_source" in fields or "building_id" in fields:
            if building_source and building_id:
                building_snapshot = build_crm_building_snapshot(conn, building_source, building_id.strip())
                set_column("building_source", building_source)
                set_column("building_id", building_id.strip())
                set_column("building_name", building_snapshot.get("building_name") or "")
                set_column("building_address", building_snapshot.get("address") or "")
                set_column("building_snapshot_json", json_dumps(building_snapshot))
            else:
                building_snapshot = {}
                set_column("building_source", "")
                set_column("building_id", "")
                set_column("building_name", "")
                set_column("building_address", "")
                set_column("building_snapshot_json", "{}")

        next_lease = payload.lease_start_date if "lease_start_date" in fields else current.get("lease_start_date")
        if (
            "lease_start_date" in fields
            and "insurance_earliest_start_date" not in fields
            and not normalize_unknown_value(current.get("insurance_earliest_start_date"))
        ):
            set_column("insurance_earliest_start_date", subtract_months_from_date(next_lease or "", 3))

        if updates:
            set_column("updated_at", timestamp)
            params.append(case_id)
            conn.execute(f"UPDATE crm_cases SET {', '.join(updates)} WHERE id = ?", params)
            ensure_crm_case_services(conn, case_id=case_id, building_snapshot=building_snapshot)
            write_audit_log(
                conn,
                request,
                actor,
                action_type="crm_case_updated",
                target_table="crm_cases",
                target_record_id=case_id,
                building_name=(building_snapshot or {}).get("building_name") or current.get("building_name") or "",
                old_value=current.get("group_name") or "",
                new_value=payload.group_name or current.get("group_name") or "",
            )
        return load_crm_case_detail(conn, case_id, actor)


@app.delete("/crm/cases/{case_id}")
async def delete_crm_case(
    case_id: str,
    payload: CrmCaseDeleteRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    reason = payload.reason.strip()
    if not reason:
        raise HTTPException(status_code=400, detail="A reason is required to delete a Case.")
    timestamp = now_iso()
    with db_connection() as conn:
        current = crm_case_row(conn, case_id)
        if not current:
            raise HTTPException(status_code=404, detail="CRM Case not found.")
        if actor.role not in {"super_admin", "admin"} and not crm_actor_can_access_case(actor, current):
            raise HTTPException(status_code=403, detail="You can delete only CRM Cases assigned to you.")
        if crm_case_is_deleted(current):
            return {"case": serialize_crm_case(current), "message": "The Case has already been deleted."}
        conn.execute(
            """
            UPDATE crm_tasks
            SET status = 'cancelled', updated_at = ?
            WHERE case_id = ?
              AND status NOT IN ('completed', 'cancelled', 'done')
            """,
            (timestamp, case_id),
        )
        conn.execute(
            """
            UPDATE crm_cases
            SET status = 'cancelled',
                deleted_at = ?,
                deleted_by = ?,
                delete_reason = ?,
                updated_at = ?
            WHERE id = ?
            """,
            (timestamp, actor.user_id, reason, timestamp, case_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_case_deleted",
            target_table="crm_cases",
            target_record_id=case_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
            old_value=current.get("status") or "",
            new_value="cancelled",
            note=reason,
        )
        updated = crm_case_row(conn, case_id) or current
        return {"case": serialize_crm_case(updated), "message": "The CRM Case was deleted and archived."}


@app.post("/crm/cases/{case_id}/restore")
async def restore_crm_case(
    case_id: str,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    if actor.role not in {"super_admin", "admin"}:
        raise HTTPException(status_code=403, detail="Only administrators can restore deleted Cases.")
    timestamp = now_iso()
    with db_connection() as conn:
        current = crm_case_row(conn, case_id)
        if not current:
            raise HTTPException(status_code=404, detail="CRM Case not found.")
        conn.execute(
            """
            UPDATE crm_cases
            SET status = 'active',
                deleted_at = '',
                deleted_by = '',
                delete_reason = '',
                updated_at = ?
            WHERE id = ?
            """,
            (timestamp, case_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_case_restored",
            target_table="crm_cases",
            target_record_id=case_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
            old_value=current.get("status") or "",
            new_value="active",
            note="Restored a deleted CRM Case.",
        )
        updated = crm_case_row(conn, case_id) or current
        return {"case": serialize_crm_case(updated), "message": "The CRM Case was restored."}


@app.post("/crm/cases/{case_id}/refresh-building-snapshot")
async def refresh_crm_building_snapshot(
    case_id: str,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        current = require_crm_case_access(conn, case_id, actor)
        if not current.get("building_source") or not current.get("building_id"):
            raise HTTPException(status_code=400, detail="This Case is not linked to a building.")
        snapshot = build_crm_building_snapshot(conn, current["building_source"], current["building_id"])
        timestamp = now_iso()
        conn.execute(
            """
            UPDATE crm_cases
            SET building_name = ?, building_address = ?, building_snapshot_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                snapshot.get("building_name") or "",
                snapshot.get("address") or "",
                json_dumps(snapshot),
                timestamp,
                case_id,
            ),
        )
        ensure_crm_case_services(conn, case_id=case_id, building_snapshot=snapshot)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_case_snapshot_refreshed",
            target_table="crm_cases",
            target_record_id=case_id,
            building_name=snapshot.get("building_name") or "",
            source=snapshot.get("source_label") or "",
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.get("/crm/cases/{case_id}/guests")
async def list_crm_case_guests(case_id: str, user: dict = Depends(get_current_user)):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        return {"guests": load_crm_case_detail(conn, case_id, actor)["guests"]}


@app.post("/crm/cases/{case_id}/guests")
async def create_crm_case_guest(
    case_id: str,
    payload: CrmGuestCreateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        require_crm_case_access(conn, case_id, actor)
        guest_id = f"crm_guest_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO crm_case_guests(
              id, case_id, full_name, phone, email, wechat, notes, status, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, 'active', ?, ?)
            """,
            (
                guest_id,
                case_id,
                payload.full_name.strip(),
                payload.phone.strip(),
                payload.email.strip(),
                payload.wechat.strip(),
                payload.notes.strip(),
                timestamp,
                timestamp,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_guest_created",
            target_table="crm_case_guests",
            target_record_id=guest_id,
            new_value=payload.full_name.strip(),
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.patch("/crm/cases/{case_id}/guests/{guest_id}")
async def update_crm_case_guest(
    case_id: str,
    guest_id: str,
    payload: CrmGuestUpdateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    fields = payload_field_names(payload)
    timestamp = now_iso()
    with db_connection() as conn:
        require_crm_case_access(conn, case_id, actor)
        current = conn.execute(
            "SELECT * FROM crm_case_guests WHERE id = ? AND case_id = ?",
            (guest_id, case_id),
        ).fetchone()
        if not current:
            raise HTTPException(status_code=404, detail="Customer record not found.")
        updates: List[str] = []
        params: List[Any] = []
        for column, value in [
            ("full_name", payload.full_name),
            ("phone", payload.phone),
            ("email", payload.email),
            ("wechat", payload.wechat),
            ("notes", payload.notes),
            ("status", payload.status),
        ]:
            if column in fields and value is not None:
                updates.append(f"{column} = ?")
                params.append(value.strip() if isinstance(value, str) else value)
        if updates:
            updates.append("updated_at = ?")
            params.extend([timestamp, guest_id])
            conn.execute(f"UPDATE crm_case_guests SET {', '.join(updates)} WHERE id = ?", params)
            write_audit_log(
                conn,
                request,
                actor,
                action_type="crm_guest_updated",
                target_table="crm_case_guests",
                target_record_id=guest_id,
                old_value=current.get("full_name") or "",
                new_value=payload.full_name or current.get("full_name") or "",
            )
        return load_crm_case_detail(conn, case_id, actor)


@app.get("/crm/cases/{case_id}/services/{service_id}/progress")
async def get_crm_service_progress(
    case_id: str,
    service_id: str,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        detail = load_crm_case_detail(conn, case_id, actor)
        service = next(
            (
                item
                for item in detail["services"]
                if item["id"] == service_id or item["service_key"] == service_id
            ),
            None,
        )
        if not service:
            raise HTTPException(status_code=404, detail="CRM service line not found.")
        return {"service": service}


@app.patch("/crm/cases/{case_id}/services/{service_id}/progress")
async def update_crm_service_progress(
    case_id: str,
    service_id: str,
    payload: CrmServiceProgressUpdateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        require_crm_case_access(conn, case_id, actor)
        service = crm_service_row_for_case(conn, case_id, service_id)
        if not service:
            raise HTTPException(status_code=404, detail="CRM service line not found.")
        group_progress = json_loads_safe(service.get("group_progress_json"), {})
        for key, value in payload.group_progress.items():
            group_progress[key] = value
        if payload.step_key.strip():
            step_key = payload.step_key.strip()
            group_progress[step_key] = payload.value
            existing_progress = conn.execute(
                """
                SELECT id
                FROM crm_case_service_progress
                WHERE case_service_id = ? AND step_key = ? AND scope = 'group'
                """,
                (service["id"], step_key),
            ).fetchone()
            if existing_progress:
                conn.execute(
                    """
                    UPDATE crm_case_service_progress
                    SET value_json = ?, note = ?, updated_by = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        json_dumps(payload.value),
                        payload.note.strip(),
                        actor.user_id,
                        timestamp,
                        existing_progress["id"],
                    ),
                )
            else:
                conn.execute(
                    """
                    INSERT INTO crm_case_service_progress(
                      id, case_service_id, step_key, scope, value_json, note, updated_by, updated_at
                    ) VALUES(?, ?, ?, 'group', ?, ?, ?, ?)
                    """,
                    (
                        f"crm_progress_{uuid.uuid4().hex}",
                        service["id"],
                        step_key,
                        json_dumps(payload.value),
                        payload.note.strip(),
                        actor.user_id,
                        timestamp,
                    ),
                )

        next_status = payload.service_status or payload.status or service.get("service_status") or service["status"]
        channel_check = group_progress.get("channel_check") or {}
        if service["service_key"] == "phone_card":
            purchased = channel_check.get("other_channel_purchased") or group_progress.get("other_channel_purchased")
            normalized_purchased = str(purchased or "").strip().lower()
            if normalized_purchased in {"是", "yes", "true", "1"}:
                next_status = "terminated"
            elif normalized_purchased in {"否", "no", "false", "0"} and service["status"] == "terminated" and not payload.status:
                next_status = "pending"
        next_applicability = payload.applicability or service["applicability"]
        responsible_customer_id = (
            payload.responsible_customer_id.strip()
            if payload.responsible_customer_id is not None
            else service.get("responsible_customer_id") or ""
        )
        if responsible_customer_id:
            customer = conn.execute(
                "SELECT id FROM crm_case_guests WHERE id = ? AND case_id = ?",
                (responsible_customer_id, case_id),
            ).fetchone()
            if not customer:
                raise HTTPException(status_code=400, detail="The responsible customer does not belong to this Case.")
        service_scope = payload.service_scope or service.get("service_scope") or crm_service_scope(service["service_key"])
        customer_rows = crm_active_customer_rows(conn, case_id)
        need_status = crm_normalize_need_status(payload.need_status or service.get("need_status") or "")
        submission_status = crm_normalize_submission_status(payload.submission_status or service.get("submission_status") or "")
        completion_status = crm_normalize_completion_status(payload.completion_status or service.get("completion_status") or "")
        if payload.status and not payload.completion_status:
            completion_status = crm_completion_from_service_status(next_status, completion_status)
        if payload.completion_status and not payload.status:
            if payload.completion_status == "completed":
                next_status = "completed"
            elif payload.completion_status == "waived":
                next_status = "not_needed"
            elif payload.completion_status == "not_applicable":
                next_status = "not_needed"
            elif payload.completion_status == "not_started":
                next_status = "pending"
            elif payload.completion_status == "in_progress":
                next_status = "in_progress"
            elif payload.completion_status == "failed":
                next_status = "terminated"
            elif payload.completion_status == "unknown":
                next_status = "pending"
        intro_status = crm_normalize_intro_status(payload.intro_status or service.get("intro_status") or "")
        follow_up_status = crm_normalize_follow_up_status(payload.follow_up_status or service.get("follow_up_status") or "")
        agent_completion_status = crm_normalize_agent_completion_status(
            payload.agent_completion_status or service.get("agent_completion_status") or ""
        )
        legacy_like_service = {
            **dict(service),
            "status": next_status,
            "need_status": need_status,
            "submission_status": submission_status,
            "completion_status": completion_status,
            "intro_status": intro_status,
            "follow_up_status": follow_up_status,
            "agent_completion_status": agent_completion_status,
        }
        legacy_flow_fields = crm_flows_from_legacy(legacy_like_service, next_applicability)
        legacy_status_payload_present = any(
            value is not None
            for value in (
                payload.status,
                payload.need_status,
                payload.submission_status,
                payload.completion_status,
                payload.intro_status,
                payload.follow_up_status,
                payload.agent_completion_status,
            )
        )
        template_for_service = conn.execute(
            "SELECT * FROM crm_service_templates WHERE id = ?",
            (service["template_id"],),
        ).fetchone()
        flow_snapshot = crm_effective_service_flow_snapshot(service.get("flow_snapshot_json"), template_for_service)
        active_flow_payload_present = payload.active_flow_step_key is not None
        staff_flow_payload_present = payload.staff_flow_status is not None
        customer_flow_payload_present = payload.customer_flow_status is not None
        active_flow_step_key = (
            payload.active_flow_step_key.strip()
            if payload.active_flow_step_key is not None
            else service.get("active_flow_step_key") or ""
        )
        staff_flow_status = crm_normalize_staff_flow_status(
            payload.staff_flow_status
            or ("" if legacy_status_payload_present else service.get("staff_flow_status") or ""),
            legacy_flow_fields["staff_flow_status"],
        )
        customer_flow_status = crm_normalize_customer_flow_status_for_profile(
            payload.customer_flow_status
            or ("" if legacy_status_payload_present else service.get("customer_flow_status") or ""),
            legacy_flow_fields["customer_flow_status"],
            flow_snapshot,
        )
        if active_flow_payload_present:
            step_fields = crm_flow_fields_from_step_key(
                flow_snapshot,
                active_flow_step_key,
                applicability=next_applicability,
                fallback_staff_flow_status=staff_flow_status,
                fallback_customer_flow_status=customer_flow_status,
            )
            active_flow_step_key = step_fields["active_flow_step_key"]
            staff_flow_status = step_fields["staff_flow_status"]
            if not customer_flow_payload_present:
                customer_flow_status = step_fields["customer_flow_status"]
            next_status = step_fields["service_status"]
        elif staff_flow_payload_present:
            active_flow_step_key = crm_default_flow_step_key(flow_snapshot, staff_flow_status)
        elif not active_flow_step_key and not legacy_status_payload_present:
            active_flow_step_key = crm_default_flow_step_key(flow_snapshot, staff_flow_status)
        if staff_flow_payload_present and not customer_flow_payload_present and not active_flow_payload_present:
            customer_flow_status = crm_customer_flow_from_staff_flow(
                staff_flow_status,
                customer_flow_status,
                flow_snapshot,
            )
        termination_reason = (
            payload.termination_reason.strip()
            if payload.termination_reason is not None
            else service.get("termination_reason") or ""
        )
        if next_applicability == "not_needed":
            customer_flow_status = "not_needed"
            staff_flow_status = "completed"
            active_flow_step_key = ""
        customer_base_status = crm_customer_flow_base_status(customer_flow_status, flow_snapshot)
        if customer_base_status == "declined":
            if not termination_reason:
                raise HTTPException(status_code=400, detail="Enter a termination reason when the customer declines the service.")
        if staff_flow_status == "terminated" and customer_base_status not in {"declined", "not_needed"}:
            customer_flow_status = "declined"
            customer_base_status = "declined"
            if not termination_reason:
                raise HTTPException(status_code=400, detail="Enter a termination reason before terminating the service.")
        next_status = crm_normalize_service_status(
            payload.service_status or payload.status or (next_status if active_flow_payload_present else ""),
            crm_service_status_from_flows(staff_flow_status, customer_flow_status, next_applicability, flow_snapshot),
        )
        legacy_fields_from_flow = crm_legacy_fields_from_flows(
            service_key=service["service_key"],
            applicability=next_applicability,
            staff_flow_status=staff_flow_status,
            customer_flow_status=customer_flow_status,
            current={
                **dict(service),
                "need_status": need_status,
                "submission_status": submission_status,
                "completion_status": completion_status,
            },
            flow_profile=flow_snapshot,
        )
        need_status = legacy_fields_from_flow["need_status"]
        submission_status = legacy_fields_from_flow["submission_status"]
        completion_status = legacy_fields_from_flow["completion_status"]
        intro_status = legacy_fields_from_flow["intro_status"]
        follow_up_status = legacy_fields_from_flow["follow_up_status"]
        agent_completion_status = legacy_fields_from_flow["agent_completion_status"]
        responsibility = crm_case_service_responsibility_state(
            customer_rows,
            service_scope=service_scope,
            applicability=next_applicability,
            completion_status=completion_status,
            current_responsible_customer_id=responsible_customer_id,
            current_responsibility_status=payload.responsibility_status or service.get("responsibility_status") or "",
        )
        responsible_changed = payload.responsible_customer_id is not None and (
            responsible_customer_id != (service.get("responsible_customer_id") or "")
        )
        if responsible_changed and responsibility["responsible_customer_id"]:
            responsibility["responsibility_status"] = "assigned"
            if payload.agent_completion_status is None and agent_completion_status == "open":
                agent_completion_status = "pending_customer"
        if crm_service_requires_responsible(service_scope, next_applicability, completion_status):
            if completion_status == "completed" and not responsibility["responsible_customer_id"]:
                raise HTTPException(status_code=400, detail="Assign a responsible customer for this group-level service before marking it complete.")
        responsible_customer_id = responsibility["responsible_customer_id"]
        blocked_reason = (
            payload.blocked_reason.strip()
            if payload.blocked_reason is not None
            else service.get("blocked_reason") or ""
        )
        old_status_snapshot = {
            "status": service.get("status") or "",
            "service_status": service.get("service_status") or "",
            "active_flow_step_key": service.get("active_flow_step_key") or "",
            "staff_flow_status": service.get("staff_flow_status") or "",
            "customer_flow_status": service.get("customer_flow_status") or "",
            "need_status": service.get("need_status") or "",
            "submission_status": service.get("submission_status") or "",
            "completion_status": service.get("completion_status") or "",
            "intro_status": service.get("intro_status") or "",
            "follow_up_status": service.get("follow_up_status") or "",
            "agent_completion_status": service.get("agent_completion_status") or "",
            "responsible_customer_id": service.get("responsible_customer_id") or "",
            "responsibility_status": service.get("responsibility_status") or "",
            "termination_reason": service.get("termination_reason") or "",
        }
        new_status_snapshot = {
            "status": next_status,
            "service_status": next_status,
            "active_flow_step_key": active_flow_step_key,
            "staff_flow_status": staff_flow_status,
            "customer_flow_status": customer_flow_status,
            "need_status": need_status,
            "submission_status": submission_status,
            "completion_status": completion_status,
            "intro_status": intro_status,
            "follow_up_status": follow_up_status,
            "agent_completion_status": agent_completion_status,
            "responsible_customer_id": responsible_customer_id,
            "responsibility_status": responsibility["responsibility_status"],
            "termination_reason": termination_reason,
        }
        conn.execute(
            """
            UPDATE crm_case_services
            SET status = ?, applicability = ?, service_scope = ?, responsible_customer_id = ?,
                covered_customer_ids = ?, responsibility_status = ?,
                active_flow_step_key = ?, staff_flow_status = ?, customer_flow_status = ?, service_status = ?, termination_reason = ?,
                need_status = ?, submission_status = ?, completion_status = ?, intro_status = ?,
                follow_up_status = ?, agent_completion_status = ?, blocked_reason = ?,
                group_progress_json = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                next_status,
                next_applicability,
                service_scope,
                responsible_customer_id,
                json_dumps(responsibility["covered_customer_ids"]),
                responsibility["responsibility_status"],
                active_flow_step_key,
                staff_flow_status,
                customer_flow_status,
                next_status,
                termination_reason,
                need_status,
                submission_status,
                completion_status,
                intro_status,
                follow_up_status,
                agent_completion_status,
                blocked_reason,
                json_dumps(group_progress),
                timestamp,
                service["id"],
            ),
        )
        if responsible_changed and responsible_customer_id:
            crm_reassign_service_tasks_to_responsible_customer(
                conn,
                case_service_id=service["id"],
                responsible_customer_id=responsible_customer_id,
                customer_rows=customer_rows,
                timestamp=timestamp,
            )
        if next_status in {"completed", "terminated"}:
            cancel_open_crm_service_tasks(conn, service["id"], timestamp)
        ensure_crm_tasks_for_case(conn, case_id)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_service_progress_updated",
            target_table="crm_case_services",
            target_record_id=service["id"],
            field_name=payload.step_key.strip() or service["service_key"],
            old_value=json_dumps(old_status_snapshot),
            new_value=json_dumps(new_status_snapshot),
            note="Service status, owner, and progress fields were updated.",
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.patch("/crm/cases/{case_id}/guests/{guest_id}/services/{service_id}/progress")
async def update_crm_guest_service_progress(
    case_id: str,
    guest_id: str,
    service_id: str,
    payload: CrmGuestProgressUpdateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        case_row = require_crm_case_access(conn, case_id, actor)
        guest = conn.execute(
            "SELECT * FROM crm_case_guests WHERE id = ? AND case_id = ?",
            (guest_id, case_id),
        ).fetchone()
        if not guest:
            raise HTTPException(status_code=404, detail="Customer record not found.")
        service = crm_service_row_for_case(conn, case_id, service_id)
        if not service:
            raise HTTPException(status_code=404, detail="CRM service line not found.")
        step_key = payload.step_key.strip()
        existing = conn.execute(
            """
            SELECT *
            FROM crm_guest_service_progress
            WHERE guest_id = ? AND case_service_id = ? AND step_key = ?
            """,
            (guest_id, service["id"], step_key),
        ).fetchone()
        next_sensitive = payload.sensitive
        if existing and not payload.sensitive:
            next_sensitive = json_loads_safe(existing.get("sensitive_json"), {})
        if existing:
            conn.execute(
                """
                UPDATE crm_guest_service_progress
                SET value_json = ?, note = ?, sensitive_json = ?, updated_by = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    json_dumps(payload.value),
                    payload.note.strip(),
                    json_dumps(next_sensitive),
                    actor.user_id,
                    timestamp,
                    existing["id"],
                ),
            )
            progress_id = existing["id"]
        else:
            progress_id = f"crm_guest_progress_{uuid.uuid4().hex}"
            conn.execute(
                """
                INSERT INTO crm_guest_service_progress(
                  id, guest_id, case_service_id, step_key, value_json, note,
                  sensitive_json, updated_by, updated_at
                ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    progress_id,
                    guest_id,
                    service["id"],
                    step_key,
                    json_dumps(payload.value),
                    payload.note.strip(),
                    json_dumps(next_sensitive),
                    actor.user_id,
                    timestamp,
                ),
            )
        if step_key == "phone_intent":
            sync_phone_card_sent_record_task(
                conn,
                case_row=case_row,
                service_row=service,
                guest_row=guest,
                progress_value=payload.value,
                actor=actor,
                timestamp=timestamp,
            )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_guest_progress_updated",
            target_table="crm_guest_service_progress",
            target_record_id=progress_id,
            field_name=step_key,
            new_value=json_dumps(payload.value),
            note="Contains sensitive fields" if payload.sensitive else "",
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.post("/crm/cases/{case_id}/generate-services")
async def generate_crm_case_services(
    case_id: str,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        current = require_crm_case_access(conn, case_id, actor)
        ensure_crm_case_services(
            conn,
            case_id=case_id,
            building_snapshot=json_loads_safe(current.get("building_snapshot_json"), {}),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_services_generated",
            target_table="crm_case_services",
            target_record_id=case_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
        )
        return load_crm_case_detail(conn, case_id, actor)


def crm_report_case_where(
    *,
    owner_user_id: str = "",
    case_status: str = "",
    building_source: str = "",
    date_from: str = "",
    date_to: str = "",
    alias: str = "c",
) -> tuple[str, List[Any]]:
    clauses: List[str] = [f"COALESCE({alias}.deleted_at, '') = ''"]
    params: List[Any] = []
    if owner_user_id.strip():
        clauses.append(f"{alias}.owner_user_id = ?")
        params.append(owner_user_id.strip())
    if case_status.strip():
        clauses.append(f"{alias}.status = ?")
        params.append(case_status.strip())
    if building_source.strip():
        clauses.append(f"{alias}.building_source = ?")
        params.append(building_source.strip())
    if date_from.strip():
        clauses.append(f"date({alias}.created_at) >= date(?)")
        params.append(date_from.strip())
    if date_to.strip():
        clauses.append(f"date({alias}.created_at) <= date(?)")
        params.append(date_to.strip())
    return " AND ".join(clauses), params


def crm_report_service_where(
    *,
    owner_user_id: str = "",
    case_status: str = "",
    service_type: str = "",
    building_source: str = "",
    date_from: str = "",
    date_to: str = "",
    case_alias: str = "c",
    service_alias: str = "s",
) -> tuple[str, List[Any]]:
    where, params = crm_report_case_where(
        owner_user_id=owner_user_id,
        case_status=case_status,
        building_source=building_source,
        date_from=date_from,
        date_to=date_to,
        alias=case_alias,
    )
    if service_type.strip():
        where += f" AND {service_alias}.service_key = ?"
        params.append(service_type.strip())
    return where, params


def crm_report_task_where(
    *,
    owner_user_id: str = "",
    case_status: str = "",
    service_type: str = "",
    priority: str = "",
    status: str = "",
    building_source: str = "",
    date_from: str = "",
    date_to: str = "",
    case_alias: str = "c",
    task_alias: str = "t",
    service_alias: str = "s",
) -> tuple[str, List[Any]]:
    clauses: List[str] = [f"COALESCE({case_alias}.deleted_at, '') = ''"]
    params: List[Any] = []
    if owner_user_id.strip():
        clauses.append(f"COALESCE(NULLIF({task_alias}.assigned_to, ''), {task_alias}.assigned_user_id) = ?")
        params.append(owner_user_id.strip())
    if case_status.strip():
        clauses.append(f"{case_alias}.status = ?")
        params.append(case_status.strip())
    if service_type.strip():
        clauses.append(f"{service_alias}.service_key = ?")
        params.append(service_type.strip())
    if priority.strip():
        clauses.append(f"{task_alias}.priority = ?")
        params.append(priority.strip())
    if status.strip():
        normalized_status = crm_normalize_task_status(status.strip())
        if normalized_status == "overdue":
            clauses.append(f"{task_alias}.status NOT IN ('completed', 'cancelled', 'done')")
            clauses.append(f"date({task_alias}.due_at) < date('now', 'localtime')")
        else:
            clauses.append(f"{task_alias}.status IN (?, ?)")
            params.extend([
                normalized_status,
                "todo" if normalized_status == "open" else "done" if normalized_status == "completed" else normalized_status,
            ])
    if building_source.strip():
        clauses.append(f"{case_alias}.building_source = ?")
        params.append(building_source.strip())
    if date_from.strip():
        clauses.append(f"date({task_alias}.due_at) >= date(?)")
        params.append(date_from.strip())
    if date_to.strip():
        clauses.append(f"date({task_alias}.due_at) <= date(?)")
        params.append(date_to.strip())
    return " AND ".join(clauses), params


def crm_report_label(mapping: dict, value: Any) -> str:
    text = normalize_unknown_value(value)
    if not text:
        return ""
    return mapping.get(text, text)


def crm_report_owner_name(row: dict) -> str:
    return normalize_unknown_value(row.get("display_name")) or normalize_unknown_value(row.get("username")) or ""


def crm_report_service_delivery_mode(row: dict) -> str:
    flow_snapshot = json_loads_safe(row.get("flow_snapshot_json"), {})
    template_config = json_loads_safe(row.get("template_config_json"), {})
    return (
        normalize_unknown_value(flow_snapshot.get("service_delivery_mode"))
        or normalize_unknown_value(template_config.get("service_delivery_mode"))
        or normalize_unknown_value(row.get("service_delivery_mode"))
    )


def crm_report_task_item(row: dict) -> dict:
    return {
        "id": row.get("id") or "",
        "title": row.get("title") or "",
        "case_id": row.get("case_id") or "",
        "case_group_name": row.get("group_name") or "",
        "service_name": row.get("service_name") or "",
        "service_key": row.get("service_key") or "",
        "customer_name": row.get("customer_name") or "",
        "assigned_to": crm_report_owner_name(
            {
                "display_name": row.get("assigned_to_name"),
                "username": row.get("assigned_username"),
            }
        ),
        "due_at": row.get("due_at") or "",
        "priority": row.get("priority") or "normal",
        "priority_label": crm_report_label(CRM_TASK_PRIORITY_LABELS, row.get("priority") or "normal"),
        "status": crm_normalize_task_status(row.get("status") or "open"),
        "status_label": crm_report_label(CRM_TASK_STATUS_LABELS, crm_normalize_task_status(row.get("status") or "open")),
        "case_status": row.get("case_status") or "",
        "case_status_label": crm_report_label(CRM_CASE_STATUS_LABELS, row.get("case_status") or ""),
    }


def crm_report_open_task_clause(task_alias: str = "t") -> str:
    return f"{task_alias}.status NOT IN ('completed', 'cancelled', 'done')"


def crm_report_open_service_clause(service_alias: str = "s") -> str:
    return (
        f"COALESCE(NULLIF({service_alias}.service_status, ''), {service_alias}.status) "
        "NOT IN ('completed', 'terminated', 'not_needed')"
    )


def crm_export_append_sheet(workbook: Workbook, title: str, headers: List[str], rows: List[List[Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(["" if value is None else value for value in row])
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 42)


@app.get("/crm/analytics")
async def crm_analytics(
    owner_user_id: str = Query(default=""),
    case_status: str = Query(default=""),
    service_type: str = Query(default=""),
    priority: str = Query(default=""),
    building_source: str = Query(default=""),
    date_from: str = Query(default="", alias="from"),
    date_to: str = Query(default="", alias="to"),
    user: dict = Depends(require_roles("super_admin")),
):
    with db_connection() as conn:
        normalize_crm_task_rows(conn)
        case_where, case_params = crm_report_case_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        task_where, task_params = crm_report_task_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            service_type=service_type,
            priority=priority,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        service_where, service_params = crm_report_service_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            service_type=service_type,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        total_cases = conn.execute(f"SELECT COUNT(*) AS total FROM crm_cases c WHERE {case_where}", case_params).fetchone()["total"]
        active_cases = conn.execute(
            f"SELECT COUNT(*) AS total FROM crm_cases c WHERE {case_where} AND c.status = 'active'",
            case_params,
        ).fetchone()["total"]
        completed_cases = conn.execute(
            f"SELECT COUNT(*) AS total FROM crm_cases c WHERE {case_where} AND c.status = 'completed'",
            case_params,
        ).fetchone()["total"]
        overdue_tasks = conn.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM crm_tasks t
            JOIN crm_cases c ON c.id = t.case_id
            LEFT JOIN crm_case_services s ON s.id = t.case_service_id
            WHERE {task_where}
              AND {crm_report_open_task_clause('t')}
              AND date(t.due_at) < date('now', 'localtime')
              AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
            """,
            task_params,
        ).fetchone()["total"]
        high_priority_tasks = conn.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM crm_tasks t
            JOIN crm_cases c ON c.id = t.case_id
            LEFT JOIN crm_case_services s ON s.id = t.case_service_id
            WHERE {task_where}
              AND {crm_report_open_task_clause('t')}
              AND t.priority IN ('high', 'urgent')
              AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
            """,
            task_params,
        ).fetchone()["total"]
        risk_services = conn.execute(
            f"""
            SELECT COUNT(*) AS total
            FROM crm_case_services s
            JOIN crm_cases c ON c.id = s.case_id
            WHERE {service_where}
              AND COALESCE(NULLIF(s.service_status, ''), s.status) = 'at_risk'
            """,
            service_params,
        ).fetchone()["total"]
        service_rows = conn.execute(
            f"""
            SELECT s.service_key, s.service_name,
                   COUNT(*) AS total,
                   SUM(CASE WHEN COALESCE(NULLIF(s.service_status, ''), s.status) = 'completed' THEN 1 ELSE 0 END) AS completed,
                   SUM(CASE WHEN COALESCE(NULLIF(s.service_status, ''), s.status) = 'at_risk' THEN 1 ELSE 0 END) AS at_risk,
                   SUM(CASE WHEN {crm_report_open_service_clause('s')} THEN 1 ELSE 0 END) AS open_count
            FROM crm_case_services s
            JOIN crm_cases c ON c.id = s.case_id
            WHERE {service_where}
            GROUP BY s.service_key, s.service_name
            ORDER BY s.service_name ASC
            """,
            service_params,
        ).fetchall()
        workload_rows = conn.execute(
            f"""
            SELECT COALESCE(NULLIF(t.assigned_to, ''), t.assigned_user_id) AS staff_id,
                   users.display_name, users.username,
                   COUNT(*) AS open_tasks,
                   SUM(CASE WHEN date(t.due_at) < date('now', 'localtime') THEN 1 ELSE 0 END) AS overdue_tasks,
                   SUM(CASE WHEN t.priority IN ('high', 'urgent') THEN 1 ELSE 0 END) AS high_priority_tasks
            FROM crm_tasks t
            JOIN crm_cases c ON c.id = t.case_id
            LEFT JOIN crm_case_services s ON s.id = t.case_service_id
            LEFT JOIN users ON users.id = COALESCE(NULLIF(t.assigned_to, ''), t.assigned_user_id)
            WHERE {task_where}
              AND {crm_report_open_task_clause('t')}
              AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
            GROUP BY staff_id, users.display_name, users.username
            ORDER BY open_tasks DESC, overdue_tasks DESC
            LIMIT 12
            """,
            task_params,
        ).fetchall()
        task_select = """
            SELECT t.*, c.group_name, c.status AS case_status,
                   s.service_key, s.service_name,
                   g.full_name AS customer_name,
                   users.display_name AS assigned_to_name, users.username AS assigned_username
            FROM crm_tasks t
            JOIN crm_cases c ON c.id = t.case_id
            LEFT JOIN crm_case_services s ON s.id = t.case_service_id
            LEFT JOIN crm_case_guests g ON g.id = COALESCE(t.target_customer_id, t.customer_id)
            LEFT JOIN users ON users.id = COALESCE(NULLIF(t.assigned_to, ''), t.assigned_user_id)
        """
        upcoming_rows = conn.execute(
            f"""
            {task_select}
            WHERE {task_where}
              AND {crm_report_open_task_clause('t')}
              AND date(t.due_at) BETWEEN date('now', 'localtime') AND date('now', '+7 days', 'localtime')
              AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
            ORDER BY datetime(t.due_at) ASC, t.priority DESC
            LIMIT 12
            """,
            task_params,
        ).fetchall()
        overdue_rows = conn.execute(
            f"""
            {task_select}
            WHERE {task_where}
              AND {crm_report_open_task_clause('t')}
              AND date(t.due_at) < date('now', 'localtime')
              AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
            ORDER BY datetime(t.due_at) ASC
            LIMIT 12
            """,
            task_params,
        ).fetchall()
        staging_where, staging_params = crm_report_case_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            building_source="staging",
            date_from=date_from,
            date_to=date_to,
        )
        staging_rows = conn.execute(
            f"""
            SELECT c.*, users.display_name, users.username,
                   (
                     SELECT COUNT(*) FROM crm_tasks t
                     LEFT JOIN crm_case_services s ON s.id = t.case_service_id
                     WHERE t.case_id = c.id
                       AND {crm_report_open_task_clause('t')}
                       AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
                   ) AS open_task_count,
                   (
                     SELECT COUNT(*) FROM crm_case_services s
                     WHERE s.case_id = c.id AND COALESCE(NULLIF(s.service_status, ''), s.status) = 'at_risk'
                   ) AS risk_count
            FROM crm_cases c
            LEFT JOIN users ON users.id = c.owner_user_id
            WHERE {staging_where}
            ORDER BY datetime(c.created_at) DESC
            LIMIT 12
            """,
            staging_params,
        ).fetchall()
        owners = conn.execute(
            """
            SELECT id, display_name, username, role
            FROM users
            WHERE is_active = 1 AND role IN ('super_admin', 'admin', 'employee')
            ORDER BY display_name ASC, username ASC
            """
        ).fetchall()
        return {
            "stats": {
                "total_cases": total_cases,
                "active_cases": active_cases,
                "completed_cases": completed_cases,
                "overdue_tasks": overdue_tasks,
                "high_priority_tasks": high_priority_tasks,
                "risk_services": risk_services,
            },
            "service_completion": [
                {
                    "service_key": row.get("service_key") or "",
                    "service_name": row.get("service_name") or "",
                    "total": row.get("total") or 0,
                    "completed": row.get("completed") or 0,
                    "at_risk": row.get("at_risk") or 0,
                    "open": row.get("open_count") or 0,
                    "completion_rate": round(((row.get("completed") or 0) / (row.get("total") or 1)) * 100, 1),
                }
                for row in service_rows
            ],
            "staff_workload": [
                {
                    "staff_id": row.get("staff_id") or "",
                    "staff_name": crm_report_owner_name(row),
                    "open_tasks": row.get("open_tasks") or 0,
                    "overdue_tasks": row.get("overdue_tasks") or 0,
                    "high_priority_tasks": row.get("high_priority_tasks") or 0,
                }
                for row in workload_rows
            ],
            "upcoming_tasks": [crm_report_task_item(row) for row in upcoming_rows],
            "overdue_tasks": [crm_report_task_item(row) for row in overdue_rows],
            "staging_cases": [
                {
                    "id": row.get("id") or "",
                    "group_name": row.get("group_name") or "",
                    "building_name": row.get("building_name") or "",
                    "unit": row.get("unit") or "",
                    "lease_start_date": row.get("lease_start_date") or "",
                    "owner_name": crm_report_owner_name(row),
                    "open_task_count": row.get("open_task_count") or 0,
                    "risk_count": row.get("risk_count") or 0,
                }
                for row in staging_rows
            ],
            "owners": owners,
        }


@app.get("/crm/export.xlsx")
async def crm_export_excel(
    owner_user_id: str = Query(default=""),
    case_status: str = Query(default=""),
    service_type: str = Query(default=""),
    priority: str = Query(default=""),
    building_source: str = Query(default=""),
    date_from: str = Query(default="", alias="from"),
    date_to: str = Query(default="", alias="to"),
    user: dict = Depends(require_roles("super_admin")),
):
    with db_connection() as conn:
        normalize_crm_task_rows(conn)
        case_where, case_params = crm_report_case_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        service_where, service_params = crm_report_service_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            service_type=service_type,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        task_where, task_params = crm_report_task_where(
            owner_user_id=owner_user_id,
            case_status=case_status,
            service_type=service_type,
            priority=priority,
            building_source=building_source,
            date_from=date_from,
            date_to=date_to,
        )
        workbook = Workbook()
        workbook.remove(workbook.active)
        case_rows = conn.execute(
            f"""
            SELECT c.*, users.display_name, users.username,
                   (SELECT COUNT(*) FROM crm_case_guests g WHERE g.case_id = c.id AND g.status != 'inactive') AS customer_count,
                   (
                     SELECT COUNT(*) FROM crm_tasks t
                     LEFT JOIN crm_case_services s ON s.id = t.case_service_id
                     WHERE t.case_id = c.id
                       AND {crm_report_open_task_clause('t')}
                       AND (t.case_service_id IS NULL OR {crm_report_open_service_clause('s')})
                   ) AS open_task_count,
                   (
                     SELECT COUNT(*) FROM crm_case_services s
                     WHERE s.case_id = c.id AND COALESCE(NULLIF(s.service_status, ''), s.status) = 'at_risk'
                   ) AS risk_count
            FROM crm_cases c
            LEFT JOIN users ON users.id = c.owner_user_id
            WHERE {case_where}
            ORDER BY datetime(c.created_at) DESC
            """,
            case_params,
        ).fetchall()
        case_ids = [row["id"] for row in case_rows]
        crm_export_append_sheet(
            workbook,
            "Cases",
            ["WeChat Group / Case", "Building", "Unit", "Lease Start Date", "Owner", "Case Status", "Customers", "Open Tasks", "Risks", "Building Source", "Created At"],
            [
                [
                    row.get("group_name"),
                    row.get("building_name"),
                    row.get("unit"),
                    row.get("lease_start_date"),
                    crm_report_owner_name(row),
                    crm_report_label(CRM_CASE_STATUS_LABELS, row.get("status")),
                    row.get("customer_count"),
                    row.get("open_task_count"),
                    row.get("risk_count"),
                    "Master" if row.get("building_source") == "master" else "Staging" if row.get("building_source") == "staging" else row.get("building_source"),
                    row.get("created_at"),
                ]
                for row in case_rows
            ],
        )
        customer_rows = conn.execute(
            f"""
            SELECT g.*, c.group_name, c.building_name, c.unit
            FROM crm_case_guests g
            JOIN crm_cases c ON c.id = g.case_id
            WHERE {case_where.replace('c.', 'c.')}
              AND g.status != 'inactive'
            ORDER BY c.group_name ASC, g.created_at ASC
            """,
            case_params,
        ).fetchall()
        crm_export_append_sheet(
            workbook,
            "Customers",
            ["Customer", "WeChat Name", "Phone", "Email", "Case", "Building", "Unit", "Notes"],
            [
                [
                    row.get("full_name"),
                    row.get("wechat"),
                    row.get("phone"),
                    row.get("email"),
                    row.get("group_name"),
                    row.get("building_name"),
                    row.get("unit"),
                    row.get("notes"),
                ]
                for row in customer_rows
            ],
        )
        service_rows = conn.execute(
            f"""
            SELECT s.*, c.group_name, c.building_name, c.unit,
                   t.config_json AS template_config_json,
                   responsible.full_name AS responsible_customer_name
            FROM crm_case_services s
            JOIN crm_cases c ON c.id = s.case_id
            LEFT JOIN crm_service_templates t ON t.id = s.template_id
            LEFT JOIN crm_case_guests responsible ON responsible.id = s.responsible_customer_id
            WHERE {service_where}
            ORDER BY c.group_name ASC, s.service_name ASC
            """,
            service_params,
        ).fetchall()
        crm_export_append_sheet(
            workbook,
            "Services",
            ["Service", "WeChat Group / Case", "Building", "Scope", "Delivery Mode", "Responsible Customer", "Staff Workflow", "Customer Workflow", "Service Status", "Completion/Termination Reason", "Blocker"],
            [
                [
                    row.get("service_name"),
                    row.get("group_name"),
                    row.get("building_name"),
                    crm_report_label(CRM_SERVICE_SCOPE_LABELS, row.get("service_scope")),
                    crm_report_label(CRM_SERVICE_DELIVERY_MODE_LABELS, crm_report_service_delivery_mode(row)),
                    row.get("responsible_customer_name"),
                    row.get("staff_flow_status") or row.get("active_flow_step_key"),
                    row.get("customer_flow_status"),
                    crm_report_label(CRM_SERVICE_STATUS_LABELS, row.get("service_status") or row.get("status")),
                    row.get("termination_reason"),
                    row.get("blocked_reason"),
                ]
                for row in service_rows
            ],
        )
        task_rows = conn.execute(
            f"""
            SELECT t.*, c.group_name, c.status AS case_status,
                   s.service_key, s.service_name,
                   g.full_name AS customer_name,
                   users.display_name AS assigned_to_name, users.username AS assigned_username
            FROM crm_tasks t
            JOIN crm_cases c ON c.id = t.case_id
            LEFT JOIN crm_case_services s ON s.id = t.case_service_id
            LEFT JOIN crm_case_guests g ON g.id = COALESCE(t.target_customer_id, t.customer_id)
            LEFT JOIN users ON users.id = COALESCE(NULLIF(t.assigned_to, ''), t.assigned_user_id)
            WHERE {task_where}
            ORDER BY datetime(t.due_at) ASC, c.group_name ASC
            """,
            task_params,
        ).fetchall()
        crm_export_append_sheet(
            workbook,
            "Tasks",
            ["Task", "Case", "Service", "Customer", "Assignee", "Due At", "Priority", "Status", "Completed At", "Description"],
            [
                [
                    row.get("title"),
                    row.get("group_name"),
                    row.get("service_name"),
                    row.get("customer_name"),
                    crm_report_task_item(row)["assigned_to"],
                    row.get("due_at"),
                    crm_report_label(CRM_TASK_PRIORITY_LABELS, row.get("priority")),
                    crm_report_label(CRM_TASK_STATUS_LABELS, crm_normalize_task_status(row.get("status") or "open")),
                    row.get("completed_at"),
                    row.get("description"),
                ]
                for row in task_rows
            ],
        )
        timeline_rows: List[List[Any]] = []
        for case_id in case_ids:
            case_name = next((row.get("group_name") for row in case_rows if row.get("id") == case_id), "")
            for item in load_crm_case_timeline(conn, case_id):
                timeline_rows.append(
                    [
                        item.get("actor_name"),
                        item.get("occurred_at"),
                        item.get("title"),
                        case_name,
                        item.get("service_name"),
                        item.get("task_title"),
                        item.get("customer_name"),
                        item.get("summary"),
                    ]
                )
        crm_export_append_sheet(
            workbook,
            "Timeline",
            ["Staff Member", "Time", "Action", "Case", "Service", "Task", "Customer", "Status-Change Summary"],
            timeline_rows,
        )
        communication_rows = conn.execute(
            f"""
            SELECT e.*, c.group_name, s.service_name, g.full_name AS customer_name,
                   users.display_name AS created_by_name, users.username AS created_username
            FROM crm_communication_events e
            JOIN crm_cases c ON c.id = e.case_id
            LEFT JOIN crm_case_services s ON s.id = e.case_service_id
            LEFT JOIN crm_case_guests g ON g.id = e.customer_id
            LEFT JOIN users ON users.id = e.created_by
            WHERE {case_where}
            ORDER BY datetime(e.created_at) DESC
            """,
            case_params,
        ).fetchall()
        crm_export_append_sheet(
            workbook,
            "Communications",
            ["Communication Summary", "Channel", "Direction", "Case", "Service", "Customer", "Recorded By", "Time"],
            [
                [
                    row.get("summary"),
                    row.get("channel"),
                    row.get("direction"),
                    row.get("group_name"),
                    row.get("service_name"),
                    row.get("customer_name"),
                    crm_report_owner_name({"display_name": row.get("created_by_name"), "username": row.get("created_username")}),
                    row.get("created_at"),
                ]
                for row in communication_rows
            ],
        )
        notification_rows = conn.execute(
            f"""
            SELECT n.*, c.group_name, s.service_name, t.title AS task_title
            FROM crm_notifications n
            JOIN crm_cases c ON c.id = n.case_id
            LEFT JOIN crm_case_services s ON s.id = n.case_service_id
            LEFT JOIN crm_tasks t ON t.id = n.task_id
            WHERE {case_where}
            ORDER BY datetime(n.created_at) DESC
            """,
            case_params,
        ).fetchall()
        crm_export_append_sheet(
            workbook,
            "Notifications",
            ["Notification Summary", "Channel", "Status", "Case", "Service", "Source Task", "Recipient", "Generation Method", "Sent At", "Created At"],
            [
                [
                    (row.get("content") or "")[:500],
                    row.get("channel"),
                    row.get("status"),
                    row.get("group_name"),
                    row.get("service_name"),
                    row.get("task_title"),
                    row.get("recipient_type"),
                    row.get("generated_by"),
                    row.get("sent_at"),
                    row.get("created_at"),
                ]
                for row in notification_rows
            ],
        )
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        filename = f"CRM_Operations_Data_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        return StreamingResponse(
            buffer,
            headers=headers,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


@app.get("/crm/tasks")
async def list_crm_tasks(
    status: str = Query(default=""),
    owner_user_id: str = Query(default=""),
    assigned_to: str = Query(default=""),
    service_type: str = Query(default=""),
    priority: str = Query(default=""),
    case_status: str = Query(default=""),
    overdue: str = Query(default=""),
    mine: str = Query(default=""),
    scope: str = Query(default="critical"),
    date_from: str = Query(default="", alias="from"),
    date_to: str = Query(default="", alias="to"),
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        normalize_crm_task_rows(conn)
        sql = """
            SELECT crm_tasks.*, crm_cases.group_name, crm_cases.status AS case_status,
                   crm_cases.building_name AS case_building_name, crm_cases.unit AS case_unit,
                   crm_case_services.service_key, crm_case_services.service_name,
                   crm_case_services.need_status AS service_need_status,
                   crm_case_services.submission_status AS service_submission_status,
                   crm_case_services.completion_status AS service_completion_status,
                   crm_case_services.intro_status AS service_intro_status,
                   crm_case_services.follow_up_status AS service_follow_up_status,
                   crm_case_services.agent_completion_status AS service_agent_completion_status,
                   crm_case_services.staff_flow_status AS service_staff_flow_status,
                   crm_case_services.customer_flow_status AS service_customer_flow_status,
                   crm_case_services.service_status,
                   crm_case_services.status AS service_legacy_status,
                   crm_case_guests.full_name AS customer_name,
                   assignee.display_name AS assigned_to_name,
                   assignee.username AS assigned_username
            FROM crm_tasks
            JOIN crm_cases ON crm_cases.id = crm_tasks.case_id
            LEFT JOIN crm_case_services ON crm_case_services.id = crm_tasks.case_service_id
            LEFT JOIN crm_case_guests ON crm_case_guests.id = COALESCE(crm_tasks.target_customer_id, crm_tasks.customer_id)
            LEFT JOIN users AS assignee ON assignee.id = COALESCE(NULLIF(crm_tasks.assigned_to, ''), crm_tasks.assigned_user_id)
            WHERE 1 = 1
              AND COALESCE(crm_cases.deleted_at, '') = ''
        """
        params: List[Any] = []
        assigned_filter = assigned_to.strip() or owner_user_id.strip()
        if actor.role != "super_admin":
            sql += " AND COALESCE(NULLIF(crm_tasks.assigned_to, ''), crm_tasks.assigned_user_id) = ?"
            params.append(actor.user_id)
        elif mine.strip().lower() in {"1", "true", "yes"}:
            sql += " AND COALESCE(NULLIF(crm_tasks.assigned_to, ''), crm_tasks.assigned_user_id) = ?"
            params.append(actor.user_id)
        elif assigned_filter:
            sql += " AND COALESCE(NULLIF(crm_tasks.assigned_to, ''), crm_tasks.assigned_user_id) = ?"
            params.append(assigned_filter)
        if (scope or "critical").strip() != "all":
            sql += """
                AND (
                  crm_tasks.task_type = 'sim_card_sent_record'
                  OR crm_tasks.created_from_rule LIKE 'renters_insurance:insurance_ddl_%'
                  OR (
                    crm_tasks.task_type = 'verify'
                    AND crm_case_services.service_key IN ('internet_setup', 'internet')
                    AND (
                      crm_tasks.title LIKE '%验证码%'
                      OR crm_tasks.description LIKE '%网络验证码%'
                      OR lower(crm_tasks.title) LIKE '%verification code%'
                      OR lower(crm_tasks.description) LIKE '%internet verification code%'
                    )
                  )
                  OR (
                    crm_tasks.task_type = 'follow_up'
                    AND crm_tasks.source = 'manual'
                    AND COALESCE(crm_tasks.created_from_rule, '') = ''
                  )
                  OR (
                    crm_tasks.task_type = 'follow_up'
                    AND (
                      crm_tasks.description LIKE '%由日历关键日期任务创建%'
                      OR crm_tasks.description LIKE '%网络开户信息稍后补充%'
                      OR crm_tasks.description LIKE '%验证码预约时间待确认%'
                      OR lower(crm_tasks.description) LIKE '%created from a calendar critical-date task%'
                      OR lower(crm_tasks.description) LIKE '%internet account information will be added later%'
                      OR lower(crm_tasks.description) LIKE '%verification-code appointment time is pending%'
                    )
                  )
                )
            """
        if status.strip():
            normalized_status = crm_normalize_task_status(status.strip())
            if normalized_status == "overdue":
                sql += " AND crm_tasks.status NOT IN ('completed', 'cancelled', 'done') AND date(crm_tasks.due_at) < date('now', 'localtime')"
            else:
                sql += " AND crm_tasks.status IN (?, ?)"
                params.extend([normalized_status, "todo" if normalized_status == "open" else "done" if normalized_status == "completed" else normalized_status])
        if service_type.strip():
            sql += " AND crm_case_services.service_key = ?"
            params.append(service_type.strip())
        if priority.strip():
            sql += " AND crm_tasks.priority = ?"
            params.append(priority.strip())
        if case_status.strip():
            sql += " AND crm_cases.status = ?"
            params.append(case_status.strip())
        if overdue.strip().lower() in {"1", "true", "yes"}:
            sql += " AND crm_tasks.status NOT IN ('completed', 'cancelled', 'done') AND date(crm_tasks.due_at) < date('now', 'localtime')"
        elif overdue.strip().lower() in {"0", "false", "no"}:
            sql += " AND NOT (crm_tasks.status NOT IN ('completed', 'cancelled', 'done') AND date(crm_tasks.due_at) < date('now', 'localtime'))"
        if date_from.strip():
            sql += " AND date(crm_tasks.due_at) >= date(?)"
            params.append(date_from.strip())
        if date_to.strip():
            sql += " AND date(crm_tasks.due_at) <= date(?)"
            params.append(date_to.strip())
        sql += " ORDER BY datetime(crm_tasks.due_at) ASC, crm_tasks.priority DESC LIMIT 500"
        rows = conn.execute(sql, params).fetchall()
        tasks = [serialize_crm_task(row) for row in rows]
        today = datetime.now().date()
        next_week = today + timedelta(days=7)
        stats = {
            "today": 0,
            "overdue": 0,
            "next_7_days": 0,
            "high_priority": 0,
            "waiting_customer": 0,
            "waiting_external": 0,
            "total": len(tasks),
        }
        for task in tasks:
            due_date = parse_iso_date((task.get("due_at") or "")[:10])
            if due_date == today and task["status"] not in {"completed", "cancelled"}:
                stats["today"] += 1
            if task.get("is_overdue"):
                stats["overdue"] += 1
            if due_date and today <= due_date <= next_week and task["status"] not in {"completed", "cancelled"}:
                stats["next_7_days"] += 1
            if task.get("priority") in {"high", "urgent"} and task["status"] not in {"completed", "cancelled"}:
                stats["high_priority"] += 1
            if task["status"] == "waiting_customer":
                stats["waiting_customer"] += 1
            if task["status"] == "waiting_external":
                stats["waiting_external"] += 1
        if actor.role == "super_admin":
            owners = conn.execute(
                """
                SELECT id, display_name, username, role
                FROM users
                WHERE is_active = 1 AND role IN ('super_admin', 'admin', 'employee')
                ORDER BY display_name ASC, username ASC
                """
            ).fetchall()
        else:
            owners = conn.execute(
                """
                SELECT id, display_name, username, role
                FROM users
                WHERE id = ?
                """,
                (actor.user_id,),
            ).fetchall()
        return {"tasks": tasks, "stats": stats, "owners": owners}


@app.post("/crm/cases/{case_id}/tasks")
async def create_crm_task(
    case_id: str,
    payload: CrmTaskCreateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        current = require_crm_case_access(conn, case_id, actor)
        task_id = f"crm_task_{uuid.uuid4().hex}"
        due_at = payload.due_at.strip() or crm_task_due_at(parse_iso_date(current.get("lease_start_date")))
        target_customer_id = payload.target_customer_id.strip() or payload.customer_id.strip()
        assigned_to = payload.assigned_to.strip() or current.get("owner_user_id") or actor.user_id
        task_status = crm_task_initial_status(due_at, payload.status)
        conn.execute(
            """
            INSERT INTO crm_tasks(
              id, case_id, case_service_id, customer_id, title, description, task_type,
              due_at, not_before_at, priority, status, assigned_user_id, assigned_to,
              target_customer_id, source, created_from_rule, dedupe_key, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'manual', '', ?, ?, ?)
            """,
            (
                task_id,
                case_id,
                payload.case_service_id.strip() or None,
                target_customer_id or None,
                payload.title.strip(),
                payload.description.strip(),
                payload.task_type.strip() or "manual",
                due_at,
                payload.not_before_at.strip() or due_at,
                payload.priority,
                task_status,
                assigned_to,
                assigned_to,
                target_customer_id or None,
                f"manual:{task_id}",
                timestamp,
                timestamp,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_task_created",
            target_table="crm_tasks",
            target_record_id=task_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
            new_value=payload.title.strip(),
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.patch("/crm/tasks/{task_id}")
async def update_crm_task(
    task_id: str,
    payload: CrmTaskUpdateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    fields = payload_field_names(payload)
    timestamp = now_iso()
    with db_connection() as conn:
        task = conn.execute("SELECT * FROM crm_tasks WHERE id = ?", (task_id,)).fetchone()
        if not task:
            raise HTTPException(status_code=404, detail="CRM Task not found.")
        current = require_crm_case_access(conn, task["case_id"], actor)
        updates: List[str] = []
        params: List[Any] = []
        for column, value in [
            ("title", payload.title),
            ("description", payload.description),
            ("task_type", payload.task_type),
            ("due_at", payload.due_at),
            ("not_before_at", payload.not_before_at),
            ("assigned_to", payload.assigned_to),
            ("target_customer_id", payload.target_customer_id),
            ("priority", payload.priority),
            ("status", payload.status),
        ]:
            if column in fields and value is not None:
                updates.append(f"{column} = ?")
                params.append(value.strip() if isinstance(value, str) else value)
        if "assigned_to" in fields and payload.assigned_to is not None:
            updates.append("assigned_user_id = ?")
            params.append(payload.assigned_to.strip())
        if "target_customer_id" in fields and payload.target_customer_id is not None:
            updates.append("customer_id = ?")
            params.append(payload.target_customer_id.strip() or None)
        if "status" in fields:
            updates.append("completed_at = ?")
            params.append(timestamp if payload.status == "completed" else "")
        if updates:
            updates.append("updated_at = ?")
            params.extend([timestamp, task_id])
            conn.execute(f"UPDATE crm_tasks SET {', '.join(updates)} WHERE id = ?", params)
            write_audit_log(
                conn,
                request,
                actor,
                action_type="crm_task_updated",
                target_table="crm_tasks",
                target_record_id=task_id,
                building_name=current.get("building_name") or current.get("group_name") or "",
                new_value=payload.status or payload.title or "",
            )
        return load_crm_case_detail(conn, task["case_id"], actor)


@app.post("/crm/cases/{case_id}/communication-events")
async def create_crm_communication_event(
    case_id: str,
    payload: CrmCommunicationEventCreateRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        current = require_crm_case_access(conn, case_id, actor)
        event_id = f"crm_comm_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO crm_communication_events(
              id, case_id, case_service_id, customer_id, channel, direction, summary,
              raw_ref_json, created_by, created_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                event_id,
                case_id,
                payload.case_service_id.strip() or None,
                payload.customer_id.strip() or None,
                payload.channel,
                payload.direction,
                payload.summary.strip(),
                json_dumps(payload.raw_ref),
                actor.user_id,
                timestamp,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_communication_created",
            target_table="crm_communication_events",
            target_record_id=event_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
            new_value=payload.summary.strip(),
        )
        return load_crm_case_detail(conn, case_id, actor)


@app.post("/crm/cases/{case_id}/communication-events/{event_id}/extract-to-review", status_code=202)
async def extract_crm_communication_to_review(
    case_id: str,
    event_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    with db_connection() as conn:
        require_crm_case_access(conn, case_id, actor)
        event = conn.execute(
            "SELECT * FROM crm_communication_events WHERE id = ? AND case_id = ?",
            (event_id, case_id),
        ).fetchone()
        if not event:
            raise HTTPException(status_code=404, detail="CRM communication record not found.")
        existing = conn.execute(
            """
            SELECT id, source_file, parse_status
            FROM source_documents
            WHERE source_type = 'chat_crm' AND source_content LIKE ?
            ORDER BY created_at DESC LIMIT 1
            """,
            (f'%"communication_event_id": "{event_id}"%',),
        ).fetchone()
        if existing:
            return {
                **queued_intake_response(existing["id"], existing.get("source_file") or "CRM Communication Summary"),
                "parse_status": existing.get("parse_status") or INTAKE_PARSE_STATUS_QUEUED,
                "reused": True,
            }
        resolved_key = resolve_chat_target_staging_key(
            conn,
            actor=actor,
            request=request,
            case_id=case_id,
        )
        raw_ref = json_loads_safe(event.get("raw_ref_json"), {})
        raw_text = raw_ref.get("raw_text") if isinstance(raw_ref, dict) else ""
        extracted_text = normalize_unknown_value(raw_text) or event.get("summary") or ""
        metadata = chat_source_metadata(
            target_staging_key=resolved_key,
            source_kind="chat_crm",
            case_id=case_id,
            communication_event_id=event_id,
            captured_at=event.get("created_at") or "",
        )
        source_file = f"CRM-Communication-{event_id}.txt"
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="text",
            parser_type="chat_crm_queued",
            source_type="chat_crm",
            source_file=source_file,
            stored_path=None,
            source_content=json_dumps(metadata),
            extracted_text=extracted_text,
            extracted_pages=pages_from_text(extracted_text),
            actor=actor,
            request=request,
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_communication_extract_requested",
            target_table="source_documents",
            target_record_id=source_document_id,
            source=source_file,
            note=f"case_id={case_id}; event_id={event_id}; target_staging_key={resolved_key}",
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return {
        **queued_intake_response(source_document_id, source_file),
        "target_staging_key": resolved_key,
        "message": "The communication record has entered the building-knowledge review queue; it will not be written directly to Master.",
    }


def build_crm_notification_draft(
    case_row: dict,
    service_row: Optional[dict],
    task_row: Optional[dict],
    prompt: str,
) -> str:
    topic = "Move-In Tasks"
    if service_row:
        topic = service_row.get("service_name") or service_row.get("service_key") or topic
    if task_row:
        topic = task_row.get("title") or topic
    extra = f"\nAdditional note: {prompt.strip()}" if prompt.strip() else ""
    return (
        f"Hello everyone, a reminder about {topic}: please confirm the relevant items by "
        f"{case_row.get('lease_start_date') or 'the move-in date'}. If everything is complete, please reply in the group. "
        f"If you need help, let us know here.{extra}"
    )


@app.post("/crm/ai/notification-drafts")
async def create_crm_ai_notification_draft(
    payload: CrmNotificationDraftRequest,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        case_row = require_crm_case_access(conn, payload.case_id, actor)
        service_row = None
        task_row = None
        if payload.case_service_id.strip():
            service_row = crm_service_row_for_case(conn, payload.case_id, payload.case_service_id.strip())
        if payload.task_id.strip():
            task_row = conn.execute(
                "SELECT * FROM crm_tasks WHERE id = ? AND case_id = ?",
                (payload.task_id.strip(), payload.case_id),
            ).fetchone()
        content = build_crm_notification_draft(case_row, service_row, task_row, payload.prompt)
        notification_id = f"crm_notification_{uuid.uuid4().hex}"
        conn.execute(
            """
            INSERT INTO crm_notifications(
              id, case_id, task_id, case_service_id, channel, recipient_type, recipient_ref,
              content, status, generated_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, 'draft', 'ai', ?, ?)
            """,
            (
                notification_id,
                payload.case_id,
                payload.task_id.strip() or None,
                payload.case_service_id.strip() or None,
                payload.channel,
                payload.recipient_type,
                payload.recipient_ref.strip(),
                content,
                timestamp,
                timestamp,
            ),
        )
        conn.execute(
            """
            INSERT INTO crm_communication_events(
              id, case_id, case_service_id, customer_id, channel, direction,
              summary, raw_ref_json, created_by, created_at
            ) VALUES(?, ?, ?, ?, 'ai_draft', 'internal', ?, ?, ?, ?)
            """,
            (
                f"crm_comm_{uuid.uuid4().hex}",
                payload.case_id,
                payload.case_service_id.strip() or None,
                None,
                f"AI-generated notification draft: {content}",
                json_dumps({"notification_id": notification_id}),
                actor.user_id,
                timestamp,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_notification_draft_created",
            target_table="crm_notifications",
            target_record_id=notification_id,
            building_name=case_row.get("building_name") or case_row.get("group_name") or "",
            new_value=content,
        )
        return load_crm_case_detail(conn, payload.case_id, actor)


@app.post("/crm/notifications/{notification_id}/approve")
async def approve_crm_notification(
    notification_id: str,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM crm_notifications WHERE id = ?", (notification_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="CRM Notification not found.")
        current = require_crm_case_access(conn, row["case_id"], actor)
        conn.execute(
            "UPDATE crm_notifications SET status = 'approved', updated_at = ? WHERE id = ?",
            (timestamp, notification_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_notification_approved",
            target_table="crm_notifications",
            target_record_id=notification_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
        )
        return load_crm_case_detail(conn, row["case_id"], actor)


@app.post("/crm/notifications/{notification_id}/send")
async def send_crm_notification(
    notification_id: str,
    request: Request,
    user: dict = Depends(get_current_user),
):
    actor = ensure_crm_actor(user)
    timestamp = now_iso()
    with db_connection() as conn:
        row = conn.execute("SELECT * FROM crm_notifications WHERE id = ?", (notification_id,)).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="CRM Notification not found.")
        current = require_crm_case_access(conn, row["case_id"], actor)
        conn.execute(
            """
            UPDATE crm_notifications
            SET status = 'sent', sent_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (timestamp, timestamp, notification_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="crm_notification_sent",
            target_table="crm_notifications",
            target_record_id=notification_id,
            building_name=current.get("building_name") or current.get("group_name") or "",
        )
        return load_crm_case_detail(conn, row["case_id"], actor)


@app.get("/fields")
async def list_fields(user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        return {"fields": field_catalog(conn)}


@app.post("/field-requests/draft-from-text")
async def draft_field_request_from_text(
    payload: FieldRequestDraftFromTextRequest,
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    draft, used_ai = await ai_generate_field_draft(payload.display_name, payload.requirement_text)
    return {
        "draft": draft,
        "used_ai": used_ai,
    }


@app.get("/field-requests")
async def list_field_requests(
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        items = load_field_change_requests(
            conn,
            requested_by=None if actor.role == "super_admin" else actor.user_id,
        )
        return {"requests": items}


@app.post("/field-requests")
async def create_field_request(
    payload: FieldRequestCreateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    normalized_draft = normalize_field_draft_payload(payload.draft)
    if actor.role == "super_admin" and payload.apply_immediately:
        ensure_excel_mirror_ready_for_write()
    request_id = f"field_request_{uuid.uuid4().hex[:12]}"
    now = now_iso()
    with db_connection() as conn:
        if actor.role == "super_admin" and payload.apply_immediately:
            applied = apply_field_draft(
                conn,
                draft_payload=normalized_draft,
                actor=actor,
                request=request,
                action_note="Field request approved immediately by super_admin.",
            )
            conn.execute(
                """
                INSERT INTO field_change_requests(
                  id, display_name, requirement_text, draft_payload_json, status, requested_by,
                  reviewer, review_comment, applied_field_key, created_at, updated_at, reviewed_at
                ) VALUES(?, ?, ?, ?, 'approved', ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    request_id,
                    payload.display_name.strip(),
                    payload.requirement_text.strip(),
                    json_dumps(normalized_draft),
                    actor.user_id,
                    actor.user_id,
                    "super_admin direct apply",
                    applied["field_key"],
                    now,
                    now,
                    now,
                ),
            )
            write_audit_log(
                conn,
                request,
                actor,
                action_type="field_request_approved",
                target_table="field_change_requests",
                target_record_id=request_id,
                field_name=applied["field_key"],
                new_value=json_dumps(normalized_draft),
                note="Super Admin approved and activated the field immediately.",
            )
            return {"ok": True, "request_id": request_id, "status": "approved", "field": applied}

        conn.execute(
            """
            INSERT INTO field_change_requests(
              id, display_name, requirement_text, draft_payload_json, status, requested_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, 'pending', ?, ?, ?)
            """,
            (
                request_id,
                payload.display_name.strip(),
                payload.requirement_text.strip(),
                json_dumps(normalized_draft),
                actor.user_id,
                now,
                now,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_request_created",
            target_table="field_change_requests",
            target_record_id=request_id,
            new_value=json_dumps(normalized_draft),
            note="Submitted a request to add a field.",
        )
        return {"ok": True, "request_id": request_id, "status": "pending"}


@app.post("/field-requests/{request_id}/approve")
async def approve_field_request(
    request_id: str,
    payload: FieldRequestDecisionRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM field_change_requests WHERE id = ?",
            (request_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Field request not found.")
        if row["status"] == "approved":
            applied_field_key = row.get("applied_field_key")
            field = find_field_definition(conn, applied_field_key) if applied_field_key else None
            return {"ok": True, "request_id": request_id, "status": "approved", "field": field}
        if row["status"] == "rejected":
            raise HTTPException(status_code=400, detail="This field request has already been rejected.")

        stored_draft = json_loads_safe(row.get("draft_payload_json"), {})
        draft_payload = payload.draft.model_dump() if payload.draft else stored_draft
        applied = apply_field_draft(
            conn,
            draft_payload=draft_payload,
            actor=actor,
            request=request,
            action_note=f"Approved field request {request_id}.",
        )
        now = now_iso()
        conn.execute(
            """
            UPDATE field_change_requests
            SET draft_payload_json = ?, status = 'approved', reviewer = ?, review_comment = ?,
                applied_field_key = ?, reviewed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (
                json_dumps(normalize_field_draft_payload(draft_payload)),
                actor.user_id,
                payload.comment.strip(),
                applied["field_key"],
                now,
                now,
                request_id,
            ),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_request_approved",
            target_table="field_change_requests",
            target_record_id=request_id,
            field_name=applied["field_key"],
            new_value=json_dumps(applied),
            note=payload.comment.strip() or "Approved the field request.",
        )
        return {"ok": True, "request_id": request_id, "status": "approved", "field": applied}


@app.post("/field-requests/{request_id}/reject")
async def reject_field_request(
    request_id: str,
    payload: FieldRequestDecisionRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM field_change_requests WHERE id = ?",
            (request_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Field request not found.")
        if row["status"] == "approved":
            raise HTTPException(status_code=400, detail="This field request has already been approved and cannot be rejected.")
        now = now_iso()
        conn.execute(
            """
            UPDATE field_change_requests
            SET status = 'rejected', reviewer = ?, review_comment = ?, reviewed_at = ?, updated_at = ?
            WHERE id = ?
            """,
            (actor.user_id, payload.comment.strip(), now, now, request_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_request_rejected",
            target_table="field_change_requests",
            target_record_id=request_id,
            note=payload.comment.strip() or "Rejected the field request.",
        )
        return {"ok": True, "request_id": request_id, "status": "rejected"}


@app.post("/fields/reset-standard")
async def reset_fields_to_standard(
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        result = reset_standard_field_catalog(conn)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_catalog_reset_to_standard",
            target_table="field_definitions",
            note=json_dumps(result),
        )
    return {"ok": True, **result}


@app.post("/fields")
async def create_field_definition(
    payload: FieldDefinitionCreateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        field = apply_field_draft(
            conn,
            draft_payload={
                "field_key": payload.field_key,
                "display_name": payload.display_name,
                "field_type": payload.field_type,
                "excel_header_name": payload.display_name,
                "description": payload.description,
                "aliases": [payload.display_name],
                "query_keywords": [payload.display_name],
            },
            actor=actor,
            request=request,
            action_note="Legacy /fields endpoint create.",
        )
        return {"field_key": field["field_key"], "field": field}


@app.post("/fields/{field_key}/aliases")
async def create_field_alias(
    field_key: str,
    payload: AliasCreateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        definition = find_field_definition(conn, field_key)
        if not definition:
            raise HTTPException(status_code=404, detail="Field not found.")
        conn.execute(
            """
            INSERT INTO field_aliases(id, field_key, alias_name, language, confidence, created_by, created_at)
            VALUES(?, ?, ?, 'mixed', 1.0, ?, ?)
            """,
            (f"alias_{field_key}_{uuid.uuid4().hex[:8]}", field_key, payload.alias_name.strip(), actor.user_id, now_iso()),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_alias_created",
            target_table="field_aliases",
            target_record_id=field_key,
            field_name=field_key,
            new_value=payload.alias_name.strip(),
        )
    return {"ok": True}


@app.patch("/fields/{field_key}")
async def update_field_definition(
    field_key: str,
    payload: FieldDefinitionUpdateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        current = find_field_definition(conn, field_key)
        if not current:
            raise HTTPException(status_code=404, detail="Field not found.")
        updates = payload.model_dump(exclude_unset=True)
        if not updates:
            return {"field": current}
        old_header = normalize_unknown_value(current.get("excel_header_name")) or current["display_name"]
        next_display_name = (normalize_unknown_value(updates.get("display_name")) or current["display_name"]).strip()
        next_description = (normalize_unknown_value(updates.get("description")) or current.get("description") or "").strip()
        next_group_key = (normalize_unknown_value(updates.get("group_key")) or current.get("group_key") or "custom").strip()
        if next_group_key not in FIELD_GROUP_OPTIONS:
            next_group_key = current.get("group_key") or "custom"
        next_excel_header = (normalize_unknown_value(updates.get("excel_header_name")) or current.get("excel_header_name") or next_display_name).strip()
        next_visible_master = updates.get("visible_in_master_detail", bool(current.get("visible_in_master_detail")))
        next_visible_staging = updates.get("visible_in_staging_detail", bool(current.get("visible_in_staging_detail")))
        next_visible_query = updates.get("visible_in_query", bool(current.get("visible_in_query")))
        next_query_keywords = updates.get("query_keywords")
        if next_query_keywords is None:
            next_query_keywords = json_loads_safe(current.get("query_keywords_json"), [])
        next_query_keywords = normalize_query_keywords(next_display_name, next_query_keywords)
        next_answer_template = (normalize_unknown_value(updates.get("answer_template")) or current.get("answer_template") or "").strip()
        next_active = updates.get("active")
        active_flag = bool(current.get("active")) if next_active is None else bool(next_active)
        status_value = "active" if active_flag else "draft"
        if next_excel_header != old_header:
            rename_excel_header_everywhere(old_header, next_excel_header)
        conn.execute(
            """
            UPDATE field_definitions
            SET display_name = ?, description = ?, group_key = ?, excel_header_name = ?,
                visible_in_master_detail = ?, visible_in_staging_detail = ?, visible_in_query = ?,
                query_keywords_json = ?, answer_template = ?, active = ?, status = ?, updated_at = ?
            WHERE field_key = ?
            """,
            (
                next_display_name,
                next_description,
                next_group_key,
                next_excel_header,
                1 if next_visible_master else 0,
                1 if next_visible_staging else 0,
                1 if next_visible_query else 0,
                json_dumps(next_query_keywords),
                next_answer_template,
                1 if active_flag else 0,
                status_value,
                now_iso(),
                field_key,
            ),
        )
        ensure_master_workbook_from_db(conn)
        ensure_staging_workbook_from_sources(conn)
        sync_excel_mirrors(conn, actor=actor, request=request)
        field = find_field_definition(conn, field_key)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="field_definition_updated",
            target_table="field_definitions",
            target_record_id=field_key,
            field_name=field_key,
            new_value=json_dumps(updates),
            note="Updated the field catalog.",
        )
        return {"field": field}


@app.get("/master-excel/status")
async def master_excel_status(user: dict = Depends(get_current_user)):
    workbook_path = resolve_master_excel_path()
    staging_workbook_path = resolve_staging_excel_path()
    with db_connection() as conn:
        try:
            if not workbook_path.exists():
                ensure_master_workbook_from_db(conn)
            if not staging_workbook_path.exists():
                ensure_staging_workbook_from_sources(conn)
            validation = validate_master_workbook(workbook_path)
            staging_validation = validate_staging_workbook(staging_workbook_path)
            reconciled_at = conn.execute(
                "SELECT value FROM app_meta WHERE key = 'excel_mirrors_refreshed_at'"
            ).fetchone()
            return {
                "ok": validation["ok"],
                "path": str(workbook_path),
                "exists": workbook_path.exists(),
                "sheet_names": validation["sheet_names"],
                "row_count": validation["row_count"],
                "missing_sheets": validation["missing_sheets"],
                "missing_headers": validation["missing_headers"],
                "duplicate_headers": validation["duplicate_headers"],
                "merged_ranges": validation["merged_ranges"],
                "main_sheet": MASTER_MAIN_SHEET,
                "help_sheet": MASTER_HELP_SHEET,
                "package_sheet": MASTER_PLAN_SHEET,
                "last_reconciled_at": reconciled_at["value"] if reconciled_at else None,
                "staging_excel": {
                    "ok": staging_validation["ok"],
                    "path": str(staging_workbook_path),
                    "exists": staging_workbook_path.exists(),
                    "sheet_names": staging_validation["sheet_names"],
                    "row_count": staging_validation["row_count"],
                    "missing_sheets": staging_validation["missing_sheets"],
                    "missing_headers": staging_validation["missing_headers"],
                    "duplicate_headers": staging_validation["duplicate_headers"],
                    "merged_ranges": staging_validation["merged_ranges"],
                },
            }
        except Exception as exc:
            return {
                "ok": False,
                "path": str(workbook_path),
                "exists": workbook_path.exists(),
                "error": str(exc),
                "main_sheet": MASTER_MAIN_SHEET,
                "help_sheet": MASTER_HELP_SHEET,
                "package_sheet": MASTER_PLAN_SHEET,
                "last_reconciled_at": None,
                "staging_excel": {
                    "path": str(staging_workbook_path),
                    "exists": staging_workbook_path.exists(),
                },
            }


@app.get("/master-excel/download")
async def master_excel_download(user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        workbook_path = ensure_master_workbook_from_db(conn)
    validation = validate_master_workbook(workbook_path)
    if not validation["ok"]:
        raise HTTPException(status_code=409, detail="The master workbook structure is invalid. Repair it before downloading.")
    return FileResponse(
        workbook_path,
        filename=workbook_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/master-excel/reconcile")
async def reconcile_master_excel(
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        return {
            "ok": True,
            **sync_excel_mirrors(
                conn,
                actor=actor,
                request=request,
            ),
        }


@app.post("/excel-mirrors/refresh")
async def refresh_excel_mirrors(
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        return {
            "ok": True,
            **sync_excel_mirrors(
                conn,
                actor=actor,
                request=request,
            ),
        }


@app.post("/master-excel/preview")
async def preview_master_excel_import(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    upload_dir = UPLOAD_ROOT / "imports" / "master_excel" / uuid.uuid4().hex
    stored_path, file_hash = save_upload_file(
        upload_dir,
        file,
        allowed_suffixes=ALLOWED_EXCEL_UPLOAD_SUFFIXES,
        label="Master workbook",
    )
    validation = validate_standard_master_upload(stored_path)
    batch_id = f"import_{uuid.uuid4().hex}"
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO import_batches(
              id, original_file_name, stored_path, file_hash, status, sheet_names, uploaded_by, confirmed_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, 'uploaded', '[]', ?, NULL, ?, ?)
            """,
            (batch_id, file.filename or stored_path.name, str(stored_path), file_hash, actor.user_id, now_iso(), now_iso()),
        )
        preview = await preview_excel_file(stored_path, conn)
        sheet_names = [item["sheet_name"] for item in preview["sheets"]]
        conn.execute(
            "UPDATE import_batches SET sheet_names = ?, updated_at = ? WHERE id = ?",
            (json_dumps(sheet_names), now_iso(), batch_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="master_excel_previewed",
            target_table="import_batches",
            target_record_id=batch_id,
            source=file.filename or stored_path.name,
            note="Previewed an uploaded master workbook.",
        )
    return {
        "batch_id": batch_id,
        "file_name": file.filename or stored_path.name,
        "validation": validation,
        **preview,
    }


@app.post("/master-excel/confirm")
async def confirm_master_excel_import(
    payload: ImportConfirmRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        batch = conn.execute(
            "SELECT * FROM import_batches WHERE id = ?",
            (payload.batch_id,),
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found.")
        validate_standard_master_upload(Path(batch["stored_path"]))
        result = persist_import_to_staging(
            conn,
            batch_id=payload.batch_id,
            file_name=batch["original_file_name"],
            sheets=payload.sheets,
            file_path=Path(batch["stored_path"]),
            actor=actor,
            request=request,
            source_type="master_excel_import",
            parser_type="excel_header_mapping",
            raw_input_type="excel",
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="master_excel_import_confirmed",
            target_table="import_batches",
            target_record_id=payload.batch_id,
            source=batch["original_file_name"],
            note=f"Imported {result['rows']} rows / {result['fields']} fields from standard workbook into staging.",
        )
        return {"batch_id": payload.batch_id, **result}


@app.post("/imports/excel/preview")
async def preview_excel_import(
    request: Request,
    file: UploadFile = File(...),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    upload_dir = UPLOAD_ROOT / "imports" / uuid.uuid4().hex
    stored_path, file_hash = save_upload_file(
        upload_dir,
        file,
        allowed_suffixes=ALLOWED_EXCEL_UPLOAD_SUFFIXES,
        label="Excel import",
    )
    batch_id = f"import_{uuid.uuid4().hex}"
    with db_connection() as conn:
        conn.execute(
            """
            INSERT INTO import_batches(
              id, original_file_name, stored_path, file_hash, status, sheet_names, uploaded_by, confirmed_by, created_at, updated_at
            ) VALUES(?, ?, ?, ?, 'uploaded', '[]', ?, NULL, ?, ?)
            """,
            (batch_id, file.filename or stored_path.name, str(stored_path), file_hash, actor.user_id, now_iso(), now_iso()),
        )
        preview = await preview_excel_file(stored_path, conn)
        sheet_names = [item["sheet_name"] for item in preview["sheets"]]
        conn.execute(
            "UPDATE import_batches SET sheet_names = ?, updated_at = ? WHERE id = ?",
            (json_dumps(sheet_names), now_iso(), batch_id),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="excel_upload_previewed",
            target_table="import_batches",
            target_record_id=batch_id,
            source=file.filename or stored_path.name,
        )
    return {"batch_id": batch_id, "file_name": file.filename or stored_path.name, **preview}


@app.post("/imports/excel/confirm")
async def confirm_excel_import(
    payload: ImportConfirmRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        batch = conn.execute(
            "SELECT * FROM import_batches WHERE id = ?",
            (payload.batch_id,),
        ).fetchone()
        if not batch:
            raise HTTPException(status_code=404, detail="Import batch not found.")
        result = persist_import_to_staging(
            conn,
            batch_id=payload.batch_id,
            file_name=batch["original_file_name"],
            sheets=payload.sheets,
            file_path=Path(batch["stored_path"]),
            actor=actor,
            request=request,
        )
        return {"batch_id": payload.batch_id, **result}


@app.get("/imports/batches")
async def list_import_batches(user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, original_file_name, status, sheet_names, uploaded_by, confirmed_by, created_at, updated_at
            FROM import_batches
            ORDER BY created_at DESC
            LIMIT 50
            """
        ).fetchall()
    return {"batches": rows}


async def create_source_document_record(
    conn: sqlite3.Connection,
    *,
    raw_input_type: str,
    parser_type: str,
    source_type: str,
    source_file: str,
    stored_path: Optional[Path],
    source_content: str,
    extracted_text: str,
    extracted_pages: Optional[List[dict]],
    actor: Actor,
    request: Request,
) -> str:
    source_document_id = f"source_{uuid.uuid4().hex}"
    timestamp = now_iso()
    conn.execute(
        """
        INSERT INTO source_documents(
          id, raw_input_type, parser_type, source_type, source_file, stored_path,
          source_content, extracted_text, extracted_pages_json, parse_status,
          created_by, created_at, updated_at
        ) VALUES(?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            source_document_id,
            raw_input_type,
            parser_type,
            source_type,
            source_file,
            str(stored_path) if stored_path else "",
            source_content,
            extracted_text,
            json_dumps(extracted_pages or []),
            INTAKE_PARSE_STATUS_QUEUED,
            actor.user_id,
            timestamp,
            timestamp,
        ),
    )
    write_audit_log(
        conn,
        request,
        actor,
        action_type="source_document_created",
        target_table="source_documents",
        target_record_id=source_document_id,
        source=source_file,
        note=f"raw_input_type={raw_input_type}",
    )
    return source_document_id


def set_source_document_parse_status(
    conn: sqlite3.Connection,
    source_document_id: str,
    status: str,
    *,
    error: str = "",
    submission_group_id: str = "",
) -> None:
    timestamp = now_iso()
    if status == INTAKE_PARSE_STATUS_RUNNING:
        conn.execute(
            """
            UPDATE source_documents
            SET parse_status = ?, parse_started_at = ?, parse_error = '', updated_at = ?
            WHERE id = ?
            """,
            (status, timestamp, timestamp, source_document_id),
        )
        return
    if status == INTAKE_PARSE_STATUS_COMPLETED:
        conn.execute(
            """
            UPDATE source_documents
            SET parse_status = ?, parse_completed_at = ?, parse_error = ?,
                submission_group_id = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, timestamp, error[:4000], submission_group_id, timestamp, source_document_id),
        )
        return
    if status == INTAKE_PARSE_STATUS_FAILED:
        conn.execute(
            """
            UPDATE source_documents
            SET parse_status = ?, parse_completed_at = ?, parse_error = ?, updated_at = ?
            WHERE id = ?
            """,
            (status, timestamp, error[:4000], timestamp, source_document_id),
        )
        return
    conn.execute(
        "UPDATE source_documents SET parse_status = ?, updated_at = ? WHERE id = ?",
        (status, timestamp, source_document_id),
    )


def source_document_row_for_job(conn: sqlite3.Connection, source_document_id: str) -> Optional[dict]:
    return conn.execute(
        "SELECT * FROM source_documents WHERE id = ?",
        (source_document_id,),
    ).fetchone()


async def run_intake_parse_job(source_document_id: str, actor_payload: dict) -> None:
    actor = Actor(
        user_id=actor_payload.get("user_id") or "system_intake_parser",
        username=actor_payload.get("username") or "system",
        role=actor_payload.get("role") or "ai_system",
    )
    with db_connection() as conn:
        row = source_document_row_for_job(conn, source_document_id)
        if not row:
            return
        set_source_document_parse_status(conn, source_document_id, INTAKE_PARSE_STATUS_RUNNING)

    try:
        with db_connection() as conn:
            row = source_document_row_for_job(conn, source_document_id)
            if not row:
                return

        raw_input_type = row.get("raw_input_type") or "text"
        parser_type = row.get("parser_type") or "welcome_letter_text"
        source_type = row.get("source_type") or "welcome_letter_text"
        source_file = row.get("source_file") or "welcome_letter.txt"
        source_options = read_intake_source_options(row.get("source_content"))
        intake_mode = normalize_intake_mode(source_options.get("intake_mode"))
        supplement_scope = normalize_supplement_scope(source_options.get("supplement_scope"))
        target_staging_key = normalize_unknown_value(source_options.get("target_staging_key")) or ""
        extracted_pages = json_loads_safe(row.get("extracted_pages_json"), [])
        extracted_text = row.get("extracted_text") or ""
        source_content = row.get("source_content") or extracted_text or ""
        vision_values: Optional[Dict[str, dict]] = None

        if raw_input_type in {"pdf", "pdf_package"}:
            pdf_paths = list_source_pdf_files(row.get("stored_path"))
            if not pdf_paths:
                raise ValueError("No parseable PDF files were found.")
            extracted_pages, extracted_text, parser_type, vision_values = await prepare_pdf_package_source_payload(
                pdf_paths,
                source_file,
                source_document_id=source_document_id,
            )
            raw_input_type = "pdf_package" if len(pdf_paths) > 1 else "pdf"
        elif raw_input_type == "image":
            image_paths = list_source_image_files(row.get("stored_path"))
            if not image_paths:
                raise ValueError("No parseable image files were found.")
            extracted_pages, extracted_text, parser_type, vision_values = await prepare_image_source_payload(
                image_paths,
                source_file,
                source_document_id=source_document_id,
            )
        else:
            if not extracted_pages:
                extracted_pages = pages_from_text(extracted_text or source_content)
            if not extracted_text:
                extracted_text = join_page_texts(extracted_pages) or source_content

        with db_connection() as conn:
            conn.execute(
                """
                UPDATE source_documents
                SET raw_input_type = ?, parser_type = ?, extracted_text = ?,
                    extracted_pages_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    raw_input_type,
                    parser_type,
                    extracted_text,
                    json_dumps(extracted_pages or []),
                    now_iso(),
                    source_document_id,
                ),
            )

        with db_connection() as conn:
            conn.execute(
                "DELETE FROM staging_update_requests WHERE submission_group_id = ?",
                (f"source:{source_document_id}",),
            )
            result = await parse_source_to_staging(
                conn,
                source_document_id=source_document_id,
                source_file=source_file,
                raw_input_type=raw_input_type,
                parser_type=parser_type,
                source_type=source_type,
                source_content=source_content,
                extracted_text=extracted_text,
                extracted_pages=extracted_pages,
                actor=actor,
                request=None,  # type: ignore[arg-type]
                vision_values=vision_values,
                intake_mode=intake_mode,
                supplement_scope=supplement_scope,
                target_staging_key=target_staging_key,
            )
            archive_warning = archive_source_document_files(
                conn,
                source_document_id,
                result,
                actor=actor,
                force_needs_review=False,
            )
            set_source_document_parse_status(
                conn,
                source_document_id,
                INTAKE_PARSE_STATUS_COMPLETED,
                error=archive_warning or "",
                submission_group_id=result.get("submission_group_id") or "",
            )
    except Exception as exc:
        detail = exc.detail if isinstance(exc, HTTPException) else str(exc)
        with db_connection() as conn:
            archive_warning = archive_source_document_files(
                conn,
                source_document_id,
                None,
                actor=actor,
                force_needs_review=True,
            )
            error_message = str(detail or "Parsing failed.")
            if archive_warning:
                error_message = f"{error_message}; {archive_warning}"
            set_source_document_parse_status(
                conn,
                source_document_id,
                INTAKE_PARSE_STATUS_FAILED,
                error=error_message,
            )
            write_audit_log(
                conn,
                None,  # type: ignore[arg-type]
                actor,
                action_type="source_parse_failed",
                target_table="source_documents",
                target_record_id=source_document_id,
                note=error_message,
            )


def queue_intake_parse(background_tasks: BackgroundTasks, source_document_id: str, actor: Actor) -> None:
    background_tasks.add_task(
        run_intake_parse_job,
        source_document_id,
        {"user_id": actor.user_id, "username": actor.username, "role": actor.role},
    )


def queued_intake_response(source_document_id: str, source_file: str) -> dict:
    return {
        "source_document_id": source_document_id,
        "source_file": source_file,
        "parse_status": INTAKE_PARSE_STATUS_QUEUED,
        "message": "Submitted for background parsing. You can continue working on other pages.",
    }


def resolve_chat_target_staging_key(
    conn: sqlite3.Connection,
    *,
    actor: Actor,
    request: Optional[Request],
    target_staging_key: str = "",
    case_id: str = "",
) -> str:
    requested_key = normalize_unknown_value(target_staging_key) or ""
    if requested_key:
        snapshot = load_staging_building_snapshot(conn, requested_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="The linked Staging building does not exist.")
        return snapshot["staging_key"]

    normalized_case_id = normalize_unknown_value(case_id) or ""
    if not normalized_case_id:
        raise HTTPException(status_code=400, detail="Chat material must be linked to a Staging building or to an existing CRM Case that already has a building.")
    case_row = require_crm_case_access(conn, normalized_case_id, actor)
    building_source = normalize_unknown_value(case_row.get("building_source")) or ""
    building_id = normalize_unknown_value(case_row.get("building_id")) or ""
    if not building_source or not building_id:
        raise HTTPException(status_code=400, detail="This CRM Case is not linked to a building. Link one in CRM first.")
    if building_source == "staging":
        snapshot = load_staging_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="The Staging building linked to this CRM Case does not exist.")
        return snapshot["staging_key"]

    if building_source != "master":
        raise HTTPException(status_code=400, detail="This CRM Case building source cannot be used for chat-material intake.")
    master = load_master_building_snapshot(conn, building_id)
    if not master:
        raise HTTPException(status_code=404, detail="The Master building linked to this CRM Case does not exist.")
    staging = resolve_staging_snapshot_for_source_sync(
        conn,
        master.get("building_name") or "",
        master.get("address") or "",
    )
    if staging:
        return staging["staging_key"]
    ensure_excel_mirror_ready_for_write()
    staging = create_manual_staging_building_snapshot(
        conn,
        building_name=master.get("building_name") or "",
        address=master.get("address") or "",
        aliases="",
        notes="Created from a CRM Case linked to a Master building to receive chat material for review.",
        insurance_required=None,
        electricity_required=None,
        internet_self_setup_required=None,
        actor=actor,
        request=request,
    )
    return staging["staging_key"]


def chat_source_metadata(
    *,
    target_staging_key: str,
    source_kind: str,
    case_id: str = "",
    communication_event_id: str = "",
    captured_at: str = "",
) -> dict:
    return build_intake_source_metadata(
        intake_mode=INTAKE_MODE_SUPPLEMENT,
        supplement_scope=SUPPLEMENT_SCOPE_ALL,
        target_staging_key=target_staging_key,
        source_kind=source_kind,
        case_id=case_id,
        communication_event_id=communication_event_id,
        captured_at=captured_at,
    )


@app.post("/intake/welcome-letter/text", status_code=202)
async def intake_text_source(
    request: Request,
    background_tasks: BackgroundTasks,
    source_text: str = Form(...),
    source_file_name: str = Form(default="email_text.txt"),
    intake_mode: str = Form(default=INTAKE_MODE_FULL_PACKAGE),
    supplement_scope: str = Form(default=SUPPLEMENT_SCOPE_ALL),
    target_staging_key: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    if len((source_text or "").encode("utf-8")) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Text content is too large. The limit is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
        )
    normalized_intake_mode = normalize_intake_mode(intake_mode)
    normalized_supplement_scope = normalize_supplement_scope(supplement_scope)
    normalized_target_staging_key = normalize_unknown_value(target_staging_key) or ""
    with db_connection() as conn:
        extracted_pages = pages_from_text(source_text)
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="text",
            parser_type="welcome_letter_text",
            source_type="welcome_letter_text",
            source_file=source_file_name,
            stored_path=None,
            source_content=json_dumps(
                build_intake_source_metadata(
                    intake_mode=normalized_intake_mode,
                    supplement_scope=normalized_supplement_scope,
                    target_staging_key=normalized_target_staging_key,
                )
            ),
            extracted_text=source_text,
            extracted_pages=extracted_pages,
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return queued_intake_response(source_document_id, source_file_name)


@app.post("/intake/welcome-letter/pdf", status_code=202)
async def intake_pdf_source(
    request: Request,
    background_tasks: BackgroundTasks,
    files: Optional[List[UploadFile]] = File(default=None),
    file: Optional[UploadFile] = File(default=None),
    intake_mode: str = Form(default=INTAKE_MODE_FULL_PACKAGE),
    supplement_scope: str = Form(default=SUPPLEMENT_SCOPE_ALL),
    target_staging_key: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    upload_files: List[UploadFile] = list(files or [])
    if file is not None:
        upload_files.append(file)
    upload_files = [item for item in upload_files if item is not None]
    if not upload_files:
        raise HTTPException(status_code=400, detail="Upload at least one PDF.")
    if len(upload_files) > MAX_PDF_UPLOAD_FILES:
        raise HTTPException(status_code=400, detail=f"You can upload at most {MAX_PDF_UPLOAD_FILES} PDFs at a time.")
    normalized_intake_mode = normalize_intake_mode(intake_mode)
    normalized_supplement_scope = normalize_supplement_scope(supplement_scope)
    normalized_target_staging_key = normalize_unknown_value(target_staging_key) or ""
    source_metadata = build_intake_source_metadata(
        intake_mode=normalized_intake_mode,
        supplement_scope=normalized_supplement_scope,
        target_staging_key=normalized_target_staging_key,
    )
    upload_dir = UPLOAD_ROOT / "sources" / uuid.uuid4().hex
    stored_paths: List[Path] = []
    total_upload_bytes = 0
    for index, upload in enumerate(upload_files, start=1):
        stored_path, _ = save_upload_file(
            upload_dir,
            upload,
            allowed_suffixes=ALLOWED_PDF_UPLOAD_SUFFIXES,
            label="PDF",
            name_prefix=f"{index:02d}",
        )
        total_upload_bytes += stored_path.stat().st_size
        if total_upload_bytes > MAX_UPLOAD_BYTES:
            shutil.rmtree(upload_dir, ignore_errors=True)
            raise HTTPException(
                status_code=413,
                detail=f"The total PDF package size cannot exceed {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
            )
        stored_paths.append(stored_path)
    source_file_label = summarize_pdf_source_files(upload_files, stored_paths)
    raw_input_type = "pdf_package" if len(stored_paths) > 1 else "pdf"
    stored_source_path = upload_dir if len(stored_paths) > 1 else stored_paths[0]
    with db_connection() as conn:
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type=raw_input_type,
            parser_type="welcome_letter_pdf_queued",
            source_type="welcome_letter_pdf",
            source_file=source_file_label,
            stored_path=stored_source_path,
            source_content=json_dumps(source_metadata) if source_metadata else "",
            extracted_text="",
            extracted_pages=[],
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return queued_intake_response(source_document_id, source_file_label)


@app.post("/intake/welcome-letter/image", status_code=202)
async def intake_image_source(
    request: Request,
    background_tasks: BackgroundTasks,
    files: Optional[List[UploadFile]] = File(default=None),
    file: Optional[UploadFile] = File(default=None),
    intake_mode: str = Form(default=INTAKE_MODE_FULL_PACKAGE),
    supplement_scope: str = Form(default=SUPPLEMENT_SCOPE_ALL),
    target_staging_key: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    upload_files: List[UploadFile] = list(files or [])
    if file is not None:
        upload_files.append(file)
    upload_files = [item for item in upload_files if item is not None]
    if not upload_files:
        raise HTTPException(status_code=400, detail="Upload at least one image.")
    upload_dir = UPLOAD_ROOT / "sources" / uuid.uuid4().hex
    stored_paths: List[Path] = []
    if len(upload_files) > MAX_IMAGE_UPLOAD_FILES:
        raise HTTPException(status_code=400, detail=f"You can upload at most {MAX_IMAGE_UPLOAD_FILES} images at a time.")
    normalized_intake_mode = normalize_intake_mode(intake_mode)
    normalized_supplement_scope = normalize_supplement_scope(supplement_scope)
    normalized_target_staging_key = normalize_unknown_value(target_staging_key) or ""
    source_metadata = build_intake_source_metadata(
        intake_mode=normalized_intake_mode,
        supplement_scope=normalized_supplement_scope,
        target_staging_key=normalized_target_staging_key,
    )
    total_upload_bytes = 0
    for upload in upload_files:
        stored_path, _ = save_upload_file(
            upload_dir,
            upload,
            allowed_suffixes=ALLOWED_IMAGE_UPLOAD_SUFFIXES,
            label="Image",
        )
        total_upload_bytes += stored_path.stat().st_size
        if total_upload_bytes > MAX_UPLOAD_BYTES:
            shutil.rmtree(upload_dir, ignore_errors=True)
            raise HTTPException(
                status_code=413,
                detail=f"The total image-group size cannot exceed {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
            )
        stored_paths.append(stored_path)
    source_file_label = summarize_image_source_files(upload_files, stored_paths)
    with db_connection() as conn:
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="image",
            parser_type="welcome_letter_image_queued",
            source_type="welcome_letter_image",
            source_file=source_file_label,
            stored_path=upload_dir,
            source_content=json_dumps(source_metadata) if source_metadata else "",
            extracted_text="",
            extracted_pages=[],
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return queued_intake_response(source_document_id, source_file_label)


@app.post("/intake/chat/text", status_code=202)
async def intake_chat_text_source(
    request: Request,
    background_tasks: BackgroundTasks,
    source_text: str = Form(...),
    source_file_name: str = Form(default="wechat_chat.txt"),
    target_staging_key: str = Form(default=""),
    case_id: str = Form(default=""),
    captured_at: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    if len((source_text or "").encode("utf-8")) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"Text content is too large. The limit is {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
        )
    with db_connection() as conn:
        resolved_key = resolve_chat_target_staging_key(
            conn,
            actor=actor,
            request=request,
            target_staging_key=target_staging_key,
            case_id=case_id,
        )
        metadata = chat_source_metadata(
            target_staging_key=resolved_key,
            source_kind="chat_text",
            case_id=case_id,
            captured_at=captured_at,
        )
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="text",
            parser_type="chat_text_queued",
            source_type="chat_text",
            source_file=source_file_name or "wechat_chat.txt",
            stored_path=None,
            source_content=json_dumps(metadata),
            extracted_text=source_text,
            extracted_pages=pages_from_text(source_text),
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return {
        **queued_intake_response(source_document_id, source_file_name or "wechat_chat.txt"),
        "target_staging_key": resolved_key,
        "privacy_notice": "Customer names, private contact details, verification codes, and identity-document information are not written to building fields. The original text is stored only in access-controlled source records.",
    }


@app.post("/intake/chat/pdf", status_code=202)
async def intake_chat_pdf_source(
    request: Request,
    background_tasks: BackgroundTasks,
    files: Optional[List[UploadFile]] = File(default=None),
    file: Optional[UploadFile] = File(default=None),
    target_staging_key: str = Form(default=""),
    case_id: str = Form(default=""),
    captured_at: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    upload_files: List[UploadFile] = list(files or [])
    if file is not None:
        upload_files.append(file)
    upload_files = [item for item in upload_files if item is not None]
    if not upload_files:
        raise HTTPException(status_code=400, detail="Upload at least one PDF.")
    if len(upload_files) > MAX_PDF_UPLOAD_FILES:
        raise HTTPException(status_code=400, detail=f"You can upload at most {MAX_PDF_UPLOAD_FILES} PDFs at a time.")
    with db_connection() as conn:
        resolved_key = resolve_chat_target_staging_key(
            conn,
            actor=actor,
            request=request,
            target_staging_key=target_staging_key,
            case_id=case_id,
        )
    upload_dir = UPLOAD_ROOT / "sources" / uuid.uuid4().hex
    stored_paths: List[Path] = []
    total_upload_bytes = 0
    for upload in upload_files:
        stored_path, _ = save_upload_file(
            upload_dir,
            upload,
            allowed_suffixes={".pdf"},
            label="PDF",
        )
        total_upload_bytes += stored_path.stat().st_size
        if total_upload_bytes > MAX_UPLOAD_BYTES:
            shutil.rmtree(upload_dir, ignore_errors=True)
            raise HTTPException(
                status_code=413,
                detail=f"The total PDF package size cannot exceed {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
            )
        stored_paths.append(stored_path)
    source_file_label = summarize_pdf_source_files(upload_files, stored_paths)
    metadata = chat_source_metadata(
        target_staging_key=resolved_key,
        source_kind="chat_pdf",
        case_id=case_id,
        captured_at=captured_at,
    )
    with db_connection() as conn:
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="pdf_package" if len(stored_paths) > 1 else "pdf",
            parser_type="chat_pdf_queued",
            source_type="chat_pdf",
            source_file=source_file_label,
            stored_path=upload_dir if len(stored_paths) > 1 else stored_paths[0],
            source_content=json_dumps(metadata),
            extracted_text="",
            extracted_pages=[],
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return {
        **queued_intake_response(source_document_id, source_file_label),
        "target_staging_key": resolved_key,
        "cloud_ocr_notice": (
            "The current OCR configuration sends files to Baidu; business-field interpretation still uses the Xiaomi vision model."
            if configured_ocr_router().primary == "baidu_unlimited_cloud"
            else (
                "OCR sends files to the configured private Unlimited-OCR service; business-field interpretation still uses the Xiaomi vision model."
                if configured_ocr_router().primary == "unlimited_ocr_local_http"
                else "OCR runs locally; if image text is insufficient, the Xiaomi vision model may still assist with business-field interpretation."
            )
        ),
    }


@app.post("/intake/chat/image", status_code=202)
async def intake_chat_image_source(
    request: Request,
    background_tasks: BackgroundTasks,
    files: Optional[List[UploadFile]] = File(default=None),
    file: Optional[UploadFile] = File(default=None),
    target_staging_key: str = Form(default=""),
    case_id: str = Form(default=""),
    captured_at: str = Form(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    upload_files: List[UploadFile] = list(files or [])
    if file is not None:
        upload_files.append(file)
    upload_files = [item for item in upload_files if item is not None]
    if not upload_files:
        raise HTTPException(status_code=400, detail="Upload at least one chat screenshot.")
    if len(upload_files) > MAX_IMAGE_UPLOAD_FILES:
        raise HTTPException(status_code=400, detail=f"You can upload at most {MAX_IMAGE_UPLOAD_FILES} images at a time.")
    with db_connection() as conn:
        resolved_key = resolve_chat_target_staging_key(
            conn,
            actor=actor,
            request=request,
            target_staging_key=target_staging_key,
            case_id=case_id,
        )
    upload_dir = UPLOAD_ROOT / "sources" / uuid.uuid4().hex
    stored_paths: List[Path] = []
    total_upload_bytes = 0
    for upload in upload_files:
        stored_path, _ = save_upload_file(
            upload_dir,
            upload,
            allowed_suffixes=ALLOWED_IMAGE_UPLOAD_SUFFIXES,
            label="Image",
        )
        total_upload_bytes += stored_path.stat().st_size
        if total_upload_bytes > MAX_UPLOAD_BYTES:
            shutil.rmtree(upload_dir, ignore_errors=True)
            raise HTTPException(
                status_code=413,
                detail=f"The total image-group size cannot exceed {MAX_UPLOAD_BYTES // (1024 * 1024)} MB.",
            )
        stored_paths.append(stored_path)
    source_file_label = summarize_image_source_files(upload_files, stored_paths)
    metadata = chat_source_metadata(
        target_staging_key=resolved_key,
        source_kind="chat_image",
        case_id=case_id,
        captured_at=captured_at,
    )
    with db_connection() as conn:
        source_document_id = await create_source_document_record(
            conn,
            raw_input_type="image",
            parser_type="chat_image_queued",
            source_type="chat_image",
            source_file=source_file_label,
            stored_path=upload_dir,
            source_content=json_dumps(metadata),
            extracted_text="",
            extracted_pages=[],
            actor=actor,
            request=request,
        )
    queue_intake_parse(background_tasks, source_document_id, actor)
    return {
        **queued_intake_response(source_document_id, source_file_label),
        "target_staging_key": resolved_key,
        "cloud_ocr_notice": (
            "The current OCR configuration sends images to Baidu; business-field interpretation still uses the Xiaomi vision model."
            if configured_ocr_router().primary == "baidu_unlimited_cloud"
            else (
                "OCR sends images to the configured private Unlimited-OCR service; business-field interpretation still uses the Xiaomi vision model."
                if configured_ocr_router().primary == "unlimited_ocr_local_http"
                else "OCR runs locally; the Xiaomi vision model participates in business-field interpretation for images."
            )
        ),
    }


@app.get("/source-files/{relative_path:path}")
async def protected_source_file(
    relative_path: str,
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    file_path = resolve_protected_file(UPLOAD_ROOT, relative_path)
    with db_connection() as conn:
        require_source_path_access(conn, actor, file_path)
    return FileResponse(file_path, filename=file_path.name)


@app.get("/legacy-files/{relative_path:path}")
async def protected_legacy_file(
    relative_path: str,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    file_path = resolve_protected_file(legacy.PDF_LIBRARY_DIR, relative_path)
    return FileResponse(file_path, filename=file_path.name)


@app.get("/image-files/{relative_path:path}")
async def protected_image_file(
    relative_path: str,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    file_path = resolve_protected_file(legacy.IMAGE_LIBRARY_DIR, relative_path)
    return FileResponse(file_path, filename=file_path.name)


@app.get("/intake/jobs/{source_document_id}")
async def get_intake_job(
    source_document_id: str,
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        row = conn.execute(
            "SELECT * FROM source_documents WHERE id = ?",
            (source_document_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Parsing job not found.")
        if actor.role == "employee" and row.get("created_by") != actor.user_id:
            raise HTTPException(status_code=403, detail="You do not have permission to view this parsing job.")
        return {"job": serialize_intake_job(row)}


@app.get("/intake/jobs")
async def list_intake_jobs(
    limit: int = Query(default=10, ge=1, le=50),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    try:
        with db_connection() as conn:
            params: List[object] = []
            where_parts = [
                "(source_type IN ('welcome_letter_text', 'welcome_letter_pdf', 'welcome_letter_image') OR source_type LIKE 'chat_%')"
            ]
            if actor.role == "employee":
                where_parts.append("created_by = ?")
                params.append(actor.user_id)
            params.append(limit)
            rows = conn.execute(
                f"""
                SELECT *
                FROM source_documents
                WHERE {' AND '.join(where_parts)}
                ORDER BY created_at DESC
                LIMIT ?
                """,
                tuple(params),
            ).fetchall()
        return {"jobs": [serialize_intake_job(row) for row in rows]}
    except sqlite3.OperationalError as exc:
        if "locked" not in str(exc).lower():
            raise
        return {
            "jobs": [],
            "warning": "Background parsing is writing to the database. The task list will recover automatically shortly.",
        }


@app.get("/source-documents")
async def list_source_documents(user: dict = Depends(require_roles("super_admin", "admin", "employee"))):
    actor = get_actor(user)
    with db_connection() as conn:
        if actor.role == "employee":
            rows = conn.execute(
                """
                SELECT id, raw_input_type, parser_type, source_type, source_file, stored_path,
                       parse_status, parse_started_at, parse_completed_at, parse_error,
                       submission_group_id, created_by, created_at, updated_at
                FROM source_documents
                WHERE created_by = ?
                ORDER BY created_at DESC
                LIMIT 100
                """,
                (actor.user_id,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id, raw_input_type, parser_type, source_type, source_file, stored_path,
                       parse_status, parse_started_at, parse_completed_at, parse_error,
                       submission_group_id, created_by, created_at, updated_at
                FROM source_documents
                ORDER BY created_at DESC
                LIMIT 100
                """
            ).fetchall()
    return {"documents": [serialize_source_document(row) for row in rows]}


@app.get("/review/groups")
async def list_review_groups(
    status: str = Query(default="actionable"),
    stage: str = Query(default=""),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT submission_group_id, MIN(building_name) AS building_name, MIN(source_file) AS source_file,
                   MIN(raw_input_type) AS raw_input_type, MIN(parser_type) AS parser_type,
                   MIN(source_type) AS source_type,
                   MIN(approval_stage) AS approval_stage, MIN(target_staging_key) AS target_staging_key,
                   MIN(submitted_by) AS submitted_by, MIN(review_status) AS review_status,
                   COUNT(*) AS item_count, MAX(created_at) AS created_at
            FROM staging_update_requests
            GROUP BY submission_group_id
            ORDER BY MAX(created_at) DESC
            """
        ).fetchall()
    for row in rows:
        row["approval_stage"] = review_group_approval_stage([row])
    if actor.role == "employee":
        rows = [row for row in rows if str(row.get("submitted_by") or "") == actor.user_id]
    normalized_status = (status or "actionable").strip().lower()
    if normalized_status == "actionable":
        rows = [row for row in rows if row["review_status"] in ACTIONABLE_REVIEW_STATUSES]
    elif normalized_status == "processed":
        rows = [row for row in rows if row["review_status"] in PROCESSED_REVIEW_STATUSES]
    elif normalized_status in {"all", "*"}:
        pass
    elif normalized_status:
        rows = [row for row in rows if row["review_status"] == normalized_status]
    normalized_stage = normalize_review_approval_stage(stage) if stage else ""
    if normalized_stage:
        rows = [row for row in rows if row.get("approval_stage") == normalized_stage]
    return {"groups": rows}


@app.get("/review/groups/{group_id}")
async def review_group_detail(
    group_id: str,
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = fetch_group_records(conn, group_id)
        if not rows:
            raise HTTPException(status_code=404, detail="Review group not found.")
        if actor.role == "employee" and any(str(row.get("submitted_by") or "") != actor.user_id for row in rows):
            raise HTTPException(status_code=403, detail="Employees can view only review records they submitted.")
        approval_stage = review_group_approval_stage(rows)
        source_document = None
        source_document_id = rows[0]["source_document_id"]
        if source_document_id:
            source_document = conn.execute(
                """
                SELECT id, raw_input_type, parser_type, source_type, source_file, stored_path,
                       source_content, extracted_text, extracted_pages_json, parse_artifacts_json, created_at
                FROM source_documents
                WHERE id = ?
                """,
                (source_document_id,),
            ).fetchone()
            if source_document and actor.role in {"super_admin", "admin"}:
                llm_logs = conn.execute(
                    """
                    SELECT id, source_document_id, stage, model, system_prompt, user_payload_json,
                           raw_response, parsed_response_json, error, created_at
                    FROM llm_call_logs
                    WHERE source_document_id = ?
                    ORDER BY created_at DESC
                    LIMIT 20
                    """,
                    (source_document_id,),
                ).fetchall()
                source_document = dict(source_document)
                source_document["llm_call_logs"] = [serialize_llm_call_log(row) for row in llm_logs]
        is_source_group = str(rows[0].get("source_type") or "").startswith("welcome_letter_")
        source_identity_allows_match = source_group_allows_existing_match(rows) if is_source_group else True

        building = None
        manual_master_building_id = review_group_manual_master_building_id(rows)
        matched_master_building = None
        if approval_stage == APPROVAL_STAGE_TO_MASTER and manual_master_building_id:
            matched_master_building = load_master_building_snapshot(conn, manual_master_building_id)
        if approval_stage == APPROVAL_STAGE_TO_MASTER and rows[0]["building_id"] and (not is_source_group or source_identity_allows_match):
            building = load_master_building_snapshot(conn, rows[0]["building_id"])
        if matched_master_building:
            building = matched_master_building
        group_address = next(
            (
                normalize_unknown_value(row.get("new_value"))
                for row in rows
                if row.get("field_name") == "address" and normalize_unknown_value(row.get("new_value"))
            ),
            "",
        )
        target_staging_key = review_group_target_staging_key(rows)
        staging_building = load_staging_building_snapshot(conn, target_staging_key) if target_staging_key else None
        if not staging_building and approval_stage == APPROVAL_STAGE_TO_STAGING and is_source_group and source_identity_allows_match:
            staging_building = resolve_staging_snapshot_for_source_sync(
                conn,
                rows[0]["building_name"],
                group_address,
            )
        serialized_rows = []
        for row in rows:
            item = serialize_review_record(row)
            if approval_stage == APPROVAL_STAGE_TO_STAGING and not str(item.get("field_name") or "").startswith("__"):
                item["old_value"] = None
                if staging_building:
                    staged_old_value = get_staging_snapshot_field_value(staging_building, item["field_name"])
                    if staged_old_value is not None:
                        item["old_value"] = staged_old_value
            serialized_rows.append(item)
        value_map = review_group_value_map(serialized_rows, building)
        return {
            "group_id": group_id,
            "approval_stage": approval_stage,
            "records": serialized_rows,
            "source_document": serialize_source_document(source_document),
            "building": building,
            "matched_staging_building": staging_building,
            "matched_master_building": matched_master_building,
            "can_confirm_staging_building": approval_stage == APPROVAL_STAGE_TO_STAGING
            and actor.role in {"super_admin", "admin"},
            "can_confirm_master_building": approval_stage == APPROVAL_STAGE_TO_MASTER and actor.role == "super_admin",
            "can_write_to_staging": approval_stage == APPROVAL_STAGE_TO_STAGING
            and actor.role in {"super_admin", "admin"},
            "can_write_to_master": approval_stage == APPROVAL_STAGE_TO_MASTER and actor.role == "super_admin",
            "insurance_database_view": build_insurance_database_view(value_map, serialized_rows),
            "insurance_mapping_view": build_insurance_mapping_view(value_map, serialized_rows),
            "insurance_staff_explanation": build_insurance_staff_explanation(value_map, serialized_rows),
        }


@app.post("/review/groups/{group_id}/confirm-staging-building")
async def confirm_review_group_staging_building(
    group_id: str,
    payload: ConfirmStagingBuildingRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = fetch_group_records(conn, group_id)
        if not rows:
            raise HTTPException(status_code=404, detail="Review group not found.")
        approval_stage = review_group_approval_stage(rows)
        if approval_stage != APPROVAL_STAGE_TO_STAGING:
            raise HTTPException(status_code=400, detail="Only a review group awaiting Staging approval can confirm a Staging building.")
        snapshot = load_staging_building_snapshot(conn, payload.staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging building not found.")
        timestamp = now_iso()
        confirmed_name = normalize_unknown_value(snapshot.get("building_name")) or ""
        confirmed_address = normalize_unknown_value(snapshot.get("address")) or ""
        marker_source_content = {
            "manual_master_building_id": snapshot["id"],
            "manual_master_building_name": confirmed_name,
            "manual_master_building_address": confirmed_address,
            "manual_master_confirmed_by": actor.user_id,
            "manual_master_confirmed_at": timestamp,
        }
        for row in rows:
            field_name = row["field_name"]
            new_value = row["new_value"]
            if field_name == "building_name":
                new_value = confirmed_name
            elif field_name == "address":
                new_value = confirmed_address
            old_value = None if field_name.startswith("__") else get_staging_snapshot_field_value(snapshot, field_name)
            normalized_new = normalize_unknown_value(new_value)
            normalized_old = normalize_unknown_value(old_value)
            conn.execute(
                """
                UPDATE staging_update_requests
                SET building_name = ?, building_id = NULL, old_value = ?, new_value = ?,
                    normalized_new_value = ?, approval_stage = ?, target_staging_key = ?,
                    conflict_with_long_term = ?, updated_at = ?
                WHERE record_id = ?
                """,
                (
                    confirmed_name,
                    old_value or "",
                    new_value or "",
                    new_value or "",
                    APPROVAL_STAGE_TO_STAGING,
                    snapshot["staging_key"],
                    1 if normalized_old and normalized_new and normalized_old != normalized_new else 0,
                    timestamp,
                    row["record_id"],
                ),
            )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="review_group_confirm_staging_building",
            target_table="staging_update_requests",
            target_record_id=group_id,
            building_name=confirmed_name,
            old_value=rows[0].get("building_name") or "",
            new_value=f"{confirmed_name} | {confirmed_address}",
            source=rows[0].get("source_file") or "",
            note="Manually linked the review group to an existing Staging building.",
        )
        rows = fetch_group_records(conn, group_id)
        return {
            "ok": True,
            "group_id": group_id,
            "matched_staging_building": snapshot,
            "approval_stage": review_group_approval_stage(rows),
            "message": "The previous-value baseline now uses the selected Staging building.",
        }


@app.post("/review/groups/{group_id}/confirm-master-building")
async def confirm_review_group_master_building(
    group_id: str,
    payload: ConfirmMasterBuildingRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = fetch_group_records(conn, group_id)
        if not rows:
            raise HTTPException(status_code=404, detail="Review group not found.")
        approval_stage = review_group_approval_stage(rows)
        if approval_stage != APPROVAL_STAGE_TO_MASTER:
            raise HTTPException(status_code=400, detail="Only a review group awaiting Master approval can select a Master building.")
        snapshot = load_master_building_snapshot(conn, payload.building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master building not found.")
        timestamp = now_iso()
        confirmed_name = normalize_unknown_value(snapshot.get("building_name")) or ""
        confirmed_address = normalize_unknown_value(snapshot.get("address")) or ""
        for row in rows:
            field_name = row["field_name"]
            new_value = row["new_value"]
            if field_name == "building_name":
                new_value = confirmed_name
            elif field_name == "address":
                new_value = confirmed_address
            old_value = None if field_name.startswith("__") else get_master_field_value(conn, snapshot["id"], field_name)
            normalized_new = normalize_unknown_value(new_value)
            normalized_old = normalize_unknown_value(old_value)
            source_content = json_loads_safe(row.get("source_content"), {})
            if not isinstance(source_content, dict):
                source_content = {}
            source_content.update(marker_source_content)
            conn.execute(
                """
                UPDATE staging_update_requests
                SET building_name = ?, building_id = ?, old_value = ?, new_value = ?,
                    normalized_new_value = ?, approval_stage = ?, source_content = ?,
                    conflict_with_long_term = ?, updated_at = ?
                WHERE record_id = ?
                """,
                (
                    confirmed_name,
                    snapshot["id"],
                    old_value or "",
                    new_value or "",
                    new_value or "",
                    APPROVAL_STAGE_TO_MASTER,
                    json_dumps(source_content),
                    1 if normalized_old and normalized_new and normalized_old != normalized_new else 0,
                    timestamp,
                    row["record_id"],
                ),
            )
        existing_fields = {str(row["field_name"] or "") for row in rows}
        for field_name, new_value in [("building_name", confirmed_name), ("address", confirmed_address)]:
            if not new_value or field_name in existing_fields:
                continue
            old_value = get_master_field_value(conn, snapshot["id"], field_name)
            normalized_old = normalize_unknown_value(old_value)
            normalized_new = normalize_unknown_value(new_value)
            create_staging_request(
                conn,
                submission_group_id=group_id,
                building_name=confirmed_name,
                building_id=snapshot["id"],
                field_name=field_name,
                old_value=old_value,
                new_value=new_value,
                source_type=rows[0]["source_type"],
                source_content=json_dumps(marker_source_content),
                source_file=rows[0]["source_file"],
                submitted_by=rows[0]["submitted_by"],
                ai_confidence=None,
                review_status=rows[0]["review_status"],
                import_batch_id=rows[0]["import_batch_id"],
                parser_type=rows[0]["parser_type"],
                raw_input_type=rows[0]["raw_input_type"],
                source_document_id=rows[0]["source_document_id"],
                approval_stage=APPROVAL_STAGE_TO_MASTER,
                target_staging_key=review_group_target_staging_key(rows),
                conflict_with_long_term=bool(normalized_old and normalized_new and normalized_old != normalized_new),
                low_confidence=False,
                missing_required_detail=False,
                priority=rows[0]["priority"] or "normal",
            )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="review_group_confirm_master_building",
            target_table="staging_update_requests",
            target_record_id=group_id,
            building_name=confirmed_name,
            old_value=rows[0].get("building_name") or "",
            new_value=f"{confirmed_name} | {confirmed_address}",
            source=rows[0].get("source_file") or "",
            note="Manually assigned the review group to an existing Master building.",
        )
        rows = fetch_group_records(conn, group_id)
        return {
            "ok": True,
            "group_id": group_id,
            "matched_master_building": snapshot,
            "building": snapshot,
            "approval_stage": review_group_approval_stage(rows),
            "message": "The Master building has been selected. Approved changes will be written to that building.",
        }


@app.post("/review/groups/{group_id}/reparse")
async def reparse_review_group_source(
    group_id: str,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = fetch_group_records(conn, group_id)
        if not rows:
            raise HTTPException(status_code=404, detail="Review group not found.")
        source_document_id = rows[0]["source_document_id"]
        if not source_document_id:
            raise HTTPException(status_code=400, detail="This review group has no original source file and cannot be reparsed.")
        source_document = conn.execute(
            """
            SELECT id, raw_input_type, parser_type, source_type, source_file, stored_path,
                   source_content, extracted_text, extracted_pages_json
            FROM source_documents
            WHERE id = ?
            """,
            (source_document_id,),
        ).fetchone()
        if not source_document:
            raise HTTPException(status_code=404, detail="Original source file not found.")

        extracted_pages = json_loads_safe(source_document.get("extracted_pages_json"), [])
        stored_path_raw = normalize_unknown_value(source_document.get("stored_path"))
        stored_path = Path(stored_path_raw) if stored_path_raw else None
        raw_input_type = source_document["raw_input_type"]
        parser_type = source_document["parser_type"]
        extracted_text = source_document.get("extracted_text") or ""
        source_content_raw = source_document.get("source_content") or ""
        source_metadata = parse_intake_source_metadata(source_content_raw)
        source_content = source_content_raw
        row_metadata = review_group_intake_metadata(rows)
        if row_metadata:
            source_metadata = row_metadata

        vision_values = None
        if raw_input_type in {"pdf", "pdf_package"} and stored_path and (stored_path.is_file() or stored_path.is_dir()):
            pdf_paths = list_source_pdf_files(stored_path)
            if not pdf_paths:
                raise HTTPException(status_code=404, detail="Original PDF file not found.")
            extracted_pages, extracted_text, parser_type, vision_values = await prepare_pdf_package_source_payload(
                pdf_paths,
                source_document["source_file"],
                source_document_id=source_document_id,
            )
            conn.execute(
                "UPDATE source_documents SET parser_type = ?, extracted_text = ?, extracted_pages_json = ?, updated_at = ? WHERE id = ?",
                (parser_type, extracted_text, json_dumps(extracted_pages), now_iso(), source_document_id),
            )
        elif raw_input_type == "image" and stored_path and (stored_path.is_file() or stored_path.is_dir()):
            image_paths = list_source_image_files(stored_path)
            extracted_pages, extracted_text, parser_type, vision_values = await prepare_image_source_payload(
                image_paths,
                source_document["source_file"],
                source_document_id=source_document_id,
            )
            conn.execute(
                "UPDATE source_documents SET parser_type = ?, extracted_text = ?, extracted_pages_json = ?, updated_at = ? WHERE id = ?",
                (parser_type, join_page_texts(extracted_pages) or extracted_text or source_content, json_dumps(extracted_pages), now_iso(), source_document_id),
            )
        elif not extracted_pages:
            extracted_pages = pages_from_text(extracted_text or source_content)
            conn.execute(
                "UPDATE source_documents SET extracted_text = ?, extracted_pages_json = ?, updated_at = ? WHERE id = ?",
                (join_page_texts(extracted_pages) or extracted_text or source_content, json_dumps(extracted_pages), now_iso(), source_document_id),
            )
            extracted_text = join_page_texts(extracted_pages) or extracted_text or source_content

        conn.execute("DELETE FROM staging_update_requests WHERE submission_group_id = ?", (group_id,))
        result = await parse_source_to_staging(
            conn,
            source_document_id=source_document_id,
            source_file=source_document["source_file"],
            raw_input_type=raw_input_type,
            parser_type=parser_type,
            source_type=source_document["source_type"],
            source_content=source_content,
            extracted_text=extracted_text,
            extracted_pages=extracted_pages,
            actor=actor,
            request=request,
            vision_values=vision_values,
            intake_mode=source_metadata.get("intake_mode", INTAKE_MODE_FULL_PACKAGE),
            supplement_scope=source_metadata.get("supplement_scope", SUPPLEMENT_SCOPE_ALL),
            target_staging_key=source_metadata.get("target_staging_key", ""),
        )
        write_audit_log(
            conn,
            request,
            actor,
            action_type="review_group_reparsed",
            target_table="staging_update_requests",
            target_record_id=group_id,
            building_name=result.get("building_name"),
            source=source_document["source_file"],
            note="Reparsed existing source document into staging.",
        )
        return {"ok": True, **result}


@app.delete("/review/groups/{group_id}")
async def delete_review_group(
    group_id: str,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        rows = fetch_group_records(conn, group_id)
        if not rows:
            raise HTTPException(status_code=404, detail="Review group not found.")
        if any(row["review_status"] == "migrated_to_master" for row in rows):
            raise HTTPException(status_code=400, detail="A review group already migrated to Master cannot be deleted directly.")
        deleted = conn.execute(
            "DELETE FROM staging_update_requests WHERE submission_group_id = ?",
            (group_id,),
        ).rowcount
        write_audit_log(
            conn,
            request,
            actor,
            action_type="review_group_deleted",
            target_table="staging_update_requests",
            target_record_id=group_id,
            building_name=rows[0]["building_name"],
            source=rows[0]["source_file"],
            note=f"Deleted {deleted} staging rows from group.",
        )
    return {"ok": True, "deleted_records": deleted}


@app.post("/review/groups/{group_id}/decision")
async def review_group_decision(
    group_id: str,
    payload: ReviewDecisionRequest,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    updates = {item.record_id: item.new_value for item in payload.updates}
    resolutions = {item.record_id: item.resolution for item in payload.updates if item.resolution}
    with db_connection() as conn:
        payload_key = {
            "group_id": group_id,
            "action": payload.action,
            "comment": payload.comment.strip(),
            "updates": {key: updates[key] for key in sorted(updates)},
            "resolutions": {key: resolutions[key] for key in sorted(resolutions)},
        }
        replay = begin_idempotent_request(
            conn,
            actor=actor,
            scope=f"review_decision:{group_id}",
            idempotency_key=idempotency_key,
            payload=payload_key,
        )
        if replay and replay.get("replay"):
            return replay["response"]
        result = apply_review_decision(
            conn,
            group_id=group_id,
            action=payload.action,
            comment=payload.comment.strip(),
            updates=updates,
            resolutions=resolutions,
            actor=actor,
            request=request,
        )
        complete_idempotent_request(
            conn,
            actor=actor,
            scope=f"review_decision:{group_id}",
            idempotency_key=idempotency_key,
            response_payload=result,
        )
    return result


@app.get("/master/buildings")
async def list_master_buildings(
    q: str = Query(default=""),
    user: dict = Depends(get_current_user),
):
    with db_connection() as conn:
        rows = conn.execute(
            """
            SELECT id, building_name, address, status, version, updated_at, last_verified_at,
                   completeness_status, completeness_score, verification_note
            FROM master_building_info
            ORDER BY building_name ASC
            """
        ).fetchall()
    if q.strip():
        query = q.strip().lower()
        rows = [
            row
            for row in rows
            if query in (row["building_name"] or "").lower() or query in (row["address"] or "").lower()
        ]
    return {"buildings": rows}


@app.get("/staging/buildings")
async def list_staging_buildings(
    q: str = Query(default=""),
    user: dict = Depends(get_current_user),
):
    with db_connection() as conn:
        snapshots = load_staging_building_snapshots(conn)
    if q.strip():
        query = q.strip().lower()
        snapshots = [
            item
            for item in snapshots
            if query in (item["building_name"] or "").lower() or query in (item["address"] or "").lower()
        ]
    return {
        "buildings": [
            {
                "id": item["staging_key"],
                "building_name": item["building_name"],
                "address": item["address"],
                "pending_count": item["pending_count"],
                "source_file": item["source_file"],
                "library_status": item.get("library_status") or STAGING_STATUS_PENDING,
                "source_date": item.get("source_date"),
            }
            for item in snapshots
        ]
        }


@app.post("/staging/buildings")
async def create_staging_building(
    payload: StagingBuildingCreateRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        snapshot = create_manual_staging_building_snapshot(
            conn,
            building_name=payload.building_name,
            address=payload.address,
            aliases=payload.aliases,
            notes=payload.notes,
            insurance_required=payload.insurance_required,
            electricity_required=payload.electricity_required,
            internet_self_setup_required=payload.internet_self_setup_required,
            actor=actor,
            request=request,
        )
        return {"building": snapshot, "message": "Building added to Staging."}


@app.get("/staging/buildings/{staging_key}")
async def staging_building_detail(staging_key: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")
        return snapshot


@app.get("/master/buildings/{building_id}")
async def master_building_detail(building_id: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
        return snapshot


@app.get("/master/buildings/{building_id}/summary")
async def master_building_summary(building_id: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
    fact_summary = build_summary_answer(snapshot, "master")
    ai_summary, cache_status, cache_key = await generate_cached_fact_explanation(
        source_mode="master",
        record_id=building_id,
        question="Create a concise, staff-friendly summary of the approved information for this building.",
        snapshot=snapshot,
        fact_answer=fact_summary,
    )
    payload = {
        "building_id": building_id,
        "fact_summary": fact_summary,
        "ai_summary": ai_summary,
        "ai_enabled": ai_explanation_enabled(),
        "cache_status": cache_status,
        "network": resolve_network_for_snapshot(snapshot),
        "message": "Master summary generated.",
    }
    if user.get("role") == "super_admin":
        payload["cache_key"] = cache_key
    return payload


@app.get("/staging/buildings/{staging_key}/summary")
async def staging_building_summary(staging_key: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")
    fact_summary = build_summary_answer(snapshot, "staging")
    ai_summary, cache_status, cache_key = await generate_cached_fact_explanation(
        source_mode="staging",
        record_id=staging_key,
        question="Create a concise, staff-friendly summary of the Staging information for this building.",
        snapshot=snapshot,
        fact_answer=fact_summary,
    )
    payload = {
        "staging_key": staging_key,
        "fact_summary": fact_summary,
        "ai_summary": ai_summary,
        "ai_enabled": ai_explanation_enabled(),
        "cache_status": cache_status,
        "network": build_network_payload_from_master(snapshot),
        "message": "Staging summary generated.",
    }
    if user.get("role") == "super_admin":
        payload["cache_key"] = cache_key
    return payload


@app.get("/master/buildings/{building_id}/network")
async def master_building_network(building_id: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
    matched = resolve_network_for_snapshot(snapshot)
    if not matched:
        return {
            "building_id": building_id,
            "matched": None,
            "message": "No structured internet information is currently available for this building.",
        }
    return {
        "building_id": building_id,
        "matched": matched,
        "message": "Structured internet information was found for this building.",
    }


@app.get("/staging/buildings/{staging_key}/network")
async def staging_building_network(staging_key: str, user: dict = Depends(get_current_user)):
    with db_connection() as conn:
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")
    matched = build_network_payload_from_master(snapshot)
    if not matched:
        return {
            "staging_key": staging_key,
            "matched": None,
            "message": "No structured internet information is currently available for this Staging building.",
        }
    return {
        "staging_key": staging_key,
        "matched": matched,
        "message": "Structured Staging internet information was found for this building.",
    }


@app.post("/staging/buildings/{staging_key}/submit-for-review")
async def submit_staging_building_for_review(
    staging_key: str,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        replay = begin_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_submit_for_review:{staging_key}",
            idempotency_key=idempotency_key,
            payload={"staging_key": staging_key},
        )
        if replay and replay.get("replay"):
            return replay["response"]
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")
        result = ensure_manual_master_review_group_for_staging_snapshot(conn, staging_snapshot=snapshot, actor=actor)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="staging_submit_for_review",
            target_table="staging_update_requests",
            target_record_id=result["group_id"],
            building_name=snapshot["building_name"],
            source=snapshot.get("source_file") or Path(resolve_staging_excel_path()).name,
            note="Submitted staging snapshot for manual master review.",
        )
        result = {
            "ok": True,
            **result,
            "message": "This building was submitted to the review queue.",
        }
        complete_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_submit_for_review:{staging_key}",
            idempotency_key=idempotency_key,
            response_payload=result,
        )
        return result


@app.post("/staging/buildings/{staging_key}/submit-master-review")
async def submit_staging_building_for_master_review(
    staging_key: str,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    return await submit_staging_building_for_review(
        staging_key=staging_key,
        request=request,
        idempotency_key=idempotency_key,
        user=user,
    )


@app.post("/staging/buildings/{staging_key}/request-update")
async def request_staging_building_update(
    staging_key: str,
    payload: MasterUpdateRequest,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin", "admin", "employee")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        replay = begin_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_request_update:{staging_key}",
            idempotency_key=idempotency_key,
            payload={
                "staging_key": staging_key,
                "updates": payload.updates,
                "note": payload.note,
            },
        )
        if replay and replay.get("replay"):
            return replay["response"]
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")

        allowed_fields = set(workbook_catalog_field_keys(conn, include_staging_only=True))
        source_file = snapshot.get("source_file") or Path(resolve_staging_excel_path()).name
        submission_group_id = f"staging_update_request:{staging_key}:{uuid.uuid4().hex}"
        created = 0
        for field_key, raw_value in dict(payload.updates or {}).items():
            if field_key not in allowed_fields or field_key in {"source_date", "source_type", "source_file", "info_cutoff_date", "library_status"}:
                continue
            definition = find_field_definition(conn, field_key)
            field_type = "boolean" if field_key in NETWORK_PROVIDER_FIELD_MAP else (definition["field_type"] if definition else "text")
            normalized_value = normalize_field_value(field_key, field_type, raw_value)
            if normalized_value is None:
                continue
            old_value = get_staging_snapshot_field_value(snapshot, field_key)
            if normalize_unknown_value(old_value) == normalize_unknown_value(normalized_value):
                continue
            create_staging_request(
                conn,
                submission_group_id=submission_group_id,
                building_name=snapshot["building_name"],
                building_id=None,
                field_name=field_key,
                old_value=old_value,
                new_value=normalized_value,
                source_type="employee_staging_update_request",
                source_content=payload.note or "Employee-submitted Staging update request",
                source_file=source_file,
                submitted_by=actor.user_id,
                ai_confidence=None,
                review_status="employee_submitted",
                import_batch_id=None,
                parser_type="manual_submission",
                raw_input_type="manual",
                source_document_id=None,
                approval_stage=APPROVAL_STAGE_TO_STAGING,
                target_staging_key=staging_key,
                conflict_with_long_term=bool(old_value and old_value != normalized_value),
                low_confidence=False,
                missing_required_detail=False,
                evidence_json=[],
                manual_review_reason=payload.note or "Employee-submitted Staging update request.",
                review_flags_json=[],
            )
            created += 1
        if created == 0:
            raise HTTPException(status_code=400, detail="There are no changes to submit.")
        write_audit_log(
            conn,
            request,
            actor,
            action_type="staging_update_request_submitted",
            target_table="staging_update_requests",
            target_record_id=submission_group_id,
            building_name=snapshot["building_name"],
            source=source_file,
            note=payload.note or "Submitted staging update request.",
        )
        result = {
            "ok": True,
            "group_id": submission_group_id,
            "created_records": created,
            "message": "The Staging update request was submitted and is awaiting administrator review.",
        }
        complete_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_request_update:{staging_key}",
            idempotency_key=idempotency_key,
            response_payload=result,
        )
        return result


@app.post("/staging/buildings/{staging_key}/promote")
async def promote_staging_building(
    staging_key: str,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin")),
):
    ensure_excel_mirror_ready_for_write()
    return await submit_staging_building_for_review(
        staging_key=staging_key,
        request=request,
        idempotency_key=idempotency_key,
        user=user,
    )


@app.patch("/staging/buildings/{staging_key}")
async def update_staging_building(
    staging_key: str,
    payload: MasterUpdateRequest,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        replay = begin_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_update:{staging_key}",
            idempotency_key=idempotency_key,
            payload={
                "staging_key": staging_key,
                "updates": payload.updates,
                "note": payload.note,
            },
        )
        if replay and replay.get("replay"):
            return replay["response"]
        snapshot = load_staging_building_snapshot(conn, staging_key)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Staging record not found.")

        previous_identity = (
            normalize_unknown_value(snapshot.get("building_name")) or "",
            normalize_unknown_value(snapshot.get("address")) or "",
        )
        updates = dict(payload.updates or {})
        for field_key in ("source_date", "source_type", "source_file", "info_cutoff_date", "library_status"):
            updates.pop(field_key, None)

        workbook_snapshot = {
            "building_name": previous_identity[0],
            "address": previous_identity[1],
            "insurance_required": snapshot.get("insurance_required"),
            "insurance_coverage_amount": snapshot.get("insurance_coverage_amount"),
            "electricity_required": snapshot.get("electricity_required"),
            "electricity_provider": snapshot.get("electricity_provider"),
            "internet_self_setup_required": snapshot.get("internet_self_setup_required"),
            "internet_provider": snapshot.get("internet_provider"),
            "internet_notes": snapshot.get("internet_notes"),
            "move_in_notes": snapshot.get("move_in_notes"),
            "extensions": dict(snapshot.get("extensions") or {}),
            "library_status": snapshot.get("library_status") or STAGING_STATUS_PENDING,
        }

        allowed_fields = set(workbook_catalog_field_keys(conn, include_staging_only=True))
        extension_field_keys = set(workbook_catalog_extension_field_keys(conn, include_staging_only=True))
        for field_key, raw_value in updates.items():
            if field_key not in allowed_fields:
                continue
            if field_key in NETWORK_PROVIDER_FIELD_MAP:
                normalized = normalize_field_value(field_key, "boolean", raw_value)
                if normalized is None:
                    workbook_snapshot["extensions"].pop(field_key, None)
                else:
                    workbook_snapshot["extensions"][field_key] = normalized
                continue
            if field_key in extension_field_keys:
                definition = find_field_definition(conn, field_key)
                field_type = definition["field_type"] if definition else "text"
                normalized = normalize_field_value(field_key, field_type, raw_value)
                if normalized is None:
                    workbook_snapshot["extensions"].pop(field_key, None)
                else:
                    workbook_snapshot["extensions"][field_key] = normalized
                continue
            definition = find_field_definition(conn, field_key)
            field_type = definition["field_type"] if definition else "text"
            workbook_snapshot[field_key] = normalize_field_value(field_key, field_type, raw_value)

        network_keys = set(NETWORK_PROVIDER_FIELD_MAP) | {"internet_provider", "internet_notes"}
        if any(field_key in updates for field_key in network_keys):
            combined_network_values: Dict[str, Optional[str]] = {
                field_key: workbook_snapshot["extensions"].get(field_key)
                for field_key in NETWORK_PROVIDER_FIELD_MAP
            }
            combined_network_values["internet_provider"] = workbook_snapshot.get("internet_provider")
            combined_network_values["internet_notes"] = workbook_snapshot.get("internet_notes")
            workbook_snapshot["internet_provider"] = build_network_provider_text_from_values(combined_network_values)

        if not normalize_unknown_value(workbook_snapshot.get("building_name")):
            raise HTTPException(status_code=400, detail="Could not save the Staging record: the building name cannot be empty.")

        sync_staging_workbook_for_building(workbook_snapshot, previous_identity=previous_identity)
        refresh_after_staging_excel_write(conn, actor=actor, request=request)
        refreshed_key = build_staging_snapshot_key(
            normalize_unknown_value(workbook_snapshot.get("building_name")) or "",
            normalize_unknown_value(workbook_snapshot.get("address")) or "",
        )
        refreshed_snapshot = load_staging_building_snapshot(conn, refreshed_key)
        if not refreshed_snapshot:
            raise HTTPException(status_code=409, detail="The Staging workbook was saved, but the corresponding building was not found after refreshing the mirror.")
        write_audit_log(
            conn,
            request,
            actor,
            action_type="staging_direct_update",
            target_table="staging_building_info",
            target_record_id=refreshed_key,
            building_name=refreshed_snapshot["building_name"],
            source=Path(resolve_staging_excel_path()).name,
            note=payload.note or "Staging edit from the frontend",
        )
        complete_idempotent_request(
            conn,
            actor=actor,
            scope=f"staging_update:{staging_key}",
            idempotency_key=idempotency_key,
            response_payload=refreshed_snapshot,
        )
        return refreshed_snapshot


@app.patch("/master/buildings/{building_id}")
async def update_master_building(
    building_id: str,
    payload: MasterUpdateRequest,
    request: Request,
    idempotency_key: Optional[str] = Header(default=None, alias="X-Idempotency-Key"),
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        replay = begin_idempotent_request(
            conn,
            actor=actor,
            scope=f"master_update:{building_id}",
            idempotency_key=idempotency_key,
            payload={
                "building_id": building_id,
                "updates": payload.updates,
                "note": payload.note,
            },
        )
        if replay and replay.get("replay"):
            return replay["response"]
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
        previous_identity = (
            normalize_unknown_value(snapshot.get("building_name")) or "",
            normalize_unknown_value(snapshot.get("address")) or "",
        )
        updates = dict(payload.updates or {})
        if "source_date" in updates:
            updates.pop("source_date", None)
        target_building_name = (
            normalize_unknown_value(updates.get("building_name")) or snapshot.get("building_name") or ""
            if "building_name" in updates
            else snapshot.get("building_name") or ""
        )
        target_address = (
            normalize_unknown_value(updates.get("address")) or ""
            if "address" in updates
            else snapshot.get("address") or ""
        )
        if not normalize_unknown_value(target_building_name):
            raise HTTPException(status_code=400, detail="Could not save the Master record: the building name cannot be empty.")

        network_keys = set(NETWORK_PROVIDER_FIELD_MAP) | {"internet_provider", "internet_notes"}
        if any(field_key in updates for field_key in network_keys):
            combined_network_values: Dict[str, Optional[str]] = {
                field_key: updates.get(field_key, snapshot.get("extensions", {}).get(field_key))
                for field_key in NETWORK_PROVIDER_FIELD_MAP
            }
            combined_network_values["internet_provider"] = updates.get(
                "internet_provider",
                snapshot.get("internet_provider"),
            )
            combined_network_values["internet_notes"] = updates.get(
                "internet_notes",
                snapshot.get("internet_notes"),
            )
            updates["internet_provider"] = build_network_provider_text_from_values(combined_network_values)

        if any(field_key in updates for field_key in DETAILED_INSURANCE_FIELD_KEYS):
            detail_payloads = {}
            for field_key in DETAILED_INSURANCE_FIELD_KEYS:
                if field_key in updates:
                    detail_payloads[field_key] = {"value": updates.get(field_key), "confidence": 1.0}
                else:
                    current_value = snapshot.get("extensions", {}).get(field_key)
                    if current_value is not None:
                        detail_payloads[field_key] = {"value": current_value, "confidence": 1.0}
            derived = derive_legacy_insurance_fields_from_detailed(detail_payloads)
            for field_key, payload in derived.items():
                updates[field_key] = payload.get("value")
            if "insurance_renters_required" in updates and not normalize_unknown_value(updates.get("insurance_renters_required")):
                updates["insurance_required"] = ""
            if "insurance_renters_minimum_coverage" in updates and not normalize_unknown_value(
                updates.get("insurance_renters_minimum_coverage")
            ):
                updates["insurance_coverage_amount"] = ""

        for field_key, raw_value in updates.items():
            definition = find_field_definition(conn, field_key)
            if not definition:
                continue
            normalized_value = normalize_field_value(field_key, definition["field_type"], raw_value)
            old_value, stored_value = upsert_master_field(
                conn,
                building_id=building_id,
                field_key=field_key,
                new_value=normalized_value,
                actor=actor,
                source_type="direct_edit",
                source_file="manual",
                source_date=None,
                info_cutoff_date=None,
            )
            write_audit_log(
                conn,
                request,
                actor,
                action_type="master_direct_update",
                target_table="master_building_info"
                if field_key in CORE_MASTER_FIELD_KEYS
                else "master_building_field_values",
                target_record_id=building_id,
                building_name=snapshot["building_name"],
                field_name=field_key,
                old_value=old_value or "",
                new_value=stored_value or "",
                note=payload.note,
            )
        refresh_master_completeness(
            conn,
            building_id,
            verification_note=payload.note.strip() or None,
        )
        bump_master_version(conn, building_id, actor)
        sync_master_workbook_for_building(conn, building_id, previous_identity=previous_identity)
        refresh_after_master_excel_write(conn, actor=actor, request=request)
        refreshed_master = lookup_master_building(conn, target_building_name, target_address)
        refreshed_snapshot = load_master_building_snapshot(conn, refreshed_master["id"]) if refreshed_master else None
        if refreshed_snapshot:
            complete_idempotent_request(
                conn,
                actor=actor,
                scope=f"master_update:{building_id}",
                idempotency_key=idempotency_key,
                response_payload=refreshed_snapshot,
            )
            return refreshed_snapshot
        result = {
            "removed_from_master": True,
            "building_name": target_building_name,
            "address": target_address,
            "message": "This record now contains only a building name and address. It was removed from the Master mirror, while its row remains in the master workbook.",
        }
        complete_idempotent_request(
            conn,
            actor=actor,
            scope=f"master_update:{building_id}",
            idempotency_key=idempotency_key,
            response_payload=result,
        )
        return result


@app.delete("/master/buildings/{building_id}")
async def delete_master_building(
    building_id: str,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
        previous_identity = (
            normalize_unknown_value(snapshot.get("building_name")) or "",
            normalize_unknown_value(snapshot.get("address")) or "",
        )
        conn.execute(
            """
            UPDATE staging_update_requests
            SET building_id = NULL, updated_at = ?
            WHERE building_id = ?
            """,
            (now_iso(), building_id),
        )
        conn.execute("DELETE FROM master_building_field_values WHERE building_id = ?", (building_id,))
        conn.execute("DELETE FROM master_building_info WHERE id = ?", (building_id,))
        workbook_path = resolve_master_excel_path()
        if workbook_path.exists():
            delete_building_snapshot(previous_identity[0], previous_identity[1])
        refresh_after_master_excel_write(conn, actor=actor, request=request)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="master_building_deleted",
            target_table="master_building_info",
            target_record_id=building_id,
            building_name=previous_identity[0],
            old_value=json_dumps(
                {
                    "building_name": snapshot.get("building_name"),
                    "address": snapshot.get("address"),
                }
            ),
            note="Direct master building deletion.",
        )
    return {"ok": True, "building_id": building_id}


@app.post("/master/buildings/{building_id}/rollback")
async def rollback_master_building(
    building_id: str,
    payload: RollbackRequest,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    actor = get_actor(user)
    ensure_excel_mirror_ready_for_write()
    with db_connection() as conn:
        snapshot = load_master_building_snapshot(conn, building_id)
        if not snapshot:
            raise HTTPException(status_code=404, detail="Master record not found.")
        previous_identity = (
            normalize_unknown_value(snapshot.get("building_name")) or "",
            normalize_unknown_value(snapshot.get("address")) or "",
        )

        sql = """
            SELECT id, action_type, target_table, target_record_id, building_name, field_name, old_value, new_value, created_at
            FROM audit_logs
            WHERE target_record_id = ?
              AND target_table IN ('master_building_info', 'master_building_field_values')
              AND field_name != ''
        """
        params: List[Any] = [building_id]
        if payload.audit_log_id:
            sql += " AND id = ?"
            params.append(payload.audit_log_id)
        sql += " ORDER BY created_at DESC LIMIT 1"
        target_log = conn.execute(sql, params).fetchone()
        if not target_log:
            raise HTTPException(status_code=404, detail="No Master change record is available to roll back.")

        definition = find_field_definition(conn, target_log["field_name"])
        if not definition:
            raise HTTPException(status_code=400, detail="The target field no longer exists, so the change cannot be rolled back automatically.")

        old_value, stored_value = upsert_master_field(
            conn,
            building_id=building_id,
            field_key=target_log["field_name"],
            new_value=target_log["old_value"],
            actor=actor,
            source_type="rollback",
            source_file=f"audit:{target_log['id']}",
            source_date=None,
            info_cutoff_date=None,
        )
        if target_log["field_name"] in DETAILED_INSURANCE_FIELD_KEYS:
            refreshed_snapshot = load_master_building_snapshot(conn, building_id)
            detail_payloads = {
                field_key: {"value": insurance_field_value(refreshed_snapshot, field_key), "confidence": 1.0}
                for field_key in DETAILED_INSURANCE_FIELD_KEYS
                if insurance_field_value(refreshed_snapshot, field_key) is not None
            }
            for field_key, payload in derive_legacy_insurance_fields_from_detailed(detail_payloads).items():
                upsert_master_field(
                    conn,
                    building_id=building_id,
                    field_key=field_key,
                    new_value=payload.get("value"),
                    actor=actor,
                    source_type="rollback",
                    source_file=f"audit:{target_log['id']}",
                    source_date=None,
                    info_cutoff_date=None,
                )
        refresh_master_completeness(
            conn,
            building_id,
            verification_note=payload.note.strip() or None,
        )
        bump_master_version(conn, building_id, actor)
        write_audit_log(
            conn,
            request,
            actor,
            action_type="master_rollback",
            target_table=target_log["target_table"],
            target_record_id=building_id,
            building_name=snapshot["building_name"],
            field_name=target_log["field_name"],
            old_value=old_value or "",
            new_value=stored_value or "",
            source=f"audit:{target_log['id']}",
            note=payload.note.strip() or f"Rollback from audit log {target_log['id']}",
        )
        if target_log["field_name"] in NETWORK_PROVIDER_FIELD_MAP or target_log["field_name"] == "internet_provider":
            refresh_master_network_provider_text(
                conn,
                building_id=building_id,
                actor=actor,
                source_type="rollback",
                source_file=f"audit:{target_log['id']}",
            )
        sync_master_workbook_for_building(conn, building_id, previous_identity=previous_identity)
        refresh_after_master_excel_write(conn, actor=actor, request=request)
        current_building_name = (
            normalize_unknown_value(target_log["old_value"])
            if target_log["field_name"] == "building_name"
            else snapshot.get("building_name")
        ) or snapshot.get("building_name") or ""
        current_address = (
            normalize_unknown_value(target_log["old_value"])
            if target_log["field_name"] == "address"
            else snapshot.get("address")
        ) or ""
        refreshed_master = lookup_master_building(conn, current_building_name, current_address)
        return {
            "building": load_master_building_snapshot(conn, refreshed_master["id"]) if refreshed_master else None,
            "rolled_back_from": dict(target_log),
        }


@app.get("/audit-logs")
async def list_audit_logs(
    limit: int = Query(default=100, ge=1, le=500),
    action_type: str = Query(default=""),
    building_name: str = Query(default=""),
    field_name: str = Query(default=""),
    user_role: str = Query(default=""),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    with db_connection() as conn:
        sql = """
            SELECT id, user_id, user_role, action_type, target_table, target_record_id, building_name,
                   field_name, old_value, new_value, source, ip_address, user_agent, note, created_at
            FROM audit_logs
            WHERE 1 = 1
        """
        params: List[Any] = []
        if action_type.strip():
            sql += " AND action_type = ?"
            params.append(action_type.strip())
        if building_name.strip():
            sql += " AND lower(building_name) LIKE ?"
            params.append(f"%{building_name.strip().lower()}%")
        if field_name.strip():
            sql += " AND field_name = ?"
            params.append(field_name.strip())
        if user_role.strip():
            sql += " AND user_role = ?"
            params.append(user_role.strip())
        sql += " ORDER BY created_at DESC LIMIT ?"
        params.append(limit)
        rows = conn.execute(sql, params).fetchall()
    return {"logs": rows}


@app.post("/bootstrap/legacy")
async def bootstrap_legacy_sources(
    request: Request,
    force: bool = Query(default=False),
    user: dict = Depends(require_roles("super_admin", "admin")),
):
    actor = get_actor(user)
    with db_connection() as conn:
        existing = conn.execute(
            "SELECT value FROM app_meta WHERE key = 'legacy_bootstrap_completed_at'"
        ).fetchone()
        if existing and not force:
            raise HTTPException(status_code=409, detail="Legacy bootstrap has already run. Pass force=true to run it again.")

        summary = {"building_spreadsheets": 0, "internet_records": 0, "pdf_sources": 0}

        for doc_path in legacy._iter_supported_document_paths():
            if doc_path.suffix.lower() not in legacy.SPREADSHEET_FILE_EXTENSIONS:
                continue
            relative_path = str(doc_path.relative_to(legacy.PDF_LIBRARY_DIR))
            category = legacy._category_from_doc_path(doc_path)
            if category == "internet":
                continue

            batch_id = f"import_{uuid.uuid4().hex}"
            preview = await preview_excel_file(doc_path, conn)
            sheets = []
            for sheet_preview in preview["sheets"]:
                mappings = []
                for header in sheet_preview["headers"]:
                    suggestion = header.get("suggested")
                    if suggestion:
                        mappings.append(
                            HeaderMappingSelection(
                                original_header=header["original_header"],
                                mapped_field_key=suggestion["field_key"],
                                action="map",
                            )
                        )
                    else:
                        mappings.append(
                            HeaderMappingSelection(
                                original_header=header["original_header"],
                                mapped_field_key=None,
                                action="create" if actor.role == "super_admin" else "ignore",
                                new_field_display_name=header["original_header"],
                            )
                        )
                sheets.append(
                    SheetImportConfirmation(
                        sheet_name=sheet_preview["sheet_name"],
                        header_row_index=sheet_preview["header_row_index"],
                        mappings=mappings,
                    )
                )

            conn.execute(
                """
                INSERT INTO import_batches(
                  id, original_file_name, stored_path, file_hash, status, sheet_names, uploaded_by, confirmed_by, created_at, updated_at
                ) VALUES(?, ?, ?, ?, 'legacy_bootstrap', ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    doc_path.name,
                    str(doc_path),
                    hashlib.sha256(doc_path.read_bytes()).hexdigest(),
                    json_dumps([item["sheet_name"] for item in preview["sheets"]]),
                    actor.user_id,
                    actor.user_id,
                    now_iso(),
                    now_iso(),
                ),
            )
            result = persist_import_to_staging(
                conn,
                batch_id=batch_id,
                file_name=doc_path.name,
                sheets=sheets,
                file_path=doc_path,
                actor=actor,
                request=request,
            )
            summary["building_spreadsheets"] += result["rows"]
            write_audit_log(
                conn,
                request,
                actor,
                action_type="legacy_excel_bootstrap",
                target_table="import_batches",
                target_record_id=batch_id,
                source=relative_path,
                note=f"Imported {result['rows']} rows from legacy spreadsheet.",
            )

        internet_status = legacy.reload_internet_store()
        for record in legacy.INTERNET_STORE.values():
            submission_group_id = f"legacy-internet:{record.internet_id}"
            building_name = record.building_name or record.address or "Unnamed"
            building_match = lookup_master_building(conn, building_name, record.address or "")
            building_id = building_match["id"] if building_match else None

            self_setup_value = None
            wifi_mode = (record.wifi_mode or "").lower()
            if any(token in wifi_mode for token in ("自行", "self", "choose")):
                self_setup_value = "true"
            elif any(token in wifi_mode for token in ("自带", "包含", "free", "强制")):
                self_setup_value = "false"

            provider_names = [
                item["label"]
                for item in record.providers.values()
                if item.get("status") in {"supported", "pending"}
            ]
            provider_text = ", ".join(provider_names) if provider_names else None
            notes = normalize_unknown_value(
                "；".join(
                    filter(
                        None,
                        [
                            record.notes,
                            record.mode_detail,
                            "；".join(
                                filter(None, [provider.get("note"), provider.get("plans")])
                                for provider in record.providers.values()
                                if provider.get("note") or provider.get("plans")
                            ),
                        ],
                    )
                )
            )
            row_values = {
                "building_name": building_name,
                "address": normalize_unknown_value(record.address),
                "internet_self_setup_required": self_setup_value,
                "internet_provider": provider_text,
                "internet_notes": notes,
            }
            source_content = json_dumps(
                {
                    "website": record.website,
                    "contact": record.contact,
                    "wifi_mode": record.wifi_mode,
                    "mode_detail": record.mode_detail,
                    "providers": record.providers,
                }
            )
            for field_key, value in row_values.items():
                if value is None:
                    continue
                old_value = get_master_field_value(conn, building_id, field_key) if building_id else None
                create_staging_request(
                    conn,
                    submission_group_id=submission_group_id,
                    building_name=building_name,
                    building_id=building_id,
                    field_name=field_key,
                    old_value=old_value,
                    new_value=value,
                    source_type="legacy_internet_excel",
                    source_content=source_content,
                    source_file=record.source_file,
                    submitted_by=actor.user_id,
                    ai_confidence=None,
                    review_status="pending",
                    import_batch_id=None,
                    parser_type="excel_header_mapping",
                    raw_input_type="excel",
                    source_document_id=None,
                    conflict_with_long_term=bool(old_value and old_value != value),
                    low_confidence=False,
                    missing_required_detail=missing_required_detail_for_field(field_key, row_values),
                )
                summary["internet_records"] += 1

        for doc_path in legacy._iter_supported_document_paths():
            if doc_path.suffix.lower() != ".pdf":
                continue
            category = legacy._category_from_doc_path(doc_path)
            if category == "internet":
                continue
            extracted_pages = extract_pdf_pages(doc_path)
            extracted_text = join_page_texts(extracted_pages) or legacy._extract_document_text(doc_path)
            source_document_id = await create_source_document_record(
                conn,
                raw_input_type="pdf",
                parser_type="legacy_pdf_text",
                source_type="legacy_pdf",
                source_file=doc_path.name,
                stored_path=doc_path,
                source_content="",
                extracted_text=extracted_text,
                extracted_pages=extracted_pages,
                actor=actor,
                request=request,
            )
            await parse_source_to_staging(
                conn,
                source_document_id=source_document_id,
                source_file=doc_path.name,
                raw_input_type="pdf",
                parser_type="legacy_pdf_text",
                source_type="legacy_pdf",
                source_content="",
                extracted_text=extracted_text,
                extracted_pages=extracted_pages,
                actor=actor,
                request=request,
            )
            summary["pdf_sources"] += 1

        write_audit_log(
            conn,
            request,
            actor,
            action_type="legacy_bootstrap_completed",
            target_table="source_documents",
            source="legacy_demo",
            note=json_dumps(summary),
        )
        conn.execute(
            "INSERT OR REPLACE INTO app_meta(key, value) VALUES('legacy_bootstrap_completed_at', ?)",
            (now_iso(),),
        )
    return {"status": internet_status, "summary": summary}


@app.post("/query/answer")
async def query_answer(
    payload: QueryAnswerRequest,
    user: dict = Depends(get_current_user),
):
    source_mode = payload.source_mode or "master"
    explanation_enabled = ai_explanation_enabled()
    network_panel_hint = False
    with db_connection() as conn:
        building = None
        suggestions = []
        building_switch_candidate = None
        selection_conflict_message = ""
        if source_mode == "staging":
            ranked = rank_staging_buildings(conn, payload.question)
            if payload.staging_key:
                building = load_staging_building_snapshot(conn, payload.staging_key)
            if building:
                building_switch_candidate, selection_conflict_message = detect_query_selection_conflict(
                    query=payload.question,
                    source_mode="staging",
                    current_record=building,
                    ranked_records=ranked,
                )
                suggestions = [row for _, row in ranked[:5]]
            if not building and ranked:
                building = ranked[0][1]
                suggestions = [row for _, row in ranked[:5]]
        else:
            ranked = rank_master_buildings(conn, payload.question)
            if payload.building_id:
                building = load_master_building_snapshot(conn, payload.building_id)
            if building:
                building_switch_candidate, selection_conflict_message = detect_query_selection_conflict(
                    query=payload.question,
                    source_mode="master",
                    current_record=building,
                    ranked_records=ranked,
                )
                suggestions = [row for _, row in ranked[:5]]
            if not building and ranked:
                building = load_master_building_snapshot(conn, ranked[0][1]["id"])
                suggestions = [row for _, row in ranked[:5]]

        if not building:
            field_keys = matched_field_keys_for_question(conn, payload.question)
            network_panel_hint = is_network_question(payload.question, field_keys)
            temporary_suggestions = []
            if source_mode == "master":
                temporary_suggestions = [
                    serialize_query_building(item, "staging")
                    for _, item in rank_staging_buildings(conn, payload.question)[:5]
                ]
            not_found_message = (
                "No matching building was found in Staging. Try another building name or submit source material to Staging first."
                if source_mode == "staging"
                else "No matching building was found in Master. Select a building or complete the review and approval workflow first."
            )
            return {
                "matched": None,
                "fact_answer": None,
                "ai_answer": None,
                "answer_mode": "database-only",
                "ai_explanation_enabled": explanation_enabled,
                "building_switch_candidate": None,
                "selection_conflict_message": "",
                "network_panel_hint": network_panel_hint,
                "source_mode": source_mode,
                "message": not_found_message,
                "answer": not_found_message,
                "suggestions": [
                    {
                        "id": item["id"],
                        "staging_key": item.get("staging_key") or item.get("id"),
                        "building_name": item["building_name"],
                        "address": item["address"],
                    }
                    for item in suggestions
                ],
                "temporary_suggestions": temporary_suggestions,
            }

        field_keys = matched_field_keys_for_question(conn, payload.question)
        network_panel_hint = is_network_question(payload.question, field_keys)
        if building_switch_candidate:
            return {
                "matched": serialize_query_building(building, source_mode),
                "field_keys": field_keys,
                "fact_answer": None,
                "ai_answer": None,
                "answer_mode": "database-only",
                "ai_explanation_enabled": explanation_enabled,
                "building_switch_candidate": building_switch_candidate,
                "selection_conflict_message": selection_conflict_message,
                "network_panel_hint": network_panel_hint,
                "source_mode": source_mode,
                "answer": selection_conflict_message,
                "message": selection_conflict_message,
                "suggestions": [
                    serialize_query_building(item, source_mode)
                    for item in suggestions[:5]
                ],
            }
        fact_answer = build_structured_answer(conn, building, payload.question, field_keys, source_mode)
        ai_answer = None
        if explanation_enabled and payload.include_ai:
            ai_answer = await generate_fact_explanation(
                question=payload.question,
                snapshot=building,
                field_keys=field_keys,
                fact_answer=fact_answer,
                source_mode=source_mode,
            )
        answer_mode = "database-plus-ai" if ai_answer else "database-only"
        return {
            "matched": {
                "id": building["id"],
                "staging_key": building.get("staging_key"),
                "building_name": building["building_name"],
                "address": building["address"],
                "source_mode": source_mode,
                "completeness_status": building.get("completeness_status"),
                "completeness_score": building.get("completeness_score"),
                "verification_note": building.get("verification_note"),
            },
            "field_keys": field_keys,
            "fact_answer": fact_answer,
            "ai_answer": ai_answer,
            "answer_mode": answer_mode,
            "ai_explanation_enabled": explanation_enabled,
            "building_switch_candidate": None,
            "selection_conflict_message": "",
            "network_panel_hint": network_panel_hint,
            "source_mode": source_mode,
            "answer": fact_answer,
            "message": fact_answer,
        }


@app.post("/query/explanation")
async def query_explanation(
    payload: QueryExplanationRequest,
    user: dict = Depends(get_current_user),
):
    if not ai_explanation_enabled():
        return {
            "ai_answer": None,
            "answer_mode": "database-only",
            "message": "AI explanation is currently unavailable.",
        }
    source_mode = payload.source_mode or "master"
    with db_connection() as conn:
        if source_mode == "staging":
            snapshot = load_staging_building_snapshot(conn, payload.staging_key) if payload.staging_key else None
            if payload.staging_key and not snapshot:
                raise HTTPException(status_code=404, detail="The Staging building snapshot has changed. Run the query again.")
            if not payload.staging_key and not snapshot:
                ranked = rank_staging_buildings(conn, payload.question)
                snapshot = ranked[0][1] if ranked else None
            record_id = snapshot.get("staging_key") if snapshot else ""
        else:
            snapshot = load_master_building_snapshot(conn, payload.building_id) if payload.building_id else None
            if payload.building_id and not snapshot:
                raise HTTPException(status_code=404, detail="The Master building snapshot has changed. Run the query again.")
            if not payload.building_id and not snapshot:
                ranked = rank_master_buildings(conn, payload.question)
                snapshot = load_master_building_snapshot(conn, ranked[0][1]["id"]) if ranked else None
            record_id = snapshot.get("id") if snapshot else ""
        if not snapshot:
            raise HTTPException(status_code=404, detail="No building record was found to explain.")
        field_keys = matched_field_keys_for_question(conn, payload.question)
        fact_answer = build_structured_answer(conn, snapshot, payload.question, field_keys, source_mode)
    ai_answer, cache_status, _cache_key = await generate_cached_fact_explanation(
        source_mode=source_mode,
        record_id=record_id,
        question=payload.question,
        snapshot=snapshot,
        fact_answer=fact_answer,
    )
    return {
        "ai_answer": ai_answer,
        "answer_mode": "database-plus-ai" if ai_answer else "database-only",
        "cache_status": cache_status,
        "message": ai_answer or "AI explanation is currently unavailable.",
        "matched": serialize_query_building(snapshot, source_mode),
    }


@app.post("/chat")
async def chat_answer(payload: ChatRequest, user: dict = Depends(get_current_user)):
    result = await query_answer(
        QueryAnswerRequest(
            building_id=payload.building_id,
            staging_key=payload.staging_key,
            source_mode=payload.source_mode,
            question=payload.question,
        ),
        user,
    )
    content = result.get("fact_answer") or result.get("answer") or result.get("message") or "An authoritative answer is not currently available."
    return stream_text_response(content)


@app.post("/chat/mobile")
async def chat_answer_mobile(payload: ChatRequest, user: dict = Depends(get_current_user)):
    result = await query_answer(
        QueryAnswerRequest(
            building_id=payload.building_id,
            staging_key=payload.staging_key,
            source_mode=payload.source_mode,
            question=payload.question,
        ),
        user,
    )
    content = result.get("fact_answer") or result.get("answer") or result.get("message") or "An authoritative answer is not currently available."
    return {
        "content": content,
        "fact_answer": result.get("fact_answer") or content,
        "ai_answer": result.get("ai_answer"),
        "answer_mode": result.get("answer_mode", "database-only"),
        "ai_explanation_enabled": result.get("ai_explanation_enabled", False),
        "building_switch_candidate": result.get("building_switch_candidate"),
        "selection_conflict_message": result.get("selection_conflict_message", ""),
        "network_panel_hint": result.get("network_panel_hint", False),
        "building_id": result.get("matched", {}).get("id", payload.building_id),
        "staging_key": result.get("matched", {}).get("staging_key", payload.staging_key),
        "source_mode": result.get("source_mode", payload.source_mode),
        "label": result.get("matched", {}).get("building_name", ""),
        "model": "database-facts",
    }


@app.get("/legacy/buildings")
async def legacy_buildings(reload: bool = Query(default=False), user: dict = Depends(get_current_user)):
    if reload:
        legacy.reload_building_store()
    return {
        "total": len(legacy.BUILDING_STORE),
        "buildings": [legacy._summary_from_doc(doc) for doc in legacy.BUILDING_STORE.values()],
    }


@app.post("/legacy/buildings/reload")
async def legacy_buildings_reload(user: dict = Depends(require_roles("super_admin", "admin"))):
    return legacy.reload_building_store()


@app.get("/legacy/internet")
async def legacy_internet(user: dict = Depends(get_current_user)):
    return {
        "total": len(legacy.INTERNET_STORE),
        "warnings": list(legacy.INTERNET_WARNINGS),
        "errors": list(legacy.INTERNET_ERRORS),
        "items": [legacy._internet_result_payload(item) for item in legacy.INTERNET_STORE.values()],
    }


def build_health_detail() -> dict:
    with db_connection() as conn:
        master_total = conn.execute("SELECT COUNT(*) AS total FROM master_building_info").fetchone()["total"]
        staging_total = conn.execute("SELECT COUNT(*) AS total FROM staging_building_info").fetchone()["total"]
        review_total = conn.execute("SELECT COUNT(*) AS total FROM staging_update_requests").fetchone()["total"]
        user_total = conn.execute("SELECT COUNT(*) AS total FROM users").fetchone()["total"]
        recent_ocr_success = conn.execute(
            "SELECT provider, completed_at, duration_ms, metadata_json FROM ocr_jobs WHERE status IN ('completed', 'success') ORDER BY completed_at DESC LIMIT 1"
        ).fetchone()
        recent_ocr_failure = conn.execute(
            "SELECT provider, completed_at, error_code FROM ocr_jobs WHERE status IN ('failed', 'empty') ORDER BY completed_at DESC LIMIT 1"
        ).fetchone()
        ocr_fallback_total = conn.execute(
            "SELECT COUNT(*) AS total FROM ocr_jobs WHERE metadata_json LIKE '%\"fallback_used\": true%'"
        ).fetchone()["total"]
    master_excel_path = resolve_master_excel_path()
    staging_excel_path = resolve_staging_excel_path()
    runtime_status = read_runtime_status()
    tunnel_health = probe_tunnel_health()
    frontend_healthy = (
        FRONTEND_INDEX_PATH.is_file() if (runtime_status.get("runtime_mode") or RUNTIME_MODE) == "daemon" else None
    )
    excel_healthy = runtime_status.get("excel_mirror_healthy")
    status = "degraded" if excel_healthy is False or frontend_healthy is False else "ok"
    return {
        "status": status,
        "runtime_mode": runtime_status.get("runtime_mode") or RUNTIME_MODE,
        "backend_healthy": True,
        "frontend_healthy": frontend_healthy,
        "excel_mirror_healthy": excel_healthy,
        "excel_mirror_last_error": runtime_status.get("excel_mirror_last_error"),
        "default_password_risk": runtime_status.get("default_password_risk", False),
        "default_password_accounts": runtime_status.get("default_password_accounts", []),
        "last_recovery_at": runtime_status.get("last_recovery_at"),
        "last_recovery_reason": runtime_status.get("last_recovery_reason"),
        "last_boot_at": runtime_status.get("last_boot_at"),
        "tunnel_health": tunnel_health,
        "ai_explanation_enabled": ai_explanation_enabled(),
        "vision_enabled": vision_client() is not None,
        "vision_model": os.getenv("VISION_MODEL", "").strip() or None,
        "vision_model_resolved": resolved_vision_model_name() or None,
        "vision_last_error": VISION_LAST_ERROR,
        "ocr_provider": configured_ocr_router().primary,
        "ocr_fallback_provider": configured_ocr_router().fallback,
        "ocr_provider_configured": configured_ocr_router().providers[configured_ocr_router().primary].configured(),
        "baidu_ocr_configured": bool(BAIDU_OCR_API_KEY and BAIDU_OCR_SECRET_KEY),
        "unlimited_ocr_local_configured": configured_ocr_router().providers[
            "unlimited_ocr_local_http"
        ].configured(),
        "unlimited_ocr_local_model": UNLIMITED_OCR_LOCAL_MODEL,
        "ocr_recent_success": dict(recent_ocr_success) if recent_ocr_success else None,
        "ocr_recent_failure": dict(recent_ocr_failure) if recent_ocr_failure else None,
        "ocr_fallback_total": ocr_fallback_total,
        "pdf_image_render_available": bool(shutil.which("pdftoppm") or fitz is not None),
        "ocr_available": shutil.which("tesseract") is not None,
        "apple_vision_ocr_available": sys.platform == "darwin" and shutil.which("swift") is not None,
        "db_path": str(DB_PATH),
        "upload_root": str(UPLOAD_ROOT),
        "master_excel_path": str(master_excel_path),
        "master_excel_exists": master_excel_path.exists(),
        "staging_excel_path": str(staging_excel_path),
        "staging_excel_exists": staging_excel_path.exists(),
        "frontend_dist_dir": str(FRONTEND_DIST_DIR),
        "packaged_frontend_available": FRONTEND_INDEX_PATH.is_file(),
        "master_buildings": master_total,
        "staging_buildings": staging_total,
        "staging_records": review_total,
        "users": user_total,
        "legacy_documents": len(legacy.BUILDING_STORE),
        "legacy_internet_records": len(legacy.INTERNET_STORE),
        "legacy_pdf_library_dir": str(legacy.PDF_LIBRARY_DIR),
        "legacy_image_library_dir": str(legacy.IMAGE_LIBRARY_DIR),
    }


@app.get("/health")
async def health_check():
    detail = build_health_detail()
    return {
        "status": detail["status"],
        "runtime_mode": detail["runtime_mode"],
        "backend_healthy": detail["backend_healthy"],
        "frontend_healthy": detail["frontend_healthy"],
        "ocr_provider": detail["ocr_provider"],
        "ocr_fallback_provider": detail["ocr_fallback_provider"],
        "ocr_provider_configured": detail["ocr_provider_configured"],
        "vision_enabled": detail["vision_enabled"],
    }


@app.get("/admin/health")
async def admin_health_check(user: dict = Depends(require_roles("super_admin", "admin"))):
    return build_health_detail()


@app.get("/system/update-status")
async def system_update_status(
    check_remote: bool = Query(default=False),
    user: dict = Depends(require_roles("super_admin")),
):
    return build_update_status(check_remote=check_remote)


@app.post("/system/update")
async def system_update(
    payload: SystemUpdateRequest,
    background_tasks: BackgroundTasks,
    request: Request,
    user: dict = Depends(require_roles("super_admin")),
):
    root = git_root_path()
    if not root:
        raise HTTPException(status_code=400, detail="The current directory is not a Git repository and cannot be updated remotely.")
    if not UPDATE_SCRIPT_PATH.is_file():
        raise HTTPException(status_code=400, detail=f"Update script not found: {UPDATE_SCRIPT_PATH}")
    status_before = build_update_status(check_remote=True)
    if status_before.get("dirty") and not payload.allow_dirty:
        raise HTTPException(status_code=400, detail="The code directory has uncommitted changes. Commit or clean them before updating to avoid overwriting work.")
    UPDATE_LOG_DIR.mkdir(parents=True, exist_ok=True)
    env = os.environ.copy()
    env.update(
        {
            "WHITEPAPER_APP_DIR": str(APP_DIR),
            "WHITEPAPER_GIT_ROOT": str(root),
            "WHITEPAPER_DB_PATH": str(DB_PATH),
            "WHITEPAPER_UPLOAD_DIR": str(UPLOAD_ROOT),
            "WHITEPAPER_FRONTEND_DIR": str(APP_DIR / "frontend"),
            "WHITEPAPER_BACKEND_DIR": str(BACKEND_DIR),
            "WHITEPAPER_UPDATE_LOG_DIR": str(UPDATE_LOG_DIR),
            "WHITEPAPER_UPDATE_ALLOW_DIRTY": "1" if payload.allow_dirty else "0",
        }
    )
    started_at = now_iso()
    try:
        result = subprocess.run(
            ["/bin/zsh", str(UPDATE_SCRIPT_PATH)],
            cwd=str(root),
            env=env,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=600,
            check=False,
        )
    except subprocess.TimeoutExpired as exc:
        result = subprocess.CompletedProcess(
            args=[str(UPDATE_SCRIPT_PATH)],
            returncode=-1,
            stdout=exc.stdout if isinstance(exc.stdout, str) else "",
            stderr="The update script timed out. Review update_logs and confirm the system state manually.",
        )
    except Exception as exc:
        result = subprocess.CompletedProcess(
            args=[str(UPDATE_SCRIPT_PATH)],
            returncode=-1,
            stdout="",
            stderr=str(exc),
        )
    completed_at = now_iso()
    last_update = {
        "started_at": started_at,
        "completed_at": completed_at,
        "ok": result.returncode == 0,
        "returncode": result.returncode,
        "stdout_tail": (result.stdout or "").splitlines()[-80:],
        "stderr_tail": (result.stderr or "").splitlines()[-80:],
    }
    write_runtime_status({"last_update": last_update, "restart_required": result.returncode == 0})
    with db_connection() as conn:
        write_audit_log(
            conn,
            request,
            get_actor(user),
            action_type="system_update_run",
            target_table="system",
            target_record_id="git_update",
            new_value=json_dumps({"ok": result.returncode == 0, "returncode": result.returncode}),
        )
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail="Update failed: " + "\n".join(last_update["stderr_tail"][-6:] or last_update["stdout_tail"][-6:] or ["Unknown error"]),
        )
    if payload.restart_after_update:
        UPDATE_RESTART_MARKER_PATH.parent.mkdir(parents=True, exist_ok=True)
        UPDATE_RESTART_MARKER_PATH.write_text(completed_at, "utf-8")
        background_tasks.add_task(delayed_process_exit, 1.0)
    return {
        "ok": True,
        "restart_scheduled": bool(payload.restart_after_update),
        "status_before": status_before,
        "last_update": last_update,
    }


@app.get("/")
async def serve_frontend_index():
    return serve_frontend("")


@app.get("/{path:path}")
async def serve_frontend_app(path: str):
    if path.startswith(
        (
            "auth",
            "dashboard",
            "fields",
            "master-excel",
            "imports",
            "intake",
            "source-documents",
            "review",
            "master",
            "audit-logs",
            "query",
            "chat",
            "legacy",
            "health",
            "admin",
            "image-files",
            "legacy-files",
            "source-files",
            "assets",
            "openapi.json",
            "docs",
            "redoc",
            "system",
        )
    ):
        raise HTTPException(status_code=404, detail="Not found")
    return serve_frontend(path)
