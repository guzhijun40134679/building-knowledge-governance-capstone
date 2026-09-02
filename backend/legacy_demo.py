import csv
import hashlib
import io
import json
import os
import random
import re
from dataclasses import dataclass
from datetime import date, datetime, time
from difflib import SequenceMatcher
from itertools import zip_longest
from pathlib import Path
from typing import Dict, List, Literal, Optional, Tuple
from urllib.parse import quote

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from openai import AsyncOpenAI
from pydantic import BaseModel, Field
from PyPDF2 import PdfReader

# Environment precedence: .env.local overrides .env so placeholders cannot replace local secrets.
BACKEND_DIR = Path(__file__).resolve().parent
DOTENV_PATH = BACKEND_DIR / ".env"
DOTENV_LOCAL_PATH = BACKEND_DIR / ".env.local"
load_dotenv(dotenv_path=DOTENV_PATH)
# Keep private local configuration authoritative even when .env contains distribution defaults.
load_dotenv(dotenv_path=DOTENV_LOCAL_PATH, override=True)

app = FastAPI(title="Property AI Demo Backend")

DEFAULT_CORS_ALLOW_ORIGINS = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
]


def _parse_cors_allow_origins() -> List[str]:
    raw = (os.getenv("CORS_ALLOW_ORIGINS", "") or "").strip()
    if not raw:
        return DEFAULT_CORS_ALLOW_ORIGINS

    origins = [item.strip() for item in raw.split(",") if item.strip()]
    return origins or DEFAULT_CORS_ALLOW_ORIGINS


# Default to local frontend origins; start.command injects CORS_ALLOW_ORIGINS for LAN mode.
app.add_middleware(
    CORSMiddleware,
    allow_origins=_parse_cors_allow_origins(),
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=False,
)


PDF_LIBRARY_DIR = Path(os.getenv("PDF_LIBRARY_DIR", "pdf_library"))
if not PDF_LIBRARY_DIR.is_absolute():
    PDF_LIBRARY_DIR = (BACKEND_DIR / PDF_LIBRARY_DIR).resolve()

SCRIPT_LIBRARY_DIR = Path(os.getenv("SCRIPT_LIBRARY_DIR", "script_library"))
if not SCRIPT_LIBRARY_DIR.is_absolute():
    SCRIPT_LIBRARY_DIR = (BACKEND_DIR / SCRIPT_LIBRARY_DIR).resolve()

IMAGE_LIBRARY_DIR = Path(os.getenv("IMAGE_LIBRARY_DIR", "image_library"))
if not IMAGE_LIBRARY_DIR.is_absolute():
    IMAGE_LIBRARY_DIR = (BACKEND_DIR / IMAGE_LIBRARY_DIR).resolve()

FRONTEND_DIST_DIR = Path(os.getenv("FRONTEND_DIST_DIR", "../frontend/dist"))
if not FRONTEND_DIST_DIR.is_absolute():
    FRONTEND_DIST_DIR = (BACKEND_DIR / FRONTEND_DIST_DIR).resolve()
FRONTEND_INDEX_PATH = FRONTEND_DIST_DIR / "index.html"
FRONTEND_ASSETS_DIR = FRONTEND_DIST_DIR / "assets"

IMAGE_FILE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".bmp"}
SPREADSHEET_FILE_EXTENSIONS = {".xlsx", ".xlsm", ".xls", ".csv", ".tsv"}
SUPPORTED_DOCUMENT_EXTENSIONS = {".pdf"} | SPREADSHEET_FILE_EXTENSIONS
BUILDING_NAME_HEADER_KEYS = {
    "大楼名称",
    "楼盘名称",
    "公寓名称",
    "building",
    "buildingname",
    "property",
    "propertyname",
}
MAX_ANSWER_TOKENS_RAW = (os.getenv("MAX_ANSWER_TOKENS", "550") or "550").strip()
try:
    MAX_ANSWER_TOKENS = max(200, int(MAX_ANSWER_TOKENS_RAW))
except ValueError:
    MAX_ANSWER_TOKENS = 550

MAX_SPREADSHEET_ROWS_PER_SHEET_RAW = (
    os.getenv("MAX_SPREADSHEET_ROWS_PER_SHEET", "1000") or "1000"
).strip()
MAX_SPREADSHEET_COLS_PER_SHEET_RAW = (
    os.getenv("MAX_SPREADSHEET_COLS_PER_SHEET", "80") or "80"
).strip()
MAX_SPREADSHEET_CELLS_PER_SHEET_RAW = (
    os.getenv("MAX_SPREADSHEET_CELLS_PER_SHEET", "20000") or "20000"
).strip()
try:
    MAX_SPREADSHEET_ROWS_PER_SHEET = max(50, int(MAX_SPREADSHEET_ROWS_PER_SHEET_RAW))
except ValueError:
    MAX_SPREADSHEET_ROWS_PER_SHEET = 1000
try:
    MAX_SPREADSHEET_COLS_PER_SHEET = max(5, int(MAX_SPREADSHEET_COLS_PER_SHEET_RAW))
except ValueError:
    MAX_SPREADSHEET_COLS_PER_SHEET = 80
try:
    MAX_SPREADSHEET_CELLS_PER_SHEET = max(500, int(MAX_SPREADSHEET_CELLS_PER_SHEET_RAW))
except ValueError:
    MAX_SPREADSHEET_CELLS_PER_SHEET = 20000

# The legacy demo has no authentication, so its static image directory is disabled by default.
if (os.getenv("WHITEPAPER_ENABLE_UNAUTH_LEGACY_IMAGE_FILES", "") or "").strip() == "1":
    app.mount(
        "/image-files",
        StaticFiles(directory=str(IMAGE_LIBRARY_DIR), check_dir=False),
        name="image-files",
    )

# Image-reference extraction supports both simple numeric references and compound identifiers.
IMAGE_REF_PATTERNS = [
    re.compile(
        r"(?:图示编号|图号|图片序号|图片编号|图片引用|图片|图|img|image|photo)\s*[:：#\-\s]*([a-z0-9][a-z0-9\-_]{0,40})",
        re.IGNORECASE,
    ),
    re.compile(
        r"\[(?:图示编号|图号|图片引用|图片|img|image|photo)?\s*[:：#\-\s]*([a-z0-9][a-z0-9\-_]{0,40})\]",
        re.IGNORECASE,
    ),
    re.compile(r"第\s*0*([1-9]\d{0,3})\s*(?:张)?图", re.IGNORECASE),
]

# General reference tokens include values such as J02-F8-07, J03-10, and AB12-C3.
# Avoid a word boundary because it fails when an identifier directly follows Chinese text.
IMAGE_CODE_PATTERN = re.compile(
    r"(?<![A-Za-z0-9])([a-z][a-z0-9]*(?:[-_][a-z0-9]+){1,8})(?![A-Za-z0-9])",
    re.IGNORECASE,
)


@dataclass
class BuildingDoc:
    # Shared representation for frontend listings, matching, and document-grounded chat.
    building_id: str
    building_name: str
    address: str
    category: str
    file_name: str
    file_path: Path
    relative_path: str
    source_type: str
    content: str


@dataclass
class InternetRecord:
    internet_id: str
    building_name: str
    address: str
    website: str
    contact: str
    wifi_mode: str
    mode_detail: str
    notes: str
    providers: Dict[str, dict]
    source_file: str
    source_sheets: List[str]
    relative_path: str


# Reference content is held in memory and rebuilt from the library when the service restarts.
BUILDING_STORE: Dict[str, BuildingDoc] = {}
INTERNET_STORE: Dict[str, InternetRecord] = {}
INTERNET_WARNINGS: List[str] = []
INTERNET_ERRORS: List[str] = []

# Script library: intent key to one or more deterministic message variants.
SCRIPT_STORE: Dict[str, List[str]] = {}

# Image index: building_id -> reference -> image list.
IMAGE_STORE: Dict[str, Dict[str, List[dict]]] = {}

# Alias index: building_id -> alternate reference -> canonical reference list.
IMAGE_REF_ALIAS_STORE: Dict[str, Dict[str, List[str]]] = {}

# Semantic image hints: building_id -> reference -> keyword list.
IMAGE_HINT_STORE: Dict[str, Dict[str, List[str]]] = {}

IMAGE_HINT_STOPWORDS = {
    "图",
    "图号",
    "图片",
    "image",
    "img",
    "photo",
    "看图",
    "示意图",
}

# Allow low-confidence image matching only when the user clearly asks to see an image.
VISUAL_INTENT_HINTS = {
    "图",
    "图片",
    "照片",
    "示意图",
    "截图",
    "看图",
    "配图",
    "在哪",
    "在哪里",
    "位置",
    "路线",
    "入口",
    "去哪里",
    "怎么走",
    "长什么样",
    "show me",
    "image",
    "photo",
    "picture",
    "where",
}

# Weak semantic tokens must not trigger an image by themselves.
IMAGE_WEAK_MATCH_TOKENS = {
    "位置",
    "哪里",
    "在哪",
    "在哪里",
    "怎么",
    "如何",
    "流程",
    "办理",
    "服务",
    "咨询",
    "联系",
    "是否",
    "可以",
    "时间",
    "tower",
    "room",
    "study",
}

EMAIL_PATTERN = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
PHONE_PATTERN = re.compile(
    r"(?:\+?1[\s\-\.]?)?(?:\(?\d{3}\)?[\s\-\.]?\d{3}[\s\-\.]?\d{4})"
)

ADDRESS_ABBR_PATTERNS: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"\bstreet\b|\bst\.?\b", re.IGNORECASE), " st "),
    (re.compile(r"\broad\b|\brd\.?\b", re.IGNORECASE), " rd "),
    (re.compile(r"\bavenue\b|\bave\.?\b", re.IGNORECASE), " ave "),
    (re.compile(r"\bdrive\b|\bdr\.?\b", re.IGNORECASE), " dr "),
    (re.compile(r"\blane\b|\bln\.?\b", re.IGNORECASE), " ln "),
    (re.compile(r"\bboulevard\b|\bblvd\.?\b", re.IGNORECASE), " blvd "),
    (re.compile(r"\bplace\b|\bpl\.?\b", re.IGNORECASE), " pl "),
    (re.compile(r"\bterrace\b|\btce\.?\b", re.IGNORECASE), " tce "),
    (re.compile(r"\bapartment\b|\bapt\.?\b", re.IGNORECASE), " apt "),
]

QUERY_STOPWORDS = {
    "我",
    "是",
    "在",
    "的",
    "想",
    "咨询",
    "请问",
    "关于",
    "这个",
    "那栋",
    "大楼",
    "楼盘",
    "地址",
    "building",
    "address",
    "about",
    "for",
    "the",
}

CATEGORY_HINTS = {
    "building": {"大楼", "楼盘", "公寓", "租房", "搬家", "停车", "车位", "物业", "门禁", "维修", "报修", "maintenance"},
    "internet": {"网络", "开网", "宽带", "wifi", "broadband", "internet", "nbn"},
    "sim": {"手机卡", "sim", "esim", "号码", "电话卡", "流量", "套餐"},
    "electric": {"开电", "电", "electric", "electricity", "coned", "con edison", "电费", "电表"},
    "electricity": {"开电", "电", "electric", "electricity", "coned", "con edison", "电费", "电表"},
}

DIRECT_CONTACT_CATEGORY_MESSAGES = {
    "electric": "Please contact the HY NY specialist in the customer-service group for help setting up electricity.",
    "electricity": "Please contact the HY NY specialist in the customer-service group for help setting up electricity.",
    "sim": "Please contact the HY NY specialist in the customer-service group for SIM-card assistance.",
    "internet": "Please contact the HY NY specialist in the customer-service group for internet or Wi-Fi assistance.",
}

INTERNET_PRIMARY_SHEET = "Final 结果"
INTERNET_REVIEW_SHEET = "复查网络2.26"
INTERNET_PROVIDER_LABELS = {
    "astound": "Astound",
    "verizon": "Verizon",
    "spectrum": "Spectrum",
    "xfinity": "Xfinity",
}
INTERNET_PROVIDER_ORDER = ["spectrum", "verizon", "astound", "xfinity"]
INTERNET_PENDING_MARKERS = {
    "#name?",
    "#value!",
    "n/a",
    "na",
    "none",
    "null",
    "无法显示",
}
INTERNET_SHEET_COLUMNS = {
    INTERNET_PRIMARY_SHEET: {
        "building_name": 2,
        "website": 4,
        "contact": 5,
        "other_notes": 6,
        "wifi_mode": 12,
        "mode_detail": 14,
        "astound": 15,
        "verizon": 16,
        "spectrum": 17,
        "xfinity": 18,
        "address": 20,
    },
    INTERNET_REVIEW_SHEET: {
        "address": 0,
        "building_name": 2,
        "astound_status": 4,
        "astound_note": 5,
        "verizon_status": 7,
        "spectrum_status": 8,
        "xfinity_status": 9,
        "verizon_plan": 11,
        "verizon_note": 12,
        "spectrum_plan": 13,
        "spectrum_note": 14,
        "xfinity_plan": 15,
    },
}


class ChatMessage(BaseModel):
    role: Literal["user", "assistant", "system"]
    content: str = Field(min_length=1)


class ChatRequest(BaseModel):
    building_id: str = Field(min_length=1)
    question: str = Field(min_length=1)
    history: List[ChatMessage] = Field(default_factory=list)


class ResolveBuildingRequest(BaseModel):
    query: str = Field(min_length=1)
    current_building_id: Optional[str] = None


class ResolveInternetRequest(BaseModel):
    query: str = Field(min_length=1)


class ResolveImagesRequest(BaseModel):
    building_id: str = Field(min_length=1)
    references: List[str] = Field(default_factory=list)


class ResolveImagesFromTextRequest(BaseModel):
    building_id: str = Field(min_length=1)
    text: str = Field(min_length=1)


def _extract_pdf_text(file_bytes: bytes) -> str:
    # Extract each page and normalize spacing, including split Chinese characters.
    pdf = PdfReader(io.BytesIO(file_bytes))
    content_list: List[str] = []
    for page in pdf.pages:
        page_text = page.extract_text() or ""
        normalized_page_text = _normalize_extracted_pdf_text(page_text)
        if normalized_page_text:
            content_list.append(normalized_page_text)
    return "\n".join(content_list).strip()


def _normalize_extracted_pdf_text(text: str) -> str:
    """
    PyPDF2 sometimes inserts spaces or line breaks between Chinese characters.
    Collapse that whitespace to improve retrieval and question matching.
    """
    normalized = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    normalized = re.sub(r"(?<=[\u4e00-\u9fff])[\s\u00a0]+(?=[\u4e00-\u9fff])", "", normalized)
    normalized = re.sub(r"[ \t]+", " ", normalized)
    normalized = re.sub(r"\n{3,}", "\n\n", normalized)
    return normalized.strip()


def _source_type_from_path(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return "pdf"
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return "spreadsheet"
    if suffix in {".csv", ".tsv"}:
        return "table"
    return "document"


def _format_spreadsheet_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, float):
        return f"{value:g}"
    text = str(value)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"\s*\n\s*", " / ", text)
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def _trim_trailing_empty_cells(cells: List[str]) -> List[str]:
    trimmed = list(cells)
    while trimmed and not trimmed[-1]:
        trimmed.pop()
    return trimmed


def _trim_trailing_empty_values(values: List[object]) -> List[object]:
    trimmed = list(values)
    while trimmed and not _format_spreadsheet_value(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _normalize_table_header(value: object) -> str:
    text = _format_spreadsheet_value(value).lower()
    return re.sub(r"[\s\-_./\\()（）:：;；&|]+", "", text)


def _find_building_name_col(headers: List[object]) -> Optional[int]:
    for index, header in enumerate(headers):
        if _normalize_table_header(header) in BUILDING_NAME_HEADER_KEYS:
            return index
    return None


def _is_template_building_name(value: str) -> bool:
    normalized = (value or "").strip().lower()
    return any(marker in normalized for marker in ["模版", "模板", "template"])


def _is_blank_table_row(row: List[object]) -> bool:
    return not any(_format_spreadsheet_value(value) for value in row)


def _format_spreadsheet_row_content(
    source_name: str,
    sheet_name: str,
    row_number: int,
    headers: List[object],
    values: List[object],
    building_name_col: int,
) -> str:
    lines = [
        "[Single-Building Spreadsheet Record]",
        f"Source file: {source_name}",
        f"Worksheet: {sheet_name}",
        f"Original row: {row_number}",
    ]

    field_lines: List[str] = []
    missing_headers: List[str] = []
    max_len = max(len(headers), len(values))
    for index in range(max_len):
        if index == building_name_col:
            continue
        header = _format_spreadsheet_value(headers[index] if index < len(headers) else "")
        value = _format_spreadsheet_value(values[index] if index < len(values) else "")
        if not header:
            continue
        if value:
            field_lines.append(f"{header}: {value}")
        else:
            missing_headers.append(header)

    building_name = _format_spreadsheet_value(
        values[building_name_col] if building_name_col < len(values) else ""
    )
    lines.append(f"Building name: {building_name}")

    if field_lines:
        lines.append("[Field Information]")
        lines.extend(field_lines)
    else:
        lines.append("[Field Information] The spreadsheet provides only the building name; all other fields are blank.")

    if missing_headers:
        lines.append("[Blank Fields] " + ", ".join(missing_headers))

    lines.append(
        "[Answering Rule] If a value is pending confirmation, blank, or not explicitly published by the official source, say that the current spreadsheet does not confirm it; do not infer a value."
    )
    return "\n".join(lines).strip()


def _format_tabular_rows(sheet_name: str, rows: object) -> str:
    lines = [f"[Worksheet: {sheet_name}]"]
    output_rows = 0
    output_cells = 0
    truncated = False
    truncated_columns = False

    for row_number, raw_row in enumerate(rows, start=1):
        raw_values = list(raw_row or [])
        if len(raw_values) > MAX_SPREADSHEET_COLS_PER_SHEET:
            raw_values = raw_values[:MAX_SPREADSHEET_COLS_PER_SHEET]
            truncated_columns = True

        cells = _trim_trailing_empty_cells([_format_spreadsheet_value(value) for value in raw_values])
        if not cells or not any(cells):
            continue

        output_rows += 1
        output_cells += len(cells)
        lines.append(f"Row {row_number}\t" + "\t".join(cells))

        if (
            output_rows >= MAX_SPREADSHEET_ROWS_PER_SHEET
            or output_cells >= MAX_SPREADSHEET_CELLS_PER_SHEET
        ):
            truncated = True
            break

    if output_rows == 0:
        return ""

    if truncated or truncated_columns:
        notes = []
        if truncated:
            notes.append("the read limit was reached; later rows were omitted")
        if truncated_columns:
            notes.append("some columns beyond the width limit were omitted")
        lines.append(f"[Read note: {'; '.join(notes)}]")

    return "\n".join(lines).strip()


def _extract_xlsx_text(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("The openpyxl dependency is required to read .xlsx or .xlsm files.") from exc

    workbook_values = load_workbook(file_path, read_only=True, data_only=True)
    workbook_formulas = None
    sections: List[str] = []
    try:
        try:
            workbook_formulas = load_workbook(file_path, read_only=True, data_only=False)
        except Exception:
            workbook_formulas = None

        for sheet in workbook_values.worksheets:
            formula_sheet = workbook_formulas[sheet.title] if workbook_formulas else None
            value_rows = sheet.iter_rows(values_only=True)
            formula_rows = formula_sheet.iter_rows(values_only=True) if formula_sheet else []

            def merged_rows():
                for value_row, formula_row in zip_longest(value_rows, formula_rows, fillvalue=()):
                    merged = []
                    for value, formula in zip_longest(value_row or (), formula_row or (), fillvalue=None):
                        if value is not None:
                            merged.append(value)
                        elif isinstance(formula, str) and formula.startswith("="):
                            merged.append(f"Formula {formula}")
                        else:
                            merged.append(formula)
                    yield merged

            section = _format_tabular_rows(sheet.title, merged_rows())
            if section:
                sections.append(section)
    finally:
        workbook_values.close()
        if workbook_formulas:
            workbook_formulas.close()

    return "\n\n".join(sections).strip()


def _load_xlsx_sheet_rows(file_path: Path) -> List[Tuple[str, List[List[object]]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("The openpyxl dependency is required to read .xlsx or .xlsm files.") from exc

    workbook_values = load_workbook(file_path, read_only=True, data_only=True)
    workbook_formulas = None
    sheets: List[Tuple[str, List[List[object]]]] = []
    try:
        try:
            workbook_formulas = load_workbook(file_path, read_only=True, data_only=False)
        except Exception:
            workbook_formulas = None

        for sheet in workbook_values.worksheets:
            formula_sheet = workbook_formulas[sheet.title] if workbook_formulas else None
            value_rows = sheet.iter_rows(values_only=True)
            formula_rows = formula_sheet.iter_rows(values_only=True) if formula_sheet else []
            merged_sheet_rows: List[List[object]] = []
            for value_row, formula_row in zip_longest(value_rows, formula_rows, fillvalue=()):
                merged = []
                for value, formula in zip_longest(value_row or (), formula_row or (), fillvalue=None):
                    if value is not None:
                        merged.append(value)
                    elif isinstance(formula, str) and formula.startswith("="):
                        merged.append(f"Formula {formula}")
                    else:
                        merged.append(formula)
                merged_sheet_rows.append(_trim_trailing_empty_values(merged))
            sheets.append((sheet.title, merged_sheet_rows))
    finally:
        workbook_values.close()
        if workbook_formulas:
            workbook_formulas.close()

    return sheets


def _extract_xls_text(file_path: Path) -> str:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("The xlrd dependency is required to read .xls files.") from exc

    workbook = xlrd.open_workbook(file_path)
    sections: List[str] = []
    for sheet in workbook.sheets():
        def rows():
            for row_index in range(sheet.nrows):
                values = []
                for col_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, col_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                        except Exception:
                            pass
                    values.append(value)
                yield values

        section = _format_tabular_rows(sheet.name, rows())
        if section:
            sections.append(section)
    return "\n\n".join(sections).strip()


def _load_xls_sheet_rows(file_path: Path) -> List[Tuple[str, List[List[object]]]]:
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("The xlrd dependency is required to read .xls files.") from exc

    workbook = xlrd.open_workbook(file_path)
    sheets: List[Tuple[str, List[List[object]]]] = []
    for sheet in workbook.sheets():
        sheet_rows: List[List[object]] = []
        for row_index in range(sheet.nrows):
            values = []
            for col_index in range(sheet.ncols):
                cell = sheet.cell(row_index, col_index)
                value = cell.value
                if cell.ctype == xlrd.XL_CELL_DATE:
                    try:
                        value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                    except Exception:
                        pass
                values.append(value)
            sheet_rows.append(_trim_trailing_empty_values(values))
        sheets.append((sheet.name, sheet_rows))
    return sheets


def _read_text_file_with_fallback(file_path: Path) -> str:
    for encoding in ["utf-8-sig", "utf-8", "gb18030", "big5", "latin-1"]:
        try:
            return file_path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    return file_path.read_text(encoding="utf-8", errors="replace")


def _extract_delimited_text(file_path: Path) -> str:
    delimiter = "\t" if file_path.suffix.lower() == ".tsv" else ","
    raw_text = _read_text_file_with_fallback(file_path)
    reader = csv.reader(io.StringIO(raw_text), delimiter=delimiter)
    return _format_tabular_rows(Path(file_path.name).stem or "Sheet1", reader)


def _load_delimited_sheet_rows(file_path: Path) -> List[Tuple[str, List[List[object]]]]:
    delimiter = "\t" if file_path.suffix.lower() == ".tsv" else ","
    raw_text = _read_text_file_with_fallback(file_path)
    rows = [_trim_trailing_empty_values(list(row)) for row in csv.reader(io.StringIO(raw_text), delimiter=delimiter)]
    return [(Path(file_path.name).stem or "Sheet1", rows)]


def _load_tabular_sheet_rows(file_path: Path) -> List[Tuple[str, List[List[object]]]]:
    suffix = file_path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return _load_xlsx_sheet_rows(file_path)
    if suffix == ".xls":
        return _load_xls_sheet_rows(file_path)
    if suffix in {".csv", ".tsv"}:
        return _load_delimited_sheet_rows(file_path)
    return []


def _extract_spreadsheet_row_entries(file_path: Path) -> List[dict]:
    if file_path.suffix.lower() not in SPREADSHEET_FILE_EXTENSIONS:
        return []

    entries: List[dict] = []
    try:
        relative_file_path = str(file_path.relative_to(PDF_LIBRARY_DIR))
    except ValueError:
        relative_file_path = file_path.name

    for sheet_name, rows in _load_tabular_sheet_rows(file_path):
        header_index: Optional[int] = None
        building_name_col: Optional[int] = None
        for index, row in enumerate(rows[:10]):
            if _is_blank_table_row(row):
                continue
            found_col = _find_building_name_col(row)
            if found_col is not None:
                header_index = index
                building_name_col = found_col
                break

        if header_index is None or building_name_col is None:
            continue

        headers = rows[header_index]
        for row_index, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
            if _is_blank_table_row(row):
                continue

            building_name = _format_spreadsheet_value(
                row[building_name_col] if building_name_col < len(row) else ""
            )
            if not building_name or _is_template_building_name(building_name):
                continue

            content = _format_spreadsheet_row_content(
                source_name=file_path.name,
                sheet_name=sheet_name,
                row_number=row_index,
                headers=headers,
                values=row,
                building_name_col=building_name_col,
            )
            row_key = re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "-", building_name.lower()).strip("-")
            row_key = row_key[:60] or f"row-{row_index}"
            entries.append(
                {
                    "building_name": building_name,
                    "address": "",
                    "relative_path": f"{relative_file_path}#{sheet_name}!R{row_index}-{row_key}",
                    "source_type": "spreadsheet-row",
                    "content": content,
                }
            )

    return entries


def _clean_internet_text(value: object, allow_link: bool = False) -> str:
    text = _format_spreadsheet_value(value)
    if not text:
        return ""

    lowered = text.strip().lower()
    if lowered in {"-", "--", "---", "----", "-----", "/", "none", "null", "nan"}:
        return ""
    if text in {"#NAME?", "#VALUE!"}:
        return ""
    if not allow_link and lowered == "link":
        return ""
    return text


def _unique_nonempty_texts(items: List[str]) -> List[str]:
    unique: List[str] = []
    seen = set()
    for item in items:
        text = (item or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        unique.append(text)
    return unique


def _internet_row_value(row: List[object], index: int) -> object:
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _internet_row_text(row: List[object], index: int, allow_link: bool = False) -> str:
    return _clean_internet_text(_internet_row_value(row, index), allow_link=allow_link)


def _normalize_internet_record_key(building_name: str, address: str) -> str:
    return _normalize_text(building_name) or _normalize_text(address)


def _normalize_internet_provider_flag(raw_value: object) -> Tuple[Optional[str], str]:
    text = _format_spreadsheet_value(raw_value)
    if not text:
        return None, ""

    normalized = _normalize_text(text)
    lowered = text.strip().lower()
    if normalized == "1":
        return "supported", ""
    if normalized == "0":
        return "unsupported", ""
    if normalized == "2":
        return "pending", "The review result is pending confirmation"
    if lowered in INTERNET_PENDING_MARKERS:
        return "pending", text
    return "pending", text


def _merge_internet_provider_status(final_value: object, review_value: object) -> Tuple[str, str]:
    final_status, final_note = _normalize_internet_provider_flag(final_value)
    review_status, review_note = _normalize_internet_provider_flag(review_value)
    notes = _unique_nonempty_texts([final_note, review_note])

    if final_status and review_status and final_status != review_status:
        notes = _unique_nonempty_texts(notes + ["The master and review worksheets disagree"])
        return "pending", "; ".join(notes)

    if final_status == "pending" or review_status == "pending":
        if not notes:
            notes = ["The official website does not state this clearly"]
        return "pending", "; ".join(notes)

    status = review_status or final_status
    if status:
        return status, ""

    return "pending", "The official website does not state this clearly"


def _parse_internet_primary_sheet(rows: List[List[object]]) -> Dict[str, dict]:
    columns = INTERNET_SHEET_COLUMNS[INTERNET_PRIMARY_SHEET]
    records: Dict[str, dict] = {}

    for row in rows[1:]:
        if _is_blank_table_row(row):
            continue

        building_name = _internet_row_text(row, columns["building_name"])
        if not building_name or _normalize_text(building_name) in {"大楼", "building"}:
            continue

        address = _internet_row_text(row, columns["address"])
        key = _normalize_internet_record_key(building_name, address)
        if not key:
            continue

        records[key] = {
            "building_name": building_name,
            "address": address,
            "website": _internet_row_text(row, columns["website"], allow_link=False),
            "contact": _internet_row_text(row, columns["contact"]),
            "notes": _internet_row_text(row, columns["other_notes"]),
            "wifi_mode": _internet_row_text(row, columns["wifi_mode"]),
            "mode_detail": _internet_row_text(row, columns["mode_detail"]),
            "providers": {
                "astound": _internet_row_value(row, columns["astound"]),
                "verizon": _internet_row_value(row, columns["verizon"]),
                "spectrum": _internet_row_value(row, columns["spectrum"]),
                "xfinity": _internet_row_value(row, columns["xfinity"]),
            },
        }

    return records


def _parse_internet_review_sheet(rows: List[List[object]]) -> Dict[str, dict]:
    columns = INTERNET_SHEET_COLUMNS[INTERNET_REVIEW_SHEET]
    records: Dict[str, dict] = {}

    for row in rows[1:]:
        if _is_blank_table_row(row):
            continue

        building_name = _internet_row_text(row, columns["building_name"])
        address = _internet_row_text(row, columns["address"])
        if not building_name and not address:
            continue

        key = _normalize_internet_record_key(building_name, address)
        if not key:
            continue

        records[key] = {
            "building_name": building_name,
            "address": address,
            "statuses": {
                "astound": _internet_row_value(row, columns["astound_status"]),
                "verizon": _internet_row_value(row, columns["verizon_status"]),
                "spectrum": _internet_row_value(row, columns["spectrum_status"]),
                "xfinity": _internet_row_value(row, columns["xfinity_status"]),
            },
            "plans": {
                "verizon": _internet_row_text(row, columns["verizon_plan"]),
                "spectrum": _internet_row_text(row, columns["spectrum_plan"]),
                "xfinity": _internet_row_text(row, columns["xfinity_plan"]),
            },
            "notes": {
                "astound": _internet_row_text(row, columns["astound_note"]),
                "verizon": _internet_row_text(row, columns["verizon_note"]),
                "spectrum": _internet_row_text(row, columns["spectrum_note"]),
                "xfinity": "",
            },
        }

    return records


def _build_internet_provider_payload(
    provider_key: str,
    final_value: object,
    review_value: object,
    review_plan: str,
    review_note: str,
) -> dict:
    status, status_note = _merge_internet_provider_status(final_value, review_value)
    notes = _unique_nonempty_texts([review_note, status_note])
    return {
        "key": provider_key,
        "label": INTERNET_PROVIDER_LABELS[provider_key],
        "status": status,
        "plans": review_plan or "",
        "note": "；".join(notes),
    }


def _build_internet_record(
    relative_path: str,
    file_name: str,
    record_key: str,
    primary: Optional[dict],
    review: Optional[dict],
) -> InternetRecord:
    building_name = ""
    address = ""
    if primary:
        building_name = primary.get("building_name", "")
        address = primary.get("address", "")
    if review:
        building_name = building_name or review.get("building_name", "")
        address = address or review.get("address", "")

    providers: Dict[str, dict] = {}
    for provider_key in INTERNET_PROVIDER_ORDER:
        final_value = primary.get("providers", {}).get(provider_key) if primary else None
        review_value = review.get("statuses", {}).get(provider_key) if review else None
        review_plan = review.get("plans", {}).get(provider_key, "") if review else ""
        review_note = review.get("notes", {}).get(provider_key, "") if review else ""
        providers[provider_key] = _build_internet_provider_payload(
            provider_key=provider_key,
            final_value=final_value,
            review_value=review_value,
            review_plan=review_plan,
            review_note=review_note,
        )

    notes = _unique_nonempty_texts(
        [
            primary.get("notes", "") if primary else "",
        ]
    )
    source_sheets = []
    if primary:
        source_sheets.append(INTERNET_PRIMARY_SHEET)
    if review:
        source_sheets.append(INTERNET_REVIEW_SHEET)

    unique_key = f"{relative_path}#{record_key}"
    internet_id = hashlib.sha1(unique_key.encode("utf-8")).hexdigest()[:12]
    return InternetRecord(
        internet_id=internet_id,
        building_name=building_name,
        address=address,
        website=primary.get("website", "") if primary else "",
        contact=primary.get("contact", "") if primary else "",
        wifi_mode=primary.get("wifi_mode", "") if primary else "",
        mode_detail=primary.get("mode_detail", "") if primary else "",
        notes="；".join(notes),
        providers=providers,
        source_file=file_name,
        source_sheets=source_sheets,
        relative_path=relative_path,
    )


def _iter_internet_document_paths() -> List[Path]:
    internet_dir = PDF_LIBRARY_DIR / "internet"
    if not internet_dir.exists():
        return []

    paths: List[Path] = []
    for file_path in internet_dir.rglob("*"):
        if not file_path.is_file():
            continue
        if file_path.name.startswith("~$"):
            continue
        if file_path.suffix.lower() in SPREADSHEET_FILE_EXTENSIONS:
            paths.append(file_path)
    return sorted(paths)


def reload_internet_store() -> dict:
    internet_dir = PDF_LIBRARY_DIR / "internet"
    internet_dir.mkdir(parents=True, exist_ok=True)

    new_store: Dict[str, InternetRecord] = {}
    warnings: List[str] = []
    errors: List[str] = []

    for doc_path in _iter_internet_document_paths():
        try:
            relative_path = str(doc_path.relative_to(PDF_LIBRARY_DIR))
            sheet_rows = {sheet_name: rows for sheet_name, rows in _load_tabular_sheet_rows(doc_path)}
            primary_rows = sheet_rows.get(INTERNET_PRIMARY_SHEET)
            review_rows = sheet_rows.get(INTERNET_REVIEW_SHEET)

            if primary_rows is None:
                errors.append(f"{relative_path}: missing worksheet {INTERNET_PRIMARY_SHEET}.")
            if review_rows is None:
                errors.append(f"{relative_path}: missing worksheet {INTERNET_REVIEW_SHEET}.")

            primary_records = _parse_internet_primary_sheet(primary_rows or [])
            review_records = _parse_internet_review_sheet(review_rows or [])

            for record_key in sorted(set(primary_records) | set(review_records)):
                record = _build_internet_record(
                    relative_path=relative_path,
                    file_name=doc_path.name,
                    record_key=record_key,
                    primary=primary_records.get(record_key),
                    review=review_records.get(record_key),
                )
                if not record.building_name and not record.address:
                    continue
                if record.internet_id in new_store:
                    warnings.append(f"{relative_path}: duplicate internet record {record.building_name or record.address}; the later row replaced it.")
                new_store[record.internet_id] = record
        except Exception as exc:
            errors.append(f"{doc_path.relative_to(PDF_LIBRARY_DIR)}: failed to parse the internet worksheet ({exc}).")

    INTERNET_STORE.clear()
    INTERNET_STORE.update(new_store)
    INTERNET_WARNINGS.clear()
    INTERNET_WARNINGS.extend(warnings)
    INTERNET_ERRORS.clear()
    INTERNET_ERRORS.extend(errors)

    return {
        "base_dir": str(internet_dir),
        "total": len(INTERNET_STORE),
        "files": len(_iter_internet_document_paths()),
        "warnings": warnings,
        "errors": errors,
    }


def _extract_document_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return _extract_pdf_text(file_path.read_bytes())
    if suffix in {".xlsx", ".xlsm"}:
        return _extract_xlsx_text(file_path)
    if suffix == ".xls":
        return _extract_xls_text(file_path)
    if suffix in {".csv", ".tsv"}:
        return _extract_delimited_text(file_path)
    return ""


def _parse_filename_to_building(file_name: str) -> Tuple[str, str]:
    """
    Parse filenames in the form building-name-address.pdf/xlsx/csv.
    Without a separator, treat the entire stem as the building name.
    """
    stem = Path(file_name).stem.strip()
    if not stem:
        return "Unnamed Building", ""

    parts = re.split(r"\s*[-—–－_]\s*", stem, maxsplit=1)
    building_name = parts[0].strip() if parts else stem
    address = parts[1].strip() if len(parts) > 1 else ""
    return building_name or stem, address


def _category_from_doc_path(doc_path: Path) -> str:
    """
    Classification rules:
    - pdf_library/xxx.pdf/xlsx/csv -> building (default)
    - pdf_library/internet/xxx.pdf/xlsx/csv -> internet
    - pdf_library/sim/xxx.pdf/xlsx/csv -> sim
    """
    relative = doc_path.relative_to(PDF_LIBRARY_DIR)
    if len(relative.parts) > 1:
        return relative.parts[0].strip().lower() or "building"
    return "building"


def _doc_label(doc: BuildingDoc) -> str:
    if doc.address:
        return f"{doc.building_name}（{doc.address}）"
    return doc.building_name


def _normalize_intent_key(value: str) -> str:
    return re.sub(r"[^a-z0-9_]+", "_", (value or "").strip().lower()).strip("_")


def _load_script_variants_from_root_file(file_path: Path) -> List[str]:
    """
    Parse root files such as initial_topic_prompt.txt as one message per line.
    """
    content = file_path.read_text(encoding="utf-8", errors="ignore")
    variants: List[str] = []
    for raw_line in content.splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        variants.append(line)
    return variants


def _load_script_variants_from_dir_file(file_path: Path) -> List[str]:
    """
    Parse each file below an intent directory as one complete message.
    """
    content = file_path.read_text(encoding="utf-8", errors="ignore").strip()
    if not content:
        return []
    return [content]


def reload_script_store() -> dict:
    # Support both line-oriented root files and one-message-per-file subdirectories.
    SCRIPT_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)

    new_store: Dict[str, List[str]] = {}
    errors: List[str] = []

    # 1) script_library/*.txt -> multiple messages, one per line.
    for txt_path in sorted(SCRIPT_LIBRARY_DIR.glob("*.txt")):
        try:
            intent = _normalize_intent_key(txt_path.stem)
            if not intent:
                continue
            variants = _load_script_variants_from_root_file(txt_path)
            if variants:
                new_store.setdefault(intent, []).extend(variants)
        except Exception as exc:
            errors.append(f"{txt_path.name}: failed to read ({exc}).")

    # 2) script_library/<intent>/*.txt -> one message per file.
    for intent_dir in sorted(SCRIPT_LIBRARY_DIR.iterdir()):
        if not intent_dir.is_dir():
            continue
        intent = _normalize_intent_key(intent_dir.name)
        if not intent:
            continue
        for txt_path in sorted(intent_dir.glob("*.txt")):
            try:
                variants = _load_script_variants_from_dir_file(txt_path)
                if variants:
                    new_store.setdefault(intent, []).extend(variants)
            except Exception as exc:
                relative = txt_path.relative_to(SCRIPT_LIBRARY_DIR)
                errors.append(f"{relative}: failed to read ({exc}).")

    # Deduplicate while preserving order.
    deduped_store: Dict[str, List[str]] = {}
    for intent, variants in new_store.items():
        seen = set()
        deduped: List[str] = []
        for item in variants:
            if item not in seen:
                seen.add(item)
                deduped.append(item)
        if deduped:
            deduped_store[intent] = deduped

    SCRIPT_STORE.clear()
    SCRIPT_STORE.update(deduped_store)

    return {
        "base_dir": str(SCRIPT_LIBRARY_DIR),
        "total_intents": len(SCRIPT_STORE),
        "total_variants": sum(len(items) for items in SCRIPT_STORE.values()),
        "errors": errors,
    }


def pick_script(intent: str, fallback: str = "") -> str:
    normalized = _normalize_intent_key(intent)
    candidates = SCRIPT_STORE.get(normalized, [])
    if not candidates:
        return fallback
    return random.choice(candidates)


def _to_posix_path(path: Path) -> str:
    return str(path).replace("\\", "/")


def _build_image_url(relative_path: Path) -> str:
    return f"/image-files/{quote(_to_posix_path(relative_path), safe='/')}"


def _normalize_image_reference(value: str) -> Optional[str]:
    raw = (value or "").strip()
    if not raw:
        return None

    cleaned = (
        raw.replace("—", "-")
        .replace("–", "-")
        .replace("－", "-")
        .replace("_", "-")
        .replace("/", "-")
    )
    cleaned = re.sub(r"[^a-zA-Z0-9\-]+", "", cleaned)
    cleaned = re.sub(r"-{2,}", "-", cleaned).strip("-")
    if not cleaned:
        return None

    if re.fullmatch(r"0*[1-9]\d{0,3}", cleaned):
        return str(int(cleaned))

    parts = [part for part in cleaned.split("-") if part]
    if not parts:
        return None

    normalized_parts: List[str] = []
    for part in parts:
        letters_digits = re.fullmatch(r"([a-zA-Z]+)(\d+)", part)
        if letters_digits:
            normalized_parts.append(
                f"{letters_digits.group(1).upper()}{int(letters_digits.group(2))}"
            )
            continue

        if re.fullmatch(r"\d+", part):
            normalized_parts.append(str(int(part)))
            continue

        if re.fullmatch(r"[a-zA-Z]+", part):
            normalized_parts.append(part.upper())
            continue

        normalized_parts.append(part.upper())

    normalized = "-".join(normalized_parts)
    if not re.search(r"\d", normalized):
        return None
    return normalized


def _get_primary_image_reference(reference: str) -> str:
    """
    For a filename such as J02-F8-07-1, the last segment usually identifies
    one of several images. Group under J02-F8-07 and retain the full identifier as an alias.
    """
    parts = reference.split("-")
    if len(parts) >= 4 and re.fullmatch(r"\d+", parts[-1]):
        return "-".join(parts[:-1])
    if (
        len(parts) == 3
        and re.fullmatch(r"\d+", parts[-1])
        and re.fullmatch(r"\d+", parts[1])
    ):
        return "-".join(parts[:-1])
    return reference


def _extract_image_reference_candidates_from_file_name(file_name: str) -> List[str]:
    """
    Extract candidate image references from compatible filename patterns such as:
    1.jpg / 01_frontdesk.png / j03-10-1.jpg / a Chinese-prefix followed by j02-f8-02-2.jpg
    """
    stem = Path(file_name).stem
    candidates: List[str] = []
    seen = set()

    for match in IMAGE_CODE_PATTERN.finditer(stem):
        normalized = _normalize_image_reference(match.group(1))
        if normalized and normalized not in seen:
            seen.add(normalized)
            candidates.append(normalized)

    if candidates:
        return candidates

    numeric = re.search(r"[1-9]\d{0,3}", stem)
    if numeric:
        normalized = _normalize_image_reference(numeric.group(0))
        if normalized:
            candidates.append(normalized)
    return candidates


def _extract_image_references_from_text(text: str) -> List[str]:
    references: List[str] = []
    seen = set()

    for pattern in IMAGE_REF_PATTERNS:
        for match in pattern.finditer(text or ""):
            ref = _normalize_image_reference(match.group(1))
            if ref and ref not in seen:
                seen.add(ref)
                references.append(ref)

    # Also recognize compound identifiers such as J02-F8-07 or J03-10 in generated text.
    for match in IMAGE_CODE_PATTERN.finditer(text or ""):
        ref = _normalize_image_reference(match.group(1))
        if ref and ref not in seen:
            seen.add(ref)
            references.append(ref)
    return references


def _image_reference_sort_key(reference: str) -> Tuple:
    parts = re.findall(r"[A-Za-z]+|\d+", reference or "")
    key = []
    for part in parts:
        if part.isdigit():
            key.append((1, int(part)))
        else:
            key.append((0, part.upper()))
    return tuple(key)


def _normalize_name_key(text: str) -> str:
    return re.sub(r"[^a-z0-9\u4e00-\u9fff]+", "", (text or "").lower())


def _find_fuzzy_building_dirs(root: Path, building_name: str) -> List[Path]:
    if not root.exists() or not root.is_dir():
        return []

    building_key = _normalize_name_key(building_name)
    if not building_key:
        return []

    matched: List[Path] = []
    for child in sorted(root.iterdir()):
        if not child.is_dir():
            continue
        child_key = _normalize_name_key(child.name)
        if not child_key:
            continue
        if building_key in child_key or child_key in building_key:
            matched.append(child)
    return matched


def _candidate_image_dirs_for_doc(doc: BuildingDoc) -> List[Path]:
    """
    Support multiple storage layouts:
    1) image_library/<category>/<pdf-stem>/
    2) image_library/<pdf-stem>/
    3) image_library/<building_id>/
    4) image_library/<category>/<building-related-folder>/
    5) image_library/<building-related-folder>/
    """
    by_relative_path = IMAGE_LIBRARY_DIR / Path(doc.relative_path).with_suffix("")
    by_file_stem = IMAGE_LIBRARY_DIR / Path(doc.file_name).stem
    by_building_id = IMAGE_LIBRARY_DIR / doc.building_id
    by_category_building_name = IMAGE_LIBRARY_DIR / doc.category / doc.building_name
    by_building_name = IMAGE_LIBRARY_DIR / doc.building_name

    candidates = [
        by_relative_path,
        by_file_stem,
        by_building_id,
        by_category_building_name,
        by_building_name,
    ]
    candidates.extend(_find_fuzzy_building_dirs(IMAGE_LIBRARY_DIR / doc.category, doc.building_name))
    candidates.extend(_find_fuzzy_building_dirs(IMAGE_LIBRARY_DIR, doc.building_name))

    unique: List[Path] = []
    seen = set()
    for item in candidates:
        key = str(item.resolve()) if item.exists() else str(item)
        if key in seen:
            continue
        seen.add(key)
        unique.append(item)
    return unique


def _build_image_hint_store_for_doc(doc: BuildingDoc, available_refs: set) -> Dict[str, List[str]]:
    """
    Extract contextual keywords from source lines that contain an image reference,
    associating nearby terms with that reference.
    """
    if not available_refs:
        return {}

    lines = [line.strip() for line in re.split(r"\n+", doc.content or "") if line.strip()]
    if not lines:
        return {}

    ref_tokens: Dict[str, set] = {}
    for idx, line in enumerate(lines):
        refs = _extract_image_references_from_text(line)
        refs = [ref for ref in refs if ref in available_refs]
        if not refs:
            continue

        # Semantic context often appears on following lines, so use a wider window.
        start = max(0, idx - 4)
        end = min(len(lines), idx + 8)
        context = " ".join(lines[start:end])
        tokens = {
            token
            for token in _tokenize_for_image(context)
            if token and token not in IMAGE_HINT_STOPWORDS
        }
        if not tokens:
            continue

        for ref in refs:
            ref_tokens.setdefault(ref, set()).update(tokens)

    return {ref: sorted(tokens) for ref, tokens in ref_tokens.items() if tokens}


def _tokenize_for_image(text: str) -> List[str]:
    """
    Image-matching tokenizer: preserve original tokens and split continuous Chinese text
    into two-to-four-character n-grams for phrase matching.
    """
    base_tokens = [token for token in _tokenize_text(text) if token]
    extras: List[str] = []
    seen = set(base_tokens)

    for chunk in re.findall(r"[\u4e00-\u9fff]{2,}", _canonicalize_text(text)):
        for n in (2, 3, 4):
            if len(chunk) < n:
                continue
            for idx in range(0, len(chunk) - n + 1):
                token = chunk[idx : idx + n]
                if token in IMAGE_HINT_STOPWORDS:
                    continue
                if token in seen:
                    continue
                seen.add(token)
                extras.append(token)

    return base_tokens + extras


def _has_visual_intent(text: str) -> bool:
    canonical = _canonicalize_text(text)
    normalized = _normalize_text(text)
    if any(hint in canonical for hint in VISUAL_INTENT_HINTS):
        return True
    # Preserve short Chinese image requests that tokenization can otherwise miss.
    return "有图" in normalized or "看图" in normalized


def _infer_image_references_from_query(building_id: str, query_text: str) -> List[str]:
    """
    When no explicit reference is present, match query keywords to reference context.
    """
    hint_map = IMAGE_HINT_STORE.get(building_id, {})
    if not hint_map:
        return []

    query_tokens = {
        token
        for token in _tokenize_for_image(query_text or "")
        if token and token not in IMAGE_HINT_STOPWORDS
    }
    if not query_tokens:
        return []

    scored: List[Tuple[int, int, str, set]] = []
    has_visual_intent = _has_visual_intent(query_text or "")
    for ref, tokens in hint_map.items():
        token_set = set(tokens)
        overlap = query_tokens & token_set
        if not overlap:
            continue

        overlap_count = len(overlap)
        only_token = next(iter(overlap)) if overlap_count == 1 else ""
        # Without image intent, a single token must be semantically strong.
        if not has_visual_intent and overlap_count == 1 and only_token in IMAGE_WEAK_MATCH_TOKENS:
            continue

        score = overlap_count * 35 + int(100 * overlap_count / max(len(query_tokens), 1))
        if has_visual_intent:
            score += 20
        elif overlap_count == 1:
            # Downweight one-token matches unless the candidate has a clear lead.
            score -= 10
        scored.append((score, overlap_count, ref, overlap))

    if not scored:
        return []

    scored.sort(key=lambda item: (-item[0], -item[1], _image_reference_sort_key(item[2])))

    # Reject ambiguous one-token matches with several tied candidates.
    if scored[0][1] == 1:
        top_score = scored[0][0]
        near_top = [item for item in scored if item[0] >= top_score - 8]
        ambiguity_limit = 3 if has_visual_intent else 2
        if len(near_top) >= ambiguity_limit:
            return []

        # Treat near-ties on the same weak token as ambiguous.
        token_groups: Dict[str, int] = {}
        for _, _, _, overlap in near_top:
            token = next(iter(overlap)) if overlap else ""
            if token:
                token_groups[token] = token_groups.get(token, 0) + 1
        if token_groups and max(token_groups.values()) >= 2 and len(near_top) >= 2:
            return []

    top_score = scored[0][0]
    top_overlap_count = scored[0][1]
    if has_visual_intent:
        threshold = max(45, int(top_score * 0.72))
        max_results = 2
    elif top_overlap_count >= 2:
        threshold = max(78, int(top_score * 0.72))
        max_results = 2
    else:
        # Without image intent, return one image group only for a high-confidence lead.
        threshold = max(52, int(top_score * 0.80))
        max_results = 1
    return [ref for score, _, ref, _ in scored if score >= threshold][:max_results]


def _resolve_images_for_doc(building_id: str, references: List[str]) -> List[dict]:
    image_map = IMAGE_STORE.get(building_id, {})
    alias_map = IMAGE_REF_ALIAS_STORE.get(building_id, {})
    if not image_map:
        return []

    resolved: List[dict] = []
    seen = set()
    for raw_ref in references:
        ref = _normalize_image_reference(raw_ref)
        if not ref:
            continue
        candidate_refs = [ref]
        if ref in alias_map:
            candidate_refs = alias_map[ref]

        for candidate_ref in candidate_refs:
            for item in image_map.get(candidate_ref, []):
                key = item.get("relative_path", "")
                if key and key not in seen:
                    seen.add(key)
                    resolved.append(item)
    return resolved


def reload_image_store() -> dict:
    IMAGE_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)

    new_store: Dict[str, Dict[str, List[dict]]] = {}
    new_alias_store: Dict[str, Dict[str, List[str]]] = {}
    new_hint_store: Dict[str, Dict[str, List[str]]] = {}
    errors: List[str] = []
    total_images = 0

    for doc in BUILDING_STORE.values():
        image_map: Dict[str, List[dict]] = {}
        alias_map: Dict[str, set] = {}
        scanned_dirs = set()
        for image_dir in _candidate_image_dirs_for_doc(doc):
            if not image_dir.exists() or not image_dir.is_dir():
                continue
            dir_key = str(image_dir.resolve())
            if dir_key in scanned_dirs:
                continue
            scanned_dirs.add(dir_key)

            for file_path in sorted(image_dir.rglob("*")):
                if not file_path.is_file():
                    continue
                if file_path.suffix.lower() not in IMAGE_FILE_EXTENSIONS:
                    continue

                references = _extract_image_reference_candidates_from_file_name(file_path.name)
                if not references:
                    continue

                try:
                    relative = file_path.relative_to(IMAGE_LIBRARY_DIR)
                except Exception as exc:
                    errors.append(f"{file_path}: invalid path ({exc}).")
                    continue

                for reference in references:
                    primary_ref = _get_primary_image_reference(reference)
                    image_map.setdefault(primary_ref, []).append(
                        {
                            "reference": primary_ref,
                            "file_name": file_path.name,
                            "relative_path": _to_posix_path(relative),
                            "url": _build_image_url(relative),
                        }
                    )

                    alias_map.setdefault(reference, set()).add(primary_ref)
                    alias_map.setdefault(primary_ref, set()).add(primary_ref)

        if image_map:
            sorted_map: Dict[str, List[dict]] = {}
            for reference, items in sorted(
                image_map.items(), key=lambda pair: _image_reference_sort_key(pair[0])
            ):
                sorted_map[reference] = sorted(items, key=lambda item: item["file_name"])
                total_images += len(sorted_map[reference])
            new_store[doc.building_id] = sorted_map
            new_alias_store[doc.building_id] = {
                key: sorted(value, key=_image_reference_sort_key)
                for key, value in alias_map.items()
            }
            hint_map = _build_image_hint_store_for_doc(doc, set(sorted_map.keys()))
            if hint_map:
                new_hint_store[doc.building_id] = hint_map

    IMAGE_STORE.clear()
    IMAGE_STORE.update(new_store)
    IMAGE_REF_ALIAS_STORE.clear()
    IMAGE_REF_ALIAS_STORE.update(new_alias_store)
    IMAGE_HINT_STORE.clear()
    IMAGE_HINT_STORE.update(new_hint_store)

    return {
        "base_dir": str(IMAGE_LIBRARY_DIR),
        "docs_with_images": len(IMAGE_STORE),
        "docs_with_image_hints": len(IMAGE_HINT_STORE),
        "total_images": total_images,
        "errors": errors,
    }


def _is_electric_doc(doc: BuildingDoc) -> bool:
    return (doc.category or "").lower() in {"electric", "electricity"}


def _is_direct_contact_doc(doc: BuildingDoc) -> bool:
    return (doc.category or "").lower() in DIRECT_CONTACT_CATEGORY_MESSAGES


def _get_direct_contact_answer(doc: BuildingDoc) -> str:
    return DIRECT_CONTACT_CATEGORY_MESSAGES.get(
        (doc.category or "").lower(),
        "Please contact the HY NY specialist in the customer-service group for assistance with this service.",
    )


def _is_manhattan_electric_doc(doc: BuildingDoc) -> bool:
    if not _is_electric_doc(doc):
        return False

    text = _canonicalize_text(
        f"{doc.building_name} {doc.address} {doc.file_name} {doc.relative_path}"
    )
    manhattan_keywords = {"manhattan", "曼哈顿", "nyc", "newyork", "new york", "hy ny"}
    return any(keyword in text for keyword in manhattan_keywords)


def _extract_contact_candidates(text: str) -> Tuple[List[str], List[str]]:
    emails = []
    phones = []
    email_seen = set()
    phone_seen = set()

    for match in EMAIL_PATTERN.findall(text or ""):
        normalized = match.strip()
        key = normalized.lower()
        if key in email_seen:
            continue
        email_seen.add(key)
        emails.append(normalized)

    for match in PHONE_PATTERN.findall(text or ""):
        normalized = re.sub(r"[^\d]", "", match or "")
        if len(normalized) < 10:
            continue
        normalized = normalized[-10:]
        formatted = f"{normalized[:3]}-{normalized[3:6]}-{normalized[6:]}"
        if formatted in phone_seen:
            continue
        phone_seen.add(formatted)
        phones.append(formatted)

    return emails[:5], phones[:5]


def _format_contact_text(text: str) -> str:
    emails, phones = _extract_contact_candidates(text)
    contacts = []
    if emails:
        contacts.append(f"Email: {', '.join(emails)}")
    if phones:
        contacts.append(f"Phone: {', '.join(phones)}")
    return "; ".join(contacts)


def _build_out_of_scope_instruction(doc: BuildingDoc) -> str:
    contact_text = _format_contact_text(doc.content) or "the source document does not provide contact details"

    category = (doc.category or "").lower()
    if category == "building":
        return (
            "If the question is outside the source document, reply: "
            f"“The source document does not cover this question. Contact building management ({contact_text}). "
            "For additional help, contact the HY NY liaison.”"
        )

    return (
        "If the question is outside the source document, reply: "
        f"“The source document does not cover this question. Contact building management ({contact_text}) "
        "or the HY NY liaison for additional help.”"
    )


def _build_system_prompt(doc: BuildingDoc) -> str:
    doc_label = _doc_label(doc)
    out_of_scope_instruction = _build_out_of_scope_instruction(doc)
    contact_priority_instruction = ""
    contact_text = _format_contact_text(doc.content)
    if contact_text:
        contact_priority_instruction = (
            "When the user asks about repairs, applications, appointments, parking, the front desk, amenities, or a service entry point, "
            f"and the source provides contact details, include those details directly: {contact_text}. "
            "Do not merely say to contact the front desk, management, or maintenance; include the relevant phone number or email. "
            "If several contacts exist, distinguish them by purpose."
        )
    else:
        contact_priority_instruction = (
            "If the source mentions a front desk, office, maintenance contact, portal, or another relevant service channel, "
            "include the specific contact or entry point instead of giving a generic instruction to contact someone."
        )
    maintenance_instruction = ""
    if (doc.category or "").lower() == "building":
        maintenance_instruction = (
            "If the user reports that something in the apartment is broken or needs repair, and the source provides a general maintenance, portal, front-desk, or emergency-maintenance process, "
            "treat the question as a request for the building's repair process and provide the documented channel. "
            "Do not mark it out of scope merely because the exact appliance or fixture is not named. "
            "Use the out-of-scope response only when the source provides no maintenance channel at all. "
            "Do not invent responsibility, fees, response times, or maintenance obligations."
        )

    return f"""
You are a professional resident-services support assistant.
Current reference: {doc_label}.
Business category: {doc.category}.
Answer strictly from the reference content below.
If the user's question is underspecified, ask for the key missing detail first.
{maintenance_instruction}
{contact_priority_instruction}
Keep the answer concise and professional: state the conclusion, then give three to five key points unless the user asks for detail.
Avoid repetition, unnecessary background, and filler.
Do not tell the user to "consult a professional agent."
For questions outside the reference, follow this exact policy: {out_of_scope_instruction}
Do not mention image-reference identifiers; the system handles image display separately.
If a spreadsheet value says it is pending confirmation, blank, or not explicitly stated by the official source, say that the current material does not confirm it. Do not fill in or guess missing facts.

[Reference Content]
{doc.content}
""".strip()


def _get_ai_client() -> AsyncOpenAI:
    api_key = os.getenv("DEEPSEEK_API_KEY", "").strip().strip('"').strip("'")
    invalid_placeholder = {
        "your_key_here",
        "your_real_key",
        "change_me",
        "xxx",
    }
    if not api_key or api_key.lower() in invalid_placeholder:
        raise HTTPException(
            status_code=500,
            detail="No valid DEEPSEEK_API_KEY is configured. Add a real key to backend/.env.local (recommended) or backend/.env; do not use a placeholder.",
        )
    return AsyncOpenAI(api_key=api_key, base_url="https://api.deepseek.com")


def _build_chat_messages(payload: ChatRequest, doc: BuildingDoc) -> List[dict]:
    system_prompt = _build_system_prompt(doc)
    messages: List[dict] = [{"role": "system", "content": system_prompt}]
    for item in payload.history:
        if item.role in {"user", "assistant"}:
            messages.append({"role": item.role, "content": item.content})
    messages.append({"role": "user", "content": payload.question})
    return messages


def _summary_from_doc(doc: BuildingDoc) -> dict:
    return {
        "building_id": doc.building_id,
        "building_name": doc.building_name,
        "address": doc.address,
        "category": doc.category,
        "source_type": doc.source_type,
        "file_name": doc.file_name,
        "relative_path": doc.relative_path,
        "label": _doc_label(doc),
        "preview": doc.content[:200],
    }


def _internet_suggestion_summary(record: InternetRecord) -> dict:
    return {
        "internet_id": record.internet_id,
        "building_name": record.building_name,
        "address": record.address,
        "label": f"{record.building_name}（{record.address}）"
        if record.address
        else record.building_name,
    }


def _internet_result_payload(record: InternetRecord) -> dict:
    payload = _internet_suggestion_summary(record)
    payload.update(
        {
            "website": record.website,
            "contact": record.contact,
            "wifi_mode": record.wifi_mode,
            "mode_detail": record.mode_detail,
            "notes": record.notes,
            "providers": [record.providers[key] for key in INTERNET_PROVIDER_ORDER],
            "source_file": record.source_file,
            "source_sheets": record.source_sheets,
            "relative_path": record.relative_path,
        }
    )
    return payload


def _canonicalize_text(text: str) -> str:
    # Normalize Chinese/English punctuation and address variants before matching.
    canonical = (text or "").lower()
    canonical = canonical.replace("，", " ").replace("。", " ").replace("、", " ")
    canonical = canonical.replace("（", " ").replace("）", " ").replace("：", " ")
    canonical = re.sub(r"(\d+)\s*号", r"\1", canonical)
    for pattern, replacement in ADDRESS_ABBR_PATTERNS:
        canonical = pattern.sub(replacement, canonical)
    return canonical


def _normalize_text(text: str) -> str:
    canonical = _canonicalize_text(text)
    return re.sub(r"[\s\-\—\–\－_.,，。:：;；/\\()（）\[\]【】]+", "", canonical)


def _tokenize_text(text: str) -> List[str]:
    canonical = _canonicalize_text(text)
    tokens = re.findall(r"[a-z0-9]+|[\u4e00-\u9fff]+", canonical)
    result: List[str] = []
    for token in tokens:
        if token in QUERY_STOPWORDS:
            continue
        if len(token) == 1 and not token.isdigit():
            continue
        result.append(token)
    return result


def _extract_numeric_tokens(text: str) -> set:
    canonical = _canonicalize_text(text)
    return set(re.findall(r"\d+[a-z]?", canonical))


def _category_hint_score(query: str, category: str) -> int:
    hints = CATEGORY_HINTS.get((category or "").lower(), set())
    if not hints:
        return 0

    canonical_query = _canonicalize_text(query)
    score = 0
    for hint in hints:
        if hint in canonical_query:
            score += 90
    return min(score, 260)


def _is_electric_query(text: str) -> bool:
    canonical = _canonicalize_text(text)
    electric_hints = CATEGORY_HINTS.get("electric", set()) | CATEGORY_HINTS.get("electricity", set())
    return any(hint in canonical for hint in electric_hints)


def _score_building_match(query: str, doc: BuildingDoc) -> int:
    # Score direct name/address matches, token overlap, street numbers, and fuzzy similarity.
    query_norm = _normalize_text(query)
    if not query_norm:
        return 0

    name_norm = _normalize_text(doc.building_name)
    addr_norm = _normalize_text(doc.address)
    display_norm = _normalize_text(f"{doc.building_name}{doc.address}{doc.category}")
    file_norm = _normalize_text(Path(doc.file_name).stem)
    category_norm = _normalize_text(doc.category)
    relative_norm = _normalize_text(doc.relative_path)
    query_tokens = set(_tokenize_text(query))
    all_tokens = set(
        _tokenize_text(
            f"{doc.building_name} {doc.address} {doc.file_name} {doc.category} {doc.relative_path}"
        )
    )
    address_tokens = set(_tokenize_text(doc.address))
    query_numbers = _extract_numeric_tokens(query)
    address_numbers = _extract_numeric_tokens(f"{doc.address} {doc.file_name}")

    score = 0
    if query_norm == name_norm:
        score += 180
    elif name_norm and query_norm in name_norm:
        score += 130
    elif query_norm and name_norm and name_norm in query_norm:
        score += 95

    if addr_norm and query_norm == addr_norm:
        score += 170
    elif addr_norm and query_norm in addr_norm:
        score += 120
    elif query_norm and addr_norm and addr_norm in query_norm:
        score += 90

    if query_norm in display_norm:
        score += 85
    if query_norm in file_norm:
        score += 50
    if category_norm and query_norm in category_norm:
        score += 80
    if relative_norm and query_norm in relative_norm:
        score += 45
    score += _category_hint_score(query, doc.category)

    # Fallback for conversational queries that contain an address inside a sentence.
    if query_tokens and all_tokens:
        overlap = query_tokens & all_tokens
        if overlap:
            overlap_count = len(overlap)
            score += overlap_count * 25
            score += int(90 * overlap_count / max(len(query_tokens), 1))

            # Give address-token matches extra weight for address-only queries.
            addr_overlap = query_tokens & address_tokens
            if addr_overlap:
                score += len(addr_overlap) * 20

    if query_numbers and address_numbers:
        common_numbers = query_numbers & address_numbers
        if common_numbers:
            score += 110
        if query_numbers.issubset(address_numbers):
            score += 30

    # Fuzzy fallback for minor misspellings and abbreviations.
    similarity_candidates = [name_norm, addr_norm, display_norm, file_norm, category_norm]
    best_similarity = 0.0
    for candidate in similarity_candidates:
        if not candidate:
            continue
        ratio = SequenceMatcher(None, query_norm, candidate).ratio()
        if ratio > best_similarity:
            best_similarity = ratio
    score += int(best_similarity * 70)

    return score


def _score_internet_match(query: str, record: InternetRecord) -> int:
    query_norm = _normalize_text(query)
    if not query_norm:
        return 0

    name_norm = _normalize_text(record.building_name)
    addr_norm = _normalize_text(record.address)
    display_norm = _normalize_text(f"{record.building_name}{record.address}{record.website}")
    relative_norm = _normalize_text(record.relative_path)
    query_tokens = set(_tokenize_text(query))
    all_tokens = set(_tokenize_text(f"{record.building_name} {record.address} {record.website}"))
    address_tokens = set(_tokenize_text(record.address))
    query_numbers = _extract_numeric_tokens(query)
    address_numbers = _extract_numeric_tokens(record.address)

    score = 0
    if query_norm == name_norm:
        score += 180
    elif name_norm and query_norm in name_norm:
        score += 130
    elif query_norm and name_norm and name_norm in query_norm:
        score += 95

    if addr_norm and query_norm == addr_norm:
        score += 170
    elif addr_norm and query_norm in addr_norm:
        score += 120
    elif query_norm and addr_norm and addr_norm in query_norm:
        score += 90

    if query_norm in display_norm:
        score += 80
    if relative_norm and query_norm in relative_norm:
        score += 35

    if query_tokens and all_tokens:
        overlap = query_tokens & all_tokens
        if overlap:
            overlap_count = len(overlap)
            score += overlap_count * 25
            score += int(90 * overlap_count / max(len(query_tokens), 1))

            addr_overlap = query_tokens & address_tokens
            if addr_overlap:
                score += len(addr_overlap) * 20

    if query_numbers and address_numbers:
        common_numbers = query_numbers & address_numbers
        if common_numbers:
            score += 110
        if query_numbers.issubset(address_numbers):
            score += 30

    best_similarity = 0.0
    for candidate in [name_norm, addr_norm, display_norm]:
        if not candidate:
            continue
        ratio = SequenceMatcher(None, query_norm, candidate).ratio()
        if ratio > best_similarity:
            best_similarity = ratio
    score += int(best_similarity * 70)

    return score


SOURCE_PRIORITY_IGNORED_TOKENS = {
    "building",
    "document",
    "excel",
    "file",
    "pdf",
    "row",
    "sheet",
    "xls",
    "xlsx",
}

GENERIC_LOOKUP_TOKENS = SOURCE_PRIORITY_IGNORED_TOKENS | {
    "broadband",
    "csv",
    "electric",
    "electricity",
    "internet",
    "maintenance",
    "nbn",
    "sim",
    "spreadsheet",
    "tsv",
    "wifi",
    "停车",
    "公寓",
    "号码",
    "大楼",
    "宽带",
    "开电",
    "手机卡",
    "报修",
    "搬家",
    "文件",
    "楼盘",
    "流量",
    "电费",
    "电话卡",
    "租房",
    "维修",
    "网络",
    "表格",
    "资料",
    "车位",
    "门禁",
    "物业",
}


def _has_specific_lookup_hint(query: str) -> bool:
    tokens = set(_tokenize_text(query)) - GENERIC_LOOKUP_TOKENS
    return bool(tokens or _extract_numeric_tokens(query))


def _source_priority_bonus(query: str, doc: BuildingDoc) -> int:
    # Prefer a building-specific PDF; keep summary spreadsheet rows as a fallback.
    if doc.category != "building" or doc.source_type != "pdf":
        return 0

    query_norm = _normalize_text(query)
    name_norm = _normalize_text(doc.building_name)
    addr_norm = _normalize_text(doc.address)
    file_norm = _normalize_text(Path(doc.file_name).stem)
    direct_hit = any(
        candidate and (query_norm in candidate or candidate in query_norm)
        for candidate in (name_norm, addr_norm, file_norm)
    )
    if direct_hit:
        return 90

    query_tokens = set(_tokenize_text(query)) - SOURCE_PRIORITY_IGNORED_TOKENS
    identity_tokens = set(
        _tokenize_text(f"{doc.building_name} {doc.address} {Path(doc.file_name).stem}")
    )
    if query_tokens & identity_tokens:
        return 90

    return 0


def _resolve_building(query: str) -> Tuple[Optional[BuildingDoc], List[Tuple[int, BuildingDoc]]]:
    # Select a credible match and retain the ranking for suggestions.
    if not _has_specific_lookup_hint(query):
        return None, []

    ranked: List[Tuple[int, BuildingDoc]] = []
    for doc in BUILDING_STORE.values():
        score = _score_building_match(query, doc)
        if score > 0:
            ranked.append((score + _source_priority_bonus(query, doc), doc))

    ranked.sort(key=lambda item: (-item[0], item[1].building_name))
    if not ranked:
        return None, []

    top_score, top_doc = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else -1

    # Accept a high score directly; medium scores require a clear margin over second place.
    matched = None
    if top_score >= 170:
        matched = top_doc
    elif top_score >= 120 and (second_score < 0 or (top_score - second_score) >= 25):
        matched = top_doc
    elif top_score >= 95 and len(ranked) == 1:
        matched = top_doc

    return matched, ranked


def _resolve_internet(query: str) -> Tuple[Optional[InternetRecord], List[Tuple[int, InternetRecord]]]:
    if not _has_specific_lookup_hint(query):
        return None, []

    ranked: List[Tuple[int, InternetRecord]] = []
    for record in INTERNET_STORE.values():
        score = _score_internet_match(query, record)
        if score > 0:
            ranked.append((score, record))

    ranked.sort(key=lambda item: (-item[0], item[1].building_name))
    if not ranked:
        return None, []

    top_score, top_record = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else -1

    matched = None
    if top_score >= 170:
        matched = top_record
    elif top_score >= 120 and (second_score < 0 or (top_score - second_score) >= 25):
        matched = top_record
    elif top_score >= 95 and len(ranked) == 1:
        matched = top_record

    return matched, ranked


def _iter_supported_document_paths() -> List[Path]:
    paths: List[Path] = []
    for file_path in PDF_LIBRARY_DIR.rglob("*"):
        if not file_path.is_file():
            continue
        if file_path.name.startswith("~$"):
            continue
        if file_path.suffix.lower() in SUPPORTED_DOCUMENT_EXTENSIONS:
            paths.append(file_path)
    return sorted(paths)


def reload_building_store() -> dict:
    # Rescan the library and rebuild the in-memory index with stable path-based identifiers.
    PDF_LIBRARY_DIR.mkdir(parents=True, exist_ok=True)

    new_store: Dict[str, BuildingDoc] = {}
    errors: List[str] = []

    for doc_path in _iter_supported_document_paths():
        try:
            if doc_path.stat().st_size == 0:
                errors.append(f"{doc_path.name}: the file is empty.")
                continue

            relative_path = str(doc_path.relative_to(PDF_LIBRARY_DIR))
            category = _category_from_doc_path(doc_path)
            if category == "internet":
                continue
            row_entries = _extract_spreadsheet_row_entries(doc_path)
            if row_entries:
                for entry in row_entries:
                    unique_key = entry["relative_path"]
                    building_id = hashlib.sha1(unique_key.encode("utf-8")).hexdigest()[:12]
                    new_store[building_id] = BuildingDoc(
                        building_id=building_id,
                        building_name=entry["building_name"],
                        address=entry["address"],
                        category=category,
                        file_name=doc_path.name,
                        file_path=doc_path,
                        relative_path=entry["relative_path"],
                        source_type=entry["source_type"],
                        content=entry["content"],
                    )
                continue

            content = _extract_document_text(doc_path)
            if not content:
                errors.append(f"{doc_path.name}: no text could be extracted.")
                continue

            building_name, address = _parse_filename_to_building(doc_path.name)
            unique_key = relative_path
            building_id = hashlib.sha1(unique_key.encode("utf-8")).hexdigest()[:12]

            new_store[building_id] = BuildingDoc(
                building_id=building_id,
                building_name=building_name,
                address=address,
                category=category,
                file_name=doc_path.name,
                file_path=doc_path,
                relative_path=relative_path,
                source_type=_source_type_from_path(doc_path),
                content=content,
            )
        except Exception as exc:
            errors.append(f"{doc_path.relative_to(PDF_LIBRARY_DIR)}: parsing failed ({exc}).")

    BUILDING_STORE.clear()
    BUILDING_STORE.update(new_store)
    image_info = reload_image_store()

    return {
        "base_dir": str(PDF_LIBRARY_DIR),
        "image_base_dir": str(IMAGE_LIBRARY_DIR),
        "supported_extensions": sorted(SUPPORTED_DOCUMENT_EXTENSIONS),
        "loaded": len(BUILDING_STORE),
        "image_docs_with_assets": image_info["docs_with_images"],
        "image_assets_loaded": image_info["total_images"],
        "errors": errors,
        "image_errors": image_info["errors"],
    }


@app.on_event("startup")
async def startup_load_documents():
    # Warm the in-memory index at startup instead of scanning on the first query.
    reload_building_store()
    reload_internet_store()
    reload_script_store()


@app.get("/buildings")
async def list_buildings(reload: bool = Query(default=False)):
    if reload:
        reload_building_store()
        reload_internet_store()

    buildings = sorted(
        (_summary_from_doc(doc) for doc in BUILDING_STORE.values()),
        key=lambda item: item["building_name"],
    )
    return {
        "base_dir": str(PDF_LIBRARY_DIR),
        "image_base_dir": str(IMAGE_LIBRARY_DIR),
        "supported_extensions": sorted(SUPPORTED_DOCUMENT_EXTENSIONS),
        "total": len(buildings),
        "internet_total": len(INTERNET_STORE),
        "internet_warnings": list(INTERNET_WARNINGS),
        "internet_errors": list(INTERNET_ERRORS),
        "docs_with_images": len(IMAGE_STORE),
        "images_total": sum(len(items) for refs in IMAGE_STORE.values() for items in refs.values()),
        "buildings": buildings,
    }


@app.post("/buildings/reload")
async def reload_buildings():
    return {
        "documents": reload_building_store(),
        "internet": reload_internet_store(),
    }


@app.post("/internet/reload")
async def reload_internet():
    return reload_internet_store()


@app.post("/internet/resolve")
async def resolve_internet(payload: ResolveInternetRequest):
    if not INTERNET_STORE:
        return {
            "matched": None,
            "suggestions": [],
            "message": INTERNET_ERRORS[0]
            if INTERNET_ERRORS
            else "No internet reference is available. Check the Excel file in the internet folder, then refresh the library.",
            "score": 0,
            "second_score": 0,
        }

    matched, ranked = _resolve_internet(payload.query)
    top_score = ranked[0][0] if ranked else 0
    second_score = ranked[1][0] if len(ranked) > 1 else 0
    suggestions = [item[1] for item in ranked[:5]]

    if matched is not None:
        return {
            "matched": _internet_result_payload(matched),
            "suggestions": [],
            "message": f"Matched internet reference: {matched.building_name}",
            "score": top_score,
            "second_score": second_score,
        }

    return {
        "matched": None,
        "suggestions": [_internet_suggestion_summary(item) for item in suggestions],
        "message": "No exact internet reference was found. Provide a more complete building name or address.",
        "score": top_score,
        "second_score": second_score,
    }


@app.get("/scripts")
async def list_scripts(reload: bool = Query(default=False)):
    if reload:
        reload_script_store()

    return {
        "base_dir": str(SCRIPT_LIBRARY_DIR),
        "total_intents": len(SCRIPT_STORE),
        "scripts": SCRIPT_STORE,
    }


@app.post("/scripts/reload")
async def reload_scripts():
    return reload_script_store()


@app.get("/scripts/{intent}/random")
async def random_script(intent: str, fallback: str = Query(default="")):
    normalized = _normalize_intent_key(intent)
    variants = SCRIPT_STORE.get(normalized, [])
    content = pick_script(normalized, fallback=fallback)
    return {
        "intent": normalized,
        "content": content,
        "available": len(variants),
    }


@app.post("/images/reload")
async def reload_images():
    return reload_image_store()


@app.get("/buildings/{building_id}/images")
async def list_building_images(building_id: str):
    doc = BUILDING_STORE.get(building_id)
    if not doc:
        raise HTTPException(status_code=404, detail="The reference does not exist, so its images cannot be listed.")

    image_map = IMAGE_STORE.get(building_id, {})
    total = sum(len(items) for items in image_map.values())
    return {
        "building_id": building_id,
        "building_name": doc.building_name,
        "address": doc.address,
        "category": doc.category,
        "image_base_dir": str(IMAGE_LIBRARY_DIR),
        "total_references": len(image_map),
        "total_images": total,
        "images_by_reference": image_map,
    }


@app.post("/images/resolve")
async def resolve_images(payload: ResolveImagesRequest):
    doc = BUILDING_STORE.get(payload.building_id)
    if not doc:
        raise HTTPException(status_code=404, detail="The reference does not exist, so images cannot be resolved.")

    normalized_refs: List[str] = []
    seen = set()
    for raw_ref in payload.references:
        normalized = _normalize_image_reference(raw_ref)
        if normalized and normalized not in seen:
            seen.add(normalized)
            normalized_refs.append(normalized)

    images = _resolve_images_for_doc(payload.building_id, normalized_refs)
    return {
        "building_id": payload.building_id,
        "references": normalized_refs,
        "images": images,
        "total": len(images),
    }


@app.post("/images/resolve-from-text")
async def resolve_images_from_text(payload: ResolveImagesFromTextRequest):
    doc = BUILDING_STORE.get(payload.building_id)
    if not doc:
        raise HTTPException(status_code=404, detail="The reference does not exist, so images cannot be resolved.")

    references = _extract_image_references_from_text(payload.text)
    if not references:
        references = _infer_image_references_from_query(payload.building_id, payload.text)
    images = _resolve_images_for_doc(payload.building_id, references)
    return {
        "building_id": payload.building_id,
        "references": references,
        "images": images,
        "total": len(images),
    }


@app.post("/buildings/resolve")
async def resolve_building(payload: ResolveBuildingRequest):
    if not BUILDING_STORE:
        return {
            "matched": None,
            "suggestions": [],
            "message": "No reference is available. Add a PDF, Excel, or CSV file to the backend pdf_library folder and refresh.",
            "score": 0,
            "second_score": 0,
            "is_switch_candidate": False,
        }

    matched, ranked = _resolve_building(payload.query)
    top_score = ranked[0][0] if ranked else 0
    second_score = ranked[1][0] if len(ranked) > 1 else 0
    suggestions = [item[1] for item in ranked[:5]]

    is_switch_candidate = False
    if matched is not None and payload.current_building_id:
        is_switch_candidate = matched.building_id != payload.current_building_id

    if matched is not None:
        return {
            "matched": _summary_from_doc(matched),
            "suggestions": [],
            "message": f"Matched reference: {_doc_label(matched)}",
            "score": top_score,
            "second_score": second_score,
            "is_switch_candidate": is_switch_candidate,
        }

    return {
        "matched": None,
        "suggestions": [_summary_from_doc(item) for item in suggestions],
        "message": "No exact reference was found. Provide a more complete building name, address, or service keyword.",
        "score": top_score,
        "second_score": second_score,
        "is_switch_candidate": False,
    }


@app.post("/chat")
async def chat_with_document(payload: ChatRequest):
    doc = BUILDING_STORE.get(payload.building_id)
    if not doc:
        raise HTTPException(
            status_code=404,
            detail="The selected reference is missing or no longer valid. Check the reference folder and refresh the library.",
        )

    if _is_direct_contact_doc(doc):
        direct_content = _get_direct_contact_answer(doc)

        async def direct_event_stream():
            yield f"data: {json.dumps({'type': 'token', 'content': direct_content}, ensure_ascii=False)}\n\n"
            yield f"data: {json.dumps({'type': 'done'}, ensure_ascii=False)}\n\n"

        return StreamingResponse(
            direct_event_stream(),
            media_type="text/event-stream",
            headers={
                "Cache-Control": "no-cache",
                "Connection": "keep-alive",
                "X-Accel-Buffering": "no",
            },
        )

    client = _get_ai_client()
    messages = _build_chat_messages(payload, doc)

    async def event_stream():
        try:
            # Stream tokens over SSE for the frontend typing effect.
            stream = await client.chat.completions.create(
                model="deepseek-chat",
                messages=messages,
                stream=True,
                temperature=0.2,
                max_tokens=MAX_ANSWER_TOKENS,
            )

            has_content = False
            async for chunk in stream:
                content = ""
                if chunk.choices and chunk.choices[0].delta:
                    content = chunk.choices[0].delta.content or ""
                if not content:
                    continue

                has_content = True
                data = {"type": "token", "content": content}
                yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"

            yield f"data: {json.dumps({'type': 'done'}, ensure_ascii=False)}\n\n"
        except Exception as exc:
            error_data = {"type": "error", "message": f"The AI service request failed: {exc}"}
            yield f"data: {json.dumps(error_data, ensure_ascii=False)}\n\n"

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/chat/mobile")
async def chat_with_document_mobile(payload: ChatRequest):
    doc = BUILDING_STORE.get(payload.building_id)
    if not doc:
        raise HTTPException(
            status_code=404,
            detail="The selected reference is missing or no longer valid. Check the reference folder and refresh the library.",
        )

    if _is_direct_contact_doc(doc):
        return {
            "content": _get_direct_contact_answer(doc),
            "building_id": doc.building_id,
            "label": _doc_label(doc),
            "model": "direct-contact",
        }

    client = _get_ai_client()
    messages = _build_chat_messages(payload, doc)

    try:
        completion = await client.chat.completions.create(
            model="deepseek-chat",
            messages=messages,
            stream=False,
            temperature=0.2,
            max_tokens=MAX_ANSWER_TOKENS,
        )
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"The AI service request failed: {exc}") from exc

    content = ""
    if completion.choices and completion.choices[0].message:
        content = completion.choices[0].message.content or ""

    return {
        "content": content.strip(),
        "building_id": doc.building_id,
        "label": _doc_label(doc),
        "model": "deepseek-chat",
    }


@app.get("/health")
async def health_check():
    images_total = sum(len(items) for refs in IMAGE_STORE.values() for items in refs.values())
    return {
        "status": "ok",
        "pdf_library_dir": str(PDF_LIBRARY_DIR),
        "script_library_dir": str(SCRIPT_LIBRARY_DIR),
        "image_library_dir": str(IMAGE_LIBRARY_DIR),
        "frontend_dist_dir": str(FRONTEND_DIST_DIR),
        "packaged_frontend_available": FRONTEND_INDEX_PATH.is_file(),
        "supported_document_extensions": sorted(SUPPORTED_DOCUMENT_EXTENSIONS),
        "max_spreadsheet_rows_per_sheet": MAX_SPREADSHEET_ROWS_PER_SHEET,
        "max_spreadsheet_cols_per_sheet": MAX_SPREADSHEET_COLS_PER_SHEET,
        "max_spreadsheet_cells_per_sheet": MAX_SPREADSHEET_CELLS_PER_SHEET,
        "buildings_in_memory": len(BUILDING_STORE),
        "documents_in_memory": len(BUILDING_STORE),
        "script_intents_in_memory": len(SCRIPT_STORE),
        "image_docs_in_memory": len(IMAGE_STORE),
        "image_assets_in_memory": images_total,
        "image_hint_docs_in_memory": len(IMAGE_HINT_STORE),
    }


if FRONTEND_ASSETS_DIR.is_dir():
    app.mount(
        "/assets",
        StaticFiles(directory=str(FRONTEND_ASSETS_DIR), check_dir=False),
        name="frontend-assets",
    )


def _resolve_frontend_candidate(path: str) -> Optional[Path]:
    if not path:
        return FRONTEND_INDEX_PATH if FRONTEND_INDEX_PATH.is_file() else None

    candidate = (FRONTEND_DIST_DIR / path.lstrip("/")).resolve()
    try:
        candidate.relative_to(FRONTEND_DIST_DIR)
    except ValueError:
        return None

    if candidate.is_file():
        return candidate
    return None


def _serve_packaged_frontend(path: str = "") -> FileResponse:
    if not FRONTEND_INDEX_PATH.is_file():
        raise HTTPException(
            status_code=404,
            detail="The packaged frontend was not found. Build the release package first.",
        )

    candidate = _resolve_frontend_candidate(path)
    if candidate is not None:
        return FileResponse(candidate)
    return FileResponse(FRONTEND_INDEX_PATH)


@app.get("/")
async def serve_frontend_index():
    return _serve_packaged_frontend("")


@app.get("/{path:path}")
async def serve_frontend_app(path: str):
    if path.startswith(
        (
            "buildings",
            "internet",
            "scripts",
            "images",
            "chat",
            "health",
            "image-files",
            "openapi.json",
            "docs",
            "redoc",
        )
    ):
        raise HTTPException(status_code=404, detail="Not found")
    return _serve_packaged_frontend(path)
